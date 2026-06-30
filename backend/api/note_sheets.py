from __future__ import annotations

import contextlib
from copy import deepcopy
from datetime import date, datetime, timedelta
import io
import json
import os
from pathlib import Path
import shutil
import sys
import threading
import time
import re
import uuid
import anyio
from math import ceil, floor, isfinite
from typing import Any, Literal, Optional
from urllib.parse import quote

from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.cron import CronTrigger
from fastapi import APIRouter, Depends, File, Form, Header, HTTPException, Query, UploadFile, WebSocket, WebSocketDisconnect
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from sqlalchemy import or_
from sqlmodel import Session, delete, func, select

from backend.core.ai.app_config import (
    AI_APP_NOTE_SHEET_CLOCKIN_LINK_DETECTION,
    AI_APP_NOTE_SHEET_EXCEL_IMPORT,
    resolve_ai_app_runtime_config,
)
from backend.core.ai.chat import OllamaClientError, chat_with_provider
from backend.core.access.auth import (
    extract_api_token,
    get_current_active_user,
    get_optional_current_user_from_token,
    validate_api_token_value,
)
from backend.api.websocket_manager import manager as ws_manager
from backend.core.attendance.service import (
    get_attendance_course_data_flow_config,
    get_attendance_course_data_step_runner_device,
    get_attendance_service_extra_config,
    get_or_create_attendance_service_config,
)
from backend.core.attendance.progress_style import (
    highlight_presence_progress,
    highlight_text_refund_progress,
    highlight_threshold_refund_progress,
    parse_compact_refund_rules,
    parse_threshold_refund_rules,
)
from backend.core.runtime.background_task_queue import background_task_queue
from backend.core.access.feature_access_guard import ensure_feature_access
from backend.core.notes.sheet_access import ensure_attendance_sheet_anonymous_viewer
from backend.core.notes import sheet_inline_links as note_sheet_inline_links
from backend.core.settings import get_settings
from backend.db import engine, get_session
from backend.models import (
    AppSetting,
    ResourceAccessGrant,
    SheetDocument,
    User,
    UserDevice,
    WorkbookDocument,
    WorkbookSheetLink,
)
from backend.core.resources.identity import (
    RESOURCE_TYPE_SHEET,
    RESOURCE_TYPE_WORKBOOK,
    ensure_resource_identity,
)
from backend.core.resources.sheet_identity import allocate_new_sheet_identity, allocate_new_workbook_identity
from backend.core.resources.sheet_refs import (
    load_sheets_by_refs,
    load_workbooks_by_refs,
    sheet_public_id,
    sheet_ref_aliases,
    workbook_public_id,
    workbook_ref_aliases,
)


router = APIRouter()

DEFAULT_NOTE_SHEET_COLUMNS = ["列1", "列2", "列3"]
DEFAULT_NOTE_SHEET_PAGE_SIZE = 50
MAX_NOTE_SHEET_PAGE_SIZE = 1000
NOTE_SHEET_EXCEL_IMPORT_PROVIDER_ID = "deepseek"
NOTE_SHEET_EXCEL_IMPORT_MODEL = "deepseek-v4-pro"
NOTE_SHEET_EXCEL_IMPORT_TIMEOUT_SECONDS = 900
NOTE_SHEET_EXCEL_IMPORT_MAX_ATTEMPTS = 3
NOTE_SHEET_EXCEL_IMPORT_RETRY_DELAY_SECONDS = 1.0
NOTE_SHEET_EXCEL_IMPORT_MAX_BYTES = 12 * 1024 * 1024
NOTE_SHEET_EXCEL_IMPORT_MAX_SHEETS = 12
NOTE_SHEET_EXCEL_IMPORT_MAX_ROWS_PER_SHEET = 600
NOTE_SHEET_EXCEL_IMPORT_MAX_COLS_PER_SHEET = 60
NOTE_SHEET_EXCEL_IMPORT_MAX_NONEMPTY_ROWS = 900
NOTE_SHEET_EXCEL_IMPORT_EXTRA_COLUMN_WIDTH = 132
NOTE_SHEET_EXCEL_IMPORT_EXTRA_COLUMN_HEADER_BACKGROUND = "#E5E7EB"
NOTE_SHEET_EXCEL_IMPORT_EXTRA_COLUMN_HEADER_TEXT = "#4B5563"
NOTE_SHEET_CELL_ACTION_EXCEL_IMPORT_RESET = "excel_import_reset"
NOTE_SHEET_CELL_ACTION_REGISTRATION_ORDER_MATCH = "registration_order_match"
NOTE_SHEET_CELL_ACTION_REGISTRATION_USER_MATCH = "registration_user_match"
NOTE_SHEET_CELL_ACTION_REGISTRATION_COMPOSITE_UPDATE = "registration_composite_update"
NOTE_SHEET_CELL_ACTION_ATTENDANCE_EXPORT = "attendance_export"
NOTE_SHEET_CELL_ACTION_CLOCKIN_LINK_DETECT = "clockin_link_detect"
NOTE_SHEET_ATTENDANCE_REFUND_FAQ_TEXT = "考勤返款常见问题解答"
NOTE_SHEET_ATTENDANCE_REFUND_FAQ_URL = "https://kdocs.cn/l/cmMznDf1i3ye"
NOTE_SHEET_ATTENDANCE_FEEDBACK_TEXT = "反馈问题"
NOTE_SHEET_ATTENDANCE_FEEDBACK_URL = "/attendance-feedback"
NOTE_SHEET_INLINE_LINK_STYLE = {"text_color": "#0000FF", "underline": True}
NOTE_SHEET_ATTENDANCE_REFUND_FAQ_LINK_STYLE = {"text_color": "#FF0000", "underline": True}
NOTE_SHEET_EXCEL_IMPORT_ACTION_TOKENS = ("导入excel", "导入Excel", "导入EXCEL")
NOTE_SHEET_REGISTRATION_ORDER_REQUIRED_COLUMNS = ["微信支付订单号", "订单日期", "商户订单号", "订单金额"]
NOTE_SHEET_REGISTRATION_ORDER_OPTIONAL_COLUMNS = ["已返款"]
NOTE_SHEET_REGISTRATION_ORDER_ACTION_OPTIONAL_COLUMNS = [*NOTE_SHEET_REGISTRATION_ORDER_OPTIONAL_COLUMNS, "备注"]
NOTE_SHEET_REGISTRATION_ORDER_COLUMNS = [
    *NOTE_SHEET_REGISTRATION_ORDER_REQUIRED_COLUMNS,
    *NOTE_SHEET_REGISTRATION_ORDER_OPTIONAL_COLUMNS,
]
NOTE_SHEET_REGISTRATION_USER_LOOKUP_COLUMNS = ["姓名", "微信昵称", "手机号", "错误手机号", "用户ID", "匹配得分"]
NOTE_SHEET_REGISTRATION_SEQUENCE_COLUMN = "序号"
NOTE_SHEET_REGISTRATION_GROUP_COLUMN = "分组"
NOTE_SHEET_REGISTRATION_SUBMITTED_AT_COLUMN = "提交时间"
NOTE_SHEET_REGISTRATION_LINKED_USER_ID_COLUMN = "关联用户ID"
NOTE_SHEET_REGISTRATION_TEMPLATE_BASE_COLUMNS = [
    NOTE_SHEET_REGISTRATION_GROUP_COLUMN,
    NOTE_SHEET_REGISTRATION_SEQUENCE_COLUMN,
    "备注",
    NOTE_SHEET_REGISTRATION_SUBMITTED_AT_COLUMN,
    "姓名",
    "微信昵称",
    "手机号",
    "错误手机号",
    *NOTE_SHEET_REGISTRATION_ORDER_REQUIRED_COLUMNS,
    "用户ID",
    "匹配得分",
    "参考信息",
    NOTE_SHEET_REGISTRATION_LINKED_USER_ID_COLUMN,
]
NOTE_SHEET_REGISTRATION_STANDARD_HEADER_BACKGROUND = "#9DC3E6"
NOTE_SHEET_REGISTRATION_LINKED_USER_ID_NOTE = (
    "有的用户账号数据源不统一，这里可以逗号隔开填写其他相关id，会合并到主id数据中汇总进度"
)
NOTE_SHEET_REGISTRATION_TRACKING_GROUP_COLUMN = "追踪分组"
NOTE_SHEET_REGISTRATION_TRACKING_STATUS_COLUMN = "追踪状态"
NOTE_SHEET_REGISTRATION_TRACKING_DEADLINE_COLUMN = "追踪截止日"
NOTE_SHEET_REGISTRATION_FROZEN_AT_COLUMN = "冻结时间"
NOTE_SHEET_TEMPLATE_RUNTIME_DERIVED_COLUMNS = {
    "规则版本",
    NOTE_SHEET_REGISTRATION_TRACKING_GROUP_COLUMN,
    NOTE_SHEET_REGISTRATION_TRACKING_STATUS_COLUMN,
    NOTE_SHEET_REGISTRATION_TRACKING_DEADLINE_COLUMN,
    NOTE_SHEET_REGISTRATION_FROZEN_AT_COLUMN,
}
NOTE_SHEET_REGISTRATION_FROZEN_GROUP = "A组"
NOTE_SHEET_REGISTRATION_FROZEN_STATUS = "已冻结"
NOTE_SHEET_REGISTRATION_ACTIVE_GROUP = "B组"
NOTE_SHEET_REGISTRATION_ACTIVE_STATUS = "追踪中"
NOTE_SHEET_REGISTRATION_ARCHIVED_BACKGROUND = "#F2F2F2"
NOTE_SHEET_REGISTRATION_ARCHIVED_TEXT = "#6B7280"
NOTE_SHEET_LEGACY_TEXT_PREFIX_STRIP_COLUMNS = {"微信支付订单号", "商户订单号", "手机号", "错误手机号", "微信号"}
NOTE_SHEET_REGISTRATION_DEFAULT_SHOP_ID = 1
NOTE_SHEET_REGISTRATION_ORDER_LOOKUP_MODE = os.environ.get("CODEYUN_NOTE_SHEET_ORDER_LOOKUP_MODE", "db_only")
NOTE_SHEET_REGISTRATION_USER_BROWSER_FALLBACK_DEFAULT = True
NOTE_SHEET_REGISTRATION_BACKGROUND_ACTIONS = {
    NOTE_SHEET_CELL_ACTION_REGISTRATION_ORDER_MATCH,
    NOTE_SHEET_CELL_ACTION_REGISTRATION_USER_MATCH,
    NOTE_SHEET_CELL_ACTION_REGISTRATION_COMPOSITE_UPDATE,
}
NOTE_SHEET_REGISTRATION_USER_BROWSER_DEVICE_NAME = os.environ.get(
    "CODEYUN_NOTE_SHEET_USER_BROWSER_DEVICE_NAME",
    "codepc_mi15",
)
NOTE_SHEET_REGISTRATION_USER_BROWSER_TIMEOUT_SECONDS = os.environ.get(
    "CODEYUN_NOTE_SHEET_USER_BROWSER_TIMEOUT_SECONDS",
    "900",
)
NOTE_SHEET_CLOCKIN_LINK_DETECTION_PROVIDER_ID = os.environ.get(
    "CODEYUN_NOTE_SHEET_CLOCKIN_LINK_PROVIDER_ID",
    "codex-cli",
)
NOTE_SHEET_CLOCKIN_LINK_DETECTION_MODEL = os.environ.get(
    "CODEYUN_NOTE_SHEET_CLOCKIN_LINK_MODEL",
    "gpt-5.3-codex-spark",
)
NOTE_SHEET_CLOCKIN_LINK_DETECTION_TIMEOUT_SECONDS = os.environ.get(
    "CODEYUN_NOTE_SHEET_CLOCKIN_LINK_TIMEOUT_SECONDS",
    "600",
)
NOTE_SHEET_DEFINED_NAMES_KEY = "defined_names"
NOTE_SHEET_WORKBOOK_DEFINED_NAMES_SETTING_PREFIX = "note_sheets.workbook.defined_names."
NOTE_SHEET_ATTENDANCE_INITIAL_ZERO_COLUMNS = {
    "禅客",
    "优秀学员评分",
    "完成视频数",
    "视频应返款",
    "打卡应返款",
    "总应返款",
    "已返款",
    "当前应返款",
}
NOTE_SHEET_ATTENDANCE_REFUND_PERIOD_NAME = "返款周期"
NOTE_SHEET_ATTENDANCE_SOURCE_OVERLAY_COLUMNS = {
    "报名日期",
    "学号",
    "分组",
    "组号",
    "姓名",
    "昵称",
    "微信昵称",
    "手机号",
    "微信支付订单号",
    "订单日期",
    "商户订单号",
    "订单金额",
    "用户ID",
    NOTE_SHEET_REGISTRATION_LINKED_USER_ID_COLUMN,
    "匹配得分",
    "规则版本",
    NOTE_SHEET_REGISTRATION_TRACKING_STATUS_COLUMN,
    NOTE_SHEET_REGISTRATION_TRACKING_GROUP_COLUMN,
    NOTE_SHEET_REGISTRATION_TRACKING_DEADLINE_COLUMN,
}
NOTE_SHEET_ATTENDANCE_GROUP_IDENTITY_BACKGROUND_COLORS = [
    "#DDEBF7",
    "#FCE4D6",
    "#E2F0D9",
    "#FFF2CC",
    "#E4DFEC",
    "#DAEEF3",
    "#F4CCCC",
    "#EADCF8",
]
NOTE_SHEET_ATTENDANCE_VIDEO_PROGRESS_STUDYING_BACKGROUND = "#FFFFBB"
NOTE_SHEET_ATTENDANCE_VIDEO_PROGRESS_COMPLETED_BACKGROUND = "#80FF80"
NOTE_SHEET_ATTENDANCE_VIDEO_ZERO_REFUND_BACKGROUND = "#D9D9D9"
NOTE_SHEET_ATTENDANCE_GROUP_IDENTITY_END_COLUMNS = (
    "禅客",
    "优秀学员评分",
    "用户ID",
    "商户订单号",
)
NOTE_SHEET_ATTENDANCE_TRAILING_META_COLUMNS = {
    "规则版本",
    NOTE_SHEET_REGISTRATION_TRACKING_GROUP_COLUMN,
    NOTE_SHEET_REGISTRATION_TRACKING_STATUS_COLUMN,
    NOTE_SHEET_REGISTRATION_TRACKING_DEADLINE_COLUMN,
    NOTE_SHEET_REGISTRATION_FROZEN_AT_COLUMN,
}
NOTE_SHEET_PERF_LOG_MAX_ENTRIES_PER_BATCH = 200
NOTE_SHEET_PERF_LOG_DIRNAME = "debug"
NOTE_SHEET_PERF_LOG_FILENAME = "note-sheet-perf.jsonl"
RESOURCE_ACCESS_ROLES = ("deny", "viewer", "editor", "manager")
RESOURCE_ACCESS_ROLE_RANK = {"deny": 0, "viewer": 1, "editor": 2, "manager": 3}
RESOURCE_ACCESS_SUBJECT_ANONYMOUS = "anonymous"
RESOURCE_ACCESS_SUBJECT_USER = "user"
RESOURCE_TYPE_WORKBOOK = "workbook"
RESOURCE_TYPE_SHEET = "sheet"
EXCEL_DATE_UNIX_EPOCH_SERIAL = 25569
ATTENDANCE_SUMMARY_WORKBOOK_ID = 2
ATTENDANCE_SUMMARY_SHEET_ID = 4
ATTENDANCE_WJX_DATA_OWNER_TYPE = "attendance_questionnaire"
ATTENDANCE_WJX_DATA_OWNER_KEY = "wjx-data"
ATTENDANCE_WJX_DATA_SHEET_KEY = "data"
ATTENDANCE_WJX_DATA_PUBLIC_EDITABLE_COLUMN_INDEXES = (9,)
ATTENDANCE_TEMPLATE_MONTHLY_SOURCE_COURSES = ("念住", "觉观")
ATTENDANCE_TEMPLATE_ODD_MONTH_SOURCE_COURSES = ("梵呗初阶",)
ATTENDANCE_TEMPLATE_EVEN_MONTH_SOURCE_COURSES = ("梵呗增益",)
ATTENDANCE_TEMPLATE_FANBEI_SOURCE_COURSES = (
    *ATTENDANCE_TEMPLATE_ODD_MONTH_SOURCE_COURSES,
    *ATTENDANCE_TEMPLATE_EVEN_MONTH_SOURCE_COURSES,
)
ATTENDANCE_TEMPLATE_FANBEI_START_DAY = 9
_note_sheet_perf_log_lock = threading.Lock()
ATTENDANCE_FIELD_BINDINGS: dict[str, tuple[str, int]] = {
    "course_type": ("课程类型", 0),
    "course_name": ("课程名称", 1),
    "online_sheet": ("在线考勤表", 2),
    "course_owner": ("考勤负责人", 3),
    "lesson_links": ("课次链接", 4),
    "clockin_links": ("打卡链接", 5),
    "start_date": ("课程开始日期", 8),
    "end_date": ("课程结束日期", 9),
    "completed_date": ("考勤实际完成结点", 10),
    "registration_count": ("报名人数", 12),
}
ATTENDANCE_FIELD_LEGACY_FALLBACKS: dict[str, int] = {
    "start_date": 6,
    "end_date": 7,
    "completed_date": 8,
    "registration_count": 10,
}
ATTENDANCE_TEMPLATE_COURSE_TEXT_RE = re.compile(
    r"(?:(?P<date>\d{8}|\d{6})\s*)?第(?P<edition>\d+)届(?P<course>\S+)",
)
ATTENDANCE_TEMPLATE_LEADING_DATE_RE = re.compile(r"^(?P<date>\d{8}|\d{6})(?P<body>.*)$")
ATTENDANCE_COURSE_SCRIPT_DIR_DEFAULT = Path(r"D:\home\chenkunze\slns\kq5034\courses")
ATTENDANCE_COURSE_SCRIPT_DIR = Path(
    os.environ.get(
        "CODEYUN_KQ5034_COURSES_DIR",
        os.fspath(ATTENDANCE_COURSE_SCRIPT_DIR_DEFAULT),
    )
)
ATTENDANCE_KQ5034_REPO_DIR = Path(
    os.environ.get(
        "CODEYUN_KQ5034_REPO_DIR",
        r"D:\home\chenkunze\slns\kq5034",
    )
)
ATTENDANCE_XLPROJECT_SRC_DIR = Path(
    os.environ.get(
        "CODEYUN_XLPROJECT_SRC_DIR",
        r"D:\home\chenkunze\slns\xlproject\src",
    )
)
ATTENDANCE_COURSE_SCRIPT_STEM_RE = re.compile(r"^d(?P<date>\d{6})(?P<body>.+)$")
ATTENDANCE_COURSE_SCRIPT_INIT_TOKEN_RE = re.compile(
    r"(super\(\).__init__\(\s*[^,]+,\s*(?:XlPath|XPath)\(__file__\)\.stem\s*,\s*)"
    r"(?P<quote>['\"])(?P<token>[^'\"]+)(?P=quote)",
    re.DOTALL,
)
ATTENDANCE_COURSE_SCRIPT_CODEYUN_ATTENDANCE_REF_RE = re.compile(
    r"attendance\s*=\s*CodeYunSheetRef\(\s*"
    r"(?P<workbook_id>\d+)\s*,\s*"
    r"(?P<sheet_id>\d+)\s*,\s*"
    r"(?P<quote>['\"])考勤表(?P=quote)\s*"
    r"\)",
    re.DOTALL,
)
ATTENDANCE_COURSE_SCRIPT_KDOCS_TOKEN_RE = re.compile(r"kdocs\.cn/l/(?P<token>[A-Za-z0-9]+)")
ATTENDANCE_COURSE_EDITION_RE = re.compile(r"第\s*(?P<edition>\d+)\s*届")
ATTENDANCE_COURSE_SCRIPT_PRODUCT_NAME_RE = re.compile(
    r"(?P<prefix>课程商品名\s*=\s*)(?P<quote>['\"])(?P<value>.*?)(?P=quote)",
)
ATTENDANCE_COURSE_SCRIPT_INVALID_FILENAME_RE = re.compile(r'[<>:"/\\|?*\x00-\x1f]')
ATTENDANCE_COURSE_SCRIPT_FILE_EXTENSION_RE = re.compile(
    r"\.(?:xlsx|xlsm|xlsb|xls|et|ett|csv|py)$",
    re.IGNORECASE,
)
NATURAL_SORT_SPLIT_RE = re.compile(r"(\d+)")
FORMULA_CELL_REFERENCE_RE = re.compile(r"(^|[^A-Za-z0-9_.$])(\$?)([A-Za-z]{1,3})(\$?)(\d+)(?![A-Za-z0-9_(!])")
A1_CELL_REFERENCE_RE = re.compile(r"^\s*(?:[^!]+!)?\$?(?P<column>[A-Za-z]{1,3})\$?(?P<row>\d+)\s*$")
R1C1_REFERENCE_RE = re.compile(r"^[Rr]\d*[Cc]\d*$")
FORMULA_DEFINED_NAME_CACHE_KEY = (-1, -1)
FORMULA_PUNCTUATION_TRANSLATION = str.maketrans({
    "，": ",",
    "（": "(",
    "）": ")",
})
DATE_PARSE_FORMULA_RE = re.compile(
    r"""^\s*=\s*(?:DATE_PARSE|日期解析)\s*\(\s*
    (?P<source>\$?[A-Za-z]{1,3}\$?\d+|"(?:[^"]|"")*"|'(?:[^']|'')*')
    (?:\s*,\s*(?P<pattern>"(?:[^"]|"")*"|'(?:[^']|'')*'))?
    (?:\s*,\s*(?P<format>"(?:[^"]|"")*"|'(?:[^']|'')*'))?
    \s*\)\s*$""",
    re.IGNORECASE | re.VERBOSE,
)
CELL_OFFSET_FORMULA_RE = re.compile(
    r"^\s*=\s*(?P<source>\$?[A-Za-z]{1,3}\$?\d+)\s*(?P<operator>[+-])\s*(?P<offset>\d+(?:\.\d+)?)\s*$",
    re.IGNORECASE,
)
FORMULA_CELL_REFERENCE_ONLY_RE = re.compile(r"^\$?([A-Za-z]{1,3})\$?(\d+)$")
NOTE_SHEET_EXCEL_IMPORT_SYSTEM_PROMPT = """你是 CodeYun 星云表格的 Excel 导入标准化代理。

你的任务是把用户上传的 Excel 工作簿标准化为当前目标 sheet 的数据行。你只做数据抽取和字段映射，不修改文件，不调用外部系统。

硬性要求：
- 最终只输出一个 JSON 对象，不要输出 Markdown 代码块或解释性正文。
- 返回的 rows 只包含真实业务数据行，不包含目标 sheet 里已有的按钮行、说明行或字段表头。
- 不确定的字段留空字符串，不要编造。
- 保持源表中真实记录的顺序。
- 所有单元格值都用字符串表示。

报名表常见规则：
- 跳过标题、二维码、空行、分组标题、说明文字、统计行、日志批阅师资名单等非报名记录，除非用户补充说明明确要求导入。
- 如果源表用“1组/2组/第1组”等行表示分组，把该分组填入后续报名记录的“分组”字段，直到遇到下一组。
- 姓名、微信昵称、手机号/手机、微信号、交易单号/微信支付订单号、订单日期、商户订单号、订单金额、退款状态、用户ID 等同义字段要映射到目标列。
- 区分两类订单字段：“微信支付订单号/交易单号/支付订单号”通常是微信支付交易单号，常见为 420... 等长数字；“商户订单号/商户单号/订单号”通常是商户侧订单号，常见为 MA... 等商户前缀。源表只有“订单号”且值是 MA... 时，必须映射到“商户订单号”，不要放入“微信支付订单号”。
- 手机号、订单号、微信号必须按文本保留；不要转成科学计数法，不要自行补全未知位。
- 不要给手机号、订单号、微信号添加用于 Excel 文本识别的反引号、单引号等前缀字符。
- “备注”只用于源表明确表达退课、退款、国际学生、人工备注等当前报名表备注语义的信息；不要把无法匹配的普通源字段塞进“备注”。
- “参考信息”是目标表人工备用字段；除非用户补充说明明确要求，否则不要自动导入到“参考信息”。
- 如果源表存在目标表没有的真实业务字段，放入 extra_columns，并在每行对象里用对应字段名保存值；常见如“选择促学金模式/自觉自律完成学修”归为“促学金模式”，“微信号”在目标表没有专门列时归为“微信号”。
- 如果目标表同时有标准列和源表同名扩展列，例如“微信昵称”和“微信昵称（必填）”、“姓名”和“真实姓名（必填）”，不要只按列名机械映射；要结合具体内容判断哪一个才是用户真实填写的字段。
- “微信昵称”这类字段要优先选择像人类昵称、中文名、英文名、常用网名的内容；不要选择明显是系统生成的数字字母代码、用户 id、提交者账号、openid/unionid 风格字符串、随机 token 或机器标识。
- 如果一个字段名更接近目标列但内容明显是系统代码，另一个字段名略有后缀但内容像真实昵称，应把真实昵称填入目标“微信昵称”，并按需保留源字段列。
- “序号”在报名表里也是考勤表“学号”。如果源表已给出 1_02、2_17、1组02号 等组内学号，必须保留为“组号_两位组内号”的形式，例如 1_02。
- 如果源表按分组独立编号（每个组都从 1 重新开始，或跨组出现重复序号），必须结合“分组”生成“组号_两位组内号”，例如“一组 + 2”写成 1_02，“二组 + 1”写成 2_01。
- 如果源表序号本身是全局唯一流水号，并且没有组内重号或重置迹象，才保留全局序号。
- 无法判断是组内编号还是全局编号时，保持源表可见语义，不要为了凑格式而编造分组前缀；完全缺失序号时再按源记录顺序从 1 开始。

返回 JSON 形状：
{
  "extra_columns": ["可选，目标表没有但源表需要保留的字段名"],
  "rows": [
    {"目标列名或 extra_columns 字段名": "标准化后的文本值"}
  ],
  "warnings": ["可选，导入风险或未能确定的信息"],
  "mapping_notes": ["可选，说明主要字段映射判断"]
}
"""
attendance_summary_scheduler = BackgroundScheduler()


class WorkbookRefItem(BaseModel):
    id: int
    title: str


class NoteSheetAccessCapabilities(BaseModel):
    can_read: bool = False
    can_use_local_view: bool = False
    can_edit_data: bool = False
    editable_data_columns: list[int] = Field(default_factory=list)
    can_edit_config: bool = False
    can_run_sheet_actions: bool = False
    can_manage_access: bool = False


class NoteSheetResourceAccess(BaseModel):
    role: Literal["none", "deny", "viewer", "editor", "manager"] = "none"
    capabilities: NoteSheetAccessCapabilities = Field(default_factory=NoteSheetAccessCapabilities)


class NoteSheetSummaryResponse(BaseModel):
    id: int
    title: str
    engine: str
    scope: str
    version: int = 1
    owner_user_id: Optional[int] = None
    created_by_user_id: Optional[int] = None
    updated_by_user_id: Optional[int] = None
    created_at: float
    updated_at: float
    deleted_at: Optional[float] = None
    deleted_by_user_id: Optional[int] = None
    parent_workbook_id: Optional[int] = None
    workbook_items: list[WorkbookRefItem] = Field(default_factory=list)
    access: Optional[NoteSheetResourceAccess] = None


class NoteSheetDetailResponse(NoteSheetSummaryResponse):
    owner_type: str
    owner_key: str
    sheet_key: str
    version: int
    document_json: dict[str, Any] = Field(default_factory=dict)
    pagination: Optional["NoteSheetPaginationResponse"] = None


class NoteSheetPaginationResponse(BaseModel):
    page: int = 1
    page_size: int = DEFAULT_NOTE_SHEET_PAGE_SIZE
    total_rows: int = 0
    unfiltered_total_rows: Optional[int] = None
    page_count: int = 1
    row_offset: int = 0
    loaded_row_count: int = 0
    row_indexes: Optional[list[int]] = None


class NoteSheetColumnOptionItemResponse(BaseModel):
    value: str
    label: str
    count: int


class NoteSheetColumnOptionsResponse(BaseModel):
    column_index: int
    header: str
    total_rows: int = 0
    options: list[NoteSheetColumnOptionItemResponse] = Field(default_factory=list)


class NoteSheetPerfLogRequest(BaseModel):
    session_id: str = ""
    url: str = ""
    user_agent: str = ""
    workbook_id: Optional[int] = None
    sheet_id: Optional[int] = None
    entries: list[dict[str, Any]] = Field(default_factory=list)


class NoteSheetPerfLogResponse(BaseModel):
    stored_count: int = 0
    path: str = ""


class AttendanceCourseUpdateDataRequest(BaseModel):
    course_type: Literal["fanbei", "nianzhu"]
    course_name: str
    include_frozen: bool = False


class AttendanceCourseUpdateDataResponse(BaseModel):
    step2: Any
    step3: dict[str, Any]


class NoteSheetPagePatchRequest(BaseModel):
    page: int = Field(default=1, ge=1)
    page_size: int = Field(default=DEFAULT_NOTE_SHEET_PAGE_SIZE, ge=1, le=MAX_NOTE_SHEET_PAGE_SIZE)
    row_offset: int = Field(default=0, ge=0)
    loaded_row_count: int = Field(default=0, ge=0)
    row_indexes: list[int] = Field(default_factory=list)
    deleted_row_indexes: list[int] = Field(default_factory=list)


class NoteSheetCreateRequest(BaseModel):
    title: str = ""
    workbook_id: Optional[int] = None
    document_json: dict[str, Any] = Field(default_factory=dict)


class NoteSheetUpdateRequest(BaseModel):
    title: Optional[str] = None
    document_json: Optional[dict[str, Any]] = None
    base_version: int | None = Field(default=None, ge=1)
    page_patch: Optional[NoteSheetPagePatchRequest] = None


class NoteSheetQueryRequest(BaseModel):
    page: int = Field(default=1, ge=1)
    page_size: int | None = Field(default=None, ge=1, le=MAX_NOTE_SHEET_PAGE_SIZE)
    paginate: bool | None = None
    include_workbook_context: bool = True
    column_filters: dict[str, Any] = Field(default_factory=dict)
    row_filter_programs: list[dict[str, Any]] = Field(default_factory=list)


class NoteSheetTableResponse(BaseModel):
    id: int
    workbook_id: int | None = None
    title: str
    version: int
    value_mode: Literal["text", "raw"] = "text"
    columns: list[str] = Field(default_factory=list)
    rows: list[dict[str, Any]] = Field(default_factory=list)
    row_count: int = 0
    data_start_row: int = 0
    field_row_index: int = 0
    grid_rows: list[list[Any]] = Field(default_factory=list)
    defined_name_values: dict[str, Any] = Field(default_factory=dict)


class NoteSheetTablePatchOperation(BaseModel):
    type: Literal["write_fields", "write_range", "set_cell", "set_note_cell"] = "write_fields"
    rows: list[dict[str, Any]] = Field(default_factory=list)
    values: list[list[Any]] = Field(default_factory=list)
    fields: list[str] = Field(default_factory=list)
    key_field: str = ""
    append_missing: bool = False
    start_row_index: int | None = Field(default=None, ge=0)
    start_sheet_row: int | None = Field(default=None, ge=1)
    row_index: int | None = Field(default=None, ge=0)
    sheet_row: int | None = Field(default=None, ge=1)
    column: str | int | None = None
    field: str = ""
    cell: str = ""
    value: Any = None


class NoteSheetTablePatchRequest(BaseModel):
    expected_version: int | None = Field(default=None, ge=1)
    operations: list[NoteSheetTablePatchOperation] = Field(default_factory=list)


class NoteSheetTablePatchResponse(BaseModel):
    sheet: NoteSheetDetailResponse
    table: NoteSheetTableResponse
    updated_cell_count: int = 0
    updated_row_count: int = 0


class NoteSheetCellPatchOperation(BaseModel):
    row_index: int = Field(ge=0)
    column_index: int = Field(ge=0)
    value: Any = None


class NoteSheetCellPatchRequest(BaseModel):
    base_version: int | None = Field(default=None, ge=1)
    operations: list[NoteSheetCellPatchOperation] = Field(default_factory=list)


class NoteSheetCellPatchResponse(BaseModel):
    sheet_id: int
    version: int
    updated_cell_count: int = 0


class NoteSheetPatchOperation(BaseModel):
    op: Literal[
        "set-cell-value",
        "set-cell-meta",
        "set-column-width",
        "set-column-hidden",
        "set-column-config",
        "merge-cells",
        "unmerge-cells",
        "insert-row",
        "delete-row",
        "insert-column",
        "delete-column",
        "move-row",
        "move-column",
    ]
    row_index: int | None = Field(default=None, ge=0)
    column_index: int | None = Field(default=None, ge=0)
    row: Any = None
    col: int | None = Field(default=None, ge=0)
    rowspan: int | None = Field(default=None, ge=1)
    colspan: int | None = Field(default=None, ge=1)
    row_id: str | None = None
    after_row_id: str | None = None
    column_id: str | None = None
    after_column_id: str | None = None
    column: Any = None
    value: Any = None
    meta: dict[str, Any] | None = None
    width: int | float | None = None
    hidden: bool | None = None
    config: dict[str, Any] | None = None


class NoteSheetPatchRequest(BaseModel):
    base_version: int = Field(ge=1)
    ops: list[NoteSheetPatchOperation] = Field(default_factory=list)


class NoteSheetPatchResponse(BaseModel):
    sheet_id: int
    version: int
    applied_op_count: int = 0
    updated_cell_count: int = 0


class NoteSheetSortRequest(BaseModel):
    base_version: int | None = Field(default=None, ge=1)
    column_index: int = Field(ge=0)
    direction: Literal["asc", "desc"] = "asc"


class NoteSheetExcelImportResponse(BaseModel):
    sheet: NoteSheetDetailResponse
    imported_count: int = 0
    preserved_row_count: int = 0
    skipped_duplicate_count: int = 0
    extra_columns: list[str] = Field(default_factory=list)
    warnings: list[str] = Field(default_factory=list)
    mapping_notes: list[str] = Field(default_factory=list)


class NoteSheetRegistrationMatchResponse(BaseModel):
    sheet: NoteSheetDetailResponse
    action: str
    updated_count: int = 0
    skipped_count: int = 0
    error_count: int = 0
    warning_count: int = 0
    message: str = ""


class NoteSheetRegistrationMatchRunRequest(BaseModel):
    action: Literal["registration_order_match", "registration_user_match", "registration_composite_update"]
    use_browser_fallback: bool | None = None
    force_restart: bool = False


class NoteSheetRegistrationMatchRunResponse(BaseModel):
    run_id: str = ""
    action: str
    sheet_id: int
    workbook_id: int | None = None
    status: Literal["idle", "pending", "running", "completed", "failed", "cancelled"] = "idle"
    phase: str = ""
    use_browser_fallback: bool = False
    already_running: bool = False
    cancel_requested: bool = False
    queued_at: float | None = None
    started_at: float | None = None
    finished_at: float | None = None
    total_count: int = 0
    processed_count: int = 0
    updated_count: int = 0
    skipped_count: int = 0
    error_count: int = 0
    warning_count: int = 0
    message: str = ""
    error_message: str | None = None
    sheet: NoteSheetDetailResponse | None = None


class NoteSheetRegistrationUserIdDetectionRequest(BaseModel):
    row_index: int = Field(ge=0)
    base_version: int | None = Field(default=None, ge=1)


class NoteSheetRegistrationUserIdDetectionCandidate(BaseModel):
    user_id: str
    video_count: int = 0
    clockin_count: int = 0
    evidence: list[str] = Field(default_factory=list)
    confidence: Literal["high", "medium", "low"] = "low"


class NoteSheetRegistrationUserIdDetectionResponse(BaseModel):
    sheet: NoteSheetDetailResponse
    attendance_sheet: NoteSheetDetailResponse | None = None
    status: Literal["applied", "review", "skipped", "error"]
    applied: bool = False
    message: str = ""
    target_user_id: str = ""
    applied_to: Literal["用户ID", "关联用户ID", ""] = ""
    candidates: list[NoteSheetRegistrationUserIdDetectionCandidate] = Field(default_factory=list)
    rebuild_summary: dict[str, Any] | None = None
    error_message: str | None = None


class NoteSheetClockinLinkDetectionRunRequest(BaseModel):
    force_restart: bool = False
    provider_id: str = ""
    model: str = ""


class NoteSheetClockinLinkDetectionRunResponse(BaseModel):
    run_id: str = ""
    action: str = NOTE_SHEET_CELL_ACTION_CLOCKIN_LINK_DETECT
    sheet_id: int
    workbook_id: int | None = None
    status: Literal["idle", "pending", "running", "completed", "failed", "cancelled"] = "idle"
    phase: str = ""
    already_running: bool = False
    cancel_requested: bool = False
    queued_at: float | None = None
    started_at: float | None = None
    finished_at: float | None = None
    total_count: int = 0
    processed_count: int = 0
    updated_count: int = 0
    skipped_count: int = 0
    error_count: int = 0
    warning_count: int = 0
    message: str = ""
    error_message: str | None = None
    provider_id: str = NOTE_SHEET_CLOCKIN_LINK_DETECTION_PROVIDER_ID
    model: str = NOTE_SHEET_CLOCKIN_LINK_DETECTION_MODEL
    results: list[dict[str, Any]] = Field(default_factory=list)
    warnings: list[str] = Field(default_factory=list)
    sheet: NoteSheetDetailResponse | None = None


class NoteSheetAttendanceTemplateGenerationRequest(BaseModel):
    base_version: Optional[int] = Field(default=None, ge=1)
    target_year: Optional[int] = Field(default=None, ge=1970, le=9999)
    target_month: Optional[int] = Field(default=None, ge=1, le=12)
    target_date: Optional[str] = None
    skip_course_types: list[str] = Field(default_factory=list)


class NoteSheetAttendanceCourseTemplateGenerationRequest(NoteSheetAttendanceTemplateGenerationRequest):
    row_index: Optional[int] = Field(default=None, ge=0)
    course_type: Optional[str] = None


class NoteSheetAttendanceTemplateActionItem(BaseModel):
    course_type: str
    course_name: str = ""
    target_date: str = ""
    row_index: Optional[int] = None
    reason: str = ""


class NoteSheetAttendanceTemplateGenerationResponse(BaseModel):
    sheet: NoteSheetDetailResponse
    generated: list[NoteSheetAttendanceTemplateActionItem] = Field(default_factory=list)
    skipped: list[NoteSheetAttendanceTemplateActionItem] = Field(default_factory=list)


class NoteSheetAttendanceCompletionRequest(BaseModel):
    base_version: Optional[int] = Field(default=None, ge=1)
    row_index: int = Field(ge=0)
    completion_date: Optional[str] = None


class NoteSheetAttendanceCompletionResponse(BaseModel):
    sheet: NoteSheetDetailResponse
    row_index: int


class NoteSheetAttendanceVideoRevisionCell(BaseModel):
    row_index: int = Field(ge=0)
    column_index: int = Field(ge=0)


class NoteSheetAttendanceVideoRevisionRequest(BaseModel):
    base_version: Optional[int] = Field(default=None, ge=1)
    revision_label: str
    cells: list[NoteSheetAttendanceVideoRevisionCell] = Field(default_factory=list)


class NoteSheetAttendanceVideoRevisionResponse(BaseModel):
    sheet: NoteSheetDetailResponse
    revision_label: str
    updated_count: int = 0
    recalculation: dict[str, Any] = Field(default_factory=dict)


class NoteSheetAttendanceCourseScriptStatusItem(BaseModel):
    row_index: int
    course_type: str = ""
    course_name: str = ""
    online_sheet: str = ""
    url: str = ""
    target_stem: str = ""
    target_filename: str = ""
    exists: bool = False
    existing_path: str = ""
    can_generate: bool = False
    reason: str = ""


class NoteSheetAttendanceCourseScriptStatusesResponse(BaseModel):
    statuses: list[NoteSheetAttendanceCourseScriptStatusItem] = Field(default_factory=list)


class NoteSheetAttendanceCourseScriptGenerationRequest(BaseModel):
    row_index: int = Field(ge=0)


class NoteSheetAttendanceCourseScriptGenerationResponse(BaseModel):
    status: NoteSheetAttendanceCourseScriptStatusItem
    source_filename: str = ""
    source_path: str = ""
    created_path: str = ""


class NoteSheetAttendanceCourseScriptOrganizeItem(BaseModel):
    row_index: int
    course_type: str = ""
    online_sheet: str = ""
    target_filename: str = ""
    completed: bool = False
    source_path: str = ""
    target_path: str = ""
    reason: str = ""


class NoteSheetAttendanceCourseScriptOrganizeResponse(BaseModel):
    moved: list[NoteSheetAttendanceCourseScriptOrganizeItem] = Field(default_factory=list)
    skipped: list[NoteSheetAttendanceCourseScriptOrganizeItem] = Field(default_factory=list)


class NoteSheetAttendanceLinkCountUpdateRequest(BaseModel):
    base_version: Optional[int] = Field(default=None, ge=1)
    field_key: Literal["lesson_links", "clockin_links"]
    row_index: Optional[int] = Field(default=None, ge=0)
    repair_with_remote_browser: bool = True


class NoteSheetAttendanceLinkCountUpdateItem(BaseModel):
    row_index: int
    course_name: str = ""
    lookup_name: str = ""
    value: str = ""
    total_count: int = 0
    linked_count: int = 0
    remote_repair_attempted: bool = False
    remote_repair_summary: dict[str, Any] | None = None
    reason: str = ""


class NoteSheetAttendanceLinkCountUpdateResponse(BaseModel):
    sheet: NoteSheetDetailResponse
    updated: list[NoteSheetAttendanceLinkCountUpdateItem] = Field(default_factory=list)
    skipped: list[NoteSheetAttendanceLinkCountUpdateItem] = Field(default_factory=list)


class WorkbookSummaryResponse(BaseModel):
    id: int
    title: str
    owner_user_id: Optional[int] = None
    created_by_user_id: Optional[int] = None
    updated_by_user_id: Optional[int] = None
    created_at: float
    updated_at: float
    deleted_at: Optional[float] = None
    deleted_by_user_id: Optional[int] = None
    sheet_count: int = 0
    access: Optional[NoteSheetResourceAccess] = None


class WorkbookDetailResponse(WorkbookSummaryResponse):
    sheets: list[NoteSheetSummaryResponse] = Field(default_factory=list)
    defined_names: list[NoteSheetDefinedNameItem] = Field(default_factory=list)


class NoteSheetTrashResponse(BaseModel):
    sheets: list[NoteSheetSummaryResponse] = Field(default_factory=list)
    workbooks: list[WorkbookSummaryResponse] = Field(default_factory=list)


class NoteSheetResourceAccessGrantItem(BaseModel):
    subject_type: Literal["anonymous", "user"]
    subject_key: str
    subject_user_id: Optional[int] = None
    username: str = ""
    nickname: str = ""
    role: Literal["deny", "viewer", "editor", "manager"]


class NoteSheetResourceAccessGrantUpdate(BaseModel):
    subject_type: Literal["anonymous", "user"]
    username: Optional[str] = None
    subject_user_id: Optional[int] = None
    role: Literal["none", "deny", "viewer", "editor", "manager"]


class NoteSheetResourceAccessUpdateRequest(BaseModel):
    grants: list[NoteSheetResourceAccessGrantUpdate] = Field(default_factory=list)


class NoteSheetResourceAccessResponse(BaseModel):
    resource_type: Literal["workbook", "sheet"]
    resource_id: int
    access: NoteSheetResourceAccess
    grants: list[NoteSheetResourceAccessGrantItem] = Field(default_factory=list)


class NoteSheetAccessUserOption(BaseModel):
    id: int
    username: str
    nickname: str = ""


class NoteSheetAccessUserOptionsResponse(BaseModel):
    users: list[NoteSheetAccessUserOption] = Field(default_factory=list)


class NoteSheetDefinedNameItem(BaseModel):
    name: str = ""
    formula: str = ""
    comment: str = ""
    scope: Literal["workbook", "worksheet"] | None = None


class NoteSheetDefinedNameWorksheetScope(BaseModel):
    sheet_id: int
    sheet_title: str = ""
    sheet_version: int | None = None
    names: list[NoteSheetDefinedNameItem] = Field(default_factory=list)


class NoteSheetDefinedNamesUpdateRequest(BaseModel):
    base_version: int | None = Field(default=None, ge=1)
    names: list[NoteSheetDefinedNameItem] = Field(default_factory=list)
    worksheets: list[NoteSheetDefinedNameWorksheetScope] = Field(default_factory=list)


class NoteSheetDefinedNamesResponse(BaseModel):
    workbook_id: int | None = None
    sheet_id: int | None = None
    sheet_version: int | None = None
    workbook: list[NoteSheetDefinedNameItem] = Field(default_factory=list)
    worksheet: list[NoteSheetDefinedNameItem] = Field(default_factory=list)
    worksheets: list[NoteSheetDefinedNameWorksheetScope] = Field(default_factory=list)
    effective: list[NoteSheetDefinedNameItem] = Field(default_factory=list)


class WorkbookCreateRequest(BaseModel):
    title: str = ""


class WorkbookUpdateRequest(BaseModel):
    title: str = ""


class WorkbookAttachSheetRequest(BaseModel):
    sheet_id: int


class WorkbookReorderSheetsRequest(BaseModel):
    sheet_ids: list[int] = Field(default_factory=list)


class WorkbookSaveAsRequest(BaseModel):
    mode: Literal["template", "duplicate"] = "duplicate"
    title: str = ""


def _create_default_sheet_document() -> dict[str, Any]:
    return {
        "schema_version": 1,
        "columns": list(DEFAULT_NOTE_SHEET_COLUMNS),
        "column_ids": [_new_sheet_column_id() for _ in DEFAULT_NOTE_SHEET_COLUMNS],
        "rows": [],
        "row_ids": [],
        "grid_rows": [list(DEFAULT_NOTE_SHEET_COLUMNS)],
        "data_start_row": 1,
        "field_row_index": 0,
        "merged_cells": [],
        "formula_reference_origin": "sheet_v2",
        "view_settings": {
            "show_row_numbers": True,
            "row_marker_numbering": "global",
            "row_marker_origin": "sheet",
            "show_column_markers": True,
            "column_marker_style": "letters",
            "frozen_column_count": 0,
            "pagination": {
                "enabled": False,
                "page_size": DEFAULT_NOTE_SHEET_PAGE_SIZE,
            },
        },
    }


def _normalize_document_json(value: Any) -> dict[str, Any]:
    if not isinstance(value, dict):
        return _create_default_sheet_document()
    return _canonicalize_sheet_document_links(dict(value), migrate_legacy_links=True)


def _normalize_created_document_json(value: Any) -> dict[str, Any]:
    if not isinstance(value, dict) or not value:
        return _create_default_sheet_document()
    return _canonicalize_sheet_document_links(dict(value), migrate_legacy_links=True)


def _is_registration_sheet(document: SheetDocument) -> bool:
    return (
        _normalize_sheet_text(document.sheet_key) == "registration"
        or _normalize_sheet_text(document.title) == "报名表"
    )


def _normalize_registration_sheet_header_document(document_json: dict[str, Any]) -> tuple[dict[str, Any], bool]:
    normalized = _normalize_document_json(document_json)
    columns = _normalize_document_columns(normalized)
    next_document = dict(normalized)
    changed = False

    if next_document.get("header_groups") != []:
        next_document["header_groups"] = []
        changed = True
    if next_document.get("merged_cells") != []:
        next_document["merged_cells"] = []
        changed = True

    data_start_row = _normalize_document_data_start_row(next_document)
    if data_start_row <= 0:
        data_start_row = 1
        next_document["data_start_row"] = data_start_row
        changed = True
    field_row_index = next_document.get("field_row_index")
    if not isinstance(field_row_index, int) or field_row_index < 0 or field_row_index >= data_start_row:
        next_document["field_row_index"] = 0
        changed = True

    grid_rows = _extract_document_grid_rows(next_document)
    if not grid_rows:
        grid_rows = [list(columns)]
        changed = True
    else:
        grid_rows = [_normalize_sheet_row(row, len(columns)) for row in grid_rows]
        if grid_rows[0] != columns:
            grid_rows[0] = list(columns)
            changed = True
    while len(grid_rows) < data_start_row:
        grid_rows.append([""] * len(columns))
        changed = True
    if next_document.get("grid_rows") != grid_rows:
        next_document["grid_rows"] = grid_rows
        changed = True

    next_document, style_changed = _apply_registration_standard_user_id_column_styles(next_document)
    changed = changed or style_changed
    next_document, order_visibility_changed = _apply_registration_standard_order_column_visibility(next_document)
    changed = changed or order_visibility_changed
    return next_document, changed


def _normalize_registration_sheet_header_persisted(
    session: Session,
    document: SheetDocument,
    document_json: dict[str, Any] | None = None,
) -> dict[str, Any]:
    source_document = dict(document_json if document_json is not None else (document.document_json or {}))
    if not _is_registration_sheet(document):
        return _normalize_document_json(source_document)
    next_document, changed = _normalize_registration_sheet_header_document(source_document)
    if changed:
        document.document_json = next_document
        document.version = max(int(document.version or 1), 1) + 1
        document.updated_at = time.time()
        session.add(document)
        session.commit()
        session.refresh(document)
        _broadcast_sheet_resource_update(document)
    return dict(document.document_json or next_document)


def _apply_registration_standard_user_id_column_styles(document_json: dict[str, Any]) -> tuple[dict[str, Any], bool]:
    normalized = _normalize_document_json(document_json)
    columns = _normalize_document_columns(normalized)
    target_indexes = [
        index
        for index, header in enumerate(columns)
        if header in {"用户ID", NOTE_SHEET_REGISTRATION_LINKED_USER_ID_COLUMN}
    ]
    if not target_indexes:
        return normalized, False

    next_document = dict(normalized)
    changed = False

    column_configs = dict(next_document.get("column_configs")) if isinstance(next_document.get("column_configs"), dict) else {}
    for header in ("用户ID", NOTE_SHEET_REGISTRATION_LINKED_USER_ID_COLUMN):
        if header not in columns:
            continue
        config = dict(column_configs.get(header)) if isinstance(column_configs.get(header), dict) else {}
        previous_config = dict(config)
        config["header_background_color"] = NOTE_SHEET_REGISTRATION_STANDARD_HEADER_BACKGROUND
        config.setdefault("font_family", "monospace")
        config.setdefault("width_mode", "fixed")
        if header == "用户ID":
            config.setdefault("duplicate_value_highlight", True)
        if header == NOTE_SHEET_REGISTRATION_LINKED_USER_ID_COLUMN:
            config.setdefault("note", NOTE_SHEET_REGISTRATION_LINKED_USER_ID_NOTE)
        if config != previous_config:
            changed = True
        column_configs[header] = config
    next_document["column_configs"] = column_configs

    field_row_index = normalized.get("field_row_index")
    header_rows = [0] if not isinstance(field_row_index, int) else [field_row_index]
    if isinstance(field_row_index, int) and field_row_index > 0:
        header_rows.insert(0, 0)

    cell_meta = dict(next_document.get("cell_meta")) if isinstance(next_document.get("cell_meta"), dict) else {}
    for row_index in header_rows:
        if row_index < 0:
            continue
        for column_index in target_indexes:
            key = f"{row_index}:{column_index}"
            meta = dict(cell_meta.get(key)) if isinstance(cell_meta.get(key), dict) else {}
            style = dict(meta.get("style")) if isinstance(meta.get("style"), dict) else {}
            previous_style = dict(style)
            style["background_color"] = NOTE_SHEET_REGISTRATION_STANDARD_HEADER_BACKGROUND
            meta["style"] = style
            if not cell_meta.get(key) or style != previous_style:
                changed = True
            cell_meta[key] = meta
    next_document["cell_meta"] = cell_meta

    entity_rows = _extract_document_entity_rows(next_document)
    entity_columns = _extract_document_entity_columns(next_document)
    entity_cells = _extract_document_entity_cells(next_document)
    if entity_rows and entity_columns and entity_cells:
        target_index_set = set(target_indexes)
        next_entity_cells = dict(entity_cells)
        for row_index in header_rows:
            if row_index < 0 or row_index >= len(entity_rows):
                continue
            row_id = _get_document_entity_row_id(entity_rows[row_index])
            if not row_id:
                continue
            source_row_cells = next_entity_cells.get(row_id)
            next_row_cells = dict(source_row_cells) if isinstance(source_row_cells, dict) else {}
            row_changed = False
            for column_index in target_index_set:
                if column_index < 0 or column_index >= len(entity_columns):
                    continue
                column_id = _get_document_entity_column_id(entity_columns[column_index])
                if not column_id:
                    continue
                source_cell = next_row_cells.get(column_id)
                next_cell = dict(source_cell) if isinstance(source_cell, dict) else {}
                style = dict(next_cell.get("style")) if isinstance(next_cell.get("style"), dict) else {}
                previous_style = dict(style)
                style["background_color"] = NOTE_SHEET_REGISTRATION_STANDARD_HEADER_BACKGROUND
                next_cell["style"] = style
                if next_cell != source_cell or style != previous_style:
                    next_row_cells[column_id] = next_cell
                    row_changed = True
            if row_changed:
                next_entity_cells[row_id] = next_row_cells
                changed = True
        if changed:
            next_document["entity_cells"] = next_entity_cells

    return next_document, changed


def _apply_registration_standard_order_column_visibility(document_json: dict[str, Any]) -> tuple[dict[str, Any], bool]:
    normalized = _normalize_document_json(document_json)
    columns = _normalize_document_columns(normalized)
    if not any(header in columns for header in NOTE_SHEET_REGISTRATION_ORDER_REQUIRED_COLUMNS):
        return normalized, False

    column_configs = dict(normalized.get("column_configs")) if isinstance(normalized.get("column_configs"), dict) else {}
    if not column_configs:
        return normalized, False

    next_configs = dict(column_configs)
    changed = False
    for header in NOTE_SHEET_REGISTRATION_ORDER_REQUIRED_COLUMNS:
        config = next_configs.get(header)
        if not isinstance(config, dict) or config.get("hidden") is not True:
            continue
        next_config = dict(config)
        next_config.pop("hidden", None)
        if next_config:
            next_configs[header] = next_config
        else:
            next_configs.pop(header, None)
        changed = True

    if not changed:
        return normalized, False
    next_document = dict(normalized)
    next_document["column_configs"] = next_configs
    return next_document, True


def _canonicalize_sheet_document_links(
    document_json: dict[str, Any],
    *,
    migrate_legacy_links: bool,
) -> dict[str, Any]:
    normalized, _stats = note_sheet_inline_links.canonicalize_sheet_document_inline_links(
        document_json,
        migrate_legacy_links=migrate_legacy_links,
        strip_legacy_links=True,
    )
    return normalized


def _coerce_inline_cell_object(value: Any) -> dict[str, Any] | None:
    return note_sheet_inline_links.coerce_inline_cell_object(value)


def _inline_cell_link_url(value: Any) -> str:
    return note_sheet_inline_links.inline_cell_link_url(value)


def _with_inline_cell_link(value: Any, url: str) -> Any:
    if not url:
        return value
    return note_sheet_inline_links.with_inline_cell_link(value, {"url": url})


def _is_defined_name_valid(name: str) -> bool:
    text = str(name or "").strip()
    if not text:
        return False
    first_char = text[0]
    if not (first_char == "_" or first_char.isalpha()):
        return False
    if not all(char == "_" or char == "." or char.isalnum() for char in text[1:]):
        return False
    if A1_CELL_REFERENCE_RE.match(text):
        return False
    return not R1C1_REFERENCE_RE.match(text)


def _normalize_defined_name_formula(value: Any) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    return text if text.startswith("=") else f"={text}"


def _normalize_defined_names(
    value: Any,
    *,
    scope: Literal["workbook", "worksheet"] | None = None,
    strict: bool = False,
) -> list[dict[str, Any]]:
    if not isinstance(value, list):
        if strict and value not in (None, ""):
            raise HTTPException(status_code=400, detail="名称管理器数据必须是列表")
        return []

    normalized: list[dict[str, Any]] = []
    seen: set[str] = set()
    for raw_item in value:
        if isinstance(raw_item, NoteSheetDefinedNameItem):
            raw = raw_item.model_dump()
        elif isinstance(raw_item, dict):
            raw = raw_item
        else:
            if strict:
                raise HTTPException(status_code=400, detail="名称管理器条目必须是对象")
            continue

        name = str(raw.get("name") or "").strip()
        formula = _normalize_defined_name_formula(raw.get("formula"))
        comment = str(raw.get("comment") or "").strip()
        if not name and not formula and not comment:
            continue
        if not name:
            if strict:
                raise HTTPException(status_code=400, detail="名称不能为空")
            continue
        if not _is_defined_name_valid(name):
            if strict:
                raise HTTPException(status_code=400, detail=f"名称不合法: {name}")
            continue
        key = name.lower()
        if key in seen:
            if strict:
                raise HTTPException(status_code=400, detail=f"名称重复: {name}")
            continue
        seen.add(key)
        item = {
            "name": name,
            "formula": formula,
            "comment": comment,
        }
        if scope is not None:
            item["scope"] = scope
        normalized.append(item)
    return normalized


def _workbook_defined_names_setting_key(workbook: WorkbookDocument | int) -> str:
    workbook_id = workbook if isinstance(workbook, int) else _require_workbook_numeric_id(workbook)
    return f"{NOTE_SHEET_WORKBOOK_DEFINED_NAMES_SETTING_PREFIX}{int(workbook_id)}"


def _get_workbook_defined_names(session: Session, workbook: WorkbookDocument | None) -> list[dict[str, Any]]:
    if workbook is None:
        return []
    setting = session.get(AppSetting, _workbook_defined_names_setting_key(workbook))
    value = setting.value if setting is not None and isinstance(setting.value, dict) else {}
    return _normalize_attendance_refund_defined_names_for_context(
        _normalize_defined_names(value.get("names"), scope="workbook"),
        context_text=workbook.title,
    )


def _is_attendance_zen_stage_context(text: str) -> bool:
    normalized = _normalize_sheet_text(text)
    return "禅宗" in normalized or "修道班" in normalized


def _normalize_zen_stage_refund_name_formula(name: str, formula: str) -> str:
    normalized_name = name.lower()
    if normalized_name == NOTE_SHEET_ATTENDANCE_REFUND_PERIOD_NAME.lower():
        return "=第几周"
    if normalized_name == "返款id后缀":
        return re.sub(r"_day", "_week", formula, flags=re.I)
    if normalized_name == "返款说明":
        return (
            formula
            .replace('天返款"', '周返款"')
            .replace('&"天"', '&"周"')
            .replace('&"天返款"', '&"周返款"')
        )
    return formula


def _normalize_attendance_refund_defined_names_for_context(
    names: list[dict[str, Any]],
    *,
    context_text: str,
) -> list[dict[str, Any]]:
    if not names or not _is_attendance_zen_stage_context(context_text):
        return names

    normalized: list[dict[str, Any]] = []
    seen: set[str] = set()
    has_day_index_name = any(str(item.get("name") or "").strip() == "第几天" for item in names)
    has_week_index_name = any(str(item.get("name") or "").strip() == "第几周" for item in names)
    inserted_week_index_name = False

    for item in names:
        name = str(item.get("name") or "").strip()
        key = name.lower()

        next_item = dict(item)
        formula = _normalize_defined_name_formula(next_item.get("formula"))
        if formula:
            next_item["formula"] = _normalize_zen_stage_refund_name_formula(name, formula)
        normalized.append(next_item)
        seen.add(key)

        if name == "第几天" and not has_week_index_name:
            week_item = {
                "name": "第几周",
                "formula": "=INT((第几天-1)/7)+1",
                "comment": "禅宗修道班按周返款",
            }
            if item.get("scope"):
                week_item["scope"] = item.get("scope")
            normalized.append(week_item)
            seen.add("第几周".lower())
            inserted_week_index_name = True

    if has_day_index_name and not has_week_index_name and not inserted_week_index_name:
        normalized.append({
            "name": "第几周",
            "formula": "=INT((第几天-1)/7)+1",
            "comment": "禅宗修道班按周返款",
        })
    return normalized


def _set_workbook_defined_names(
    session: Session,
    workbook: WorkbookDocument,
    names: list[dict[str, Any]],
) -> None:
    key = _workbook_defined_names_setting_key(workbook)
    setting = session.get(AppSetting, key)
    normalized = _normalize_defined_names(names, strict=True)
    if not normalized:
        if setting is not None:
            session.delete(setting)
        return
    if setting is None:
        setting = AppSetting(key=key)
    setting.value = {"names": normalized}
    setting.updated_at = time.time()
    session.add(setting)


def _get_sheet_defined_names(document_json: dict[str, Any]) -> list[dict[str, Any]]:
    return _normalize_defined_names(document_json.get(NOTE_SHEET_DEFINED_NAMES_KEY), scope="worksheet")


def _list_workbook_defined_name_worksheets(
    session: Session,
    workbook: WorkbookDocument,
    current_user: User | None = None,
) -> list[dict[str, Any]]:
    links = session.exec(
        select(WorkbookSheetLink)
        .where(WorkbookSheetLink.workbook_id.in_(workbook_ref_aliases(workbook)))
        .order_by(WorkbookSheetLink.order_index, WorkbookSheetLink.created_at)
    ).all()
    sheet_ids = [link.sheet_id for link in links]
    sheet_map = load_sheets_by_refs(session, sheet_ids)
    result: list[dict[str, Any]] = []
    seen_sheet_ids: set[str] = set()
    for link in links:
        document = sheet_map.get(str(link.sheet_id))
        if document is None:
            continue
        document_key = str(document.id)
        if document_key in seen_sheet_ids:
            continue
        seen_sheet_ids.add(document_key)
        access = _resolve_sheet_resource_access(session, document, current_user, workbook=workbook)
        if not access.capabilities.can_read:
            continue
        result.append({
            "sheet_id": _require_sheet_numeric_id(document),
            "sheet_title": document.title or "未命名工作表",
            "sheet_version": int(document.version or 1),
            "names": _get_sheet_defined_names(dict(document.document_json or {})),
        })
    return result


def _replace_sheet_defined_names(document_json: dict[str, Any], names: list[dict[str, Any]]) -> dict[str, Any]:
    next_document = dict(document_json)
    normalized = _normalize_defined_names(names, strict=True)
    if normalized:
        next_document[NOTE_SHEET_DEFINED_NAMES_KEY] = normalized
    else:
        next_document.pop(NOTE_SHEET_DEFINED_NAMES_KEY, None)
    return next_document


def _merge_effective_defined_names(
    workbook_names: list[dict[str, Any]],
    worksheet_names: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    by_key: dict[str, dict[str, Any]] = {}
    order: list[str] = []
    for item in [*workbook_names, *worksheet_names]:
        name = str(item.get("name") or "").strip()
        if not name:
            continue
        key = name.lower()
        if key not in by_key:
            order.append(key)
        by_key[key] = dict(item)
    return [by_key[key] for key in order]


def _defined_names_for_formula(
    session: Session,
    document: SheetDocument,
    workbook: WorkbookDocument | None,
) -> dict[str, str]:
    names: dict[str, str] = {}
    context_text = " ".join(
        item
        for item in [
            _normalize_sheet_text(workbook.title if workbook is not None else ""),
            _normalize_sheet_text(document.title),
        ]
        if item
    )
    effective_names = _merge_effective_defined_names(
        _get_workbook_defined_names(session, workbook),
        _get_sheet_defined_names(dict(document.document_json or {})),
    )
    for item in _normalize_attendance_refund_defined_names_for_context(
        effective_names,
        context_text=context_text,
    ):
        name = str(item.get("name") or "").strip()
        formula = _normalize_defined_name_formula(item.get("formula"))
        if name and formula:
            names[name.lower()] = formula
    return names


def _get_note_sheet_perf_log_path() -> Path:
    return get_settings().data_dir / NOTE_SHEET_PERF_LOG_DIRNAME / NOTE_SHEET_PERF_LOG_FILENAME


def _write_note_sheet_perf_log_batch(
    payload: NoteSheetPerfLogRequest,
    current_user: User | None,
) -> NoteSheetPerfLogResponse:
    entries = payload.entries[:NOTE_SHEET_PERF_LOG_MAX_ENTRIES_PER_BATCH]
    log_path = _get_note_sheet_perf_log_path()
    if not entries:
        return NoteSheetPerfLogResponse(stored_count=0, path=os.fspath(log_path))

    record = {
        "received_at": time.time(),
        "session_id": payload.session_id,
        "url": payload.url,
        "user_agent": payload.user_agent,
        "workbook_id": payload.workbook_id,
        "sheet_id": payload.sheet_id,
        "user_id": current_user.id if current_user else None,
        "entries": entries,
    }
    log_path.parent.mkdir(parents=True, exist_ok=True)
    line = json.dumps(record, ensure_ascii=False, default=str)
    with _note_sheet_perf_log_lock:
        with log_path.open("a", encoding="utf-8") as file:
            file.write(line)
            file.write("\n")
    return NoteSheetPerfLogResponse(stored_count=len(entries), path=os.fspath(log_path))


def _normalize_document_data_start_row(document_json: dict[str, Any]) -> int:
    raw_value = document_json.get("data_start_row")
    try:
        numeric = int(raw_value)
    except (TypeError, ValueError):
        return 0
    return max(numeric, 0)


def _uses_sheet_formula_reference_origin(document_json: dict[str, Any]) -> bool:
    return document_json.get("formula_reference_origin") == "sheet_v2"


def _get_formula_reference_row_offset(document_json: dict[str, Any]) -> int:
    return _normalize_document_data_start_row(document_json) if _uses_sheet_formula_reference_origin(document_json) else 0


def _extract_document_grid_rows(document_json: dict[str, Any]) -> list[Any]:
    grid_rows = document_json.get("grid_rows")
    return list(grid_rows) if isinstance(grid_rows, list) else []


def _replace_document_data_rows(document_json: dict[str, Any], rows: list[Any]) -> dict[str, Any]:
    next_document = dict(document_json)
    next_document["rows"] = rows
    grid_rows = _extract_document_grid_rows(next_document)
    if grid_rows:
        data_start_row = min(_normalize_document_data_start_row(next_document), len(grid_rows))
        next_document["grid_rows"] = [*grid_rows[:data_start_row], *rows]
    return next_document


def _normalize_document_columns(document_json: dict[str, Any]) -> list[str]:
    columns = document_json.get("columns")
    if not isinstance(columns, list):
        return list(DEFAULT_NOTE_SHEET_COLUMNS)
    normalized = [str(column or "").strip() for column in columns]
    return normalized or list(DEFAULT_NOTE_SHEET_COLUMNS)


def _row_has_meaningful_cell(row: Any) -> bool:
    if isinstance(row, list):
        return any(str(cell or "").strip() for cell in row)
    if isinstance(row, dict):
        return any(str(cell or "").strip() for cell in row.values())
    return bool(str(row or "").strip())


def _document_has_structural_customization(document_json: dict[str, Any]) -> bool:
    for key in (
        "header_groups",
        "cell_meta",
        "entity_columns",
        "entity_rows",
        "entity_cells",
        "column_configs",
        "grid_rows",
        "merged_cells",
        "data_start_row",
        "field_row_index",
    ):
        value = document_json.get(key)
        if key == "grid_rows":
            columns = _normalize_document_columns(document_json)
            if isinstance(value, list) and value and value != [columns]:
                return True
            continue
        if key == "data_start_row":
            if value not in (None, 1):
                return True
            continue
        if key == "field_row_index":
            if value not in (None, 0):
                return True
            continue
        if isinstance(value, dict) and value:
            return True
        if isinstance(value, list) and value:
            return True
    return False


def _document_has_meaningful_content(document_json: dict[str, Any]) -> bool:
    normalized = _normalize_document_json(document_json)
    if _normalize_document_columns(normalized) != DEFAULT_NOTE_SHEET_COLUMNS:
        return True
    if any(_row_has_meaningful_cell(row) for row in _extract_document_rows(normalized)):
        return True
    return _document_has_structural_customization(normalized)


def _is_default_blank_document(document_json: dict[str, Any]) -> bool:
    normalized = _normalize_document_json(document_json)
    return (
        _normalize_document_columns(normalized) == DEFAULT_NOTE_SHEET_COLUMNS
        and not any(_row_has_meaningful_cell(row) for row in _extract_document_rows(normalized))
        and not _document_has_structural_customization(normalized)
    )


def _reject_default_blank_overwrite(current_document: dict[str, Any], incoming_document: dict[str, Any]) -> None:
    if _document_has_meaningful_content(current_document) and _is_default_blank_document(incoming_document):
        raise HTTPException(status_code=409, detail="拒绝使用默认空表覆盖已有表格数据")


def _normalize_page_size(value: int | None) -> int:
    numeric = int(value or DEFAULT_NOTE_SHEET_PAGE_SIZE)
    return min(max(numeric, 1), MAX_NOTE_SHEET_PAGE_SIZE)


def _extract_document_rows(document_json: dict[str, Any]) -> list[Any]:
    rows = document_json.get("rows")
    return list(rows) if isinstance(rows, list) else []


def _new_sheet_row_id() -> str:
    return f"row_{uuid.uuid4().hex[:12]}"


def _new_sheet_column_id() -> str:
    return f"col_{uuid.uuid4().hex[:12]}"


def _normalize_sheet_identity_list(source: Any, *, count: int, prefix: str) -> tuple[list[str], bool]:
    raw_values = list(source) if isinstance(source, list) else []
    ids: list[str] = []
    seen: set[str] = set()
    changed = not isinstance(source, list) or len(raw_values) != count
    generator = _new_sheet_row_id if prefix == "row" else _new_sheet_column_id

    for index in range(count):
        candidate = str(raw_values[index]).strip() if index < len(raw_values) else ""
        if not candidate or candidate in seen:
            candidate = generator()
            changed = True
        seen.add(candidate)
        ids.append(candidate)
    return ids, changed


def _ensure_sheet_document_identity(document_json: dict[str, Any]) -> tuple[dict[str, Any], bool]:
    normalized = _normalize_document_json(document_json)
    columns = _normalize_document_columns(normalized)
    rows = _extract_document_rows(normalized)
    row_ids, row_changed = _normalize_sheet_identity_list(normalized.get("row_ids"), count=len(rows), prefix="row")
    column_ids, column_changed = _normalize_sheet_identity_list(
        normalized.get("column_ids"),
        count=len(columns),
        prefix="column",
    )
    if not row_changed and not column_changed:
        return normalized, False
    return {
        **normalized,
        "row_ids": row_ids,
        "column_ids": column_ids,
    }, True


def _ensure_sheet_document_identity_persisted(session: Session, document: SheetDocument) -> dict[str, Any]:
    next_document, changed = _ensure_sheet_document_identity(dict(document.document_json or {}))
    if changed:
        document.document_json = next_document
        document.updated_at = time.time()
        session.add(document)
        session.commit()
        session.refresh(document)
        return dict(document.document_json or {})
    return next_document


def _find_sheet_identity_index(ids: list[str], value: str | None, *, detail: str) -> int:
    needle = str(value or "").strip()
    if not needle:
        raise HTTPException(status_code=400, detail=detail)
    try:
        return ids.index(needle)
    except ValueError:
        raise HTTPException(status_code=400, detail=f"{detail}不存在") from None


def _find_sheet_identity_insert_index(ids: list[str], after_id: str | None) -> int:
    if after_id is None or str(after_id).strip() == "":
        return 0
    return _find_sheet_identity_index(ids, after_id, detail="目标位置") + 1


def _coerce_sheet_row_payload(row: Any, *, column_count: int) -> Any:
    if isinstance(row, dict):
        return dict(row)
    if isinstance(row, list):
        return _normalize_sheet_row(row, column_count)
    return [""] * column_count


def _coerce_sheet_column_header(column: Any, *, existing_columns: list[str]) -> str:
    if isinstance(column, dict):
        raw_header = column.get("header") or column.get("name") or column.get("title") or column.get("key")
    else:
        raw_header = column
    base = str(raw_header or "新列").strip() or "新列"
    if base not in existing_columns:
        return base
    suffix = 2
    while f"{base}{suffix}" in existing_columns:
        suffix += 1
    return f"{base}{suffix}"


def _extract_sheet_column_width(column: Any) -> int:
    if isinstance(column, dict):
        try:
            width = float(column.get("width"))
        except (TypeError, ValueError):
            width = 96
        if isfinite(width) and width > 0:
            return int(round(width))
    return 96


def _extract_sheet_column_config(column: Any) -> dict[str, Any]:
    if not isinstance(column, dict):
        return {}
    config = column.get("config")
    return dict(config) if isinstance(config, dict) else {}


def _shift_merged_cell_rows_for_insert(merged_cells: Any, insert_index: int, amount: int, *, row_offset: int = 0) -> list[Any]:
    if not isinstance(merged_cells, list) or amount <= 0:
        return list(merged_cells) if isinstance(merged_cells, list) else []
    effective_insert_index = insert_index + row_offset
    shifted: list[Any] = []
    for cell in merged_cells:
        normalized = dict(cell) if isinstance(cell, dict) else None
        if normalized is None:
            continue
        row = int(normalized.get("row") or 0)
        col = int(normalized.get("col") or 0)
        rowspan = max(int(normalized.get("rowspan") or 1), 1)
        colspan = max(int(normalized.get("colspan") or 1), 1)
        if row < effective_insert_index < row + rowspan:
            rowspan += amount
        elif row >= effective_insert_index:
            row += amount
        if rowspan > 1 or colspan > 1:
            shifted.append({"row": row, "col": col, "rowspan": rowspan, "colspan": colspan})
    return shifted


def _shift_merged_cell_rows_for_delete(merged_cells: Any, delete_index: int, *, row_offset: int = 0) -> list[Any]:
    if not isinstance(merged_cells, list):
        return []
    effective_delete_index = delete_index + row_offset
    shifted: list[Any] = []
    for cell in merged_cells:
        if not isinstance(cell, dict):
            continue
        row = int(cell.get("row") or 0)
        col = int(cell.get("col") or 0)
        rowspan = max(int(cell.get("rowspan") or 1), 1)
        colspan = max(int(cell.get("colspan") or 1), 1)
        if row <= effective_delete_index < row + rowspan:
            rowspan -= 1
        elif row > effective_delete_index:
            row -= 1
        if rowspan > 1 or colspan > 1:
            shifted.append({"row": row, "col": col, "rowspan": rowspan, "colspan": colspan})
    return shifted


def _shift_cell_meta_rows_for_delete(cell_meta: Any, delete_index: int, *, row_offset: int = 0) -> dict[str, Any]:
    if not isinstance(cell_meta, dict):
        return {}
    shifted: dict[str, Any] = {}
    effective_delete_index = delete_index + row_offset
    for key, meta in cell_meta.items():
        parsed = _parse_cell_meta_key(key)
        if parsed is None:
            shifted[str(key)] = meta
            continue
        row_index, column_index = parsed
        if row_index == effective_delete_index:
            continue
        next_row_index = row_index - 1 if row_index > effective_delete_index else row_index
        shifted[f"{next_row_index}:{column_index}"] = meta
    return shifted


def _remap_cell_meta_columns(cell_meta: Any, column_index_map: dict[int, int | None]) -> dict[str, Any]:
    if not isinstance(cell_meta, dict):
        return {}
    remapped: dict[str, Any] = {}
    for key, meta in cell_meta.items():
        parsed = _parse_cell_meta_key(key)
        if parsed is None:
            remapped[str(key)] = meta
            continue
        row_index, column_index = parsed
        next_column_index = column_index_map.get(column_index)
        if next_column_index is None:
            continue
        remapped[f"{row_index}:{next_column_index}"] = meta
    return remapped


def _move_list_item(values: list[Any], source_index: int, insert_index: int) -> list[Any]:
    next_values = list(values)
    item = next_values.pop(source_index)
    adjusted_insert_index = insert_index - 1 if source_index < insert_index else insert_index
    next_values.insert(min(max(adjusted_insert_index, 0), len(next_values)), item)
    return next_values


def _sheet_patch_data_row_meta_key(document_json: dict[str, Any], row_index: int, column_index: int) -> str:
    return f"{_normalize_document_data_start_row(document_json) + row_index}:{column_index}"


def _sheet_patch_require_int(value: Any, detail: str) -> int:
    if value is None:
        raise HTTPException(status_code=400, detail=detail)
    try:
        return int(value)
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail=detail) from None


def _sheet_patch_validate_data_cell(
    *,
    row_index: int,
    column_index: int,
    rows: list[Any],
    columns: list[str],
) -> None:
    if column_index < 0 or column_index >= len(columns):
        raise HTTPException(status_code=400, detail="列号超出表格范围")
    if row_index < 0 or row_index >= len(rows):
        raise HTTPException(status_code=400, detail="行号超出表格范围")


def _sheet_patch_validate_grid_range(
    *,
    row: int,
    col: int,
    rowspan: int,
    colspan: int,
    row_count: int,
    column_count: int,
) -> None:
    if row < 0 or col < 0 or row >= row_count or col >= column_count:
        raise HTTPException(status_code=400, detail="合并范围超出表格范围")
    if rowspan <= 1 and colspan <= 1:
        raise HTTPException(status_code=400, detail="合并范围至少需要跨两格")
    if row + rowspan > row_count or col + colspan > column_count:
        raise HTTPException(status_code=400, detail="合并范围超出表格范围")


def _sheet_ranges_overlap(left: dict[str, int], right: dict[str, int]) -> bool:
    left_row_end = int(left["row"]) + int(left["rowspan"])
    left_col_end = int(left["col"]) + int(left["colspan"])
    right_row_end = int(right["row"]) + int(right["rowspan"])
    right_col_end = int(right["col"]) + int(right["colspan"])
    return (
        int(left["row"]) < right_row_end
        and int(right["row"]) < left_row_end
        and int(left["col"]) < right_col_end
        and int(right["col"]) < left_col_end
    )


def _normalize_sheet_merged_cell(item: Any, *, row_count: int, column_count: int) -> dict[str, int] | None:
    if not isinstance(item, dict):
        return None
    try:
        row = int(item.get("row"))
        col = int(item.get("col"))
        rowspan = int(item.get("rowspan") or 1)
        colspan = int(item.get("colspan") or 1)
    except (TypeError, ValueError):
        return None
    if row < 0 or col < 0 or row >= row_count or col >= column_count or (rowspan <= 1 and colspan <= 1):
        return None
    if row + rowspan > row_count or col + colspan > column_count:
        return None
    return {"row": row, "col": col, "rowspan": rowspan, "colspan": colspan}


def _normalize_sheet_merged_cells(source: Any, *, row_count: int, column_count: int) -> list[dict[str, int]]:
    if not isinstance(source, list):
        return []
    cells: list[dict[str, int]] = []
    for item in source:
        cell = _normalize_sheet_merged_cell(item, row_count=row_count, column_count=column_count)
        if cell is None:
            continue
        if any(_sheet_ranges_overlap(cell, existing) for existing in cells):
            continue
        cells.append(cell)
    return cells


def _apply_note_sheet_patch_ops(document_json: dict[str, Any], ops: list[NoteSheetPatchOperation]) -> tuple[dict[str, Any], int]:
    next_document, _identity_changed = _ensure_sheet_document_identity(deepcopy(_normalize_document_json(document_json)))
    columns = _normalize_document_columns(next_document)
    rows = _extract_document_rows(next_document)
    row_ids = list(next_document.get("row_ids")) if isinstance(next_document.get("row_ids"), list) else []
    column_ids = list(next_document.get("column_ids")) if isinstance(next_document.get("column_ids"), list) else []
    grid_rows = _extract_document_grid_rows(next_document)
    row_count = max(len(grid_rows), _normalize_document_data_start_row(next_document) + len(rows))
    column_count = len(columns)
    updated_cell_count = 0

    for operation in ops:
        if operation.op == "set-cell-value":
            row_index = _sheet_patch_require_int(operation.row_index, "缺少行号")
            column_index = _sheet_patch_require_int(operation.column_index, "缺少列号")
            _sheet_patch_validate_data_cell(row_index=row_index, column_index=column_index, rows=rows, columns=columns)
            current_row = rows[row_index]
            current_cells = _normalize_sheet_row(current_row, len(columns))
            if _normalize_restricted_cell_value(current_cells[column_index]) == _normalize_restricted_cell_value(operation.value):
                continue
            rows[row_index] = _set_row_cell_value(current_row, columns, column_index, operation.value)
            next_document = _replace_document_data_rows(next_document, rows)
            grid_rows = _extract_document_grid_rows(next_document)
            row_count = max(len(grid_rows), _normalize_document_data_start_row(next_document) + len(rows))
            updated_cell_count += 1
            continue

        if operation.op == "set-cell-meta":
            row_index = _sheet_patch_require_int(operation.row_index, "缺少行号")
            column_index = _sheet_patch_require_int(operation.column_index, "缺少列号")
            _sheet_patch_validate_data_cell(row_index=row_index, column_index=column_index, rows=rows, columns=columns)
            cell_meta = _extract_document_cell_meta(next_document)
            key = _sheet_patch_data_row_meta_key(next_document, row_index, column_index)
            meta = dict(operation.meta or {})
            if meta:
                cell_meta[key] = meta
            else:
                cell_meta.pop(key, None)
            if cell_meta:
                next_document["cell_meta"] = cell_meta
            else:
                next_document.pop("cell_meta", None)
            continue

        if operation.op == "set-column-width":
            column_index = _sheet_patch_require_int(operation.column_index, "缺少列号")
            if column_index < 0 or column_index >= column_count:
                raise HTTPException(status_code=400, detail="列号超出表格范围")
            try:
                width = float(operation.width)
            except (TypeError, ValueError):
                raise HTTPException(status_code=400, detail="列宽无效") from None
            if not isfinite(width) or width <= 0:
                raise HTTPException(status_code=400, detail="列宽无效")
            widths = list(next_document.get("column_widths")) if isinstance(next_document.get("column_widths"), list) else []
            if len(widths) < column_count:
                widths.extend([None] * (column_count - len(widths)))
            widths[column_index] = int(round(width))
            next_document["column_widths"] = widths[:column_count]
            continue

        if operation.op in {"set-column-hidden", "set-column-config"}:
            column_index = _sheet_patch_require_int(operation.column_index, "缺少列号")
            if column_index < 0 or column_index >= column_count:
                raise HTTPException(status_code=400, detail="列号超出表格范围")
            header = columns[column_index]
            column_configs = dict(next_document.get("column_configs")) if isinstance(next_document.get("column_configs"), dict) else {}
            current_config = dict(column_configs.get(header)) if isinstance(column_configs.get(header), dict) else {}
            if operation.op == "set-column-hidden":
                if operation.hidden is None:
                    raise HTTPException(status_code=400, detail="缺少隐藏状态")
                if operation.hidden:
                    current_config["hidden"] = True
                else:
                    current_config.pop("hidden", None)
                    current_config.pop("restore_index", None)
            else:
                current_config = dict(operation.config or {})
            if current_config:
                column_configs[header] = current_config
            else:
                column_configs.pop(header, None)
            if column_configs:
                next_document["column_configs"] = column_configs
            else:
                next_document.pop("column_configs", None)
            continue

        if operation.op == "merge-cells":
            row = _sheet_patch_require_int(operation.row, "缺少合并起始行")
            col = _sheet_patch_require_int(operation.col, "缺少合并起始列")
            rowspan = _sheet_patch_require_int(operation.rowspan, "缺少合并行数")
            colspan = _sheet_patch_require_int(operation.colspan, "缺少合并列数")
            _sheet_patch_validate_grid_range(row=row, col=col, rowspan=rowspan, colspan=colspan, row_count=row_count, column_count=column_count)
            next_cell = {"row": row, "col": col, "rowspan": rowspan, "colspan": colspan}
            merged_cells = _normalize_sheet_merged_cells(next_document.get("merged_cells"), row_count=row_count, column_count=column_count)
            if any(_sheet_ranges_overlap(next_cell, existing) for existing in merged_cells):
                raise HTTPException(status_code=400, detail="合并范围与已有合并单元格重叠")
            next_document["merged_cells"] = [*merged_cells, next_cell]
            continue

        if operation.op == "unmerge-cells":
            row = _sheet_patch_require_int(operation.row, "缺少取消合并行")
            col = _sheet_patch_require_int(operation.col, "缺少取消合并列")
            merged_cells = _normalize_sheet_merged_cells(next_document.get("merged_cells"), row_count=row_count, column_count=column_count)
            next_document["merged_cells"] = [
                cell
                for cell in merged_cells
                if not (
                    int(cell["row"]) <= row < int(cell["row"]) + int(cell["rowspan"])
                    and int(cell["col"]) <= col < int(cell["col"]) + int(cell["colspan"])
                )
            ]
            continue

        if operation.op == "insert-row":
            insert_index = _find_sheet_identity_insert_index(row_ids, operation.after_row_id)
            next_row = _coerce_sheet_row_payload(operation.row, column_count=column_count)
            next_row_id = str(operation.row_id or "").strip() or _new_sheet_row_id()
            if next_row_id in row_ids:
                raise HTTPException(status_code=400, detail="row_id 已存在")
            rows = [*rows[:insert_index], next_row, *rows[insert_index:]]
            row_ids = [*row_ids[:insert_index], next_row_id, *row_ids[insert_index:]]
            next_document = _replace_document_data_rows(next_document, rows)
            next_document["row_ids"] = row_ids
            data_start_row = _normalize_document_data_start_row(next_document)
            if "cell_meta" in next_document:
                next_document["cell_meta"] = _shift_cell_meta_rows_for_insert(
                    next_document.get("cell_meta"),
                    insert_index,
                    1,
                    row_offset=data_start_row,
                )
            if "merged_cells" in next_document:
                next_document["merged_cells"] = _shift_merged_cell_rows_for_insert(
                    next_document.get("merged_cells"),
                    insert_index,
                    1,
                    row_offset=data_start_row,
                )
            grid_rows = _extract_document_grid_rows(next_document)
            row_count = max(len(grid_rows), data_start_row + len(rows))
            continue

        if operation.op == "delete-row":
            delete_index = _find_sheet_identity_index(row_ids, operation.row_id, detail="行 ID")
            rows = [*rows[:delete_index], *rows[delete_index + 1:]]
            row_ids = [*row_ids[:delete_index], *row_ids[delete_index + 1:]]
            next_document = _replace_document_data_rows(next_document, rows)
            next_document["row_ids"] = row_ids
            data_start_row = _normalize_document_data_start_row(next_document)
            if "cell_meta" in next_document:
                next_document["cell_meta"] = _shift_cell_meta_rows_for_delete(
                    next_document.get("cell_meta"),
                    delete_index,
                    row_offset=data_start_row,
                )
            if "merged_cells" in next_document:
                next_document["merged_cells"] = _shift_merged_cell_rows_for_delete(
                    next_document.get("merged_cells"),
                    delete_index,
                    row_offset=data_start_row,
                )
            grid_rows = _extract_document_grid_rows(next_document)
            row_count = max(len(grid_rows), data_start_row + len(rows))
            continue

        if operation.op == "move-row":
            if _normalize_sheet_merged_cells(next_document.get("merged_cells"), row_count=row_count, column_count=column_count):
                raise HTTPException(status_code=400, detail="存在合并单元格时暂不支持移动行")
            source_index = _find_sheet_identity_index(row_ids, operation.row_id, detail="行 ID")
            insert_index = _find_sheet_identity_insert_index(row_ids, operation.after_row_id)
            if insert_index == source_index or insert_index == source_index + 1:
                continue
            order = _move_list_item(list(range(len(rows))), source_index, insert_index)
            row_index_map = {old_index: new_index for new_index, old_index in enumerate(order)}
            rows = [rows[old_index] for old_index in order]
            row_ids = [row_ids[old_index] for old_index in order]
            next_document = _replace_document_data_rows(next_document, rows)
            next_document["row_ids"] = row_ids
            if "cell_meta" in next_document:
                next_document["cell_meta"] = _remap_cell_meta_rows(
                    next_document.get("cell_meta"),
                    row_index_map,
                    row_offset=_normalize_document_data_start_row(next_document),
                )
            grid_rows = _extract_document_grid_rows(next_document)
            row_count = max(len(grid_rows), _normalize_document_data_start_row(next_document) + len(rows))
            continue

        if operation.op == "insert-column":
            insert_index = _find_sheet_identity_insert_index(column_ids, operation.after_column_id)
            header = _coerce_sheet_column_header(operation.column, existing_columns=columns)
            next_column_id = str(operation.column_id or "").strip() or _new_sheet_column_id()
            if next_column_id in column_ids:
                raise HTTPException(status_code=400, detail="column_id 已存在")
            next_document = _insert_document_column(
                next_document,
                insert_index=insert_index,
                header=header,
                width=_extract_sheet_column_width(operation.column),
            )
            column_ids = [*column_ids[:insert_index], next_column_id, *column_ids[insert_index:]]
            next_document["column_ids"] = column_ids
            config = _extract_sheet_column_config(operation.column)
            if config:
                column_configs = dict(next_document.get("column_configs")) if isinstance(next_document.get("column_configs"), dict) else {}
                column_configs[header] = config
                next_document["column_configs"] = column_configs
            columns = _normalize_document_columns(next_document)
            rows = _extract_document_rows(next_document)
            grid_rows = _extract_document_grid_rows(next_document)
            column_count = len(columns)
            row_count = max(len(grid_rows), _normalize_document_data_start_row(next_document) + len(rows))
            continue

        if operation.op == "delete-column":
            delete_index = _find_sheet_identity_index(column_ids, operation.column_id, detail="列 ID")
            if len(columns) <= 1:
                raise HTTPException(status_code=400, detail="至少保留一列")
            next_document = _delete_document_column(next_document, delete_index=delete_index)
            column_ids = [*column_ids[:delete_index], *column_ids[delete_index + 1:]]
            next_document["column_ids"] = column_ids
            columns = _normalize_document_columns(next_document)
            rows = _extract_document_rows(next_document)
            grid_rows = _extract_document_grid_rows(next_document)
            column_count = len(columns)
            row_count = max(len(grid_rows), _normalize_document_data_start_row(next_document) + len(rows))
            continue

        if operation.op == "move-column":
            if _normalize_sheet_merged_cells(next_document.get("merged_cells"), row_count=row_count, column_count=column_count):
                raise HTTPException(status_code=400, detail="存在合并单元格时暂不支持移动列")
            source_index = _find_sheet_identity_index(column_ids, operation.column_id, detail="列 ID")
            insert_index = _find_sheet_identity_insert_index(column_ids, operation.after_column_id)
            if insert_index == source_index or insert_index == source_index + 1:
                continue
            order = _move_list_item(list(range(column_count)), source_index, insert_index)
            column_index_map = {old_index: new_index for new_index, old_index in enumerate(order)}
            next_columns = [columns[old_index] for old_index in order]
            original_grid_rows = _extract_document_grid_rows(next_document)
            next_rows: list[Any] = []
            for row in rows:
                remapped_row = _remap_row_formula_cell_references(row, columns=columns, column_index_map=column_index_map)
                normalized_row = _normalize_sheet_row(remapped_row, column_count)
                next_rows.append([normalized_row[old_index] for old_index in order])
            next_document = {
                **next_document,
                "columns": next_columns,
            }
            next_document = _replace_document_data_rows(next_document, next_rows)
            next_document["column_ids"] = [column_ids[old_index] for old_index in order]
            source_widths = next_document.get("column_widths")
            if isinstance(source_widths, list):
                widths = [*source_widths, *([None] * max(column_count - len(source_widths), 0))]
                next_document["column_widths"] = [widths[old_index] for old_index in order]
            if original_grid_rows:
                next_document["grid_rows"] = [
                    [_normalize_sheet_row(_remap_row_formula_cell_references(row, columns=columns, column_index_map=column_index_map), column_count)[old_index] for old_index in order]
                    for row in original_grid_rows
                ]
            if "cell_meta" in next_document:
                next_document["cell_meta"] = _remap_cell_meta_columns(next_document.get("cell_meta"), column_index_map)
            columns = _normalize_document_columns(next_document)
            rows = _extract_document_rows(next_document)
            column_ids = list(next_document.get("column_ids")) if isinstance(next_document.get("column_ids"), list) else []
            grid_rows = _extract_document_grid_rows(next_document)
            column_count = len(columns)
            row_count = max(len(grid_rows), _normalize_document_data_start_row(next_document) + len(rows))
            continue

    return next_document, updated_cell_count


def _validate_note_sheet_patch_access(
    *,
    access: NoteSheetResourceAccess,
    current_user: User | None,
    operations: list[NoteSheetPatchOperation],
    columns: list[str],
    rows: list[Any],
) -> None:
    editable_columns = {
        int(index)
        for index in access.capabilities.editable_data_columns
        if 0 <= int(index) < len(columns)
    }
    can_edit_all_data = bool(access.capabilities.can_edit_data)
    if not can_edit_all_data and not editable_columns:
        raise HTTPException(status_code=403, detail="没有该资源权限")
    if can_edit_all_data and current_user is None:
        raise HTTPException(status_code=403, detail="没有该资源权限")

    for operation in operations:
        if operation.op == "set-cell-value":
            row_index = _sheet_patch_require_int(operation.row_index, "缺少行号")
            column_index = _sheet_patch_require_int(operation.column_index, "缺少列号")
            _sheet_patch_validate_data_cell(row_index=row_index, column_index=column_index, rows=rows, columns=columns)
            if not can_edit_all_data and column_index not in editable_columns:
                raise HTTPException(status_code=403, detail="游客只能编辑开放列")
            continue

        if not can_edit_all_data:
            raise HTTPException(status_code=403, detail="没有该资源权限")


def _get_document_pagination_settings(document_json: dict[str, Any]) -> tuple[bool, int]:
    normalized = _normalize_document_json(document_json)
    return _get_normalized_document_pagination_settings(normalized)


def _get_normalized_document_pagination_settings(normalized: dict[str, Any]) -> tuple[bool, int]:
    view_settings = normalized.get("view_settings")
    if not isinstance(view_settings, dict):
        return False, DEFAULT_NOTE_SHEET_PAGE_SIZE

    pagination = view_settings.get("pagination")
    if not isinstance(pagination, dict):
        return False, DEFAULT_NOTE_SHEET_PAGE_SIZE

    enabled = pagination.get("enabled") is True
    page_size = _normalize_page_size(pagination.get("page_size"))
    return enabled, page_size


def _is_attendance_questionnaire_data_sheet(document: SheetDocument) -> bool:
    return (
        document.owner_type == ATTENDANCE_WJX_DATA_OWNER_TYPE
        and document.owner_key == ATTENDANCE_WJX_DATA_OWNER_KEY
        and document.sheet_key == ATTENDANCE_WJX_DATA_SHEET_KEY
    )


def _sync_attendance_questionnaire_course_links(
    session: Session,
    document_json: dict[str, Any],
) -> tuple[dict[str, Any], bool]:
    from backend.api.attendance import _sync_attendance_wjx_sheet_course_links

    return _sync_attendance_wjx_sheet_course_links(session, document_json)


def _sync_attendance_questionnaire_entry_statuses(
    session: Session,
    document_json: dict[str, Any],
) -> int:
    from backend.api.attendance import _sync_attendance_wjx_entries_from_sheet_document

    return _sync_attendance_wjx_entries_from_sheet_document(session, document_json)


_attendance_questionnaire_sync_cache_lock = threading.Lock()
_attendance_questionnaire_sync_cache: dict[str, tuple[Any, ...]] = {}


def _get_attendance_questionnaire_sync_signature(
    session: Session,
    document: SheetDocument,
) -> tuple[Any, ...] | None:
    if not _is_attendance_questionnaire_data_sheet(document):
        return None

    from backend.api.attendance import (
        FEEDBACK_COURSE_SOURCE_SHEET_ID,
        FIXED_WJX_TEMPLATE_ACTIVITY_ID,
        LOCAL_FEEDBACK_ACTIVITY_ID,
    )
    from backend.models import AttendanceWjxDataEntry

    activity_ids = [FIXED_WJX_TEMPLATE_ACTIVITY_ID, LOCAL_FEEDBACK_ACTIVITY_ID]
    entry_stats = session.exec(
        select(
            func.count(AttendanceWjxDataEntry.id),
            func.max(AttendanceWjxDataEntry.updated_at),
            func.max(AttendanceWjxDataEntry.synced_at),
        ).where(AttendanceWjxDataEntry.activity_id.in_(activity_ids))
    ).one()
    course_source_stats = session.exec(
        select(SheetDocument.version, SheetDocument.updated_at)
        .where(SheetDocument.numeric_id == FEEDBACK_COURSE_SOURCE_SHEET_ID)
    ).first()
    return (
        str(document.id),
        int(document.version or 0),
        float(document.updated_at or 0.0),
        int(entry_stats[0] or 0),
        float(entry_stats[1] or 0.0),
        float(entry_stats[2] or 0.0),
        int(course_source_stats[0] or 0) if course_source_stats else 0,
        float(course_source_stats[1] or 0.0) if course_source_stats else 0.0,
    )


def _sync_attendance_questionnaire_sheet_document_if_needed(
    session: Session,
    document: SheetDocument,
    *,
    sync_entry_statuses: bool = False,
) -> None:
    signature = _get_attendance_questionnaire_sync_signature(session, document)
    if signature is None:
        return

    cache_key = str(document.id)
    with _attendance_questionnaire_sync_cache_lock:
        if _attendance_questionnaire_sync_cache.get(cache_key) == signature:
            return

    _sync_attendance_questionnaire_sheet_document(
        session,
        document,
        sync_entry_statuses=sync_entry_statuses,
    )

    next_signature = _get_attendance_questionnaire_sync_signature(session, document)
    if next_signature is not None:
        with _attendance_questionnaire_sync_cache_lock:
            _attendance_questionnaire_sync_cache[cache_key] = next_signature


def _sync_attendance_questionnaire_sheet_document(
    session: Session,
    document: SheetDocument,
    *,
    sync_entry_statuses: bool = False,
) -> None:
    if not _is_attendance_questionnaire_data_sheet(document):
        return

    current_document = deepcopy(dict(document.document_json or {}))
    from backend.api.attendance import _sync_attendance_wjx_sheet_rows_from_entries

    next_document, rows_changed = _sync_attendance_wjx_sheet_rows_from_entries(
        session,
        deepcopy(current_document),
    )
    next_document, links_changed = _sync_attendance_questionnaire_course_links(
        session,
        deepcopy(next_document),
    )
    changed = rows_changed or links_changed
    if changed and next_document != current_document:
        document.document_json = next_document
        document.version = max(int(document.version or 1), 1) + 1
        document.updated_at = time.time()
        session.add(document)
        session.commit()
        session.refresh(document)
        _broadcast_sheet_resource_update(document)
        next_document = dict(document.document_json or {})

    if sync_entry_statuses:
        _sync_attendance_questionnaire_entry_statuses(session, next_document)


def _build_paged_document(
    document_json: dict[str, Any],
    *,
    page: int,
    page_size: int,
    assume_normalized: bool = False,
) -> tuple[dict[str, Any], NoteSheetPaginationResponse]:
    normalized = document_json if assume_normalized else _normalize_document_json(document_json)
    all_rows = _extract_document_rows(normalized)
    safe_page_size = _normalize_page_size(page_size)
    actual_page_count = max(1, ceil(len(all_rows) / safe_page_size) if all_rows else 1)
    safe_page = min(max(int(page or 1), 1), actual_page_count)
    row_offset = min((safe_page - 1) * safe_page_size, len(all_rows))
    page_rows = all_rows[row_offset: row_offset + safe_page_size]
    page_document = _slice_paged_document_row_metadata(
        normalized,
        page_rows=page_rows,
        page_data_indexes=list(range(row_offset, row_offset + len(page_rows))),
    )
    grid_rows = _extract_document_grid_rows(normalized)
    if grid_rows:
        data_start_row = min(_normalize_document_data_start_row(normalized), len(grid_rows))
        page_document["grid_rows"] = [*grid_rows[:data_start_row], *page_rows]

    return (
        page_document,
        NoteSheetPaginationResponse(
            page=safe_page,
            page_size=safe_page_size,
            total_rows=len(all_rows),
            page_count=actual_page_count,
            row_offset=row_offset,
            loaded_row_count=len(page_rows),
        ),
    )


def _slice_paged_document_row_metadata(
    normalized: dict[str, Any],
    *,
    page_rows: list[Any],
    page_data_indexes: list[int],
) -> dict[str, Any]:
    """裁剪分页文档里的行级元数据。

    ``rows`` 本身只返回当前页，但历史实现会继续把整表的
    ``cell_meta``、``entity_rows`` 和 ``entity_cells`` 一并带回前端。
    这些字段在大表里可能远大于可见数据，所以分页响应只保留：

    - 表头区，也就是 ``data_start_row`` 之前的文档行；
    - 当前页实际包含的数据行，行号仍使用完整文档里的绝对行号；
    - 当前页实体行引用到的实体单元格。

    ``merged_cells`` 等结构字段暂时保持完整返回，因为分页保存合并仍依赖
    客户端把结构配置原样带回，贸然裁剪会有丢失未加载结构的风险。
    """
    page_document = {
        **normalized,
        "rows": page_rows,
    }
    data_start_row = _normalize_document_data_start_row(normalized)
    page_document_rows = {data_start_row + index for index in page_data_indexes if index >= 0}

    cell_meta = normalized.get("cell_meta")
    if isinstance(cell_meta, dict):
        next_cell_meta: dict[str, Any] = {}
        for key, value in cell_meta.items():
            position = _parse_cell_meta_key(key)
            if position is None:
                next_cell_meta[str(key)] = value
                continue
            row_index, _column_index = position
            if row_index < data_start_row or row_index in page_document_rows:
                next_cell_meta[str(key)] = value
        page_document["cell_meta"] = next_cell_meta

    entity_rows = _extract_document_entity_rows(normalized)
    if entity_rows:
        next_entity_rows = [
            entity_rows[index]
            for index in range(min(data_start_row, len(entity_rows)))
        ]
        for data_index in page_data_indexes:
            entity_index = data_start_row + data_index
            if 0 <= entity_index < len(entity_rows):
                next_entity_rows.append(entity_rows[entity_index])
        page_document["entity_rows"] = next_entity_rows

        entity_cells = normalized.get("entity_cells")
        if isinstance(entity_cells, dict):
            visible_row_ids = {
                row_id
                for row_id in (_get_document_entity_row_id(row) for row in next_entity_rows)
                if row_id
            }
            page_document["entity_cells"] = {
                row_id: cells
                for row_id, cells in entity_cells.items()
                if str(row_id) in visible_row_ids
            }

    return page_document


def _normalize_filter_text(value: Any) -> str:
    return _normalize_sheet_text(value)


def _normalize_filter_search_text(value: Any) -> str:
    return _normalize_filter_text(value).lower()


def _read_filter_field(record: dict[str, Any], camel_key: str, snake_key: str | None = None, default: Any = None) -> Any:
    if camel_key in record:
        return record.get(camel_key)
    if snake_key and snake_key in record:
        return record.get(snake_key)
    return default


def _normalize_filter_excluded_values(value: Any) -> set[str]:
    if not isinstance(value, list):
        return set()
    return {_normalize_filter_text(item) for item in value}


def _parse_note_sheet_filter_number(value: Any) -> float | None:
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value) if isfinite(float(value)) else None

    text = _normalize_filter_text(value)
    if not text:
        return None

    sign = 1.0
    parenthesized = re.fullmatch(r"\((.*)\)", text)
    if parenthesized:
        sign = -1.0
        text = parenthesized.group(1).strip()

    multiplier = 1.0
    if text.endswith("%"):
        multiplier *= 0.01
        text = text[:-1]
    if text.endswith("万"):
        multiplier *= 10_000
        text = text[:-1]
    elif text.endswith("亿"):
        multiplier *= 100_000_000
        text = text[:-1]

    normalized = re.sub(r"[￥¥,，\s]", "", text)
    if not re.fullmatch(r"[-+]?(?:\d+(?:\.\d+)?|\.\d+)", normalized):
        return None
    return float(normalized) * multiplier * sign


def _parse_note_sheet_filter_date_day(value: Any) -> int | None:
    if isinstance(value, datetime):
        return value.date().toordinal()
    if isinstance(value, date):
        return value.toordinal()
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        parts = _serial_to_date_parts(float(value))
        return date(parts[0], parts[1], parts[2]).toordinal() if parts else None

    text = _normalize_filter_text(value)
    if not text:
        return None
    if re.fullmatch(r"-?\d+(?:\.\d+)?", text):
        parts = _serial_to_date_parts(float(text))
        return date(parts[0], parts[1], parts[2]).toordinal() if parts else None

    separated = re.match(r"^(\d{4})[-/年](\d{1,2})[-/月](\d{1,2})日?(?:[ T]+\d{1,2}(?::\d{1,2}(?::\d{1,2})?)?)?$", text)
    if separated:
        try:
            return date(int(separated.group(1)), int(separated.group(2)), int(separated.group(3))).toordinal()
        except ValueError:
            return None

    compact = re.match(r"^(\d{4})(\d{2})(\d{2})(?:\d{2}\d{2}\d{0,2})?$", text)
    if compact:
        try:
            return date(int(compact.group(1)), int(compact.group(2)), int(compact.group(3))).toordinal()
        except ValueError:
            return None
    return None


def _normalize_note_sheet_filter_date(value: Any) -> str:
    day = _parse_note_sheet_filter_date_day(value)
    if day is None:
        return ""
    parsed = date.fromordinal(day)
    return parsed.isoformat()


def _is_column_filter_enabled(column_configs: dict[str, Any], header: str) -> bool:
    config = column_configs.get(header)
    return isinstance(config, dict) and config.get("filter_enabled") is True


def _is_column_filter_option_mode(column_configs: dict[str, Any], header: str) -> bool:
    config = column_configs.get(header)
    return isinstance(config, dict) and config.get("value_mode") == "fixed_options"


def _is_column_filter_date_mode(column_configs: dict[str, Any], header: str) -> bool:
    config = column_configs.get(header)
    return isinstance(config, dict) and config.get("value_type") == "date"


def _is_column_filter_number_mode(column_configs: dict[str, Any], header: str) -> bool:
    config = column_configs.get(header)
    return isinstance(config, dict) and config.get("value_type") in {"number", "percent"}


def _normalize_note_sheet_column_filters(
    column_filters: dict[str, Any],
    *,
    columns: list[str],
    column_configs: dict[str, Any],
) -> list[dict[str, Any]]:
    normalized: list[dict[str, Any]] = []
    header_indexes = {header: index for index, header in enumerate(columns)}
    for header, raw_state in column_filters.items():
        if header not in header_indexes or not _is_column_filter_enabled(column_configs, header):
            continue
        state = raw_state if isinstance(raw_state, dict) else {"query": raw_state}
        option_mode = _is_column_filter_option_mode(column_configs, header)
        date_mode = _is_column_filter_date_mode(column_configs, header)
        number_mode = _is_column_filter_number_mode(column_configs, header)
        query = "" if date_mode or number_mode else _normalize_filter_text(state.get("query"))
        excluded_values = _normalize_filter_excluded_values(
            _read_filter_field(state, "excludedValues", "excluded_values", []),
        ) if option_mode else set()
        date_start = _normalize_note_sheet_filter_date(_read_filter_field(state, "dateStart", "date_start", ""))
        date_end = _normalize_note_sheet_filter_date(_read_filter_field(state, "dateEnd", "date_end", ""))
        if date_start and date_end and date_start > date_end:
            date_start, date_end = date_end, date_start
        number_operator = _read_filter_field(state, "numberOperator", "number_operator", "gte")
        if number_operator not in {"eq", "neq", "gt", "gte", "lt", "lte", "between"}:
            number_operator = "gte"
        number_value = _parse_note_sheet_filter_number(_read_filter_field(state, "numberValue", "number_value", ""))
        number_value_to = _parse_note_sheet_filter_number(_read_filter_field(state, "numberValueTo", "number_value_to", ""))
        if number_operator == "between" and number_value is not None and number_value_to is not None and number_value > number_value_to:
            number_value, number_value_to = number_value_to, number_value

        active = (
            bool(query)
            or bool(excluded_values)
            or (date_mode and (bool(date_start) or bool(date_end)))
            or (
                number_mode
                and number_value is not None
                and (number_operator != "between" or number_value_to is not None)
            )
        )
        if not active:
            continue
        normalized.append({
            "header": header,
            "column_index": header_indexes[header],
            "query": query,
            "excluded_values": excluded_values,
            "date_start_day": _parse_note_sheet_filter_date_day(date_start) if date_start else None,
            "date_end_day": _parse_note_sheet_filter_date_day(date_end) if date_end else None,
            "number_operator": number_operator,
            "number_value": number_value,
            "number_value_to": number_value_to,
            "number_mode": number_mode,
        })
    return normalized


def _does_number_match_column_filter(value: float, operator: str, left: float, right: float | None) -> bool:
    if operator == "eq":
        return value == left
    if operator == "neq":
        return value != left
    if operator == "gt":
        return value > left
    if operator == "gte":
        return value >= left
    if operator == "lt":
        return value < left
    if operator == "lte":
        return value <= left
    if operator == "between":
        return right is not None and left <= value <= right
    return True


def _does_row_match_column_filters(
    raw_row: list[Any],
    display_row: list[Any],
    filters: list[dict[str, Any]],
) -> bool:
    for item in filters:
        column_index = int(item["column_index"])
        display_text = _normalize_filter_text(display_row[column_index] if column_index < len(display_row) else "")
        if item["query"] and _normalize_filter_search_text(item["query"]) not in _normalize_filter_search_text(display_text):
            return False
        if display_text in item["excluded_values"]:
            return False
        if item["date_start_day"] is not None or item["date_end_day"] is not None:
            raw_day = _parse_note_sheet_filter_date_day(raw_row[column_index] if column_index < len(raw_row) else "")
            display_day = _parse_note_sheet_filter_date_day(display_text)
            day_value = raw_day if raw_day is not None else display_day
            if day_value is None:
                return False
            if item["date_start_day"] is not None and day_value < item["date_start_day"]:
                return False
            if item["date_end_day"] is not None and day_value > item["date_end_day"]:
                return False
        if item["number_value"] is not None:
            raw_number = _parse_note_sheet_filter_number(raw_row[column_index] if column_index < len(raw_row) else "")
            display_number = _parse_note_sheet_filter_number(display_text)
            number_value = raw_number if raw_number is not None else display_number
            if number_value is None:
                return False
            if not _does_number_match_column_filter(
                number_value,
                item["number_operator"],
                item["number_value"],
                item["number_value_to"],
            ):
                return False
    return True


def _normalize_external_row_filter_program(value: Any) -> dict[str, Any]:
    program = value if isinstance(value, dict) else {}
    rules: list[dict[str, Any]] = []
    for raw_rule in program.get("rules") if isinstance(program.get("rules"), list) else []:
        rule = raw_rule if isinstance(raw_rule, dict) else {}
        action = rule.get("action")
        if action not in {"include", "exclude", "filter"}:
            action = "include"
        matcher = rule.get("matcher") if isinstance(rule.get("matcher"), dict) else {}
        kind = matcher.get("kind")
        if kind not in {"all", "none", "field", "full_text_contains"}:
            kind = "all"
        op = matcher.get("op")
        if op not in {"eq", "neq", "in", "not_in", "contains", "not_contains", "gte", "lte", "between"}:
            op = "contains" if kind == "field" else None
        rules.append({
            "action": action,
            "matcher": {
                "kind": kind,
                "field": matcher.get("field") if isinstance(matcher.get("field"), str) else None,
                "op": op,
                "value": matcher.get("value", ""),
                "values": matcher.get("values") if isinstance(matcher.get("values"), list) else [],
            },
        })
    return {
        "default": program.get("default") if isinstance(program.get("default"), bool) else False,
        "rules": rules,
    }


def _is_external_row_filter_rule_meaningful(rule: dict[str, Any]) -> bool:
    matcher = rule.get("matcher") if isinstance(rule.get("matcher"), dict) else {}
    kind = matcher.get("kind")
    if kind == "none":
        return True
    if kind == "all":
        return rule.get("action") != "include"
    if kind == "full_text_contains":
        return _normalize_filter_text(matcher.get("value")) != ""
    if kind != "field" or not matcher.get("field"):
        return False
    op = matcher.get("op")
    if op in {"between", "in", "not_in"}:
        return any(_normalize_filter_text(item) for item in matcher.get("values") or [])
    return _normalize_filter_text(matcher.get("value")) != ""


def _is_external_row_filter_program_active(program: dict[str, Any]) -> bool:
    return any(_is_external_row_filter_rule_meaningful(rule) for rule in program.get("rules", []))


def _compare_external_filter_values(left: str, right: str) -> int:
    left_number = _parse_note_sheet_filter_number(left)
    right_number = _parse_note_sheet_filter_number(right)
    if left_number is not None and right_number is not None:
        return (left_number > right_number) - (left_number < right_number)

    left_day = _parse_note_sheet_filter_date_day(left)
    right_day = _parse_note_sheet_filter_date_day(right)
    if left_day is not None and right_day is not None:
        return (left_day > right_day) - (left_day < right_day)

    left_text = _normalize_filter_search_text(left)
    right_text = _normalize_filter_search_text(right)
    return (left_text > right_text) - (left_text < right_text)


def _does_external_filter_value_match(text: str, matcher: dict[str, Any]) -> bool:
    op = matcher.get("op") or "contains"
    value = _normalize_filter_text(matcher.get("value"))
    values = [_normalize_filter_text(item) for item in (matcher.get("values") or [])]
    if op == "contains":
        return _normalize_filter_search_text(value) in _normalize_filter_search_text(text)
    if op == "not_contains":
        return _normalize_filter_search_text(value) not in _normalize_filter_search_text(text)
    if op == "eq":
        return text == value
    if op == "neq":
        return text != value
    if op == "in":
        return text in values
    if op == "not_in":
        return text not in values
    if op == "between":
        if len(values) < 2 or not values[0] or not values[1]:
            return True
        return _compare_external_filter_values(text, values[0]) >= 0 and _compare_external_filter_values(text, values[1]) <= 0
    if op == "gte":
        return _compare_external_filter_values(text, value) >= 0
    if op == "lte":
        return _compare_external_filter_values(text, value) <= 0
    return True


def _does_row_match_external_filter_matcher(
    display_row: list[Any],
    *,
    columns: list[str],
    matcher: dict[str, Any],
) -> bool:
    kind = matcher.get("kind")
    if kind == "all":
        return True
    if kind == "none":
        return False
    if kind == "full_text_contains":
        keyword = _normalize_filter_search_text(matcher.get("value"))
        if not keyword:
            return True
        return any(keyword in _normalize_filter_search_text(value) for value in display_row)
    if kind != "field" or not matcher.get("field"):
        return True
    field = _normalize_filter_text(matcher.get("field"))
    try:
        column_index = columns.index(field)
    except ValueError:
        return True
    text = _normalize_filter_text(display_row[column_index] if column_index < len(display_row) else "")
    return _does_external_filter_value_match(text, matcher)


def _does_row_match_external_filter_program(
    display_row: list[Any],
    *,
    columns: list[str],
    program: dict[str, Any],
) -> bool:
    decision = bool(program.get("default"))
    for rule in program.get("rules", []):
        if not _is_external_row_filter_rule_meaningful(rule):
            continue
        matcher = rule.get("matcher") if isinstance(rule.get("matcher"), dict) else {}
        matched = _does_row_match_external_filter_matcher(display_row, columns=columns, matcher=matcher)
        action = rule.get("action")
        if action == "filter":
            decision = decision and matched
        elif action == "exclude":
            decision = decision and not matched
        else:
            decision = decision or matched
    return decision


def _build_filtered_paged_document(
    document_json: dict[str, Any],
    *,
    page: int,
    page_size: int,
    column_filters: dict[str, Any],
    row_filter_programs: list[dict[str, Any]],
    assume_normalized: bool = False,
) -> tuple[dict[str, Any], dict[str, Any]]:
    normalized = document_json if assume_normalized else _normalize_document_json(document_json)
    columns = _normalize_document_columns(normalized)
    column_configs = normalized.get("column_configs")
    column_configs = column_configs if isinstance(column_configs, dict) else {}
    all_rows = _extract_document_rows(normalized)
    data_start_row = _normalize_document_data_start_row(normalized)
    text_grid = _build_table_text_grid(normalized, columns=columns, rows=all_rows)
    text_rows = text_grid[data_start_row:data_start_row + len(all_rows)]
    active_column_filters = _normalize_note_sheet_column_filters(
        column_filters,
        columns=columns,
        column_configs=column_configs,
    )
    active_row_filter_programs = [
        program
        for program in (_normalize_external_row_filter_program(item) for item in row_filter_programs)
        if _is_external_row_filter_program_active(program)
    ]

    filtered_indexes: list[int] = []
    for row_index, row in enumerate(all_rows):
        raw_row = _normalize_sheet_row(row, len(columns))
        display_row = _normalize_sheet_row(text_rows[row_index] if row_index < len(text_rows) else raw_row, len(columns))
        if not _does_row_match_column_filters(raw_row, display_row, active_column_filters):
            continue
        if any(
            not _does_row_match_external_filter_program(display_row, columns=columns, program=program)
            for program in active_row_filter_programs
        ):
            continue
        filtered_indexes.append(row_index)

    safe_page_size = _normalize_page_size(page_size)
    actual_page_count = max(1, ceil(len(filtered_indexes) / safe_page_size) if filtered_indexes else 1)
    safe_page = min(max(int(page or 1), 1), actual_page_count)
    row_offset = min((safe_page - 1) * safe_page_size, len(filtered_indexes))
    page_row_indexes = filtered_indexes[row_offset:row_offset + safe_page_size]
    page_rows = [all_rows[index] for index in page_row_indexes]
    page_document = _slice_paged_document_row_metadata(
        normalized,
        page_rows=page_rows,
        page_data_indexes=page_row_indexes,
    )
    grid_rows = _extract_document_grid_rows(normalized)
    if grid_rows:
        prefix_count = min(data_start_row, len(grid_rows))
        page_document["grid_rows"] = [*grid_rows[:prefix_count], *page_rows]

    return (
        page_document,
        {
            "page": safe_page,
            "page_size": safe_page_size,
            "total_rows": len(filtered_indexes),
            "unfiltered_total_rows": len(all_rows),
            "page_count": actual_page_count,
            "row_offset": row_offset,
            "loaded_row_count": len(page_rows),
            "row_indexes": page_row_indexes,
        },
    )


def _build_workspace_pagination(
    *,
    page_patch: NoteSheetPagePatchRequest,
    total_rows: int,
    current_row_count: int,
) -> NoteSheetPaginationResponse:
    safe_page_size = _normalize_page_size(page_patch.page_size)
    actual_page_count = max(1, ceil(total_rows / safe_page_size) if total_rows else 1)
    display_page = max(int(page_patch.page or 1), 1)
    return NoteSheetPaginationResponse(
        page=display_page,
        page_size=safe_page_size,
        total_rows=total_rows,
        page_count=max(actual_page_count, display_page),
        row_offset=min(int(page_patch.row_offset or 0), total_rows),
        loaded_row_count=max(current_row_count, 0),
    )


def _extract_document_entity_rows(document_json: dict[str, Any]) -> list[Any]:
    entity_rows = document_json.get("entity_rows")
    return list(entity_rows) if isinstance(entity_rows, list) else []


def _extract_document_entity_columns(document_json: dict[str, Any]) -> list[Any]:
    entity_columns = document_json.get("entity_columns")
    return list(entity_columns) if isinstance(entity_columns, list) else []


def _extract_document_entity_cells(document_json: dict[str, Any]) -> dict[str, Any]:
    entity_cells = document_json.get("entity_cells")
    return dict(entity_cells) if isinstance(entity_cells, dict) else {}


def _get_document_entity_row_id(row: Any) -> str:
    if not isinstance(row, dict):
        return ""
    return str(row.get("id") or "").strip()


def _remove_orphan_document_entity_cells(document_json: dict[str, Any]) -> dict[str, Any]:
    entity_cells = _extract_document_entity_cells(document_json)
    if not entity_cells:
        return document_json

    row_ids = {
        row_id
        for row_id in (_get_document_entity_row_id(row) for row in _extract_document_entity_rows(document_json))
        if row_id
    }
    next_entity_cells = {
        row_id: cells
        for row_id, cells in entity_cells.items()
        if row_id in row_ids
    }
    if len(next_entity_cells) == len(entity_cells):
        return document_json

    next_document = dict(document_json)
    next_document["entity_cells"] = next_entity_cells
    return next_document


def _get_document_entity_column_id(column: Any) -> str:
    if not isinstance(column, dict):
        return ""
    return str(column.get("id") or "").strip()


def _coerce_document_entity_row(row: Any, *, fallback_id: str, kind: str) -> dict[str, Any]:
    if isinstance(row, dict):
        coerced = dict(row)
        coerced["id"] = str(coerced.get("id") or fallback_id).strip() or fallback_id
        coerced["kind"] = str(coerced.get("kind") or kind).strip() or kind
        return coerced
    return {"id": fallback_id, "kind": kind}


def _merge_paged_entity_model(
    next_document: dict[str, Any],
    current_document: dict[str, Any],
    incoming_document: dict[str, Any],
    *,
    target_data_indexes: list[int],
    merged_row_count: int,
) -> None:
    current_entity_rows = _extract_document_entity_rows(current_document)
    incoming_entity_rows = _extract_document_entity_rows(incoming_document)
    current_entity_cells = _extract_document_entity_cells(current_document)
    incoming_entity_cells = _extract_document_entity_cells(incoming_document)
    if not (current_entity_rows or incoming_entity_rows or current_entity_cells or incoming_entity_cells):
        return

    data_start_row = _normalize_document_data_start_row(next_document)
    merged_entity_rows: list[dict[str, Any]] = []
    for row_index in range(data_start_row):
        source_row = (
            incoming_entity_rows[row_index]
            if row_index < len(incoming_entity_rows)
            else current_entity_rows[row_index]
            if row_index < len(current_entity_rows)
            else None
        )
        merged_entity_rows.append(
            _coerce_document_entity_row(
                source_row,
                fallback_id=f"header_{row_index + 1}",
                kind="field" if row_index == max(0, data_start_row - 1) else "header_group",
            )
        )

    for row_index in range(merged_row_count):
        source_index = data_start_row + row_index
        source_row = current_entity_rows[source_index] if source_index < len(current_entity_rows) else None
        merged_entity_rows.append(
            _coerce_document_entity_row(source_row, fallback_id=f"row_{row_index + 1}", kind="data")
        )

    for local_index, target_index in enumerate(target_data_indexes):
        if target_index < 0 or target_index >= merged_row_count:
            continue
        incoming_index = data_start_row + local_index
        source_row = incoming_entity_rows[incoming_index] if incoming_index < len(incoming_entity_rows) else None
        merged_entity_rows[data_start_row + target_index] = _coerce_document_entity_row(
            source_row,
            fallback_id=f"row_{target_index + 1}",
            kind="data",
        )

    next_entity_cells = dict(current_entity_cells)
    for row_index in range(data_start_row):
        target_row_id = _get_document_entity_row_id(merged_entity_rows[row_index])
        incoming_row_id = _get_document_entity_row_id(incoming_entity_rows[row_index]) if row_index < len(incoming_entity_rows) else ""
        current_row_id = _get_document_entity_row_id(current_entity_rows[row_index]) if row_index < len(current_entity_rows) else ""
        source_row_id = incoming_row_id if incoming_row_id in incoming_entity_cells else current_row_id
        source_cells = (
            incoming_entity_cells.get(source_row_id)
            if source_row_id == incoming_row_id
            else current_entity_cells.get(source_row_id)
        )
        if isinstance(source_cells, dict) and target_row_id:
            next_entity_cells[target_row_id] = source_cells

    for local_index, target_index in enumerate(target_data_indexes):
        if target_index < 0 or target_index >= merged_row_count:
            continue
        incoming_index = data_start_row + local_index
        target_index_with_header = data_start_row + target_index
        target_row_id = _get_document_entity_row_id(merged_entity_rows[target_index_with_header])
        incoming_row_id = (
            _get_document_entity_row_id(incoming_entity_rows[incoming_index])
            if incoming_index < len(incoming_entity_rows)
            else ""
        )
        current_row_id = (
            _get_document_entity_row_id(current_entity_rows[target_index_with_header])
            if target_index_with_header < len(current_entity_rows)
            else ""
        )
        if current_row_id and current_row_id != target_row_id:
            next_entity_cells.pop(current_row_id, None)
        if incoming_row_id in incoming_entity_cells and target_row_id:
            next_entity_cells[target_row_id] = incoming_entity_cells[incoming_row_id]
        elif target_row_id:
            next_entity_cells.pop(target_row_id, None)

    next_document["entity_rows"] = merged_entity_rows
    next_document["entity_cells"] = next_entity_cells


def _delete_document_entity_data_rows(
    document_json: dict[str, Any],
    deleted_data_indexes: set[int],
) -> dict[str, Any]:
    if not deleted_data_indexes:
        return document_json

    current_entity_rows = _extract_document_entity_rows(document_json)
    current_entity_cells = _extract_document_entity_cells(document_json)
    if not current_entity_rows and not current_entity_cells:
        return document_json

    data_start_row = _normalize_document_data_start_row(document_json)
    next_entity_rows: list[Any] = []
    removed_row_ids: set[str] = set()
    for row_index, row in enumerate(current_entity_rows):
        data_index = row_index - data_start_row
        if data_index >= 0 and data_index in deleted_data_indexes:
            row_id = _get_document_entity_row_id(row)
            if row_id:
                removed_row_ids.add(row_id)
            continue
        next_entity_rows.append(row)

    next_document = dict(document_json)
    if current_entity_rows:
        next_document["entity_rows"] = next_entity_rows
    if current_entity_cells:
        next_entity_cells = dict(current_entity_cells)
        for row_id in removed_row_ids:
            next_entity_cells.pop(row_id, None)
        next_document["entity_cells"] = next_entity_cells
    return next_document


def _extract_document_cell_meta(document_json: dict[str, Any]) -> dict[str, Any]:
    cell_meta = document_json.get("cell_meta")
    return dict(cell_meta) if isinstance(cell_meta, dict) else {}


def _get_document_grid_cell_value(document_json: dict[str, Any], row_index: int, column_index: int) -> Any:
    if row_index < 0 or column_index < 0:
        return ""

    columns = _normalize_document_columns(document_json)
    column_count = len(columns)
    if column_index >= column_count:
        return ""

    grid_rows = _extract_document_grid_rows(document_json)
    if row_index < len(grid_rows):
        return _normalize_sheet_row(grid_rows[row_index], column_count)[column_index]

    data_start_row = _normalize_document_data_start_row(document_json)
    data_row_index = row_index - data_start_row
    rows = _extract_document_rows(document_json)
    if 0 <= data_row_index < len(rows):
        return _extract_row_cell_value(rows[data_row_index], column_index, columns)
    return ""


def _remove_rich_text_from_meta_entry(entry: Any) -> Any:
    if not isinstance(entry, dict):
        return entry
    if "rich_text" not in entry and entry.get("cell_type") != "rich_text":
        return entry
    next_entry = dict(entry)
    if next_entry.get("cell_type") == "rich_text":
        next_entry.pop("cell_type", None)
    next_entry.pop("rich_text", None)
    return next_entry or None


def _strip_formula_cell_rich_text(document_json: dict[str, Any]) -> dict[str, Any]:
    normalized = _normalize_document_json(document_json)
    changed = False

    cell_meta = _extract_document_cell_meta(normalized)
    if cell_meta:
        next_cell_meta: dict[str, Any] = {}
        for key, entry in cell_meta.items():
            position = _parse_cell_meta_key(key)
            if position is None:
                next_cell_meta[key] = entry
                continue
            row_index, column_index = position
            if _is_formula_expression(_get_document_grid_cell_value(normalized, row_index, column_index)):
                next_entry = _remove_rich_text_from_meta_entry(entry)
                changed = changed or next_entry is not entry
                if next_entry is not None:
                    next_cell_meta[key] = next_entry
                continue
            next_cell_meta[key] = entry
        if next_cell_meta != cell_meta:
            normalized["cell_meta"] = next_cell_meta
            changed = True

    entity_cells = _extract_document_entity_cells(normalized)
    entity_rows = _extract_document_entity_rows(normalized)
    entity_columns = _extract_document_entity_columns(normalized)
    if entity_cells and entity_rows and entity_columns:
        next_entity_cells = dict(entity_cells)
        for row_index, row in enumerate(entity_rows):
            if not isinstance(row, dict):
                continue
            row_id = str(row.get("id") or "")
            if not row_id or not isinstance(next_entity_cells.get(row_id), dict):
                continue
            source_row_cells = next_entity_cells[row_id]
            next_row_cells = dict(source_row_cells)
            for column_index, column in enumerate(entity_columns):
                if not isinstance(column, dict):
                    continue
                column_id = str(column.get("id") or "")
                entry = next_row_cells.get(column_id)
                if not isinstance(entry, dict) or ("rich_text" not in entry and entry.get("cell_type") != "rich_text"):
                    continue
                value = entry.get("value", _get_document_grid_cell_value(normalized, row_index, column_index))
                if not _is_formula_expression(value):
                    continue
                next_entry = _remove_rich_text_from_meta_entry(entry)
                if next_entry is None:
                    next_row_cells.pop(column_id, None)
                else:
                    next_row_cells[column_id] = next_entry
                changed = True
            if next_row_cells:
                next_entity_cells[row_id] = next_row_cells
            else:
                next_entity_cells.pop(row_id, None)
        if next_entity_cells != entity_cells:
            normalized["entity_cells"] = next_entity_cells
            changed = True

    return normalized if changed else document_json


def _merge_paged_cell_meta_by_rows(
    current_document: dict[str, Any],
    incoming_document: dict[str, Any],
    *,
    target_data_indexes: list[int],
    source_data_indexes: list[int],
    deleted_data_indexes: list[int],
) -> dict[str, Any] | None:
    has_cell_meta = isinstance(current_document.get("cell_meta"), dict) or isinstance(incoming_document.get("cell_meta"), dict)
    if not has_cell_meta:
        return None

    current_meta = _extract_document_cell_meta(current_document)
    incoming_meta = _extract_document_cell_meta(incoming_document)

    data_start_row = _normalize_document_data_start_row(current_document)
    deleted_indexes = sorted({index for index in deleted_data_indexes if index >= 0})
    replacement_pairs = [
        (target_index, source_index)
        for target_index, source_index in zip(target_data_indexes, source_data_indexes)
        if target_index >= 0 and source_index >= 0
    ]
    replacement_target_rows = {data_start_row + target_index for target_index, _source_index in replacement_pairs}
    replacement_source_rows = {data_start_row + source_index for _target_index, source_index in replacement_pairs}

    next_meta: dict[str, Any] = {}
    for key, value in current_meta.items():
        position = _parse_cell_meta_key(key)
        if position is None:
            continue
        row, column = position
        if row < data_start_row:
            next_meta[key] = value
            continue
        data_index = row - data_start_row
        if data_index in deleted_indexes:
            continue
        shifted_data_index = data_index - sum(1 for deleted_index in deleted_indexes if deleted_index < data_index)
        shifted_row = data_start_row + shifted_data_index
        if shifted_row in replacement_target_rows:
            continue
        next_meta[f"{shifted_row}:{column}"] = value

    for key, value in incoming_meta.items():
        position = _parse_cell_meta_key(key)
        if position is None:
            continue
        row, column = position
        if row < data_start_row:
            next_meta[key] = value
            continue
        if row not in replacement_source_rows:
            continue
        for target_index, source_index in replacement_pairs:
            if row == data_start_row + source_index:
                next_meta[f"{data_start_row + target_index}:{column}"] = value
                break

    return next_meta


def _merge_paged_document(
    current_document: dict[str, Any],
    incoming_document: dict[str, Any],
    page_patch: NoteSheetPagePatchRequest,
) -> dict[str, Any]:
    normalized_current = _normalize_document_json(current_document)
    normalized_incoming = _normalize_document_json(incoming_document)

    current_rows = _extract_document_rows(normalized_current)
    incoming_rows = _extract_document_rows(normalized_incoming)
    current_row_ids = list(normalized_current.get("row_ids")) if isinstance(normalized_current.get("row_ids"), list) else []
    incoming_row_ids = list(normalized_incoming.get("row_ids")) if isinstance(normalized_incoming.get("row_ids"), list) else []
    row_indexes = [
        int(index)
        for index in page_patch.row_indexes
        if isinstance(index, int) and 0 <= int(index) < len(current_rows)
    ]
    deleted_row_indexes = sorted({
        int(index)
        for index in page_patch.deleted_row_indexes
        if isinstance(index, int) and 0 <= int(index) < len(current_rows)
    })
    if row_indexes or deleted_row_indexes:
        if len(row_indexes) != len(incoming_rows) or len(set(row_indexes)) != len(row_indexes):
            raise HTTPException(status_code=409, detail="筛选分页保存需要保持行数一致")
        if set(row_indexes).intersection(deleted_row_indexes):
            raise HTTPException(status_code=409, detail="筛选分页保存的更新行和删除行不能重叠")
        merged_rows = list(current_rows)
        for source_index, incoming_row in zip(row_indexes, incoming_rows):
            merged_rows[source_index] = incoming_row
        for deleted_index in sorted(deleted_row_indexes, reverse=True):
            del merged_rows[deleted_index]
        mapped_row_indexes = [
            source_index - sum(1 for deleted_index in deleted_row_indexes if deleted_index < source_index)
            for source_index in row_indexes
        ]
        next_document = {
            **normalized_current,
            **{
                key: value
                for key, value in normalized_incoming.items()
                if key not in {"rows", "row_ids", "grid_rows", "entity_rows", "entity_cells", "cell_meta"}
            },
            "rows": merged_rows,
        }
        if current_row_ids or incoming_row_ids:
            merged_row_ids = [str(value) for value in current_row_ids[:len(current_rows)]]
            if len(merged_row_ids) < len(current_rows):
                merged_row_ids.extend([""] * (len(current_rows) - len(merged_row_ids)))
            for incoming_offset, target_index in enumerate(row_indexes):
                if incoming_offset < len(incoming_row_ids) and 0 <= target_index < len(merged_row_ids):
                    merged_row_ids[target_index] = str(incoming_row_ids[incoming_offset])
            for deleted_index in sorted(deleted_row_indexes, reverse=True):
                if 0 <= deleted_index < len(merged_row_ids):
                    del merged_row_ids[deleted_index]
            next_document["row_ids"] = merged_row_ids
        merged_cell_meta = _merge_paged_cell_meta_by_rows(
            normalized_current,
            normalized_incoming,
            target_data_indexes=mapped_row_indexes,
            source_data_indexes=row_indexes,
            deleted_data_indexes=deleted_row_indexes,
        )
        if merged_cell_meta is not None:
            next_document["cell_meta"] = merged_cell_meta
        current_grid_rows = _extract_document_grid_rows(normalized_current)
        incoming_grid_rows = _extract_document_grid_rows(normalized_incoming)
        if current_grid_rows or incoming_grid_rows:
            data_start_row = _normalize_document_data_start_row(next_document)
            source_grid_rows = incoming_grid_rows or current_grid_rows
            next_document["grid_rows"] = [*source_grid_rows[:data_start_row], *merged_rows]
        entity_current_document = _delete_document_entity_data_rows(
            normalized_current,
            set(deleted_row_indexes),
        )
        _merge_paged_entity_model(
            next_document,
            entity_current_document,
            normalized_incoming,
            target_data_indexes=mapped_row_indexes,
            merged_row_count=len(merged_rows),
        )
        return next_document

    row_offset = min(max(int(page_patch.row_offset or 0), 0), len(current_rows))
    loaded_row_count = max(int(page_patch.loaded_row_count or 0), 0)
    tail_start = min(row_offset + loaded_row_count, len(current_rows))

    merged_rows = [
        *current_rows[:row_offset],
        *incoming_rows,
        *current_rows[tail_start:],
    ]
    next_document = {
        **normalized_current,
            **{
                key: value
                for key, value in normalized_incoming.items()
                if key not in {"rows", "row_ids", "grid_rows", "entity_rows", "entity_cells", "cell_meta"}
            },
            "rows": merged_rows,
        }
    if current_row_ids or incoming_row_ids:
        next_document["row_ids"] = [
            *[str(value) for value in current_row_ids[:row_offset]],
            *[str(value) for value in incoming_row_ids[:len(incoming_rows)]],
            *[str(value) for value in current_row_ids[tail_start:len(current_rows)]],
        ]
    replacement_indexes = list(range(row_offset, row_offset + len(incoming_rows)))
    deleted_indexes = list(range(row_offset + len(incoming_rows), tail_start))
    merged_cell_meta = _merge_paged_cell_meta_by_rows(
        normalized_current,
        normalized_incoming,
        target_data_indexes=replacement_indexes,
        source_data_indexes=replacement_indexes,
        deleted_data_indexes=deleted_indexes,
    )
    if merged_cell_meta is not None:
        next_document["cell_meta"] = merged_cell_meta
    current_grid_rows = _extract_document_grid_rows(normalized_current)
    incoming_grid_rows = _extract_document_grid_rows(normalized_incoming)
    if current_grid_rows or incoming_grid_rows:
        data_start_row = _normalize_document_data_start_row(next_document)
        source_grid_rows = incoming_grid_rows or current_grid_rows
        next_document["grid_rows"] = [*source_grid_rows[:data_start_row], *merged_rows]
    _merge_paged_entity_model(
        next_document,
        normalized_current,
        normalized_incoming,
        target_data_indexes=replacement_indexes,
        merged_row_count=len(merged_rows),
    )
    return next_document


def _normalize_restricted_cell_value(value: Any) -> str:
    value = _extract_cell_value(value)
    return "" if value is None else str(value)


RESTRICTED_DATA_UPDATE_STRUCTURAL_KEYS = (
    "schema_version",
    "data_start_row",
    "field_row_index",
    "merged_cells",
    "formula_reference_origin",
    "header_groups",
    "entity_columns",
    "row_ids",
    "column_ids",
    "column_widths",
    "view_settings",
)


def _apply_restricted_data_column_update(
    current_document: dict[str, Any],
    incoming_document: dict[str, Any],
    editable_columns: list[int],
) -> dict[str, Any]:
    current_normalized = _normalize_document_json(current_document)
    incoming_normalized = _normalize_document_json(incoming_document)
    current_columns = _normalize_document_columns(current_normalized)
    incoming_columns = _normalize_document_columns(incoming_normalized)
    if incoming_columns != current_columns:
        raise HTTPException(status_code=403, detail="游客只能编辑开放列，不能修改表格结构")

    for key in RESTRICTED_DATA_UPDATE_STRUCTURAL_KEYS:
        if current_normalized.get(key) != incoming_normalized.get(key):
            raise HTTPException(status_code=403, detail="游客只能编辑开放列，不能修改表格结构")

    allowed_columns = {
        int(index)
        for index in editable_columns
        if 0 <= int(index) < len(current_columns)
    }
    if not allowed_columns:
        raise HTTPException(status_code=403, detail="没有该资源权限")

    current_rows = _extract_document_rows(current_normalized)
    incoming_rows = _extract_document_rows(incoming_normalized)
    if len(current_rows) != len(incoming_rows):
        raise HTTPException(status_code=403, detail="游客只能编辑开放列，不能新增或删除行")

    next_rows: list[Any] = []
    column_count = len(current_columns)
    for row_index, current_row in enumerate(current_rows):
        incoming_row = incoming_rows[row_index]
        current_cells = _normalize_sheet_row(current_row, column_count)
        incoming_cells = _normalize_sheet_row(incoming_row, column_count)
        for column_index in range(column_count):
            if column_index in allowed_columns:
                continue
            if _normalize_restricted_cell_value(incoming_cells[column_index]) != _normalize_restricted_cell_value(current_cells[column_index]):
                raise HTTPException(status_code=403, detail="游客只能编辑开放列")

        next_row = current_row
        for column_index in sorted(allowed_columns):
            next_row = _set_row_cell_value(next_row, current_columns, column_index, incoming_cells[column_index])
        next_rows.append(next_row)

    return _replace_document_data_rows(current_normalized, next_rows)


def _normalize_title(value: str | None, *, default_value: str) -> str:
    normalized = str(value or "").strip()
    return normalized or default_value


def _extract_sort_cell_text(row: Any, column_index: int, columns: list[Any]) -> str:
    if isinstance(row, list):
        raw_value = row[column_index] if column_index < len(row) else ""
    elif isinstance(row, dict):
        column_key = str(columns[column_index]) if column_index < len(columns) else ""
        raw_value = row.get(column_key, "")
    else:
        raw_value = ""
    raw_value = _extract_cell_value(raw_value)
    return "" if raw_value is None else str(raw_value).strip()


def _normalize_column_option_value(value: Any) -> str:
    return _normalize_sheet_text(value)


def _get_column_option_label(value: str) -> str:
    return value or "(空白)"


def _build_note_sheet_column_options_response(
    document: SheetDocument,
    *,
    column_index: int,
) -> NoteSheetColumnOptionsResponse:
    normalized = _normalize_document_json(dict(document.document_json or {}))
    columns = _normalize_document_columns(normalized)
    if column_index < 0 or column_index >= len(columns):
        raise HTTPException(status_code=400, detail="列索引超出范围")

    rows = _extract_document_rows(normalized)
    data_start_row = _normalize_document_data_start_row(normalized)
    text_grid = _build_table_text_grid(normalized, columns=columns, rows=rows)
    text_rows = text_grid[data_start_row: data_start_row + len(rows)]

    counts: dict[str, int] = {}
    for row in text_rows:
        cells = _normalize_sheet_row(row, len(columns))
        value = _normalize_column_option_value(cells[column_index])
        counts[value] = counts.get(value, 0) + 1

    options = [
        NoteSheetColumnOptionItemResponse(
            value=value,
            label=_get_column_option_label(value),
            count=count,
        )
        for value, count in counts.items()
    ]
    options.sort(key=lambda item: (-item.count, item.label))

    return NoteSheetColumnOptionsResponse(
        column_index=column_index,
        header=columns[column_index],
        total_rows=len(rows),
        options=options,
    )


def _extract_row_cell_value(row: Any, column_index: int, columns: list[Any]) -> Any:
    if isinstance(row, list):
        return _extract_cell_value(row[column_index]) if column_index < len(row) else ""
    if isinstance(row, dict):
        column_key = str(columns[column_index]) if column_index < len(columns) else ""
        return _extract_cell_value(row.get(column_key, ""))
    return ""


def _is_formula_expression(value: Any) -> bool:
    value = _extract_cell_value(value)
    return isinstance(value, str) and value.startswith("=")


def _normalize_formula_for_sort(value: str) -> str:
    return value.translate(FORMULA_PUNCTUATION_TRANSLATION)


def _unquote_formula_string(value: str | None) -> str:
    if not value:
        return ""
    stripped = value.strip()
    if len(stripped) >= 2 and stripped[0] == stripped[-1] and stripped[0] in {"'", '"'}:
        quote = stripped[0]
        inner = stripped[1:-1]
        return inner.replace(quote * 2, quote)
    return stripped


def _date_parts_to_serial(year: int, month: int, day: int) -> float | None:
    try:
        parsed = date(year, month, day)
    except ValueError:
        return None
    unix_epoch = date(1970, 1, 1)
    return float((parsed - unix_epoch).days + EXCEL_DATE_UNIX_EPOCH_SERIAL)


def _serial_to_date_parts(value: float) -> tuple[int, int, int] | None:
    try:
        parsed = date.fromordinal(date(1970, 1, 1).toordinal() + round(value - EXCEL_DATE_UNIX_EPOCH_SERIAL))
    except (OverflowError, ValueError):
        return None
    return parsed.year, parsed.month, parsed.day


def _normalize_two_digit_year(value: int) -> int:
    return 1900 + value if value >= 70 else 2000 + value


def _parse_compact_date_serial(value: Any, pattern: str = "yyyymmdd") -> float | None:
    text = _normalize_sheet_text(value)
    normalized_pattern = re.sub(r"[^ymd]", "", str(pattern or "yyyymmdd").lower()) or "yyyymmdd"
    if normalized_pattern not in {"yyyymmdd", "yymmdd"}:
        return None

    match = re.search(r"\d{8}", text)
    if match:
        digits = match.group(0)
        return _date_parts_to_serial(int(digits[:4]), int(digits[4:6]), int(digits[6:8]))

    match = re.search(r"\d{6}", text)
    if match:
        digits = match.group(0)
        return _date_parts_to_serial(
            _normalize_two_digit_year(int(digits[:2])),
            int(digits[2:4]),
            int(digits[4:6]),
        )
    return None


def _parse_date_sort_value(value: Any) -> float | None:
    value = _extract_cell_value(value)
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)

    text = "" if value is None else str(value).strip()
    if not text:
        return None

    if re.fullmatch(r"-?\d+(?:\.\d+)?", text):
        return float(text)

    separated = re.fullmatch(r"(\d{4})[-/年](\d{1,2})[-/月](\d{1,2})日?", text)
    if separated:
        return _date_parts_to_serial(
            int(separated.group(1)),
            int(separated.group(2)),
            int(separated.group(3)),
        )

    compact = re.fullmatch(r"(\d{4})(\d{2})(\d{2})", text)
    if compact:
        return _date_parts_to_serial(
            int(compact.group(1)),
            int(compact.group(2)),
            int(compact.group(3)),
        )
    return None


def _parse_number_sort_value(value: Any) -> float | None:
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    text = "" if value is None else str(value).strip()
    if re.fullmatch(r"[-+]?(?:\d+(?:\.\d+)?|\.\d+)", text):
        return float(text)
    return None


def _parse_percent_sort_value(value: Any) -> float | None:
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    text = "" if value is None else str(value).strip()
    percent_match = re.fullmatch(r"([-+]?(?:\d+(?:\.\d+)?|\.\d+))%", text)
    if percent_match:
        return float(percent_match.group(1)) / 100
    return _parse_number_sort_value(value)


def _parse_table_formula_cell_reference(value: str) -> tuple[int, int] | None:
    match = FORMULA_CELL_REFERENCE_ONLY_RE.fullmatch(value.strip())
    if not match:
        return None
    column_index = _excel_column_index(match.group(1))
    row_number = int(match.group(2))
    if column_index is None or row_number < 1:
        return None
    return row_number - 1, column_index


def _resolve_formula_reference_value(
    token: str,
    *,
    rows: list[Any],
    columns: list[Any],
    depth: int,
    reference_row_offset: int = 0,
    grid_rows: list[Any] | None = None,
) -> Any:
    reference = _parse_formula_cell_reference(token)
    if reference is None:
        return _unquote_formula_string(token)

    reference_row_index, column_index = reference
    row_index = reference_row_index - reference_row_offset
    if row_index < 0:
        if grid_rows and 0 <= reference_row_index < len(grid_rows):
            return _extract_row_cell_value(grid_rows[reference_row_index], column_index, columns)
        return ""
    if row_index >= len(rows):
        return ""
    raw_value = _extract_row_cell_value(rows[row_index], column_index, columns)
    return _evaluate_formula_sort_value(
        raw_value,
        row_index=row_index,
        column_index=column_index,
        rows=rows,
        columns=columns,
        reference_row_offset=reference_row_offset,
        grid_rows=grid_rows,
        depth=depth + 1,
    )


def _evaluate_formula_sort_value(
    value: Any,
    *,
    row_index: int,
    column_index: int,
    rows: list[Any],
    columns: list[Any],
    reference_row_offset: int = 0,
    grid_rows: list[Any] | None = None,
    depth: int = 0,
) -> Any:
    if depth > 8 or not _is_formula_expression(value):
        return value

    formula = _normalize_formula_for_sort(str(value))
    date_match = DATE_PARSE_FORMULA_RE.match(formula)
    if date_match:
        source_value = _resolve_formula_reference_value(
            date_match.group("source"),
            rows=rows,
            columns=columns,
            reference_row_offset=reference_row_offset,
            grid_rows=grid_rows,
            depth=depth,
        )
        pattern = _unquote_formula_string(date_match.group("pattern")) or "yyyymmdd"
        parsed = _parse_compact_date_serial(source_value, pattern)
        return parsed if parsed is not None else value

    offset_match = CELL_OFFSET_FORMULA_RE.match(formula)
    if offset_match:
        source_value = _resolve_formula_reference_value(
            offset_match.group("source"),
            rows=rows,
            columns=columns,
            reference_row_offset=reference_row_offset,
            grid_rows=grid_rows,
            depth=depth,
        )
        numeric_value = _parse_number_sort_value(source_value)
        if numeric_value is None:
            numeric_value = _parse_date_sort_value(source_value)
        if numeric_value is None:
            return value
        offset = float(offset_match.group("offset"))
        return numeric_value + offset if offset_match.group("operator") == "+" else numeric_value - offset

    return value


def _excel_column_index(label: str) -> int | None:
    index = 0
    for char in label.upper():
        char_code = ord(char)
        if char_code < 65 or char_code > 90:
            return None
        index = index * 26 + char_code - 64
    return index - 1


def _excel_column_label(index: int) -> str:
    current = index
    label = ""
    while True:
        label = chr(65 + (current % 26)) + label
        current = current // 26 - 1
        if current < 0:
            return label


def _apply_formula_reference_column_case(label: str, source_label: str) -> str:
    return label.lower() if source_label == source_label.lower() else label


def _remap_formula_cell_references(
    formula: str,
    *,
    row_index_map: dict[int, int | None] | None = None,
    column_index_map: dict[int, int | None] | None = None,
    row_index_offset: int = 0,
) -> str:
    if not _is_formula_expression(formula):
        return formula

    def replace(match: re.Match[str]) -> str:
        prefix = match.group(1)
        column_absolute_marker = match.group(2)
        column_label = match.group(3)
        row_absolute_marker = match.group(4)
        row_label = match.group(5)
        source_column_index = _excel_column_index(column_label)
        source_row_number = int(row_label)
        if source_column_index is None or source_row_number < 1:
            return f"{prefix}#REF!"

        source_row_index = source_row_number - 1
        next_column_index = (
            column_index_map.get(source_column_index, source_column_index)
            if column_index_map is not None
            else source_column_index
        )
        if row_index_map is None:
            next_row_index = source_row_index
        elif source_row_index < row_index_offset:
            next_row_index = source_row_index
        else:
            source_data_row_index = source_row_index - row_index_offset
            next_data_row_index = row_index_map.get(source_data_row_index, source_data_row_index)
            next_row_index = None if next_data_row_index is None else next_data_row_index + row_index_offset
        if next_column_index is None or next_row_index is None or next_column_index < 0 or next_row_index < 0:
            return f"{prefix}#REF!"

        next_column_label = _apply_formula_reference_column_case(
            _excel_column_label(next_column_index),
            column_label,
        )
        return f"{prefix}{column_absolute_marker}{next_column_label}{row_absolute_marker}{next_row_index + 1}"

    return FORMULA_CELL_REFERENCE_RE.sub(replace, formula)


def _remap_formula_value_references(
    value: Any,
    *,
    row_index_map: dict[int, int | None] | None = None,
    column_index_map: dict[int, int | None] | None = None,
    row_index_offset: int = 0,
) -> Any:
    if not _is_formula_expression(value):
        return value
    return _remap_formula_cell_references(
        value,
        row_index_map=row_index_map,
        column_index_map=column_index_map,
        row_index_offset=row_index_offset,
    )


def _remap_row_formula_cell_references(
    row: Any,
    *,
    columns: list[Any],
    row_index_map: dict[int, int | None] | None = None,
    column_index_map: dict[int, int | None] | None = None,
    row_index_offset: int = 0,
) -> Any:
    if isinstance(row, list):
        return [
            _remap_formula_value_references(
                value,
                row_index_map=row_index_map,
                column_index_map=column_index_map,
                row_index_offset=row_index_offset,
            )
            for value in row
        ]

    if isinstance(row, dict):
        next_row = dict(row)
        for column in columns:
            column_key = str(column)
            if column_key not in next_row:
                continue
            next_row[column_key] = _remap_formula_value_references(
                next_row[column_key],
                row_index_map=row_index_map,
                column_index_map=column_index_map,
                row_index_offset=row_index_offset,
            )
        return next_row

    return row


def _shift_formula_cell_references(formula: str, *, row_delta: int = 0, column_delta: int = 0) -> str:
    if not _is_formula_expression(formula) or (row_delta == 0 and column_delta == 0):
        return formula

    def replace(match: re.Match[str]) -> str:
        prefix = match.group(1)
        column_absolute_marker = match.group(2)
        column_label = match.group(3)
        row_absolute_marker = match.group(4)
        row_label = match.group(5)
        source_column_index = _excel_column_index(column_label)
        source_row_number = int(row_label)
        if source_column_index is None or source_row_number < 1:
            return f"{prefix}#REF!"

        next_column_index = source_column_index if column_absolute_marker else source_column_index + column_delta
        next_row_number = source_row_number if row_absolute_marker else source_row_number + row_delta
        if next_column_index < 0 or next_row_number < 1:
            return f"{prefix}#REF!"

        next_column_label = _apply_formula_reference_column_case(
            _excel_column_label(next_column_index),
            column_label,
        )
        return f"{prefix}{column_absolute_marker}{next_column_label}{row_absolute_marker}{next_row_number}"

    return FORMULA_CELL_REFERENCE_RE.sub(replace, formula)


def _shift_formula_value_references(value: Any, *, row_delta: int = 0, column_delta: int = 0) -> Any:
    if not _is_formula_expression(value):
        return value
    shifted_value = _shift_formula_cell_references(str(_extract_cell_value(value)), row_delta=row_delta, column_delta=column_delta)
    if isinstance(value, dict):
        next_cell = dict(value)
        next_cell["value"] = shifted_value
        return next_cell
    return shifted_value


def _normalize_sheet_row(row: Any, column_count: int) -> list[Any]:
    if isinstance(row, list):
        return [*(row[:column_count]), *([""] * max(column_count - len(row), 0))]
    if isinstance(row, dict):
        return [row.get(str(index), "") for index in range(column_count)]
    return [""] * column_count


def _extract_cell_value(value: Any) -> Any:
    return note_sheet_inline_links.extract_inline_cell_value(value)


def _normalize_sheet_text(value: Any) -> str:
    value = _extract_cell_value(value)
    return "" if value is None else str(value).strip()


def _normalize_excel_import_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _strip_legacy_text_prefix(value: Any) -> str:
    return _normalize_sheet_text(value).lstrip("`'")


def _normalize_excel_import_cell_for_column(column: str, value: Any) -> str:
    normalized_value = _normalize_excel_import_cell(value)
    return (
        _strip_legacy_text_prefix(normalized_value)
        if column in NOTE_SHEET_LEGACY_TEXT_PREFIX_STRIP_COLUMNS
        else normalized_value
    )


def _compact_action_token(value: Any) -> str:
    return re.sub(r"\s+", "", _normalize_sheet_text(value)).lower()


def _is_legacy_action_text(value: Any, tokens: tuple[str, ...]) -> bool:
    action_tokens = {_compact_action_token(token) for token in tokens}
    return _compact_action_token(value) in action_tokens


def _is_legacy_excel_import_action_text(value: Any) -> bool:
    return _is_legacy_action_text(value, NOTE_SHEET_EXCEL_IMPORT_ACTION_TOKENS)


def _sheet_row_has_legacy_excel_import_action(row: Any) -> bool:
    values = row if isinstance(row, list) else list(row.values()) if isinstance(row, dict) else [row]
    return any(_is_legacy_excel_import_action_text(value) for value in values)


def _cell_meta_has_excel_import_action(meta: Any) -> bool:
    if not isinstance(meta, dict):
        return False
    action = meta.get("action")
    if action == NOTE_SHEET_CELL_ACTION_EXCEL_IMPORT_RESET:
        return True
    if not isinstance(action, dict):
        return False
    return _normalize_sheet_text(action.get("type") or action.get("name")) == NOTE_SHEET_CELL_ACTION_EXCEL_IMPORT_RESET


def _find_excel_import_action_data_row(
    document_json: dict[str, Any],
    rows: list[list[Any]],
    *,
    action_document_row: int | None = None,
    action_column: int | None = None,
) -> int | None:
    data_start_row = _normalize_document_data_start_row(document_json)
    column_count = len(_normalize_document_columns(document_json))
    cell_meta = document_json.get("cell_meta")

    def is_valid_action_cell(document_row: int, column_index: int | None) -> bool:
        if column_index is None or column_index < 0 or column_index >= column_count:
            return False
        meta = cell_meta.get(f"{document_row}:{column_index}") if isinstance(cell_meta, dict) else None
        if _cell_meta_has_excel_import_action(meta):
            return True
        data_row_index = document_row - data_start_row
        if data_row_index < 0:
            grid_rows = _extract_document_grid_rows(document_json)
            if 0 <= document_row < len(grid_rows):
                return _is_legacy_excel_import_action_text(_normalize_sheet_row(grid_rows[document_row], column_count)[column_index])
            return False
        if data_row_index >= len(rows):
            return False
        return _is_legacy_excel_import_action_text(rows[data_row_index][column_index])

    if action_document_row is not None or action_column is not None:
        if action_document_row is None or action_column is None:
            raise HTTPException(status_code=400, detail="导入按钮坐标不完整")
        if is_valid_action_cell(action_document_row, action_column):
            return action_document_row - data_start_row
        raise HTTPException(status_code=400, detail="当前单元格不是导入 Excel 按钮")

    if isinstance(cell_meta, dict):
        action_positions: list[tuple[int, int]] = []
        for key, meta in cell_meta.items():
            parsed = _parse_cell_meta_key(key)
            if parsed is None or not _cell_meta_has_excel_import_action(meta):
                continue
            row_index, column_index = parsed
            if row_index >= 0 and 0 <= column_index < column_count:
                action_positions.append((row_index, column_index))
        for row_index, column_index in sorted(action_positions):
            data_row_index = row_index - data_start_row
            if data_row_index < 0:
                return data_row_index
            if 0 <= data_row_index < len(rows) and 0 <= column_index < column_count:
                return data_row_index

    grid_rows = _extract_document_grid_rows(document_json)
    for index, row in enumerate(grid_rows[:data_start_row]):
        if _sheet_row_has_legacy_excel_import_action(row):
            return index - data_start_row

    for index, row in enumerate(rows):
        if _sheet_row_has_legacy_excel_import_action(row):
            return index
    return None


def _get_excel_import_preserved_data_rows(
    document_json: dict[str, Any],
    *,
    action_document_row: int | None = None,
    action_column: int | None = None,
) -> list[list[Any]]:
    normalized = _normalize_document_json(document_json)
    columns = _normalize_document_columns(normalized)
    rows = [_normalize_sheet_row(row, len(columns)) for row in _extract_document_rows(normalized)]
    action_data_row = _find_excel_import_action_data_row(
        normalized,
        rows,
        action_document_row=action_document_row,
        action_column=action_column,
    )
    return rows[: max(action_data_row + 1, 0)] if action_data_row is not None else []


def _filter_cell_meta_for_document_row_prefix(cell_meta: Any, *, max_document_row: int, column_count: int) -> dict[str, Any]:
    if not isinstance(cell_meta, dict):
        return {}
    filtered: dict[str, Any] = {}
    for key, meta in cell_meta.items():
        parsed = _parse_cell_meta_key(key)
        if parsed is None:
            continue
        row_index, column_index = parsed
        if 0 <= row_index < max_document_row and 0 <= column_index < column_count:
            filtered[str(key)] = meta
    return filtered


def _filter_merged_cells_for_document_row_prefix(merged_cells: Any, *, max_document_row: int, column_count: int) -> list[Any]:
    if not isinstance(merged_cells, list):
        return []
    filtered: list[Any] = []
    for item in merged_cells:
        if not isinstance(item, dict):
            continue
        try:
            row = int(item.get("row") or 0)
            col = int(item.get("col") or 0)
            rowspan = max(int(item.get("rowspan") or 1), 1)
            colspan = max(int(item.get("colspan") or 1), 1)
        except (TypeError, ValueError):
            continue
        if row < 0 or col < 0 or col >= column_count:
            continue
        if row + rowspan <= max_document_row:
            filtered.append({"row": row, "col": col, "rowspan": rowspan, "colspan": min(colspan, column_count - col)})
    return filtered


def _filter_entity_model_for_document_row_prefix(
    document_json: dict[str, Any],
    *,
    max_document_row: int,
) -> dict[str, Any]:
    entity_rows = _extract_document_entity_rows(document_json)
    entity_cells = _extract_document_entity_cells(document_json)
    if not entity_rows and not entity_cells:
        return document_json

    next_document = dict(document_json)
    kept_rows = entity_rows[:max(max_document_row, 0)]
    kept_row_ids = {
        row_id
        for row_id in (_get_document_entity_row_id(row) for row in kept_rows)
        if row_id
    }
    if entity_rows:
        next_document["entity_rows"] = kept_rows
    if entity_cells:
        next_document["entity_cells"] = {
            row_id: cells
            for row_id, cells in entity_cells.items()
            if row_id in kept_row_ids
        }
    return next_document


def _get_excel_import_field_row_index(document_json: dict[str, Any], data_start_row: int) -> int:
    try:
        index = int(document_json.get("field_row_index") or 0)
    except (TypeError, ValueError):
        index = 0
    if 0 <= index < data_start_row:
        return index
    return max(data_start_row - 1, 0)


def _append_document_extra_columns_for_excel_import(
    document_json: dict[str, Any],
    extra_columns: list[str],
) -> tuple[dict[str, Any], list[str]]:
    normalized = _normalize_document_json(document_json)
    columns = _normalize_document_columns(normalized)
    appended_columns: list[str] = []
    used_keys = {_normalize_import_record_key(column) for column in columns}
    for header in extra_columns:
        _append_excel_import_extra_column_header(appended_columns, used_keys, header)

    if not appended_columns:
        return normalized, []

    extra_count = len(appended_columns)
    next_columns = [*columns, *appended_columns]
    next_document = {
        **normalized,
        "columns": next_columns,
    }

    next_document["rows"] = [
        [*_normalize_sheet_row(row, len(columns)), *([""] * extra_count)]
        for row in _extract_document_rows(normalized)
    ]

    source_widths = normalized.get("column_widths")
    if isinstance(source_widths, list):
        next_document["column_widths"] = [
            *source_widths[:len(columns)],
            *([NOTE_SHEET_EXCEL_IMPORT_EXTRA_COLUMN_WIDTH] * extra_count),
        ]
    else:
        next_document["column_widths"] = [NOTE_SHEET_EXCEL_IMPORT_EXTRA_COLUMN_WIDTH] * len(next_columns)

    grid_rows = _extract_document_grid_rows(normalized)
    if grid_rows:
        data_start_row = min(_normalize_document_data_start_row(normalized), len(grid_rows))
        field_row_index = _get_excel_import_field_row_index(normalized, data_start_row)
        next_grid_rows: list[list[Any]] = []
        for row_index, row in enumerate(grid_rows):
            suffix = appended_columns if row_index == field_row_index else [""] * extra_count
            next_grid_rows.append([*_normalize_sheet_row(row, len(columns)), *suffix])
        next_document["grid_rows"] = next_grid_rows

    source_configs = normalized.get("column_configs")
    next_configs = dict(source_configs) if isinstance(source_configs, dict) else {}
    for header in appended_columns:
        current = next_configs.get(header)
        config = dict(current) if isinstance(current, dict) else {}
        config.setdefault("header_background_color", NOTE_SHEET_EXCEL_IMPORT_EXTRA_COLUMN_HEADER_BACKGROUND)
        config.setdefault("header_text_color", NOTE_SHEET_EXCEL_IMPORT_EXTRA_COLUMN_HEADER_TEXT)
        next_configs[header] = config
    next_document["column_configs"] = next_configs

    if "header_groups" in normalized:
        next_document["header_groups"] = _insert_columns_into_header_groups(
            normalized.get("header_groups"),
            len(columns),
            extra_count,
        )

    return next_document, appended_columns


def _replace_document_rows_for_excel_import(
    document_json: dict[str, Any],
    import_rows: list[list[Any]],
    *,
    extra_columns: list[str] | None = None,
    action_document_row: int | None = None,
    action_column: int | None = None,
) -> tuple[dict[str, Any], int]:
    normalized = _normalize_document_json(document_json)
    normalized, _appended_columns = _append_document_extra_columns_for_excel_import(normalized, extra_columns or [])
    columns = _normalize_document_columns(normalized)
    column_count = len(columns)
    preserved_rows = _get_excel_import_preserved_data_rows(
        normalized,
        action_document_row=action_document_row,
        action_column=action_column,
    )
    normalized_import_rows = [_normalize_sheet_row(row, column_count) for row in import_rows]
    if _is_registration_append_sheet(columns):
        normalized_import_rows = _normalize_registration_import_sequences(
            preserved_rows,
            normalized_import_rows,
            columns,
            append_mode=False,
        )
    next_document = _replace_document_data_rows(normalized, [*preserved_rows, *normalized_import_rows])

    data_start_row = _normalize_document_data_start_row(next_document)
    max_preserved_document_row = data_start_row + len(preserved_rows)
    next_document["cell_meta"] = _filter_cell_meta_for_document_row_prefix(
        next_document.get("cell_meta"),
        max_document_row=max_preserved_document_row,
        column_count=column_count,
    )
    next_document["merged_cells"] = _filter_merged_cells_for_document_row_prefix(
        next_document.get("merged_cells"),
        max_document_row=max_preserved_document_row,
        column_count=column_count,
    )
    next_document = _filter_entity_model_for_document_row_prefix(
        next_document,
        max_document_row=max_preserved_document_row,
    )
    if _is_registration_append_sheet(columns):
        next_document = _repair_registration_sequence_column_config(next_document)
    return next_document, len(preserved_rows)


def _get_column_index(columns: list[str], header: str) -> int:
    try:
        return columns.index(header)
    except ValueError:
        return -1


def _registration_cutoff_date(now: date | None = None) -> date:
    current = now or date.today()
    month_index = current.year * 12 + (current.month - 1) - 2
    year = month_index // 12
    month = month_index % 12 + 1
    if month == 12:
        next_month = date(year + 1, 1, 1)
    else:
        next_month = date(year, month + 1, 1)
    last_day = (next_month - timedelta(days=1)).day
    return date(year, month, min(current.day, last_day))


def _add_registration_months(value: date, months: int) -> date:
    month_index = value.year * 12 + (value.month - 1) + months
    year = month_index // 12
    month = month_index % 12 + 1
    if month == 12:
        next_month = date(year + 1, 1, 1)
    else:
        next_month = date(year, month + 1, 1)
    last_day = (next_month - timedelta(days=1)).day
    return date(year, month, min(value.day, last_day))


def _parse_registration_submitted_at_datetime(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        parts = _serial_to_date_parts(float(value))
        return datetime(parts[0], parts[1], parts[2]) if parts else None

    text = _normalize_sheet_text(value)
    if not text:
        return None
    normalized = text.replace("T", " ")
    if re.fullmatch(r"-?\d+(?:\.\d+)?", normalized):
        parts = _serial_to_date_parts(float(normalized))
        return datetime(parts[0], parts[1], parts[2]) if parts else None
    for fmt in (
        "%Y-%m-%d %H:%M:%S",
        "%Y/%m/%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y/%m/%d %H:%M",
        "%Y年%m月%d日 %H:%M:%S",
        "%Y年%m月%d日 %H:%M",
        "%Y-%m-%d",
        "%Y/%m/%d",
        "%Y年%m月%d日",
        "%Y%m%d%H%M%S",
        "%Y%m%d%H%M",
        "%Y%m%d",
    ):
        try:
            return datetime.strptime(normalized, fmt)
        except ValueError:
            continue
    return None


def _format_registration_tracking_date(value: date) -> str:
    return value.isoformat()


def _format_registration_tracking_datetime(value: datetime) -> str:
    return value.strftime("%Y-%m-%d %H:%M:%S")


def _apply_registration_tracking_values_to_row(
    row: list[Any],
    columns: list[str],
    *,
    now: date | None = None,
) -> bool:
    group_index = _get_column_index(columns, NOTE_SHEET_REGISTRATION_TRACKING_GROUP_COLUMN)
    status_index = _get_column_index(columns, NOTE_SHEET_REGISTRATION_TRACKING_STATUS_COLUMN)
    deadline_index = _get_column_index(columns, NOTE_SHEET_REGISTRATION_TRACKING_DEADLINE_COLUMN)
    frozen_at_index = _get_column_index(columns, NOTE_SHEET_REGISTRATION_FROZEN_AT_COLUMN)
    if group_index < 0 and status_index < 0 and deadline_index < 0 and frozen_at_index < 0:
        return False

    submitted_at_index = _get_column_index(columns, NOTE_SHEET_REGISTRATION_SUBMITTED_AT_COLUMN)
    submitted_at = (
        _parse_registration_submitted_at_datetime(row[submitted_at_index])
        if submitted_at_index >= 0
        else None
    )
    current_date = now or date.today()
    current_datetime = datetime.combine(current_date, datetime.now().time()) if now is not None else datetime.now()
    changed = False

    def set_cell(column_index: int, value: str) -> None:
        nonlocal changed
        if column_index < 0 or column_index >= len(row):
            return
        if _normalize_sheet_text(row[column_index]) == value:
            return
        row[column_index] = value
        changed = True

    if submitted_at is None:
        set_cell(group_index, "")
        set_cell(status_index, "")
        set_cell(deadline_index, "")
        set_cell(frozen_at_index, "")
        return changed

    submitted_date = submitted_at.date()
    archived = submitted_date < _registration_cutoff_date(current_date)
    set_cell(group_index, NOTE_SHEET_REGISTRATION_FROZEN_GROUP if archived else NOTE_SHEET_REGISTRATION_ACTIVE_GROUP)
    set_cell(status_index, NOTE_SHEET_REGISTRATION_FROZEN_STATUS if archived else NOTE_SHEET_REGISTRATION_ACTIVE_STATUS)
    set_cell(deadline_index, _format_registration_tracking_date(_add_registration_months(submitted_date, 2)))
    if archived:
        if frozen_at_index >= 0 and not _normalize_sheet_text(row[frozen_at_index]):
            set_cell(frozen_at_index, _format_registration_tracking_datetime(current_datetime))
    else:
        set_cell(frozen_at_index, "")
    return changed


def _is_registration_append_sheet(columns: list[str]) -> bool:
    required = {
        NOTE_SHEET_REGISTRATION_SEQUENCE_COLUMN,
        NOTE_SHEET_REGISTRATION_GROUP_COLUMN,
        NOTE_SHEET_REGISTRATION_SUBMITTED_AT_COLUMN,
        "姓名",
    }
    return required.issubset(set(columns))


def _is_archived_registration_row(row: list[Any], columns: list[str], *, now: date | None = None) -> bool:
    submitted_at_index = _get_column_index(columns, NOTE_SHEET_REGISTRATION_SUBMITTED_AT_COLUMN)
    if submitted_at_index < 0:
        return False
    submitted_at = _parse_registration_submitted_at_datetime(row[submitted_at_index])
    return submitted_at is not None and submitted_at.date() < _registration_cutoff_date(now)


def _registration_submitted_at_sort_parts(value: Any) -> tuple[int, int] | None:
    submitted_at = _parse_registration_submitted_at_datetime(value)
    if submitted_at is None:
        return None
    return submitted_at.date().toordinal(), (
        submitted_at.hour * 3600
        + submitted_at.minute * 60
        + submitted_at.second
    )


def _registration_dynamic_group_sort_key(
    row: list[Any],
    columns: list[str],
    source_index: int,
    *,
    now: date | None = None,
) -> tuple[Any, ...]:
    submitted_at_index = _get_column_index(columns, NOTE_SHEET_REGISTRATION_SUBMITTED_AT_COLUMN)
    submitted_parts = (
        _registration_submitted_at_sort_parts(row[submitted_at_index])
        if submitted_at_index >= 0
        else None
    )
    if submitted_parts is None:
        return 0, 1, source_index

    day, second = submitted_parts
    if day >= _registration_cutoff_date(now).toordinal():
        return 0, 0, day, second, source_index
    return 1, 0, -day, -second, source_index


def _is_registration_archived_style(style: Any) -> bool:
    if not isinstance(style, dict):
        return False
    return (
        _normalize_sheet_text(style.get("background_color")).upper() == NOTE_SHEET_REGISTRATION_ARCHIVED_BACKGROUND
        and _normalize_sheet_text(style.get("text_color")).upper() == NOTE_SHEET_REGISTRATION_ARCHIVED_TEXT
    )


def _apply_registration_archived_row_styles(
    document_json: dict[str, Any],
    *,
    now: date | None = None,
) -> dict[str, Any]:
    normalized = _normalize_document_json(document_json)
    columns = _normalize_document_columns(normalized)
    rows = [_normalize_sheet_row(row, len(columns)) for row in _extract_document_rows(normalized)]
    if not rows or not _is_registration_append_sheet(columns):
        return normalized

    data_start_row = _normalize_document_data_start_row(normalized)
    cell_meta = dict(normalized.get("cell_meta")) if isinstance(normalized.get("cell_meta"), dict) else {}
    changed = False

    for data_row_index, row in enumerate(rows):
        document_row = data_start_row + data_row_index
        archived = _is_archived_registration_row(row, columns, now=now)
        for column_index in range(len(columns)):
            key = f"{document_row}:{column_index}"
            current_meta = cell_meta.get(key)
            next_meta = dict(current_meta) if isinstance(current_meta, dict) else {}
            current_style = next_meta.get("style")
            next_style = dict(current_style) if isinstance(current_style, dict) else {}
            if archived:
                if (
                    next_style.get("background_color") == NOTE_SHEET_REGISTRATION_ARCHIVED_BACKGROUND
                    and next_style.get("text_color") == NOTE_SHEET_REGISTRATION_ARCHIVED_TEXT
                ):
                    continue
                next_style["background_color"] = NOTE_SHEET_REGISTRATION_ARCHIVED_BACKGROUND
                next_style["text_color"] = NOTE_SHEET_REGISTRATION_ARCHIVED_TEXT
                next_meta["style"] = next_style
                cell_meta[key] = next_meta
                changed = True
                continue

            if not _is_registration_archived_style(next_style):
                continue
            next_style.pop("background_color", None)
            next_style.pop("text_color", None)
            if next_style:
                next_meta["style"] = next_style
            else:
                next_meta.pop("style", None)
            if next_meta:
                cell_meta[key] = next_meta
            else:
                cell_meta.pop(key, None)
            changed = True

    if not changed and normalized.get("cell_meta") == cell_meta:
        return normalized
    next_document = dict(normalized)
    next_document["cell_meta"] = cell_meta
    return next_document


def _remap_document_entity_data_rows(
    document_json: dict[str, Any],
    source_document_json: dict[str, Any],
    *,
    row_index_map: dict[int, int],
    data_row_count: int,
) -> dict[str, Any]:
    source_entity_rows = _extract_document_entity_rows(source_document_json)
    if not source_entity_rows:
        return document_json

    data_start_row = _normalize_document_data_start_row(source_document_json)
    source_data_entity_rows = source_entity_rows[data_start_row:data_start_row + data_row_count]
    if not source_data_entity_rows:
        return document_json

    next_data_entity_rows: list[Any | None] = [None] * data_row_count
    for source_index, target_index in row_index_map.items():
        if (
            0 <= source_index < len(source_data_entity_rows)
            and 0 <= target_index < data_row_count
        ):
            next_data_entity_rows[target_index] = source_data_entity_rows[source_index]

    if any(row is None for row in next_data_entity_rows):
        return document_json

    next_document = dict(document_json)
    next_document["entity_rows"] = [
        *source_entity_rows[:data_start_row],
        *(row for row in next_data_entity_rows if row is not None),
        *source_entity_rows[data_start_row + data_row_count:],
    ]
    if "entity_cells" in source_document_json:
        next_document["entity_cells"] = source_document_json["entity_cells"]
    if "entity_columns" in source_document_json:
        next_document["entity_columns"] = source_document_json["entity_columns"]
    return next_document


def _order_registration_rows_by_dynamic_expiration(
    document_json: dict[str, Any],
    *,
    now: date | None = None,
) -> dict[str, Any]:
    normalized = _normalize_document_json(document_json)
    columns = _normalize_document_columns(normalized)
    if not _is_registration_append_sheet(columns):
        return normalized
    if _get_column_index(columns, NOTE_SHEET_REGISTRATION_SUBMITTED_AT_COLUMN) < 0:
        return normalized

    rows: list[list[Any]] = []
    tracking_changed = False
    for row in _extract_document_rows(normalized):
        next_row = _normalize_sheet_row(row, len(columns))
        tracking_changed = _apply_registration_tracking_values_to_row(next_row, columns, now=now) or tracking_changed
        rows.append(next_row)
    ordered_source_indexes = sorted(
        range(len(rows)),
        key=lambda index: _registration_dynamic_group_sort_key(rows[index], columns, index, now=now),
    )
    if ordered_source_indexes == list(range(len(rows))) and not tracking_changed:
        return _apply_registration_archived_row_styles(normalized, now=now)

    row_index_map = {
        source_index: target_index
        for target_index, source_index in enumerate(ordered_source_indexes)
    }
    formula_row_offset = _get_formula_reference_row_offset(normalized)
    next_rows = [
        _remap_row_formula_cell_references(
            rows[source_index],
            columns=columns,
            row_index_map=row_index_map,
            row_index_offset=formula_row_offset,
        )
        for source_index in ordered_source_indexes
    ]
    next_document = _replace_document_data_rows({
        **normalized,
        "columns": columns,
    }, next_rows)
    if isinstance(normalized.get("cell_meta"), dict):
        next_document["cell_meta"] = _remap_cell_meta_rows(
            normalized.get("cell_meta"),
            row_index_map,
            row_offset=_normalize_document_data_start_row(normalized),
        )
    next_document = _remap_document_entity_data_rows(
        next_document,
        normalized,
        row_index_map=row_index_map,
        data_row_count=len(rows),
    )
    return _apply_registration_archived_row_styles(next_document, now=now)


def _get_registration_active_row_count(rows: list[list[Any]], columns: list[str]) -> int:
    if not _is_registration_append_sheet(columns):
        return len(rows)
    return sum(1 for row in rows if not _is_archived_registration_row(row, columns))


CHINESE_INTEGER_DIGITS = {
    "零": 0,
    "〇": 0,
    "一": 1,
    "二": 2,
    "两": 2,
    "三": 3,
    "四": 4,
    "五": 5,
    "六": 6,
    "七": 7,
    "八": 8,
    "九": 9,
}
REGISTRATION_GROUP_SEQUENCE_RE = re.compile(
    r"^\s*(?P<group>\d{1,3})\s*(?:[_\-－–—]|组)\s*(?P<number>\d{1,4})\s*(?:号)?\s*$"
)


def _parse_chinese_integer(value: str) -> int | None:
    text = value.strip()
    if not text:
        return None
    if text.isdigit():
        return int(text)
    if any(char not in (set(CHINESE_INTEGER_DIGITS) | {"十"}) for char in text):
        return None
    if "十" in text:
        left, _sep, right = text.partition("十")
        tens = 1 if not left else CHINESE_INTEGER_DIGITS.get(left)
        ones = 0 if not right else CHINESE_INTEGER_DIGITS.get(right)
        if tens is None or ones is None:
            return None
        return tens * 10 + ones
    if len(text) == 1:
        return CHINESE_INTEGER_DIGITS.get(text)
    digits = [CHINESE_INTEGER_DIGITS.get(char) for char in text]
    if any(digit is None for digit in digits):
        return None
    return int("".join(str(digit) for digit in digits if digit is not None))


def _parse_registration_group_number(value: Any) -> str | None:
    text = _normalize_sheet_text(value)
    if not text:
        return None
    match = re.search(r"(?:第\s*)?(\d{1,3})\s*(?:小组|组)", text)
    if match:
        return str(int(match.group(1)))
    match = re.search(r"(?:第\s*)?([零〇一二两三四五六七八九十]{1,4})\s*(?:小组|组)", text)
    if match:
        number = _parse_chinese_integer(match.group(1))
        return str(number) if number and number > 0 else None
    return None


def _parse_registration_group_sequence(value: Any) -> tuple[str, int, int] | None:
    text = _normalize_sheet_text(value)
    if not text:
        return None
    match = REGISTRATION_GROUP_SEQUENCE_RE.fullmatch(text)
    if not match:
        return None
    member_text = match.group("number")
    member = int(member_text)
    return str(int(match.group("group"))), member, max(len(member_text), 2)


def _parse_registration_plain_sequence(value: Any) -> int | None:
    text = _normalize_sheet_text(value)
    if not text:
        return None
    if re.fullmatch(r"\d+(?:\.0+)?", text):
        return int(float(text))
    return None


def _format_registration_group_sequence(group_number: str, member_number: int, *, width: int = 2) -> str:
    return f"{int(group_number)}_{int(member_number):0{max(width, 2)}d}"


def _registration_rows_have_group_scoped_sequences(rows: list[list[Any]], sequence_index: int) -> bool:
    if sequence_index < 0:
        return False
    return any(
        _parse_registration_group_sequence(row[sequence_index] if sequence_index < len(row) else "") is not None
        for row in rows
    )


def _registration_import_has_group_sequence_reset(
    import_rows: list[list[Any]],
    columns: list[str],
) -> bool:
    sequence_index = _get_column_index(columns, NOTE_SHEET_REGISTRATION_SEQUENCE_COLUMN)
    group_index = _get_column_index(columns, NOTE_SHEET_REGISTRATION_GROUP_COLUMN)
    if sequence_index < 0 or group_index < 0:
        return False

    pairs: list[tuple[str, int]] = []
    for row in import_rows:
        group_number = _parse_registration_group_number(row[group_index] if group_index < len(row) else "")
        sequence_number = _parse_registration_plain_sequence(row[sequence_index] if sequence_index < len(row) else "")
        if group_number and sequence_number is not None:
            pairs.append((group_number, sequence_number))

    if len({group for group, _sequence in pairs}) < 2:
        return False

    sequence_groups: dict[int, set[str]] = {}
    for group_number, sequence_number in pairs:
        sequence_groups.setdefault(sequence_number, set()).add(group_number)
    if any(len(groups) > 1 for groups in sequence_groups.values()):
        return True

    previous_group = ""
    previous_sequence: int | None = None
    for group_number, sequence_number in pairs:
        if previous_group and group_number != previous_group and previous_sequence is not None and sequence_number <= previous_sequence:
            return True
        previous_group = group_number
        previous_sequence = sequence_number
    return False


def _should_use_registration_group_scoped_sequences(
    existing_rows: list[list[Any]],
    import_rows: list[list[Any]],
    columns: list[str],
) -> bool:
    sequence_index = _get_column_index(columns, NOTE_SHEET_REGISTRATION_SEQUENCE_COLUMN)
    return (
        _registration_rows_have_group_scoped_sequences(existing_rows, sequence_index)
        or _registration_rows_have_group_scoped_sequences(import_rows, sequence_index)
        or _registration_import_has_group_sequence_reset(import_rows, columns)
    )


def _normalize_registration_import_sequences(
    existing_rows: list[list[Any]],
    import_rows: list[list[Any]],
    columns: list[str],
    *,
    append_mode: bool,
) -> list[list[Any]]:
    sequence_index = _get_column_index(columns, NOTE_SHEET_REGISTRATION_SEQUENCE_COLUMN)
    if sequence_index < 0 or not import_rows:
        return import_rows

    group_index = _get_column_index(columns, NOTE_SHEET_REGISTRATION_GROUP_COLUMN)
    use_group_scoped = _should_use_registration_group_scoped_sequences(existing_rows, import_rows, columns)
    used_sequences = {
        _normalize_sheet_text(row[sequence_index])
        for row in existing_rows
        if sequence_index < len(row) and _normalize_sheet_text(row[sequence_index])
    }

    next_global_sequence = 1
    group_state: dict[str, tuple[int, int]] = {}
    for row in existing_rows:
        value = row[sequence_index] if sequence_index < len(row) else ""
        grouped = _parse_registration_group_sequence(value)
        if grouped is not None:
            group_number, member_number, width = grouped
            current_member, current_width = group_state.get(group_number, (0, 2))
            group_state[group_number] = (max(current_member, member_number), max(current_width, width))
            continue
        plain = _parse_registration_plain_sequence(value)
        if plain is not None:
            next_global_sequence = max(next_global_sequence, plain + 1)

    def next_group_sequence(group_number: str) -> str:
        current_member, current_width = group_state.get(group_number, (0, 2))
        next_member = current_member + 1
        group_state[group_number] = (next_member, current_width)
        return _format_registration_group_sequence(group_number, next_member, width=current_width)

    normalized_rows: list[list[Any]] = []
    for raw_row in import_rows:
        row = list(raw_row)
        while len(row) <= sequence_index:
            row.append("")

        current_sequence = _normalize_sheet_text(row[sequence_index])
        if use_group_scoped:
            grouped = _parse_registration_group_sequence(current_sequence)
            group_number = (
                _parse_registration_group_number(row[group_index])
                if group_index >= 0 and group_index < len(row)
                else None
            )
            if grouped is not None:
                sequence_group, member_number, width = grouped
                candidate_group = group_number or sequence_group
                candidate = _format_registration_group_sequence(candidate_group, member_number, width=width)
                current_member, current_width = group_state.get(candidate_group, (0, 2))
                group_state[candidate_group] = (max(current_member, member_number), max(current_width, width))
            else:
                plain = _parse_registration_plain_sequence(current_sequence)
                candidate = (
                    _format_registration_group_sequence(group_number, plain, width=2)
                    if group_number and plain is not None
                    else current_sequence
                )
                if group_number and plain is not None:
                    current_member, current_width = group_state.get(group_number, (0, 2))
                    group_state[group_number] = (max(current_member, plain), current_width)

            if append_mode and (not candidate or candidate in used_sequences):
                fallback_group = (
                    group_number
                    or (grouped[0] if grouped is not None else None)
                )
                if fallback_group:
                    candidate = next_group_sequence(fallback_group)
            row[sequence_index] = candidate
            if candidate:
                used_sequences.add(candidate)
        elif append_mode:
            row[sequence_index] = str(next_global_sequence)
            used_sequences.add(row[sequence_index])
            next_global_sequence += 1

        normalized_rows.append(row)
    return normalized_rows


def _coerce_registration_import_rows(
    existing_rows: list[list[Any]],
    import_rows: list[list[Any]],
    columns: list[str],
) -> list[list[Any]]:
    if not import_rows or not _is_registration_append_sheet(columns):
        return import_rows

    sequence_index = _get_column_index(columns, NOTE_SHEET_REGISTRATION_SEQUENCE_COLUMN)
    group_index = _get_column_index(columns, NOTE_SHEET_REGISTRATION_GROUP_COLUMN)
    order_index = _get_column_index(columns, "微信支付订单号")
    existing_order_ids: set[str] = set()
    if order_index >= 0:
        existing_order_ids = {
            order_id
            for order_id in (_strip_legacy_text_prefix(row[order_index]) for row in existing_rows)
            if order_id and order_id != "/"
        }
    default_group = ""
    if group_index >= 0:
        for row in reversed(existing_rows):
            if _is_archived_registration_row(row, columns):
                continue
            default_group = _normalize_sheet_text(row[group_index])
            if default_group:
                break

    coerced_rows: list[list[Any]] = []
    for row in import_rows:
        if order_index >= 0:
            order_id = _strip_legacy_text_prefix(row[order_index])
            if order_id and order_id != "/":
                if order_id in existing_order_ids:
                    continue
                existing_order_ids.add(order_id)
        next_row = list(row)
        if group_index >= 0 and not _normalize_sheet_text(next_row[group_index]) and default_group:
            next_row[group_index] = default_group
        coerced_rows.append(next_row)
    return _normalize_registration_import_sequences(
        existing_rows,
        coerced_rows,
        columns,
        append_mode=True,
    )


def _filter_duplicate_excel_import_payment_order_rows(
    document_json: dict[str, Any],
    import_rows: list[list[Any]],
    columns: list[str],
) -> tuple[list[list[Any]], int]:
    order_index = _get_column_index(columns, "微信支付订单号")
    if order_index < 0 or not import_rows:
        return import_rows, 0

    existing_order_ids = {
        order_id
        for row in (
            _normalize_sheet_row(raw_row, len(columns))
            for raw_row in _extract_document_rows(_normalize_document_json(document_json))
        )
        if (order_id := _strip_legacy_text_prefix(row[order_index])) and order_id != "/"
    }

    filtered_rows: list[list[Any]] = []
    skipped_count = 0
    for raw_row in import_rows:
        row = _normalize_sheet_row(raw_row, len(columns))
        order_id = _strip_legacy_text_prefix(row[order_index])
        if order_id and order_id != "/":
            if order_id in existing_order_ids:
                skipped_count += 1
                continue
            existing_order_ids.add(order_id)
        filtered_rows.append(row)

    return filtered_rows, skipped_count


def _repair_group_scoped_sequence_column_config(document_json: dict[str, Any], column_name: str) -> dict[str, Any]:
    normalized = _normalize_document_json(document_json)
    columns = _normalize_document_columns(normalized)

    sequence_index = _get_column_index(columns, column_name)
    if sequence_index < 0:
        return document_json

    rows = [_normalize_sheet_row(row, len(columns)) for row in _extract_document_rows(normalized)]
    if not _registration_rows_have_group_scoped_sequences(rows, sequence_index):
        return document_json

    configs = normalized.get("column_configs")
    if not isinstance(configs, dict):
        return document_json

    current = configs.get(column_name)
    if not isinstance(current, dict) or "value_type" not in current:
        return document_json

    next_config = dict(current)
    next_config.pop("value_type", None)
    next_configs = dict(configs)
    if next_config:
        next_configs[column_name] = next_config
    else:
        next_configs.pop(column_name, None)

    next_document = dict(normalized)
    next_document["column_configs"] = next_configs
    return next_document


def _repair_registration_sequence_column_config(document_json: dict[str, Any]) -> dict[str, Any]:
    normalized = _normalize_document_json(document_json)
    if not _is_registration_append_sheet(_normalize_document_columns(normalized)):
        return document_json
    return _repair_group_scoped_sequence_column_config(normalized, NOTE_SHEET_REGISTRATION_SEQUENCE_COLUMN)


def _repair_attendance_student_id_column_config(document_json: dict[str, Any]) -> dict[str, Any]:
    return _repair_group_scoped_sequence_column_config(document_json, "学号")


def _insert_document_data_rows_for_excel_import(
    document_json: dict[str, Any],
    insert_index: int,
    import_rows: list[list[Any]],
) -> dict[str, Any]:
    normalized = _normalize_document_json(document_json)
    columns = _normalize_document_columns(normalized)
    existing_rows = [_normalize_sheet_row(row, len(columns)) for row in _extract_document_rows(normalized)]
    safe_insert_index = max(0, min(int(insert_index), len(existing_rows)))
    normalized_import_rows = [_normalize_sheet_row(row, len(columns)) for row in import_rows]
    next_rows = [
        *existing_rows[:safe_insert_index],
        *normalized_import_rows,
        *existing_rows[safe_insert_index:],
    ]
    next_document = _replace_document_data_rows(normalized, next_rows)
    if isinstance(normalized.get("cell_meta"), dict) and normalized_import_rows:
        next_document["cell_meta"] = _shift_cell_meta_rows_for_insert(
            normalized.get("cell_meta"),
            safe_insert_index,
            len(normalized_import_rows),
            row_offset=_normalize_document_data_start_row(normalized),
        )
    if safe_insert_index < len(existing_rows):
        next_document = _filter_entity_model_for_document_row_prefix(
            next_document,
            max_document_row=_normalize_document_data_start_row(next_document),
        )
    return next_document


def _append_document_rows_for_excel_import(
    document_json: dict[str, Any],
    import_rows: list[list[Any]],
    *,
    extra_columns: list[str] | None = None,
) -> tuple[dict[str, Any], int]:
    normalized = _normalize_document_json(document_json)
    normalized, _appended_columns = _append_document_extra_columns_for_excel_import(normalized, extra_columns or [])
    columns = _normalize_document_columns(normalized)
    column_count = len(columns)
    existing_rows = [_normalize_sheet_row(row, column_count) for row in _extract_document_rows(normalized)]
    normalized_import_rows = [_normalize_sheet_row(row, column_count) for row in import_rows]
    if _is_registration_append_sheet(columns):
        preserved_row_count = _get_registration_active_row_count(existing_rows, columns)
        coerced_rows = _coerce_registration_import_rows(existing_rows, normalized_import_rows, columns)
        next_document = _replace_document_data_rows(normalized, [*existing_rows, *coerced_rows])
        next_document = _filter_entity_model_for_document_row_prefix(
            next_document,
            max_document_row=_normalize_document_data_start_row(next_document),
        )
        next_document = _order_registration_rows_by_dynamic_expiration(next_document)
        return _repair_registration_sequence_column_config(next_document), preserved_row_count

    return (
        _insert_document_data_rows_for_excel_import(normalized, len(existing_rows), normalized_import_rows),
        len(existing_rows),
    )


def _load_attendance_order_lookup_provider():
    _load_attendance_kqdb_provider()
    try:
        from kq5034.attendance_api import lookup_order  # type: ignore

        return lookup_order
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"无法加载订单匹配工具：{exc}") from exc


def _load_attendance_user_lookup_provider():
    try:
        from kq5034.attendance_api import lookup_registration_user_db  # type: ignore

        return lookup_registration_user_db
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"无法加载用户匹配工具：{exc}") from exc


def _format_registration_match_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _find_required_registration_column_indexes(columns: list[str], required_columns: list[str]) -> dict[str, int]:
    indexes: dict[str, int] = {}
    missing: list[str] = []
    for header in required_columns:
        index = _find_column_index(columns, header)
        if index is None:
            missing.append(header)
        else:
            indexes[header] = index
    if missing:
        raise HTTPException(status_code=400, detail=f"报名表缺少字段：{', '.join(missing)}")
    return indexes


def _find_registration_column_indexes(
    columns: list[str],
    required_columns: list[str],
    *,
    optional_columns: list[str] | tuple[str, ...] = (),
) -> dict[str, int]:
    indexes = _find_required_registration_column_indexes(columns, required_columns)
    for header in optional_columns:
        index = _find_column_index(columns, header)
        if index is not None:
            indexes[header] = index
    return indexes


def _normalize_registration_order_lookup_mode() -> str:
    value = str(NOTE_SHEET_REGISTRATION_ORDER_LOOKUP_MODE or "").strip().lower()
    return value if value in {"hybrid", "db_only", "browser_only"} else "db_only"


def _build_registration_match_summary(**overrides: int) -> dict[str, int]:
    summary = {
        "target_count": 0,
        "updated_count": 0,
        "skipped_count": 0,
        "error_count": 0,
        "warning_count": 0,
        "matched_count": 0,
        "unmatched_count": 0,
        "invalid_count": 0,
        "already_complete_count": 0,
    }
    for key, value in overrides.items():
        summary[key] = int(value or 0)
    return summary


def _format_registration_match_clause(
    label: str,
    action: str,
    summary: dict[str, int],
    *,
    zero_text: str,
    unmatched_text: str,
) -> str:
    target_count = int(summary.get("target_count") or 0)
    matched_count = int(summary.get("matched_count") or 0)
    updated_count = int(summary.get("updated_count") or 0)
    if target_count <= 0:
        return f"{label}：{zero_text}"

    details: list[str] = []
    unmatched_count = int(summary.get("unmatched_count") or 0)
    error_count = int(summary.get("error_count") or 0)
    if unmatched_count:
        details.append(f"{unmatched_count} 条{unmatched_text}")
    if error_count:
        details.append(f"{error_count} 条异常")
    suffix = f"（{'，'.join(details)}）" if details else ""
    if matched_count:
        return f"{label}：{action} {matched_count}/{target_count}{suffix}"
    return f"{label}：{action} 0/{target_count}，更新 {updated_count} 行{suffix}"


def _format_registration_order_match_message(summary: dict[str, int]) -> str:
    return _format_registration_match_clause(
        "订单匹配",
        "补全",
        summary,
        zero_text="没有待补全订单",
        unmatched_text="本地订单库未命中",
    )


def _format_registration_user_match_message(summary: dict[str, int]) -> str:
    return _format_registration_match_clause(
        "用户匹配",
        "命中",
        summary,
        zero_text="没有待匹配用户",
        unmatched_text="未命中用户",
    )


def _format_registration_attendance_sync_message(summary: dict[str, int]) -> str:
    updated_count = int(summary.get("updated_count") or 0)
    inserted_count = int(summary.get("inserted_count") if "inserted_count" in summary else updated_count)
    repaired_count = int(summary.get("repaired_count") or 0)
    skipped_count = int(summary.get("skipped_count") or 0)
    error_count = int(summary.get("error_count") or 0)
    details: list[str] = []
    if skipped_count:
        details.append(f"{skipped_count} 条已存在或缺少用户ID")
    if error_count:
        details.append(f"{error_count} 条异常")
    suffix = f"（{'，'.join(details)}）" if details else ""
    if repaired_count:
        return f"考勤同步：新增 {inserted_count} 行，修复 {repaired_count} 行{suffix}"
    return f"考勤新增 {inserted_count} 行{suffix}"


def _format_registration_composite_update_message(
    order_summary: dict[str, int],
    user_summary: dict[str, int],
    attendance_summary: dict[str, int],
) -> str:
    return "综合更新完成：" + "；".join([
        _format_registration_order_match_message(order_summary),
        _format_registration_user_match_message(user_summary),
        _format_registration_attendance_sync_message(attendance_summary),
    ])


def _registration_attendance_row_identity(row: list[Any], columns: list[str]) -> tuple[str, str, str]:
    user_id_index = _get_column_index(columns, "用户ID")
    student_id_index = _get_column_index(columns, NOTE_SHEET_REGISTRATION_SEQUENCE_COLUMN)
    merchant_order_index = _get_column_index(columns, "商户订单号")
    user_id = _normalize_sheet_text(row[user_id_index]) if user_id_index >= 0 else ""
    student_id = _normalize_sheet_text(row[student_id_index]) if student_id_index >= 0 else ""
    merchant_order_id = (
        _strip_legacy_text_prefix(row[merchant_order_index])
        if merchant_order_index >= 0
        else ""
    )
    return user_id, student_id, merchant_order_id


def _find_registration_attendance_existing_index(
    *,
    user_id: str,
    student_id: str,
    merchant_order_id: str,
    existing_user_id_rows: dict[str, int],
    existing_student_id_rows: dict[str, int],
    existing_merchant_order_id_rows: dict[str, int],
) -> int | None:
    existing_index = existing_user_id_rows.get(user_id) if user_id else None
    if existing_index is None and student_id:
        existing_index = existing_student_id_rows.get(student_id)
    if existing_index is None and merchant_order_id:
        existing_index = existing_merchant_order_id_rows.get(merchant_order_id)
    return existing_index


def _registration_user_browser_timeout_seconds() -> float:
    try:
        return max(float(NOTE_SHEET_REGISTRATION_USER_BROWSER_TIMEOUT_SECONDS), 1.0)
    except (TypeError, ValueError):
        return 900.0


def _build_remote_device_headers(entry: UserDevice) -> dict[str, str]:
    return {
        "Authorization": f"Bearer {entry.token}",
        "X-Device-Token": entry.token,
    }


def _resolve_registration_user_browser_device(
    session: Session,
    current_user: User,
) -> UserDevice:
    config = get_or_create_attendance_service_config(session)
    configured_entry_id = _normalize_sheet_text(config.execution_device_entry_id)
    if configured_entry_id:
        entry = session.get(UserDevice, configured_entry_id)
        if entry is None:
            raise HTTPException(status_code=400, detail="考勤配置的执行设备不存在，请在考勤配置页重新选择")
        if not entry.is_active:
            raise HTTPException(status_code=400, detail="考勤配置的执行设备已停用，请在考勤配置页重新选择")
    else:
        target_name = _normalize_sheet_text(NOTE_SHEET_REGISTRATION_USER_BROWSER_DEVICE_NAME)
        statement = select(UserDevice).where(
            UserDevice.user_id == current_user.id,
            UserDevice.is_active == True,  # noqa: E712
        )
        entry = next(
            (
                item
                for item in session.exec(statement).all()
                if item.name == target_name or item.device_id == target_name
            ),
            None,
        )
        if entry is None:
            raise HTTPException(status_code=400, detail=f"未找到可用的远程设备：{target_name}")

    if entry.mode != "remote" or not _normalize_sheet_text(entry.server_url):
        raise HTTPException(status_code=400, detail="小鹅通兜底回查必须使用远程执行设备，请在考勤配置页选择 codepc_mi15")
    if not _normalize_sheet_text(entry.token):
        raise HTTPException(status_code=400, detail="远程执行设备缺少访问令牌，请检查设备清单配置")
    return entry


def _lookup_registration_users_with_remote_browser(
    session: Session,
    current_user: User,
    *,
    course_name: str,
    shop_id: int,
    items: list[dict[str, Any]],
) -> dict[str, dict[str, Any]]:
    if not items:
        return {}

    entry = _resolve_registration_user_browser_device(session, current_user)
    server_url = _normalize_sheet_text(entry.server_url).rstrip("/")
    payload = {
        "course_name": course_name,
        "course_product_name": "",
        "shop_id": shop_id,
        "close_browser": True,
        "items": [
            {
                "key": _normalize_sheet_text(item.get("key")),
                "names": [text for text in item.get("names", []) if _normalize_sheet_text(text)],
                "phones": [text for text in item.get("phones", []) if _normalize_sheet_text(text)],
            }
            for item in items
        ],
    }

    try:
        import requests

        with requests.Session() as request_session:
            request_session.trust_env = False
            response = request_session.post(
                f"{server_url}/api/device-control/attendance/user-match/lookup",
                json=payload,
                headers=_build_remote_device_headers(entry),
                timeout=_registration_user_browser_timeout_seconds(),
            )
    except Exception as exc:
        raise HTTPException(status_code=502, detail=f"调用 {entry.name} 小鹅通回查失败：{exc}") from exc

    if response.status_code >= 400:
        try:
            detail = response.json().get("detail")
        except Exception:
            detail = response.text.strip()
        raise HTTPException(status_code=502, detail=detail or f"{entry.name} 小鹅通回查失败，HTTP {response.status_code}")

    try:
        data = response.json()
    except Exception as exc:
        raise HTTPException(status_code=502, detail=f"{entry.name} 小鹅通回查返回了无法解析的响应") from exc

    result_map: dict[str, dict[str, Any]] = {}
    result_items = data.get("results") if isinstance(data, dict) else []
    if not isinstance(result_items, list):
        result_items = []
    for item in result_items:
        if not isinstance(item, dict):
            continue
        key = _normalize_sheet_text(item.get("key"))
        if key:
            result_map[key] = item
    return result_map


def _lookup_registration_orders_with_remote_browser(
    session: Session,
    current_user: User,
    *,
    order_ids: list[str],
) -> list[dict[str, Any]]:
    if not order_ids:
        return []

    entry = _resolve_registration_user_browser_device(session, current_user)
    server_url = _normalize_sheet_text(entry.server_url).rstrip("/")
    extra_config = get_attendance_service_extra_config(session)
    payload = {
        "action": "inspect",
        "rows": [{"微信支付订单号": order_id, "商户订单号": order_id} for order_id in order_ids],
        "login_users": list(extra_config.get("scan_reminder_users") or []),
        "lookup_mode": "browser_only",
    }

    try:
        import requests

        with requests.Session() as request_session:
            request_session.trust_env = False
            response = request_session.post(
                f"{server_url}/api/device-control/attendance/order/execute",
                json=payload,
                headers=_build_remote_device_headers(entry),
                timeout=600,
            )
    except Exception as exc:
        raise HTTPException(status_code=502, detail=f"调用 {entry.name} 实时查单失败：{exc}") from exc

    if response.status_code >= 400:
        try:
            detail = response.json().get("detail")
        except Exception:
            detail = response.text.strip()
        raise HTTPException(status_code=502, detail=detail or f"{entry.name} 实时查单失败，HTTP {response.status_code}")

    try:
        data = response.json()
    except Exception as exc:
        raise HTTPException(status_code=502, detail=f"{entry.name} 实时查单返回了无法解析的响应") from exc

    rows = data.get("rows") if isinstance(data, dict) else []
    if not isinstance(rows, list):
        rows = []
    result_rows: list[dict[str, Any]] = []
    for index in range(len(order_ids)):
        row = rows[index] if index < len(rows) else {}
        result_rows.append(dict(row) if isinstance(row, dict) else {})
    return result_rows


def _clockin_link_detection_timeout_seconds() -> float:
    try:
        return max(float(NOTE_SHEET_CLOCKIN_LINK_DETECTION_TIMEOUT_SECONDS), 30.0)
    except (TypeError, ValueError):
        return 600.0


def _defined_name_literal_value(formula: Any) -> str:
    text = _normalize_defined_name_formula(formula)
    if text.startswith("="):
        text = text[1:].strip()
    return _unquote_formula_string(text).strip()


def _get_effective_defined_name_literal(
    session: Session,
    document: SheetDocument,
    workbook: WorkbookDocument | None,
    name: str,
) -> str:
    target_key = _normalize_sheet_text(name).lower()
    for item in reversed(_merge_effective_defined_names(
        _get_workbook_defined_names(session, workbook),
        _get_sheet_defined_names(dict(document.document_json or {})),
    )):
        if _normalize_sheet_text(item.get("name")).lower() == target_key:
            return _defined_name_literal_value(item.get("formula"))
    return ""


def _normalize_clockin_target_name(value: Any) -> str:
    text = re.sub(r"\s+", "", _normalize_sheet_text(value))
    for suffix in ("打卡链接", "打卡数据", "打卡"):
        if text.endswith(suffix):
            text = text[: -len(suffix)]
    return text


def _clockin_config_column_indexes(columns: list[str]) -> dict[str, int]:
    required = ["name", "url"]
    missing: list[str] = []
    indexes: dict[str, int] = {}
    for header in required:
        index = _find_column_index(columns, header)
        if index is None:
            missing.append(header)
        else:
            indexes[header] = index
    if missing:
        raise HTTPException(status_code=400, detail=f"打卡配置缺少字段：{', '.join(missing)}")
    for optional in ("start_date", "end_date", "days", "clockin_user_num", "total_user_num"):
        index = _find_column_index(columns, optional)
        if index is not None:
            indexes[optional] = index
    return indexes


def _extract_clockin_detection_targets(document_json: dict[str, Any]) -> list[dict[str, Any]]:
    normalized = _normalize_document_json(document_json)
    columns = _normalize_document_columns(normalized)
    indexes = _clockin_config_column_indexes(columns)
    rows = [_normalize_sheet_row(row, len(columns)) for row in _extract_document_rows(normalized)]
    result: list[dict[str, Any]] = []
    for row_index, row in enumerate(rows):
        name = _normalize_sheet_text(row[indexes["name"]])
        target = _normalize_clockin_target_name(name)
        if not target:
            continue
        result.append({
            "row_index": row_index,
            "name": name,
            "target": target,
        })
    return result


def _detect_clockin_links_with_remote_browser(
    session: Session,
    current_user: User,
    *,
    root_url: str,
    targets: list[str],
    provider_id: str,
    model: str,
) -> dict[str, Any]:
    entry = _resolve_registration_user_browser_device(session, current_user)
    server_url = _normalize_sheet_text(entry.server_url).rstrip("/")
    payload = {
        "root_url": root_url,
        "targets": targets,
        "provider_id": provider_id,
        "model": model,
        "close_tabs": True,
    }
    try:
        import requests

        with requests.Session() as request_session:
            request_session.trust_env = False
            response = request_session.post(
                f"{server_url}/api/device-control/attendance/clockin-links/detect",
                json=payload,
                headers=_build_remote_device_headers(entry),
                timeout=_clockin_link_detection_timeout_seconds(),
            )
    except Exception as exc:
        try:
            return _detect_clockin_links_with_remote_python_run(
                entry,
                root_url=root_url,
                targets=targets,
                provider_id=provider_id,
                model=model,
            )
        except HTTPException as fallback_exc:
            raise HTTPException(
                status_code=502,
                detail=f"调用 {entry.name} 自动检测打卡链接失败：{exc}；python-run 兜底也失败：{fallback_exc.detail}",
            ) from exc

    if response.status_code in {404, 405}:
        return _detect_clockin_links_with_remote_python_run(
            entry,
            root_url=root_url,
            targets=targets,
            provider_id=provider_id,
            model=model,
        )

    if response.status_code >= 400:
        try:
            detail = response.json().get("detail")
        except Exception:
            detail = response.text.strip()
        raise HTTPException(status_code=502, detail=detail or f"{entry.name} 自动检测打卡链接失败，HTTP {response.status_code}")

    try:
        data = response.json()
    except Exception as exc:
        raise HTTPException(status_code=502, detail=f"{entry.name} 自动检测打卡链接返回了无法解析的响应") from exc
    return dict(data) if isinstance(data, dict) else {}


def _detect_clockin_links_with_remote_python_run(
    entry: UserDevice,
    *,
    root_url: str,
    targets: list[str],
    provider_id: str,
    model: str,
) -> dict[str, Any]:
    detector_path = Path(__file__).resolve().parents[1] / "core" / "clockin_link_detector.py"
    try:
        detector_source = detector_path.read_text(encoding="utf-8")
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"无法读取打卡链接探测脚本：{exc}") from exc

    script = (
        "import json\n"
        f"_source = {json.dumps(detector_source)}\n"
        "_ns = {}\n"
        "exec(_source, _ns)\n"
        "result = _ns['detect_clockin_links_browser'](\n"
        f"    root_url={json.dumps(root_url, ensure_ascii=False)},\n"
        f"    targets={json.dumps(targets, ensure_ascii=False)},\n"
        f"    provider_id={json.dumps(provider_id, ensure_ascii=False)},\n"
        f"    model={json.dumps(model, ensure_ascii=False)},\n"
        "    close_tabs=True,\n"
        ")\n"
        "print(json.dumps(result, ensure_ascii=False))\n"
    )
    server_url = _normalize_sheet_text(entry.server_url).rstrip("/")
    try:
        import requests

        with requests.Session() as request_session:
            request_session.trust_env = False
            response = request_session.post(
                f"{server_url}/api/device-control/python-runs",
                json={
                    "mode": "script",
                    "script": script,
                    "timeout": int(_clockin_link_detection_timeout_seconds()),
                },
                headers=_build_remote_device_headers(entry),
                timeout=_clockin_link_detection_timeout_seconds() + 30,
            )
    except Exception as exc:
        raise HTTPException(status_code=502, detail=f"调用 {entry.name} python-run 检测打卡链接失败：{exc}") from exc

    if response.status_code >= 400:
        try:
            detail = response.json().get("detail")
        except Exception:
            detail = response.text.strip()
        raise HTTPException(status_code=502, detail=detail or f"{entry.name} python-run 检测打卡链接失败，HTTP {response.status_code}")

    try:
        payload = response.json()
    except Exception as exc:
        raise HTTPException(status_code=502, detail=f"{entry.name} python-run 返回了无法解析的响应") from exc
    if payload.get("status") != "completed" or payload.get("returncode") not in (0, None):
        error = _normalize_sheet_text(payload.get("error")) or _normalize_sheet_text(payload.get("stderr"))
        raise HTTPException(status_code=502, detail=error or f"{entry.name} python-run 检测打卡链接未成功")
    stdout = _normalize_sheet_text(payload.get("stdout"))
    try:
        return json.loads(stdout)
    except Exception as exc:
        match = re.search(r"\{.*\}", stdout, re.S)
        if match:
            try:
                return json.loads(match.group(0))
            except Exception:
                pass
        raise HTTPException(status_code=502, detail=f"{entry.name} python-run 检测结果不是有效 JSON") from exc


def _apply_clockin_link_detection_results(
    document_json: dict[str, Any],
    targets: list[dict[str, Any]],
    detection_result: dict[str, Any],
) -> tuple[dict[str, Any], dict[str, Any]]:
    normalized = _normalize_document_json(document_json)
    columns = _normalize_document_columns(normalized)
    indexes = _clockin_config_column_indexes(columns)
    rows = [_normalize_sheet_row(row, len(columns)) for row in _extract_document_rows(normalized)]
    result_items = detection_result.get("results") if isinstance(detection_result.get("results"), list) else []
    result_by_target = {
        _normalize_clockin_target_name(item.get("target")): item
        for item in result_items
        if isinstance(item, dict) and _normalize_clockin_target_name(item.get("target"))
    }

    updated_count = 0
    skipped_count = 0
    error_count = 0
    warnings: list[str] = []
    written: list[dict[str, Any]] = []

    for target in targets:
        row_index = int(target["row_index"])
        row = rows[row_index] if 0 <= row_index < len(rows) else None
        if row is None:
            skipped_count += 1
            continue
        target_name = _normalize_clockin_target_name(target.get("target"))
        item = result_by_target.get(target_name)
        if not item:
            error_count += 1
            warnings.append(f"未检测到 {target.get('name') or target_name} 的打卡链接")
            continue
        url = _normalize_sheet_text(item.get("url"))
        if not url:
            error_count += 1
            warnings.append(f"{target.get('name') or target_name} 的检测结果没有 URL")
            continue

        before = list(row)
        row[indexes["url"]] = url
        for field in ("start_date", "end_date", "days"):
            if field in indexes and item.get(field) not in (None, ""):
                row[indexes[field]] = item.get(field)
        if row != before:
            updated_count += 1
            rows[row_index] = row
        else:
            skipped_count += 1
        written.append({
            "row_index": row_index,
            "name": target.get("name"),
            "target": target_name,
            "url": url,
        })

    next_document = _replace_document_data_rows(normalized, rows)
    remote_warnings = [str(item) for item in (detection_result.get("warnings") or [])]
    return next_document, {
        "updated_count": updated_count,
        "skipped_count": skipped_count,
        "error_count": error_count,
        "warning_count": len(warnings) + len(remote_warnings),
        "warnings": [*warnings, *remote_warnings],
        "written": written,
    }


def _order_info_has_fill_values(order_info: Any, fill_columns: list[str] | None = None) -> bool:
    if not isinstance(order_info, dict):
        return False
    return any(
        _normalize_sheet_text(order_info.get(column))
        for column in (fill_columns or NOTE_SHEET_REGISTRATION_ORDER_COLUMNS[1:])
    )


def _derive_registration_order_month(order_id: Any) -> str:
    normalized = _strip_legacy_text_prefix(order_id)
    if not normalized:
        return ""
    merchant_timestamp_match = re.search(r"\bMA(20\d{2})(\d{2})(\d{2})", normalized, re.I)
    if merchant_timestamp_match:
        year = int(merchant_timestamp_match.group(1))
        month = int(merchant_timestamp_match.group(2))
        day = int(merchant_timestamp_match.group(3))
        if 1 <= month <= 12 and 1 <= day <= 31:
            return f"{year:04d}{month:02d}"
    for match in re.finditer(r"(\d{6})", normalized):
        value = match.group(1)
        year = int(value[:2])
        month = int(value[2:4])
        day = int(value[4:6])
        if 20 <= year <= 39 and 1 <= month <= 12 and 1 <= day <= 31:
            return f"20{year:02d}{month:02d}"
    return ""


def _registration_order_lookup_id(row: list[Any], indexes: dict[str, int]) -> str:
    for column in ("微信支付订单号", "商户订单号"):
        index = indexes.get(column)
        if index is None or index >= len(row):
            continue
        value = _strip_legacy_text_prefix(row[index])
        if value:
            return value
    return ""


def _registration_row_has_identity_payload(row: list[Any], columns: list[str]) -> bool:
    """Return whether a registration row contains real learner/source data.

    Order fields alone are not enough to make a row active: blank template rows
    can accidentally retain merchant order ids or amounts. Those rows must not
    be sent through order matching or attendance synchronization.
    """

    indexes = {column: index for index, column in enumerate(columns)}
    identity_columns = (
        "序号",
        "提交时间",
        "姓名",
        "微信昵称",
        "手机号",
        "错误手机号",
        "用户ID",
        NOTE_SHEET_REGISTRATION_LINKED_USER_ID_COLUMN,
    )
    for column in identity_columns:
        index = indexes.get(column)
        if index is not None and index < len(row) and _normalize_sheet_text(row[index]):
            return True
    return False


def _registration_row_has_completed_refund_remark(row: list[Any], indexes: dict[str, int]) -> bool:
    remark_index = indexes.get("备注")
    if remark_index is None or remark_index >= len(row):
        return False
    remark = _normalize_sheet_text(row[remark_index])
    if not remark:
        return False
    return bool(re.search(r"(?:已\s*退\s*(?:费|款|课)|退款\s*成功|退费\s*完成|退款\s*完成)", remark))


def _registration_order_amount_from_row_or_info(
    row: list[Any],
    indexes: dict[str, int],
    order_info: Any = None,
) -> float | None:
    if isinstance(order_info, dict):
        amount = _parse_number_sort_value(order_info.get("订单金额"))
        if amount is not None:
            return amount
    amount_index = indexes.get("订单金额")
    if amount_index is not None and amount_index < len(row):
        return _parse_number_sort_value(row[amount_index])
    return None


def _registration_refunded_amount_from_row_or_info(
    row: list[Any],
    indexes: dict[str, int],
    order_info: Any = None,
) -> float | None:
    if isinstance(order_info, dict):
        refunded = _parse_number_sort_value(order_info.get("已返款"))
        if refunded is not None:
            return refunded
    refunded_index = indexes.get("已返款")
    if refunded_index is not None and refunded_index < len(row):
        return _parse_number_sort_value(row[refunded_index])
    return None


def _registration_row_needs_refund_payment_audit(
    row: list[Any],
    indexes: dict[str, int],
    *,
    order_info: Any = None,
) -> bool:
    if not _registration_row_has_completed_refund_remark(row, indexes):
        return False
    refunded = _registration_refunded_amount_from_row_or_info(row, indexes, order_info)
    amount = _registration_order_amount_from_row_or_info(row, indexes, order_info)
    if refunded is None:
        return True
    if amount is not None and amount > 0:
        return refunded + 1e-9 < amount
    return refunded <= 0


_REGISTRATION_REFUND_AUDIT_NOTE_PREFIXES = (
    "支付复核异常：",
    "支付复核待确认：",
)


def _set_registration_refund_audit_note(row: list[Any], indexes: dict[str, int], note: str | None) -> bool:
    remark_index = indexes.get("备注")
    if remark_index is None or remark_index >= len(row):
        return False
    current = _normalize_sheet_text(row[remark_index])
    parts = [
        part.strip()
        for part in re.split(r"[;；]", current)
        if part.strip() and not part.strip().startswith(_REGISTRATION_REFUND_AUDIT_NOTE_PREFIXES)
    ]
    if note:
        parts.append(note)
    next_value = "；".join(parts)
    if next_value == current:
        return False
    row[remark_index] = next_value
    return True


def _format_registration_refund_audit_note(row: list[Any], indexes: dict[str, int], *, prefix: str) -> str:
    refunded = _registration_refunded_amount_from_row_or_info(row, indexes)
    amount = _registration_order_amount_from_row_or_info(row, indexes)
    if refunded is not None and amount is not None and amount > 0:
        refunded_text = _format_registration_match_cell(refunded)
        amount_text = _format_registration_match_cell(amount)
        return f"{prefix}源表标退费完成，但支付侧已返款{refunded_text}/{amount_text}"
    if refunded is not None:
        return f"{prefix}源表标退费完成，但支付侧已返款{_format_registration_match_cell(refunded)}"
    return f"{prefix}源表标退费完成，但支付侧未确认已返款"


def _apply_registration_order_info_to_row(
    row: list[Any],
    indexes: dict[str, int],
    order_info: dict[str, Any],
) -> None:
    for column in NOTE_SHEET_REGISTRATION_ORDER_COLUMNS:
        if column not in indexes:
            continue
        value = order_info.get(column) if isinstance(order_info, dict) else ""
        if value is not None and value != "":
            formatted_value = _format_registration_match_cell(value)
            if column == "微信支付订单号":
                formatted_value = _strip_legacy_text_prefix(formatted_value)
            row[indexes[column]] = formatted_value
    if not _normalize_sheet_text(row[indexes["订单日期"]]):
        row[indexes["订单日期"]] = _derive_registration_order_month(row[indexes["微信支付订单号"]])


def _clear_registration_order_optional_refund_value(row: list[Any], indexes: dict[str, int]) -> None:
    refunded_index = indexes.get("已返款")
    if refunded_index is not None:
        row[refunded_index] = ""


def _update_registration_order_match_document(
    document_json: dict[str, Any],
    *,
    session: Session | None = None,
    current_user: User | None = None,
    use_browser_fallback: bool = False,
) -> tuple[dict[str, Any], dict[str, int]]:
    normalized = _normalize_document_json(document_json)
    columns = _normalize_document_columns(normalized)
    indexes = _find_registration_column_indexes(
        columns,
        NOTE_SHEET_REGISTRATION_ORDER_REQUIRED_COLUMNS,
        optional_columns=NOTE_SHEET_REGISTRATION_ORDER_ACTION_OPTIONAL_COLUMNS,
    )
    rows = [_normalize_sheet_row(row, len(columns)) for row in _extract_document_rows(normalized)]
    if not rows:
        return normalized, _build_registration_match_summary()

    get_kqdb = _load_attendance_kqdb_provider()
    lookup_order = _load_attendance_order_lookup_provider()
    kqdb = get_kqdb()
    lookup_mode = _normalize_registration_order_lookup_mode()

    updated_count = 0
    skipped_count = 0
    error_count = 0
    warning_count = 0
    target_count = 0
    matched_count = 0
    unmatched_count = 0
    invalid_count = 0
    already_complete_count = 0
    next_rows: list[list[Any]] = []
    browser_candidates: list[dict[str, Any]] = []
    fill_columns = [column for column in NOTE_SHEET_REGISTRATION_ORDER_COLUMNS if column in indexes]
    required_completeness_columns = [column for column in NOTE_SHEET_REGISTRATION_ORDER_REQUIRED_COLUMNS if column in indexes]
    completeness_columns = fill_columns[1:]

    for source_row in rows:
        row = list(source_row)
        if not _registration_row_has_identity_payload(row, columns):
            skipped_count += 1
            invalid_count += 1
            next_rows.append(row)
            continue
        order_id = _registration_order_lookup_id(row, indexes)
        if len(order_id) < 8:
            skipped_count += 1
            invalid_count += 1
            next_rows.append(row)
            continue

        needs_refund_audit = _registration_row_needs_refund_payment_audit(row, indexes)
        has_complete_order_info = all(_normalize_sheet_text(row[indexes[column]]) for column in required_completeness_columns)
        if has_complete_order_info and not needs_refund_audit:
            skipped_count += 1
            already_complete_count += 1
            next_rows.append(row)
            continue

        target_count += 1
        before = list(row)
        if (
            not _normalize_sheet_text(row[indexes["订单日期"]])
            and all(_normalize_sheet_text(row[indexes[column]]) for column in completeness_columns if column != "订单日期")
            and not needs_refund_audit
        ):
            derived_order_month = _derive_registration_order_month(order_id)
            if derived_order_month:
                row[indexes["订单日期"]] = derived_order_month
                updated_count += 1
                matched_count += 1
                next_rows.append(row)
                continue

        try:
            order_info = lookup_order(
                order_id,
                kqdb=kqdb,
                lookup_mode=lookup_mode,
                use_browser=lookup_mode != "db_only",
            )
        except Exception as exc:
            row[indexes["订单金额"]] = str(exc)
            _clear_registration_order_optional_refund_value(row, indexes)
            error_count += 1
            next_rows.append(row)
            if row != before:
                updated_count += 1
            continue

        if not order_info:
            if use_browser_fallback and session is not None and current_user is not None:
                browser_candidates.append({
                    "row_position": len(next_rows),
                    "order_id": order_id,
                    "audit": needs_refund_audit,
                    "already_matched": False,
                })
            else:
                skipped_count += 1
                unmatched_count += 1
                warning_count += 1
                if needs_refund_audit:
                    if _set_registration_refund_audit_note(
                        row,
                        indexes,
                        _format_registration_refund_audit_note(row, indexes, prefix="支付复核待确认："),
                    ):
                        updated_count += 1
            next_rows.append(row)
            continue
        if isinstance(order_info, dict) and "error" in order_info:
            row[indexes["订单金额"]] = _format_registration_match_cell(order_info.get("error") or "订单不存在")
            _clear_registration_order_optional_refund_value(row, indexes)
            error_count += 1
            next_rows.append(row)
            if row != before:
                updated_count += 1
            continue

        _apply_registration_order_info_to_row(row, indexes, order_info)
        matched_count += 1
        needs_refund_audit = _registration_row_needs_refund_payment_audit(row, indexes, order_info=order_info)
        if needs_refund_audit:
            if use_browser_fallback and session is not None and current_user is not None:
                browser_candidates.append({
                    "row_position": len(next_rows),
                    "order_id": order_id,
                    "audit": True,
                    "already_matched": True,
                })
            else:
                warning_count += 1
                _set_registration_refund_audit_note(
                    row,
                    indexes,
                    _format_registration_refund_audit_note(row, indexes, prefix="支付复核异常："),
                )
        elif _registration_row_has_completed_refund_remark(row, indexes):
            _set_registration_refund_audit_note(row, indexes, None)
        if row != before:
            updated_count += 1
        next_rows.append(row)

    if browser_candidates:
        try:
            browser_rows = _lookup_registration_orders_with_remote_browser(
                session,
                current_user,
                order_ids=[str(candidate["order_id"]) for candidate in browser_candidates],
            )
        except Exception as exc:
            error_text = str(exc.detail if isinstance(exc, HTTPException) else exc)
            for candidate in browser_candidates:
                row = next_rows[int(candidate["row_position"])]
                before = list(row)
                if candidate.get("audit"):
                    _set_registration_refund_audit_note(
                        row,
                        indexes,
                        _format_registration_refund_audit_note(row, indexes, prefix="支付复核待确认："),
                    )
                    warning_count += 1
                    if row != before:
                        updated_count += 1
                    continue
                row[indexes["订单金额"]] = error_text
                _clear_registration_order_optional_refund_value(row, indexes)
                error_count += 1
                if row != before:
                    updated_count += 1
        else:
            for candidate, order_info in zip(browser_candidates, browser_rows, strict=False):
                row = next_rows[int(candidate["row_position"])]
                before = list(row)
                if isinstance(order_info, dict) and _normalize_sheet_text(order_info.get("error")):
                    if candidate.get("audit"):
                        _set_registration_refund_audit_note(
                            row,
                            indexes,
                            _format_registration_refund_audit_note(row, indexes, prefix="支付复核待确认："),
                        )
                        warning_count += 1
                        if row != before:
                            updated_count += 1
                        continue
                    row[indexes["订单金额"]] = _normalize_sheet_text(order_info.get("error"))
                    _clear_registration_order_optional_refund_value(row, indexes)
                    error_count += 1
                    if row != before:
                        updated_count += 1
                    continue
                if not _order_info_has_fill_values(order_info, completeness_columns):
                    if candidate.get("audit"):
                        _set_registration_refund_audit_note(
                            row,
                            indexes,
                            _format_registration_refund_audit_note(row, indexes, prefix="支付复核待确认："),
                        )
                        warning_count += 1
                        if row != before:
                            updated_count += 1
                    else:
                        skipped_count += 1
                        unmatched_count += 1
                        warning_count += 1
                    continue

                _apply_registration_order_info_to_row(row, indexes, order_info)
                if not candidate.get("already_matched"):
                    matched_count += 1
                if _registration_row_needs_refund_payment_audit(row, indexes, order_info=order_info):
                    warning_count += 1
                    _set_registration_refund_audit_note(
                        row,
                        indexes,
                        _format_registration_refund_audit_note(row, indexes, prefix="支付复核异常："),
                    )
                elif _registration_row_has_completed_refund_remark(row, indexes):
                    _set_registration_refund_audit_note(row, indexes, None)
                if row != before:
                    updated_count += 1

    return _replace_document_data_rows(normalized, next_rows), _build_registration_match_summary(
        target_count=target_count,
        updated_count=updated_count,
        skipped_count=skipped_count,
        error_count=error_count,
        warning_count=warning_count,
        matched_count=matched_count,
        unmatched_count=unmatched_count,
        invalid_count=invalid_count,
        already_complete_count=already_complete_count,
    )


def _count_registration_order_match_targets(document_json: dict[str, Any]) -> int:
    normalized = _normalize_document_json(document_json)
    columns = _normalize_document_columns(normalized)
    indexes = _find_registration_column_indexes(
        columns,
        NOTE_SHEET_REGISTRATION_ORDER_REQUIRED_COLUMNS,
        optional_columns=NOTE_SHEET_REGISTRATION_ORDER_ACTION_OPTIONAL_COLUMNS,
    )
    rows = [_normalize_sheet_row(row, len(columns)) for row in _extract_document_rows(normalized)]
    count = 0
    completeness_columns = [
        column
        for column in NOTE_SHEET_REGISTRATION_ORDER_COLUMNS[1:]
        if column in indexes
    ]
    for row in rows:
        order_id = _registration_order_lookup_id(row, indexes)
        if len(order_id) < 8:
            continue
        has_complete_order_info = all(_normalize_sheet_text(row[indexes[column]]) for column in completeness_columns)
        if has_complete_order_info and not _registration_row_needs_refund_payment_audit(row, indexes):
            continue
        count += 1
    return count


def _get_registration_course_name(document: SheetDocument, workbook: WorkbookDocument | None) -> str:
    def normalize_course_name(value: str) -> str:
        text = _normalize_sheet_text(value)
        text = re.sub(r"\.[^.]+$", "", text)
        text = re.sub(r"^\d{2}(\d{6})", r"d\1", text)
        return text.replace(".", "点").replace(",", "")

    if workbook is not None and _normalize_sheet_text(workbook.title):
        return normalize_course_name(workbook.title)
    return normalize_course_name(document.title)


def _resolve_registration_shop_id(
    session: Session,
    document: SheetDocument,
    workbook: WorkbookDocument | None,
) -> int:
    for name in ("店铺ID", "店铺编号", "shop_id"):
        raw_value = _get_effective_defined_name_literal(session, document, workbook, name)
        if raw_value:
            try:
                shop_id = int(float(raw_value))
            except (TypeError, ValueError):
                continue
            if shop_id in {1, 2}:
                return shop_id

    shop_name = _get_effective_defined_name_literal(session, document, workbook, "店铺名")
    if shop_name == "宗门学府":
        return 2
    if shop_name == "5034山中薪":
        return 1

    text = " ".join(
        item
        for item in [
            _get_registration_course_name(document, workbook),
            _normalize_sheet_text(workbook.title if workbook is not None else ""),
            _normalize_sheet_text(document.title),
        ]
        if item
    )
    if "禅宗" in text or "修道班" in text:
        return 2
    return NOTE_SHEET_REGISTRATION_DEFAULT_SHOP_ID


def _update_registration_user_match_document(
    document_json: dict[str, Any],
    *,
    session: Session,
    current_user: User,
    course_name: str,
    shop_id: int,
    use_browser_fallback: bool,
) -> tuple[dict[str, Any], dict[str, int]]:
    normalized = _normalize_document_json(document_json)
    columns = _normalize_document_columns(normalized)
    indexes = _find_required_registration_column_indexes(columns, NOTE_SHEET_REGISTRATION_USER_LOOKUP_COLUMNS)
    rows = [_normalize_sheet_row(row, len(columns)) for row in _extract_document_rows(normalized)]
    if not rows:
        return normalized, _build_registration_match_summary()

    get_kqdb = _load_attendance_kqdb_provider()
    lookup_user = _load_attendance_user_lookup_provider()
    kqdb = get_kqdb()

    updated_count = 0
    skipped_count = 0
    error_count = 0
    warning_count = 0
    target_count = 0
    matched_count = 0
    unmatched_count = 0
    already_complete_count = 0
    invalid_count = 0
    next_rows: list[list[Any]] = []
    browser_candidates: list[dict[str, Any]] = []
    updated_row_positions: set[int] = set()

    def mark_updated(row_position: int) -> None:
        nonlocal updated_count
        if row_position in updated_row_positions:
            return
        updated_row_positions.add(row_position)
        updated_count += 1

    for source_row in rows:
        row = list(source_row)
        if _normalize_sheet_text(row[indexes["用户ID"]]):
            skipped_count += 1
            already_complete_count += 1
            next_rows.append(row)
            continue

        names = [
            _normalize_sheet_text(row[indexes["姓名"]]),
            _normalize_sheet_text(row[indexes["微信昵称"]]),
        ]
        phones = [
            _strip_legacy_text_prefix(row[indexes["手机号"]]),
            _strip_legacy_text_prefix(row[indexes["错误手机号"]]),
        ]
        names = [item for item in names if item]
        phones = [item for item in phones if item and item.lower() != "none"]
        if not names and not phones:
            skipped_count += 1
            invalid_count += 1
            next_rows.append(row)
            continue

        target_count += 1
        before = list(row)
        try:
            user_id, weight = lookup_user(
                names,
                phones,
                course_name=course_name,
                course_product_name="",
                shop_id=shop_id,
                return_mode=1,
                kqdb=kqdb,
            )
        except Exception as exc:
            row[indexes["匹配得分"]] = str(exc)
            error_count += 1
            next_rows.append(row)
            if row != before:
                updated_count += 1
            continue

        row[indexes["用户ID"]] = _format_registration_match_cell(user_id)
        row[indexes["匹配得分"]] = _format_registration_match_cell(weight) if weight is not None else ""
        row_position = len(next_rows)
        initial_changed = row != before
        if initial_changed and (user_id or not use_browser_fallback):
            mark_updated(row_position)
        if user_id:
            matched_count += 1
        elif not use_browser_fallback:
            unmatched_count += 1
            warning_count += 1
        next_rows.append(row)
        if not user_id and use_browser_fallback:
            browser_candidates.append(
                {
                    "key": str(len(browser_candidates)),
                    "row_position": row_position,
                    "initial_changed": initial_changed,
                    "names": names,
                    "phones": phones,
                }
            )

    if use_browser_fallback and browser_candidates:
        try:
            browser_results = _lookup_registration_users_with_remote_browser(
                session,
                current_user,
                course_name=course_name,
                shop_id=shop_id,
                items=[
                    {
                        "key": candidate["key"],
                        "names": candidate["names"],
                        "phones": candidate["phones"],
                    }
                    for candidate in browser_candidates
                ],
            )
        except Exception as exc:
            if isinstance(exc, HTTPException):
                error_text = str(exc.detail)
            else:
                error_text = str(exc)
            for candidate in browser_candidates:
                row = next_rows[int(candidate["row_position"])]
                before = list(row)
                row[indexes["匹配得分"]] = error_text
                error_count += 1
                if row != before:
                    mark_updated(int(candidate["row_position"]))
        else:
            for candidate in browser_candidates:
                result = browser_results.get(str(candidate["key"]), {})
                row = next_rows[int(candidate["row_position"])]
                before = list(row)
                result_error = _normalize_sheet_text(result.get("error"))
                remote_user_id = _normalize_sheet_text(result.get("user_id"))
                if result_error:
                    row[indexes["匹配得分"]] = result_error
                    error_count += 1
                elif remote_user_id:
                    row[indexes["用户ID"]] = remote_user_id
                    row[indexes["匹配得分"]] = "95"
                    matched_count += 1
                else:
                    unmatched_count += 1
                    warning_count += 1
                if row != before or bool(candidate.get("initial_changed")):
                    mark_updated(int(candidate["row_position"]))

    return _replace_document_data_rows(normalized, next_rows), _build_registration_match_summary(
        target_count=target_count,
        updated_count=updated_count,
        skipped_count=skipped_count,
        error_count=error_count,
        warning_count=warning_count,
        matched_count=matched_count,
        unmatched_count=unmatched_count,
        invalid_count=invalid_count,
        already_complete_count=already_complete_count,
    )


def _document_dict_rows_for_detection(document_json: dict[str, Any]) -> list[dict[str, Any]]:
    normalized = _normalize_document_json(document_json)
    columns = _normalize_document_columns(normalized)
    rows = [_normalize_sheet_row(row, len(columns)) for row in _extract_document_rows(normalized)]
    return [dict(zip(columns, row)) for row in rows]


def _ensure_registration_linked_user_id_column(document_json: dict[str, Any]) -> tuple[dict[str, Any], int]:
    normalized = _normalize_document_json(document_json)
    columns = _normalize_document_columns(normalized)
    if NOTE_SHEET_REGISTRATION_LINKED_USER_ID_COLUMN in columns:
        return normalized, columns.index(NOTE_SHEET_REGISTRATION_LINKED_USER_ID_COLUMN)

    if "参考信息" in columns:
        insert_index = columns.index("参考信息") + 1
    elif "匹配得分" in columns:
        insert_index = columns.index("匹配得分") + 1
    else:
        insert_index = len(columns)
    next_document = _insert_document_column(
        normalized,
        insert_index=insert_index,
        header=NOTE_SHEET_REGISTRATION_LINKED_USER_ID_COLUMN,
        width=220,
    )
    next_document, _style_changed = _apply_registration_standard_user_id_column_styles(next_document)
    next_columns = _normalize_document_columns(next_document)
    return next_document, next_columns.index(NOTE_SHEET_REGISTRATION_LINKED_USER_ID_COLUMN)


def _candidate_progress_total(candidate: NoteSheetRegistrationUserIdDetectionCandidate) -> int:
    return int(candidate.video_count or 0) + int(candidate.clockin_count or 0)


def _normalize_detection_phone(value: Any) -> str:
    text = re.sub(r"\D+", "", _strip_legacy_text_prefix(value))
    return text if len(text) >= 7 else ""


def _normalize_detection_label(value: Any) -> str:
    return re.sub(r"[\s_\-—－]+", "", _normalize_sheet_text(value)).lower()


def _registration_detection_student_label_variants(value: Any) -> set[str]:
    text = _normalize_sheet_text(value)
    if not text:
        return set()
    variants = {text, text.replace("_", "-"), text.replace("-", "_"), text.replace("_", ""), text.replace("-", "")}
    compact = _normalize_detection_label(text)
    if compact:
        variants.add(compact)
    return {item for item in variants if item}


def _append_registration_detection_note(row: list[Any], indexes: dict[str, int], message: str) -> None:
    index = indexes.get("参考信息")
    if index is None:
        return
    current = _normalize_sheet_text(row[index])
    lines = [line for line in current.splitlines() if not line.startswith("检测用户ID：")]
    lines.append(f"检测用户ID：{message}")
    row[index] = "\n".join(line for line in lines if line).strip()


def _collect_registration_course_user_progress(
    video_document: dict[str, Any] | None,
    clockin_document: dict[str, Any] | None,
) -> dict[str, dict[str, Any]]:
    progress: dict[str, dict[str, Any]] = {}

    def ensure(user_id: str) -> dict[str, Any]:
        item = progress.setdefault(user_id, {"video_count": 0, "clockin_count": 0, "labels": set()})
        return item

    for row in _document_dict_rows_for_detection(video_document or {}):
        user_id = _normalize_sheet_text(row.get("user_id2") or row.get("用户ID") or row.get("user_id"))
        if not user_id:
            continue
        item = ensure(user_id)
        item["video_count"] = int(item.get("video_count") or 0) + 1
        for key in ("remark_nm", "nickname", "用户昵称", "姓名", "微信昵称"):
            label = _normalize_sheet_text(row.get(key))
            if label:
                item["labels"].add(label)

    for row in _document_dict_rows_for_detection(clockin_document or {}):
        user_id = _normalize_sheet_text(row.get("user_id2") or row.get("用户ID") or row.get("user_id"))
        if not user_id:
            continue
        item = ensure(user_id)
        item["clockin_count"] = int(item.get("clockin_count") or 0) + 1
        for key in ("nickname", "remark_nm", "groupname", "extra_姓名", "extra_打卡昵称", "姓名", "微信昵称"):
            label = _normalize_sheet_text(row.get(key))
            if label:
                item["labels"].add(label)
    return progress


def _query_registration_detection_user_ids_by_phone(phone: str) -> list[str]:
    if not phone:
        return []
    try:
        kqdb = _load_attendance_kqdb_provider()()
    except HTTPException:
        raise
    except Exception:
        return []

    sql = (
        "SELECT user_id2, user_id, bind_phone, collect_phone "
        "FROM user_table WHERE bind_phone = %s OR collect_phone = %s LIMIT 50"
    )
    records: list[dict[str, Any]]
    try:
        records = kqdb.exec2dict(sql, [phone, phone])
    except Exception:
        safe_phone = phone.replace("'", "''")
        try:
            records = kqdb.exec2dict(sql.replace("%s", f"'{safe_phone}'"))
        except Exception:
            return []
    user_ids: list[str] = []
    seen: set[str] = set()
    for record in records or []:
        user_id = _normalize_sheet_text(record.get("user_id2") or record.get("user_id"))
        if user_id and user_id not in seen:
            seen.add(user_id)
            user_ids.append(user_id)
    return user_ids


def _make_registration_detection_candidate(
    user_id: str,
    progress: dict[str, dict[str, Any]],
    *,
    evidence: list[str],
    confidence: Literal["high", "medium", "low"],
) -> NoteSheetRegistrationUserIdDetectionCandidate | None:
    user_progress = progress.get(user_id)
    if not user_progress:
        return None
    video_count = int(user_progress.get("video_count") or 0)
    clockin_count = int(user_progress.get("clockin_count") or 0)
    if video_count + clockin_count <= 0:
        return None
    return NoteSheetRegistrationUserIdDetectionCandidate(
        user_id=user_id,
        video_count=video_count,
        clockin_count=clockin_count,
        evidence=evidence,
        confidence=confidence,
    )


def _merge_registration_detection_candidates(
    candidates: list[NoteSheetRegistrationUserIdDetectionCandidate],
) -> list[NoteSheetRegistrationUserIdDetectionCandidate]:
    by_user_id: dict[str, NoteSheetRegistrationUserIdDetectionCandidate] = {}
    rank = {"low": 0, "medium": 1, "high": 2}
    for candidate in candidates:
        existing = by_user_id.get(candidate.user_id)
        if existing is None:
            by_user_id[candidate.user_id] = candidate
            continue
        existing.video_count = max(existing.video_count, candidate.video_count)
        existing.clockin_count = max(existing.clockin_count, candidate.clockin_count)
        for evidence in candidate.evidence:
            if evidence not in existing.evidence:
                existing.evidence.append(evidence)
        if rank[candidate.confidence] > rank[existing.confidence]:
            existing.confidence = candidate.confidence
    return sorted(
        by_user_id.values(),
        key=lambda item: (rank[item.confidence], _candidate_progress_total(item)),
        reverse=True,
    )


def _build_registration_user_id_detection_candidates(
    row_map: dict[str, Any],
    progress: dict[str, dict[str, Any]],
) -> list[NoteSheetRegistrationUserIdDetectionCandidate]:
    candidates: list[NoteSheetRegistrationUserIdDetectionCandidate] = []
    phones = [
        ("手机号", _normalize_detection_phone(row_map.get("手机号"))),
        ("错误手机号", _normalize_detection_phone(row_map.get("错误手机号"))),
    ]
    for field, phone in phones:
        if not phone:
            continue
        for user_id in _query_registration_detection_user_ids_by_phone(phone):
            candidate = _make_registration_detection_candidate(
                user_id,
                progress,
                evidence=[f"{field}命中 {phone}"],
                confidence="high",
            )
            if candidate is not None:
                candidates.append(candidate)

    student_variants = _registration_detection_student_label_variants(
        row_map.get(NOTE_SHEET_REGISTRATION_SEQUENCE_COLUMN) or row_map.get("学号")
    )
    if student_variants:
        normalized_variants = {_normalize_detection_label(variant) for variant in student_variants}
        for user_id, item in progress.items():
            labels = {_normalize_detection_label(label) for label in item.get("labels") or []}
            if labels and any(
                variant and any(variant in label for label in labels)
                for variant in normalized_variants
            ):
                candidate = _make_registration_detection_candidate(
                    user_id,
                    progress,
                    evidence=["课程数据昵称/备注命中学号"],
                    confidence="high",
                )
                if candidate is not None:
                    candidates.append(candidate)

    identity_labels = {
        _normalize_detection_label(row_map.get("姓名")),
        _normalize_detection_label(row_map.get("微信昵称")),
        _normalize_detection_label(row_map.get("昵称")),
    }
    identity_labels.discard("")
    if identity_labels:
        for user_id, item in progress.items():
            labels = {_normalize_detection_label(label) for label in item.get("labels") or []}
            if labels.intersection(identity_labels):
                candidate = _make_registration_detection_candidate(
                    user_id,
                    progress,
                    evidence=["课程数据姓名/昵称命中"],
                    confidence="medium",
                )
                if candidate is not None:
                    candidates.append(candidate)
    return _merge_registration_detection_candidates(candidates)


def _format_registration_detection_candidate_summary(
    candidates: list[NoteSheetRegistrationUserIdDetectionCandidate],
    *,
    limit: int = 3,
) -> str:
    parts: list[str] = []
    for candidate in candidates[:limit]:
        evidence = "、".join(candidate.evidence[:2])
        parts.append(
            f"{candidate.user_id}(视频{candidate.video_count}/打卡{candidate.clockin_count}"
            f"/{candidate.confidence}{('，' + evidence) if evidence else ''})"
        )
    if len(candidates) > limit:
        parts.append(f"另有{len(candidates) - limit}个")
    return "；".join(parts)


def _rebuild_registration_attendance_after_user_id_detection(
    session: Session,
    *,
    attendance: SheetDocument,
    course_name: str,
) -> dict[str, Any]:
    attendance_sheet_id = int(attendance.numeric_id or attendance.id)
    course_text = _normalize_sheet_text(course_name)
    if "梵呗" in course_text:
        from backend.core.attendance.fanbei_course_sheets import rebuild_fanbei_attendance_from_course_sheets

        return rebuild_fanbei_attendance_from_course_sheets(
            session,
            attendance_sheet_id=attendance_sheet_id,
        )

    from backend.core.attendance.nianzhu_course_sheets import rebuild_nianzhu_attendance_from_course_sheets

    return rebuild_nianzhu_attendance_from_course_sheets(
        session,
        attendance_sheet_id=attendance_sheet_id,
        active_only=True,
        course_name=course_name,
    )


def _format_registration_attendance_submitted_at(value: Any) -> str:
    parsed = _parse_registration_submitted_at_datetime(value)
    if parsed is None:
        return _normalize_sheet_text(value)
    if parsed.hour or parsed.minute or parsed.second:
        return parsed.strftime("%Y-%m-%d %H:%M")
    return parsed.strftime("%Y-%m-%d")


def _get_workbook_sheet_by_key_or_title(
    session: Session,
    workbook: WorkbookDocument,
    *,
    sheet_key: str,
    title: str,
) -> SheetDocument | None:
    links = session.exec(
        select(WorkbookSheetLink)
        .where(WorkbookSheetLink.workbook_id.in_(workbook_ref_aliases(workbook)))
        .order_by(WorkbookSheetLink.order_index, WorkbookSheetLink.created_at)
    ).all()
    if not links:
        return None
    sheet_map = load_sheets_by_refs(session, [link.sheet_id for link in links])
    for link in links:
        sheet = sheet_map.get(str(link.sheet_id))
        if sheet is None or sheet.scope != "notes":
            continue
        if _normalize_sheet_text(sheet.sheet_key) == sheet_key or _normalize_sheet_text(sheet.title) == title:
            return sheet
    return None


def _resolve_registration_attendance_sheet(
    session: Session,
    registration_document: SheetDocument,
    workbook: WorkbookDocument | None,
) -> tuple[SheetDocument | None, WorkbookDocument | None]:
    target_workbook = workbook
    if target_workbook is None:
        workbooks = _get_workbooks_for_sheet(session, registration_document)
        target_workbook = workbooks[0] if workbooks else None
    if target_workbook is None:
        return None, None
    return (
        _get_workbook_sheet_by_key_or_title(
            session,
            target_workbook,
            sheet_key="attendance",
            title="考勤表",
        ),
        target_workbook,
    )


def _is_archived_attendance_row(row: list[Any], columns: list[str]) -> bool:
    tracking_group_index = _get_column_index(columns, NOTE_SHEET_REGISTRATION_TRACKING_GROUP_COLUMN)
    tracking_status_index = _get_column_index(columns, NOTE_SHEET_REGISTRATION_TRACKING_STATUS_COLUMN)
    frozen_at_index = _get_column_index(columns, NOTE_SHEET_REGISTRATION_FROZEN_AT_COLUMN)
    if tracking_group_index >= 0 and _normalize_sheet_text(row[tracking_group_index]) == NOTE_SHEET_REGISTRATION_FROZEN_GROUP:
        return True
    if tracking_status_index >= 0 and _normalize_sheet_text(row[tracking_status_index]) == NOTE_SHEET_REGISTRATION_FROZEN_STATUS:
        return True
    if frozen_at_index >= 0 and _normalize_sheet_text(row[frozen_at_index]):
        return True
    return False


def _attendance_row_has_identity_payload(row: list[Any], columns: list[str]) -> bool:
    indexes = {column: index for index, column in enumerate(columns)}
    identity_columns = (
        "报名日期",
        "分组",
        "学号",
        "姓名",
        "昵称",
        "用户ID",
        NOTE_SHEET_REGISTRATION_LINKED_USER_ID_COLUMN,
        "商户订单号",
    )
    for column in identity_columns:
        index = indexes.get(column)
        if index is not None and index < len(row) and _normalize_sheet_text(row[index]):
            return True
    return False


def _remove_empty_attendance_identity_rows(document_json: dict[str, Any]) -> tuple[dict[str, Any], int]:
    normalized = _normalize_document_json(document_json)
    columns = _normalize_document_columns(normalized)
    rows = [
        _normalize_sheet_row(row, len(columns))
        for row in _extract_document_rows(normalized)
    ]
    if not rows:
        return normalized, 0
    next_rows: list[list[Any]] = []
    removed_count = 0
    for row in rows:
        if not _is_archived_attendance_row(row, columns) and not _attendance_row_has_identity_payload(row, columns):
            removed_count += 1
            continue
        next_rows.append(row)
    if removed_count <= 0:
        return normalized, 0
    return _replace_document_data_rows(normalized, next_rows), removed_count


def _parse_attendance_registration_datetime(value: Any, *, now: date | None = None) -> datetime | None:
    parsed = _parse_registration_submitted_at_datetime(value)
    if parsed is not None:
        return parsed
    text = _normalize_sheet_text(value)
    if not text:
        return None
    current = now or date.today()
    for fmt in ("%Y/%m/%d %H:%M:%S", "%Y/%m/%d %H:%M", "%Y/%m/%d"):
        try:
            parsed_without_year = datetime.strptime(f"{current.year}/{text}", fmt)
        except ValueError:
            continue
        return parsed_without_year
    return None


def _is_attendance_tracking_sheet(columns: list[str]) -> bool:
    if _get_column_index(columns, "报名日期") < 0:
        return False
    return any(
        _get_column_index(columns, header) >= 0
        for header in (
            NOTE_SHEET_REGISTRATION_TRACKING_GROUP_COLUMN,
            NOTE_SHEET_REGISTRATION_TRACKING_STATUS_COLUMN,
            NOTE_SHEET_REGISTRATION_TRACKING_DEADLINE_COLUMN,
            NOTE_SHEET_REGISTRATION_FROZEN_AT_COLUMN,
        )
    )


def _apply_attendance_tracking_values_to_row(
    row: list[Any],
    columns: list[str],
    *,
    now: date | None = None,
) -> bool:
    submitted_at_index = _get_column_index(columns, "报名日期")
    group_index = _get_column_index(columns, NOTE_SHEET_REGISTRATION_TRACKING_GROUP_COLUMN)
    status_index = _get_column_index(columns, NOTE_SHEET_REGISTRATION_TRACKING_STATUS_COLUMN)
    deadline_index = _get_column_index(columns, NOTE_SHEET_REGISTRATION_TRACKING_DEADLINE_COLUMN)
    frozen_at_index = _get_column_index(columns, NOTE_SHEET_REGISTRATION_FROZEN_AT_COLUMN)
    if submitted_at_index < 0:
        return False

    submitted_at = _parse_attendance_registration_datetime(row[submitted_at_index], now=now)
    current_date = now or date.today()
    current_datetime = datetime.combine(current_date, datetime.now().time()) if now is not None else datetime.now()
    changed = False

    def set_cell(column_index: int, value: str) -> None:
        nonlocal changed
        if column_index < 0 or column_index >= len(row):
            return
        if _normalize_sheet_text(row[column_index]) == value:
            return
        row[column_index] = value
        changed = True

    if submitted_at is None:
        return False

    submitted_date = submitted_at.date()
    archived = submitted_date < _registration_cutoff_date(current_date)
    if archived:
        set_cell(group_index, NOTE_SHEET_REGISTRATION_FROZEN_GROUP)
        set_cell(status_index, NOTE_SHEET_REGISTRATION_FROZEN_STATUS)
        if frozen_at_index >= 0 and not _normalize_sheet_text(row[frozen_at_index]):
            set_cell(frozen_at_index, _format_registration_tracking_datetime(current_datetime))
    else:
        if group_index >= 0 and _normalize_sheet_text(row[group_index]) in {"", NOTE_SHEET_REGISTRATION_FROZEN_GROUP}:
            set_cell(group_index, NOTE_SHEET_REGISTRATION_ACTIVE_GROUP)
        set_cell(status_index, NOTE_SHEET_REGISTRATION_ACTIVE_STATUS)
        set_cell(frozen_at_index, "")
    set_cell(deadline_index, _format_registration_tracking_date(_add_registration_months(submitted_date, 2)))
    return changed


def _attendance_dynamic_group_sort_key(
    row: list[Any],
    columns: list[str],
    source_index: int,
    *,
    now: date | None = None,
) -> tuple[Any, ...]:
    submitted_at_index = _get_column_index(columns, "报名日期")
    submitted_at = (
        _parse_attendance_registration_datetime(row[submitted_at_index], now=now)
        if submitted_at_index >= 0
        else None
    )
    if submitted_at is None:
        return 0, 1, source_index
    day = submitted_at.date().toordinal()
    second = submitted_at.hour * 3600 + submitted_at.minute * 60 + submitted_at.second
    if day >= _registration_cutoff_date(now).toordinal():
        return 0, 0, day, second, source_index
    return 1, 0, -day, -second, source_index


def _order_attendance_rows_by_dynamic_expiration(
    document_json: dict[str, Any],
    *,
    now: date | None = None,
) -> tuple[dict[str, Any], int]:
    normalized = _normalize_document_json(document_json)
    columns = _normalize_document_columns(normalized)
    if not _is_attendance_tracking_sheet(columns):
        return normalized, 0

    rows: list[list[Any]] = []
    tracking_changed_count = 0
    for row in _extract_document_rows(normalized):
        next_row = _normalize_sheet_row(row, len(columns))
        if _apply_attendance_tracking_values_to_row(next_row, columns, now=now):
            tracking_changed_count += 1
        rows.append(next_row)

    ordered_source_indexes = sorted(
        range(len(rows)),
        key=lambda index: _attendance_dynamic_group_sort_key(rows[index], columns, index, now=now),
    )
    order_changed = ordered_source_indexes != list(range(len(rows)))
    if not order_changed and tracking_changed_count <= 0:
        return normalized, 0

    row_index_map = {
        source_index: target_index
        for target_index, source_index in enumerate(ordered_source_indexes)
    }
    formula_row_offset = _get_formula_reference_row_offset(normalized)
    next_rows = [
        _remap_row_formula_cell_references(
            rows[source_index],
            columns=columns,
            row_index_map=row_index_map,
            row_index_offset=formula_row_offset,
        )
        for source_index in ordered_source_indexes
    ]
    next_document = _replace_document_data_rows({
        **normalized,
        "columns": columns,
    }, next_rows)
    if isinstance(normalized.get("cell_meta"), dict):
        next_document["cell_meta"] = _remap_cell_meta_rows(
            normalized.get("cell_meta"),
            row_index_map,
            row_offset=_normalize_document_data_start_row(normalized),
        )
    next_document = _remap_document_entity_data_rows(
        next_document,
        normalized,
        row_index_map=row_index_map,
        data_row_count=len(rows),
    )
    next_document, style_repaired_count = _apply_attendance_archived_row_styles(
        next_document,
        next_rows,
        columns,
    )
    return next_document, tracking_changed_count + style_repaired_count + (1 if order_changed else 0)


def _is_refunded_registration_row(row: list[Any], columns: list[str]) -> bool:
    remark_index = _get_column_index(columns, "备注")
    if remark_index < 0:
        return False
    remark = _normalize_sheet_text(row[remark_index])
    if not remark:
        return False
    return bool(re.search(r"(?:已\s*退\s*(?:费|款|课)|退款\s*成功|退费\s*完成|退款\s*完成)", remark))


def _get_attendance_append_insert_index(rows: list[list[Any]], columns: list[str]) -> int:
    for index, row in enumerate(rows):
        if _is_archived_attendance_row(row, columns):
            return index
    return len(rows)


def _find_attendance_row_template_index(
    rows: list[list[Any]],
    columns: list[str],
    target_index: int,
) -> int | None:
    before_limit = max(0, min(target_index, len(rows)))
    for index in range(before_limit - 1, -1, -1):
        if not _is_archived_attendance_row(rows[index], columns):
            return index
    for index, row in enumerate(rows):
        if index == target_index:
            continue
        if not _is_archived_attendance_row(row, columns):
            return index
    return None


def _build_attendance_row_from_template(
    attendance_columns: list[str],
    *,
    template_row: list[Any] | None = None,
    template_row_index: int | None = None,
    target_row_index: int | None = None,
) -> list[Any]:
    next_row = [""] * len(attendance_columns)
    if template_row is None:
        for index, header in enumerate(attendance_columns):
            if header in NOTE_SHEET_ATTENDANCE_INITIAL_ZERO_COLUMNS:
                next_row[index] = "0"
        return next_row

    normalized_template = _normalize_sheet_row(template_row, len(attendance_columns))
    row_delta = (
        target_row_index - template_row_index
        if template_row_index is not None and target_row_index is not None
        else 0
    )
    for index, header in enumerate(attendance_columns):
        template_value = normalized_template[index]
        if _is_formula_expression(template_value):
            next_row[index] = _shift_formula_value_references(template_value, row_delta=row_delta)
        elif header in NOTE_SHEET_ATTENDANCE_INITIAL_ZERO_COLUMNS:
            next_row[index] = "0"
    return next_row


def _build_attendance_row_from_registration(
    registration_row: list[Any],
    registration_columns: list[str],
    attendance_columns: list[str],
    *,
    template_row: list[Any] | None = None,
    template_row_index: int | None = None,
    target_row_index: int | None = None,
) -> list[Any]:
    source = dict(zip(registration_columns, _normalize_sheet_row(registration_row, len(registration_columns))))

    def source_value(*headers: str) -> str:
        for header in headers:
            text = _normalize_sheet_text(source.get(header))
            if text:
                return text
        return ""

    def source_order_amount() -> str:
        merchant_order = _strip_legacy_text_prefix(source_value("商户订单号"))
        amount = _parse_number_sort_value(source.get("订单金额"))
        if merchant_order and amount is not None and amount > 0:
            return _format_registration_match_cell(amount)
        return "0"

    submitted_at = _format_registration_attendance_submitted_at(source.get(NOTE_SHEET_REGISTRATION_SUBMITTED_AT_COLUMN))
    next_row = _build_attendance_row_from_template(
        attendance_columns,
        template_row=template_row,
        template_row_index=template_row_index,
        target_row_index=target_row_index,
    )
    for index, header in enumerate(attendance_columns):
        if header == "报名日期":
            next_row[index] = submitted_at
        elif header == "学号":
            next_row[index] = source_value(NOTE_SHEET_REGISTRATION_SEQUENCE_COLUMN)
        elif header in {"分组", "组号"}:
            next_row[index] = source_value(NOTE_SHEET_REGISTRATION_GROUP_COLUMN)
        elif header == "姓名":
            next_row[index] = source_value("姓名")
        elif header in {"昵称", "微信昵称"}:
            next_row[index] = source_value("微信昵称", "昵称")
        elif header == "手机号":
            next_row[index] = source_value("手机号")
        elif header == "微信支付订单号":
            next_row[index] = source_value("微信支付订单号")
        elif header == "商户订单号":
            next_row[index] = source_value("商户订单号")
        elif header == "订单日期":
            next_row[index] = source_value("订单日期")
        elif header == "订单金额":
            next_row[index] = source_order_amount()
        elif header == "已返款" and not _is_formula_expression(next_row[index]):
            next_row[index] = "0"
        elif header == "用户ID":
            next_row[index] = source_value("用户ID")
        elif header == NOTE_SHEET_REGISTRATION_LINKED_USER_ID_COLUMN:
            next_row[index] = source_value(NOTE_SHEET_REGISTRATION_LINKED_USER_ID_COLUMN)
        elif header == "匹配得分":
            next_row[index] = source_value("匹配得分")
        elif header == "规则版本":
            next_row[index] = "当前规则"
        elif header == NOTE_SHEET_REGISTRATION_TRACKING_STATUS_COLUMN:
            next_row[index] = "追踪中"
        elif header == NOTE_SHEET_REGISTRATION_TRACKING_GROUP_COLUMN:
            next_row[index] = source_value(NOTE_SHEET_REGISTRATION_GROUP_COLUMN)
    return next_row


def _extract_attendance_clockin_refund_period_reference(formula: Any) -> str:
    text = _normalize_sheet_text(formula)
    if NOTE_SHEET_ATTENDANCE_REFUND_PERIOD_NAME in text:
        return NOTE_SHEET_ATTENDANCE_REFUND_PERIOD_NAME
    match = re.search(r",\s*(\$?[A-Za-z]{1,3}\$?\d+)\s*\)\s*\*", text, re.I)
    if match:
        candidate = match.group(1).strip()
        if candidate:
            return candidate
    return "$J$3"


def _build_attendance_dual_clockin_refund_formula(
    columns: list[str],
    *,
    row_number: int,
    period_reference: str,
) -> str | None:
    study_index = _get_column_index(columns, "共学打卡")
    practice_index = _get_column_index(columns, "共修打卡")
    if study_index < 0 or practice_index < 0:
        return None
    study_ref = f"{_excel_column_label(study_index)}{row_number}"
    practice_ref = f"{_excel_column_label(practice_index)}{row_number}"
    period_ref = period_reference.strip() or "$J$3"
    return (
        f"=MIN(IFERROR(VALUE({study_ref}),0),{period_ref})*5"
        f"+MIN(IFERROR(VALUE({practice_ref}),0),{period_ref})*5"
    )


def _build_attendance_total_refund_formula(
    columns: list[str],
    *,
    row_number: int,
    standard_amount_row_number: int,
) -> str | None:
    video_index = _get_column_index(columns, "视频应返款")
    clockin_index = _get_column_index(columns, "打卡应返款")
    order_amount_index = _get_column_index(columns, "订单金额")
    if video_index < 0 or clockin_index < 0 or order_amount_index < 0:
        return None
    video_ref = f"{_excel_column_label(video_index)}{row_number}"
    clockin_ref = f"{_excel_column_label(clockin_index)}{row_number}"
    order_ref = f"{_excel_column_label(order_amount_index)}{row_number}"
    standard_amount_ref = f"${_excel_column_label(order_amount_index)}${standard_amount_row_number}"
    return f"=MIN(IFERROR({video_ref}+{clockin_ref}+{order_ref}-IF({standard_amount_ref}>0,{standard_amount_ref},{order_ref}),0),{order_ref})"


def _build_attendance_zen_guest_formula(columns: list[str], *, row_number: int) -> str | None:
    completed_video_index = _get_column_index(columns, "完成视频数")
    clockin_index = _get_column_index(columns, "打卡数")
    if completed_video_index < 0 or clockin_index < 0:
        return None
    completed_video_ref = f"{_excel_column_label(completed_video_index)}{row_number}"
    clockin_ref = f"{_excel_column_label(clockin_index)}{row_number}"
    return f'=IF(AND({completed_video_ref}>=11,{clockin_ref}>=7),"是","")'


def _formula_row_ranges(column_indexes: list[int], row_number: int) -> list[str]:
    sorted_indexes = sorted({index for index in column_indexes if index >= 0})
    if not sorted_indexes:
        return []

    ranges: list[str] = []
    start = previous = sorted_indexes[0]
    for index in sorted_indexes[1:]:
        if index == previous + 1:
            previous = index
            continue
        start_ref = f"{_excel_column_label(start)}{row_number}"
        previous_ref = f"{_excel_column_label(previous)}{row_number}"
        ranges.append(start_ref if start == previous else f"{start_ref}:{previous_ref}")
        start = previous = index
    start_ref = f"{_excel_column_label(start)}{row_number}"
    previous_ref = f"{_excel_column_label(previous)}{row_number}"
    ranges.append(start_ref if start == previous else f"{start_ref}:{previous_ref}")
    return ranges


def _attendance_video_progress_column_indexes(columns: list[str]) -> list[int]:
    return [
        index
        for index, header in enumerate(columns)
        if _is_attendance_video_progress_column(header)
    ]


def _build_attendance_completed_video_formula(
    columns: list[str],
    *,
    row_number: int,
) -> str | None:
    ranges = _formula_row_ranges(_attendance_video_progress_column_indexes(columns), row_number)
    if not ranges:
        return None
    parts = [f'COUNTIF({range_ref},"*完成*")' for range_ref in ranges]
    parts.extend(f'COUNTIF({range_ref},"*回放*")' for range_ref in ranges)
    return "=" + "+".join(parts)


def _build_attendance_video_refund_formula(
    document_json: dict[str, Any],
    columns: list[str],
    *,
    row_number: int,
) -> str | None:
    ranges = _formula_row_ranges(_attendance_video_progress_column_indexes(columns), row_number)
    if not ranges:
        return None
    rules = {
        key: amount
        for key, amount in _attendance_video_refund_rules(document_json).items()
        if key and amount > 0
    }
    if not rules:
        return None
    parts: list[str] = []
    for key, amount in rules.items():
        amount_text = _format_registration_match_cell(amount)
        parts.extend(f'COUNTIF({range_ref},"*{key}*")*{amount_text}' for range_ref in ranges)
    return "=" + "+".join(parts)


def _build_attendance_clockin_refund_formula(
    document_json: dict[str, Any],
    columns: list[str],
    *,
    row_number: int,
) -> str | None:
    clockin_index = _get_column_index(columns, "打卡数")
    refund_index = _get_column_index(columns, "打卡应返款")
    if clockin_index < 0 or refund_index < 0:
        return None
    rules = _attendance_clockin_rules_by_column(document_json, columns).get(clockin_index)
    if not rules:
        grid_rows = _extract_document_grid_rows(document_json)
        data_start_row = _normalize_document_data_start_row(document_json)
        for raw_row in reversed(grid_rows[:data_start_row]):
            row = _normalize_sheet_row(raw_row, len(columns))
            rules = parse_threshold_refund_rules(row[refund_index])
            if rules:
                break
    if not rules:
        return None
    clockin_ref = f"{_excel_column_label(clockin_index)}{row_number}"
    parts: list[str] = []
    for rule in sorted(rules, key=lambda item: item.threshold, reverse=True):
        parts.extend([
            f"{clockin_ref}>={_format_registration_match_cell(rule.threshold)}",
            _format_registration_match_cell(rule.refund_amount),
        ])
    parts.append("0")
    return "=SWITCH(TRUE," + ",".join(parts) + ")"


def _build_attendance_current_refund_formula(columns: list[str], *, row_number: int) -> str | None:
    total_index = _get_column_index(columns, "总应返款")
    refunded_index = _get_column_index(columns, "已返款")
    order_amount_index = _get_column_index(columns, "订单金额")
    if total_index < 0 or refunded_index < 0 or order_amount_index < 0:
        return None
    total_ref = f"{_excel_column_label(total_index)}{row_number}"
    refunded_ref = f"{_excel_column_label(refunded_index)}{row_number}"
    order_ref = f"{_excel_column_label(order_amount_index)}{row_number}"
    return f"=({order_ref}>0)*({total_ref}-{refunded_ref})"


def _build_attendance_refund_config_formula(columns: list[str], *, row_number: int) -> str | None:
    merchant_order_index = _get_column_index(columns, "商户订单号")
    current_refund_index = _get_column_index(columns, "当前应返款")
    if merchant_order_index < 0 or current_refund_index < 0:
        return None
    merchant_order_ref = f"{_excel_column_label(merchant_order_index)}{row_number}"
    current_refund_ref = f"{_excel_column_label(current_refund_index)}{row_number}"
    period_ref = f"${_excel_column_label(current_refund_index)}$1"
    return (
        f'=IF({current_refund_ref}>0,TEXTJOIN(",",TRUE,{merchant_order_ref},{current_refund_ref},'
        f'"念住闯关每日返款",{merchant_order_ref}&"_day"&{period_ref}),"")'
    )


def _normalize_attendance_managed_refund_formulas(
    document_json: dict[str, Any],
) -> tuple[dict[str, Any], int]:
    normalized = _normalize_document_json(document_json)
    columns = _normalize_document_columns(normalized)
    rows = [
        _normalize_sheet_row(row, len(columns))
        for row in _extract_document_rows(normalized)
    ]
    if not rows:
        return document_json, 0

    row_reference_offset = _get_formula_reference_row_offset(normalized)
    standard_amount_row_number = row_reference_offset if row_reference_offset > 0 else 3
    formula_builders = {
        "禅客": lambda row_number: _build_attendance_zen_guest_formula(columns, row_number=row_number),
        "完成视频数": lambda row_number: _build_attendance_completed_video_formula(columns, row_number=row_number),
        "视频应返款": lambda row_number: _build_attendance_video_refund_formula(normalized, columns, row_number=row_number),
        "打卡应返款": lambda row_number: _build_attendance_clockin_refund_formula(normalized, columns, row_number=row_number),
        "总应返款": lambda row_number: _build_attendance_total_refund_formula(
            columns,
            row_number=row_number,
            standard_amount_row_number=standard_amount_row_number,
        ),
        "当前应返款": lambda row_number: _build_attendance_current_refund_formula(columns, row_number=row_number),
        "返款配置": lambda row_number: _build_attendance_refund_config_formula(columns, row_number=row_number),
    }
    changed_cells: list[tuple[int, int]] = []
    for row_index, row in enumerate(rows):
        if _is_archived_attendance_row(row, columns) or not _attendance_row_has_identity_payload(row, columns):
            continue
        row_number = row_reference_offset + row_index + 1
        for header, builder in formula_builders.items():
            column_index = _get_column_index(columns, header)
            if column_index < 0:
                continue
            formula = builder(row_number)
            if not formula:
                continue
            if row[column_index] != formula:
                row[column_index] = formula
                changed_cells.append((row_index, column_index))

    if not changed_cells:
        return document_json, 0

    next_document = _replace_document_data_rows(normalized, rows)
    data_start_row = _normalize_document_data_start_row(normalized)
    for row_index, column_index in changed_cells:
        next_document = _set_document_entity_cell_value(
            next_document,
            document_row=data_start_row + row_index,
            column_index=column_index,
            value=rows[row_index][column_index],
        )
    return next_document, len(changed_cells)


def _is_legacy_nianzhu_zen_guest_formula(value: Any) -> bool:
    text = _normalize_sheet_text(value)
    if not _is_formula_expression(text):
        return False
    compact = re.sub(r"\s+", "", text).upper()
    return ">=11" in compact and ">=7" in compact and (
        compact.startswith("=AND(")
        or compact.startswith("=IF(AND(")
    )


def _find_attendance_standard_order_amount(rows: list[list[Any]], columns: list[str]) -> str:
    order_amount_index = _get_column_index(columns, "订单金额")
    if order_amount_index < 0:
        return ""
    for row in rows:
        amount = _parse_number_sort_value(row[order_amount_index] if order_amount_index < len(row) else "")
        if amount is not None and amount > 0:
            return _format_registration_match_cell(amount)
    return ""


def _set_document_entity_cell_value(
    document_json: dict[str, Any],
    *,
    document_row: int,
    column_index: int,
    value: Any,
) -> dict[str, Any]:
    entity_rows = _extract_document_entity_rows(document_json)
    entity_columns = _extract_document_entity_columns(document_json)
    if (
        document_row < 0
        or column_index < 0
        or document_row >= len(entity_rows)
        or column_index >= len(entity_columns)
    ):
        return document_json
    row_id = _get_document_entity_row_id(entity_rows[document_row])
    column_id = _get_document_entity_column_id(entity_columns[column_index])
    if not row_id or not column_id:
        return document_json

    entity_cells = _extract_document_entity_cells(document_json)
    row_cells = entity_cells.get(row_id)
    next_row_cells = dict(row_cells) if isinstance(row_cells, dict) else {}
    source_cell = next_row_cells.get(column_id)
    next_cell = dict(source_cell) if isinstance(source_cell, dict) else {}
    next_cell["value"] = value
    if _is_formula_expression(value):
        next_cell.pop("rich_text", None)
        if next_cell.get("cell_type") == "rich_text":
            next_cell.pop("cell_type", None)
    if next_cell == source_cell:
        return document_json

    next_row_cells[column_id] = next_cell
    next_entity_cells = dict(entity_cells)
    next_entity_cells[row_id] = next_row_cells
    next_document = dict(document_json)
    next_document["entity_cells"] = next_entity_cells
    return next_document


def _normalize_attendance_current_refund_note(value: Any) -> str:
    text = _normalize_sheet_text(value)
    if not text:
        return text
    should_normalize = "返款" in text or "最近运行更新时间" in text
    if not should_normalize:
        return text
    timestamp_match = re.search(
        r"\d{4}/\d{1,2}/\d{1,2}\s+\d{1,2}:\d{2}(?::\d{2})?(?:,\d+)?",
        text,
    )
    if timestamp_match:
        return timestamp_match.group(0)
    if "待首次同步" in text:
        return "待首次同步"
    normalized = re.sub(
        r"^\s*[^\r\n]*返款\s*(?:[\r\n]+|(?=最近运行更新时间\s*[:：]?))",
        "",
        text,
        count=1,
    ).strip()
    normalized = re.sub(
        r"^\s*最近运行更新时间\s*[:：]?\s*",
        "",
        normalized,
        count=1,
    ).strip()
    if normalized != text:
        return normalized
    if re.fullmatch(r"\s*[^\r\n]*返款\s*", text):
        return ""
    return text


def _normalize_attendance_refund_config_row(
    document_json: dict[str, Any],
    columns: list[str],
) -> tuple[dict[str, Any], int]:
    grid_rows = _extract_document_grid_rows(document_json)
    config_row_index = _normalize_document_data_start_row(document_json) - 1
    field_row_index = int(document_json.get("field_row_index") or 0)
    period_display_index = _get_column_index(columns, "已返款")
    order_amount_index = _get_column_index(columns, "订单金额")
    current_refund_index = _get_column_index(columns, "当前应返款")
    if not grid_rows or config_row_index < 0 or config_row_index >= len(grid_rows) or config_row_index <= field_row_index:
        return document_json, 0

    next_grid_rows = [list(row) if isinstance(row, list) else [] for row in grid_rows]
    config_row = _normalize_sheet_row(next_grid_rows[config_row_index], len(columns))
    config_changed_cells: list[int] = []
    if period_display_index >= 0:
        period_label_formula = f'="第"&{NOTE_SHEET_ATTENDANCE_REFUND_PERIOD_NAME}&"周"'
        current_period_label = _normalize_sheet_text(config_row[period_display_index])
        current_period_label_compact = re.sub(r"\s+", "", current_period_label)
        legacy_day_label_formula = f'="第"&{NOTE_SHEET_ATTENDANCE_REFUND_PERIOD_NAME}&"天"'
        if (
            not current_period_label
            or current_period_label_compact == legacy_day_label_formula
        ):
            config_row[period_display_index] = period_label_formula
            config_changed_cells.append(period_display_index)
    standard_amount = _find_attendance_standard_order_amount(
        [_normalize_sheet_row(row, len(columns)) for row in _extract_document_rows(document_json)],
        columns,
    )
    if (
        order_amount_index >= 0
        and standard_amount
        and (
            not _normalize_sheet_text(config_row[order_amount_index])
            or _is_formula_expression(config_row[order_amount_index])
        )
    ):
        config_row[order_amount_index] = standard_amount
        config_changed_cells.append(order_amount_index)
    if current_refund_index >= 0:
        current_refund_note = _normalize_attendance_current_refund_note(config_row[current_refund_index])
        if current_refund_note != _normalize_sheet_text(config_row[current_refund_index]):
            config_row[current_refund_index] = current_refund_note
            config_changed_cells.append(current_refund_index)
    if not config_changed_cells:
        return document_json, 0

    next_grid_rows[config_row_index] = config_row
    next_document = dict(document_json)
    next_document["grid_rows"] = next_grid_rows
    for column_index in config_changed_cells:
        next_document = _set_document_entity_cell_value(
            next_document,
            document_row=config_row_index,
            column_index=column_index,
            value=config_row[column_index],
        )
    return next_document, len(config_changed_cells)


def _normalize_attendance_current_refund_config_note(
    document_json: dict[str, Any],
    columns: list[str],
) -> tuple[dict[str, Any], int]:
    grid_rows = _extract_document_grid_rows(document_json)
    config_row_index = _normalize_document_data_start_row(document_json) - 1
    field_row_index = int(document_json.get("field_row_index") or 0)
    current_refund_index = _get_column_index(columns, "当前应返款")
    if (
        current_refund_index < 0
        or not grid_rows
        or config_row_index < 0
        or config_row_index >= len(grid_rows)
        or config_row_index <= field_row_index
    ):
        return document_json, 0

    next_grid_rows = [list(row) if isinstance(row, list) else [] for row in grid_rows]
    config_row = _normalize_sheet_row(next_grid_rows[config_row_index], len(columns))
    current_refund_note = _normalize_attendance_current_refund_note(config_row[current_refund_index])
    if current_refund_note == _normalize_sheet_text(config_row[current_refund_index]):
        return document_json, 0

    config_row[current_refund_index] = current_refund_note
    next_grid_rows[config_row_index] = config_row
    next_document = dict(document_json)
    next_document["grid_rows"] = next_grid_rows
    next_document = _set_document_entity_cell_value(
        next_document,
        document_row=config_row_index,
        column_index=current_refund_index,
        value=current_refund_note,
    )
    return next_document, 1


def _normalize_attendance_dual_clockin_refund_formulas(
    document_json: dict[str, Any],
) -> tuple[dict[str, Any], int]:
    normalized = _normalize_document_json(document_json)
    columns = _normalize_document_columns(normalized)
    normalized, note_changed_count = _normalize_attendance_current_refund_config_note(normalized, columns)
    refund_index = _get_column_index(columns, "打卡应返款")
    total_index = _get_column_index(columns, "总应返款")
    has_dual_clockin_refund = (
        refund_index >= 0
        and _get_column_index(columns, "共学打卡") >= 0
        and _get_column_index(columns, "共修打卡") >= 0
    )
    config_row_changed_count = 0
    if has_dual_clockin_refund:
        normalized, config_row_changed_count = _normalize_attendance_refund_config_row(normalized, columns)

    rows = [
        _normalize_sheet_row(row, len(columns))
        for row in _extract_document_rows(normalized)
    ]
    changed_formula_cells: list[tuple[int, int]] = []
    row_reference_offset = _get_formula_reference_row_offset(normalized)
    standard_amount_row_number = row_reference_offset if row_reference_offset > 0 else 3
    zen_guest_index = _get_column_index(columns, "禅客")
    for row_index, row in enumerate(rows):
        row_number = row_reference_offset + row_index + 1
        if zen_guest_index >= 0 and _is_legacy_nianzhu_zen_guest_formula(row[zen_guest_index]):
            next_zen_guest_formula = _build_attendance_zen_guest_formula(columns, row_number=row_number)
            if next_zen_guest_formula and next_zen_guest_formula != row[zen_guest_index]:
                row[zen_guest_index] = next_zen_guest_formula
                changed_formula_cells.append((row_index, zen_guest_index))

        current_formula = row[refund_index] if has_dual_clockin_refund else ""
        if has_dual_clockin_refund and _is_formula_expression(current_formula):
            next_formula = _build_attendance_dual_clockin_refund_formula(
                columns,
                row_number=row_number,
                period_reference=_extract_attendance_clockin_refund_period_reference(current_formula),
            )
            if next_formula and next_formula != current_formula:
                row[refund_index] = next_formula
                changed_formula_cells.append((row_index, refund_index))

        if has_dual_clockin_refund and total_index >= 0 and _is_formula_expression(row[total_index]):
            next_total_formula = _build_attendance_total_refund_formula(
                columns,
                row_number=row_number,
                standard_amount_row_number=standard_amount_row_number,
            )
            if next_total_formula and next_total_formula != row[total_index]:
                row[total_index] = next_total_formula
                changed_formula_cells.append((row_index, total_index))

    changed_count = note_changed_count + config_row_changed_count
    next_document = normalized
    if changed_formula_cells:
        next_document = _replace_document_data_rows(next_document, rows)
        data_start_row = _normalize_document_data_start_row(normalized)
        for row_index, column_index in changed_formula_cells:
            next_document = _set_document_entity_cell_value(
                next_document,
                document_row=data_start_row + row_index,
                column_index=column_index,
                value=rows[row_index][column_index],
            )
        changed_count += len(changed_formula_cells)

    if has_dual_clockin_refund:
        next_document, next_config_row_changed_count = _normalize_attendance_refund_config_row(next_document, columns)
        changed_count += next_config_row_changed_count

    if changed_count <= 0:
        return document_json, 0
    return next_document, changed_count


def _normalize_attendance_dual_clockin_refund_formulas_persisted(
    session: Session,
    document: SheetDocument,
    document_json: dict[str, Any],
) -> dict[str, Any]:
    next_document, changed_count = _normalize_attendance_dual_clockin_refund_formulas(document_json)
    if not changed_count:
        return next_document
    document.document_json = next_document
    document.updated_at = time.time()
    session.add(document)
    session.commit()
    session.refresh(document)
    return dict(document.document_json or {})


def _merge_attendance_registration_defaults(
    current_row: list[Any],
    candidate_row: list[Any],
    attendance_columns: list[str],
) -> tuple[list[Any], bool]:
    next_row = _normalize_sheet_row(current_row, len(attendance_columns))
    candidate = _normalize_sheet_row(candidate_row, len(attendance_columns))
    changed = False
    for index, header in enumerate(attendance_columns):
        candidate_value = candidate[index]
        if _normalize_sheet_text(candidate_value) == "":
            continue
        current_value = next_row[index]
        current_text = _normalize_sheet_text(current_value)
        should_fill = False
        if _is_formula_expression(candidate_value):
            should_fill = not current_text or not _is_formula_expression(current_value)
        elif header in NOTE_SHEET_ATTENDANCE_INITIAL_ZERO_COLUMNS:
            should_fill = not current_text
        elif header == "报名日期":
            candidate_datetime = _parse_registration_submitted_at_datetime(candidate_value)
            should_fill = candidate_datetime is not None and (
                not current_text
                or _parse_registration_submitted_at_datetime(current_value) is None
                or _format_registration_attendance_submitted_at(current_value) != current_text
            )
        elif header == "订单金额":
            candidate_amount = _parse_number_sort_value(candidate_value)
            current_amount = _parse_number_sort_value(current_value)
            if candidate_amount is not None:
                should_fill = (
                    not current_text
                    or current_amount is None
                    or abs(candidate_amount - current_amount) > 1e-9
                )
        elif header in NOTE_SHEET_ATTENDANCE_SOURCE_OVERLAY_COLUMNS:
            should_fill = not current_text
        if should_fill and current_value != candidate_value:
            next_row[index] = candidate_value
            changed = True
    return next_row, changed


def _attendance_progress_style_start_column(columns: list[str]) -> int:
    candidate_indexes = [
        index
        for header in ("打卡数", "共学打卡", "共修打卡")
        for index in [_get_column_index(columns, header)]
        if index >= 0
    ]
    if candidate_indexes:
        return min(candidate_indexes)

    current_refund_index = _get_column_index(columns, "当前应返款")
    if current_refund_index >= 0:
        return min(current_refund_index + 1, len(columns))
    return len(columns)


def _attendance_cell_meta_template_column_limit(columns: list[str]) -> int:
    return _attendance_progress_style_start_column(columns)


def _attendance_group_identity_end_column_index(columns: list[str]) -> int:
    for header in NOTE_SHEET_ATTENDANCE_GROUP_IDENTITY_END_COLUMNS:
        index = _get_column_index(columns, header)
        if index >= 0:
            return index
    return -1


def _attendance_group_background_key(row: list[Any], columns: list[str]) -> str:
    group_index = _get_column_index(columns, NOTE_SHEET_REGISTRATION_GROUP_COLUMN)
    group_number = _parse_registration_group_number(row[group_index] if group_index >= 0 else "")
    if group_number:
        return group_number

    student_id_index = _get_column_index(columns, "学号")
    parsed_sequence = _parse_registration_group_sequence(row[student_id_index] if student_id_index >= 0 else "")
    if parsed_sequence is not None:
        return parsed_sequence[0]

    return _normalize_sheet_text(row[group_index] if group_index >= 0 else "")


def _attendance_group_color_for_key(group_key: str, color_by_group: dict[str, str]) -> str:
    if group_key not in color_by_group:
        palette = NOTE_SHEET_ATTENDANCE_GROUP_IDENTITY_BACKGROUND_COLORS
        color_by_group[group_key] = palette[len(color_by_group) % len(palette)]
    return color_by_group[group_key]


def _set_cell_meta_background(
    cell_meta: dict[str, Any],
    *,
    document_row: int,
    column_index: int,
    color: str,
) -> bool:
    key = f"{document_row}:{column_index}"
    previous_meta = cell_meta.get(key)
    next_meta = dict(previous_meta) if isinstance(previous_meta, dict) else {}
    style = dict(next_meta.get("style")) if isinstance(next_meta.get("style"), dict) else {}
    previous_color = style.get("background_color")
    style["background_color"] = color
    next_meta["style"] = style
    cell_meta[key] = next_meta
    return previous_color != color


def _set_entity_cell_background(
    document_json: dict[str, Any],
    *,
    document_row: int,
    column_index: int,
    color: str,
) -> bool:
    entity_rows = _extract_document_entity_rows(document_json)
    entity_columns = _extract_document_entity_columns(document_json)
    if document_row < 0 or document_row >= len(entity_rows) or column_index < 0 or column_index >= len(entity_columns):
        return False

    row_id = _get_document_entity_row_id(entity_rows[document_row])
    column_id = _get_document_entity_column_id(entity_columns[column_index])
    if not row_id or not column_id:
        return False

    entity_cells = _extract_document_entity_cells(document_json)
    row_cells = dict(entity_cells.get(row_id)) if isinstance(entity_cells.get(row_id), dict) else {}
    previous_cell = row_cells.get(column_id)
    next_cell = dict(previous_cell) if isinstance(previous_cell, dict) else {}
    style = dict(next_cell.get("style")) if isinstance(next_cell.get("style"), dict) else {}
    previous_color = style.get("background_color")
    style["background_color"] = color
    next_cell["style"] = style
    row_cells[column_id] = next_cell
    entity_cells[row_id] = row_cells
    document_json["entity_cells"] = entity_cells
    return previous_color != color


def _set_entity_cell_style(
    document_json: dict[str, Any],
    *,
    document_row: int,
    column_index: int,
    style: dict[str, Any],
) -> bool:
    entity_rows = _extract_document_entity_rows(document_json)
    entity_columns = _extract_document_entity_columns(document_json)
    if document_row < 0 or document_row >= len(entity_rows) or column_index < 0 or column_index >= len(entity_columns):
        return False

    row_id = _get_document_entity_row_id(entity_rows[document_row])
    column_id = _get_document_entity_column_id(entity_columns[column_index])
    if not row_id or not column_id:
        return False

    entity_cells = _extract_document_entity_cells(document_json)
    row_cells = dict(entity_cells.get(row_id)) if isinstance(entity_cells.get(row_id), dict) else {}
    previous_cell = row_cells.get(column_id)
    next_cell = dict(previous_cell) if isinstance(previous_cell, dict) else {}
    previous_style = next_cell.get("style")
    next_cell["style"] = dict(style)
    row_cells[column_id] = next_cell
    entity_cells[row_id] = row_cells
    document_json["entity_cells"] = entity_cells
    return previous_style != next_cell["style"]


def _set_entity_cell_background_optional(
    document_json: dict[str, Any],
    *,
    document_row: int,
    column_index: int,
    color: str | None,
) -> bool:
    entity_rows = _extract_document_entity_rows(document_json)
    entity_columns = _extract_document_entity_columns(document_json)
    if document_row < 0 or document_row >= len(entity_rows) or column_index < 0 or column_index >= len(entity_columns):
        return False

    row_id = _get_document_entity_row_id(entity_rows[document_row])
    column_id = _get_document_entity_column_id(entity_columns[column_index])
    if not row_id or not column_id:
        return False

    entity_cells = _extract_document_entity_cells(document_json)
    row_cells = dict(entity_cells.get(row_id)) if isinstance(entity_cells.get(row_id), dict) else {}
    previous_cell = row_cells.get(column_id)
    next_cell = dict(previous_cell) if isinstance(previous_cell, dict) else {}
    style = dict(next_cell.get("style")) if isinstance(next_cell.get("style"), dict) else {}
    previous_color = style.get("background_color")

    if color:
        style["background_color"] = color
    else:
        style.pop("background_color", None)

    if style:
        next_cell["style"] = style
    else:
        next_cell.pop("style", None)

    if next_cell:
        row_cells[column_id] = next_cell
    else:
        row_cells.pop(column_id, None)

    if row_cells:
        entity_cells[row_id] = row_cells
    else:
        entity_cells.pop(row_id, None)
    document_json["entity_cells"] = entity_cells
    return previous_color != color


def _set_cell_meta_background_optional(
    cell_meta: dict[str, Any],
    *,
    document_row: int,
    column_index: int,
    color: str | None,
) -> bool:
    key = f"{document_row}:{column_index}"
    previous_meta = cell_meta.get(key)
    next_meta = dict(previous_meta) if isinstance(previous_meta, dict) else {}
    style = dict(next_meta.get("style")) if isinstance(next_meta.get("style"), dict) else {}
    previous_color = style.get("background_color")

    if color:
        style["background_color"] = color
    else:
        style.pop("background_color", None)

    if style:
        next_meta["style"] = style
    else:
        next_meta.pop("style", None)

    if next_meta:
        cell_meta[key] = next_meta
    else:
        cell_meta.pop(key, None)
    return previous_color != color


def _remove_entity_cell_archived_style(
    document_json: dict[str, Any],
    *,
    document_row: int,
    column_index: int,
) -> bool:
    entity_rows = _extract_document_entity_rows(document_json)
    entity_columns = _extract_document_entity_columns(document_json)
    if document_row < 0 or document_row >= len(entity_rows) or column_index < 0 or column_index >= len(entity_columns):
        return False

    row_id = _get_document_entity_row_id(entity_rows[document_row])
    column_id = _get_document_entity_column_id(entity_columns[column_index])
    if not row_id or not column_id:
        return False

    entity_cells = _extract_document_entity_cells(document_json)
    row_cells = dict(entity_cells.get(row_id)) if isinstance(entity_cells.get(row_id), dict) else {}
    current_cell = row_cells.get(column_id)
    if not isinstance(current_cell, dict) or not _is_registration_archived_style(current_cell.get("style")):
        return False

    next_cell = dict(current_cell)
    next_cell.pop("style", None)
    if next_cell:
        row_cells[column_id] = next_cell
    else:
        row_cells.pop(column_id, None)
    if row_cells:
        entity_cells[row_id] = row_cells
    else:
        entity_cells.pop(row_id, None)
    document_json["entity_cells"] = entity_cells
    return True


def _apply_attendance_archived_row_styles(
    document_json: dict[str, Any],
    rows: list[list[Any]],
    columns: list[str],
) -> tuple[dict[str, Any], int]:
    if not rows or not columns:
        return document_json, 0

    next_document = dict(document_json)
    cell_meta = dict(next_document.get("cell_meta")) if isinstance(next_document.get("cell_meta"), dict) else {}
    row_offset = _normalize_document_data_start_row(next_document) if _extract_document_grid_rows(next_document) else 0
    archived_style = {
        "background_color": NOTE_SHEET_REGISTRATION_ARCHIVED_BACKGROUND,
        "text_color": NOTE_SHEET_REGISTRATION_ARCHIVED_TEXT,
    }
    changed_rows: set[int] = set()

    for row_index, row in enumerate(rows):
        archived = _is_archived_attendance_row(_normalize_sheet_row(row, len(columns)), columns)
        document_row = row_offset + row_index
        for column_index in range(len(columns)):
            key = f"{document_row}:{column_index}"
            current_meta = cell_meta.get(key)
            next_meta = dict(current_meta) if isinstance(current_meta, dict) else {}
            current_style = next_meta.get("style")
            next_style = dict(current_style) if isinstance(current_style, dict) else {}
            if archived:
                merged_style = {**next_style, **archived_style}
                if next_style != merged_style:
                    next_meta["style"] = merged_style
                    cell_meta[key] = next_meta
                    changed_rows.add(row_index)
                if _set_entity_cell_style(
                    next_document,
                    document_row=document_row,
                    column_index=column_index,
                    style=merged_style,
                ):
                    changed_rows.add(row_index)
                continue

            if _is_registration_archived_style(next_style):
                next_style.pop("background_color", None)
                next_style.pop("text_color", None)
                if next_style:
                    next_meta["style"] = next_style
                else:
                    next_meta.pop("style", None)
                if next_meta:
                    cell_meta[key] = next_meta
                else:
                    cell_meta.pop(key, None)
                changed_rows.add(row_index)
            if _remove_entity_cell_archived_style(
                next_document,
                document_row=document_row,
                column_index=column_index,
            ):
                changed_rows.add(row_index)

    next_document["cell_meta"] = cell_meta
    return next_document, len(changed_rows)


def _apply_attendance_group_identity_backgrounds(
    document_json: dict[str, Any],
    rows: list[list[Any]],
    columns: list[str],
) -> tuple[dict[str, Any], int]:
    end_column_index = _attendance_group_identity_end_column_index(columns)
    if end_column_index < 0:
        return document_json, 0

    next_document = dict(document_json)
    cell_meta = dict(next_document.get("cell_meta")) if isinstance(next_document.get("cell_meta"), dict) else {}
    row_offset = _normalize_document_data_start_row(next_document) if _extract_document_grid_rows(next_document) else 0
    changed_rows: set[int] = set()
    color_by_group: dict[str, str] = {}
    for row_index, row in enumerate(rows):
        normalized_row = _normalize_sheet_row(row, len(columns))
        if _is_archived_attendance_row(normalized_row, columns):
            continue
        group_key = _attendance_group_background_key(normalized_row, columns)
        if not group_key:
            continue
        color = _attendance_group_color_for_key(group_key, color_by_group)
        document_row = row_offset + row_index
        for column_index in range(0, min(end_column_index + 1, len(columns))):
            legacy_changed = _set_cell_meta_background(
                cell_meta,
                document_row=document_row,
                column_index=column_index,
                color=color,
            )
            entity_changed = _set_entity_cell_background(
                next_document,
                document_row=document_row,
                column_index=column_index,
                color=color,
            )
            if legacy_changed or entity_changed:
                changed_rows.add(row_index)

    next_document["cell_meta"] = cell_meta
    return next_document, len(changed_rows)


def _attendance_progress_style_column_range(columns: list[str]) -> tuple[int, int]:
    start_index = _attendance_progress_style_start_column(columns)
    if start_index >= len(columns):
        return len(columns), len(columns)
    end_index = len(columns)
    for column_index in range(start_index + 1, len(columns)):
        if columns[column_index] in NOTE_SHEET_ATTENDANCE_TRAILING_META_COLUMNS:
            end_index = column_index
            break
    return start_index, end_index


def _is_attendance_video_progress_column(header: str) -> bool:
    return re.search(r"第\s*0*\d+\s*课", _normalize_sheet_text(header)) is not None


def _attendance_video_refund_rules(document_json: dict[str, Any]) -> dict[str, int]:
    def normalize_rules(raw_rules: Any) -> dict[str, int]:
        if not isinstance(raw_rules, dict):
            return {}
        rules: dict[str, int] = {}
        for key, value in raw_rules.items():
            text_key = _normalize_sheet_text(key)
            if not text_key:
                continue
            try:
                amount = int(float(value))
            except (TypeError, ValueError):
                continue
            rules[text_key] = amount
        return rules

    source_meta = document_json.get("source_meta") if isinstance(document_json.get("source_meta"), dict) else {}
    raw_rules = source_meta.get("timed_video_rules") if isinstance(source_meta, dict) else None
    rules = normalize_rules(raw_rules)
    if rules:
        return rules

    columns = _normalize_document_columns(document_json)
    video_refund_index = _get_column_index(columns, "视频应返款")
    if video_refund_index < 0:
        return {}

    grid_rows = _extract_document_grid_rows(document_json)
    data_start_row = _normalize_document_data_start_row(document_json)
    candidates: list[Any] = []
    if grid_rows:
        for row in grid_rows[:data_start_row]:
            if video_refund_index < len(row):
                candidates.append(row[video_refund_index])

    for row in _extract_document_rows(document_json):
        if video_refund_index < len(row):
            candidates.append(row[video_refund_index])

    for candidate in candidates:
        parsed_rules = parse_compact_refund_rules(candidate)
        if parsed_rules:
            return parsed_rules

        text = _normalize_sheet_text(candidate)
        formula_rules: dict[str, int] = {}
        for key, amount in re.findall(r'"\*([^"*]+)\*"\)\s*\*\s*(\d+(?:\.\d+)?)', text):
            text_key = _normalize_sheet_text(key)
            if not text_key:
                continue
            formula_rules[text_key] = int(float(amount))
        if formula_rules:
            formula_rules.setdefault("回放", 0)
            return formula_rules

    return {}


def _attendance_video_progress_background(value: Any, refund_rules: dict[str, int] | None = None) -> tuple[bool, str | None]:
    text = _normalize_sheet_text(value)
    if not text:
        return True, None
    if re.search(r"学习中|^进度\s*\d+(?:\.\d+)?\s*%|^观看\s*\d+", text):
        return True, None
    if refund_rules and re.search(r"当堂完成|第\s*\d+\s*天回放|回放|已完成", text):
        refund_amount, color = highlight_text_refund_progress(refund_rules, text)
        if refund_amount <= 0 and color is None:
            has_zero_refund_match = any(
                key and amount <= 0 and key in text
                for key, amount in refund_rules.items()
            )
            return True, NOTE_SHEET_ATTENDANCE_VIDEO_ZERO_REFUND_BACKGROUND if has_zero_refund_match else None
        return True, color
    if re.search(r"\d+\s*遍", text):
        return True, NOTE_SHEET_ATTENDANCE_VIDEO_PROGRESS_COMPLETED_BACKGROUND
    if re.search(r"准时完成|当周完成", text):
        return True, NOTE_SHEET_ATTENDANCE_VIDEO_PROGRESS_COMPLETED_BACKGROUND
    if re.search(r"延\s*\d+\s*周完成|延迟完成", text):
        return True, NOTE_SHEET_ATTENDANCE_VIDEO_ZERO_REFUND_BACKGROUND
    if re.search(r"当堂完成|第\s*\d+\s*天回放|回放|已完成", text):
        return True, highlight_presence_progress(text) or NOTE_SHEET_ATTENDANCE_VIDEO_PROGRESS_COMPLETED_BACKGROUND
    return False, None


def _is_attendance_clockin_count_column(header: str) -> bool:
    return _normalize_sheet_text(header) in {"打卡数", "共学打卡", "共修打卡"}


def _attendance_clockin_rules_by_column(document_json: dict[str, Any], columns: list[str]) -> dict[int, Any]:
    grid_rows = _extract_document_grid_rows(document_json)
    if not grid_rows:
        return {}

    data_start_row = _normalize_document_data_start_row(document_json)
    header_rows = [
        _normalize_sheet_row(row, len(columns))
        for row in grid_rows[:data_start_row]
    ]
    if not header_rows:
        return {}

    rules_by_column: dict[int, Any] = {}
    for column_index, header in enumerate(columns):
        if not _is_attendance_clockin_count_column(header):
            continue
        candidates = [
            row[column_index]
            for row in header_rows
            if column_index < len(row)
        ]
        for candidate in reversed(candidates):
            rules = parse_threshold_refund_rules(candidate)
            if rules:
                rules_by_column[column_index] = rules
                break
    return rules_by_column


def _attendance_clockin_count_background(value: Any, rules: Any) -> tuple[bool, str | None]:
    text = _normalize_sheet_text(value)
    if not text:
        return True, None
    if not rules:
        return False, None
    _refund_amount, color = highlight_threshold_refund_progress(rules, value)
    return True, color


def _apply_attendance_progress_backgrounds(
    document_json: dict[str, Any],
    *,
    assume_normalized: bool = False,
) -> tuple[dict[str, Any], int]:
    normalized = document_json if assume_normalized else _normalize_document_json(document_json)
    columns = _normalize_document_columns(normalized)
    video_column_indexes = [
        index
        for index, header in enumerate(columns)
        if _is_attendance_video_progress_column(header)
    ]
    video_refund_rules = _attendance_video_refund_rules(normalized)
    clockin_rules_by_column = _attendance_clockin_rules_by_column(normalized, columns)
    if not video_column_indexes and not clockin_rules_by_column:
        return document_json, 0

    rows = [_normalize_sheet_row(row, len(columns)) for row in _extract_document_rows(normalized)]
    if not rows:
        return document_json, 0

    next_document = dict(normalized)
    cell_meta = dict(next_document.get("cell_meta")) if isinstance(next_document.get("cell_meta"), dict) else {}
    row_offset = _normalize_document_data_start_row(next_document) if _extract_document_grid_rows(next_document) else 0
    changed_cells = 0

    for row_index, row in enumerate(rows):
        if _is_archived_attendance_row(row, columns):
            continue
        document_row = row_offset + row_index
        for column_index, rules in clockin_rules_by_column.items():
            recognized, color = _attendance_clockin_count_background(row[column_index], rules)
            if not recognized:
                continue
            legacy_changed = _set_cell_meta_background_optional(
                cell_meta,
                document_row=document_row,
                column_index=column_index,
                color=color,
            )
            entity_changed = _set_entity_cell_background_optional(
                next_document,
                document_row=document_row,
                column_index=column_index,
                color=color,
            )
            if legacy_changed or entity_changed:
                changed_cells += 1
        for column_index in video_column_indexes:
            recognized, color = _attendance_video_progress_background(row[column_index], video_refund_rules)
            if not recognized:
                continue
            legacy_changed = _set_cell_meta_background_optional(
                cell_meta,
                document_row=document_row,
                column_index=column_index,
                color=color,
            )
            entity_changed = _set_entity_cell_background_optional(
                next_document,
                document_row=document_row,
                column_index=column_index,
                color=color,
            )
            if legacy_changed or entity_changed:
                changed_cells += 1

    next_document["cell_meta"] = cell_meta
    return next_document, changed_cells


def _apply_course_attendance_header_links_for_response(
    session: Session,
    document: SheetDocument,
    document_json: dict[str, Any],
) -> tuple[dict[str, Any], int]:
    next_document = document_json
    total_count = 0
    try:
        from backend.core.attendance.nianzhu_course_sheets import apply_course_attendance_header_links_for_response as apply_nianzhu_links
    except Exception:
        apply_nianzhu_links = None
    if apply_nianzhu_links is not None:
        next_document, count = apply_nianzhu_links(
            session,
            attendance=document,
            document_json=next_document,
        )
        total_count += count

    try:
        from backend.core.attendance.fanbei_course_sheets import apply_course_attendance_header_links_for_response as apply_fanbei_links
    except Exception:
        apply_fanbei_links = None
    if apply_fanbei_links is not None:
        next_document, count = apply_fanbei_links(
            session,
            attendance=document,
            document_json=next_document,
        )
        total_count += count

    return next_document, total_count


def _cell_meta_has_style(meta: Any) -> bool:
    return isinstance(meta, dict) and isinstance(meta.get("style"), dict) and bool(meta.get("style"))


def _cell_meta_has_archived_attendance_style(meta: Any) -> bool:
    if not isinstance(meta, dict):
        return False
    style = meta.get("style")
    if not isinstance(style, dict):
        return False
    background_color = _normalize_sheet_text(style.get("background_color")).upper()
    text_color = _normalize_sheet_text(style.get("text_color")).upper()
    return background_color == "#F2F2F2" and text_color == "#6B7280"


def _attendance_row_has_archived_cell_meta(
    cell_meta: Any,
    *,
    document_row: int,
    column_limit: int,
) -> bool:
    if not isinstance(cell_meta, dict):
        return False
    for column_index in range(max(column_limit, 0)):
        if _cell_meta_has_archived_attendance_style(cell_meta.get(f"{document_row}:{column_index}")):
            return True
    return False


def _attendance_row_has_empty_progress_cell_meta(
    row: list[Any],
    cell_meta: Any,
    *,
    document_row: int,
    start_column: int,
    end_column: int,
) -> bool:
    if not isinstance(cell_meta, dict) or start_column >= end_column:
        return False
    if any(_normalize_sheet_text(row[column_index]) for column_index in range(start_column, min(end_column, len(row)))):
        return False
    for column_index in range(start_column, end_column):
        if _cell_meta_has_style(cell_meta.get(f"{document_row}:{column_index}")):
            return True
    return False


def _apply_attendance_template_cell_meta(
    cell_meta: Any,
    *,
    template_document_row: int | None,
    target_document_row: int,
    column_limit: int,
) -> dict[str, Any]:
    if not isinstance(cell_meta, dict):
        return {}
    next_meta: dict[str, Any] = {}
    for key, meta in cell_meta.items():
        parsed = _parse_cell_meta_key(key)
        if parsed is None:
            next_meta[str(key)] = meta
            continue
        row_index, _column_index = parsed
        if row_index != target_document_row:
            next_meta[str(key)] = meta

    if template_document_row is None:
        return next_meta
    for column_index in range(max(column_limit, 0)):
        meta = cell_meta.get(f"{template_document_row}:{column_index}")
        if meta is not None:
            next_meta[f"{target_document_row}:{column_index}"] = deepcopy(meta)
    return next_meta


def _entity_cell_has_archived_attendance_style(cells: Any, column_id: str) -> bool:
    if not isinstance(cells, dict) or not column_id:
        return False
    return _cell_meta_has_archived_attendance_style(cells.get(column_id))


def _attendance_row_has_archived_entity_cell_meta(
    document_json: dict[str, Any],
    *,
    document_row: int,
    column_limit: int,
) -> bool:
    entity_rows = _extract_document_entity_rows(document_json)
    entity_columns = _extract_document_entity_columns(document_json)
    if document_row < 0 or document_row >= len(entity_rows):
        return False
    row_id = _get_document_entity_row_id(entity_rows[document_row])
    row_cells = _extract_document_entity_cells(document_json).get(row_id)
    if not isinstance(row_cells, dict):
        return False
    for column_index in range(min(max(column_limit, 0), len(entity_columns))):
        column_id = _get_document_entity_column_id(entity_columns[column_index])
        if _entity_cell_has_archived_attendance_style(row_cells, column_id):
            return True
    return False


def _attendance_row_has_empty_progress_entity_cell_meta(
    row: list[Any],
    document_json: dict[str, Any],
    *,
    document_row: int,
    start_column: int,
    end_column: int,
) -> bool:
    if start_column >= end_column:
        return False
    if any(_normalize_sheet_text(row[column_index]) for column_index in range(start_column, min(end_column, len(row)))):
        return False

    entity_rows = _extract_document_entity_rows(document_json)
    entity_columns = _extract_document_entity_columns(document_json)
    if document_row < 0 or document_row >= len(entity_rows):
        return False
    row_id = _get_document_entity_row_id(entity_rows[document_row])
    row_cells = _extract_document_entity_cells(document_json).get(row_id)
    if not isinstance(row_cells, dict):
        return False
    for column_index in range(start_column, min(end_column, len(entity_columns))):
        column_id = _get_document_entity_column_id(entity_columns[column_index])
        if column_id and _cell_meta_has_style(row_cells.get(column_id)):
            return True
    return False


def _attendance_row_needs_registration_meta_repair(
    row: list[Any],
    document_json: dict[str, Any],
    *,
    document_row: int,
    column_limit: int,
    progress_style_start_column: int,
    progress_style_end_column: int,
) -> bool:
    cell_meta = document_json.get("cell_meta")
    return (
        _attendance_row_has_archived_cell_meta(
            cell_meta,
            document_row=document_row,
            column_limit=column_limit,
        )
        or _attendance_row_has_archived_entity_cell_meta(
            document_json,
            document_row=document_row,
            column_limit=column_limit,
        )
        or _attendance_row_has_empty_progress_cell_meta(
            row,
            cell_meta,
            document_row=document_row,
            start_column=progress_style_start_column,
            end_column=progress_style_end_column,
        )
        or _attendance_row_has_empty_progress_entity_cell_meta(
            row,
            document_json,
            document_row=document_row,
            start_column=progress_style_start_column,
            end_column=progress_style_end_column,
        )
    )


def _remove_attendance_row_entity_cell_styles(
    document_json: dict[str, Any],
    *,
    target_document_row: int,
) -> dict[str, Any]:
    entity_rows = _extract_document_entity_rows(document_json)
    if target_document_row < 0 or target_document_row >= len(entity_rows):
        return document_json
    target_row_id = _get_document_entity_row_id(entity_rows[target_document_row])
    if not target_row_id:
        return document_json
    entity_cells = _extract_document_entity_cells(document_json)
    target_cells = entity_cells.get(target_row_id)
    if not isinstance(target_cells, dict):
        return document_json

    next_target_cells: dict[str, Any] = {}
    changed = False
    for column_id, cell in target_cells.items():
        if not isinstance(cell, dict) or "style" not in cell:
            next_target_cells[column_id] = cell
            continue
        next_cell = dict(cell)
        next_cell.pop("style", None)
        changed = True
        if next_cell:
            next_target_cells[column_id] = next_cell
    if not changed:
        return document_json

    next_document = dict(document_json)
    next_entity_cells = dict(entity_cells)
    if next_target_cells:
        next_entity_cells[target_row_id] = next_target_cells
    else:
        next_entity_cells.pop(target_row_id, None)
    next_document["entity_cells"] = next_entity_cells
    return next_document


def _sync_registration_rows_to_attendance_document(
    registration_json: dict[str, Any],
    attendance_json: dict[str, Any],
) -> tuple[dict[str, Any], dict[str, int]]:
    registration_document = _normalize_document_json(registration_json)
    attendance_document = _normalize_document_json(attendance_json)
    registration_columns = _normalize_document_columns(registration_document)
    attendance_columns = _normalize_document_columns(attendance_document)
    if "用户ID" not in registration_columns:
        return attendance_document, _build_registration_match_summary(error_count=1)

    attendance_document, empty_attendance_rows_removed = _remove_empty_attendance_identity_rows(attendance_document)
    attendance_columns = _normalize_document_columns(attendance_document)
    registration_rows = [
        _normalize_sheet_row(row, len(registration_columns))
        for row in _extract_document_rows(registration_document)
    ]
    attendance_rows = [
        _normalize_sheet_row(row, len(attendance_columns))
        for row in _extract_document_rows(attendance_document)
    ]
    attendance_document, formula_repaired_count = _normalize_attendance_dual_clockin_refund_formulas(attendance_document)
    attendance_document, managed_formula_repaired_count = _normalize_attendance_managed_refund_formulas(attendance_document)
    formula_repaired_count += managed_formula_repaired_count
    if formula_repaired_count:
        attendance_rows = [
            _normalize_sheet_row(row, len(attendance_columns))
            for row in _extract_document_rows(attendance_document)
        ]
    user_id_index = _get_column_index(attendance_columns, "用户ID")
    student_id_index = _get_column_index(attendance_columns, "学号")
    merchant_order_index = _get_column_index(attendance_columns, "商户订单号")
    existing_user_id_rows = {
        _normalize_sheet_text(row[user_id_index]): index
        for index, row in enumerate(attendance_rows)
        if user_id_index >= 0 and _normalize_sheet_text(row[user_id_index])
    }
    existing_student_id_rows = {
        _normalize_sheet_text(row[student_id_index]): index
        for index, row in enumerate(attendance_rows)
        if student_id_index >= 0 and _normalize_sheet_text(row[student_id_index])
    }
    existing_merchant_order_id_rows = {
        _strip_legacy_text_prefix(row[merchant_order_index]): index
        for index, row in enumerate(attendance_rows)
        if merchant_order_index >= 0 and _strip_legacy_text_prefix(row[merchant_order_index])
    }
    existing_user_ids = set(existing_user_id_rows)
    existing_student_ids = set(existing_student_id_rows)
    existing_merchant_order_ids = set(existing_merchant_order_id_rows)

    skipped_count = 0
    repaired_count = empty_attendance_rows_removed
    repair_meta_targets: list[tuple[int, int | None]] = []
    pending_registration_rows: list[list[Any]] = []
    cell_meta_row_offset = _normalize_document_data_start_row(attendance_document) if _extract_document_grid_rows(attendance_document) else 0
    cell_meta_column_limit = _attendance_cell_meta_template_column_limit(attendance_columns)
    progress_style_start_column, progress_style_end_column = _attendance_progress_style_column_range(attendance_columns)
    for row in registration_rows:
        if not _registration_row_has_identity_payload(row, registration_columns):
            skipped_count += 1
            continue
        if _is_archived_registration_row(row, registration_columns):
            skipped_count += 1
            continue
        if _is_refunded_registration_row(row, registration_columns):
            skipped_count += 1
            continue
        user_id, student_id, merchant_order_id = _registration_attendance_row_identity(row, registration_columns)
        if not user_id and not student_id and not merchant_order_id:
            skipped_count += 1
            continue

        existing_index = _find_registration_attendance_existing_index(
            user_id=user_id,
            student_id=student_id,
            merchant_order_id=merchant_order_id,
            existing_user_id_rows=existing_user_id_rows,
            existing_student_id_rows=existing_student_id_rows,
            existing_merchant_order_id_rows=existing_merchant_order_id_rows,
        )
        if existing_index is not None:
            if 0 <= existing_index < len(attendance_rows) and not _is_archived_attendance_row(
                attendance_rows[existing_index],
                attendance_columns,
            ):
                template_index = _find_attendance_row_template_index(attendance_rows, attendance_columns, existing_index)
                candidate_row = _build_attendance_row_from_registration(
                    row,
                    registration_columns,
                    attendance_columns,
                    template_row=attendance_rows[template_index] if template_index is not None else None,
                    template_row_index=template_index,
                    target_row_index=existing_index,
                )
                repaired_row, changed = _merge_attendance_registration_defaults(
                    attendance_rows[existing_index],
                    candidate_row,
                    attendance_columns,
                )
                needs_meta_repair = _attendance_row_needs_registration_meta_repair(
                    attendance_rows[existing_index],
                    attendance_document,
                    document_row=cell_meta_row_offset + existing_index,
                    column_limit=cell_meta_column_limit,
                    progress_style_start_column=progress_style_start_column,
                    progress_style_end_column=progress_style_end_column,
                )
                if changed:
                    attendance_rows[existing_index] = repaired_row
                if changed or needs_meta_repair:
                    repaired_count += 1
                    repair_meta_targets.append((existing_index, template_index))
            skipped_count += 1
            continue
        if student_id and student_id in existing_student_ids:
            skipped_count += 1
            continue
        if merchant_order_id and merchant_order_id in existing_merchant_order_ids:
            skipped_count += 1
            continue

        pending_registration_rows.append(row)
        pending_row_index = len(attendance_rows) + len(pending_registration_rows) - 1
        if user_id:
            existing_user_ids.add(user_id)
            existing_user_id_rows[user_id] = pending_row_index
        if student_id:
            existing_student_ids.add(student_id)
            existing_student_id_rows[student_id] = pending_row_index
        if merchant_order_id:
            existing_merchant_order_ids.add(merchant_order_id)
            existing_merchant_order_id_rows[merchant_order_id] = pending_row_index

    tracking_repaired_count = 0
    if pending_registration_rows:
        attendance_document = _replace_document_data_rows(attendance_document, attendance_rows)
        attendance_document, tracking_repaired_count = _order_attendance_rows_by_dynamic_expiration(attendance_document)
        attendance_rows = [
            _normalize_sheet_row(row, len(attendance_columns))
            for row in _extract_document_rows(attendance_document)
        ]

    insert_index = _get_attendance_append_insert_index(attendance_rows, attendance_columns)
    template_index = _find_attendance_row_template_index(attendance_rows, attendance_columns, insert_index)
    inserted_rows = [
        _build_attendance_row_from_registration(
            row,
            registration_columns,
            attendance_columns,
            template_row=attendance_rows[template_index] if template_index is not None else None,
            template_row_index=template_index,
            target_row_index=insert_index + offset,
        )
        for offset, row in enumerate(pending_registration_rows)
    ]

    if not inserted_rows and repaired_count <= 0 and formula_repaired_count <= 0 and tracking_repaired_count <= 0:
        styled_document, group_style_repaired_count = _apply_attendance_group_identity_backgrounds(
            attendance_document,
            attendance_rows,
            attendance_columns,
        )
        if group_style_repaired_count > 0:
            return styled_document, _build_registration_match_summary(
                updated_count=group_style_repaired_count,
                matched_count=group_style_repaired_count,
                skipped_count=skipped_count,
                repaired_count=group_style_repaired_count,
            )
        return attendance_document, _build_registration_match_summary(skipped_count=skipped_count)

    if inserted_rows:
        formula_row_offset = _get_formula_reference_row_offset(attendance_document)
        existing_rows = _remap_existing_rows_for_insert(
            attendance_rows,
            columns=attendance_columns,
            insert_index=insert_index,
            amount=len(inserted_rows),
            row_index_offset=formula_row_offset,
        )
    else:
        existing_rows = attendance_rows
    next_rows = [
        *existing_rows[:insert_index],
        *inserted_rows,
        *existing_rows[insert_index:],
    ]
    next_document = _replace_document_data_rows(attendance_document, next_rows)
    if inserted_rows:
        next_document = _filter_entity_model_for_document_row_prefix(
            next_document,
            max_document_row=_normalize_document_data_start_row(next_document),
        )
    else:
        for target_index, _source_template_index in repair_meta_targets:
            next_document = _remove_attendance_row_entity_cell_styles(
                next_document,
                target_document_row=cell_meta_row_offset + target_index,
            )
    if isinstance(attendance_document.get("cell_meta"), dict):
        row_offset = cell_meta_row_offset
        column_limit = cell_meta_column_limit
        next_cell_meta = attendance_document.get("cell_meta")
        for target_index, source_template_index in repair_meta_targets:
            next_cell_meta = _apply_attendance_template_cell_meta(
                next_cell_meta,
                template_document_row=row_offset + source_template_index if source_template_index is not None else None,
                target_document_row=row_offset + target_index,
                column_limit=column_limit,
            )
        if inserted_rows and insert_index < len(attendance_rows):
            next_cell_meta = _shift_cell_meta_rows_for_insert(
                next_cell_meta,
                insert_index,
                len(inserted_rows),
                row_offset=row_offset,
            )
        for offset in range(len(inserted_rows)):
            target_index = insert_index + offset
            shifted_template_index = template_index
            if shifted_template_index is not None and shifted_template_index >= insert_index:
                shifted_template_index += len(inserted_rows)
            next_cell_meta = _apply_attendance_template_cell_meta(
                next_cell_meta,
                template_document_row=row_offset + shifted_template_index if shifted_template_index is not None else None,
                target_document_row=row_offset + target_index,
                column_limit=column_limit,
            )
        next_document["cell_meta"] = next_cell_meta
    next_document = _repair_attendance_student_id_column_config(next_document)
    next_document, managed_formula_repaired_count = _normalize_attendance_managed_refund_formulas(next_document)
    if managed_formula_repaired_count:
        next_rows = [
            _normalize_sheet_row(row, len(attendance_columns))
            for row in _extract_document_rows(next_document)
        ]
    next_document, group_style_repaired_count = _apply_attendance_group_identity_backgrounds(
        next_document,
        next_rows,
        attendance_columns,
    )
    return next_document, _build_registration_match_summary(
        updated_count=len(inserted_rows) + repaired_count + formula_repaired_count + tracking_repaired_count + managed_formula_repaired_count + group_style_repaired_count,
        matched_count=len(inserted_rows) + repaired_count + formula_repaired_count + tracking_repaired_count + managed_formula_repaired_count + group_style_repaired_count,
        skipped_count=skipped_count,
        inserted_count=len(inserted_rows),
        repaired_count=repaired_count + formula_repaired_count + tracking_repaired_count + managed_formula_repaired_count + group_style_repaired_count,
    )


def _serialize_note_sheet_action_detail(
    session: Session,
    document: SheetDocument,
    access: NoteSheetResourceAccess,
    current_user: User | None,
) -> NoteSheetDetailResponse:
    workbook_items = _list_workbook_refs_for_sheet_ids(session, [document.id], current_user).get(document.id, [])
    parent_workbook_id = _get_parent_workbook_id_for_sheet(session, document)
    return NoteSheetDetailResponse.model_validate(
        _serialize_sheet_detail(
            document,
            workbook_items=workbook_items,
            parent_workbook_id=parent_workbook_id,
            document_json=dict(document.document_json or {}),
            access=access,
        ),
    )


def _run_registration_match_action(
    *,
    sheet_id: int,
    action: str,
    workbook_id: int | None,
    session: Session,
    current_user: User,
    use_browser_fallback: bool = False,
) -> NoteSheetRegistrationMatchResponse:
    document, access, workbook = _get_note_sheet_or_404(
        session,
        current_user,
        sheet_id,
        required_role="editor",
        workbook_id=workbook_id,
    )
    if not access.capabilities.can_edit_data or not access.capabilities.can_run_sheet_actions:
        raise HTTPException(status_code=403, detail="没有执行报名表动作的权限")

    current_document = _normalize_document_json(dict(document.document_json or {}))
    if action == NOTE_SHEET_CELL_ACTION_REGISTRATION_ORDER_MATCH:
        next_document, summary = _update_registration_order_match_document(
            current_document,
            session=session,
            current_user=current_user,
            use_browser_fallback=use_browser_fallback,
        )
        message = _format_registration_order_match_message(summary)
    elif action == NOTE_SHEET_CELL_ACTION_REGISTRATION_USER_MATCH:
        registration_shop_id = _resolve_registration_shop_id(session, document, workbook)
        next_document, summary = _update_registration_user_match_document(
            current_document,
            session=session,
            current_user=current_user,
            course_name=_get_registration_course_name(document, workbook),
            shop_id=registration_shop_id,
            use_browser_fallback=use_browser_fallback,
        )
        message = _format_registration_user_match_message(summary)
    else:
        raise HTTPException(status_code=400, detail="不支持的报名表动作")

    if current_document != next_document:
        document.document_json = next_document
        document.version = max(int(document.version or 1), 1) + 1
        document.updated_by_user_id = current_user.id
        document.updated_at = time.time()
        session.add(document)
        session.commit()
        session.refresh(document)
        _broadcast_sheet_resource_update(document)

    return NoteSheetRegistrationMatchResponse(
        sheet=_serialize_note_sheet_action_detail(session, document, access, current_user),
        action=action,
        updated_count=summary["updated_count"],
        skipped_count=summary["skipped_count"],
        error_count=summary["error_count"],
        warning_count=summary["warning_count"],
        message=message,
    )


def _default_registration_match_browser_fallback(action: str) -> bool:
    if action in {
        NOTE_SHEET_CELL_ACTION_REGISTRATION_ORDER_MATCH,
        NOTE_SHEET_CELL_ACTION_REGISTRATION_COMPOSITE_UPDATE,
    }:
        return True
    return bool(NOTE_SHEET_REGISTRATION_USER_BROWSER_FALLBACK_DEFAULT)


_REGISTRATION_MATCH_RUN_LOCK = threading.RLock()
_REGISTRATION_MATCH_RUNS: dict[str, dict[str, Any]] = {}
_REGISTRATION_MATCH_ACTIVE_BY_KEY: dict[tuple[int, str], str] = {}
_REGISTRATION_MATCH_TERMINAL_STATUSES = {"completed", "failed", "cancelled"}
_REGISTRATION_MATCH_ACTIVE_STATUSES = {"pending", "running"}


def _registration_match_run_key(sheet_id: int, action: str) -> tuple[int, str]:
    return int(sheet_id), str(action)


def _is_registration_match_run_active(run: dict[str, Any] | None) -> bool:
    return bool(run and run.get("status") in _REGISTRATION_MATCH_ACTIVE_STATUSES and not run.get("cancel_requested"))


def _serialize_registration_match_run(
    run: dict[str, Any] | None,
    *,
    sheet: NoteSheetDetailResponse | None = None,
    already_running: bool = False,
    sheet_id: int | None = None,
    action: str | None = None,
    workbook_id: int | None = None,
) -> NoteSheetRegistrationMatchRunResponse:
    if run is None:
        return NoteSheetRegistrationMatchRunResponse(
            sheet_id=int(sheet_id or 0),
            workbook_id=workbook_id,
            action=str(action or ""),
            sheet=sheet,
        )
    return NoteSheetRegistrationMatchRunResponse(
        run_id=str(run.get("run_id") or ""),
        action=str(run.get("action") or ""),
        sheet_id=int(run.get("sheet_id") or 0),
        workbook_id=run.get("workbook_id"),
        status=run.get("status") or "idle",
        phase=str(run.get("phase") or ""),
        use_browser_fallback=bool(run.get("use_browser_fallback")),
        already_running=already_running,
        cancel_requested=bool(run.get("cancel_requested")),
        queued_at=run.get("queued_at"),
        started_at=run.get("started_at"),
        finished_at=run.get("finished_at"),
        total_count=int(run.get("total_count") or 0),
        processed_count=int(run.get("processed_count") or 0),
        updated_count=int(run.get("updated_count") or 0),
        skipped_count=int(run.get("skipped_count") or 0),
        error_count=int(run.get("error_count") or 0),
        warning_count=int(run.get("warning_count") or 0),
        message=str(run.get("message") or ""),
        error_message=run.get("error_message"),
        sheet=sheet,
    )


def _get_registration_match_run_snapshot(run_id: str) -> dict[str, Any] | None:
    with _REGISTRATION_MATCH_RUN_LOCK:
        run = _REGISTRATION_MATCH_RUNS.get(run_id)
        return dict(run) if run else None


def _get_active_registration_match_run_snapshot(sheet_id: int, action: str | None = None) -> dict[str, Any] | None:
    with _REGISTRATION_MATCH_RUN_LOCK:
        actions = [action] if action else sorted(NOTE_SHEET_REGISTRATION_BACKGROUND_ACTIONS)
        for item in actions:
            key = _registration_match_run_key(sheet_id, str(item))
            run_id = _REGISTRATION_MATCH_ACTIVE_BY_KEY.get(key)
            run = _REGISTRATION_MATCH_RUNS.get(run_id or "")
            if _is_registration_match_run_active(run):
                return dict(run)
            if run_id and run and run.get("status") in _REGISTRATION_MATCH_TERMINAL_STATUSES:
                _REGISTRATION_MATCH_ACTIVE_BY_KEY.pop(key, None)
        return None


def _get_conflicting_registration_match_runs(sheet_id: int) -> list[dict[str, Any]]:
    runs: list[dict[str, Any]] = []
    with _REGISTRATION_MATCH_RUN_LOCK:
        for action in sorted(NOTE_SHEET_REGISTRATION_BACKGROUND_ACTIONS):
            key = _registration_match_run_key(sheet_id, action)
            run_id = _REGISTRATION_MATCH_ACTIVE_BY_KEY.get(key)
            run = _REGISTRATION_MATCH_RUNS.get(run_id or "")
            if _is_registration_match_run_active(run):
                runs.append(dict(run))
            elif run_id and run and run.get("status") in _REGISTRATION_MATCH_TERMINAL_STATUSES:
                _REGISTRATION_MATCH_ACTIVE_BY_KEY.pop(key, None)
    return runs


def _update_registration_match_run(run_id: str, **updates: Any) -> dict[str, Any] | None:
    with _REGISTRATION_MATCH_RUN_LOCK:
        run = _REGISTRATION_MATCH_RUNS.get(run_id)
        if run is None:
            return None
        run.update(updates)
        return dict(run)


def _request_cancel_registration_match_run(run_id: str) -> None:
    now = time.time()
    with _REGISTRATION_MATCH_RUN_LOCK:
        run = _REGISTRATION_MATCH_RUNS.get(run_id)
        if run is None or run.get("status") in _REGISTRATION_MATCH_TERMINAL_STATUSES:
            return
        run["cancel_requested"] = True
        run["status"] = "cancelled"
        run["finished_at"] = now
        run["message"] = "已请求停止并准备重新开始"
        key = _registration_match_run_key(int(run.get("sheet_id") or 0), str(run.get("action") or ""))
        if _REGISTRATION_MATCH_ACTIVE_BY_KEY.get(key) == run_id:
            _REGISTRATION_MATCH_ACTIVE_BY_KEY.pop(key, None)


def _is_registration_match_run_current(run_id: str) -> bool:
    with _REGISTRATION_MATCH_RUN_LOCK:
        run = _REGISTRATION_MATCH_RUNS.get(run_id)
        if not run or run.get("cancel_requested"):
            return False
        key = _registration_match_run_key(int(run.get("sheet_id") or 0), str(run.get("action") or ""))
        return _REGISTRATION_MATCH_ACTIVE_BY_KEY.get(key) == run_id


def _finish_registration_match_run(run_id: str, status: str, *, message: str = "", error_message: str | None = None) -> None:
    now = time.time()
    with _REGISTRATION_MATCH_RUN_LOCK:
        run = _REGISTRATION_MATCH_RUNS.get(run_id)
        if run is None:
            return
        if run.get("status") == "cancelled" and status != "failed":
            status = "cancelled"
        run["status"] = status
        run["finished_at"] = now
        if message:
            run["message"] = message
        if error_message:
            run["error_message"] = error_message
        key = _registration_match_run_key(int(run.get("sheet_id") or 0), str(run.get("action") or ""))
        if _REGISTRATION_MATCH_ACTIVE_BY_KEY.get(key) == run_id:
            _REGISTRATION_MATCH_ACTIVE_BY_KEY.pop(key, None)


def _count_registration_user_match_targets(document_json: dict[str, Any]) -> int:
    normalized = _normalize_document_json(document_json)
    columns = _normalize_document_columns(normalized)
    indexes = _find_required_registration_column_indexes(columns, NOTE_SHEET_REGISTRATION_USER_LOOKUP_COLUMNS)
    rows = [_normalize_sheet_row(row, len(columns)) for row in _extract_document_rows(normalized)]
    count = 0
    for row in rows:
        if _normalize_sheet_text(row[indexes["用户ID"]]):
            continue
        names = [
            _normalize_sheet_text(row[indexes["姓名"]]),
            _normalize_sheet_text(row[indexes["微信昵称"]]),
        ]
        phones = [
            _strip_legacy_text_prefix(row[indexes["手机号"]]),
            _strip_legacy_text_prefix(row[indexes["错误手机号"]]),
        ]
        if any(names) or any(phone and phone.lower() != "none" for phone in phones):
            count += 1
    return count


def _save_registration_match_row(
    *,
    session: Session,
    sheet_id: int,
    row_position: int,
    row: list[Any],
    current_user_id: int | None,
    run_id: str,
) -> bool:
    if not _is_registration_match_run_current(run_id):
        return False
    document = _get_sheet_by_numeric_id_or_404(session, sheet_id)
    current_document = _normalize_document_json(dict(document.document_json or {}))
    columns = _normalize_document_columns(current_document)
    current_rows = [_normalize_sheet_row(item, len(columns)) for item in _extract_document_rows(current_document)]
    if row_position < 0 or row_position >= len(current_rows):
        return False
    next_row = _normalize_sheet_row(row, len(columns))
    if current_rows[row_position] == next_row:
        return False
    current_rows[row_position] = next_row
    document.document_json = _replace_document_data_rows(current_document, current_rows)
    document.version = max(int(document.version or 1), 1) + 1
    document.updated_by_user_id = current_user_id
    document.updated_at = time.time()
    session.add(document)
    session.commit()
    session.refresh(document)
    _broadcast_sheet_resource_update(document)
    return True


def _run_registration_order_match_background(
    *,
    run_id: str,
    sheet_id: int,
    workbook_id: int | None,
    current_user_snapshot: dict[str, Any],
    use_browser_fallback: bool,
) -> None:
    _ = use_browser_fallback
    try:
        _update_registration_match_run(
            run_id,
            status="running",
            phase=NOTE_SHEET_CELL_ACTION_REGISTRATION_ORDER_MATCH,
            started_at=time.time(),
            message="正在更新订单匹配",
        )
        with Session(engine) as session:
            current_user = User(
                id=int(current_user_snapshot.get("id") or 0),
                username=str(current_user_snapshot.get("username") or ""),
                hashed_password="",
                is_active=True,
                is_superuser=bool(current_user_snapshot.get("is_superuser")),
            )
            document, access, _workbook = _get_note_sheet_or_404(
                session,
                current_user,
                sheet_id,
                required_role="editor",
                workbook_id=workbook_id,
            )
            if not access.capabilities.can_edit_data or not access.capabilities.can_run_sheet_actions:
                raise HTTPException(status_code=403, detail="没有执行报名表动作的权限")

            current_document = _normalize_document_json(dict(document.document_json or {}))
            total_count = _count_registration_order_match_targets(current_document)
            _update_registration_match_run(run_id, total_count=total_count)
            if not _is_registration_match_run_current(run_id):
                _finish_registration_match_run(run_id, "cancelled", message="订单匹配已停止")
                return

            next_document, summary = _update_registration_order_match_document(
                current_document,
                session=session,
                current_user=current_user,
                use_browser_fallback=use_browser_fallback,
            )
            if current_document != next_document:
                document.document_json = next_document
                document.version = max(int(document.version or 1), 1) + 1
                document.updated_by_user_id = current_user.id
                document.updated_at = time.time()
                session.add(document)
                session.commit()
                _broadcast_sheet_resource_update(document)
            _update_registration_match_run(
                run_id,
                processed_count=total_count,
                updated_count=summary["updated_count"],
                skipped_count=summary["skipped_count"],
                error_count=summary["error_count"],
                warning_count=summary["warning_count"],
            )
            _finish_registration_match_run(
                run_id,
                "completed",
                message=_format_registration_order_match_message(summary),
            )
    except Exception as exc:
        _finish_registration_match_run(
            run_id,
            "failed",
            message="订单匹配任务失败",
            error_message=str(exc.detail if isinstance(exc, HTTPException) else exc),
        )


def _run_registration_user_match_background(
    *,
    run_id: str,
    sheet_id: int,
    workbook_id: int | None,
    current_user_snapshot: dict[str, Any],
    use_browser_fallback: bool,
) -> None:
    updated_positions: set[int] = set()
    try:
        _update_registration_match_run(
            run_id,
            status="running",
            phase=NOTE_SHEET_CELL_ACTION_REGISTRATION_USER_MATCH,
            started_at=time.time(),
            message="正在匹配用户",
        )
        with Session(engine) as session:
            current_user_id = int(current_user_snapshot.get("id") or 0)
            current_user = User(
                id=current_user_id,
                username=str(current_user_snapshot.get("username") or ""),
                hashed_password="",
                is_active=True,
                is_superuser=bool(current_user_snapshot.get("is_superuser")),
            )
            document, access, workbook = _get_note_sheet_or_404(
                session,
                current_user,
                sheet_id,
                required_role="editor",
                workbook_id=workbook_id,
            )
            if not access.capabilities.can_edit_data or not access.capabilities.can_run_sheet_actions:
                raise HTTPException(status_code=403, detail="没有执行报名表动作的权限")

            normalized = _normalize_document_json(dict(document.document_json or {}))
            columns = _normalize_document_columns(normalized)
            indexes = _find_required_registration_column_indexes(columns, NOTE_SHEET_REGISTRATION_USER_LOOKUP_COLUMNS)
            rows = [_normalize_sheet_row(row, len(columns)) for row in _extract_document_rows(normalized)]
            course_name = _get_registration_course_name(document, workbook)
            registration_shop_id = _resolve_registration_shop_id(session, document, workbook)
            total_count = _count_registration_user_match_targets(normalized)
            _update_registration_match_run(run_id, total_count=total_count)

            get_kqdb = _load_attendance_kqdb_provider()
            lookup_user = _load_attendance_user_lookup_provider()
            kqdb = get_kqdb()

            processed_count = 0
            skipped_count = 0
            error_count = 0
            warning_count = 0
            matched_count = 0
            unmatched_count = 0

            def mark_updated(row_position: int) -> None:
                updated_positions.add(row_position)
                _update_registration_match_run(run_id, updated_count=len(updated_positions))

            for row_position, source_row in enumerate(rows):
                if not _is_registration_match_run_current(run_id):
                    _finish_registration_match_run(run_id, "cancelled", message="用户匹配已停止")
                    return

                row = list(source_row)
                if _normalize_sheet_text(row[indexes["用户ID"]]):
                    skipped_count += 1
                    _update_registration_match_run(run_id, skipped_count=skipped_count)
                    continue

                names = [
                    _normalize_sheet_text(row[indexes["姓名"]]),
                    _normalize_sheet_text(row[indexes["微信昵称"]]),
                ]
                phones = [
                    _strip_legacy_text_prefix(row[indexes["手机号"]]),
                    _strip_legacy_text_prefix(row[indexes["错误手机号"]]),
                ]
                names = [item for item in names if item]
                phones = [item for item in phones if item and item.lower() != "none"]
                if not names and not phones:
                    skipped_count += 1
                    _update_registration_match_run(run_id, skipped_count=skipped_count)
                    continue

                before = list(row)
                try:
                    user_id, weight = lookup_user(
                        names,
                        phones,
                        course_name=course_name,
                        course_product_name="",
                        shop_id=registration_shop_id,
                        return_mode=1,
                        kqdb=kqdb,
                    )
                    row[indexes["用户ID"]] = _format_registration_match_cell(user_id)
                    row[indexes["匹配得分"]] = _format_registration_match_cell(weight) if weight is not None else ""
                    if user_id:
                        matched_count += 1
                    elif not use_browser_fallback:
                        unmatched_count += 1
                        warning_count += 1
                    if _save_registration_match_row(
                        session=session,
                        sheet_id=sheet_id,
                        row_position=row_position,
                        row=row,
                        current_user_id=current_user_id,
                        run_id=run_id,
                    ) or row != before:
                        mark_updated(row_position)

                    if not user_id and use_browser_fallback:
                        result_map = _lookup_registration_users_with_remote_browser(
                            session,
                            current_user,
                            course_name=course_name,
                            shop_id=registration_shop_id,
                            items=[{"key": "0", "names": names, "phones": phones}],
                        )
                        result = result_map.get("0", {})
                        row_before_remote = list(row)
                        result_error = _normalize_sheet_text(result.get("error"))
                        remote_user_id = _normalize_sheet_text(result.get("user_id"))
                        if result_error:
                            row[indexes["匹配得分"]] = result_error
                            error_count += 1
                        elif remote_user_id:
                            row[indexes["用户ID"]] = remote_user_id
                            row[indexes["匹配得分"]] = "95"
                            matched_count += 1
                        else:
                            unmatched_count += 1
                            warning_count += 1
                        if _save_registration_match_row(
                            session=session,
                            sheet_id=sheet_id,
                            row_position=row_position,
                            row=row,
                            current_user_id=current_user_id,
                            run_id=run_id,
                        ) or row != row_before_remote:
                            mark_updated(row_position)
                except Exception as exc:
                    row[indexes["匹配得分"]] = str(exc.detail if isinstance(exc, HTTPException) else exc)
                    error_count += 1
                    if _save_registration_match_row(
                        session=session,
                        sheet_id=sheet_id,
                        row_position=row_position,
                        row=row,
                        current_user_id=current_user_id,
                        run_id=run_id,
                    ) or row != before:
                        mark_updated(row_position)
                finally:
                    processed_count += 1
                    _update_registration_match_run(
                        run_id,
                        processed_count=processed_count,
                        skipped_count=skipped_count,
                        error_count=error_count,
                        warning_count=warning_count,
                        message=f"已处理 {processed_count}/{total_count} 行用户匹配",
                    )

            summary = _build_registration_match_summary(
                target_count=total_count,
                updated_count=len(updated_positions),
                skipped_count=skipped_count,
                error_count=error_count,
                warning_count=warning_count,
                matched_count=matched_count,
                unmatched_count=unmatched_count,
            )
            _finish_registration_match_run(
                run_id,
                "completed",
                message=_format_registration_user_match_message(summary),
            )
    except Exception as exc:
        _finish_registration_match_run(
            run_id,
            "failed",
            message="用户匹配任务失败",
            error_message=str(exc.detail if isinstance(exc, HTTPException) else exc),
        )


def _run_registration_composite_update_background(
    *,
    run_id: str,
    sheet_id: int,
    workbook_id: int | None,
    current_user_snapshot: dict[str, Any],
    use_browser_fallback: bool,
) -> None:
    try:
        _update_registration_match_run(
            run_id,
            status="running",
            phase=NOTE_SHEET_CELL_ACTION_REGISTRATION_ORDER_MATCH,
            started_at=time.time(),
            total_count=3,
            processed_count=0,
            message="正在更新订单匹配",
        )
        with Session(engine) as session:
            current_user = User(
                id=int(current_user_snapshot.get("id") or 0),
                username=str(current_user_snapshot.get("username") or ""),
                hashed_password="",
                is_active=True,
                is_superuser=bool(current_user_snapshot.get("is_superuser")),
            )
            document, access, workbook = _get_note_sheet_or_404(
                session,
                current_user,
                sheet_id,
                required_role="editor",
                workbook_id=workbook_id,
            )
            if not access.capabilities.can_edit_data or not access.capabilities.can_run_sheet_actions:
                raise HTTPException(status_code=403, detail="没有执行报名表动作的权限")

            updated_count = 0
            skipped_count = 0
            error_count = 0
            warning_count = 0
            attendance_summary = _build_registration_match_summary()

            if not _is_registration_match_run_current(run_id):
                _finish_registration_match_run(run_id, "cancelled", message="综合更新已停止")
                return
            current_document = _normalize_document_json(dict(document.document_json or {}))
            next_document, order_summary = _update_registration_order_match_document(
                current_document,
                session=session,
                current_user=current_user,
                use_browser_fallback=use_browser_fallback,
            )
            if current_document != next_document:
                document.document_json = next_document
                document.version = max(int(document.version or 1), 1) + 1
                document.updated_by_user_id = current_user.id
                document.updated_at = time.time()
                session.add(document)
                session.commit()
                session.refresh(document)
                _broadcast_sheet_resource_update(document)
            updated_count += order_summary["updated_count"]
            skipped_count += order_summary["skipped_count"]
            error_count += order_summary["error_count"]
            warning_count += order_summary.get("warning_count", 0)
            _update_registration_match_run(
                run_id,
                processed_count=1,
                updated_count=updated_count,
                skipped_count=skipped_count,
                error_count=error_count,
                warning_count=warning_count,
                phase=NOTE_SHEET_CELL_ACTION_REGISTRATION_USER_MATCH,
                message="正在更新用户匹配",
            )

            if not _is_registration_match_run_current(run_id):
                _finish_registration_match_run(run_id, "cancelled", message="综合更新已停止")
                return
            current_document = _normalize_document_json(dict(document.document_json or {}))
            next_document, user_summary = _update_registration_user_match_document(
                current_document,
                session=session,
                current_user=current_user,
                course_name=_get_registration_course_name(document, workbook),
                shop_id=_resolve_registration_shop_id(session, document, workbook),
                use_browser_fallback=use_browser_fallback,
            )
            if current_document != next_document:
                document.document_json = next_document
                document.version = max(int(document.version or 1), 1) + 1
                document.updated_by_user_id = current_user.id
                document.updated_at = time.time()
                session.add(document)
                session.commit()
                session.refresh(document)
                _broadcast_sheet_resource_update(document)
            updated_count += user_summary["updated_count"]
            skipped_count += user_summary["skipped_count"]
            error_count += user_summary["error_count"]
            warning_count += user_summary.get("warning_count", 0)
            _update_registration_match_run(
                run_id,
                processed_count=2,
                updated_count=updated_count,
                skipped_count=skipped_count,
                error_count=error_count,
                warning_count=warning_count,
                phase="registration_attendance_sync",
                message="正在同步考勤表",
            )

            if not _is_registration_match_run_current(run_id):
                _finish_registration_match_run(run_id, "cancelled", message="综合更新已停止")
                return
            attendance, attendance_workbook = _resolve_registration_attendance_sheet(session, document, workbook)
            if attendance is None:
                skipped_count += 1
                error_count += 1
                attendance_summary = _build_registration_match_summary(
                    skipped_count=1,
                    error_count=1,
                )
                _update_registration_match_run(
                    run_id,
                    processed_count=3,
                    updated_count=updated_count,
                    skipped_count=skipped_count,
                    error_count=error_count,
                    warning_count=warning_count,
                    message="未找到同工作簿的考勤表",
                )
            else:
                attendance_access = _resolve_sheet_resource_access(
                    session,
                    attendance,
                    current_user,
                    workbook=attendance_workbook,
                )
                if not attendance_access.capabilities.can_edit_data:
                    raise HTTPException(status_code=403, detail="没有编辑考勤表的权限")
                registration_document = _normalize_document_json(dict(document.document_json or {}))
                attendance_document = _normalize_document_json(dict(attendance.document_json or {}))
                next_attendance_document, attendance_summary = _sync_registration_rows_to_attendance_document(
                    registration_document,
                    attendance_document,
                )
                if attendance_document != next_attendance_document:
                    attendance.document_json = next_attendance_document
                    attendance.version = max(int(attendance.version or 1), 1) + 1
                    attendance.updated_by_user_id = current_user.id
                    attendance.updated_at = time.time()
                    session.add(attendance)
                    session.commit()
                updated_count += attendance_summary["updated_count"]
                skipped_count += attendance_summary["skipped_count"]
                error_count += attendance_summary["error_count"]
                warning_count += attendance_summary.get("warning_count", 0)
                _update_registration_match_run(
                    run_id,
                    processed_count=3,
                    updated_count=updated_count,
                    skipped_count=skipped_count,
                    error_count=error_count,
                    warning_count=warning_count,
                )

            _finish_registration_match_run(
                run_id,
                "completed",
                message=_format_registration_composite_update_message(order_summary, user_summary, attendance_summary),
            )
    except Exception as exc:
        _finish_registration_match_run(
            run_id,
            "failed",
            message="综合更新任务失败",
            error_message=str(exc.detail if isinstance(exc, HTTPException) else exc),
        )


def _start_registration_match_run(
    *,
    sheet_id: int,
    workbook_id: int | None,
    action: str,
    current_user: User,
    use_browser_fallback: bool,
    force_restart: bool,
) -> NoteSheetRegistrationMatchRunResponse:
    if action not in NOTE_SHEET_REGISTRATION_BACKGROUND_ACTIONS:
        raise HTTPException(status_code=400, detail="该动作暂未接入后台任务")
    with _REGISTRATION_MATCH_RUN_LOCK:
        conflicting_runs = _get_conflicting_registration_match_runs(sheet_id)
        if conflicting_runs:
            if not force_restart:
                return _serialize_registration_match_run(dict(conflicting_runs[0]), already_running=True)
            for active_run in conflicting_runs:
                _request_cancel_registration_match_run(str(active_run.get("run_id") or ""))

        run_id = uuid.uuid4().hex
        run = {
            "run_id": run_id,
            "action": action,
            "sheet_id": int(sheet_id),
            "workbook_id": workbook_id,
            "status": "pending",
            "phase": "",
            "use_browser_fallback": bool(use_browser_fallback),
            "current_user_id": current_user.id,
            "cancel_requested": False,
            "queued_at": time.time(),
            "started_at": None,
            "finished_at": None,
            "total_count": 0,
            "processed_count": 0,
            "updated_count": 0,
            "skipped_count": 0,
            "error_count": 0,
            "warning_count": 0,
            "message": "任务已排队",
            "error_message": None,
        }
        _REGISTRATION_MATCH_RUNS[run_id] = run
        _REGISTRATION_MATCH_ACTIVE_BY_KEY[_registration_match_run_key(sheet_id, action)] = run_id

    if action == NOTE_SHEET_CELL_ACTION_REGISTRATION_USER_MATCH:
        target = _run_registration_user_match_background
    elif action == NOTE_SHEET_CELL_ACTION_REGISTRATION_ORDER_MATCH:
        target = _run_registration_order_match_background
    elif action == NOTE_SHEET_CELL_ACTION_REGISTRATION_COMPOSITE_UPDATE:
        target = _run_registration_composite_update_background
    else:
        raise HTTPException(status_code=400, detail="该动作暂未接入后台任务")

    thread = threading.Thread(
        target=target,
        kwargs={
            "run_id": run_id,
            "sheet_id": sheet_id,
            "workbook_id": workbook_id,
            "current_user_snapshot": {
                "id": current_user.id,
                "username": current_user.username,
                "is_superuser": current_user.is_superuser,
            },
            "use_browser_fallback": use_browser_fallback,
        },
        name=f"note-sheet-registration-match-{run_id[:8]}",
        daemon=True,
    )
    thread.start()
    return _serialize_registration_match_run(_get_registration_match_run_snapshot(run_id))


_CLOCKIN_LINK_DETECTION_RUN_LOCK = threading.Lock()
_CLOCKIN_LINK_DETECTION_RUNS: dict[str, dict[str, Any]] = {}
_CLOCKIN_LINK_DETECTION_ACTIVE_BY_SHEET: dict[int, str] = {}
_CLOCKIN_LINK_DETECTION_TERMINAL_STATUSES = {"completed", "failed", "cancelled"}
_CLOCKIN_LINK_DETECTION_ACTIVE_STATUSES = {"pending", "running"}


def _is_clockin_link_detection_run_active(run: dict[str, Any] | None) -> bool:
    return bool(run and run.get("status") in _CLOCKIN_LINK_DETECTION_ACTIVE_STATUSES and not run.get("cancel_requested"))


def _serialize_clockin_link_detection_run(
    run: dict[str, Any] | None,
    *,
    sheet: NoteSheetDetailResponse | None = None,
    already_running: bool = False,
    sheet_id: int | None = None,
    workbook_id: int | None = None,
) -> NoteSheetClockinLinkDetectionRunResponse:
    if run is None:
        return NoteSheetClockinLinkDetectionRunResponse(
            sheet_id=int(sheet_id or 0),
            workbook_id=workbook_id,
            sheet=sheet,
        )
    return NoteSheetClockinLinkDetectionRunResponse(
        run_id=str(run.get("run_id") or ""),
        action=NOTE_SHEET_CELL_ACTION_CLOCKIN_LINK_DETECT,
        sheet_id=int(run.get("sheet_id") or 0),
        workbook_id=run.get("workbook_id"),
        status=run.get("status") or "idle",
        phase=str(run.get("phase") or ""),
        already_running=already_running,
        cancel_requested=bool(run.get("cancel_requested")),
        queued_at=run.get("queued_at"),
        started_at=run.get("started_at"),
        finished_at=run.get("finished_at"),
        total_count=int(run.get("total_count") or 0),
        processed_count=int(run.get("processed_count") or 0),
        updated_count=int(run.get("updated_count") or 0),
        skipped_count=int(run.get("skipped_count") or 0),
        error_count=int(run.get("error_count") or 0),
        warning_count=int(run.get("warning_count") or 0),
        message=str(run.get("message") or ""),
        error_message=run.get("error_message"),
        provider_id=str(run.get("provider_id") or NOTE_SHEET_CLOCKIN_LINK_DETECTION_PROVIDER_ID),
        model=str(run.get("model") or NOTE_SHEET_CLOCKIN_LINK_DETECTION_MODEL),
        results=list(run.get("results") or []),
        warnings=list(run.get("warnings") or []),
        sheet=sheet,
    )


def _get_clockin_link_detection_run_snapshot(run_id: str) -> dict[str, Any] | None:
    with _CLOCKIN_LINK_DETECTION_RUN_LOCK:
        run = _CLOCKIN_LINK_DETECTION_RUNS.get(run_id)
        return dict(run) if run else None


def _get_active_clockin_link_detection_run_snapshot(sheet_id: int) -> dict[str, Any] | None:
    with _CLOCKIN_LINK_DETECTION_RUN_LOCK:
        run_id = _CLOCKIN_LINK_DETECTION_ACTIVE_BY_SHEET.get(int(sheet_id))
        run = _CLOCKIN_LINK_DETECTION_RUNS.get(run_id or "")
        if _is_clockin_link_detection_run_active(run):
            return dict(run)
        if run_id and run and run.get("status") in _CLOCKIN_LINK_DETECTION_TERMINAL_STATUSES:
            _CLOCKIN_LINK_DETECTION_ACTIVE_BY_SHEET.pop(int(sheet_id), None)
        return None


def _update_clockin_link_detection_run(run_id: str, **updates: Any) -> dict[str, Any] | None:
    with _CLOCKIN_LINK_DETECTION_RUN_LOCK:
        run = _CLOCKIN_LINK_DETECTION_RUNS.get(run_id)
        if run is None:
            return None
        run.update(updates)
        return dict(run)


def _request_cancel_clockin_link_detection_run(run_id: str) -> None:
    now = time.time()
    with _CLOCKIN_LINK_DETECTION_RUN_LOCK:
        run = _CLOCKIN_LINK_DETECTION_RUNS.get(run_id)
        if run is None or run.get("status") in _CLOCKIN_LINK_DETECTION_TERMINAL_STATUSES:
            return
        run["cancel_requested"] = True
        run["status"] = "cancelled"
        run["finished_at"] = now
        run["message"] = "已请求停止并准备重新开始"
        sheet_id = int(run.get("sheet_id") or 0)
        if _CLOCKIN_LINK_DETECTION_ACTIVE_BY_SHEET.get(sheet_id) == run_id:
            _CLOCKIN_LINK_DETECTION_ACTIVE_BY_SHEET.pop(sheet_id, None)


def _is_clockin_link_detection_run_current(run_id: str) -> bool:
    with _CLOCKIN_LINK_DETECTION_RUN_LOCK:
        run = _CLOCKIN_LINK_DETECTION_RUNS.get(run_id)
        if not run or run.get("cancel_requested"):
            return False
        sheet_id = int(run.get("sheet_id") or 0)
        return _CLOCKIN_LINK_DETECTION_ACTIVE_BY_SHEET.get(sheet_id) == run_id


def _finish_clockin_link_detection_run(
    run_id: str,
    status: str,
    *,
    message: str = "",
    error_message: str | None = None,
) -> None:
    now = time.time()
    with _CLOCKIN_LINK_DETECTION_RUN_LOCK:
        run = _CLOCKIN_LINK_DETECTION_RUNS.get(run_id)
        if run is None:
            return
        if run.get("status") == "cancelled" and status != "failed":
            status = "cancelled"
        run["status"] = status
        run["finished_at"] = now
        if message:
            run["message"] = message
        if error_message:
            run["error_message"] = error_message
        sheet_id = int(run.get("sheet_id") or 0)
        if _CLOCKIN_LINK_DETECTION_ACTIVE_BY_SHEET.get(sheet_id) == run_id:
            _CLOCKIN_LINK_DETECTION_ACTIVE_BY_SHEET.pop(sheet_id, None)


def _run_clockin_link_detection_background(
    *,
    run_id: str,
    sheet_id: int,
    workbook_id: int | None,
    current_user_snapshot: dict[str, Any],
    provider_id: str,
    model: str,
) -> None:
    try:
        _update_clockin_link_detection_run(
            run_id,
            status="running",
            phase="read_config",
            started_at=time.time(),
            message="正在读取打卡配置",
        )
        with Session(engine) as session:
            current_user = User(
                id=int(current_user_snapshot.get("id") or 0),
                username=str(current_user_snapshot.get("username") or ""),
                hashed_password="",
                is_active=True,
                is_superuser=bool(current_user_snapshot.get("is_superuser")),
            )
            document, access, workbook = _get_note_sheet_or_404(
                session,
                current_user,
                sheet_id,
                required_role="editor",
                workbook_id=workbook_id,
            )
            if not access.capabilities.can_edit_data or not access.capabilities.can_run_sheet_actions:
                raise HTTPException(status_code=403, detail="没有执行打卡链接检测的权限")

            current_document = _normalize_document_json(dict(document.document_json or {}))
            root_url = _get_effective_defined_name_literal(session, document, workbook, "打卡根目录")
            if not root_url:
                raise HTTPException(status_code=400, detail="名称管理器缺少“打卡根目录”")
            if not root_url.startswith(("http://", "https://")):
                raise HTTPException(status_code=400, detail="“打卡根目录”不是有效 URL")

            targets = _extract_clockin_detection_targets(current_document)
            if not targets:
                raise HTTPException(status_code=400, detail="打卡配置表没有可检测的打卡行")
            _update_clockin_link_detection_run(
                run_id,
                total_count=len(targets),
                phase="remote_detect",
                message=f"正在调用 mi15 检测 {len(targets)} 个打卡链接",
            )
            if not _is_clockin_link_detection_run_current(run_id):
                _finish_clockin_link_detection_run(run_id, "cancelled", message="自动检测打卡链接已停止")
                return

            detection_result = _detect_clockin_links_with_remote_browser(
                session,
                current_user,
                root_url=root_url,
                targets=[str(item["target"]) for item in targets],
                provider_id=provider_id,
                model=model,
            )
            result_items = detection_result.get("results") if isinstance(detection_result.get("results"), list) else []
            _update_clockin_link_detection_run(
                run_id,
                processed_count=len(result_items),
                results=[dict(item) for item in result_items if isinstance(item, dict)],
                warnings=[str(item) for item in (detection_result.get("warnings") or [])],
                phase="write_sheet",
                message="正在写回打卡配置",
            )
            if not _is_clockin_link_detection_run_current(run_id):
                _finish_clockin_link_detection_run(run_id, "cancelled", message="自动检测打卡链接已停止")
                return

            next_document, summary = _apply_clockin_link_detection_results(
                current_document,
                targets,
                detection_result,
            )
            if next_document != current_document:
                document.document_json = next_document
                document.version = max(int(document.version or 1), 1) + 1
                document.updated_by_user_id = current_user.id
                document.updated_at = time.time()
                session.add(document)
                session.commit()
                _broadcast_sheet_resource_update(document)
            _update_clockin_link_detection_run(
                run_id,
                processed_count=len(targets),
                updated_count=summary["updated_count"],
                skipped_count=summary["skipped_count"],
                error_count=summary["error_count"],
                warning_count=summary["warning_count"],
                warnings=summary["warnings"],
            )
            message = f"自动检测打卡链接完成：写入 {summary['updated_count']} 条"
            if summary["error_count"]:
                message += f"，{summary['error_count']} 条异常"
            _finish_clockin_link_detection_run(run_id, "completed", message=message)
    except Exception as exc:
        _finish_clockin_link_detection_run(
            run_id,
            "failed",
            message="自动检测打卡链接失败",
            error_message=str(exc.detail if isinstance(exc, HTTPException) else exc),
        )


def _start_clockin_link_detection_run(
    *,
    sheet_id: int,
    workbook_id: int | None,
    current_user: User,
    provider_id: str,
    model: str,
    force_restart: bool,
) -> NoteSheetClockinLinkDetectionRunResponse:
    active_run = _get_active_clockin_link_detection_run_snapshot(sheet_id)
    if active_run:
        if not force_restart:
            return _serialize_clockin_link_detection_run(active_run, already_running=True)
        _request_cancel_clockin_link_detection_run(str(active_run.get("run_id") or ""))

    run_id = uuid.uuid4().hex
    run = {
        "run_id": run_id,
        "action": NOTE_SHEET_CELL_ACTION_CLOCKIN_LINK_DETECT,
        "sheet_id": int(sheet_id),
        "workbook_id": workbook_id,
        "status": "pending",
        "phase": "",
        "current_user_id": current_user.id,
        "cancel_requested": False,
        "queued_at": time.time(),
        "started_at": None,
        "finished_at": None,
        "total_count": 0,
        "processed_count": 0,
        "updated_count": 0,
        "skipped_count": 0,
        "error_count": 0,
        "warning_count": 0,
        "message": "任务已排队",
        "error_message": None,
        "provider_id": provider_id,
        "model": model,
        "results": [],
        "warnings": [],
    }
    with _CLOCKIN_LINK_DETECTION_RUN_LOCK:
        _CLOCKIN_LINK_DETECTION_RUNS[run_id] = run
        _CLOCKIN_LINK_DETECTION_ACTIVE_BY_SHEET[int(sheet_id)] = run_id

    thread = threading.Thread(
        target=_run_clockin_link_detection_background,
        kwargs={
            "run_id": run_id,
            "sheet_id": sheet_id,
            "workbook_id": workbook_id,
            "current_user_snapshot": {
                "id": current_user.id,
                "username": current_user.username,
                "is_superuser": current_user.is_superuser,
            },
            "provider_id": provider_id,
            "model": model,
        },
        name=f"note-sheet-clockin-links-{run_id[:8]}",
        daemon=True,
    )
    thread.start()
    return _serialize_clockin_link_detection_run(_get_clockin_link_detection_run_snapshot(run_id))


def _extract_excel_workbook_payload(raw_bytes: bytes, filename: str) -> dict[str, Any]:
    if len(raw_bytes) > NOTE_SHEET_EXCEL_IMPORT_MAX_BYTES:
        raise HTTPException(status_code=413, detail="Excel 文件过大，当前导入上限为 12MB")

    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="服务端缺少 openpyxl，无法解析 Excel") from exc

    try:
        workbook = load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Excel 文件无法读取：{exc}") from exc

    source_sheet_count = len(workbook.sheetnames)
    sheets: list[dict[str, Any]] = []
    total_nonempty_rows = 0
    try:
        for worksheet in workbook.worksheets[:NOTE_SHEET_EXCEL_IMPORT_MAX_SHEETS]:
            max_row = min(int(worksheet.max_row or 0), NOTE_SHEET_EXCEL_IMPORT_MAX_ROWS_PER_SHEET)
            max_col = min(int(worksheet.max_column or 0), NOTE_SHEET_EXCEL_IMPORT_MAX_COLS_PER_SHEET)
            sheet_rows: list[dict[str, Any]] = []
            if max_row <= 0 or max_col <= 0:
                sheets.append({
                    "name": worksheet.title,
                    "max_row": int(worksheet.max_row or 0),
                    "max_column": int(worksheet.max_column or 0),
                    "rows": sheet_rows,
                })
                continue

            for row_number, row in enumerate(worksheet.iter_rows(max_row=max_row, max_col=max_col), start=1):
                values = [_normalize_excel_import_cell(cell.value) for cell in row]
                while values and not values[-1]:
                    values.pop()
                if not any(values):
                    continue
                total_nonempty_rows += 1
                if total_nonempty_rows > NOTE_SHEET_EXCEL_IMPORT_MAX_NONEMPTY_ROWS:
                    raise HTTPException(status_code=413, detail="Excel 非空行过多，当前首版导入上限为 900 行")
                sheet_rows.append({
                    "row_number": row_number,
                    "values": values,
                })

            sheets.append({
                "name": worksheet.title,
                "max_row": int(worksheet.max_row or 0),
                "max_column": int(worksheet.max_column or 0),
                "scanned_row_count": max_row,
                "scanned_column_count": max_col,
                "rows": sheet_rows,
            })
    finally:
        workbook.close()

    if not any(sheet.get("rows") for sheet in sheets):
        raise HTTPException(status_code=400, detail="Excel 中没有可读取的非空行")

    return {
        "file_name": filename,
        "sheet_count": source_sheet_count,
        "sheets": sheets,
    }


def _get_excel_import_column_notes(document_json: dict[str, Any], columns: list[str]) -> dict[str, str]:
    column_configs = document_json.get("column_configs")
    if not isinstance(column_configs, dict):
        return {}
    notes: dict[str, str] = {}
    for header in columns:
        config = column_configs.get(header)
        if isinstance(config, dict):
            note = _normalize_sheet_text(config.get("note"))
            if note:
                notes[header] = note
    return notes


def _build_note_sheet_excel_import_prompt(
    *,
    document_json: dict[str, Any],
    workbook_payload: dict[str, Any],
    instruction: str,
    action_document_row: int | None = None,
    action_column: int | None = None,
) -> str:
    normalized = _normalize_document_json(document_json)
    columns = _normalize_document_columns(normalized)
    data_start_row = _normalize_document_data_start_row(normalized)
    grid_rows = _extract_document_grid_rows(normalized)
    preserved_rows = _get_excel_import_preserved_data_rows(
        normalized,
        action_document_row=action_document_row,
        action_column=action_column,
    )
    target_context = {
        "columns": columns,
        "column_notes": _get_excel_import_column_notes(normalized, columns),
        "header_rows": grid_rows[:data_start_row],
        "preserved_leading_data_rows": preserved_rows,
        "imported_rows_should_start_after_preserved_count": len(preserved_rows),
        "trigger_action": {
            "type": NOTE_SHEET_CELL_ACTION_EXCEL_IMPORT_RESET,
            "document_row": action_document_row,
            "column": action_column,
        },
    }
    user_instruction = instruction.strip() or "无补充说明。"
    return "\n\n".join([
        "请把 Excel 工作簿标准化为目标 sheet 的数据行。",
        "注意：Excel 文件名只作为来源信息，可能不准确；课程、日期和字段判断必须以用户补充说明、目标 sheet 结构、工作表表头和数据内容为准。",
        "目标 sheet 结构：",
        json.dumps(target_context, ensure_ascii=False, indent=2),
        "用户补充说明：",
        user_instruction,
        "Excel 工作簿抽取结果：",
        json.dumps(workbook_payload, ensure_ascii=False, indent=2),
    ])


def _parse_note_sheet_excel_import_json(content: str) -> dict[str, Any]:
    text = content.strip()
    try:
        payload = json.loads(text)
    except json.JSONDecodeError:
        start = text.find("{")
        end = text.rfind("}")
        if start < 0 or end <= start:
            raise HTTPException(status_code=502, detail="DeepSeek 未返回 JSON 对象")
        try:
            payload = json.loads(text[start:end + 1])
        except json.JSONDecodeError as exc:
            raise HTTPException(status_code=502, detail=f"DeepSeek 返回的 JSON 无法解析：{exc}") from exc
    if not isinstance(payload, dict):
        raise HTTPException(status_code=502, detail="DeepSeek 返回 JSON 必须是对象")
    return payload


def _normalize_import_record_key(value: Any) -> str:
    return re.sub(r"\s+", "", str(value or "")).lower()


def _normalize_excel_import_extra_column_header(value: Any) -> str:
    header = _normalize_sheet_text(value)
    header = re.sub(r"[\r\n\t]+", " ", header)
    header = re.sub(r"\s{2,}", " ", header).strip()
    return header[:60].strip()


def _append_excel_import_extra_column_header(
    headers: list[str],
    used_keys: set[str],
    value: Any,
) -> None:
    header = _normalize_excel_import_extra_column_header(value)
    if not header:
        return
    key = _normalize_import_record_key(header)
    if not key or key in used_keys:
        return
    headers.append(header)
    used_keys.add(key)


def _extract_note_sheet_excel_import_extra_columns(
    payload: dict[str, Any],
    columns: list[str],
    raw_rows: list[Any],
) -> list[str]:
    extra_headers: list[str] = []
    used_keys = {_normalize_import_record_key(column) for column in columns}

    raw_extra_columns = payload.get("extra_columns")
    if isinstance(raw_extra_columns, list):
        for raw_header in raw_extra_columns:
            if isinstance(raw_header, dict):
                raw_header = (
                    raw_header.get("header")
                    or raw_header.get("name")
                    or raw_header.get("label")
                    or raw_header.get("field")
                )
            _append_excel_import_extra_column_header(extra_headers, used_keys, raw_header)

    target_keys = {_normalize_import_record_key(column) for column in columns}
    for raw_row in raw_rows:
        if not isinstance(raw_row, dict):
            continue
        for raw_key, raw_value in raw_row.items():
            key = _normalize_import_record_key(raw_key)
            if key in target_keys or not _normalize_sheet_text(raw_value):
                continue
            _append_excel_import_extra_column_header(extra_headers, used_keys, raw_key)

    return extra_headers


def _coerce_note_sheet_excel_import_rows(payload: dict[str, Any], columns: list[str]) -> tuple[list[list[Any]], list[str]]:
    raw_rows = payload.get("rows")
    if not isinstance(raw_rows, list):
        raise HTTPException(status_code=502, detail="DeepSeek 返回 JSON 缺少 rows 数组")

    extra_columns = _extract_note_sheet_excel_import_extra_columns(payload, columns, raw_rows)
    output_columns = [*columns, *extra_columns]
    normalized_rows: list[list[Any]] = []
    for raw_row in raw_rows:
        if isinstance(raw_row, dict):
            exact_row = {_normalize_sheet_text(key): value for key, value in raw_row.items()}
            normalized_key_row = {_normalize_import_record_key(key): value for key, value in raw_row.items()}
            row = [
                _normalize_excel_import_cell_for_column(
                    column,
                    exact_row.get(column, normalized_key_row.get(_normalize_import_record_key(column), "")),
                )
                for column in output_columns
            ]
        elif isinstance(raw_row, list):
            row = [
                _normalize_excel_import_cell_for_column(column, value)
                for column, value in zip(output_columns, _normalize_sheet_row(raw_row, len(output_columns)))
            ]
        else:
            continue
        if any(_normalize_sheet_text(cell) for cell in row):
            normalized_rows.append(row)

    if not normalized_rows:
        raise HTTPException(status_code=422, detail="DeepSeek 没有识别到可导入的数据行")
    return normalized_rows, extra_columns


def _normalize_excel_import_header_key(value: Any) -> str:
    header = _normalize_excel_import_extra_column_header(value)
    header = re.sub(r"（\s*(?:必填|自动)\s*）|\(\s*(?:必填|自动)\s*\)", "", header)
    return _normalize_import_record_key(header)


REGISTRATION_GROUP_SEQUENCE_SOURCE_ALIASES: dict[str, tuple[str, ...]] = {
    "group": ("分组", "组别", "组号"),
    "sequence": ("序号", "编号", "学号"),
    "remark": ("备注", "说明"),
    "submitted_at": ("提交时间", "提交日期", "报名时间"),
    "name": ("姓名", "真实姓名", "真实姓名（必填）", "提交者", "提交者（自动）"),
    "nickname": ("微信昵称", "昵称", "微信名"),
    "phone": ("手机号", "手机", "手机号码", "联系电话"),
    "payment_order": ("微信支付订单号", "交易单号", "支付订单号"),
    "merchant_order": ("商户订单号", "商户单号", "订单号"),
    "amount": ("订单金额", "支付金额", "金额", "订单金额 "),
}
REGISTRATION_GROUP_SEQUENCE_SOURCE_KEYS = {
    field: {_normalize_excel_import_header_key(alias) for alias in aliases}
    for field, aliases in REGISTRATION_GROUP_SEQUENCE_SOURCE_ALIASES.items()
}


def _registration_source_header_field(value: Any) -> str | None:
    key = _normalize_excel_import_header_key(value)
    if not key:
        return None
    for field, keys in REGISTRATION_GROUP_SEQUENCE_SOURCE_KEYS.items():
        if key in keys:
            return field
    return None


def _registration_group_sequence_row_keys(row: dict[str, str]) -> list[tuple[str, str]]:
    keys: list[tuple[str, str]] = []
    phone = _strip_legacy_text_prefix(row.get("phone"))
    merchant_order = _strip_legacy_text_prefix(row.get("merchant_order"))
    payment_order = _strip_legacy_text_prefix(row.get("payment_order"))
    name = _normalize_sheet_text(row.get("name"))
    nickname = _normalize_sheet_text(row.get("nickname"))
    if phone:
        keys.append(("phone", phone))
    if merchant_order:
        keys.append(("merchant_order", merchant_order))
        keys.append(("order", merchant_order))
    if payment_order:
        keys.append(("payment_order", payment_order))
        keys.append(("order", payment_order))
    if name and nickname:
        keys.append(("name_nickname", f"{name}\n{nickname}"))
    if name:
        keys.append(("name", name))
    return keys


def _normalize_registration_source_order_fields(row: dict[str, str]) -> dict[str, str]:
    next_row = dict(row)
    payment_order = _strip_legacy_text_prefix(next_row.get("payment_order"))
    merchant_order = _strip_legacy_text_prefix(next_row.get("merchant_order"))
    if payment_order and not merchant_order and re.match(r"^MA\d{8,}", payment_order, re.I):
        next_row["merchant_order"] = payment_order
        next_row["payment_order"] = ""
    elif merchant_order and not payment_order and re.match(r"^42\d{16,}", merchant_order):
        next_row["payment_order"] = merchant_order
        next_row["merchant_order"] = ""
    return next_row


def _extract_registration_group_sequence_source_rows(workbook_payload: dict[str, Any]) -> list[dict[str, str]]:
    source_rows: list[dict[str, str]] = []
    sheets = workbook_payload.get("sheets")
    if not isinstance(sheets, list):
        return source_rows

    for sheet in sheets:
        sheet_rows = sheet.get("rows") if isinstance(sheet, dict) else None
        if not isinstance(sheet_rows, list) or not sheet_rows:
            continue

        best_header_index = -1
        best_fields: dict[int, str] = {}
        best_score = 0
        for index, raw_row in enumerate(sheet_rows[:12]):
            values = raw_row.get("values") if isinstance(raw_row, dict) else None
            if not isinstance(values, list):
                continue
            fields = {
                column_index: field
                for column_index, value in enumerate(values)
                if (field := _registration_source_header_field(value))
            }
            score = len(set(fields.values()))
            if "sequence" in fields.values():
                score += 3
            if {"phone", "merchant_order", "payment_order"} & set(fields.values()):
                score += 2
            if score > best_score:
                best_header_index = index
                best_fields = fields
                best_score = score

        if best_header_index < 0 or "sequence" not in best_fields.values() or best_score < 4:
            continue

        current_group = ""
        for raw_row in sheet_rows[best_header_index + 1:]:
            values = raw_row.get("values") if isinstance(raw_row, dict) else None
            if not isinstance(values, list):
                continue
            row = {
                field: _normalize_excel_import_cell(values[column_index] if column_index < len(values) else "")
                for column_index, field in best_fields.items()
            }
            row = _normalize_registration_source_order_fields(row)
            visible_values = [_normalize_sheet_text(value) for value in row.values()]
            if not any(visible_values):
                continue
            if any(
                _normalize_excel_import_header_key(value) in REGISTRATION_GROUP_SEQUENCE_SOURCE_KEYS["sequence"]
                for value in visible_values
            ):
                continue

            row_group_text = _normalize_sheet_text(row.get("group"))
            if _parse_registration_group_number(row_group_text):
                current_group = row_group_text

            grouped = _parse_registration_group_sequence(row.get("sequence"))
            if grouped is None:
                continue

            group_number, member_number, width = grouped
            row["sequence"] = _format_registration_group_sequence(group_number, member_number, width=width)
            row_group_number = _parse_registration_group_number(row_group_text)
            current_group_number = _parse_registration_group_number(current_group)
            if row_group_number and row_group_number != group_number:
                row["group"] = f"{group_number}组"
            elif not row_group_number and current_group_number and current_group_number != group_number:
                row["group"] = f"{group_number}组"
            elif not row_group_number:
                row["group"] = current_group or f"{group_number}组"
            if _registration_group_sequence_row_keys(row):
                source_rows.append(row)

    return source_rows


def _build_unique_registration_group_sequence_lookup(
    source_rows: list[dict[str, str]],
) -> dict[tuple[str, str], dict[str, str]]:
    grouped: dict[tuple[str, str], list[dict[str, str]]] = {}
    for row in source_rows:
        for key in _registration_group_sequence_row_keys(row):
            grouped.setdefault(key, []).append(row)

    lookup: dict[tuple[str, str], dict[str, str]] = {}
    for key, rows in grouped.items():
        sequence_values = {_normalize_sheet_text(row.get("sequence")) for row in rows}
        if len(sequence_values) == 1:
            lookup[key] = rows[0]
    return lookup


def _prefer_registration_group_sequences_from_workbook(
    *,
    workbook_payload: dict[str, Any],
    rows: list[list[Any]],
    columns: list[str],
) -> tuple[list[list[Any]], int, int]:
    if not rows or not _is_registration_append_sheet(columns):
        return rows, 0, 0

    sequence_index = _get_column_index(columns, NOTE_SHEET_REGISTRATION_SEQUENCE_COLUMN)
    group_index = _get_column_index(columns, NOTE_SHEET_REGISTRATION_GROUP_COLUMN)
    if sequence_index < 0:
        return rows, 0, 0

    source_rows = _extract_registration_group_sequence_source_rows(workbook_payload)
    if not source_rows:
        return rows, 0, 0
    lookup = _build_unique_registration_group_sequence_lookup(source_rows)
    if not lookup:
        return rows, 0, 0

    field_indexes = {
        "phone": _get_column_index(columns, "手机号"),
        "merchant_order": _get_column_index(columns, "商户订单号"),
        "payment_order": _get_column_index(columns, "微信支付订单号"),
        "amount": _get_column_index(columns, "订单金额"),
        "submitted_at": _get_column_index(columns, "提交时间"),
        "remark": _get_column_index(columns, "备注"),
        "name": _get_column_index(columns, "姓名"),
        "nickname": _get_column_index(columns, "微信昵称"),
    }

    def cell(row: list[Any], field: str) -> str:
        index = field_indexes.get(field, -1)
        return _normalize_sheet_text(row[index]) if index >= 0 and index < len(row) else ""

    changed_count = 0
    source_field_count = 0
    next_rows: list[list[Any]] = []
    for raw_row in rows:
        row = _normalize_sheet_row(raw_row, len(columns))
        candidate = {
            "phone": _strip_legacy_text_prefix(cell(row, "phone")),
            "merchant_order": _strip_legacy_text_prefix(cell(row, "merchant_order")),
            "payment_order": _strip_legacy_text_prefix(cell(row, "payment_order")),
            "name": cell(row, "name"),
            "nickname": cell(row, "nickname"),
        }
        matched_source: dict[str, str] | None = None
        for key in _registration_group_sequence_row_keys(candidate):
            matched_source = lookup.get(key)
            if matched_source is not None:
                break
        if matched_source is None:
            next_rows.append(row)
            continue

        next_sequence = _normalize_sheet_text(matched_source.get("sequence"))
        if next_sequence and _normalize_sheet_text(row[sequence_index]) != next_sequence:
            row[sequence_index] = next_sequence
            changed_count += 1
        if group_index >= 0 and not _normalize_sheet_text(row[group_index]):
            group_value = _normalize_sheet_text(matched_source.get("group"))
            if group_value:
                row[group_index] = group_value

        source_field_changed = False
        for field, target_column in (
            ("submitted_at", "提交时间"),
            ("remark", "备注"),
            ("amount", "订单金额"),
        ):
            index = field_indexes.get(field, -1)
            value = _normalize_sheet_text(matched_source.get(field))
            if index >= 0 and value and _normalize_sheet_text(row[index]) != value:
                row[index] = value
                source_field_changed = True

        merchant_index = field_indexes.get("merchant_order", -1)
        payment_index = field_indexes.get("payment_order", -1)
        source_merchant_order = _strip_legacy_text_prefix(matched_source.get("merchant_order"))
        source_payment_order = _strip_legacy_text_prefix(matched_source.get("payment_order"))
        if merchant_index >= 0 and source_merchant_order:
            if _strip_legacy_text_prefix(row[merchant_index]) != source_merchant_order:
                row[merchant_index] = source_merchant_order
                source_field_changed = True
            if (
                payment_index >= 0
                and not source_payment_order
                and (
                    _strip_legacy_text_prefix(row[payment_index]) == source_merchant_order
                    or re.match(r"^MA\d{8,}", _strip_legacy_text_prefix(row[payment_index]), re.I)
                )
            ):
                row[payment_index] = ""
                source_field_changed = True
        if (
            payment_index >= 0
            and source_payment_order
            and _strip_legacy_text_prefix(row[payment_index]) != source_payment_order
        ):
            row[payment_index] = source_payment_order
            source_field_changed = True
        if source_field_changed:
            source_field_count += 1
        next_rows.append(row)

    return next_rows, changed_count, source_field_count


def _normalize_import_message_list(value: Any) -> list[str]:
    if not isinstance(value, list):
        return []
    return [_normalize_sheet_text(item) for item in value if _normalize_sheet_text(item)]


def _run_note_sheet_excel_import_deepseek(
    *,
    document_json: dict[str, Any],
    workbook_payload: dict[str, Any],
    instruction: str,
    session: Session,
    current_user: User,
    action_document_row: int | None = None,
    action_column: int | None = None,
) -> tuple[list[list[Any]], list[str], list[str], list[str]]:
    columns = _normalize_document_columns(document_json)
    prompt = _build_note_sheet_excel_import_prompt(
        document_json=document_json,
        workbook_payload=workbook_payload,
        instruction=instruction,
        action_document_row=action_document_row,
        action_column=action_column,
    )
    runtime = resolve_ai_app_runtime_config(
        session=session,
        current_user=current_user,
        app_id=AI_APP_NOTE_SHEET_EXCEL_IMPORT,
    )
    attempts = max(1, int(os.environ.get("CODEYUN_NOTE_SHEET_EXCEL_IMPORT_MAX_ATTEMPTS") or NOTE_SHEET_EXCEL_IMPORT_MAX_ATTEMPTS))
    provider_id = str(runtime.get("provider") or NOTE_SHEET_EXCEL_IMPORT_PROVIDER_ID)
    model = os.environ.get("CODEYUN_NOTE_SHEET_EXCEL_IMPORT_MODEL") or runtime.get("model") or NOTE_SHEET_EXCEL_IMPORT_MODEL
    last_error_status = 502
    last_error_detail = ""
    last_error_attempt = 0

    for attempt in range(1, attempts + 1):
        retry_suffix = ""
        if attempt > 1 and last_error_detail:
            retry_suffix = "\n\n上一次 AI 导入失败，请严格返回合法 JSON 对象，并确保 rows 是可导入的数据行。上一次错误：" + last_error_detail
        try:
            response = chat_with_provider(
                provider_id=provider_id,
                base_url=runtime.get("base_url"),
                api_key=runtime.get("api_key"),
                model=model,
                messages=[{"role": "user", "content": f"{prompt}{retry_suffix}"}],
                system_prompt=NOTE_SHEET_EXCEL_IMPORT_SYSTEM_PROMPT,
                response_format="json",
                timeout_seconds=NOTE_SHEET_EXCEL_IMPORT_TIMEOUT_SECONDS,
                extra_providers=tuple(runtime.get("extra_providers") or ()),
            )
            payload = _parse_note_sheet_excel_import_json(str(response.get("content") or ""))
            import_rows, extra_columns = _coerce_note_sheet_excel_import_rows(payload, columns)
            return (
                import_rows,
                extra_columns,
                _normalize_import_message_list(payload.get("warnings")),
                _normalize_import_message_list(payload.get("mapping_notes")),
            )
        except OllamaClientError as exc:
            last_error_status = 502
            last_error_detail = f"模型调用异常：{_normalize_sheet_text(str(exc)) or exc.__class__.__name__}"
            last_error_attempt = attempt
        except HTTPException as exc:
            if exc.status_code not in {502, 422}:
                raise
            last_error_status = exc.status_code
            detail = _normalize_sheet_text(exc.detail)
            if exc.status_code == 422:
                last_error_detail = f"AI 返回数据无效：{detail}"
            else:
                last_error_detail = f"AI 返回解析错误：{detail}"
            last_error_attempt = attempt

        if attempt < attempts:
            time.sleep(NOTE_SHEET_EXCEL_IMPORT_RETRY_DELAY_SECONDS * attempt)

    if not last_error_detail:
        raise HTTPException(status_code=502, detail="AI 导入未完成：没有收到模型响应或解析结果")
    raise HTTPException(
        status_code=last_error_status,
        detail=(
            f"AI 导入未完成：已尝试 {attempts} 次，最后错误发生在第 {last_error_attempt or attempts} 次；"
            f"模型 {provider_id}/{model}；{last_error_detail}"
        ),
    )


def _find_column_index(columns: list[Any], header: str) -> int | None:
    for index, column in enumerate(columns):
        if _normalize_sheet_text(column) == header:
            return index
    return None


def _normalize_table_field_token(value: Any) -> str:
    return re.sub(r"\s+", "", _normalize_sheet_text(value)).lower()


def _find_table_column_index(columns: list[Any], field: str | int | None) -> int | None:
    if isinstance(field, int):
        return field if 0 <= field < len(columns) else None

    field_text = _normalize_sheet_text(field)
    if not field_text:
        return None

    exact_index = _find_column_index(columns, field_text)
    if exact_index is not None:
        return exact_index

    token = _normalize_table_field_token(field_text)
    token_matches = [
        index
        for index, column in enumerate(columns)
        if _normalize_table_field_token(column) == token
    ]
    if len(token_matches) == 1:
        return token_matches[0]

    lesson_match = re.search(r"第\s*0*(?P<number>\d+)\s*课", field_text)
    if lesson_match:
        lesson_number = int(lesson_match.group("number"))
        lesson_tokens = {
            f"第{lesson_number}课",
            f"第{lesson_number:02d}课",
        }
        lesson_matches = [
            index
            for index, column in enumerate(columns)
            if any(token in _normalize_sheet_text(column) for token in lesson_tokens)
        ]
        if len(lesson_matches) == 1:
            return lesson_matches[0]

    contains_matches = [
        index
        for index, column in enumerate(columns)
        if token and token in _normalize_table_field_token(column)
    ]
    if len(contains_matches) == 1:
        return contains_matches[0]

    return None


def _find_bound_column_index(columns: list[Any], header: str, fallback_index: int | None = None) -> int | None:
    header_index = _find_column_index(columns, header)
    if header_index is not None:
        return header_index
    if fallback_index is not None and 0 <= fallback_index < len(columns):
        return fallback_index
    return None


def _find_attendance_column_index(columns: list[Any], field_key: str) -> int | None:
    binding = ATTENDANCE_FIELD_BINDINGS.get(field_key)
    if binding is None:
        return None
    header, fallback_index = binding
    header_index = _find_column_index(columns, header)
    if header_index is not None:
        return header_index
    if field_key in {"lesson_links", "clockin_links"}:
        return None

    has_link_count_columns = (
        _find_column_index(columns, ATTENDANCE_FIELD_BINDINGS["lesson_links"][0]) is not None
        or _find_column_index(columns, ATTENDANCE_FIELD_BINDINGS["clockin_links"][0]) is not None
    )
    effective_fallback_index = fallback_index if has_link_count_columns else ATTENDANCE_FIELD_LEGACY_FALLBACKS.get(
        field_key,
        fallback_index,
    )
    if 0 <= effective_fallback_index < len(columns):
        return effective_fallback_index
    return None


def _shift_cell_meta_columns_for_insert(cell_meta: Any, insert_index: int, amount: int) -> dict[str, Any]:
    if not isinstance(cell_meta, dict) or amount <= 0:
        return dict(cell_meta) if isinstance(cell_meta, dict) else {}

    shifted: dict[str, Any] = {}
    for key, meta in cell_meta.items():
        parsed = _parse_cell_meta_key(key)
        if parsed is None:
            shifted[str(key)] = meta
            continue
        row_index, column_index = parsed
        next_column_index = column_index + amount if column_index >= insert_index else column_index
        shifted[f"{row_index}:{next_column_index}"] = meta
    return shifted


def _insert_columns_into_header_groups(header_groups: Any, insert_index: int, amount: int) -> list[Any]:
    if not isinstance(header_groups, list) or amount <= 0:
        return list(header_groups) if isinstance(header_groups, list) else []

    next_groups: list[Any] = []
    for row in header_groups:
        if not isinstance(row, list):
            next_groups.append(row)
            continue

        next_row: list[Any] = []
        current_index = 0
        expanded = False
        for cell in row:
            if isinstance(cell, dict):
                next_cell = dict(cell)
                colspan = int(next_cell.get("colspan") or 1)
            else:
                next_cell = cell
                colspan = 1

            colspan = max(colspan, 1)
            if not expanded and current_index <= insert_index < current_index + colspan:
                if isinstance(next_cell, dict):
                    next_cell["colspan"] = colspan + amount
                else:
                    next_cell = {"label": str(next_cell), "colspan": colspan + amount}
                expanded = True

            next_row.append(next_cell)
            current_index += colspan

        if not expanded and insert_index >= current_index:
            next_row.append({"label": "", "colspan": amount})
        next_groups.append(next_row)
    return next_groups


def _shift_merged_cell_columns_for_insert(merged_cells: Any, insert_index: int, amount: int) -> list[Any]:
    if not isinstance(merged_cells, list) or amount <= 0:
        return list(merged_cells) if isinstance(merged_cells, list) else []

    shifted: list[Any] = []
    for cell in merged_cells:
        if not isinstance(cell, dict):
            continue
        row = int(cell.get("row") or 0)
        col = int(cell.get("col") or 0)
        rowspan = max(int(cell.get("rowspan") or 1), 1)
        colspan = max(int(cell.get("colspan") or 1), 1)
        if col < insert_index < col + colspan:
            colspan += amount
        elif col >= insert_index:
            col += amount
        if rowspan > 1 or colspan > 1:
            shifted.append({"row": row, "col": col, "rowspan": rowspan, "colspan": colspan})
    return shifted


def _insert_cell_into_grid_row(row: Any, column_count: int, insert_index: int) -> list[Any]:
    normalized_row = _normalize_sheet_row(row, column_count)
    return [
        *normalized_row[:insert_index],
        "",
        *normalized_row[insert_index:],
    ]


def _delete_cell_from_grid_row(row: Any, column_count: int, delete_index: int) -> list[Any]:
    normalized_row = _normalize_sheet_row(row, column_count)
    return [
        *normalized_row[:delete_index],
        *normalized_row[delete_index + 1:],
    ]


def _insert_document_column(
    document_json: dict[str, Any],
    *,
    insert_index: int,
    header: str,
    width: int = 96,
) -> dict[str, Any]:
    normalized = _normalize_document_json(document_json)
    columns = _normalize_document_columns(normalized)
    bounded_insert_index = min(max(insert_index, 0), len(columns))
    column_index_map = {
        index: index + 1 if index >= bounded_insert_index else index
        for index in range(len(columns))
    }

    remapped_rows = [
        _remap_row_formula_cell_references(row, columns=columns, column_index_map=column_index_map)
        for row in _extract_document_rows(normalized)
    ]
    next_rows: list[Any] = []
    for row in remapped_rows:
        if isinstance(row, list):
            normalized_row = _normalize_sheet_row(row, len(columns))
            next_rows.append([
                *normalized_row[:bounded_insert_index],
                "",
                *normalized_row[bounded_insert_index:],
            ])
        elif isinstance(row, dict):
            next_row = dict(row)
            next_row.setdefault(header, "")
            next_rows.append(next_row)
        else:
            next_rows.append([""] * bounded_insert_index + [""] + [""] * (len(columns) - bounded_insert_index))

    source_widths = normalized.get("column_widths")
    if isinstance(source_widths, list):
        next_widths = [*source_widths[:bounded_insert_index], width, *source_widths[bounded_insert_index:]]
    else:
        next_widths = [width] * (len(columns) + 1)

    next_document = {
        **normalized,
        "columns": [*columns[:bounded_insert_index], header, *columns[bounded_insert_index:]],
        "rows": next_rows,
        "column_widths": next_widths,
    }
    grid_rows = _extract_document_grid_rows(normalized)
    if grid_rows:
        next_grid_rows = [
            _insert_cell_into_grid_row(row, len(columns), bounded_insert_index)
            for row in grid_rows
        ]
        field_row_index = int(normalized.get("field_row_index") or 0)
        if 0 <= field_row_index < len(next_grid_rows):
            next_grid_rows[field_row_index][bounded_insert_index] = header
        next_document["grid_rows"] = next_grid_rows
    if "cell_meta" in normalized:
        next_document["cell_meta"] = _shift_cell_meta_columns_for_insert(
            normalized.get("cell_meta"),
            bounded_insert_index,
            1,
        )
    if "merged_cells" in normalized:
        next_document["merged_cells"] = _shift_merged_cell_columns_for_insert(
            normalized.get("merged_cells"),
            bounded_insert_index,
            1,
        )
    if "header_groups" in normalized:
        next_document["header_groups"] = _insert_columns_into_header_groups(
            normalized.get("header_groups"),
            bounded_insert_index,
            1,
        )
    return next_document


def _shift_cell_meta_columns_for_delete(cell_meta: Any, delete_index: int) -> dict[str, Any]:
    if not isinstance(cell_meta, dict):
        return {}

    shifted: dict[str, Any] = {}
    for key, meta in cell_meta.items():
        parsed = _parse_cell_meta_key(key)
        if parsed is None:
            shifted[str(key)] = meta
            continue
        row_index, column_index = parsed
        if column_index == delete_index:
            continue
        next_column_index = column_index - 1 if column_index > delete_index else column_index
        shifted[f"{row_index}:{next_column_index}"] = meta
    return shifted


def _shift_merged_cell_columns_for_delete(merged_cells: Any, delete_index: int) -> list[Any]:
    if not isinstance(merged_cells, list):
        return []

    shifted: list[Any] = []
    for cell in merged_cells:
        if not isinstance(cell, dict):
            continue
        row = int(cell.get("row") or 0)
        col = int(cell.get("col") or 0)
        rowspan = max(int(cell.get("rowspan") or 1), 1)
        colspan = max(int(cell.get("colspan") or 1), 1)
        if col <= delete_index < col + colspan:
            colspan -= 1
        elif col > delete_index:
            col -= 1
        if rowspan > 1 or colspan > 1:
            shifted.append({"row": row, "col": col, "rowspan": rowspan, "colspan": colspan})
    return shifted


def _delete_columns_from_header_groups(header_groups: Any, delete_index: int) -> list[Any]:
    if not isinstance(header_groups, list):
        return []

    next_groups: list[Any] = []
    for row in header_groups:
        if not isinstance(row, list):
            next_groups.append(row)
            continue

        next_row: list[Any] = []
        current_index = 0
        for cell in row:
            if isinstance(cell, dict):
                next_cell = dict(cell)
                colspan = int(next_cell.get("colspan") or 1)
            else:
                next_cell = cell
                colspan = 1
            colspan = max(colspan, 1)
            cell_end = current_index + colspan
            if current_index <= delete_index < cell_end:
                if colspan > 1:
                    if isinstance(next_cell, dict):
                        next_cell["colspan"] = colspan - 1
                    else:
                        next_cell = {"label": str(next_cell), "colspan": colspan - 1}
                    next_row.append(next_cell)
            else:
                next_row.append(next_cell)
            current_index = cell_end
        next_groups.append(next_row)
    return next_groups


def _delete_document_column(
    document_json: dict[str, Any],
    *,
    delete_index: int,
) -> dict[str, Any]:
    normalized = _normalize_document_json(document_json)
    columns = _normalize_document_columns(normalized)
    if delete_index < 0 or delete_index >= len(columns):
        return normalized
    deleted_header = columns[delete_index]
    column_index_map = {
        index: (None if index == delete_index else index - 1 if index > delete_index else index)
        for index in range(len(columns))
    }

    remapped_rows = [
        _remap_row_formula_cell_references(row, columns=columns, column_index_map=column_index_map)
        for row in _extract_document_rows(normalized)
    ]
    next_rows: list[Any] = []
    for row in remapped_rows:
        if isinstance(row, list):
            normalized_row = _normalize_sheet_row(row, len(columns))
            next_rows.append([*normalized_row[:delete_index], *normalized_row[delete_index + 1:]])
        elif isinstance(row, dict):
            next_row = dict(row)
            next_row.pop(str(deleted_header), None)
            next_rows.append(next_row)
        else:
            next_rows.append([""] * (len(columns) - 1))

    source_widths = normalized.get("column_widths")
    if isinstance(source_widths, list):
        next_widths = [*source_widths[:delete_index], *source_widths[delete_index + 1:]]
    else:
        next_widths = [96] * (len(columns) - 1)

    next_document = {
        **normalized,
        "columns": [*columns[:delete_index], *columns[delete_index + 1:]],
        "rows": next_rows,
        "column_widths": next_widths,
    }
    grid_rows = _extract_document_grid_rows(normalized)
    if grid_rows:
        next_grid_rows: list[Any] = []
        for row in grid_rows:
            remapped_row = _remap_row_formula_cell_references(row, columns=columns, column_index_map=column_index_map)
            next_grid_rows.append(_delete_cell_from_grid_row(remapped_row, len(columns), delete_index))
        next_document["grid_rows"] = next_grid_rows
    if "cell_meta" in normalized:
        next_document["cell_meta"] = _shift_cell_meta_columns_for_delete(
            normalized.get("cell_meta"),
            delete_index,
        )
    if "merged_cells" in normalized:
        next_document["merged_cells"] = _shift_merged_cell_columns_for_delete(
            normalized.get("merged_cells"),
            delete_index,
        )
    if "header_groups" in normalized:
        next_document["header_groups"] = _delete_columns_from_header_groups(
            normalized.get("header_groups"),
            delete_index,
        )
    if isinstance(normalized.get("column_configs"), dict):
        configs = dict(normalized.get("column_configs") or {})
        configs.pop(str(deleted_header), None)
        next_document["column_configs"] = configs
    elif isinstance(normalized.get("column_configs"), list):
        configs = list(normalized.get("column_configs") or [])
        next_document["column_configs"] = [*configs[:delete_index], *configs[delete_index + 1:]]
    entity_columns = normalized.get("entity_columns")
    entity_cells = normalized.get("entity_cells")
    if isinstance(entity_columns, list):
        deleted_column = entity_columns[delete_index] if delete_index < len(entity_columns) else None
        deleted_column_id = _get_document_entity_column_id(deleted_column)
        next_document["entity_columns"] = [*entity_columns[:delete_index], *entity_columns[delete_index + 1:]]
        if isinstance(entity_cells, dict) and deleted_column_id:
            next_entity_cells: dict[str, Any] = {}
            for row_id, row_cells in entity_cells.items():
                if not isinstance(row_cells, dict):
                    next_entity_cells[row_id] = row_cells
                    continue
                next_row_cells = dict(row_cells)
                next_row_cells.pop(deleted_column_id, None)
                if next_row_cells:
                    next_entity_cells[row_id] = next_row_cells
            next_document["entity_cells"] = next_entity_cells
    return next_document


def _get_attendance_link_count_field_insert_index(columns: list[Any], field_key: str) -> int:
    if field_key == "lesson_links":
        clockin_index = _find_column_index(columns, ATTENDANCE_FIELD_BINDINGS["clockin_links"][0])
        if clockin_index is not None:
            return clockin_index
        owner_index = _find_attendance_column_index(columns, "course_owner")
        return min((owner_index + 1) if owner_index is not None else 4, len(columns))

    lesson_index = _find_column_index(columns, ATTENDANCE_FIELD_BINDINGS["lesson_links"][0])
    if lesson_index is not None:
        return min(lesson_index + 1, len(columns))
    owner_index = _find_attendance_column_index(columns, "course_owner")
    return min((owner_index + 1) if owner_index is not None else 4, len(columns))


def _ensure_attendance_link_count_columns(document_json: dict[str, Any]) -> dict[str, Any]:
    next_document = _normalize_document_json(document_json)
    for field_key in ("lesson_links", "clockin_links"):
        columns = _normalize_document_columns(next_document)
        header, _fallback_index = ATTENDANCE_FIELD_BINDINGS[field_key]
        if _find_column_index(columns, header) is not None:
            continue
        next_document = _insert_document_column(
            next_document,
            insert_index=_get_attendance_link_count_field_insert_index(columns, field_key),
            header=header,
            width=96,
        )
    return next_document


def _parse_attendance_template_course(value: Any) -> dict[str, Any] | None:
    match = ATTENDANCE_TEMPLATE_COURSE_TEXT_RE.search(_normalize_sheet_text(value))
    if not match:
        return None

    raw_date = match.group("date")
    course_date: date | None = None
    if raw_date:
        if len(raw_date) == 6:
            year = _normalize_two_digit_year(int(raw_date[:2]))
            month = int(raw_date[2:4])
            day = int(raw_date[4:6])
        else:
            year = int(raw_date[:4])
            month = int(raw_date[4:6])
            day = int(raw_date[6:8])

        try:
            course_date = date(year, month, day)
        except ValueError:
            return None

    return {
        "date": course_date,
        "edition": int(match.group("edition")),
        "course": match.group("course"),
    }


def _get_next_month_first_day(today: date | None = None) -> date:
    current = today or date.today()
    year = current.year + (1 if current.month == 12 else 0)
    month = 1 if current.month == 12 else current.month + 1
    return date(year, month, 1)


def _get_next_sunday(today: date | None = None) -> date:
    current = today or date.today()
    days_until_sunday = (6 - current.weekday()) % 7
    if days_until_sunday == 0:
        days_until_sunday = 7
    return current + timedelta(days=days_until_sunday)


def _is_attendance_fanbei_course_type(course_type: str) -> bool:
    return _normalize_sheet_text(course_type) in ATTENDANCE_TEMPLATE_FANBEI_SOURCE_COURSES


def _get_attendance_course_month_start_date(course_type: str, month_date: date) -> date:
    if _is_attendance_fanbei_course_type(course_type):
        return date(month_date.year, month_date.month, ATTENDANCE_TEMPLATE_FANBEI_START_DAY)
    return month_date


def _add_months(value: date, month_delta: int) -> date:
    month_index = value.year * 12 + (value.month - 1) + month_delta
    year = month_index // 12
    month = month_index % 12 + 1
    return date(year, month, value.day)


def _get_next_attendance_fanbei_cycle_date(source_date: date) -> date:
    next_cycle_month = _add_months(source_date.replace(day=ATTENDANCE_TEMPLATE_FANBEI_START_DAY), 2)
    return date(next_cycle_month.year, next_cycle_month.month, ATTENDANCE_TEMPLATE_FANBEI_START_DAY)


def _parse_attendance_date_text(value: Any) -> date | None:
    text = _normalize_sheet_text(value)
    if not text:
        return None

    try:
        return date.fromisoformat(text)
    except ValueError:
        pass

    separated = re.fullmatch(r"(\d{4})[-/年](\d{1,2})[-/月](\d{1,2})日?", text)
    if separated:
        try:
            return date(int(separated.group(1)), int(separated.group(2)), int(separated.group(3)))
        except ValueError:
            return None

    compact = re.fullmatch(r"(\d{6}|\d{8})", text)
    if compact:
        digits = compact.group(1)
        try:
            if len(digits) == 6:
                return date(_normalize_two_digit_year(int(digits[:2])), int(digits[2:4]), int(digits[4:6]))
            return date(int(digits[:4]), int(digits[4:6]), int(digits[6:8]))
        except ValueError:
            return None
    return None


def _get_attendance_template_target_date(payload: NoteSheetAttendanceTemplateGenerationRequest | None = None) -> date:
    if payload and payload.target_date is not None:
        if payload.target_year is not None or payload.target_month is not None:
            raise HTTPException(status_code=400, detail="target_date 不能和 target_year/target_month 同时提供")
        parsed_date = _parse_attendance_date_text(payload.target_date)
        if parsed_date is None:
            raise HTTPException(status_code=400, detail="target_date 格式不正确")
        return parsed_date
    if payload and payload.target_year is not None and payload.target_month is not None:
        return date(int(payload.target_year), int(payload.target_month), 1)
    if payload and (payload.target_year is not None or payload.target_month is not None):
        raise HTTPException(status_code=400, detail="target_year 和 target_month 必须同时提供")
    return _get_next_month_first_day()


def _is_attendance_zen_course_type(course_type: str) -> bool:
    return "禅宗" in course_type


def _has_attendance_template_target_payload(
    payload: NoteSheetAttendanceTemplateGenerationRequest | None,
) -> bool:
    return bool(
        payload
        and (
            payload.target_date is not None
            or payload.target_year is not None
            or payload.target_month is not None
        )
    )


def _get_attendance_course_template_payload_target_date(
    course_type: str,
    payload: NoteSheetAttendanceCourseTemplateGenerationRequest | None = None,
) -> date:
    target_date = _get_attendance_template_target_date(payload)
    return _get_attendance_course_month_start_date(course_type, target_date)


def _get_attendance_course_template_reference_start_date(
    document_json: dict[str, Any],
    payload: NoteSheetAttendanceCourseTemplateGenerationRequest | None,
    *,
    course_type: str,
) -> date | None:
    normalized = _normalize_document_json(document_json)
    columns = _normalize_document_columns(normalized)
    rows = _extract_document_rows(normalized)
    start_date_index = _find_attendance_column_index(columns, "start_date")
    formula_row_offset = _get_formula_reference_row_offset(normalized)
    formula_grid_rows = _extract_document_grid_rows(normalized)

    if payload and payload.row_index is not None:
        if payload.row_index >= len(rows):
            return None
        return _extract_attendance_row_start_date(
            _normalize_sheet_row(rows[payload.row_index], len(columns)),
            row_index=payload.row_index,
            columns=columns,
            rows=rows,
            start_date_index=start_date_index,
            reference_row_offset=formula_row_offset,
            grid_rows=formula_grid_rows,
        )

    source = _find_attendance_template_source_row(
        rows,
        columns=columns,
        course_type=course_type,
        target_date=date.max,
        reference_row_offset=formula_row_offset,
        grid_rows=formula_grid_rows,
    )
    return source[2]["date"] if source is not None else None


def _get_attendance_course_template_target_date(
    course_type: str,
    payload: NoteSheetAttendanceCourseTemplateGenerationRequest | None = None,
    document_json: dict[str, Any] | None = None,
) -> date:
    if _has_attendance_template_target_payload(payload):
        return _get_attendance_course_template_payload_target_date(course_type, payload)
    if _is_attendance_fanbei_course_type(course_type):
        source_start_date = (
            _get_attendance_course_template_reference_start_date(
                document_json,
                payload,
                course_type=course_type,
            )
            if document_json is not None
            else None
        )
        if source_start_date is not None:
            return _get_next_attendance_fanbei_cycle_date(source_start_date)
        return _get_attendance_course_month_start_date(course_type, _get_next_month_first_day())
    return _get_next_sunday() if _is_attendance_zen_course_type(course_type) else _get_next_month_first_day()


def _get_attendance_batch_course_targets(target_date: date) -> tuple[tuple[str, date], ...]:
    targets = [
        (course_type, target_date)
        for course_type in ATTENDANCE_TEMPLATE_MONTHLY_SOURCE_COURSES
    ]
    fanbei_target_date = _get_attendance_course_month_start_date("梵呗初阶", target_date)
    if target_date.month % 2 == 1:
        targets.extend(
            (course_type, fanbei_target_date)
            for course_type in ATTENDANCE_TEMPLATE_ODD_MONTH_SOURCE_COURSES
        )
    else:
        targets.extend(
            (course_type, fanbei_target_date)
            for course_type in ATTENDANCE_TEMPLATE_EVEN_MONTH_SOURCE_COURSES
        )
    return tuple(targets)


ATTENDANCE_TEMPLATE_SKIP_MONTHS_SETTING_KEY = "attendance_summary.template_skip_months"


def _normalize_attendance_template_skip_course_types(values: Any) -> set[str]:
    if not isinstance(values, list):
        return set()
    return {
        _normalize_sheet_text(value)
        for value in values
        if _normalize_sheet_text(value)
    }


def _read_attendance_template_skip_course_types(session: Session, target_date: date) -> set[str]:
    row = session.get(AppSetting, ATTENDANCE_TEMPLATE_SKIP_MONTHS_SETTING_KEY)
    if row is None or not isinstance(row.value, dict):
        return set()
    month_key = f"{target_date.year:04d}-{target_date.month:02d}"
    return _normalize_attendance_template_skip_course_types(row.value.get(month_key))


def _format_attendance_course_date(value: date) -> str:
    return f"{value.year:04d}{value.month:02d}{value.day:02d}"


def _format_attendance_date_serial(value: date) -> str:
    serial = _date_parts_to_serial(value.year, value.month, value.day)
    if serial is None:
        return ""
    return str(int(serial))


def _coerce_attendance_date(value: Any) -> date | None:
    text = _normalize_sheet_text(value)
    if re.fullmatch(r"\d{6}|\d{8}", text):
        parsed_text_date = _parse_attendance_date_text(text)
        if parsed_text_date is not None:
            return parsed_text_date

    sort_value = _parse_date_sort_value(value)
    if sort_value is not None:
        parts = _serial_to_date_parts(sort_value)
        if parts is not None:
            return date(*parts)

    compact_serial = _parse_compact_date_serial(value)
    if compact_serial is not None:
        parts = _serial_to_date_parts(compact_serial)
        if parts is not None:
            return date(*parts)
    return None


def _extract_attendance_row_start_date(
    row: list[Any],
    *,
    row_index: int,
    columns: list[Any],
    rows: list[Any],
    start_date_index: int | None,
    reference_row_offset: int = 0,
    grid_rows: list[Any] | None = None,
) -> date | None:
    if start_date_index is None or start_date_index >= len(row):
        return None
    raw_value = row[start_date_index]
    evaluated_value = _evaluate_formula_sort_value(
        raw_value,
        row_index=row_index,
        column_index=start_date_index,
        rows=rows,
        columns=columns,
        reference_row_offset=reference_row_offset,
        grid_rows=grid_rows,
    )
    return _coerce_attendance_date(evaluated_value)


def _find_attendance_template_source_row(
    rows: list[Any],
    *,
    columns: list[Any],
    course_type: str,
    target_date: date,
    reference_row_offset: int = 0,
    grid_rows: list[Any] | None = None,
) -> tuple[int, list[Any], dict[str, Any]] | None:
    # Monthly course workbooks evolve in real use. The next workbook must inherit
    # the nearest previous workbook of the same course type, not a fixed old base.
    column_count = len(columns)
    type_index = _find_attendance_column_index(columns, "course_type")
    name_index = _find_attendance_column_index(columns, "course_name")
    online_sheet_index = _find_attendance_column_index(columns, "online_sheet")
    start_date_index = _find_attendance_column_index(columns, "start_date")
    candidates: list[tuple[date, int, list[Any], dict[str, Any]]] = []

    for row_index, source_row in enumerate(rows):
        row = _normalize_sheet_row(source_row, column_count)
        if type_index is None or _normalize_sheet_text(row[type_index]) != course_type:
            continue

        source_date = _extract_attendance_row_start_date(
            row,
            row_index=row_index,
            columns=columns,
            rows=rows,
            start_date_index=start_date_index,
            reference_row_offset=reference_row_offset,
            grid_rows=grid_rows,
        )
        if source_date is None:
            continue
        if source_date >= target_date:
            continue

        name_value = row[name_index] if name_index is not None and name_index < len(row) else ""
        online_sheet_value = (
            row[online_sheet_index]
            if online_sheet_index is not None and online_sheet_index < len(row)
            else ""
        )
        info = {
            "date": source_date,
            "course_name": _normalize_sheet_text(name_value),
            "online_sheet": _normalize_sheet_text(online_sheet_value),
        }
        candidates.append((source_date, row_index, row, info))

    if not candidates:
        return None

    _source_date, row_index, row, info = max(
        candidates,
        key=lambda item: (item[0], -item[1]),
    )
    return row_index, row, info


def _attendance_template_row_exists(
    rows: list[Any],
    *,
    columns: list[Any],
    course_type: str,
    target_date: date,
    reference_row_offset: int = 0,
    grid_rows: list[Any] | None = None,
) -> bool:
    type_index = _find_attendance_column_index(columns, "course_type")
    start_date_index = _find_attendance_column_index(columns, "start_date")
    column_count = len(columns)

    for row_index, source_row in enumerate(rows):
        row = _normalize_sheet_row(source_row, column_count)
        type_value = _normalize_sheet_text(row[type_index]) if type_index is not None else ""
        if type_value != course_type:
            continue
        start_date = _extract_attendance_row_start_date(
            row,
            row_index=row_index,
            columns=columns,
            rows=rows,
            start_date_index=start_date_index,
            reference_row_offset=reference_row_offset,
            grid_rows=grid_rows,
        )
        if start_date == target_date:
            return True
    return False


def _set_row_cell_value(row: Any, columns: list[Any], column_index: int, value: Any) -> Any:
    if isinstance(row, list):
        next_row = [*row, *([""] * max(column_index + 1 - len(row), 0))]
        current = next_row[column_index]
        if isinstance(current, dict) and "link" in current and not isinstance(value, dict):
            next_cell = dict(current)
            next_cell["value"] = value
            next_row[column_index] = next_cell
        else:
            next_row[column_index] = value
        return next_row
    if isinstance(row, dict):
        next_row = dict(row)
        column_key = str(columns[column_index]) if column_index < len(columns) else str(column_index)
        current = next_row.get(column_key)
        if isinstance(current, dict) and "link" in current and not isinstance(value, dict):
            next_cell = dict(current)
            next_cell["value"] = value
            next_row[column_key] = next_cell
        else:
            next_row[column_key] = value
        return next_row
    next_row = [""] * max(len(columns), column_index + 1)
    next_row[column_index] = value
    return next_row


def _attendance_row_has_completion(row: Any, columns: list[Any], completed_index: int) -> bool:
    return _normalize_sheet_text(_extract_row_cell_value(row, completed_index, columns)) != ""


def _set_attendance_summary_row_completed(
    document_json: dict[str, Any],
    *,
    row_index: int,
    completion_date: date,
) -> tuple[dict[str, Any], int]:
    normalized = _normalize_document_json(document_json)
    columns = _normalize_document_columns(normalized)
    rows = _extract_document_rows(normalized)
    completed_index = _find_attendance_column_index(columns, "completed_date")
    if completed_index is None:
        raise HTTPException(status_code=400, detail="当前表缺少考勤实际完成结点字段")
    if row_index < 0 or row_index >= len(rows):
        raise HTTPException(status_code=400, detail="row_index 超出表格范围")

    completed_row = _set_row_cell_value(
        rows[row_index],
        columns,
        completed_index,
        _format_attendance_date_serial(completion_date),
    )
    remaining_rows = [
        (source_index, row)
        for source_index, row in enumerate(rows)
        if source_index != row_index
    ]
    insert_index = next(
        (
            index
            for index, (_source_index, row) in enumerate(remaining_rows)
            if _attendance_row_has_completion(row, columns, completed_index)
        ),
        len(remaining_rows),
    )
    ordered_rows = [
        *remaining_rows[:insert_index],
        (row_index, completed_row),
        *remaining_rows[insert_index:],
    ]
    row_index_map = {
        source_index: target_index
        for target_index, (source_index, _row) in enumerate(ordered_rows)
    }
    formula_row_offset = _get_formula_reference_row_offset(normalized)
    next_rows = [
        _remap_row_formula_cell_references(
            row,
            columns=columns,
            row_index_map=row_index_map,
            row_index_offset=formula_row_offset,
        )
        for _source_index, row in ordered_rows
    ]
    next_document = _replace_document_data_rows({
        **normalized,
        "columns": columns,
    }, next_rows)
    if isinstance(normalized.get("cell_meta"), dict):
        next_document["cell_meta"] = _remap_cell_meta_rows(
            normalized.get("cell_meta"),
            row_index_map,
            row_offset=_normalize_document_data_start_row(normalized),
        )
    return next_document, row_index_map[row_index]


def _shift_cell_meta_rows_for_insert(cell_meta: Any, insert_index: int, amount: int, *, row_offset: int = 0) -> dict[str, Any]:
    if not isinstance(cell_meta, dict) or amount <= 0:
        return dict(cell_meta) if isinstance(cell_meta, dict) else {}

    shifted: dict[str, Any] = {}
    for key, meta in cell_meta.items():
        parsed = _parse_cell_meta_key(key)
        if parsed is None:
            continue
        row_index, column_index = parsed
        effective_insert_index = insert_index + row_offset
        next_row_index = row_index + amount if row_index >= effective_insert_index else row_index
        shifted[f"{next_row_index}:{column_index}"] = meta
    return shifted


def _increment_attendance_template_edition(text: str, course_type: str) -> str:
    def replace(match: re.Match[str]) -> str:
        matched_course = match.group("course")
        if matched_course != course_type:
            return match.group(0)
        return f"第{int(match.group('edition')) + 1}届{matched_course}"

    return ATTENDANCE_TEMPLATE_COURSE_TEXT_RE.sub(replace, text, count=1)


def _should_strip_attendance_template_date_prefix(course_type: str) -> bool:
    return _normalize_sheet_text(course_type) in {"念住", "觉观"}


def _derive_attendance_template_text(value: Any, *, course_type: str, target_date: date) -> str:
    text = _normalize_sheet_text(value)
    if not text:
        return ""

    date_prefix = ""
    body = text
    leading_date = ATTENDANCE_TEMPLATE_LEADING_DATE_RE.match(text)
    if leading_date:
        date_prefix = "" if _should_strip_attendance_template_date_prefix(course_type) else _format_attendance_course_date(target_date)
        body = leading_date.group("body")

    body = _increment_attendance_template_edition(body, course_type)
    return f"{date_prefix}{body}"


def _build_inserted_attendance_template_row(
    source_row: list[Any],
    *,
    columns: list[Any],
    source_row_index: int,
    target_row_index: int,
    target_date: date,
    course_type: str,
    source_start_date: date | None = None,
) -> list[Any]:
    column_count = len(columns)
    row_delta = target_row_index - source_row_index
    next_row = [
        _shift_formula_value_references(value, row_delta=row_delta)
        for value in _normalize_sheet_row(source_row, column_count)
    ]

    replacements: dict[str, Any] = {
        "course_type": course_type,
        "start_date": _format_attendance_date_serial(target_date),
        "completed_date": "",
    }
    end_date_index = _find_attendance_column_index(columns, "end_date")
    if end_date_index is not None and end_date_index < len(source_row) and source_start_date is not None:
        source_end_value = source_row[end_date_index]
        source_end_date = None if _is_formula_expression(source_end_value) else _coerce_attendance_date(source_end_value)
        if source_end_date is not None:
            replacements["end_date"] = _format_attendance_date_serial(
                target_date + (source_end_date - source_start_date)
            )

    for field_key in ("course_name", "online_sheet"):
        column_index = _find_attendance_column_index(columns, field_key)
        if column_index is not None:
            replacements[field_key] = _derive_attendance_template_text(
                source_row[column_index] if column_index < len(source_row) else "",
                course_type=course_type,
                target_date=target_date,
            )

    for field_key, value in replacements.items():
        column_index = _find_attendance_column_index(columns, field_key)
        if column_index is not None:
            next_row[column_index] = value

    clear_from_index = _find_attendance_column_index(columns, "registration_count")
    if clear_from_index is not None:
        for column_index in range(clear_from_index, len(columns)):
            if not _is_formula_expression(next_row[column_index]):
                next_row[column_index] = ""

    return next_row


def _remap_existing_rows_for_insert(
    rows: list[Any],
    *,
    columns: list[Any],
    insert_index: int,
    amount: int,
    row_index_offset: int = 0,
) -> list[Any]:
    if amount <= 0:
        return rows
    row_index_map = {
        index: index + amount if index >= insert_index else index
        for index in range(len(rows))
    }
    return [
        _remap_row_formula_cell_references(
            row,
            columns=columns,
            row_index_map=row_index_map,
            row_index_offset=row_index_offset,
        )
        for row in rows
    ]


def _build_attendance_template_detail_response(
    session: Session,
    document: SheetDocument,
    response_document: dict[str, Any],
    *,
    access: NoteSheetResourceAccess | None = None,
    current_user: User | None = None,
) -> NoteSheetDetailResponse:
    workbook_items = _list_workbook_refs_for_sheet_ids(session, [document.id], current_user).get(document.id, [])
    parent_workbook_id = _get_parent_workbook_id_for_sheet(session, document)
    paginate_enabled, page_size = _get_document_pagination_settings(response_document)
    if paginate_enabled:
        page_document, pagination = _build_paged_document(response_document, page=1, page_size=page_size)
    else:
        page_document = _normalize_document_json(response_document)
        pagination = None

    return NoteSheetDetailResponse.model_validate(
        _serialize_sheet_detail(
            document,
            workbook_items=workbook_items,
            parent_workbook_id=parent_workbook_id,
            document_json=page_document,
            pagination=pagination,
            access=access,
        ),
    )


def _is_attendance_summary_document(session: Session, document: SheetDocument) -> bool:
    if int(document.numeric_id or 0) != ATTENDANCE_SUMMARY_SHEET_ID:
        return False

    workbook = session.exec(
        select(WorkbookDocument)
        .where(WorkbookDocument.numeric_id == ATTENDANCE_SUMMARY_WORKBOOK_ID)
        .where(_active_workbook_condition())
    ).first()
    if workbook is None:
        return False

    link = session.exec(
        select(WorkbookSheetLink)
        .where(WorkbookSheetLink.workbook_id.in_(workbook_ref_aliases(workbook)))
        .where(WorkbookSheetLink.sheet_id.in_(sheet_ref_aliases(document)))
    ).first()
    return link is not None


def run_attendance_summary_template_job() -> tuple[int, int]:
    with Session(engine) as session:
        document = session.exec(
            select(SheetDocument)
            .where(SheetDocument.numeric_id == ATTENDANCE_SUMMARY_SHEET_ID)
            .where(_active_sheet_condition())
        ).first()
        if document is None or document.scope != "notes" or not _is_attendance_summary_document(session, document):
            return 0, 0

        current_document = _normalize_document_json(dict(document.document_json or {}))
        current_document, repaired = _repair_attendance_summary_cell_meta(current_document)
        current_document, links_repaired = _repair_attendance_summary_online_sheet_links(current_document)
        if repaired or links_repaired:
            document.document_json = current_document
            document.version = max(int(document.version or 1), 1) + 1
            document.updated_by_user_id = document.owner_user_id
            document.updated_at = time.time()
            session.add(document)
            session.commit()
            session.refresh(document)
            _broadcast_sheet_resource_update(document)
            current_document = _normalize_document_json(dict(document.document_json or {}))

        job_target_date = _get_next_month_first_day()
        skip_course_types = _read_attendance_template_skip_course_types(session, job_target_date)
        job_targets = [
            (course_type, course_target_date)
            for course_type, course_target_date in _get_attendance_batch_course_targets(job_target_date)
            if course_type not in skip_course_types
        ]
        next_document, generated, skipped = _generate_attendance_next_month_templates(
            current_document,
            target_date=job_target_date,
            skip_course_types=skip_course_types,
        )
        next_document, materialized_count = _materialize_attendance_template_workbooks_for_targets(
            session,
            source_document_json=current_document,
            generated_document_json=next_document,
            generated=generated,
            targets=job_targets,
            owner_user_id=document.owner_user_id,
        )
        if (generated or materialized_count) and current_document != next_document:
            document.document_json = next_document
            document.version = max(int(document.version or 1), 1) + 1
            document.updated_by_user_id = document.owner_user_id
            document.updated_at = time.time()
            session.add(document)
            session.commit()
            _broadcast_sheet_resource_update(document)

        if generated or skipped:
            print(
                "Attendance summary template job finished: "
                f"generated={len(generated)} skipped={len(skipped)}"
            )
        return len(generated), len(skipped)


def _repair_attendance_summary_cell_meta(document_json: dict[str, Any]) -> tuple[dict[str, Any], bool]:
    """修复因插入课程模板导致的 cell_meta/entity_rows 行错位。

    链接已经迁移为行内单元格；这里仅处理样式、动作等表格元数据
    以及实体行编号的旧错位。
    返回 (修复后文档, 是否执行了修复)。
    """
    normalized = _normalize_document_json(document_json)
    if "cell_meta" not in normalized:
        return normalized, False

    columns = _normalize_document_columns(normalized)
    rows = _extract_document_rows(normalized)
    data_start_row = _normalize_document_data_start_row(normalized)
    cell_meta = dict(normalized["cell_meta"])
    if not isinstance(cell_meta, dict) or not rows:
        return normalized, False

    type_index = _find_attendance_column_index(columns, "course_type")
    start_date_index = _find_attendance_column_index(columns, "start_date")
    formula_row_offset = _get_formula_reference_row_offset(normalized)
    grid_rows = _extract_document_grid_rows(normalized)
    column_count = len(columns)

    # 找到所有本应有模板 target_date 的新插入课程行
    next_month = _get_next_month_first_day()
    targets = _get_attendance_batch_course_targets(next_month)
    new_course_data_indices: list[int] = []
    for course_type, target_date in targets:
        for row_index, row in enumerate(rows):
            row_values = _normalize_sheet_row(row, column_count)
            if type_index is not None and _normalize_sheet_text(row_values[type_index]) != course_type:
                continue
            row_date = _extract_attendance_row_start_date(
                row_values,
                row_index=row_index,
                columns=columns,
                rows=rows,
                start_date_index=start_date_index,
                reference_row_offset=formula_row_offset,
                grid_rows=grid_rows,
            )
            if row_date == target_date:
                new_course_data_indices.append(row_index)
                break

    if not new_course_data_indices:
        return normalized, False

    # 检测错位: 新课程行不应有任何 cell_meta 或 entity_row，如果存在则说明移位未生效
    corrupted = False
    for data_idx in new_course_data_indices:
        doc_row = data_start_row + data_idx
        prefix = f"{doc_row}:"
        if any(key.startswith(prefix) for key in cell_meta):
            corrupted = True
            break

    entity_rows = _extract_document_entity_rows(normalized)
    if not corrupted and entity_rows:
        for data_idx in new_course_data_indices:
            doc_row = data_start_row + data_idx
            if doc_row < len(entity_rows) and _get_document_entity_row_id(entity_rows[doc_row]):
                corrupted = True
                break

    if not corrupted:
        return normalized, False

    # 执行修复: 以正确的 row_offset 重做移位
    # 新课程行一定在数据区最前面（insert_index=0）
    amount = len(new_course_data_indices)
    repaired_meta = _shift_cell_meta_rows_for_insert(
        cell_meta,
        insert_index=0,
        amount=amount,
        row_offset=data_start_row,
    )
    normalized["cell_meta"] = repaired_meta

    # 同步修复 entity_rows: 在 data_start_row 处插入空条目，旧条目后移
    if entity_rows:
        normalized["entity_rows"] = [
            *entity_rows[:data_start_row],
            *([{}] * amount),
            *entity_rows[data_start_row:],
        ]

    return normalized, True


def init_attendance_summary_scheduler() -> None:
    if get_settings().is_test:
        return

    from backend.db import engine
    from backend.models import AppSetting
    from sqlmodel import Session
    with Session(engine) as session:
        row = session.get(AppSetting, "background_task.attendance_summary_monthly_templates.enabled")
        enabled = bool(row.value.get("enabled", False)) if row and isinstance(row.value, dict) else False

    if not enabled:
        return

    if not attendance_summary_scheduler.running:
        attendance_summary_scheduler.start()
    attendance_summary_scheduler.add_job(
        lambda: background_task_queue.enqueue(
            "attendance_summary_monthly_templates",
            run_attendance_summary_template_job,
        ),
        CronTrigger.from_crontab("5 0 27 * *"),
        id="attendance_summary_monthly_templates",
        replace_existing=True,
    )
    print("Attendance summary template job scheduled: 5 0 27 * *")


def shutdown_attendance_summary_scheduler() -> None:
    if attendance_summary_scheduler.running:
        attendance_summary_scheduler.shutdown(wait=False)


def _get_attendance_course_name_from_row(row: list[Any], columns: list[Any]) -> str:
    name_index = _find_attendance_column_index(columns, "course_name")
    if name_index is None or name_index >= len(row):
        return ""
    return _normalize_sheet_text(row[name_index])


def _get_attendance_template_course_run_bucket(text: str) -> tuple[int, int]:
    # Keep this in sync with kq5034._课程类型排序值 so generated rows keep the same run order.
    if "念住闯关" in text:
        return (1, 0)
    if "念住" in text:
        return (2, 0)
    if "觉观" in text:
        return (2, 1)
    if "禅宗" in text or "修道班" in text:
        return (3, 0)
    if "梵呗初阶" in text:
        return (10, 0)
    if "梵呗增益" in text:
        return (10, 1)
    return (10, 0)


def _get_attendance_template_row_run_bucket(row: Any, columns: list[Any]) -> tuple[int, int]:
    normalized_row = _normalize_sheet_row(row, len(columns))
    values: list[str] = []
    for field_key in ("course_type", "course_name", "online_sheet"):
        column_index = _find_attendance_column_index(columns, field_key)
        if column_index is not None and column_index < len(normalized_row):
            value = _normalize_sheet_text(normalized_row[column_index])
            if value:
                values.append(value)
    return _get_attendance_template_course_run_bucket(" ".join(values))


def _get_attendance_template_item_run_bucket(item: NoteSheetAttendanceTemplateActionItem) -> tuple[int, int]:
    values = [
        _normalize_sheet_text(item.course_type),
        _normalize_sheet_text(item.course_name),
    ]
    return _get_attendance_template_course_run_bucket(" ".join(value for value in values if value))


def _get_attendance_template_insert_index(
    rows: list[Any],
    *,
    columns: list[Any],
    pending_items: list[NoteSheetAttendanceTemplateActionItem],
) -> int:
    if not pending_items:
        return 0

    first_pending_bucket = min(_get_attendance_template_item_run_bucket(item) for item in pending_items)
    insert_index = 0
    for row_index, row in enumerate(rows):
        if _get_attendance_template_row_run_bucket(row, columns) < first_pending_bucket:
            insert_index = row_index + 1
            continue
        break
    return insert_index


def _generate_attendance_course_templates(
    document_json: dict[str, Any],
    *,
    targets: list[tuple[str, date]],
) -> tuple[dict[str, Any], list[NoteSheetAttendanceTemplateActionItem], list[NoteSheetAttendanceTemplateActionItem]]:
    normalized = _normalize_document_json(document_json)
    columns = _normalize_document_columns(normalized)
    rows = _extract_document_rows(normalized)
    formula_row_offset = _get_formula_reference_row_offset(normalized)
    formula_grid_rows = _extract_document_grid_rows(normalized)
    generated: list[NoteSheetAttendanceTemplateActionItem] = []
    skipped: list[NoteSheetAttendanceTemplateActionItem] = []

    if not rows:
        for course_type, _target_date in targets:
            skipped.append(NoteSheetAttendanceTemplateActionItem(course_type=course_type, reason="当前表没有可复制的模板行"))
        return normalized, generated, skipped

    pending_rows: list[tuple[int, int, list[Any], date, NoteSheetAttendanceTemplateActionItem]] = []
    seen_targets: set[tuple[str, date]] = set()
    for course_type, target_date in targets:
        target_order = len(seen_targets)
        target_key = (course_type, target_date)
        if target_key in seen_targets:
            continue
        seen_targets.add(target_key)

        if _attendance_template_row_exists(
            rows,
            columns=columns,
            course_type=course_type,
            target_date=target_date,
            reference_row_offset=formula_row_offset,
            grid_rows=formula_grid_rows,
        ):
            skipped.append(NoteSheetAttendanceTemplateActionItem(
                course_type=course_type,
                target_date=target_date.isoformat(),
                reason="目标课程已存在",
            ))
            continue

        source = _find_attendance_template_source_row(
            rows,
            columns=columns,
            course_type=course_type,
            target_date=target_date,
            reference_row_offset=formula_row_offset,
            grid_rows=formula_grid_rows,
        )
        if source is None:
            skipped.append(NoteSheetAttendanceTemplateActionItem(course_type=course_type, reason="没有找到上一次课程模板"))
            continue

        source_row_index, source_row, source_info = source
        preview_row = _build_inserted_attendance_template_row(
            source_row,
            columns=columns,
            source_row_index=source_row_index,
            target_row_index=source_row_index,
            target_date=target_date,
            course_type=course_type,
            source_start_date=source_info["date"],
        )
        target_course_name = _get_attendance_course_name_from_row(preview_row, columns)
        pending_rows.append((
            target_order,
            source_row_index,
            source_row,
            source_info["date"],
            NoteSheetAttendanceTemplateActionItem(
                course_type=course_type,
                course_name=target_course_name,
                target_date=target_date.isoformat(),
            ),
        ))

    if not pending_rows:
        return normalized, generated, skipped

    pending_rows.sort(key=lambda item: (_get_attendance_template_item_run_bucket(item[4]), item[0]))
    pending_items = [item for *_row_info, item in pending_rows]
    insert_index = _get_attendance_template_insert_index(
        rows,
        columns=columns,
        pending_items=pending_items,
    )
    existing_rows = _remap_existing_rows_for_insert(
        rows,
        columns=columns,
        insert_index=insert_index,
        amount=len(pending_rows),
        row_index_offset=formula_row_offset,
    )

    inserted_rows: list[list[Any]] = []
    for offset, (_target_order, source_row_index, source_row, source_start_date, item) in enumerate(pending_rows):
        target_row_index = insert_index + offset
        item_target_date = _parse_attendance_date_text(item.target_date) or _get_next_month_first_day()
        inserted_rows.append(_build_inserted_attendance_template_row(
            source_row,
            columns=columns,
            source_row_index=source_row_index,
            target_row_index=target_row_index,
            target_date=item_target_date,
            course_type=item.course_type,
            source_start_date=source_start_date,
        ))
        item.row_index = target_row_index
        generated.append(item)

    if not inserted_rows:
        return normalized, generated, skipped

    next_rows = [
        *existing_rows[:insert_index],
        *inserted_rows,
        *existing_rows[insert_index:],
    ]
    next_document = _replace_document_data_rows({
        **normalized,
        "columns": columns,
    }, next_rows)
    if "cell_meta" in normalized:
        next_document["cell_meta"] = _shift_cell_meta_rows_for_insert(
            normalized.get("cell_meta"),
            insert_index,
            len(inserted_rows),
            row_offset=_normalize_document_data_start_row(normalized),
        )

    entity_rows = _extract_document_entity_rows(normalized)
    if entity_rows:
        data_start_row = _normalize_document_data_start_row(normalized)
        next_document["entity_rows"] = [
            *entity_rows[:data_start_row + insert_index],
            *([{}] * len(inserted_rows)),
            *entity_rows[data_start_row + insert_index:],
        ]

    return next_document, generated, skipped


def _generate_attendance_next_month_templates(
    document_json: dict[str, Any],
    *,
    target_date: date,
    skip_course_types: set[str] | None = None,
) -> tuple[dict[str, Any], list[NoteSheetAttendanceTemplateActionItem], list[NoteSheetAttendanceTemplateActionItem]]:
    skipped: list[NoteSheetAttendanceTemplateActionItem] = []
    normalized_skip_course_types = {
        _normalize_sheet_text(course_type)
        for course_type in (skip_course_types or set())
        if _normalize_sheet_text(course_type)
    }
    targets: list[tuple[str, date]] = []
    for course_type, course_target_date in _get_attendance_batch_course_targets(target_date):
        if course_type in normalized_skip_course_types:
            skipped.append(NoteSheetAttendanceTemplateActionItem(
                course_type=course_type,
                target_date=course_target_date.isoformat(),
                reason="本月未排课",
            ))
            continue
        targets.append((course_type, course_target_date))
    next_document, generated, generated_skipped = _generate_attendance_course_templates(
        document_json,
        targets=targets,
    )
    return next_document, generated, [*skipped, *generated_skipped]


def _parse_local_workbook_sheet_url(value: Any) -> tuple[int | None, int | None]:
    text = _normalize_sheet_text(value)
    if not text:
        return None, None
    workbook_match = re.search(r"(?:^|/)workbook(?:s)?/(?P<workbook_id>\d+)", text)
    if workbook_match is None:
        return None, None
    sheet_match = re.search(r"(?:[?&]sheet=|/sheet/)(?P<sheet_id>\d+)", text)
    return (
        int(workbook_match.group("workbook_id")),
        int(sheet_match.group("sheet_id")) if sheet_match else None,
    )


def _attendance_course_slug(course_type: str, course_name: str) -> str:
    normalized_type = _normalize_sheet_text(course_type)
    normalized_name = _normalize_sheet_text(course_name)
    if normalized_type == "念住" or "念住" in normalized_name:
        return "nianzhu"
    if normalized_type == "觉观" or "觉观" in normalized_name:
        return "jueguan"
    if "梵呗" in normalized_type or "梵呗" in normalized_name:
        return "fanbei"
    slug = re.sub(r"[^0-9A-Za-z]+", "-", normalized_type).strip("-").lower()
    return slug or "course"


def _derive_attendance_course_owner_key(
    *,
    target_date: date,
    course_type: str,
    course_name: str,
) -> str:
    parsed = _parse_attendance_template_course(course_name)
    slug = _attendance_course_slug(course_type, course_name)
    edition = parsed.get("edition") if parsed else None
    suffix = f"-{edition}" if edition is not None else ""
    return f"{target_date:%Y%m%d}-{slug}{suffix}"


def _find_course_template_workbook_by_owner_key(
    session: Session,
    *,
    owner_key: str,
) -> tuple[WorkbookDocument, SheetDocument] | None:
    attendance = session.exec(
        select(SheetDocument)
        .where(SheetDocument.scope == "notes")
        .where(SheetDocument.owner_type == "course_workbook")
        .where(SheetDocument.owner_key == owner_key)
        .where(SheetDocument.sheet_key == "attendance")
        .where(_active_sheet_condition())
    ).first()
    if attendance is None:
        return None

    links = session.exec(
        select(WorkbookSheetLink)
        .where(WorkbookSheetLink.sheet_id.in_(sheet_ref_aliases(attendance)))
        .order_by(WorkbookSheetLink.created_at)
    ).all()
    if not links:
        return None
    workbook_map = load_workbooks_by_refs(session, [link.workbook_id for link in links])
    workbook = next(
        (
            workbook_map.get(str(link.workbook_id))
            for link in links
            if workbook_map.get(str(link.workbook_id)) is not None
        ),
        None,
    )
    return (workbook, attendance) if workbook is not None else None


def _copy_resource_access_grants(
    session: Session,
    *,
    source_resource_type: str,
    source_resource_id: str,
    target_resource_type: str,
    target_resource_id: str,
    now: float,
) -> None:
    source_grants = session.exec(
        select(ResourceAccessGrant)
        .where(ResourceAccessGrant.resource_type == source_resource_type)
        .where(ResourceAccessGrant.resource_id == source_resource_id)
    ).all()
    if not source_grants:
        return

    existing = {
        grant.subject_key: grant
        for grant in session.exec(
            select(ResourceAccessGrant)
            .where(ResourceAccessGrant.resource_type == target_resource_type)
            .where(ResourceAccessGrant.resource_id == target_resource_id)
        ).all()
    }
    for source_grant in source_grants:
        target_grant = existing.get(source_grant.subject_key)
        if target_grant is None:
            target_grant = ResourceAccessGrant(
                resource_type=target_resource_type,
                resource_id=target_resource_id,
                subject_key=source_grant.subject_key,
                subject_type=source_grant.subject_type,
                subject_user_id=source_grant.subject_user_id,
                role=source_grant.role,
                created_at=now,
                updated_at=now,
                updated_by_user_id=source_grant.updated_by_user_id,
            )
        else:
            target_grant.subject_type = source_grant.subject_type
            target_grant.subject_user_id = source_grant.subject_user_id
            target_grant.role = source_grant.role
            target_grant.updated_at = now
            target_grant.updated_by_user_id = source_grant.updated_by_user_id
        session.add(target_grant)


def _should_clear_course_template_sheet_data(source_sheet: SheetDocument) -> bool:
    # This is currently used by the monthly 念住/觉观 workbook generator, where
    # runtime video/clockin URLs must come from the target month. 禅宗/修道班 have
    # different inheritance rules and should get an explicit course-type policy
    # before they reuse this path.
    return True


def _clone_course_template_sheet_document_json(source_sheet: SheetDocument) -> dict[str, Any]:
    if _should_clear_course_template_sheet_data(source_sheet):
        return _clone_sheet_document_json(
            dict(source_sheet.document_json or {}),
            mode="template",
        )
    return _clone_sheet_document_json(
        dict(source_sheet.document_json or {}),
        mode="duplicate",
    )


def _parse_attendance_course_owner_key_date(owner_key: Any) -> date | None:
    match = re.match(r"^(?P<year>\d{4})(?P<month>\d{2})(?P<day>\d{2})(?:-|$)", _normalize_sheet_text(owner_key))
    if match is None:
        return None
    with contextlib.suppress(ValueError):
        return date(
            int(match.group("year")),
            int(match.group("month")),
            int(match.group("day")),
        )
    return None


def _shift_chinese_month_day_text(value: str, *, source_start: date, day_delta: int) -> str:
    if not value or day_delta == 0:
        return value

    def replace(match: re.Match[str]) -> str:
        month = int(match.group("month"))
        day = int(match.group("day"))
        year = source_start.year
        if source_start.month == 12 and month == 1:
            year += 1
        elif source_start.month == 1 and month == 12:
            year -= 1
        with contextlib.suppress(ValueError):
            shifted = date(year, month, day) + timedelta(days=day_delta)
            return f"{shifted.month}月{shifted.day}日"
        return match.group(0)

    return re.sub(r"(?P<month>\d{1,2})月(?P<day>\d{1,2})日", replace, value)


def _shift_chinese_month_day_cell(value: Any, *, source_start: date, day_delta: int) -> Any:
    if isinstance(value, dict):
        next_value = dict(value)
        if "value" in next_value and isinstance(next_value["value"], str):
            next_value["value"] = _shift_chinese_month_day_text(
                next_value["value"],
                source_start=source_start,
                day_delta=day_delta,
            )
        return next_value
    if isinstance(value, str):
        return _shift_chinese_month_day_text(value, source_start=source_start, day_delta=day_delta)
    return value


def _shift_date_function_text(value: str, *, day_delta: int) -> str:
    if not value or day_delta == 0:
        return value

    def replace(match: re.Match[str]) -> str:
        with contextlib.suppress(ValueError):
            shifted = date(
                int(match.group("year")),
                int(match.group("month")),
                int(match.group("day")),
            ) + timedelta(days=day_delta)
            return f"DATE({shifted.year},{shifted.month},{shifted.day})"
        return match.group(0)

    return re.sub(
        r"DATE\(\s*(?P<year>\d{4})\s*,\s*(?P<month>\d{1,2})\s*,\s*(?P<day>\d{1,2})\s*\)",
        replace,
        value,
        flags=re.IGNORECASE,
    )


def _adapt_course_template_defined_names(document_json: dict[str, Any], *, day_delta: int) -> dict[str, Any]:
    defined_names = document_json.get(NOTE_SHEET_DEFINED_NAMES_KEY)
    if not isinstance(defined_names, list) or day_delta == 0:
        return document_json

    next_names: list[Any] = []
    changed = False
    for item in defined_names:
        if not isinstance(item, dict):
            next_names.append(item)
            continue
        next_item = dict(item)
        formula = next_item.get("formula")
        if isinstance(formula, str):
            next_formula = _shift_date_function_text(formula, day_delta=day_delta)
            if next_formula != formula:
                next_item["formula"] = next_formula
                changed = True
        next_names.append(next_item)

    if not changed:
        return document_json
    next_document = dict(document_json)
    next_document[NOTE_SHEET_DEFINED_NAMES_KEY] = next_names
    return next_document


def _set_cell_display_value_preserving_link(cell: Any, value: Any) -> Any:
    if isinstance(cell, dict):
        next_cell = dict(cell)
        next_cell["value"] = value
        return next_cell
    return value


def _sync_header_grid_rows_from_entity_cells(document_json: dict[str, Any]) -> dict[str, Any]:
    grid_rows = _extract_document_grid_rows(document_json)
    entity_rows = _extract_document_entity_rows(document_json)
    entity_columns = _extract_document_entity_columns(document_json)
    entity_cells = _extract_document_entity_cells(document_json)
    if not grid_rows or not entity_rows or not entity_columns or not entity_cells:
        return document_json

    column_count = len(_normalize_document_columns(document_json))
    data_start_row = min(_normalize_document_data_start_row(document_json), len(grid_rows), len(entity_rows))
    next_grid_rows: list[Any] = []
    changed = False
    for row_index, row in enumerate(grid_rows):
        if row_index >= data_start_row:
            next_grid_rows.append(row)
            continue
        row_id = _get_document_entity_row_id(entity_rows[row_index])
        row_cells = entity_cells.get(row_id) if row_id else None
        if not isinstance(row_cells, dict):
            next_grid_rows.append(row)
            continue
        next_row = _normalize_sheet_row(row, column_count)
        for column_index, column in enumerate(entity_columns[:column_count]):
            if not isinstance(column, dict):
                continue
            column_id = _normalize_sheet_text(column.get("id"))
            entry = row_cells.get(column_id) if column_id else None
            if not isinstance(entry, dict) or "value" not in entry:
                continue
            next_cell = _set_cell_display_value_preserving_link(next_row[column_index], entry.get("value"))
            if next_cell != next_row[column_index]:
                next_row[column_index] = next_cell
                changed = True
        next_grid_rows.append(next_row)

    if not changed:
        return document_json
    next_document = dict(document_json)
    next_document["grid_rows"] = next_grid_rows
    return next_document


def _adapt_course_template_header_dates(
    document_json: dict[str, Any],
    *,
    source_owner_key: Any,
    target_owner_key: str,
) -> dict[str, Any]:
    source_start = _parse_attendance_course_owner_key_date(source_owner_key)
    target_start = _parse_attendance_course_owner_key_date(target_owner_key)
    if source_start is None or target_start is None or source_start == target_start:
        return document_json

    day_delta = (target_start - source_start).days
    next_document = dict(document_json)
    data_start_row = _normalize_document_data_start_row(next_document)
    grid_rows = _extract_document_grid_rows(next_document)
    if grid_rows:
        next_grid_rows: list[Any] = []
        for row_index, row in enumerate(grid_rows):
            if row_index >= data_start_row:
                next_grid_rows.append(row)
                continue
            if not isinstance(row, list):
                next_grid_rows.append(row)
                continue
            next_grid_rows.append([
                _shift_chinese_month_day_cell(cell, source_start=source_start, day_delta=day_delta)
                for cell in row
            ])
        next_document["grid_rows"] = next_grid_rows

    header_groups = next_document.get("header_groups")
    if isinstance(header_groups, list):
        next_header_groups: list[Any] = []
        for row in header_groups:
            if not isinstance(row, list):
                next_header_groups.append(row)
                continue
            next_row: list[Any] = []
            for cell in row:
                if not isinstance(cell, dict):
                    next_row.append(cell)
                    continue
                next_cell = dict(cell)
                if isinstance(next_cell.get("label"), str):
                    next_cell["label"] = _shift_chinese_month_day_text(
                        next_cell["label"],
                        source_start=source_start,
                        day_delta=day_delta,
                    )
                next_row.append(next_cell)
            next_header_groups.append(next_row)
        next_document["header_groups"] = next_header_groups

    entity_rows = _extract_document_entity_rows(next_document)
    entity_columns = _extract_document_entity_columns(next_document)
    entity_cells = _extract_document_entity_cells(next_document)
    if entity_rows and entity_columns and entity_cells:
        header_row_ids = {
            row_id
            for row_id in (_get_document_entity_row_id(row) for row in entity_rows[:data_start_row])
            if row_id
        }
        next_entity_cells = dict(entity_cells)
        changed = False
        for row_id in header_row_ids:
            row_cells = entity_cells.get(row_id)
            if not isinstance(row_cells, dict):
                continue
            next_row_cells = dict(row_cells)
            for column in entity_columns:
                if not isinstance(column, dict):
                    continue
                column_id = _normalize_sheet_text(column.get("id"))
                entry = next_row_cells.get(column_id)
                if not isinstance(entry, dict) or not isinstance(entry.get("value"), str):
                    continue
                next_value = _shift_chinese_month_day_text(
                    entry["value"],
                    source_start=source_start,
                    day_delta=day_delta,
                )
                if next_value == entry["value"]:
                    continue
                next_entry = dict(entry)
                next_entry["value"] = next_value
                next_row_cells[column_id] = next_entry
                changed = True
            next_entity_cells[row_id] = next_row_cells
        if changed:
            next_document["entity_cells"] = next_entity_cells
    next_document = _adapt_course_template_defined_names(next_document, day_delta=day_delta)
    next_document = _sync_header_grid_rows_from_entity_cells(next_document)
    return next_document


def _reset_course_template_runtime_header_values(document_json: dict[str, Any]) -> dict[str, Any]:
    next_document = _normalize_document_json(document_json)
    columns = _normalize_document_columns(next_document)
    grid_rows = _extract_document_grid_rows(next_document)
    data_start_row = _normalize_document_data_start_row(next_document)
    config_row_index = data_start_row - 1
    field_row_index = int(next_document.get("field_row_index") or 0)
    if config_row_index < 0 or config_row_index <= field_row_index:
        return next_document

    period_display_index = _get_column_index(columns, "已返款")
    current_refund_index = _get_column_index(columns, "当前应返款")
    if period_display_index < 0 and current_refund_index < 0:
        return next_document

    next_grid_rows = [list(row) if isinstance(row, list) else [] for row in grid_rows]
    while len(next_grid_rows) <= config_row_index:
        next_grid_rows.append([])
    config_row = _normalize_sheet_row(next_grid_rows[config_row_index], len(columns))
    changed_cells: list[int] = []
    if period_display_index >= 0:
        period_formula = f'="第"&{NOTE_SHEET_ATTENDANCE_REFUND_PERIOD_NAME}&"天"'
        if _normalize_sheet_text(config_row[period_display_index]) != period_formula:
            config_row[period_display_index] = period_formula
            changed_cells.append(period_display_index)
    if current_refund_index >= 0 and _normalize_sheet_text(config_row[current_refund_index]):
        config_row[current_refund_index] = ""
        changed_cells.append(current_refund_index)

    column_configs_changed = False
    column_configs = next_document.get("column_configs")
    next_column_configs = dict(column_configs) if isinstance(column_configs, dict) else {}
    if current_refund_index >= 0 and current_refund_index < len(columns):
        current_refund_header = str(columns[current_refund_index])
        current_refund_config = next_column_configs.get(current_refund_header)
        if isinstance(current_refund_config, dict) and "note" in current_refund_config:
            next_config = dict(current_refund_config)
            next_config.pop("note", None)
            if next_config:
                next_column_configs[current_refund_header] = next_config
            else:
                next_column_configs.pop(current_refund_header, None)
            column_configs_changed = True

    if not changed_cells and not column_configs_changed:
        return next_document

    next_grid_rows[config_row_index] = config_row
    next_document = dict(next_document)
    next_document["grid_rows"] = next_grid_rows
    if column_configs_changed:
        if next_column_configs:
            next_document["column_configs"] = next_column_configs
        else:
            next_document.pop("column_configs", None)
    for column_index in changed_cells:
        next_document = _set_document_entity_cell_value(
            next_document,
            document_row=config_row_index,
            column_index=column_index,
            value=config_row[column_index],
        )
    return next_document


def _normalize_course_template_refund_header_styles(document_json: dict[str, Any]) -> dict[str, Any]:
    next_document = _normalize_document_json(document_json)
    columns = _normalize_document_columns(next_document)
    start_index = _get_column_index(columns, "完成视频数")
    end_index = _get_column_index(columns, "当前应返款")
    if start_index < 0 or end_index < start_index:
        return next_document

    field_row_index = int(next_document.get("field_row_index") or 0)
    note_row_index = max(_normalize_document_data_start_row(next_document) - 1, 0)
    styles_by_row = {
        max(field_row_index - 1, 0): {"background_color": "#FFBA84"},
        field_row_index: {"background_color": "#FFDCC4"},
        note_row_index: {"background_color": "#D8D8D8"},
    }

    cell_meta = deepcopy(dict(next_document.get("cell_meta") or {}))
    changed = False
    for row_index, style in styles_by_row.items():
        for column_index in range(start_index, end_index + 1):
            key = f"{row_index}:{column_index}"
            meta = deepcopy(cell_meta.get(key)) if isinstance(cell_meta.get(key), dict) else {}
            previous_style = meta.get("style") if isinstance(meta.get("style"), dict) else {}
            next_style = dict(previous_style)
            for stale_key in ("background_color", "text_color"):
                next_style.pop(stale_key, None)
            next_style.update(style)
            if next_style != previous_style:
                meta["style"] = next_style
                cell_meta[key] = meta
                changed = True

    if changed:
        next_document = dict(next_document)
        next_document["cell_meta"] = cell_meta

    entity_rows = _extract_document_entity_rows(next_document)
    entity_columns = _extract_document_entity_columns(next_document)
    entity_cells = _extract_document_entity_cells(next_document)
    if entity_rows and entity_columns and entity_cells:
        next_entity_cells = deepcopy(entity_cells)
        for row_index, style in styles_by_row.items():
            if row_index < 0 or row_index >= len(entity_rows):
                continue
            row_id = _get_document_entity_row_id(entity_rows[row_index])
            if not row_id:
                continue
            row_cells = dict(next_entity_cells.get(row_id) or {})
            for column_index in range(start_index, min(end_index + 1, len(entity_columns))):
                column = entity_columns[column_index]
                if not isinstance(column, dict):
                    continue
                column_id = _normalize_sheet_text(column.get("id"))
                if not column_id:
                    continue
                entry = dict(row_cells.get(column_id) or {})
                previous_style = entry.get("style") if isinstance(entry.get("style"), dict) else {}
                next_style = dict(previous_style)
                for stale_key in ("background_color", "text_color"):
                    next_style.pop(stale_key, None)
                next_style.update(style)
                if next_style == previous_style:
                    continue
                entry["style"] = next_style
                row_cells[column_id] = entry
                changed = True
            if row_cells:
                next_entity_cells[row_id] = row_cells
        if changed:
            next_document = dict(next_document)
            next_document["entity_cells"] = next_entity_cells

    return next_document


def _set_header_grid_cell_inline_link(
    document_json: dict[str, Any],
    *,
    row_index: int,
    column_index: int,
    text: str,
    url: str,
    style: dict[str, Any] | None = None,
) -> tuple[dict[str, Any], bool]:
    columns = _normalize_document_columns(document_json)
    if row_index < 0 or column_index < 0 or column_index >= len(columns):
        return document_json, False

    next_document = deepcopy(document_json)
    grid_rows = _extract_document_grid_rows(next_document)
    if row_index >= len(grid_rows):
        return next_document, False

    changed = False
    next_grid_rows = [list(row) if isinstance(row, list) else [] for row in grid_rows]
    target_row = _normalize_sheet_row(next_grid_rows[row_index], len(columns))
    source_grid_cell = target_row[column_index]
    linked_cell = _with_inline_cell_link(text, url)
    link_style = style or NOTE_SHEET_INLINE_LINK_STYLE
    if isinstance(linked_cell, dict):
        source_style = source_grid_cell.get("style") if isinstance(source_grid_cell, dict) else None
        linked_cell["style"] = {
            **(source_style if isinstance(source_style, dict) else {}),
            **link_style,
        }
    if target_row[column_index] != linked_cell:
        target_row[column_index] = linked_cell
        next_grid_rows[row_index] = target_row
        next_document = dict(next_document)
        next_document["grid_rows"] = next_grid_rows
        changed = True

    entity_rows = _extract_document_entity_rows(next_document)
    entity_columns = _extract_document_entity_columns(next_document)
    entity_cells = _extract_document_entity_cells(next_document)
    if (
        entity_rows
        and entity_columns
        and row_index < len(entity_rows)
        and column_index < len(entity_columns)
    ):
        row_id = _get_document_entity_row_id(entity_rows[row_index])
        column_id = _get_document_entity_column_id(entity_columns[column_index])
        if row_id and column_id:
            next_entity_cells = dict(entity_cells)
            row_cells = dict(next_entity_cells.get(row_id) or {})
            source_entry = row_cells.get(column_id)
            next_entry = dict(source_entry) if isinstance(source_entry, dict) else {}
            next_entry["value"] = text
            next_entry["link"] = {"url": url}
            source_style = next_entry.get("style")
            next_entry["style"] = {
                **(source_style if isinstance(source_style, dict) else {}),
                **link_style,
            }
            if next_entry != source_entry:
                row_cells[column_id] = next_entry
                next_entity_cells[row_id] = row_cells
                next_document = dict(next_document)
                next_document["entity_cells"] = next_entity_cells
                changed = True

    return next_document, changed


def _normalize_attendance_refund_faq_link(document_json: dict[str, Any]) -> dict[str, Any]:
    next_document = deepcopy(document_json)
    columns = _normalize_document_columns(next_document)
    grid_rows = _extract_document_grid_rows(next_document)
    data_start_row = _normalize_document_data_start_row(next_document)
    if not columns or not grid_rows or data_start_row <= 0:
        return next_document

    target_row_index = -1
    target_column_index = -1
    for row_index, row in enumerate(grid_rows[:data_start_row]):
        if not isinstance(row, list):
            continue
        normalized_row = _normalize_sheet_row(row, len(columns))
        for column_index, cell in enumerate(normalized_row):
            if _normalize_sheet_text(_extract_cell_value(cell)) == NOTE_SHEET_ATTENDANCE_REFUND_FAQ_TEXT:
                target_row_index = row_index
                target_column_index = column_index
                break
        if target_row_index >= 0:
            break
    if target_row_index < 0 or target_column_index < 0:
        return next_document

    next_document, changed = _set_header_grid_cell_inline_link(
        next_document,
        row_index=target_row_index,
        column_index=target_column_index,
        text=NOTE_SHEET_ATTENDANCE_REFUND_FAQ_TEXT,
        url=NOTE_SHEET_ATTENDANCE_REFUND_FAQ_URL,
        style=NOTE_SHEET_ATTENDANCE_REFUND_FAQ_LINK_STYLE,
    )
    return next_document if changed else document_json


def _normalize_attendance_feedback_link(document_json: dict[str, Any]) -> dict[str, Any]:
    next_document = deepcopy(document_json)
    columns = _normalize_document_columns(next_document)
    total_refund_index = _get_column_index(columns, "总应返款")
    config_row_index = _normalize_document_data_start_row(next_document) - 1
    field_row_index = int(next_document.get("field_row_index") or 0)
    if total_refund_index < 0 or config_row_index < 0 or config_row_index <= field_row_index:
        return next_document

    next_document, changed = _set_header_grid_cell_inline_link(
        next_document,
        row_index=config_row_index,
        column_index=total_refund_index,
        text=NOTE_SHEET_ATTENDANCE_FEEDBACK_TEXT,
        url=NOTE_SHEET_ATTENDANCE_FEEDBACK_URL,
    )
    return next_document if changed else document_json


def _cell_without_inline_link(cell: Any) -> Any:
    if not isinstance(cell, dict):
        return cell
    next_cell = dict(cell)
    next_cell.pop("link", None)
    if set(next_cell) == {"value"}:
        return next_cell.get("value")
    return next_cell


def _strip_course_runtime_header_links(document_json: dict[str, Any]) -> dict[str, Any]:
    next_document = deepcopy(document_json)
    columns = _normalize_document_columns(next_document)
    field_row_index = int(next_document.get("field_row_index") or 0)
    if not columns or field_row_index < 0:
        return next_document

    target_indexes = {
        index
        for index, column in enumerate(columns)
        if _normalize_sheet_text(column) == "打卡数" or _attendance_lesson_number_from_text(column) is not None
    }
    if not target_indexes:
        return next_document

    changed = False
    grid_rows = _extract_document_grid_rows(next_document)
    if field_row_index < len(grid_rows):
        next_grid_rows = [list(row) if isinstance(row, list) else [] for row in grid_rows]
        row = _normalize_sheet_row(next_grid_rows[field_row_index], len(columns))
        for column_index in target_indexes:
            source_cell = row[column_index]
            next_cell = _cell_without_inline_link(source_cell)
            if next_cell != source_cell:
                row[column_index] = next_cell
                changed = True
        if changed:
            next_grid_rows[field_row_index] = row
            next_document["grid_rows"] = next_grid_rows

    cell_meta = deepcopy(dict(next_document.get("cell_meta") or {}))
    cell_meta_changed = False
    for column_index in target_indexes:
        key = f"{field_row_index}:{column_index}"
        meta = cell_meta.get(key)
        if not isinstance(meta, dict) or "link" not in meta:
            continue
        next_meta = dict(meta)
        next_meta.pop("link", None)
        if next_meta:
            cell_meta[key] = next_meta
        else:
            cell_meta.pop(key, None)
        cell_meta_changed = True
    if cell_meta_changed:
        if cell_meta:
            next_document["cell_meta"] = cell_meta
        else:
            next_document.pop("cell_meta", None)
        changed = True

    entity_rows = _extract_document_entity_rows(next_document)
    entity_columns = _extract_document_entity_columns(next_document)
    entity_cells = _extract_document_entity_cells(next_document)
    if entity_rows and entity_columns and entity_cells and field_row_index < len(entity_rows):
        row_id = _get_document_entity_row_id(entity_rows[field_row_index])
        if row_id:
            next_entity_cells = deepcopy(entity_cells)
            row_cells = dict(next_entity_cells.get(row_id) or {})
            row_changed = False
            for column_index in target_indexes:
                if column_index >= len(entity_columns):
                    continue
                column_id = _get_document_entity_column_id(entity_columns[column_index])
                if not column_id:
                    continue
                entry = row_cells.get(column_id)
                if not isinstance(entry, dict) or "link" not in entry:
                    continue
                next_entry = dict(entry)
                next_entry.pop("link", None)
                row_cells[column_id] = next_entry
                row_changed = True
            if row_changed:
                next_entity_cells[row_id] = row_cells
                next_document["entity_cells"] = next_entity_cells
                changed = True

    return next_document if changed else document_json


def _attendance_lesson_number_from_text(value: Any) -> int | None:
    match = re.search(r"第\s*0*(?P<number>\d+)\s*课", _normalize_sheet_text(_extract_cell_value(value)))
    if match is None:
        return None
    with contextlib.suppress(ValueError):
        return int(match.group("number"))
    return None


def _is_plain_attendance_lesson_header(value: Any) -> bool:
    return re.fullmatch(r"第\s*0*\d+\s*课", _normalize_sheet_text(_extract_cell_value(value))) is not None


def _remove_duplicate_plain_lesson_columns(document_json: dict[str, Any]) -> dict[str, Any]:
    next_document = _normalize_document_json(document_json)
    columns = _normalize_document_columns(next_document)
    if not columns:
        return next_document

    decorated_lesson_numbers: set[int] = set()
    for column in columns:
        if _is_plain_attendance_lesson_header(column):
            continue
        lesson_number = _attendance_lesson_number_from_text(column)
        if lesson_number is not None:
            decorated_lesson_numbers.add(lesson_number)

    if not decorated_lesson_numbers:
        return next_document

    for index in range(len(columns) - 1, -1, -1):
        if not _is_plain_attendance_lesson_header(columns[index]):
            continue
        lesson_number = _attendance_lesson_number_from_text(columns[index])
        if lesson_number not in decorated_lesson_numbers:
            continue
        next_document = _delete_document_column(next_document, delete_index=index)
        columns = _normalize_document_columns(next_document)
    return next_document


def _strip_course_template_column_header_colors(document_json: dict[str, Any]) -> dict[str, Any]:
    next_document = _normalize_document_json(document_json)
    column_configs = next_document.get("column_configs")
    if not isinstance(column_configs, dict):
        return next_document

    next_configs: dict[str, Any] = {}
    changed = False
    for column, config in column_configs.items():
        if not isinstance(config, dict):
            next_configs[column] = config
            continue
        next_config = dict(config)
        for key in ("header_background_color", "header_text_color"):
            if key in next_config:
                next_config.pop(key, None)
                changed = True
        if next_config:
            next_configs[column] = next_config
        else:
            changed = True
    if not changed:
        return next_document
    next_document["column_configs"] = next_configs
    return next_document


def _maybe_materialize_zen_course_data_sheets(
    session: Session,
    *,
    workbook: WorkbookDocument,
    attendance_sheet: SheetDocument,
    course_name: str,
    allow_header_link_fallback: bool = True,
) -> None:
    normalized_course_name = _normalize_sheet_text(course_name)
    if "念住" not in normalized_course_name and "觉观" not in normalized_course_name:
        return

    from backend.core.attendance.nianzhu_course_sheets import (
        CLOCKIN_CONFIG_SHEET_KEY,
        VIDEO_CONFIG_SHEET_KEY,
        _course_sheet_documents_from_attendance,
        materialize_nianzhu_course_sheets,
    )

    documents = _course_sheet_documents_from_attendance(
        deepcopy(dict(attendance_sheet.document_json or {})),
        course_name=normalized_course_name,
    )
    video_meta = dict(documents[VIDEO_CONFIG_SHEET_KEY].get("source_meta") or {})
    clockin_meta = dict(documents[CLOCKIN_CONFIG_SHEET_KEY].get("source_meta") or {})
    video_rows = _extract_document_rows(documents[VIDEO_CONFIG_SHEET_KEY])
    clockin_rows = _extract_document_rows(documents[CLOCKIN_CONFIG_SHEET_KEY])
    has_video_source = int(video_meta.get("legacy_lesson_rows") or 0) > 0 or (allow_header_link_fallback and any(
        _normalize_sheet_text(row[4] if isinstance(row, list) and len(row) > 4 else "")
        for row in video_rows
    ))
    has_clockin_source = int(clockin_meta.get("legacy_clockin_rows") or 0) > 0 or (allow_header_link_fallback and any(
        _normalize_sheet_text(row[2] if isinstance(row, list) and len(row) > 2 else "")
        for row in clockin_rows
    ))
    if not has_video_source:
        return
    if not has_clockin_source:
        return

    summary = materialize_nianzhu_course_sheets(
        session,
        workbook_id=_require_workbook_numeric_id(workbook),
        attendance_sheet_id=_require_sheet_numeric_id(attendance_sheet),
        course_name=normalized_course_name,
        replace=True,
    )
    if int(summary.get("legacy_lesson_rows") or 0) > 0 or has_video_source:
        _prune_no_attendance_video_config_rows_for_course_template(
            session,
            attendance_sheet=attendance_sheet,
            owner_key=_normalize_sheet_text(attendance_sheet.owner_key),
        )
        attendance_sheet.document_json = _remove_duplicate_plain_lesson_columns(
            dict(attendance_sheet.document_json or {})
        )
        attendance_sheet.document_json = _strip_course_template_column_header_colors(
            dict(attendance_sheet.document_json or {})
        )
        attendance_sheet.document_json = _normalize_course_template_refund_header_styles(
            dict(attendance_sheet.document_json or {})
        )
        attendance_sheet.document_json = _normalize_attendance_feedback_link(
            dict(attendance_sheet.document_json or {})
        )
        attendance_sheet.document_json = _normalize_attendance_refund_faq_link(
            dict(attendance_sheet.document_json or {})
        )
        attendance_sheet.version = max(int(attendance_sheet.version or 1), 1) + 1
        attendance_sheet.updated_at = time.time()
        session.add(attendance_sheet)
        _normalize_no_attendance_color_boundary(session, attendance_sheet)


def _no_attendance_video_column_names(attendance_sheet: SheetDocument) -> set[str]:
    document = _normalize_document_json(dict(attendance_sheet.document_json or {}))
    grid_rows = _extract_document_grid_rows(document)
    if not grid_rows:
        return set()
    field_row_index = int(document.get("field_row_index") or 0)
    note_row_index = max(_normalize_document_data_start_row(document) - 1, 0)
    if field_row_index < 0 or field_row_index >= len(grid_rows) or note_row_index >= len(grid_rows):
        return set()

    field_row = grid_rows[field_row_index]
    note_row = grid_rows[note_row_index]
    result: set[str] = set()
    for column_index, header_cell in enumerate(field_row):
        note_text = _normalize_sheet_text(note_row[column_index] if column_index < len(note_row) else "")
        if "不做考勤" not in note_text:
            continue
        header_text = _normalize_sheet_text(_extract_cell_value(header_cell))
        if header_text:
            result.add(header_text)
    return result


def _cell_background_color(cell_meta: dict[str, Any], row_index: int, column_index: int) -> str:
    meta = cell_meta.get(f"{row_index}:{column_index}")
    if not isinstance(meta, dict):
        return ""
    style = meta.get("style")
    if not isinstance(style, dict):
        return ""
    return _normalize_sheet_text(style.get("background_color"))


def _copy_cell_meta_style(
    cell_meta: dict[str, Any],
    *,
    source_row_index: int,
    source_column_index: int,
    target_row_index: int,
    target_column_index: int,
) -> bool:
    source_meta = cell_meta.get(f"{source_row_index}:{source_column_index}")
    if not isinstance(source_meta, dict):
        return False
    source_style = source_meta.get("style")
    if not isinstance(source_style, dict):
        return False

    target_key = f"{target_row_index}:{target_column_index}"
    target_meta = deepcopy(cell_meta.get(target_key)) if isinstance(cell_meta.get(target_key), dict) else {}
    target_style = target_meta.get("style") if isinstance(target_meta.get("style"), dict) else {}
    if target_style == source_style:
        return False

    target_meta["style"] = deepcopy(source_style)
    cell_meta[target_key] = target_meta
    return True


def _copy_entity_cell_style(
    document: dict[str, Any],
    *,
    source_row_index: int,
    source_column_index: int,
    target_row_index: int,
    target_column_index: int,
) -> bool:
    entity_rows = _extract_document_entity_rows(document)
    entity_columns = _extract_document_entity_columns(document)
    entity_cells = _extract_document_entity_cells(document)
    if (
        source_row_index < 0
        or target_row_index < 0
        or source_column_index < 0
        or target_column_index < 0
        or source_row_index >= len(entity_rows)
        or target_row_index >= len(entity_rows)
        or source_column_index >= len(entity_columns)
        or target_column_index >= len(entity_columns)
        or not entity_cells
    ):
        return False

    source_row_id = _get_document_entity_row_id(entity_rows[source_row_index])
    target_row_id = _get_document_entity_row_id(entity_rows[target_row_index])
    source_column = entity_columns[source_column_index]
    target_column = entity_columns[target_column_index]
    if not isinstance(source_column, dict) or not isinstance(target_column, dict):
        return False
    source_column_id = _normalize_sheet_text(source_column.get("id"))
    target_column_id = _normalize_sheet_text(target_column.get("id"))
    if not source_row_id or not target_row_id or not source_column_id or not target_column_id:
        return False

    source_entry = entity_cells.get(source_row_id, {}).get(source_column_id)
    if not isinstance(source_entry, dict) or not isinstance(source_entry.get("style"), dict):
        return False

    next_entity_cells = dict(entity_cells)
    target_row_cells = dict(next_entity_cells.get(target_row_id) or {})
    target_entry = dict(target_row_cells.get(target_column_id) or {})
    if target_entry.get("style") == source_entry["style"]:
        return False
    target_entry["style"] = deepcopy(source_entry["style"])
    target_row_cells[target_column_id] = target_entry
    next_entity_cells[target_row_id] = target_row_cells
    document["entity_cells"] = next_entity_cells
    return True


def _normalize_no_attendance_color_boundary(session: Session, attendance_sheet: SheetDocument) -> None:
    document = _normalize_document_json(dict(attendance_sheet.document_json or {}))
    grid_rows = _extract_document_grid_rows(document)
    if not grid_rows:
        return

    field_row_index = int(document.get("field_row_index") or 0)
    note_row_index = max(_normalize_document_data_start_row(document) - 1, 0)
    if field_row_index < 0 or field_row_index >= len(grid_rows) or note_row_index >= len(grid_rows):
        return

    note_row = grid_rows[note_row_index]
    no_attendance_indexes = [
        index
        for index, cell in enumerate(note_row)
        if "不做考勤" in _normalize_sheet_text(cell)
    ]
    if not no_attendance_indexes:
        return

    first_no_attendance_index = min(no_attendance_indexes)
    boundary_previous_index = first_no_attendance_index - 1
    source_index = boundary_previous_index - 1
    if boundary_previous_index < 0 or source_index < 0:
        return

    cell_meta = deepcopy(dict(document.get("cell_meta") or {}))
    changed = False
    for row_index in range(0, note_row_index + 1):
        changed = _copy_cell_meta_style(
            cell_meta,
            source_row_index=row_index,
            source_column_index=source_index,
            target_row_index=row_index,
            target_column_index=boundary_previous_index,
        ) or changed
        changed = _copy_entity_cell_style(
            document,
            source_row_index=row_index,
            source_column_index=source_index,
            target_row_index=row_index,
            target_column_index=boundary_previous_index,
        ) or changed

    if not changed:
        return
    document["cell_meta"] = cell_meta
    attendance_sheet.document_json = document
    attendance_sheet.version = max(int(attendance_sheet.version or 1), 1) + 1
    attendance_sheet.updated_at = time.time()
    session.add(attendance_sheet)


def _prune_no_attendance_video_config_rows_for_course_template(
    session: Session,
    *,
    attendance_sheet: SheetDocument,
    owner_key: str,
) -> None:
    if not owner_key:
        return
    no_attendance_names = _no_attendance_video_column_names(attendance_sheet)
    if not no_attendance_names:
        return
    video_config = session.exec(
        select(SheetDocument)
        .where(SheetDocument.scope == "notes")
        .where(SheetDocument.owner_type == "course_workbook")
        .where(SheetDocument.owner_key == owner_key)
        .where(SheetDocument.sheet_key == "video_config")
        .where(_active_sheet_condition())
    ).first()
    if video_config is None:
        return
    document = _normalize_document_json(dict(video_config.document_json or {}))
    columns = _normalize_document_columns(document)
    if "lesson_id2" not in columns:
        return
    lesson_url_index = columns.index("lesson_id2")
    lesson_name_index = columns.index("lesson_name") if "lesson_name" in columns else None
    rows = [_normalize_sheet_row(row, len(columns)) for row in _extract_document_rows(document)]
    next_rows = [
        row for row in rows
        if _normalize_sheet_text(row[lesson_url_index])
        or lesson_name_index is None
        or _normalize_sheet_text(row[lesson_name_index]) not in no_attendance_names
    ]
    if len(next_rows) == len(rows):
        return
    next_document = dict(document)
    next_document["rows"] = next_rows
    grid_rows = _extract_document_grid_rows(document)
    if grid_rows:
        data_start_row = _normalize_document_data_start_row(document)
        header_rows = [_normalize_sheet_row(row, len(columns)) for row in grid_rows[:data_start_row]]
        next_document["grid_rows"] = [*header_rows, *next_rows]
    video_config.document_json = next_document
    video_config.version = max(int(video_config.version or 1), 1) + 1
    video_config.updated_at = time.time()
    session.add(video_config)


def _clone_attendance_course_template_workbook(
    session: Session,
    *,
    source_workbook_id: int,
    title: str,
    owner_key: str,
    owner_user_id: int | None,
) -> tuple[WorkbookDocument, SheetDocument] | None:
    existing = _find_course_template_workbook_by_owner_key(session, owner_key=owner_key)
    if existing is not None:
        return existing

    source_workbook = session.exec(
        select(WorkbookDocument)
        .where(WorkbookDocument.numeric_id == int(source_workbook_id))
        .where(_active_workbook_condition())
    ).first()
    if source_workbook is None:
        return None

    links = session.exec(
        select(WorkbookSheetLink)
        .where(WorkbookSheetLink.workbook_id.in_(workbook_ref_aliases(source_workbook)))
        .order_by(WorkbookSheetLink.order_index, WorkbookSheetLink.created_at)
    ).all()
    source_sheet_map = load_sheets_by_refs(session, [link.sheet_id for link in links])
    source_sheets = [
        (link, source_sheet_map.get(str(link.sheet_id)))
        for link in links
        if source_sheet_map.get(str(link.sheet_id)) is not None
    ]
    if not source_sheets:
        return None

    now = time.time()
    effective_owner_user_id = owner_user_id or source_workbook.owner_user_id
    workbook_identity = allocate_new_workbook_identity(session)
    workbook = WorkbookDocument(
        id=workbook_identity.primary_id,
        numeric_id=workbook_identity.numeric_id,
        legacy_id=workbook_identity.legacy_id,
        title=_normalize_title(title, default_value="未命名工作簿"),
        owner_user_id=effective_owner_user_id,
        created_by_user_id=effective_owner_user_id,
        updated_by_user_id=effective_owner_user_id,
        created_at=now,
        updated_at=now,
    )
    session.add(workbook)
    session.flush()
    _ensure_workbook_identity(session, workbook)
    _set_workbook_defined_names(session, workbook, _get_workbook_defined_names(session, source_workbook))
    _copy_resource_access_grants(
        session,
        source_resource_type=RESOURCE_TYPE_WORKBOOK,
        source_resource_id=_workbook_resource_id(source_workbook),
        target_resource_type=RESOURCE_TYPE_WORKBOOK,
        target_resource_id=_workbook_resource_id(workbook),
        now=now,
    )

    attendance_sheet: SheetDocument | None = None
    for link, source_sheet in source_sheets:
        if source_sheet is None:
            continue
        cloned_document_json = _clone_course_template_sheet_document_json(source_sheet)
        if _normalize_sheet_text(source_sheet.sheet_key) == "attendance":
            cloned_document_json = _adapt_course_template_header_dates(
                cloned_document_json,
                source_owner_key=source_sheet.owner_key,
                target_owner_key=owner_key,
            )
            cloned_document_json = _strip_course_runtime_header_links(cloned_document_json)
            cloned_document_json = _reset_course_template_runtime_header_values(cloned_document_json)
            cloned_document_json = _normalize_course_template_refund_header_styles(cloned_document_json)
            cloned_document_json = _normalize_attendance_feedback_link(cloned_document_json)
            cloned_document_json = _normalize_attendance_refund_faq_link(cloned_document_json)
        document_identity = allocate_new_sheet_identity(session)
        document = SheetDocument(
            id=document_identity.primary_id,
            numeric_id=document_identity.numeric_id,
            legacy_id=document_identity.legacy_id,
            scope=source_sheet.scope,
            owner_type="course_workbook",
            owner_key=owner_key,
            sheet_key=source_sheet.sheet_key,
            title=source_sheet.title,
            engine=source_sheet.engine,
            document_json=cloned_document_json,
            version=1,
            owner_user_id=effective_owner_user_id,
            created_by_user_id=effective_owner_user_id,
            updated_by_user_id=effective_owner_user_id,
            created_at=now,
            updated_at=now,
        )
        session.add(document)
        session.flush()
        ensure_attendance_sheet_anonymous_viewer(session, document)
        _copy_resource_access_grants(
            session,
            source_resource_type=RESOURCE_TYPE_SHEET,
            source_resource_id=_sheet_resource_id(source_sheet),
            target_resource_type=RESOURCE_TYPE_SHEET,
            target_resource_id=_sheet_resource_id(document),
            now=now,
        )
        session.add(
            WorkbookSheetLink(
                workbook_id=_workbook_link_ref(workbook),
                sheet_id=_sheet_link_ref(document),
                order_index=link.order_index,
                created_at=now,
            )
        )
        if _normalize_sheet_text(source_sheet.sheet_key) == "attendance":
            attendance_sheet = document

    if attendance_sheet is None:
        return None
    _normalize_no_attendance_color_boundary(session, attendance_sheet)
    _maybe_materialize_zen_course_data_sheets(
        session,
        workbook=workbook,
        attendance_sheet=attendance_sheet,
        course_name=title,
        allow_header_link_fallback=False,
    )
    return workbook, attendance_sheet


def _materialize_generated_attendance_template_workbooks(
    session: Session,
    *,
    source_document_json: dict[str, Any],
    generated_document_json: dict[str, Any],
    generated: list[NoteSheetAttendanceTemplateActionItem],
    owner_user_id: int | None,
) -> tuple[dict[str, Any], int]:
    if not generated:
        return generated_document_json, 0

    source_document = _normalize_document_json(source_document_json)
    source_columns = _normalize_document_columns(source_document)
    source_rows = _extract_document_rows(source_document)
    source_formula_row_offset = _get_formula_reference_row_offset(source_document)
    source_grid_rows = _extract_document_grid_rows(source_document)

    next_document = _normalize_document_json(generated_document_json)
    next_columns = _normalize_document_columns(next_document)
    next_rows = list(_extract_document_rows(next_document))
    online_sheet_index = _find_attendance_column_index(next_columns, "online_sheet")
    start_date_index = _find_attendance_column_index(next_columns, "start_date")
    if online_sheet_index is None or start_date_index is None:
        return next_document, 0

    materialized = 0
    for item in generated:
        if item.row_index is None or item.row_index < 0 or item.row_index >= len(next_rows):
            continue
        target_row = _normalize_sheet_row(next_rows[item.row_index], len(next_columns))
        if note_sheet_inline_links.inline_cell_link(target_row[online_sheet_index]):
            continue
        target_date = _extract_attendance_row_start_date(
            target_row,
            row_index=item.row_index,
            columns=next_columns,
            rows=next_rows,
            start_date_index=start_date_index,
            reference_row_offset=_get_formula_reference_row_offset(next_document),
            grid_rows=_extract_document_grid_rows(next_document),
        )
        if target_date is None:
            target_date = _parse_attendance_date_text(item.target_date)
        if target_date is None:
            continue

        source = _find_attendance_template_source_row(
            source_rows,
            columns=source_columns,
            course_type=item.course_type,
            target_date=target_date,
            reference_row_offset=source_formula_row_offset,
            grid_rows=source_grid_rows,
        )
        if source is None:
            continue
        source_row_index, _source_row, _source_info = source
        source_online_index = _find_attendance_column_index(source_columns, "online_sheet")
        if source_online_index is None:
            continue
        source_url = _get_document_cell_link_url(source_document, source_row_index, source_online_index)
        source_workbook_id, _source_sheet_id = _parse_local_workbook_sheet_url(source_url)
        if source_workbook_id is None:
            continue

        target_title = _normalize_sheet_text(target_row[online_sheet_index]) or item.course_name
        owner_key = _derive_attendance_course_owner_key(
            target_date=target_date,
            course_type=item.course_type,
            course_name=target_title or item.course_name,
        )
        cloned = _clone_attendance_course_template_workbook(
            session,
            source_workbook_id=source_workbook_id,
            title=target_title or item.course_name,
            owner_key=owner_key,
            owner_user_id=owner_user_id,
        )
        if cloned is None:
            continue
        workbook, attendance_sheet = cloned
        link_url = f"/workbook/{_require_workbook_numeric_id(workbook)}?sheet={_require_sheet_numeric_id(attendance_sheet)}"
        next_rows[item.row_index] = _set_row_cell_value(
            next_rows[item.row_index],
            next_columns,
            online_sheet_index,
            _with_inline_cell_link(target_title or item.course_name, link_url),
        )
        materialized += 1

    if not materialized:
        return next_document, 0
    return _normalize_document_json(_replace_document_data_rows(next_document, next_rows)), materialized


def _materialize_attendance_template_workbooks_for_targets(
    session: Session,
    *,
    source_document_json: dict[str, Any],
    generated_document_json: dict[str, Any],
    generated: list[NoteSheetAttendanceTemplateActionItem],
    targets: list[tuple[str, date]],
    owner_user_id: int | None,
) -> tuple[dict[str, Any], int]:
    if not targets and not generated:
        return generated_document_json, 0

    next_document = _normalize_document_json(generated_document_json)
    columns = _normalize_document_columns(next_document)
    rows = _extract_document_rows(next_document)
    type_index = _find_attendance_column_index(columns, "course_type")
    online_sheet_index = _find_attendance_column_index(columns, "online_sheet")
    start_date_index = _find_attendance_column_index(columns, "start_date")
    if type_index is None or online_sheet_index is None or start_date_index is None:
        return next_document, 0

    formula_row_offset = _get_formula_reference_row_offset(next_document)
    grid_rows = _extract_document_grid_rows(next_document)
    materialize_items = list(generated)
    existing_keys = {
        (item.course_type, _parse_attendance_date_text(item.target_date))
        for item in generated
    }

    for course_type, target_date in targets:
        if (course_type, target_date) in existing_keys:
            continue
        for row_index, raw_row in enumerate(rows):
            row = _normalize_sheet_row(raw_row, len(columns))
            if _normalize_sheet_text(row[type_index]) != course_type:
                continue
            row_date = _extract_attendance_row_start_date(
                row,
                row_index=row_index,
                columns=columns,
                rows=rows,
                start_date_index=start_date_index,
                reference_row_offset=formula_row_offset,
                grid_rows=grid_rows,
            )
            if row_date != target_date:
                continue
            if note_sheet_inline_links.inline_cell_link(row[online_sheet_index]):
                break
            materialize_items.append(NoteSheetAttendanceTemplateActionItem(
                course_type=course_type,
                course_name=_normalize_sheet_text(row[online_sheet_index]),
                target_date=target_date.isoformat(),
                row_index=row_index,
                reason="目标课程已存在，补建工作簿",
            ))
            break

    return _materialize_generated_attendance_template_workbooks(
        session,
        source_document_json=source_document_json,
        generated_document_json=next_document,
        generated=materialize_items,
        owner_user_id=owner_user_id,
    )


def _resolve_attendance_course_template_type(
    document_json: dict[str, Any],
    payload: NoteSheetAttendanceCourseTemplateGenerationRequest,
) -> str:
    if payload.course_type is not None:
        course_type = _normalize_sheet_text(payload.course_type)
        if course_type:
            return course_type

    if payload.row_index is None:
        raise HTTPException(status_code=400, detail="必须提供 row_index 或 course_type")

    normalized = _normalize_document_json(document_json)
    columns = _normalize_document_columns(normalized)
    rows = _extract_document_rows(normalized)
    type_index = _find_attendance_column_index(columns, "course_type")
    if type_index is None:
        raise HTTPException(status_code=400, detail="当前表缺少课程类型字段")
    if payload.row_index >= len(rows):
        raise HTTPException(status_code=400, detail="row_index 超出表格范围")

    row = _normalize_sheet_row(rows[payload.row_index], len(columns))
    course_type = _normalize_sheet_text(row[type_index])
    if not course_type:
        raise HTTPException(status_code=400, detail="选中的课程类型为空")
    return course_type


def _extract_attendance_course_script_token(url: Any) -> str:
    text = _normalize_sheet_text(url)
    if not text:
        return ""

    match = ATTENDANCE_COURSE_SCRIPT_KDOCS_TOKEN_RE.search(text)
    if match:
        return match.group("token")

    tail = text.split("?", 1)[0].rstrip("/").rsplit("/", 1)[-1]
    if re.fullmatch(r"[A-Za-z0-9]{8,}", tail):
        return tail
    return ""


def _format_attendance_course_script_prefix(value: date) -> str:
    return f"d{value.year % 100:02d}{value.month:02d}{value.day:02d}"


def _parse_attendance_course_script_stem(stem: str) -> tuple[date, str] | None:
    match = ATTENDANCE_COURSE_SCRIPT_STEM_RE.match(stem)
    if not match:
        return None

    raw_date = match.group("date")
    try:
        parsed_date = date(
            _normalize_two_digit_year(int(raw_date[:2])),
            int(raw_date[2:4]),
            int(raw_date[4:6]),
        )
    except ValueError:
        return None

    body = match.group("body").strip()
    return (parsed_date, body) if body else None


def _parse_attendance_course_text_date(value: Any) -> tuple[date | None, str]:
    text = _normalize_sheet_text(value)
    leading_date = ATTENDANCE_TEMPLATE_LEADING_DATE_RE.match(text)
    if not leading_date:
        chinese_month = re.match(r"^(?P<year>\d{4})年(?P<month>\d{1,2})月(?P<body>.+)$", text)
        if not chinese_month:
            return None, text
        return None, chinese_month.group("body").strip()

    parsed_date = _parse_attendance_date_text(leading_date.group("date"))
    return parsed_date, leading_date.group("body").strip()


def _sanitize_attendance_course_script_body(value: Any) -> str:
    text = _normalize_sheet_text(value)
    text = text.replace("4.5阶", "4点5阶")
    text = ATTENDANCE_COURSE_SCRIPT_INVALID_FILENAME_RE.sub("_", text)
    return text.strip(" .")


def _standardize_attendance_course_script_stem(value: Any) -> str:
    text = _normalize_sheet_text(value)
    if not text:
        return ""

    text = ATTENDANCE_COURSE_SCRIPT_FILE_EXTENSION_RE.sub("", text)
    text = re.sub(r"^\d{2}(\d{6})", r"d\1", text, count=1)
    text = re.sub(r"^(\d{6})(?=\D|$)", r"d\1", text, count=1)
    text = text.replace(".", "点")
    text = ATTENDANCE_COURSE_SCRIPT_INVALID_FILENAME_RE.sub("_", text)
    return text.strip(" .")


def _get_document_cell_link_url(document_json: dict[str, Any], row_index: int, column_index: int) -> str:
    data_start_row = _normalize_document_data_start_row(document_json)
    rows = _extract_document_rows(document_json)
    if 0 <= row_index < len(rows):
        row = _normalize_sheet_row(rows[row_index], len(_normalize_document_columns(document_json)))
        if 0 <= column_index < len(row):
            inline_url = _inline_cell_link_url(row[column_index])
            if inline_url:
                return inline_url
    grid_rows = _extract_document_grid_rows(document_json)
    document_row_index = data_start_row + row_index
    if 0 <= document_row_index < len(grid_rows):
        row = _normalize_sheet_row(grid_rows[document_row_index], len(_normalize_document_columns(document_json)))
        if 0 <= column_index < len(row):
            return _inline_cell_link_url(row[column_index])
    return ""


def _attendance_course_script_dir_candidates() -> list[Path]:
    configured_dir = os.fspath(ATTENDANCE_COURSE_SCRIPT_DIR) != os.fspath(ATTENDANCE_COURSE_SCRIPT_DIR_DEFAULT)
    if configured_dir and ATTENDANCE_COURSE_SCRIPT_DIR.exists() and ATTENDANCE_COURSE_SCRIPT_DIR.is_dir():
        return [ATTENDANCE_COURSE_SCRIPT_DIR]

    candidate_paths = [
        ATTENDANCE_COURSE_SCRIPT_DIR,
        ATTENDANCE_XLPROJECT_SRC_DIR / "xlsln" / "kq5034" / "courses",
        Path(__file__).resolve().parents[3] / "xlproject" / "src" / "xlsln" / "kq5034" / "courses",
    ]
    candidates: list[Path] = []
    seen: set[str] = set()
    for path in candidate_paths:
        try:
            key = os.fspath(path.resolve())
        except OSError:
            key = os.fspath(path)
        if key in seen:
            continue
        seen.add(key)
        candidates.append(path)
    return candidates


def _get_attendance_course_script_dir() -> Path:
    return next(
        (path for path in _attendance_course_script_dir_candidates() if path.exists() and path.is_dir()),
        ATTENDANCE_COURSE_SCRIPT_DIR,
    )


def _iter_attendance_course_script_files() -> list[Path]:
    paths: list[Path] = []
    for base_dir in _attendance_course_script_dir_candidates():
        for directory in (base_dir, base_dir / "已完结"):
            if not directory.exists() or not directory.is_dir():
                continue
            paths.extend(path for path in directory.glob("*.py") if _parse_attendance_course_script_stem(path.stem))
    return paths


def _find_attendance_course_script_by_stem(stem: str) -> Path | None:
    filename = f"{stem}.py"
    for base_dir in _attendance_course_script_dir_candidates():
        for directory in (base_dir, base_dir / "已完结"):
            path = directory / filename
            if path.exists() and path.is_file():
                return path
    return None


def _normalize_attendance_zen_stage_tokens(value: str) -> set[str]:
    normalized = value.replace("点", ".")
    tokens: set[str] = set()
    if "4.5阶" in normalized:
        tokens.update({"4.5阶", "4点5阶"})
    for token in ("一阶", "二阶", "三阶", "四阶", "五阶", "二三阶"):
        if token in value:
            tokens.add(token)
    return tokens


def _attendance_course_script_matches_type(course_type: str, script_body: str) -> bool:
    normalized_type = _normalize_sheet_text(course_type)
    normalized_body = _normalize_sheet_text(script_body)
    if not normalized_type or not normalized_body:
        return False

    if "禅宗" in normalized_type:
        if "禅宗" not in normalized_body:
            return False
        stage_tokens = _normalize_attendance_zen_stage_tokens(normalized_type)
        if not stage_tokens:
            return True
        return any(token in normalized_body for token in stage_tokens)

    if normalized_type == "念住":
        return "念住" in normalized_body and "闯关" not in normalized_body
    if normalized_type == "念住闯关":
        return "念住闯关" in normalized_body
    if normalized_type == "觉观":
        return "觉观" in normalized_body
    if normalized_type in {"梵呗初阶", "梵呗增益"}:
        return normalized_type in normalized_body
    return normalized_type in normalized_body


def _find_attendance_course_script_template(
    course_type: str,
    *,
    target_date: date,
    target_stem: str,
    script_files: list[Path] | None = None,
) -> Path | None:
    candidates: list[tuple[int, int, str, Path]] = []
    for path in script_files if script_files is not None else _iter_attendance_course_script_files():
        parsed = _parse_attendance_course_script_stem(path.stem)
        if parsed is None or path.stem == target_stem:
            continue
        script_date, script_body = parsed
        if not _attendance_course_script_matches_type(course_type, script_body):
            continue

        date_delta = abs((script_date - target_date).days)
        future_penalty = 1 if script_date > target_date else 0
        candidates.append((date_delta, future_penalty, path.name, path))

    if not candidates:
        return None
    return min(candidates, key=lambda item: (item[0], item[1], item[2]))[3]


def _build_attendance_course_script_status(
    document_json: dict[str, Any],
    *,
    row_index: int,
    script_files: list[Path] | None = None,
) -> NoteSheetAttendanceCourseScriptStatusItem:
    normalized = _normalize_document_json(document_json)
    columns = _normalize_document_columns(normalized)
    rows = _extract_document_rows(normalized)
    formula_row_offset = _get_formula_reference_row_offset(normalized)
    formula_grid_rows = _extract_document_grid_rows(normalized)
    if row_index < 0 or row_index >= len(rows):
        raise HTTPException(status_code=400, detail="row_index 超出表格范围")

    column_count = len(columns)
    row = _normalize_sheet_row(rows[row_index], column_count)
    type_index = _find_attendance_column_index(columns, "course_type")
    name_index = _find_attendance_column_index(columns, "course_name")
    online_sheet_index = _find_attendance_column_index(columns, "online_sheet")
    start_date_index = _find_attendance_column_index(columns, "start_date")

    if online_sheet_index is None:
        return NoteSheetAttendanceCourseScriptStatusItem(row_index=row_index, reason="当前表缺少在线考勤表字段")

    course_type = _normalize_sheet_text(row[type_index]) if type_index is not None else ""
    course_name = _normalize_sheet_text(row[name_index]) if name_index is not None else ""
    online_sheet = _normalize_sheet_text(row[online_sheet_index])
    url = _get_document_cell_link_url(normalized, row_index, online_sheet_index)

    status = NoteSheetAttendanceCourseScriptStatusItem(
        row_index=row_index,
        course_type=course_type,
        course_name=course_name,
        online_sheet=online_sheet,
        url=url,
    )
    if not course_type:
        status.reason = "课程类型为空"
        return status
    if not online_sheet:
        status.reason = "在线考勤表为空"
        return status
    if not url:
        status.reason = "在线考勤表没有超链接"
        return status

    target_stem = _standardize_attendance_course_script_stem(online_sheet)
    parsed_target_stem = _parse_attendance_course_script_stem(target_stem)
    target_date = parsed_target_stem[0] if parsed_target_stem is not None else None

    if parsed_target_stem is None:
        text_date, text_body = _parse_attendance_course_text_date(online_sheet)
        target_date = text_date
        if target_date is None and start_date_index is not None:
            target_date = _extract_attendance_row_start_date(
                row,
                row_index=row_index,
                columns=columns,
                rows=rows,
                start_date_index=start_date_index,
                reference_row_offset=formula_row_offset,
                grid_rows=formula_grid_rows,
            )
        if target_date is None:
            status.reason = "无法识别课程日期"
            return status

        script_body = _sanitize_attendance_course_script_body(text_body or course_name)
        if not script_body:
            status.reason = "无法识别脚本名称"
            return status

        target_stem = f"{_format_attendance_course_script_prefix(target_date)}{script_body}"

    status.target_stem = target_stem
    status.target_filename = f"{target_stem}.py"

    existing = _find_attendance_course_script_by_stem(target_stem)
    if existing is not None:
        status.exists = True
        status.existing_path = str(existing)
        status.reason = "脚本已存在"
        return status

    source = _find_attendance_course_script_template(
        course_type,
        target_date=target_date,
        target_stem=target_stem,
        script_files=script_files,
    )
    if source is None:
        status.reason = "没有找到同类型脚本模板"
        return status

    status.can_generate = True
    return status


def _load_attendance_course_link_provider():
    for path in (ATTENDANCE_XLPROJECT_SRC_DIR, ATTENDANCE_KQ5034_REPO_DIR):
        path_text = str(path)
        if path.exists() and path_text not in sys.path:
            sys.path.insert(0, path_text)

    try:
        from kqmain import 获取课程链接  # type: ignore

        return 获取课程链接
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"无法加载考勤链接查询工具：{exc}") from exc


def _load_attendance_kqdb_provider():
    try:
        from kq5034.attendance_api import get_kqdb  # type: ignore

        return get_kqdb
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"无法加载考勤数据库配置：{exc}") from exc


def _resolve_attendance_course_lookup_name(
    document_json: dict[str, Any],
    *,
    row: list[Any],
    row_index: int,
    columns: list[Any],
) -> str:
    normalized = _normalize_document_json(document_json)
    name_index = _find_attendance_column_index(columns, "course_name")
    online_sheet_index = _find_attendance_column_index(columns, "online_sheet")
    start_date_index = _find_attendance_column_index(columns, "start_date")

    course_name = _normalize_sheet_text(row[name_index]) if name_index is not None else ""
    online_sheet = _normalize_sheet_text(row[online_sheet_index]) if online_sheet_index is not None else ""

    standardized_stem = _standardize_attendance_course_script_stem(online_sheet)
    if _parse_attendance_course_script_stem(standardized_stem) is not None:
        return standardized_stem

    text_date, text_body = _parse_attendance_course_text_date(online_sheet)
    target_date = text_date
    if target_date is None and start_date_index is not None:
        rows = _extract_document_rows(normalized)
        target_date = _extract_attendance_row_start_date(
            row,
            row_index=row_index,
            columns=columns,
            rows=rows,
            start_date_index=start_date_index,
            reference_row_offset=_get_formula_reference_row_offset(normalized),
            grid_rows=_extract_document_grid_rows(normalized),
        )
    body = _sanitize_attendance_course_script_body(text_body or course_name)
    if target_date is not None and body:
        return f"{_format_attendance_course_script_prefix(target_date)}{body}"
    return _sanitize_attendance_course_script_body(online_sheet or course_name)


def _extract_attendance_course_script_attendance_url(source_text: str) -> str:
    codeyun_match = ATTENDANCE_COURSE_SCRIPT_CODEYUN_ATTENDANCE_REF_RE.search(source_text)
    if codeyun_match:
        return (
            f"/workbook/{int(codeyun_match.group('workbook_id'))}"
            f"?sheet={int(codeyun_match.group('sheet_id'))}"
        )

    init_token_match = ATTENDANCE_COURSE_SCRIPT_INIT_TOKEN_RE.search(source_text)
    if init_token_match:
        token = init_token_match.group("token").strip()
        return f"https://www.kdocs.cn/l/{token}" if token else ""

    kdocs_match = ATTENDANCE_COURSE_SCRIPT_KDOCS_TOKEN_RE.search(source_text)
    if kdocs_match:
        return f"https://www.kdocs.cn/l/{kdocs_match.group('token')}"
    return ""


def _resolve_attendance_summary_online_sheet_url(
    document_json: dict[str, Any],
    *,
    row: list[Any],
    row_index: int,
    columns: list[Any],
) -> str:
    lookup_name = _resolve_attendance_course_lookup_name(
        document_json,
        row=row,
        row_index=row_index,
        columns=columns,
    )
    if not lookup_name:
        return ""

    script_path = _find_attendance_course_script_by_stem(lookup_name)
    if script_path is None:
        return ""

    try:
        source_text = script_path.read_text(encoding="utf-8")
    except OSError:
        return ""
    return _extract_attendance_course_script_attendance_url(source_text)


def _repair_attendance_summary_online_sheet_links(document_json: dict[str, Any]) -> tuple[dict[str, Any], int]:
    normalized = _normalize_document_json(document_json)
    columns = _normalize_document_columns(normalized)
    rows = _extract_document_rows(normalized)
    online_sheet_index = _find_attendance_column_index(columns, "online_sheet")
    if online_sheet_index is None:
        return normalized, 0

    next_rows = list(rows)
    repaired = 0
    for row_index, source_row in enumerate(rows):
        row = _normalize_sheet_row(source_row, len(columns))
        online_sheet = _normalize_sheet_text(row[online_sheet_index])
        if not online_sheet:
            continue
        if _get_document_cell_link_url(normalized, row_index, online_sheet_index):
            continue

        url = _resolve_attendance_summary_online_sheet_url(
            normalized,
            row=row,
            row_index=row_index,
            columns=columns,
        )
        if not url:
            continue

        next_rows[row_index] = _set_row_cell_value(
            source_row,
            columns,
            online_sheet_index,
            _with_inline_cell_link(online_sheet, url),
        )
        repaired += 1

    if not repaired:
        return normalized, 0
    return _normalize_document_json(_replace_document_data_rows(normalized, next_rows)), repaired


def _preserve_attendance_summary_online_sheet_links(
    current_document: dict[str, Any],
    incoming_document: dict[str, Any],
) -> tuple[dict[str, Any], int]:
    normalized_current = _normalize_document_json(current_document)
    normalized_incoming = _normalize_document_json(incoming_document)
    current_columns = _normalize_document_columns(normalized_current)
    incoming_columns = _normalize_document_columns(normalized_incoming)
    current_online_index = _find_attendance_column_index(current_columns, "online_sheet")
    incoming_online_index = _find_attendance_column_index(incoming_columns, "online_sheet")
    if current_online_index is None or incoming_online_index is None:
        return normalized_incoming, 0

    current_rows = _extract_document_rows(normalized_current)
    incoming_rows = _extract_document_rows(normalized_incoming)
    next_rows = list(incoming_rows)
    preserved = 0
    for row_index, incoming_source_row in enumerate(incoming_rows):
        if row_index >= len(current_rows):
            continue
        current_row = _normalize_sheet_row(current_rows[row_index], len(current_columns))
        incoming_row = _normalize_sheet_row(incoming_source_row, len(incoming_columns))
        current_link = note_sheet_inline_links.inline_cell_link(current_row[current_online_index])
        if not current_link or note_sheet_inline_links.inline_cell_link(incoming_row[incoming_online_index]):
            continue

        current_value = _normalize_sheet_text(current_row[current_online_index])
        incoming_value = _normalize_sheet_text(incoming_row[incoming_online_index])
        if not current_value or current_value != incoming_value:
            continue

        next_rows[row_index] = _set_row_cell_value(
            incoming_source_row,
            incoming_columns,
            incoming_online_index,
            note_sheet_inline_links.with_inline_cell_link(incoming_value, current_link),
        )
        preserved += 1

    if not preserved:
        return normalized_incoming, 0
    return _normalize_document_json(_replace_document_data_rows(normalized_incoming, next_rows)), preserved


def _extract_attendance_link_item_url(item: Any) -> str:
    if isinstance(item, dict):
        return _normalize_sheet_text(item.get("url"))
    if isinstance(item, (list, tuple)) and len(item) >= 2:
        return _normalize_sheet_text(item[1])
    return ""


def _format_attendance_link_count_value(total_count: int, linked_count: int) -> str:
    if total_count <= 0:
        return "0"
    return str(total_count) if linked_count >= total_count else f"{linked_count}/{total_count}"


def _query_attendance_link_count(
    field_key: Literal["lesson_links", "clockin_links"],
    lookup_name: str,
) -> tuple[int, int]:
    try:
        xldb = _load_attendance_kqdb_provider()()
        if field_key == "lesson_links":
            records = xldb.exec2dict(
                "SELECT lesson_id2 AS url FROM lesson_table WHERE lesson_name LIKE %s ORDER BY lesson_id",
                [f"{lookup_name}-%"],
            )
        else:
            records = xldb.exec2dict(
                "SELECT url FROM clockin_table WHERE name LIKE %s ORDER BY clockin_id",
                [f"{lookup_name}-%"],
            )
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"查询考勤链接失败：{exc}") from exc

    items = list(records or [])
    total_count = len(items)
    linked_count = sum(1 for item in items if _extract_attendance_link_item_url(item))
    return total_count, linked_count


def _attendance_summary_course_type_for_row(row: list[Any], columns: list[Any]) -> str:
    type_index = _find_attendance_column_index(columns, "course_type")
    return _normalize_sheet_text(row[type_index]) if type_index is not None and type_index < len(row) else ""


def _can_remote_repair_nianzhu_jueguan_links(course_type: str, lookup_name: str) -> bool:
    normalized_type = _normalize_sheet_text(course_type)
    normalized_name = _normalize_sheet_text(lookup_name)
    return normalized_type in {"念住", "觉观"} or "念住" in normalized_name or "觉观" in normalized_name


def _remote_device_error_detail(response: Any) -> str:
    try:
        detail = response.json().get("detail")
    except Exception:
        detail = getattr(response, "text", "").strip()
    return str(detail or f"远程执行失败，HTTP {getattr(response, 'status_code', '')}").strip()


def _post_remote_attendance_device_json(entry: UserDevice, *, path: str, payload: dict[str, Any], timeout: int) -> dict[str, Any]:
    server_url = _normalize_sheet_text(entry.server_url).rstrip("/")
    token = _normalize_sheet_text(entry.token)
    if not server_url or not token:
        raise RuntimeError("远程课程数据浏览器设备缺少后端地址或访问令牌")
    try:
        import requests

        with requests.Session() as request_session:
            request_session.trust_env = False
            response = request_session.post(
                f"{server_url}{path}",
                json=payload,
                headers={
                    "Authorization": f"Bearer {token}",
                    "X-Device-Token": token,
                },
                timeout=timeout,
            )
    except Exception as exc:
        raise RuntimeError(f"调用远程课程数据浏览器设备失败：{exc}") from exc
    if response.status_code >= 400:
        raise RuntimeError(_remote_device_error_detail(response))
    try:
        data = response.json()
    except Exception as exc:
        raise RuntimeError("远程课程数据浏览器设备返回了无法解析的响应") from exc
    return dict(data) if isinstance(data, dict) else {"result": data}


def _repair_nianzhu_jueguan_links_with_remote_step1(
    session: Session,
    *,
    course_type: str,
    lookup_name: str,
    field_key: Literal["lesson_links", "clockin_links"],
) -> dict[str, Any]:
    if not _can_remote_repair_nianzhu_jueguan_links(course_type, lookup_name):
        return {"attempted": False, "reason": "课程类型不需要念住/觉观远程补抓"}

    config = get_or_create_attendance_service_config(session)
    entry = get_attendance_course_data_step_runner_device(
        session,
        config,
        step_number=1,
        data_flow_config=get_attendance_course_data_flow_config(session),
    )
    if entry is None:
        return {"attempted": False, "reason": "课程数据 step1 未配置浏览器设备"}
    if _normalize_sheet_text(entry.mode) != "remote":
        return {
            "attempted": False,
            "reason": f"课程数据 step1 当前是本机设备 {entry.name or entry.device_id}，已跳过本机爬虫",
        }

    payload = {
        "course_name": lookup_name,
        "shop_id": 1,
        "update_lessons": field_key == "lesson_links",
        "update_clockins": field_key == "clockin_links",
        "clockin_pattern": "",
        "dynamic_clockin_plugin": "",
        "close_browser": True,
    }
    result = _post_remote_attendance_device_json(
        entry,
        path="/api/device-control/attendance/nianzhu/step1",
        payload=payload,
        timeout=1200,
    )
    return {
        "attempted": True,
        "device_entry_id": entry.entry_id,
        "device_name": entry.name,
        "device_id": entry.device_id,
        "result": result,
    }


def _update_attendance_link_counts(
    session: Session,
    document_json: dict[str, Any],
    *,
    field_key: Literal["lesson_links", "clockin_links"],
    row_index: int | None = None,
    repair_with_remote_browser: bool = True,
) -> tuple[dict[str, Any], list[NoteSheetAttendanceLinkCountUpdateItem], list[NoteSheetAttendanceLinkCountUpdateItem]]:
    normalized = _ensure_attendance_link_count_columns(document_json)
    columns = _normalize_document_columns(normalized)
    rows = _extract_document_rows(normalized)
    target_column_index = _find_column_index(columns, ATTENDANCE_FIELD_BINDINGS[field_key][0])
    if target_column_index is None:
        raise HTTPException(status_code=400, detail="当前表缺少链接统计字段")

    if row_index is not None and (row_index < 0 or row_index >= len(rows)):
        raise HTTPException(status_code=400, detail="row_index 超出表格范围")

    row_indices = [row_index] if row_index is not None else list(range(len(rows)))
    next_rows = list(rows)
    updated: list[NoteSheetAttendanceLinkCountUpdateItem] = []
    skipped: list[NoteSheetAttendanceLinkCountUpdateItem] = []

    for source_row_index in row_indices:
        if source_row_index is None:
            continue
        row = _normalize_sheet_row(next_rows[source_row_index], len(columns))
        lookup_name = _resolve_attendance_course_lookup_name(
            normalized,
            row=row,
            row_index=source_row_index,
            columns=columns,
        )
        course_name_index = _find_attendance_column_index(columns, "course_name")
        course_name = _normalize_sheet_text(row[course_name_index]) if course_name_index is not None else ""
        course_type = _attendance_summary_course_type_for_row(row, columns)
        item = NoteSheetAttendanceLinkCountUpdateItem(
            row_index=source_row_index,
            course_name=course_name,
            lookup_name=lookup_name,
        )
        if not lookup_name:
            item.reason = "课程名为空"
            skipped.append(item)
            continue

        total_count, linked_count = _query_attendance_link_count(field_key, lookup_name)
        remote_repair_error = False
        if repair_with_remote_browser and (total_count <= 0 or linked_count < total_count):
            try:
                repair_summary = _repair_nianzhu_jueguan_links_with_remote_step1(
                    session,
                    course_type=course_type,
                    lookup_name=lookup_name,
                    field_key=field_key,
                )
            except Exception as exc:
                repair_summary = {"attempted": True, "error": str(exc)}
            item.remote_repair_attempted = bool(repair_summary.get("attempted"))
            item.remote_repair_summary = repair_summary
            remote_repair_error = bool(repair_summary.get("attempted") and repair_summary.get("error"))
            if repair_summary.get("attempted") and not remote_repair_error:
                total_count, linked_count = _query_attendance_link_count(field_key, lookup_name)
        if remote_repair_error:
            item.reason = "远程补抓失败，已保留原链接数"
            skipped.append(item)
            continue
        value = _format_attendance_link_count_value(total_count, linked_count)
        item.total_count = total_count
        item.linked_count = linked_count
        item.value = value
        next_rows[source_row_index] = _set_row_cell_value(
            row,
            columns,
            target_column_index,
            value,
        )
        updated.append(item)

    return _replace_document_data_rows({
        **normalized,
        "columns": columns,
    }, next_rows), updated, skipped


def _list_attendance_course_script_statuses(document_json: dict[str, Any]) -> list[NoteSheetAttendanceCourseScriptStatusItem]:
    normalized = _normalize_document_json(document_json)
    rows = _extract_document_rows(normalized)
    script_files = _iter_attendance_course_script_files()
    statuses: list[NoteSheetAttendanceCourseScriptStatusItem] = []
    for row_index in range(len(rows)):
        status = _build_attendance_course_script_status(
            normalized,
            row_index=row_index,
            script_files=script_files,
        )
        if status.online_sheet or status.url or status.target_filename:
            statuses.append(status)
    return statuses


def _replace_attendance_course_script_token(source_text: str, next_token: str) -> str:
    match = ATTENDANCE_COURSE_SCRIPT_INIT_TOKEN_RE.search(source_text)
    if not match:
        raise HTTPException(status_code=400, detail="脚本模板中没有找到在线表格 token 参数")

    old_token = match.group("token")

    def replace(match_obj: re.Match[str]) -> str:
        return f"{match_obj.group(1)}{match_obj.group('quote')}{next_token}{match_obj.group('quote')}"

    next_text = ATTENDANCE_COURSE_SCRIPT_INIT_TOKEN_RE.sub(replace, source_text, count=1)
    if old_token and old_token != next_token:
        next_text = next_text.replace(old_token, next_token)
    return next_text


def _extract_attendance_course_edition(value: Any) -> str:
    match = ATTENDANCE_COURSE_EDITION_RE.search(_normalize_sheet_text(value))
    return match.group("edition") if match else ""


def _replace_attendance_course_script_product_edition(source_text: str, *target_texts: Any) -> str:
    next_edition = next(
        (edition for edition in (_extract_attendance_course_edition(text) for text in target_texts) if edition),
        "",
    )
    if not next_edition:
        return source_text

    def replace(match_obj: re.Match[str]) -> str:
        product_name = match_obj.group("value")
        if not ATTENDANCE_COURSE_EDITION_RE.search(product_name):
            return match_obj.group(0)

        next_product_name = ATTENDANCE_COURSE_EDITION_RE.sub(
            f"第{next_edition}届",
            product_name,
            count=1,
        )
        return f"{match_obj.group('prefix')}{match_obj.group('quote')}{next_product_name}{match_obj.group('quote')}"

    return ATTENDANCE_COURSE_SCRIPT_PRODUCT_NAME_RE.sub(replace, source_text, count=1)


def _generate_attendance_course_script(
    document_json: dict[str, Any],
    *,
    row_index: int,
) -> NoteSheetAttendanceCourseScriptGenerationResponse:
    course_script_dir = _get_attendance_course_script_dir()
    if not course_script_dir.exists() or not course_script_dir.is_dir():
        raise HTTPException(status_code=400, detail="考勤脚本目录不存在")

    script_files = _iter_attendance_course_script_files()
    status = _build_attendance_course_script_status(
        document_json,
        row_index=row_index,
        script_files=script_files,
    )
    if status.exists:
        raise HTTPException(status_code=409, detail="对应 py 脚本已存在")
    if not status.can_generate or not status.target_stem:
        raise HTTPException(status_code=400, detail=status.reason or "无法生成 py 脚本")

    target_date, _target_body = _parse_attendance_course_script_stem(status.target_stem) or (None, "")
    if target_date is None:
        raise HTTPException(status_code=400, detail="无法识别目标脚本日期")

    source_path = _find_attendance_course_script_template(
        status.course_type,
        target_date=target_date,
        target_stem=status.target_stem,
        script_files=script_files,
    )
    if source_path is None:
        raise HTTPException(status_code=400, detail="没有找到同类型脚本模板")

    next_token = _extract_attendance_course_script_token(status.url)
    if not next_token:
        raise HTTPException(status_code=400, detail="在线考勤表链接不是可识别的 KDocs 链接")

    target_path = course_script_dir / status.target_filename
    if target_path.exists():
        raise HTTPException(status_code=409, detail="对应 py 脚本已存在")

    try:
        source_text = source_path.read_text(encoding="utf-8")
        target_text = _replace_attendance_course_script_token(source_text, next_token)
        target_text = _replace_attendance_course_script_product_edition(
            target_text,
            status.target_stem,
            status.course_name,
            status.online_sheet,
        )
        target_path.write_text(target_text, encoding="utf-8")
        shutil.copystat(source_path, target_path)
    except OSError as exc:
        raise HTTPException(status_code=500, detail=f"写入 py 脚本失败：{exc}") from exc

    status.exists = True
    status.can_generate = False
    status.existing_path = str(target_path)
    status.reason = "脚本已生成"
    return NoteSheetAttendanceCourseScriptGenerationResponse(
        status=status,
        source_filename=source_path.name,
        source_path=str(source_path),
        created_path=str(target_path),
    )


def _is_attendance_script_in_directory(path: Path, directory: Path) -> bool:
    try:
        return path.resolve().parent == directory.resolve()
    except OSError:
        return path.parent == directory


def _organize_attendance_course_scripts(
    document_json: dict[str, Any],
) -> NoteSheetAttendanceCourseScriptOrganizeResponse:
    course_script_dir = _get_attendance_course_script_dir()
    if not course_script_dir.exists() or not course_script_dir.is_dir():
        raise HTTPException(status_code=400, detail="考勤脚本目录不存在")

    normalized = _normalize_document_json(document_json)
    columns = _normalize_document_columns(normalized)
    rows = _extract_document_rows(normalized)
    completed_index = _find_attendance_column_index(columns, "completed_date")
    if completed_index is None:
        raise HTTPException(status_code=400, detail="当前表缺少考勤实际完成结点字段")

    completed_dir = course_script_dir / "已完结"
    moved: list[NoteSheetAttendanceCourseScriptOrganizeItem] = []
    skipped: list[NoteSheetAttendanceCourseScriptOrganizeItem] = []

    for row_index, raw_row in enumerate(rows):
        row = _normalize_sheet_row(raw_row, len(columns))
        completed = bool(_normalize_sheet_text(row[completed_index]))
        status = _build_attendance_course_script_status(normalized, row_index=row_index)
        if not status.target_filename:
            continue

        item = NoteSheetAttendanceCourseScriptOrganizeItem(
            row_index=row_index,
            course_type=status.course_type,
            online_sheet=status.online_sheet,
            target_filename=status.target_filename,
            completed=completed,
        )
        if not status.exists or not status.existing_path:
            item.reason = "脚本不存在"
            skipped.append(item)
            continue

        source_path = Path(status.existing_path)
        desired_dir = completed_dir if completed else course_script_dir
        target_path = desired_dir / status.target_filename
        item.source_path = str(source_path)
        item.target_path = str(target_path)

        if _is_attendance_script_in_directory(source_path, desired_dir):
            item.reason = "位置已正确"
            skipped.append(item)
            continue
        if not source_path.exists() or not source_path.is_file():
            item.reason = "脚本不存在"
            skipped.append(item)
            continue
        if target_path.exists():
            item.reason = "目标位置已存在"
            skipped.append(item)
            continue

        try:
            desired_dir.mkdir(parents=True, exist_ok=True)
            shutil.move(str(source_path), str(target_path))
        except OSError as exc:
            item.reason = f"移动失败：{exc}"
            skipped.append(item)
            continue

        moved.append(item)

    return NoteSheetAttendanceCourseScriptOrganizeResponse(moved=moved, skipped=skipped)


def _natural_sort_key(value: str) -> tuple[tuple[int, Any], ...]:
    normalized = value.casefold()
    parts = NATURAL_SORT_SPLIT_RE.split(normalized)
    key: list[tuple[int, Any]] = []
    for part in parts:
        if not part:
            continue
        if part.isdigit():
            key.append((0, int(part)))
        else:
            key.append((1, part))
    return tuple(key)


def _get_column_config(document_json: dict[str, Any], column_index: int, columns: list[Any]) -> dict[str, Any]:
    column_configs = document_json.get("column_configs")
    if not isinstance(column_configs, dict) or column_index < 0 or column_index >= len(columns):
        return {}
    config = column_configs.get(str(columns[column_index]))
    return dict(config) if isinstance(config, dict) else {}


def _get_column_value_type(config: dict[str, Any]) -> str:
    value_type = config.get("value_type")
    if value_type in {"multi_text", "number", "percent", "date", "phone"}:
        return str(value_type)
    return "text"


def _extract_sort_cell_value(
    row: Any,
    *,
    row_index: int,
    column_index: int,
    columns: list[Any],
    rows: list[Any],
    reference_row_offset: int = 0,
    grid_rows: list[Any] | None = None,
) -> Any:
    raw_value = _extract_row_cell_value(row, column_index, columns)
    return _evaluate_formula_sort_value(
        raw_value,
        row_index=row_index,
        column_index=column_index,
        rows=rows,
        columns=columns,
        reference_row_offset=reference_row_offset,
        grid_rows=grid_rows,
    )


def _sort_value_is_empty(value: Any) -> bool:
    return value is None or str(value).strip() == ""


def _build_sort_key(value: Any, value_type: str) -> tuple[int, Any]:
    if value_type == "date":
        date_value = _parse_date_sort_value(value)
        if date_value is not None:
            return 0, date_value
    if value_type == "number":
        number_value = _parse_number_sort_value(value)
        if number_value is not None:
            return 0, number_value
    if value_type == "percent":
        percent_value = _parse_percent_sort_value(value)
        if percent_value is not None:
            return 0, percent_value
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return 0, float(value)
    return 1, _natural_sort_key("" if value is None else str(value).strip())


def _parse_cell_meta_key(key: Any) -> tuple[int, int] | None:
    if not isinstance(key, str):
        return None
    row_text, separator, column_text = key.partition(":")
    if separator != ":" or not row_text.isdigit() or not column_text.isdigit():
        return None
    return int(row_text), int(column_text)


def _remap_cell_meta_rows(cell_meta: Any, row_index_map: dict[int, int], *, row_offset: int = 0) -> dict[str, Any]:
    if not isinstance(cell_meta, dict):
        return {}

    remapped: dict[str, Any] = {}
    for key, value in cell_meta.items():
        position = _parse_cell_meta_key(key)
        if position is None:
            remapped[str(key)] = value
            continue

        row_index, column_index = position
        lookup_row_index = row_index - row_offset
        next_row_index = row_index_map.get(lookup_row_index)
        if next_row_index is None:
            remapped[str(key)] = value
        else:
            remapped[f"{next_row_index + row_offset}:{column_index}"] = value
    return remapped


def _sort_sheet_document_rows(
    document_json: dict[str, Any],
    *,
    column_index: int,
    direction: Literal["asc", "desc"],
) -> dict[str, Any]:
    normalized = _normalize_document_json(document_json)
    all_rows = _extract_document_rows(normalized)
    columns = list(normalized.get("columns") or [])
    data_start_row = _normalize_document_data_start_row(normalized)
    formula_row_offset = _get_formula_reference_row_offset(normalized)
    formula_grid_rows = _extract_document_grid_rows(normalized)
    merged_cells = normalized.get("merged_cells")
    if isinstance(merged_cells, list):
        for cell in merged_cells:
            if not isinstance(cell, dict):
                continue
            if int(cell.get("row") or 0) >= data_start_row and int(cell.get("rowspan") or 1) > 1:
                raise HTTPException(status_code=400, detail="数据区存在跨行合并，不能排序")
    column_config = _get_column_config(normalized, column_index, columns)
    column_value_type = _get_column_value_type(column_config)

    sortable_rows: list[tuple[int, Any]] = []
    empty_value_rows: list[tuple[int, Any]] = []
    blank_rows: list[tuple[int, Any]] = []

    for row_index, row in enumerate(all_rows):
        if isinstance(row, list):
            row_values = row
        elif isinstance(row, dict):
            row_values = list(row.values())
        else:
            row_values = []
        if not any(str(cell or "").strip() for cell in row_values):
            blank_rows.append((row_index, row))
            continue

        cell_value = _extract_sort_cell_value(
            row,
            row_index=row_index,
            column_index=column_index,
            columns=columns,
            rows=all_rows,
            reference_row_offset=formula_row_offset,
            grid_rows=formula_grid_rows,
        )
        if _sort_value_is_empty(cell_value):
            empty_value_rows.append((row_index, row))
        else:
            sortable_rows.append((row_index, row))

    sorted_rows = sorted(
        sortable_rows,
        key=lambda item: _build_sort_key(
            _extract_sort_cell_value(
                item[1],
                row_index=item[0],
                column_index=column_index,
                columns=columns,
                rows=all_rows,
                reference_row_offset=formula_row_offset,
                grid_rows=formula_grid_rows,
            ),
            column_value_type,
        ),
        reverse=direction == "desc",
    )
    ordered_rows = [
        *sorted_rows,
        *empty_value_rows,
        *blank_rows,
    ]
    row_index_map = {
        source_index: target_index
        for target_index, (source_index, _row) in enumerate(ordered_rows)
    }

    next_rows = [
        _remap_row_formula_cell_references(
            row,
            columns=columns,
            row_index_map=row_index_map,
            row_index_offset=formula_row_offset,
        )
        for _source_index, row in ordered_rows
    ]
    next_document = _replace_document_data_rows(normalized, next_rows)
    if isinstance(normalized.get("cell_meta"), dict):
        next_document["cell_meta"] = _remap_cell_meta_rows(
            normalized.get("cell_meta"),
            row_index_map,
            row_offset=_normalize_document_data_start_row(normalized),
        )
    return next_document


def _ensure_workbook_identity(session: Session, workbook: WorkbookDocument) -> int:
    return ensure_resource_identity(
        session,
        RESOURCE_TYPE_WORKBOOK,
        str(getattr(workbook, "legacy_id", None) or workbook.id),
        None,
    )


def _require_sheet_numeric_id(document: SheetDocument) -> int:
    numeric_id = int(document.numeric_id or 0)
    if numeric_id <= 0:
        raise HTTPException(status_code=500, detail="表格编号缺失")
    return numeric_id


def _require_workbook_numeric_id(workbook: WorkbookDocument) -> int:
    numeric_id = int(workbook.numeric_id or 0)
    if numeric_id <= 0:
        raise HTTPException(status_code=500, detail="工作簿编号缺失")
    return numeric_id


def _sheet_resource_id(document: SheetDocument) -> str:
    return sheet_public_id(document)


def _workbook_resource_id(workbook: WorkbookDocument) -> str:
    return workbook_public_id(workbook)


def _sheet_link_ref(document: SheetDocument) -> str:
    return _sheet_resource_id(document)


def _workbook_link_ref(workbook: WorkbookDocument) -> str:
    return _workbook_resource_id(workbook)


def _normalize_resource_role(value: Any) -> Literal["deny", "viewer", "editor", "manager"] | None:
    role = str(value or "").strip()
    if role in RESOURCE_ACCESS_ROLE_RANK:
        return role  # type: ignore[return-value]
    return None


def _build_resource_access(role: str | None) -> NoteSheetResourceAccess:
    normalized_role = _normalize_resource_role(role) if role else None
    if normalized_role is None:
        return NoteSheetResourceAccess(role="none")

    rank = RESOURCE_ACCESS_ROLE_RANK[normalized_role]
    return NoteSheetResourceAccess(
        role=normalized_role,
        capabilities=NoteSheetAccessCapabilities(
            can_read=rank >= RESOURCE_ACCESS_ROLE_RANK["viewer"],
            can_use_local_view=rank >= RESOURCE_ACCESS_ROLE_RANK["viewer"],
            can_edit_data=rank >= RESOURCE_ACCESS_ROLE_RANK["editor"],
            can_edit_config=rank >= RESOURCE_ACCESS_ROLE_RANK["editor"],
            can_run_sheet_actions=rank >= RESOURCE_ACCESS_ROLE_RANK["editor"],
            can_manage_access=rank >= RESOURCE_ACCESS_ROLE_RANK["manager"],
        ),
    )


def _is_attendance_wjx_data_sheet(document: SheetDocument) -> bool:
    return (
        document.scope == "notes"
        and document.owner_type == ATTENDANCE_WJX_DATA_OWNER_TYPE
        and document.owner_key == ATTENDANCE_WJX_DATA_OWNER_KEY
        and document.sheet_key == ATTENDANCE_WJX_DATA_SHEET_KEY
    )


def _apply_sheet_specific_access_capabilities(
    access: NoteSheetResourceAccess,
    document: SheetDocument,
) -> NoteSheetResourceAccess:
    if (
        access.capabilities.can_read
        and not access.capabilities.can_edit_data
        and _is_attendance_wjx_data_sheet(document)
    ):
        access.capabilities.editable_data_columns = list(ATTENDANCE_WJX_DATA_PUBLIC_EDITABLE_COLUMN_INDEXES)
    return access


def _resource_role_allows(role: str | None, required_role: Literal["viewer", "editor", "manager"]) -> bool:
    normalized_role = _normalize_resource_role(role)
    if normalized_role is None:
        return False
    return RESOURCE_ACCESS_ROLE_RANK[normalized_role] >= RESOURCE_ACCESS_ROLE_RANK[required_role]


def _active_sheet_condition():
    return or_(SheetDocument.deleted_at.is_(None), SheetDocument.deleted_at <= 0)


def _deleted_sheet_condition():
    return SheetDocument.deleted_at > 0


def _sheet_resource_update_room(sheet_id: int) -> str:
    return f"resource:sheet:{sheet_id}"


def _broadcast_sheet_resource_update(document: SheetDocument) -> None:
    message = {
        "type": "resource-updated",
        "resource_type": "sheet",
        "resource_id": str(_require_sheet_numeric_id(document)),
        "version": int(document.version or 1),
        "updated_at": float(document.updated_at or time.time()),
        "updated_by_user_id": document.updated_by_user_id,
    }
    try:
        anyio.from_thread.run(ws_manager.broadcast, _sheet_resource_update_room(_require_sheet_numeric_id(document)), message)
    except RuntimeError:
        # No active AnyIO worker context, which can happen in unit tests or direct script calls.
        pass


def _check_sheet_base_version(document: SheetDocument, base_version: int | None) -> None:
    if base_version is not None and int(base_version) != int(document.version or 1):
        raise HTTPException(status_code=409, detail="表格已被其他人更新，请刷新后重试")


def _active_workbook_condition():
    return or_(WorkbookDocument.deleted_at.is_(None), WorkbookDocument.deleted_at <= 0)


def _deleted_workbook_condition():
    return WorkbookDocument.deleted_at > 0


def _highest_resource_role(left: str | None, right: str | None) -> str | None:
    left_role = _normalize_resource_role(left) if left else None
    right_role = _normalize_resource_role(right) if right else None
    if left_role is None:
        return right_role
    if right_role is None:
        return left_role
    return left_role if RESOURCE_ACCESS_ROLE_RANK[left_role] >= RESOURCE_ACCESS_ROLE_RANK[right_role] else right_role


def _build_resource_subject_key(subject_type: str, subject_user_id: int | None = None) -> str:
    if subject_type == RESOURCE_ACCESS_SUBJECT_ANONYMOUS:
        return RESOURCE_ACCESS_SUBJECT_ANONYMOUS
    if subject_type == RESOURCE_ACCESS_SUBJECT_USER and subject_user_id is not None:
        return f"user:{subject_user_id}"
    raise HTTPException(status_code=400, detail="非法权限主体")


def _current_user_subject_keys(current_user: User | None) -> list[str]:
    if current_user is None:
        return [RESOURCE_ACCESS_SUBJECT_ANONYMOUS]
    return [
        _build_resource_subject_key(RESOURCE_ACCESS_SUBJECT_USER, current_user.id),
        RESOURCE_ACCESS_SUBJECT_ANONYMOUS,
    ]


def _fetch_resource_grants(session: Session, resource_type: str, resource_id: str) -> list[ResourceAccessGrant]:
    return list(session.exec(
        select(ResourceAccessGrant)
        .where(ResourceAccessGrant.resource_type == resource_type)
        .where(ResourceAccessGrant.resource_id == resource_id)
    ).all())


def _delete_resource_access_grants(
    session: Session,
    *,
    resource_type: str,
    resource_ids: list[str],
) -> None:
    if not resource_ids:
        return
    session.exec(
        delete(ResourceAccessGrant)
        .where(ResourceAccessGrant.resource_type == resource_type)
        .where(ResourceAccessGrant.resource_id.in_(resource_ids))
    )


def _get_sheet_ids_owned_only_by_workbook(
    session: Session,
    *,
    workbook_id: str,
    sheet_ids: list[str],
) -> list[str]:
    if not sheet_ids:
        return []

    sheet_map = load_sheets_by_refs(session, sheet_ids)
    unique_sheets = {str(sheet.id): sheet for sheet in sheet_map.values()}.values()
    sheet_refs = sorted({ref for sheet in unique_sheets for ref in sheet_ref_aliases(sheet)})
    if not sheet_refs:
        return []
    workbook_map = load_workbooks_by_refs(session, [workbook_id])
    workbook_refs = {
        ref
        for workbook in {str(workbook.id): workbook for workbook in workbook_map.values()}.values()
        for ref in workbook_ref_aliases(workbook)
    } or {str(workbook_id)}
    links = session.exec(
        select(WorkbookSheetLink).where(WorkbookSheetLink.sheet_id.in_(sheet_refs))
    ).all()
    linked_workbook_map = load_workbooks_by_refs(session, [link.workbook_id for link in links])
    link_counts: dict[str, int] = {}
    target_sheet_ids: set[str] = set()
    for link in links:
        sheet = sheet_map.get(str(link.sheet_id))
        linked_workbook = linked_workbook_map.get(str(link.workbook_id))
        if sheet is None or linked_workbook is None:
            continue
        sheet_id = str(sheet.id)
        link_counts[sheet_id] = link_counts.get(sheet_id, 0) + 1
        if str(link.workbook_id) in workbook_refs:
            target_sheet_ids.add(sheet_id)
    return [
        str(sheet.id)
        for sheet in {str(sheet.id): sheet for sheet in sheet_map.values()}.values()
        if link_counts.get(str(sheet.id), 0) <= 1
        and str(sheet.id) in target_sheet_ids
    ]


def _resolve_subject_grant_role(
    grants: list[ResourceAccessGrant],
    current_user: User | None,
) -> str | None:
    grant_map = {grant.subject_key: grant for grant in grants}
    for subject_key in _current_user_subject_keys(current_user):
        grant = grant_map.get(subject_key)
        if grant is not None:
            return _normalize_resource_role(grant.role)
    return None


def _get_workbooks_for_sheet(session: Session, document: SheetDocument) -> list[WorkbookDocument]:
    links = session.exec(
        select(WorkbookSheetLink).where(WorkbookSheetLink.sheet_id.in_(sheet_ref_aliases(document)))
    ).all()
    workbook_ids = [link.workbook_id for link in links]
    if not workbook_ids:
        return []
    workbook_map = load_workbooks_by_refs(session, workbook_ids)
    return list({str(workbook.id): workbook for workbook in workbook_map.values()}.values())


def _get_workbook_by_numeric_id_or_404(
    session: Session,
    workbook_id: int,
    *,
    include_deleted: bool = False,
) -> WorkbookDocument:
    query = select(WorkbookDocument).where(WorkbookDocument.numeric_id == workbook_id)
    if not include_deleted:
        query = query.where(_active_workbook_condition())
    workbook = session.exec(query).first()
    if workbook is None:
        raise HTTPException(status_code=404, detail="工作簿不存在")
    return workbook


def _get_sheet_by_numeric_id_or_404(
    session: Session,
    sheet_id: int,
    *,
    include_deleted: bool = False,
) -> SheetDocument:
    query = select(SheetDocument).where(SheetDocument.numeric_id == sheet_id)
    if not include_deleted:
        query = query.where(_active_sheet_condition())
    document = session.exec(query).first()
    if document is None or document.scope != "notes":
        raise HTTPException(status_code=404, detail="表格不存在")
    return document


def _is_superuser_or_user_id(current_user: User | None, *user_ids: int | None) -> bool:
    if current_user is None:
        return False
    if current_user.is_superuser:
        return True
    return any(user_id == current_user.id for user_id in user_ids if user_id is not None)


def _is_workbook_resource_principal(current_user: User | None, workbook: WorkbookDocument) -> bool:
    return _is_superuser_or_user_id(
        current_user,
        workbook.owner_user_id,
        workbook.created_by_user_id,
    )


def _sheet_owner_key_user_id(document: SheetDocument) -> int | None:
    if document.owner_type != "user":
        return None
    try:
        return int(str(document.owner_key).strip())
    except (TypeError, ValueError):
        return None


def _is_sheet_resource_principal(current_user: User | None, document: SheetDocument) -> bool:
    return _is_superuser_or_user_id(
        current_user,
        document.owner_user_id,
        document.created_by_user_id,
        _sheet_owner_key_user_id(document),
    )


def _resolve_workbook_resource_access(
    session: Session,
    workbook: WorkbookDocument,
    current_user: User | None,
) -> NoteSheetResourceAccess:
    if _is_workbook_resource_principal(current_user, workbook):
        return _build_resource_access("manager")

    role = _resolve_subject_grant_role(
        _fetch_resource_grants(session, RESOURCE_TYPE_WORKBOOK, _workbook_resource_id(workbook)),
        current_user,
    )
    return _build_resource_access(role)


def _resolve_sheet_resource_access(
    session: Session,
    document: SheetDocument,
    current_user: User | None,
    *,
    workbook: WorkbookDocument | None = None,
) -> NoteSheetResourceAccess:
    if _is_sheet_resource_principal(current_user, document):
        return _apply_sheet_specific_access_capabilities(_build_resource_access("manager"), document)
    if workbook is not None and _is_workbook_resource_principal(current_user, workbook):
        return _apply_sheet_specific_access_capabilities(_build_resource_access("manager"), document)
    if workbook is None and any(
        _is_workbook_resource_principal(current_user, item)
        for item in _get_workbooks_for_sheet(session, document)
    ):
        return _apply_sheet_specific_access_capabilities(_build_resource_access("manager"), document)

    direct_role = _resolve_subject_grant_role(
        _fetch_resource_grants(session, RESOURCE_TYPE_SHEET, _sheet_resource_id(document)),
        current_user,
    )
    if direct_role == "deny":
        return _apply_sheet_specific_access_capabilities(_build_resource_access(direct_role), document)

    inherited_role: str | None = None
    if workbook is not None:
        inherited_role = _resolve_workbook_resource_access(session, workbook, current_user).role
    else:
        for candidate in _get_workbooks_for_sheet(session, document):
            candidate_access = _resolve_workbook_resource_access(session, candidate, current_user)
            inherited_role = _highest_resource_role(inherited_role, candidate_access.role)
    return _apply_sheet_specific_access_capabilities(_build_resource_access(
        _highest_resource_role(direct_role, inherited_role),
    ), document)


def _require_resource_access(
    access: NoteSheetResourceAccess,
    required_role: Literal["viewer", "editor", "manager"],
) -> None:
    if not _resource_role_allows(access.role, required_role):
        raise HTTPException(status_code=403, detail="没有该资源权限")


def _require_note_sheets_feature(session: Session, current_user: User) -> None:
    ensure_feature_access(session, feature_key="notes.sheets", current_user=current_user)


def _get_note_sheet_or_404(
    session: Session,
    current_user: User | None,
    sheet_id: int,
    *,
    required_role: Literal["viewer", "editor", "manager"] = "viewer",
    workbook_id: int | None = None,
) -> tuple[SheetDocument, NoteSheetResourceAccess, WorkbookDocument | None]:
    document = _get_sheet_by_numeric_id_or_404(session, sheet_id)
    workbook = _get_workbook_by_numeric_id_or_404(session, workbook_id) if workbook_id is not None else None
    if workbook is not None:
        link = session.exec(
            select(WorkbookSheetLink)
            .where(WorkbookSheetLink.workbook_id.in_(workbook_ref_aliases(workbook)))
            .where(WorkbookSheetLink.sheet_id.in_(sheet_ref_aliases(document)))
        ).first()
        if link is None:
            raise HTTPException(status_code=404, detail="工作簿中不存在该工作表")

    access = _resolve_sheet_resource_access(session, document, current_user, workbook=workbook)
    _require_resource_access(access, required_role)
    return document, access, workbook


def _get_optional_trusted_device(
    authorization: str | None = Header(default=None),
    x_device_token: str | None = Header(default=None),
    token: str | None = Query(default=None),
    sec_websocket_protocol: str | None = Header(default=None),
) -> Any | None:
    final_token = extract_api_token(
        authorization=authorization,
        x_device_token=x_device_token,
        token=token,
        sec_websocket_protocol=sec_websocket_protocol,
    )
    if not final_token:
        return None

    try:
        return validate_api_token_value(final_token)
    except HTTPException:
        return None


def _get_note_sheet_for_table_or_404(
    session: Session,
    current_user: User | None,
    trusted_device: Any | None,
    sheet_id: int,
    *,
    required_role: Literal["viewer", "editor", "manager"] = "viewer",
    workbook_id: int | None = None,
) -> tuple[SheetDocument, NoteSheetResourceAccess, WorkbookDocument | None]:
    if trusted_device is None:
        return _get_note_sheet_or_404(
            session,
            current_user,
            sheet_id,
            required_role=required_role,
            workbook_id=workbook_id,
        )

    document = _get_sheet_by_numeric_id_or_404(session, sheet_id)
    workbook = _get_workbook_by_numeric_id_or_404(session, workbook_id) if workbook_id is not None else None
    if workbook is not None:
        link = session.exec(
            select(WorkbookSheetLink)
            .where(WorkbookSheetLink.workbook_id.in_(workbook_ref_aliases(workbook)))
            .where(WorkbookSheetLink.sheet_id.in_(sheet_ref_aliases(document)))
        ).first()
        if link is None:
            raise HTTPException(status_code=404, detail="工作簿中不存在该工作表")

    access = _apply_sheet_specific_access_capabilities(_build_resource_access("manager"), document)
    _require_resource_access(access, required_role)
    return document, access, workbook


def _get_workbook_or_404(
    session: Session,
    current_user: User | None,
    workbook_id: int,
    *,
    required_role: Literal["viewer", "editor", "manager"] = "viewer",
) -> tuple[WorkbookDocument, NoteSheetResourceAccess]:
    workbook = _get_workbook_by_numeric_id_or_404(session, workbook_id)
    access = _resolve_workbook_resource_access(session, workbook, current_user)
    _require_resource_access(access, required_role)
    return workbook, access


def _list_workbook_refs_for_sheet_ids(
    session: Session,
    sheet_ids: list[str],
    current_user: User | None,
    *,
    include_deleted: bool = False,
) -> dict[str, list[WorkbookRefItem]]:
    if not sheet_ids:
        return {}

    sheet_map = load_sheets_by_refs(session, sheet_ids, include_deleted=include_deleted)
    unique_sheets = {str(sheet.id): sheet for sheet in sheet_map.values()}.values()
    sheet_refs = sorted({ref for sheet in unique_sheets for ref in sheet_ref_aliases(sheet)})
    if not sheet_refs:
        return {sheet_id: [] for sheet_id in sheet_ids}
    links = session.exec(
        select(WorkbookSheetLink)
        .where(WorkbookSheetLink.sheet_id.in_(sheet_refs))
        .order_by(WorkbookSheetLink.order_index, WorkbookSheetLink.created_at)
    ).all()
    if not links:
        return {}

    workbook_ids = sorted({link.workbook_id for link in links})
    workbook_map = load_workbooks_by_refs(session, workbook_ids, include_deleted=include_deleted)
    workbook_readable_cache: dict[str, bool] = {}

    result: dict[str, list[WorkbookRefItem]] = {sheet_id: [] for sheet_id in sheet_ids}
    for link in links:
        workbook = workbook_map.get(str(link.workbook_id))
        sheet = sheet_map.get(str(link.sheet_id))
        if workbook is None or sheet is None:
            continue
        workbook_key = str(workbook.id)
        can_read_workbook = workbook_readable_cache.get(workbook_key)
        if can_read_workbook is None:
            can_read_workbook = _resolve_workbook_resource_access(
                session,
                workbook,
                current_user,
            ).capabilities.can_read
            workbook_readable_cache[workbook_key] = can_read_workbook
        if not can_read_workbook:
            continue
        result.setdefault(str(sheet.id), []).append(
            WorkbookRefItem(id=_require_workbook_numeric_id(workbook), title=workbook.title),
        )
    return result


def _list_parent_workbook_ids_for_sheet_ids(
    session: Session,
    sheet_ids: list[str],
    *,
    include_deleted: bool = False,
) -> dict[str, int]:
    if not sheet_ids:
        return {}

    sheet_map = load_sheets_by_refs(session, sheet_ids, include_deleted=include_deleted)
    unique_sheets = {str(sheet.id): sheet for sheet in sheet_map.values()}.values()
    sheet_refs = sorted({ref for sheet in unique_sheets for ref in sheet_ref_aliases(sheet)})
    if not sheet_refs:
        return {}
    links = session.exec(
        select(WorkbookSheetLink)
        .where(WorkbookSheetLink.sheet_id.in_(sheet_refs))
        .order_by(WorkbookSheetLink.order_index, WorkbookSheetLink.created_at)
    ).all()
    if not links:
        return {}

    workbook_map = load_workbooks_by_refs(session, sorted({link.workbook_id for link in links}), include_deleted=include_deleted)
    result: dict[str, int] = {}
    for link in links:
        sheet = sheet_map.get(str(link.sheet_id))
        workbook = workbook_map.get(str(link.workbook_id))
        if sheet is None or workbook is None:
            continue
        result.setdefault(str(sheet.id), _require_workbook_numeric_id(workbook))
    return result


def _get_parent_workbook_id_for_sheet(session: Session, document: SheetDocument) -> int | None:
    return _list_parent_workbook_ids_for_sheet_ids(session, [document.id]).get(str(document.id))


def _get_sheet_workbook_context(
    session: Session,
    document: SheetDocument,
    current_user: User | None,
    *,
    workbook: WorkbookDocument | None = None,
    include_workbook_context: bool = True,
) -> tuple[list[WorkbookRefItem], int | None]:
    if not include_workbook_context:
        if workbook is None:
            return [], None
        numeric_workbook_id = _require_workbook_numeric_id(workbook)
        return [WorkbookRefItem(id=numeric_workbook_id, title=workbook.title)], numeric_workbook_id

    sheet_refs = sheet_ref_aliases(document)
    if not sheet_refs:
        return [], _require_workbook_numeric_id(workbook) if workbook is not None else None

    links = session.exec(
        select(WorkbookSheetLink)
        .where(WorkbookSheetLink.sheet_id.in_(sheet_refs))
        .order_by(WorkbookSheetLink.order_index, WorkbookSheetLink.created_at)
    ).all()
    parent_workbook_id = _require_workbook_numeric_id(workbook) if workbook is not None else None
    if not links:
        return [], parent_workbook_id

    workbook_map = load_workbooks_by_refs(session, sorted({link.workbook_id for link in links}))
    workbook_readable_cache: dict[str, bool] = {}
    workbook_items: list[WorkbookRefItem] = []
    seen_workbook_ids: set[int] = set()
    for link in links:
        linked_workbook = workbook_map.get(str(link.workbook_id))
        if linked_workbook is None:
            continue
        numeric_workbook_id = _require_workbook_numeric_id(linked_workbook)
        if parent_workbook_id is None:
            parent_workbook_id = numeric_workbook_id
        workbook_key = str(linked_workbook.id)
        can_read_workbook = workbook_readable_cache.get(workbook_key)
        if can_read_workbook is None:
            can_read_workbook = _resolve_workbook_resource_access(
                session,
                linked_workbook,
                current_user,
            ).capabilities.can_read
            workbook_readable_cache[workbook_key] = can_read_workbook
        if not can_read_workbook or numeric_workbook_id in seen_workbook_ids:
            continue
        seen_workbook_ids.add(numeric_workbook_id)
        workbook_items.append(WorkbookRefItem(id=numeric_workbook_id, title=linked_workbook.title))
    return workbook_items, parent_workbook_id


def _serialize_sheet_summary(
    document: SheetDocument,
    *,
    workbook_items: list[WorkbookRefItem] | None = None,
    parent_workbook_id: int | None = None,
    access: NoteSheetResourceAccess | None = None,
) -> dict[str, Any]:
    return {
        "id": _require_sheet_numeric_id(document),
        "title": document.title,
        "engine": document.engine,
        "scope": document.scope,
        "version": int(document.version or 1),
        "owner_user_id": document.owner_user_id,
        "created_by_user_id": document.created_by_user_id,
        "updated_by_user_id": document.updated_by_user_id,
        "created_at": float(document.created_at or 0.0),
        "updated_at": float(document.updated_at or 0.0),
        "deleted_at": document.deleted_at,
        "deleted_by_user_id": document.deleted_by_user_id,
        "parent_workbook_id": parent_workbook_id,
        "workbook_items": workbook_items or [],
        "access": access,
    }


def _serialize_sheet_detail(
    document: SheetDocument,
    *,
    workbook_items: list[WorkbookRefItem] | None = None,
    parent_workbook_id: int | None = None,
    document_json: dict[str, Any] | None = None,
    pagination: NoteSheetPaginationResponse | None = None,
    access: NoteSheetResourceAccess | None = None,
) -> dict[str, Any]:
    return {
        **_serialize_sheet_summary(
            document,
            workbook_items=workbook_items,
            parent_workbook_id=parent_workbook_id,
            access=access,
        ),
        "owner_type": document.owner_type,
        "owner_key": document.owner_key,
        "sheet_key": document.sheet_key,
        "version": int(document.version or 1),
        "document_json": dict(document_json if document_json is not None else (document.document_json or {})),
        "pagination": pagination,
    }


def _parse_table_cell_reference(cell: str) -> tuple[int, int]:
    match = A1_CELL_REFERENCE_RE.match(str(cell or ""))
    if not match:
        raise HTTPException(status_code=400, detail=f"无法解析单元格坐标: {cell}")
    column_index = _excel_column_index(match.group("column"))
    if column_index is None:
        raise HTTPException(status_code=400, detail=f"无法解析单元格列: {cell}")
    return int(match.group("row")), column_index


def _table_sheet_row_to_data_index(sheet_row: int, data_start_row: int) -> int:
    return int(sheet_row) - int(data_start_row) - 1


def _table_data_index_to_sheet_row(row_index: int, data_start_row: int) -> int:
    return int(data_start_row) + int(row_index) + 1


def _table_row_to_mapping(
    row: Any,
    *,
    row_index: int,
    data_start_row: int,
    columns: list[str],
) -> dict[str, Any]:
    values = _normalize_sheet_row(row, len(columns))
    item = {
        "_row_index": row_index,
        "_sheet_row": _table_data_index_to_sheet_row(row_index, data_start_row),
    }
    for column_index, column in enumerate(columns):
        item[str(column)] = values[column_index] if column_index < len(values) else ""
    return item


def _split_formula_top_level(value: str, separator: str) -> list[str]:
    parts: list[str] = []
    current: list[str] = []
    depth = 0
    in_string = False
    index = 0
    while index < len(value):
        char = value[index]
        if char == '"':
            current.append(char)
            if in_string and index + 1 < len(value) and value[index + 1] == '"':
                current.append(value[index + 1])
                index += 2
                continue
            in_string = not in_string
            index += 1
            continue
        if not in_string:
            if char == "(":
                depth += 1
            elif char == ")":
                depth = max(depth - 1, 0)
            elif char == separator and depth == 0:
                parts.append("".join(current).strip())
                current = []
                index += 1
                continue
        current.append(char)
        index += 1
    parts.append("".join(current).strip())
    return parts


def _split_formula_top_level_operators(value: str, operators: set[str]) -> tuple[list[str], list[str]]:
    parts: list[str] = []
    found_operators: list[str] = []
    current: list[str] = []
    depth = 0
    in_string = False
    for index, char in enumerate(value):
        if char == '"':
            in_string = not in_string
        if not in_string:
            if char == "(":
                depth += 1
            elif char == ")":
                depth = max(depth - 1, 0)
            elif char in operators and depth == 0 and current:
                if char in "+-":
                    previous = value[index - 1] if index > 0 else ""
                    if previous in "+-*/(":
                        current.append(char)
                        continue
                parts.append("".join(current).strip())
                found_operators.append(char)
                current = []
                continue
        current.append(char)
    if found_operators:
        parts.append("".join(current).strip())
    return parts, found_operators


def _strip_formula_outer_parentheses(value: str) -> str:
    text = value.strip()
    while text.startswith("(") and text.endswith(")"):
        depth = 0
        in_string = False
        balanced_outer = True
        for index, char in enumerate(text):
            if char == '"':
                in_string = not in_string
                continue
            if in_string:
                continue
            if char == "(":
                depth += 1
            elif char == ")":
                depth -= 1
                if depth == 0 and index != len(text) - 1:
                    balanced_outer = False
                    break
        if not balanced_outer or depth != 0:
            return text
        text = text[1:-1].strip()
    return text


def _split_formula_args(value: str) -> list[str]:
    return _split_formula_top_level(value, ",")


def _parse_formula_function(value: str) -> tuple[str, list[str]] | None:
    match = re.match(r"^\s*([A-Za-z_][A-Za-z0-9_]*)\s*\((.*)\)\s*$", value, re.S)
    if not match:
        return None
    return match.group(1).upper(), _split_formula_args(match.group(2))


def _coerce_formula_number(value: Any) -> float | None:
    if isinstance(value, bool):
        return float(int(value))
    if isinstance(value, int | float):
        return float(value)
    if isinstance(value, datetime):
        return float(value.toordinal())
    if isinstance(value, date):
        return float(value.toordinal())
    text = str(value or "").strip()
    if not text:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return None


def _format_formula_number(value: float) -> int | float:
    rounded = round(value)
    if abs(value - rounded) < 1e-9:
        return int(rounded)
    return value


def _parse_formula_string_literal(value: str) -> str | None:
    text = value.strip()
    if len(text) < 2 or not (text.startswith('"') and text.endswith('"')):
        return None
    return text[1:-1].replace('""', '"')


def _parse_formula_cell_reference(value: str) -> tuple[int, int] | None:
    match = A1_CELL_REFERENCE_RE.match(value)
    if not match:
        return None
    column_index = _excel_column_index(match.group("column"))
    if column_index is None:
        return None
    return int(match.group("row")) - 1, column_index


def _parse_table_formula_range_reference(value: str) -> tuple[int, int, int, int] | None:
    match = re.match(
        r"^\s*\$?([A-Za-z]{1,3})\$?(\d+)\s*:\s*\$?([A-Za-z]{1,3})\$?(\d+)\s*$",
        value,
    )
    if not match:
        return None
    start_column = _excel_column_index(match.group(1))
    end_column = _excel_column_index(match.group(3))
    if start_column is None or end_column is None:
        return None
    start_row = int(match.group(2)) - 1
    end_row = int(match.group(4)) - 1
    return (
        min(start_row, end_row),
        min(start_column, end_column),
        max(start_row, end_row),
        max(start_column, end_column),
    )


def _get_formula_defined_name_cache(cache: dict[tuple[int, int], Any]) -> dict[str, Any]:
    value = cache.get(FORMULA_DEFINED_NAME_CACHE_KEY)
    if isinstance(value, dict):
        return value
    value = {}
    cache[FORMULA_DEFINED_NAME_CACHE_KEY] = value
    return value


def _get_formula_grid_cell(
    grid_rows: list[list[Any]],
    row_index: int,
    column_index: int,
    cache: dict[tuple[int, int], Any],
    defined_names: dict[str, str] | None = None,
) -> Any:
    if row_index < 0 or column_index < 0 or row_index >= len(grid_rows):
        return ""
    row = grid_rows[row_index]
    if column_index >= len(row):
        return ""
    return _evaluate_table_formula_text_value(
        row[column_index],
        grid_rows=grid_rows,
        row_index=row_index,
        column_index=column_index,
        cache=cache,
        defined_names=defined_names,
    )


def _get_formula_grid_range(
    grid_rows: list[list[Any]],
    range_ref: tuple[int, int, int, int],
    cache: dict[tuple[int, int], Any],
    defined_names: dict[str, str] | None = None,
) -> list[Any]:
    start_row, start_col, end_row, end_col = range_ref
    values: list[Any] = []
    for row_index in range(start_row, end_row + 1):
        for column_index in range(start_col, end_col + 1):
            values.append(_get_formula_grid_cell(grid_rows, row_index, column_index, cache, defined_names))
    return values


def _evaluate_formula_comparison(
    expr: str,
    *,
    grid_rows: list[list[Any]],
    cache: dict[tuple[int, int], Any],
    defined_names: dict[str, str] | None = None,
    name_stack: tuple[str, ...] = (),
) -> bool | None:
    text = expr.strip()
    in_string = False
    depth = 0
    operators = (">=", "<=", "<>", "=", ">", "<")
    index = 0
    while index < len(text):
        char = text[index]
        if char == '"':
            in_string = not in_string
            index += 1
            continue
        if in_string:
            index += 1
            continue
        if char == "(":
            depth += 1
            index += 1
            continue
        if char == ")":
            depth = max(depth - 1, 0)
            index += 1
            continue
        if depth == 0:
            for operator in operators:
                if text.startswith(operator, index):
                    left = _evaluate_table_formula_expr(
                        text[:index],
                        grid_rows=grid_rows,
                        cache=cache,
                        defined_names=defined_names,
                        name_stack=name_stack,
                    )
                    right = _evaluate_table_formula_expr(
                        text[index + len(operator):],
                        grid_rows=grid_rows,
                        cache=cache,
                        defined_names=defined_names,
                        name_stack=name_stack,
                    )
                    left_number = _coerce_formula_number(left)
                    right_number = _coerce_formula_number(right)
                    if left_number is not None and right_number is not None:
                        left_value: Any = left_number
                        right_value: Any = right_number
                    else:
                        left_value = str(left or "")
                        right_value = str(right or "")
                    if operator == ">=":
                        return left_value >= right_value
                    if operator == "<=":
                        return left_value <= right_value
                    if operator == "<>":
                        return left_value != right_value
                    if operator == "=":
                        return left_value == right_value
                    if operator == ">":
                        return left_value > right_value
                    if operator == "<":
                        return left_value < right_value
        index += 1
    return None


def _evaluate_formula_arithmetic(
    expr: str,
    *,
    grid_rows: list[list[Any]],
    cache: dict[tuple[int, int], Any],
    defined_names: dict[str, str] | None = None,
    name_stack: tuple[str, ...] = (),
) -> Any | None:
    terms, operators = _split_formula_top_level_operators(expr, {"+", "-"})
    if operators:
        first = _evaluate_table_formula_expr(
            terms[0],
            grid_rows=grid_rows,
            cache=cache,
            defined_names=defined_names,
            name_stack=name_stack,
        )
        result = _coerce_formula_number(first)
        if result is None:
            return None
        for operator, term in zip(operators, terms[1:]):
            value = _evaluate_table_formula_expr(
                term,
                grid_rows=grid_rows,
                cache=cache,
                defined_names=defined_names,
                name_stack=name_stack,
            )
            number = _coerce_formula_number(value)
            if number is None:
                return None
            result = result + number if operator == "+" else result - number
        return _format_formula_number(result)

    multiply_parts, multiply_operators = _split_formula_top_level_operators(expr, {"*", "/"})
    if multiply_operators:
        first = _evaluate_table_formula_expr(
            multiply_parts[0],
            grid_rows=grid_rows,
            cache=cache,
            defined_names=defined_names,
            name_stack=name_stack,
        )
        result = _coerce_formula_number(first)
        if result is None:
            return None
        for operator, part in zip(multiply_operators, multiply_parts[1:]):
            value = _evaluate_table_formula_expr(
                part,
                grid_rows=grid_rows,
                cache=cache,
                defined_names=defined_names,
                name_stack=name_stack,
            )
            number = _coerce_formula_number(value)
            if number is None:
                return None
            if operator == "*":
                result *= number
            else:
                if number == 0:
                    return None
                result /= number
        return _format_formula_number(result)
    return None


def _evaluate_table_formula_expr(
    expr: str,
    *,
    grid_rows: list[list[Any]],
    cache: dict[tuple[int, int], Any],
    defined_names: dict[str, str] | None = None,
    name_stack: tuple[str, ...] = (),
) -> Any:
    text = _strip_formula_outer_parentheses(expr)
    if not text:
        return ""

    concat_parts = _split_formula_top_level(text, "&")
    if len(concat_parts) > 1:
        return "".join(str(_evaluate_table_formula_expr(
            part,
            grid_rows=grid_rows,
            cache=cache,
            defined_names=defined_names,
            name_stack=name_stack,
        )) for part in concat_parts)

    literal = _parse_formula_string_literal(text)
    if literal is not None:
        return literal

    upper = text.upper()
    if upper in {"TRUE", "TRUE()"}:
        return True
    if upper in {"FALSE", "FALSE()"}:
        return False
    if upper == "TODAY()":
        return date.today()
    if upper == "NOW()":
        return datetime.now()

    cell_ref = _parse_table_formula_cell_reference(text)
    if cell_ref is not None:
        return _get_formula_grid_cell(grid_rows, cell_ref[0], cell_ref[1], cache, defined_names)

    range_ref = _parse_table_formula_range_reference(text)
    if range_ref is not None:
        return _get_formula_grid_range(grid_rows, range_ref, cache, defined_names)

    try:
        numeric = float(text)
    except ValueError:
        numeric = None
    if numeric is not None:
        return _format_formula_number(numeric)

    name_key = text.lower()
    if defined_names and name_key in defined_names and name_key not in name_stack:
        name_cache = _get_formula_defined_name_cache(cache)
        if name_key in name_cache:
            return name_cache[name_key]
        name_formula = defined_names[name_key]
        result = _evaluate_table_formula_expr(
            name_formula[1:] if name_formula.startswith("=") else name_formula,
            grid_rows=grid_rows,
            cache=cache,
            defined_names=defined_names,
            name_stack=(*name_stack, name_key),
        )
        name_cache[name_key] = result
        return result

    comparison = _evaluate_formula_comparison(
        text,
        grid_rows=grid_rows,
        cache=cache,
        defined_names=defined_names,
        name_stack=name_stack,
    )
    if comparison is not None:
        return comparison

    arithmetic = _evaluate_formula_arithmetic(
        text,
        grid_rows=grid_rows,
        cache=cache,
        defined_names=defined_names,
        name_stack=name_stack,
    )
    if arithmetic is not None:
        return arithmetic

    func = _parse_formula_function(text)
    if func is not None:
        name, args = func
        if name == "IF" and len(args) >= 2:
            condition = bool(_evaluate_table_formula_expr(
                args[0],
                grid_rows=grid_rows,
                cache=cache,
                defined_names=defined_names,
                name_stack=name_stack,
            ))
            return _evaluate_table_formula_expr(
                args[1] if condition else (args[2] if len(args) > 2 else ""),
                grid_rows=grid_rows,
                cache=cache,
                defined_names=defined_names,
                name_stack=name_stack,
            )
        if name == "IFERROR" and args:
            try:
                return _evaluate_table_formula_expr(
                    args[0],
                    grid_rows=grid_rows,
                    cache=cache,
                    defined_names=defined_names,
                    name_stack=name_stack,
                )
            except Exception:
                return _evaluate_table_formula_expr(
                    args[1] if len(args) > 1 else "",
                    grid_rows=grid_rows,
                    cache=cache,
                    defined_names=defined_names,
                    name_stack=name_stack,
                )
        if name == "AND":
            return all(bool(_evaluate_table_formula_expr(
                arg,
                grid_rows=grid_rows,
                cache=cache,
                defined_names=defined_names,
                name_stack=name_stack,
            )) for arg in args)
        if name == "OR":
            return any(bool(_evaluate_table_formula_expr(
                arg,
                grid_rows=grid_rows,
                cache=cache,
                defined_names=defined_names,
                name_stack=name_stack,
            )) for arg in args)
        if name == "LEN" and args:
            return len(str(_evaluate_table_formula_expr(
                args[0],
                grid_rows=grid_rows,
                cache=cache,
                defined_names=defined_names,
                name_stack=name_stack,
            ) or ""))
        if name == "INT" and args:
            value = _evaluate_table_formula_expr(args[0], grid_rows=grid_rows, cache=cache, defined_names=defined_names, name_stack=name_stack)
            number = _coerce_formula_number(value)
            return floor(number) if number is not None else text
        if name == "VALUE" and args:
            value = _evaluate_table_formula_expr(args[0], grid_rows=grid_rows, cache=cache, defined_names=defined_names, name_stack=name_stack)
            number = _coerce_formula_number(value)
            return _format_formula_number(number) if number is not None else text
        if name == "DATE" and len(args) >= 3:
            year = int(_coerce_formula_number(_evaluate_table_formula_expr(args[0], grid_rows=grid_rows, cache=cache, defined_names=defined_names, name_stack=name_stack)) or 0)
            month = int(_coerce_formula_number(_evaluate_table_formula_expr(args[1], grid_rows=grid_rows, cache=cache, defined_names=defined_names, name_stack=name_stack)) or 0)
            day = int(_coerce_formula_number(_evaluate_table_formula_expr(args[2], grid_rows=grid_rows, cache=cache, defined_names=defined_names, name_stack=name_stack)) or 0)
            return date(year, month, day)
        if name in {"TEXTJOIN", "TEXTJOIN_COMPAT"} and len(args) >= 3:
            delimiter = str(_evaluate_table_formula_expr(args[0], grid_rows=grid_rows, cache=cache, defined_names=defined_names, name_stack=name_stack))
            ignore_empty = bool(_evaluate_table_formula_expr(args[1], grid_rows=grid_rows, cache=cache, defined_names=defined_names, name_stack=name_stack))
            values = [
                _evaluate_table_formula_expr(arg, grid_rows=grid_rows, cache=cache, defined_names=defined_names, name_stack=name_stack)
                for arg in args[2:]
            ]
            parts = [str(value) for value in values if not ignore_empty or str(value) != ""]
            return delimiter.join(parts)
        if name in {"DATEDIF", "DATEDIF_COMPAT"} and len(args) >= 3:
            start_value = _evaluate_table_formula_expr(args[0], grid_rows=grid_rows, cache=cache, defined_names=defined_names, name_stack=name_stack)
            end_value = _evaluate_table_formula_expr(args[1], grid_rows=grid_rows, cache=cache, defined_names=defined_names, name_stack=name_stack)
            unit = str(_evaluate_table_formula_expr(args[2], grid_rows=grid_rows, cache=cache, defined_names=defined_names, name_stack=name_stack)).lower()
            start_date = date.fromisoformat(str(start_value))
            if isinstance(end_value, date):
                end_date = end_value
            else:
                end_date = date.fromisoformat(str(end_value))
            if unit == "d":
                return (end_date - start_date).days
            return text
        if name == "COUNTIF" and len(args) >= 2:
            values = _evaluate_table_formula_expr(args[0], grid_rows=grid_rows, cache=cache, defined_names=defined_names, name_stack=name_stack)
            pattern = str(_evaluate_table_formula_expr(args[1], grid_rows=grid_rows, cache=cache, defined_names=defined_names, name_stack=name_stack))
            needle = pattern.strip("*")
            if not isinstance(values, list):
                values = [values]
            if pattern.startswith("*") and pattern.endswith("*"):
                return sum(1 for value in values if needle in str(value or ""))
            return sum(1 for value in values if str(value or "") == pattern)
        if name in {"MIN", "MAX", "SUM"} and args:
            numbers: list[float] = []
            for arg in args:
                value = _evaluate_table_formula_expr(arg, grid_rows=grid_rows, cache=cache, defined_names=defined_names, name_stack=name_stack)
                values = value if isinstance(value, list) else [value]
                for item in values:
                    number = _coerce_formula_number(item)
                    if number is not None:
                        numbers.append(number)
            if not numbers:
                return 0
            if name == "MIN":
                return _format_formula_number(min(numbers))
            if name == "MAX":
                return _format_formula_number(max(numbers))
            return _format_formula_number(sum(numbers))
        if name == "SWITCH" and len(args) >= 3:
            target = _evaluate_table_formula_expr(args[0], grid_rows=grid_rows, cache=cache, defined_names=defined_names, name_stack=name_stack)
            pairs = args[1:]
            default_value = None
            if len(pairs) % 2 == 1:
                default_value = pairs[-1]
                pairs = pairs[:-1]
            for condition_expr, value_expr in zip(pairs[0::2], pairs[1::2]):
                condition = _evaluate_table_formula_expr(condition_expr, grid_rows=grid_rows, cache=cache, defined_names=defined_names, name_stack=name_stack)
                if condition == target:
                    return _evaluate_table_formula_expr(value_expr, grid_rows=grid_rows, cache=cache, defined_names=defined_names, name_stack=name_stack)
            return _evaluate_table_formula_expr(default_value, grid_rows=grid_rows, cache=cache, defined_names=defined_names, name_stack=name_stack) if default_value is not None else ""

    return text


def _evaluate_table_formula_text_value(
    value: Any,
    *,
    grid_rows: list[list[Any]],
    row_index: int,
    column_index: int,
    cache: dict[tuple[int, int], Any],
    defined_names: dict[str, str] | None = None,
) -> Any:
    if not _is_formula_expression(value):
        return value
    key = (row_index, column_index)
    if key in cache:
        return cache[key]
    cache[key] = ""
    try:
        result = _evaluate_table_formula_expr(
            str(value)[1:],
            grid_rows=grid_rows,
            cache=cache,
            defined_names=defined_names,
        )
    except Exception:
        result = value
    cache[key] = result
    return result


def _build_table_text_grid(
    normalized: dict[str, Any],
    *,
    columns: list[str],
    rows: list[Any],
    defined_names: dict[str, str] | None = None,
) -> list[list[Any]]:
    data_start_row = _normalize_document_data_start_row(normalized)
    raw_grid_rows = _extract_document_grid_rows(normalized)
    prefix_rows = raw_grid_rows[:data_start_row] if raw_grid_rows else [[""] * len(columns) for _ in range(data_start_row)]
    source_grid = [
        _normalize_sheet_row(row, len(columns))
        for row in [*prefix_rows, *rows]
    ]
    cache: dict[tuple[int, int], Any] = {}
    return [
        [
            _evaluate_table_formula_text_value(
                cell,
                grid_rows=source_grid,
                row_index=row_index,
                column_index=column_index,
                cache=cache,
                defined_names=defined_names,
            )
            for column_index, cell in enumerate(row)
        ]
        for row_index, row in enumerate(source_grid)
    ]


def _build_table_defined_name_values(
    *,
    text_grid: list[list[Any]],
    defined_names: dict[str, str] | None,
) -> dict[str, Any]:
    if not defined_names:
        return {}
    cache: dict[tuple[int, int], Any] = {}
    values: dict[str, Any] = {}
    for name in defined_names:
        try:
            values[name] = _evaluate_table_formula_expr(
                name,
                grid_rows=text_grid,
                cache=cache,
                defined_names=defined_names,
            )
        except Exception:
            continue
    return values


def _build_note_sheet_table_response(
    document: SheetDocument,
    *,
    workbook: WorkbookDocument | None = None,
    include_grid: bool = False,
    value_mode: Literal["text", "raw"] = "text",
    defined_names: dict[str, str] | None = None,
) -> NoteSheetTableResponse:
    normalized = _normalize_document_json(dict(document.document_json or {}))
    if _is_registration_sheet(document):
        normalized, _registration_header_changed = _normalize_registration_sheet_header_document(normalized)
    normalized, _formula_repaired_count = _normalize_attendance_dual_clockin_refund_formulas(normalized)
    columns = _normalize_document_columns(normalized)
    rows = _extract_document_rows(normalized)
    data_start_row = _normalize_document_data_start_row(normalized)
    field_row_index = int(normalized.get("field_row_index") or 0)
    raw_grid_rows = _extract_document_grid_rows(normalized)
    response_rows = rows
    response_grid_rows = raw_grid_rows
    defined_name_values: dict[str, Any] = {}
    if value_mode == "text":
        text_grid = _build_table_text_grid(normalized, columns=columns, rows=rows, defined_names=defined_names)
        defined_name_values = _build_table_defined_name_values(text_grid=text_grid, defined_names=defined_names)
        response_rows = text_grid[data_start_row:data_start_row + len(rows)]
        if raw_grid_rows:
            response_grid_rows = text_grid[:len(raw_grid_rows)]
    table_rows = [
        _table_row_to_mapping(
            row,
            row_index=row_index,
            data_start_row=data_start_row,
            columns=columns,
        )
        for row_index, row in enumerate(response_rows)
    ]
    grid_rows = [
        _normalize_sheet_row(row, len(columns))
        for row in response_grid_rows
    ] if include_grid else []
    return NoteSheetTableResponse(
        id=_require_sheet_numeric_id(document),
        workbook_id=_require_workbook_numeric_id(workbook) if workbook is not None else None,
        title=document.title,
        version=int(document.version or 1),
        value_mode=value_mode,
        columns=columns,
        rows=table_rows,
        row_count=len(rows),
        data_start_row=data_start_row,
        field_row_index=field_row_index,
        grid_rows=grid_rows,
        defined_name_values=defined_name_values,
    )


def _resolve_workbook_for_sheet(
    session: Session,
    document: SheetDocument,
    workbook: WorkbookDocument | None,
) -> WorkbookDocument | None:
    if workbook is not None:
        return workbook
    workbooks = _get_workbooks_for_sheet(session, document)
    return workbooks[0] if workbooks else None


def _build_registration_phone_map(registration_document: SheetDocument | None) -> dict[str, str]:
    if registration_document is None:
        return {}
    registration_json = _normalize_document_json(dict(registration_document.document_json or {}))
    registration_columns = _normalize_document_columns(registration_json)
    sequence_index = _get_column_index(registration_columns, NOTE_SHEET_REGISTRATION_SEQUENCE_COLUMN)
    phone_index = _get_column_index(registration_columns, "手机号")
    if sequence_index < 0 or phone_index < 0:
        return {}

    phone_map: dict[str, str] = {}
    for raw_row in _extract_document_rows(registration_json):
        row = _normalize_sheet_row(raw_row, len(registration_columns))
        student_id = _normalize_sheet_text(row[sequence_index])
        phone = _strip_legacy_text_prefix(row[phone_index])
        if student_id and phone:
            phone_map[student_id] = phone
    return phone_map


def _attendance_export_columns(columns: list[str]) -> list[str]:
    export_columns = [column for column in columns if column != "手机号"]
    nickname_index = _get_column_index(export_columns, "昵称")
    if nickname_index >= 0:
        export_columns.insert(nickname_index + 1, "手机号")
    return export_columns


def _attendance_export_rows(
    table: NoteSheetTableResponse,
    *,
    phone_map: dict[str, str],
) -> list[list[Any]]:
    source_columns = table.columns
    export_columns = _attendance_export_columns(source_columns)
    student_id_column = "学号"
    rows: list[list[Any]] = []
    for item in table.rows:
        student_id = _normalize_sheet_text(item.get(student_id_column))
        if not student_id and not _normalize_sheet_text(item.get("姓名")) and not _normalize_sheet_text(item.get("昵称")):
            continue
        row_values: list[Any] = []
        for column in export_columns:
            if column == "手机号":
                value = phone_map.get(student_id, "")
            else:
                value = item.get(column, "")
            row_values.append(value)
        rows.append(row_values)
    return rows


def _normalize_attendance_export_cell(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, int | float):
        return value
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    return str(value)


def _build_attendance_export_workbook_bytes(
    table: NoteSheetTableResponse,
    *,
    phone_map: dict[str, str],
) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="服务端缺少 openpyxl，无法导出 Excel") from exc

    export_columns = _attendance_export_columns(table.columns)
    export_rows = _attendance_export_rows(table, phone_map=phone_map)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "考勤表"
    header_fill = PatternFill(fill_type="solid", fgColor="D9E1F2")
    for column_index, header in enumerate(export_columns, start=1):
        cell = worksheet.cell(row=1, column=column_index, value=header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for row_index, row in enumerate(export_rows, start=2):
        for column_index, value in enumerate(row, start=1):
            header = export_columns[column_index - 1]
            cell = worksheet.cell(row=row_index, column=column_index, value=_normalize_attendance_export_cell(value))
            if header in {"手机号", "学号", "微信支付订单号", "商户订单号"}:
                cell.number_format = "@"

    for column_index, header in enumerate(export_columns, start=1):
        values = [header, *[str(row[column_index - 1] or "") for row in export_rows[:200]]]
        width = min(max(max((len(value) for value in values), default=0) + 2, 10), 32)
        worksheet.column_dimensions[get_column_letter(column_index)].width = width

    stream = io.BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _attendance_export_filename(workbook: WorkbookDocument | None, document: SheetDocument) -> str:
    base_title = _normalize_sheet_text(workbook.title if workbook is not None else "") or _normalize_sheet_text(document.title)
    safe_title = re.sub(r'[\\/:*?"<>|]+', "_", base_title).strip() or "考勤表"
    suffix = "" if safe_title.endswith("考勤表") else "_考勤表"
    return f"{safe_title}{suffix}.xlsx"


def _ensure_table_data_row(rows: list[Any], row_index: int, columns: list[str]) -> None:
    while len(rows) <= row_index:
        rows.append([""] * len(columns))


def _set_table_data_cell(
    rows: list[Any],
    *,
    row_index: int,
    column_index: int,
    value: Any,
    columns: list[str],
) -> bool:
    if row_index < 0:
        raise HTTPException(status_code=400, detail="数据行号超出表格范围")
    _ensure_table_data_row(rows, row_index, columns)
    current_value = _extract_row_cell_value(rows[row_index], column_index, columns)
    if current_value == value:
        return False
    rows[row_index] = _set_row_cell_value(rows[row_index], columns, column_index, value)
    return True


def _ensure_table_grid_row(grid_rows: list[Any], row_index: int, column_count: int) -> None:
    while len(grid_rows) <= row_index:
        grid_rows.append([""] * column_count)


def _set_table_grid_cell(
    grid_rows: list[Any],
    *,
    grid_row_index: int,
    column_index: int,
    value: Any,
    column_count: int,
) -> bool:
    if grid_row_index < 0:
        raise HTTPException(status_code=400, detail="表格行号超出范围")
    _ensure_table_grid_row(grid_rows, grid_row_index, column_count)
    current_cells = _normalize_sheet_row(grid_rows[grid_row_index], column_count)
    current_value = current_cells[column_index] if column_index < len(current_cells) else ""
    if current_value == value:
        return False
    current_cells[column_index] = value
    grid_rows[grid_row_index] = current_cells
    return True


def _set_table_sheet_cell(
    rows: list[Any],
    grid_rows: list[Any],
    *,
    sheet_row: int,
    column_index: int,
    value: Any,
    columns: list[str],
    data_start_row: int,
) -> tuple[bool, bool]:
    if sheet_row <= 0:
        raise HTTPException(status_code=400, detail="sheet_row 必须从 1 开始")
    if column_index < 0 or column_index >= len(columns):
        raise HTTPException(status_code=400, detail="列号超出表格范围")
    if sheet_row <= data_start_row:
        changed = _set_table_grid_cell(
            grid_rows,
            grid_row_index=sheet_row - 1,
            column_index=column_index,
            value=value,
            column_count=len(columns),
        )
        return changed, False
    changed = _set_table_data_cell(
        rows,
        row_index=_table_sheet_row_to_data_index(sheet_row, data_start_row),
        column_index=column_index,
        value=value,
        columns=columns,
    )
    return changed, True


def _resolve_table_operation_column(columns: list[str], operation: NoteSheetTablePatchOperation) -> int:
    column = operation.column if operation.column is not None else operation.field
    column_index = _find_table_column_index(columns, column)
    if column_index is None:
        raise HTTPException(status_code=400, detail=f"找不到字段列: {column}")
    return column_index


def _apply_table_write_fields_operation(
    rows: list[Any],
    *,
    columns: list[str],
    data_start_row: int,
    operation: NoteSheetTablePatchOperation,
) -> tuple[int, set[int]]:
    if not operation.rows:
        return 0, set()

    writable_fields = [
        field
        for field in (operation.fields or sorted({
            str(key)
            for item in operation.rows
            for key in item
            if not str(key).startswith("_") and str(key) != operation.key_field
        }))
        if not str(field).startswith("_")
    ]
    column_indexes: dict[str, int] = {}
    for field in writable_fields:
        index = _find_table_column_index(columns, field)
        if index is None:
            raise HTTPException(status_code=400, detail=f"找不到字段列: {field}")
        column_indexes[field] = index

    key_row_map: dict[str, int] = {}
    if operation.key_field:
        key_column_index = _find_table_column_index(columns, operation.key_field)
        if key_column_index is None:
            raise HTTPException(status_code=400, detail=f"找不到 key_field 字段列: {operation.key_field}")
        for row_index, row in enumerate(rows):
            key = _normalize_sheet_text(_extract_row_cell_value(row, key_column_index, columns))
            if key and key not in key_row_map:
                key_row_map[key] = row_index

    if operation.start_row_index is not None:
        current_row_index = operation.start_row_index
    elif operation.start_sheet_row is not None:
        current_row_index = _table_sheet_row_to_data_index(operation.start_sheet_row, data_start_row)
    else:
        current_row_index = 0

    updated_cells = 0
    updated_rows: set[int] = set()
    for offset, incoming in enumerate(operation.rows):
        target_row_index: int | None = None
        if operation.key_field:
            key_value = _normalize_sheet_text(incoming.get(operation.key_field))
            target_row_index = key_row_map.get(key_value)
            if target_row_index is None and operation.append_missing:
                target_row_index = len(rows)
                key_row_map[key_value] = target_row_index
        else:
            target_row_index = current_row_index + offset

        if target_row_index is None or target_row_index < 0:
            continue

        for field, column_index in column_indexes.items():
            if field not in incoming:
                continue
            if _set_table_data_cell(
                rows,
                row_index=target_row_index,
                column_index=column_index,
                value=incoming.get(field),
                columns=columns,
            ):
                updated_cells += 1
                updated_rows.add(target_row_index)

    return updated_cells, updated_rows


def _apply_table_write_range_operation(
    rows: list[Any],
    grid_rows: list[Any],
    *,
    columns: list[str],
    data_start_row: int,
    operation: NoteSheetTablePatchOperation,
) -> tuple[int, set[int]]:
    if operation.cell:
        start_sheet_row, start_column_index = _parse_table_cell_reference(operation.cell)
    else:
        start_column_index = _resolve_table_operation_column(columns, operation)
        if operation.start_sheet_row is not None:
            start_sheet_row = operation.start_sheet_row
        elif operation.start_row_index is not None:
            start_sheet_row = _table_data_index_to_sheet_row(operation.start_row_index, data_start_row)
        else:
            start_sheet_row = _table_data_index_to_sheet_row(0, data_start_row)

    updated_cells = 0
    updated_rows: set[int] = set()
    for row_offset, row_values in enumerate(operation.values):
        for column_offset, value in enumerate(row_values):
            column_index = start_column_index + column_offset
            if column_index < 0 or column_index >= len(columns):
                continue
            changed, is_data_row = _set_table_sheet_cell(
                rows,
                grid_rows,
                sheet_row=start_sheet_row + row_offset,
                column_index=column_index,
                value=value,
                columns=columns,
                data_start_row=data_start_row,
            )
            if changed:
                updated_cells += 1
                if is_data_row:
                    updated_rows.add(_table_sheet_row_to_data_index(start_sheet_row + row_offset, data_start_row))
    return updated_cells, updated_rows


def _apply_table_set_cell_operation(
    rows: list[Any],
    grid_rows: list[Any],
    *,
    columns: list[str],
    data_start_row: int,
    operation: NoteSheetTablePatchOperation,
) -> tuple[int, set[int]]:
    if operation.cell:
        sheet_row, column_index = _parse_table_cell_reference(operation.cell)
    else:
        column_index = _resolve_table_operation_column(columns, operation)
        if operation.sheet_row is not None:
            sheet_row = operation.sheet_row
        elif operation.row_index is not None:
            sheet_row = _table_data_index_to_sheet_row(operation.row_index, data_start_row)
        else:
            raise HTTPException(status_code=400, detail="set_cell 需要 cell、sheet_row 或 row_index")

    changed, is_data_row = _set_table_sheet_cell(
        rows,
        grid_rows,
        sheet_row=sheet_row,
        column_index=column_index,
        value=operation.value,
        columns=columns,
        data_start_row=data_start_row,
    )
    if not changed:
        return 0, set()
    return 1, {_table_sheet_row_to_data_index(sheet_row, data_start_row)} if is_data_row else set()


def _apply_table_set_note_cell_operation(
    grid_rows: list[Any],
    *,
    columns: list[str],
    field_row_index: int,
    operation: NoteSheetTablePatchOperation,
) -> int:
    column_index = _resolve_table_operation_column(columns, operation)
    if operation.sheet_row is not None:
        grid_row_index = operation.sheet_row - 1
    elif operation.row_index is not None:
        grid_row_index = operation.row_index
    else:
        grid_row_index = field_row_index + 1
    return int(_set_table_grid_cell(
        grid_rows,
        grid_row_index=grid_row_index,
        column_index=column_index,
        value=operation.value,
        column_count=len(columns),
    ))


def _apply_note_sheet_table_patch(
    document_json: dict[str, Any],
    operations: list[NoteSheetTablePatchOperation],
) -> tuple[dict[str, Any], int, int]:
    normalized = _normalize_document_json(document_json)
    columns = _normalize_document_columns(normalized)
    rows = list(_extract_document_rows(normalized))
    data_start_row = _normalize_document_data_start_row(normalized)
    field_row_index = int(normalized.get("field_row_index") or 0)
    raw_grid_rows = _extract_document_grid_rows(normalized)
    grid_rows = [
        _normalize_sheet_row(row, len(columns))
        for row in raw_grid_rows[:data_start_row]
    ]
    if not grid_rows and data_start_row:
        grid_rows = [[""] * len(columns) for _ in range(data_start_row)]

    updated_cells = 0
    updated_rows: set[int] = set()
    for operation in operations:
        if operation.type == "write_fields":
            cell_count, row_indexes = _apply_table_write_fields_operation(
                rows,
                columns=columns,
                data_start_row=data_start_row,
                operation=operation,
            )
        elif operation.type == "write_range":
            cell_count, row_indexes = _apply_table_write_range_operation(
                rows,
                grid_rows,
                columns=columns,
                data_start_row=data_start_row,
                operation=operation,
            )
        elif operation.type == "set_cell":
            cell_count, row_indexes = _apply_table_set_cell_operation(
                rows,
                grid_rows,
                columns=columns,
                data_start_row=data_start_row,
                operation=operation,
            )
        elif operation.type == "set_note_cell":
            cell_count = _apply_table_set_note_cell_operation(
                grid_rows,
                columns=columns,
                field_row_index=field_row_index,
                operation=operation,
            )
            row_indexes = set()
        else:  # pragma: no cover - guarded by pydantic Literal
            raise HTTPException(status_code=400, detail=f"未知表格操作: {operation.type}")
        updated_cells += cell_count
        updated_rows.update(row_indexes)

    next_document = {
        **normalized,
        "columns": columns,
        "grid_rows": [*grid_rows, *rows] if grid_rows else raw_grid_rows,
    }
    return _replace_document_data_rows(next_document, rows), updated_cells, len(updated_rows)


def _serialize_workbook_summary(
    workbook: WorkbookDocument,
    *,
    sheet_count: int = 0,
    access: NoteSheetResourceAccess | None = None,
) -> dict[str, Any]:
    return {
        "id": _require_workbook_numeric_id(workbook),
        "title": workbook.title,
        "owner_user_id": workbook.owner_user_id,
        "created_by_user_id": workbook.created_by_user_id,
        "updated_by_user_id": workbook.updated_by_user_id,
        "created_at": float(workbook.created_at or 0.0),
        "updated_at": float(workbook.updated_at or 0.0),
        "deleted_at": workbook.deleted_at,
        "deleted_by_user_id": workbook.deleted_by_user_id,
        "sheet_count": sheet_count,
        "access": access,
    }


def _serialize_workbook_detail(
    session: Session,
    workbook: WorkbookDocument,
    *,
    current_user: User | None = None,
    access: NoteSheetResourceAccess | None = None,
) -> dict[str, Any]:
    links = session.exec(
        select(WorkbookSheetLink)
        .where(WorkbookSheetLink.workbook_id.in_(workbook_ref_aliases(workbook)))
        .order_by(WorkbookSheetLink.order_index, WorkbookSheetLink.created_at)
    ).all()
    sheet_ids = [link.sheet_id for link in links]
    sheet_map = load_sheets_by_refs(session, sheet_ids)
    sheets = list({str(sheet.id): sheet for sheet in sheet_map.values()}.values())
    sheet_workbook_refs = _list_workbook_refs_for_sheet_ids(session, sheet_ids, current_user)
    parent_workbook_id = _require_workbook_numeric_id(workbook)
    ordered_sheets: list[dict[str, Any]] = []
    for link in links:
        sheet = sheet_map.get(str(link.sheet_id))
        if sheet is None:
            continue
        sheet_access = _resolve_sheet_resource_access(session, sheet, current_user, workbook=workbook)
        if not sheet_access.capabilities.can_read:
            continue
        ordered_sheets.append(_serialize_sheet_summary(
            sheet,
            workbook_items=sheet_workbook_refs.get(str(sheet.id), []),
            parent_workbook_id=parent_workbook_id,
            access=sheet_access,
        ))
    return {
        **_serialize_workbook_summary(workbook, sheet_count=len(ordered_sheets), access=access),
        "sheets": ordered_sheets,
        "defined_names": _get_workbook_defined_names(session, workbook),
    }


def _serialize_resource_access_grants(
    session: Session,
    *,
    resource_type: str,
    resource_id: str,
) -> list[NoteSheetResourceAccessGrantItem]:
    grants = _fetch_resource_grants(session, resource_type, resource_id)
    user_ids = sorted({
        int(grant.subject_user_id)
        for grant in grants
        if grant.subject_user_id is not None
    })
    users = session.exec(select(User).where(User.id.in_(user_ids))).all() if user_ids else []
    user_map = {user.id: user for user in users}

    result: list[NoteSheetResourceAccessGrantItem] = []
    for grant in sorted(grants, key=lambda item: (item.subject_type, item.subject_key)):
        role = _normalize_resource_role(grant.role)
        if role is None:
            continue
        user = user_map.get(grant.subject_user_id) if grant.subject_user_id is not None else None
        result.append(NoteSheetResourceAccessGrantItem(
            subject_type=grant.subject_type,  # type: ignore[arg-type]
            subject_key=grant.subject_key,
            subject_user_id=grant.subject_user_id,
            username=user.username if user is not None else "",
            nickname=user.nickname if user is not None else "",
            role=role,
        ))
    return result


def _resolve_resource_access_update_subject(
    session: Session,
    item: NoteSheetResourceAccessGrantUpdate,
) -> tuple[str, str, int | None]:
    if item.subject_type == RESOURCE_ACCESS_SUBJECT_ANONYMOUS:
        return RESOURCE_ACCESS_SUBJECT_ANONYMOUS, RESOURCE_ACCESS_SUBJECT_ANONYMOUS, None

    user: User | None = None
    if item.subject_user_id is not None:
        user = session.get(User, item.subject_user_id)
    if user is None and item.username:
        user = session.exec(select(User).where(User.username == item.username.strip())).first()
    if user is None:
        raise HTTPException(status_code=400, detail="用户不存在")
    return RESOURCE_ACCESS_SUBJECT_USER, _build_resource_subject_key(RESOURCE_ACCESS_SUBJECT_USER, user.id), user.id


def _save_resource_access_grants(
    session: Session,
    *,
    resource_type: str,
    resource_id: str,
    payload: NoteSheetResourceAccessUpdateRequest,
    current_user: User,
) -> None:
    now = time.time()
    existing = {
        grant.subject_key: grant
        for grant in _fetch_resource_grants(session, resource_type, resource_id)
    }
    desired_subject_keys: set[str] = set()
    normalized_items: dict[str, tuple[str, int | None, str]] = {}

    for item in payload.grants:
        subject_type, subject_key, subject_user_id = _resolve_resource_access_update_subject(session, item)
        desired_subject_keys.add(subject_key)
        if item.role == "none":
            continue
        role = _normalize_resource_role(item.role)
        if role is None:
            raise HTTPException(status_code=400, detail="非法权限角色")
        if subject_type == RESOURCE_ACCESS_SUBJECT_ANONYMOUS and role in {"editor", "manager"}:
            raise HTTPException(status_code=400, detail="游客不能拥有编辑权限")
        normalized_items[subject_key] = (subject_type, subject_user_id, role)

    for subject_key, grant in list(existing.items()):
        if subject_key not in normalized_items:
            session.delete(grant)

    for subject_key, (subject_type, subject_user_id, role) in normalized_items.items():
        grant = existing.get(subject_key)
        if grant is None:
            grant = ResourceAccessGrant(
                resource_type=resource_type,
                resource_id=resource_id,
                subject_key=subject_key,
                subject_type=subject_type,
                subject_user_id=subject_user_id,
                role=role,
                created_at=now,
                updated_at=now,
                updated_by_user_id=current_user.id,
            )
        else:
            grant.subject_type = subject_type
            grant.subject_user_id = subject_user_id
            grant.role = role
            grant.updated_at = now
            grant.updated_by_user_id = current_user.id
        session.add(grant)


def _build_resource_access_response(
    session: Session,
    *,
    resource_type: Literal["workbook", "sheet"],
    numeric_id: int,
    resource_id: str,
    access: NoteSheetResourceAccess,
) -> NoteSheetResourceAccessResponse:
    return NoteSheetResourceAccessResponse(
        resource_type=resource_type,
        resource_id=numeric_id,
        access=access,
        grants=_serialize_resource_access_grants(session, resource_type=resource_type, resource_id=resource_id),
    )


def _clear_template_document_data_area(document_json: dict[str, Any]) -> dict[str, Any]:
    normalized = _normalize_document_json(document_json)
    next_document = dict(normalized)
    next_document["rows"] = []

    data_start_row = _normalize_document_data_start_row(normalized)
    grid_rows = _extract_document_grid_rows(normalized)
    if grid_rows:
        next_document["grid_rows"] = [*grid_rows[: min(data_start_row, len(grid_rows))]]

    cell_meta = normalized.get("cell_meta")
    if isinstance(cell_meta, dict):
        next_cell_meta: dict[str, Any] = {}
        for key, value in cell_meta.items():
            position = _parse_cell_meta_key(key)
            if position is None:
                next_cell_meta[str(key)] = value
                continue
            row_index, _column_index = position
            if row_index < data_start_row:
                next_cell_meta[str(key)] = value
        next_document["cell_meta"] = next_cell_meta

    next_document = _filter_entity_model_for_document_row_prefix(
        next_document,
        max_document_row=data_start_row,
    )

    return next_document


def _is_template_runtime_derived_column_context(columns: list[str]) -> bool:
    if _is_registration_append_sheet(columns):
        return True
    column_set = set(columns)
    if "用户ID" in column_set and (
        "匹配得分" in column_set
        or "参考信息" in column_set
        or "报名日期" in column_set
        or "订单金额" in column_set
    ):
        return True
    return False


def _clean_template_runtime_derived_columns(document_json: dict[str, Any]) -> dict[str, Any]:
    next_document = _normalize_document_json(document_json)
    columns = _normalize_document_columns(next_document)
    if not _is_template_runtime_derived_column_context(columns):
        return next_document

    for index in range(len(columns) - 1, -1, -1):
        if columns[index] not in NOTE_SHEET_TEMPLATE_RUNTIME_DERIVED_COLUMNS:
            continue
        next_document = _delete_document_column(next_document, delete_index=index)
        columns = _normalize_document_columns(next_document)

    next_document, _changed = _apply_registration_standard_user_id_column_styles(next_document)
    return next_document


def _clean_registration_template_import_source_columns(document_json: dict[str, Any]) -> dict[str, Any]:
    next_document = _normalize_document_json(document_json)
    columns = _normalize_document_columns(next_document)
    if not _is_registration_append_sheet(columns):
        return next_document

    base_columns = set(NOTE_SHEET_REGISTRATION_TEMPLATE_BASE_COLUMNS)
    for index in range(len(columns) - 1, -1, -1):
        if columns[index] in base_columns:
            continue
        next_document = _delete_document_column(next_document, delete_index=index)
        columns = _normalize_document_columns(next_document)

    next_document, _changed = _apply_registration_standard_user_id_column_styles(next_document)
    return next_document


def _clone_sheet_document_json(
    document_json: dict[str, Any],
    *,
    mode: Literal["template", "duplicate"],
) -> dict[str, Any]:
    cloned = deepcopy(_normalize_document_json(document_json))
    if mode == "template":
        cloned = _clear_template_document_data_area(cloned)
        cloned = _clean_template_runtime_derived_columns(cloned)
        cloned = _clean_registration_template_import_source_columns(cloned)
    return cloned


def _get_next_workbook_link_order(session: Session, workbook_id: str) -> int:
    workbook_map = load_workbooks_by_refs(session, [workbook_id])
    workbook_refs = {
        ref
        for workbook in {str(workbook.id): workbook for workbook in workbook_map.values()}.values()
        for ref in workbook_ref_aliases(workbook)
    } or {str(workbook_id)}
    links = session.exec(
        select(WorkbookSheetLink)
        .where(WorkbookSheetLink.workbook_id.in_(workbook_refs))
        .order_by(WorkbookSheetLink.order_index.desc(), WorkbookSheetLink.created_at.desc())
    ).all()
    if not links:
        return 10
    return max(int(links[0].order_index or 0), 0) + 10


@router.get("/access-users", response_model=NoteSheetAccessUserOptionsResponse)
def list_note_sheet_access_users(
    q: str = Query(default="", max_length=100),
    limit: int = Query(default=50, ge=1, le=100),
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_active_user),
):
    _require_note_sheets_feature(session, current_user)
    query_text = q.strip()
    statement = select(User).where(User.is_active == True)  # noqa: E712
    if query_text:
        pattern = f"%{query_text}%"
        statement = statement.where(or_(User.username.like(pattern), User.nickname.like(pattern)))
    statement = statement.order_by(User.username.asc(), User.id.asc()).limit(limit)

    users = session.exec(statement).all()
    return NoteSheetAccessUserOptionsResponse(
        users=[
            NoteSheetAccessUserOption(
                id=user.id or 0,
                username=user.username,
                nickname=user.nickname or "",
            )
            for user in users
            if user.id is not None
        ],
    )


@router.get("/sheets", response_model=list[NoteSheetSummaryResponse])
def list_note_sheets(
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_active_user),
):
    _require_note_sheets_feature(session, current_user)
    documents = session.exec(
        select(SheetDocument)
        .where(SheetDocument.scope == "notes")
        .where(SheetDocument.owner_user_id == current_user.id)
        .where(_active_sheet_condition())
        .order_by(SheetDocument.updated_at.desc(), SheetDocument.created_at.desc())
    ).all()
    workbook_refs = _list_workbook_refs_for_sheet_ids(session, [document.id for document in documents], current_user)
    parent_workbook_ids = _list_parent_workbook_ids_for_sheet_ids(session, [document.id for document in documents])
    return [
        NoteSheetSummaryResponse.model_validate(
            _serialize_sheet_summary(
                document,
                workbook_items=workbook_refs.get(document.id, []),
                parent_workbook_id=parent_workbook_ids.get(str(document.id)),
                access=_resolve_sheet_resource_access(session, document, current_user),
            ),
        )
        for document in documents
    ]


@router.get("/trash", response_model=NoteSheetTrashResponse)
def list_note_sheet_trash(
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_active_user),
):
    _require_note_sheets_feature(session, current_user)
    sheet_query = (
        select(SheetDocument)
        .where(SheetDocument.scope == "notes")
        .where(_deleted_sheet_condition())
        .order_by(SheetDocument.deleted_at.desc(), SheetDocument.updated_at.desc())
    )
    workbook_query = (
        select(WorkbookDocument)
        .where(_deleted_workbook_condition())
        .order_by(WorkbookDocument.deleted_at.desc(), WorkbookDocument.updated_at.desc())
    )
    if not current_user.is_superuser:
        sheet_query = sheet_query.where(SheetDocument.owner_user_id == current_user.id)
        workbook_query = workbook_query.where(WorkbookDocument.owner_user_id == current_user.id)

    sheets = session.exec(sheet_query).all()
    workbooks = session.exec(workbook_query).all()
    sheet_ids = [sheet.id for sheet in sheets]
    workbook_items = _list_workbook_refs_for_sheet_ids(
        session,
        sheet_ids,
        current_user,
        include_deleted=True,
    )
    parent_workbook_ids = _list_parent_workbook_ids_for_sheet_ids(
        session,
        sheet_ids,
        include_deleted=True,
    )

    workbook_refs = [ref for workbook in workbooks for ref in workbook_ref_aliases(workbook)]
    links = session.exec(
        select(WorkbookSheetLink).where(WorkbookSheetLink.workbook_id.in_(workbook_refs))
    ).all() if workbook_refs else []
    linked_sheet_map = load_sheets_by_refs(session, [link.sheet_id for link in links], include_deleted=True)
    workbook_ref_map = {
        ref: workbook
        for workbook in workbooks
        for ref in workbook_ref_aliases(workbook)
    }
    workbook_sheet_counts: dict[str, int] = {}
    for link in links:
        workbook = workbook_ref_map.get(str(link.workbook_id))
        if workbook is None or linked_sheet_map.get(str(link.sheet_id)) is None:
            continue
        workbook_key = str(workbook.id)
        workbook_sheet_counts[workbook_key] = workbook_sheet_counts.get(workbook_key, 0) + 1

    return NoteSheetTrashResponse(
        sheets=[
            NoteSheetSummaryResponse.model_validate(
                _serialize_sheet_summary(
                    sheet,
                    workbook_items=workbook_items.get(str(sheet.id), []),
                    parent_workbook_id=parent_workbook_ids.get(str(sheet.id)),
                    access=_resolve_sheet_resource_access(session, sheet, current_user),
                ),
            )
            for sheet in sheets
        ],
        workbooks=[
            WorkbookSummaryResponse.model_validate(
                _serialize_workbook_summary(
                    workbook,
                    sheet_count=workbook_sheet_counts.get(str(workbook.id), 0),
                    access=_resolve_workbook_resource_access(session, workbook, current_user),
                ),
            )
            for workbook in workbooks
        ],
    )


@router.post("/perf-logs", response_model=NoteSheetPerfLogResponse)
def append_note_sheet_perf_logs(
    payload: NoteSheetPerfLogRequest,
    current_user: User | None = Depends(get_optional_current_user_from_token),
):
    return _write_note_sheet_perf_log_batch(payload, current_user)


@router.post("/sheets", response_model=NoteSheetDetailResponse)
def create_note_sheet(
    payload: NoteSheetCreateRequest,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_active_user),
):
    _require_note_sheets_feature(session, current_user)
    workbook: WorkbookDocument | None = None
    if payload.workbook_id is not None:
        workbook, _workbook_access = _get_workbook_or_404(
            session,
            current_user,
            payload.workbook_id,
            required_role="editor",
        )

    now = time.time()
    document_identity = allocate_new_sheet_identity(session)
    document = SheetDocument(
        id=document_identity.primary_id,
        numeric_id=document_identity.numeric_id,
        legacy_id=document_identity.legacy_id,
        scope="notes",
        owner_type="user",
        owner_key=str(current_user.id),
        sheet_key="pending",
        title=_normalize_title(payload.title, default_value="未命名表格"),
        engine="handsontable",
        document_json=_strip_formula_cell_rich_text(_normalize_created_document_json(payload.document_json)),
        version=1,
        owner_user_id=current_user.id,
        created_by_user_id=current_user.id,
        updated_by_user_id=current_user.id,
        created_at=now,
        updated_at=now,
    )
    session.add(document)
    session.flush()
    document.sheet_key = str(_require_sheet_numeric_id(document))
    session.add(document)
    if workbook is not None:
        workbook_ref = _workbook_link_ref(workbook)
        sheet_ref = _sheet_link_ref(document)
        session.add(
            WorkbookSheetLink(
                workbook_id=workbook_ref,
                sheet_id=sheet_ref,
                order_index=_get_next_workbook_link_order(session, workbook_ref),
                created_at=now,
            ),
        )
        workbook.updated_by_user_id = current_user.id
        workbook.updated_at = now
        session.add(workbook)
    session.commit()
    session.refresh(document)
    workbook_items = _list_workbook_refs_for_sheet_ids(session, [document.id], current_user).get(document.id, [])
    parent_workbook_id = _get_parent_workbook_id_for_sheet(session, document)
    return NoteSheetDetailResponse.model_validate(
        _serialize_sheet_detail(
            document,
            workbook_items=workbook_items,
            parent_workbook_id=parent_workbook_id,
            access=_resolve_sheet_resource_access(session, document, current_user, workbook=workbook),
        ),
    )


@router.get("/sheets/{sheet_id}", response_model=NoteSheetDetailResponse)
def get_note_sheet(
    sheet_id: int,
    page: int = Query(default=1, ge=1),
    page_size: int | None = Query(default=None, ge=1, le=MAX_NOTE_SHEET_PAGE_SIZE),
    paginate: bool | None = Query(default=None),
    workbook_id: int | None = Query(default=None, ge=1),
    include_workbook_context: bool = Query(default=True),
    session: Session = Depends(get_session),
    current_user: User | None = Depends(get_optional_current_user_from_token),
):
    document, access, workbook = _get_note_sheet_or_404(
        session,
        current_user,
        sheet_id,
        required_role="viewer",
        workbook_id=workbook_id,
    )
    _sync_attendance_questionnaire_sheet_document_if_needed(session, document)
    stored_document = _ensure_sheet_document_identity_persisted(session, document)
    stored_document = _normalize_registration_sheet_header_persisted(session, document, stored_document)
    stored_document = _normalize_attendance_dual_clockin_refund_formulas_persisted(session, document, stored_document)
    workbook_items, parent_workbook_id = _get_sheet_workbook_context(
        session,
        document,
        current_user,
        workbook=workbook,
        include_workbook_context=include_workbook_context,
    )
    full_document = dict(stored_document)
    full_document, _header_link_count = _apply_course_attendance_header_links_for_response(session, document, full_document)
    full_document = _normalize_document_json(full_document)
    full_document, _progress_style_count = _apply_attendance_progress_backgrounds(full_document, assume_normalized=True)
    document_paginate_enabled, document_page_size = _get_normalized_document_pagination_settings(full_document)
    effective_paginate = document_paginate_enabled if paginate is None else paginate

    if effective_paginate:
        page_document, pagination = _build_paged_document(
            full_document,
            page=page,
            page_size=page_size if page_size is not None else document_page_size,
            assume_normalized=True,
        )
    else:
        page_document = full_document
        pagination = None

    return NoteSheetDetailResponse.model_validate(
        _serialize_sheet_detail(
            document,
            workbook_items=workbook_items,
            parent_workbook_id=parent_workbook_id,
            document_json=page_document,
            pagination=pagination,
            access=access,
        ),
    )


@router.post("/sheets/{sheet_id}/query")
def query_note_sheet(
    sheet_id: int,
    payload: NoteSheetQueryRequest,
    workbook_id: int | None = Query(default=None, ge=1),
    session: Session = Depends(get_session),
    current_user: User | None = Depends(get_optional_current_user_from_token),
):
    document, access, workbook = _get_note_sheet_or_404(
        session,
        current_user,
        sheet_id,
        required_role="viewer",
        workbook_id=workbook_id,
    )
    _sync_attendance_questionnaire_sheet_document(session, document)
    full_document = _normalize_registration_sheet_header_persisted(session, document)
    workbook_items, parent_workbook_id = _get_sheet_workbook_context(
        session,
        document,
        current_user,
        workbook=workbook,
        include_workbook_context=payload.include_workbook_context,
    )
    full_document, _header_link_count = _apply_course_attendance_header_links_for_response(session, document, full_document)
    full_document = _normalize_document_json(full_document)
    full_document, _progress_style_count = _apply_attendance_progress_backgrounds(full_document, assume_normalized=True)
    document_paginate_enabled, document_page_size = _get_normalized_document_pagination_settings(full_document)
    effective_paginate = document_paginate_enabled if payload.paginate is None else payload.paginate

    if effective_paginate:
        page_document, pagination = _build_filtered_paged_document(
            full_document,
            page=payload.page,
            page_size=payload.page_size if payload.page_size is not None else document_page_size,
            column_filters=payload.column_filters,
            row_filter_programs=payload.row_filter_programs,
            assume_normalized=True,
        )
    else:
        page_document = full_document
        pagination = None

    return _serialize_sheet_detail(
        document,
        workbook_items=workbook_items,
        parent_workbook_id=parent_workbook_id,
        document_json=page_document,
        pagination=pagination,
        access=access,
    )


@router.get("/sheets/{sheet_id}/column-options", response_model=NoteSheetColumnOptionsResponse)
def get_note_sheet_column_options(
    sheet_id: int,
    column_index: int = Query(..., ge=0),
    workbook_id: int | None = Query(default=None, ge=1),
    session: Session = Depends(get_session),
    current_user: User | None = Depends(get_optional_current_user_from_token),
):
    document, _access, _workbook = _get_note_sheet_or_404(
        session,
        current_user,
        sheet_id,
        required_role="viewer",
        workbook_id=workbook_id,
    )
    _sync_attendance_questionnaire_sheet_document(session, document)
    return _build_note_sheet_column_options_response(document, column_index=column_index)


@router.get("/sheets/{sheet_id}/table", response_model=NoteSheetTableResponse)
def get_note_sheet_table(
    sheet_id: int,
    workbook_id: int | None = Query(default=None, ge=1),
    include_grid: bool = Query(default=False),
    value_mode: Literal["text", "raw"] = Query(default="text"),
    session: Session = Depends(get_session),
    current_user: User | None = Depends(get_optional_current_user_from_token),
    trusted_device: Any | None = Depends(_get_optional_trusted_device),
):
    document, _access, workbook = _get_note_sheet_for_table_or_404(
        session,
        current_user,
        trusted_device,
        sheet_id,
        required_role="viewer",
        workbook_id=workbook_id,
    )
    return _build_note_sheet_table_response(
        document,
        workbook=workbook,
        include_grid=include_grid,
        value_mode=value_mode,
        defined_names=_defined_names_for_formula(session, document, workbook),
    )


@router.get("/sheets/{sheet_id}/attendance-export")
def export_note_sheet_attendance_table(
    sheet_id: int,
    workbook_id: int | None = Query(default=None, ge=1),
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_active_user),
):
    document, access, workbook = _get_note_sheet_or_404(
        session,
        current_user,
        sheet_id,
        required_role="editor",
        workbook_id=workbook_id,
    )
    if not access.capabilities.can_edit_data or not access.capabilities.can_run_sheet_actions:
        raise HTTPException(status_code=403, detail="没有执行表格动作的权限")

    target_workbook = _resolve_workbook_for_sheet(session, document, workbook)
    registration_document = (
        _get_workbook_sheet_by_key_or_title(
            session,
            target_workbook,
            sheet_key="registration",
            title="报名表",
        )
        if target_workbook is not None
        else None
    )
    if registration_document is not None:
        registration_access = _resolve_sheet_resource_access(
            session,
            registration_document,
            current_user,
            workbook=target_workbook,
        )
        if not registration_access.capabilities.can_read:
            raise HTTPException(status_code=403, detail="没有读取报名表的权限")

    table = _build_note_sheet_table_response(
        document,
        workbook=target_workbook,
        include_grid=False,
        value_mode="text",
        defined_names=_defined_names_for_formula(session, document, target_workbook),
    )
    raw_bytes = _build_attendance_export_workbook_bytes(
        table,
        phone_map=_build_registration_phone_map(registration_document),
    )
    filename = _attendance_export_filename(target_workbook, document)
    return StreamingResponse(
        io.BytesIO(raw_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}",
        },
    )


@router.post("/sheets/{sheet_id}/attendance/course-update-data", response_model=AttendanceCourseUpdateDataResponse)
def update_note_sheet_attendance_course_data(
    sheet_id: int,
    payload: AttendanceCourseUpdateDataRequest,
    workbook_id: int | None = Query(default=None, ge=1),
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_active_user),
):
    document, access, workbook = _get_note_sheet_or_404(
        session,
        current_user,
        sheet_id,
        required_role="editor",
        workbook_id=workbook_id,
    )
    if not access.capabilities.can_edit_data or not access.capabilities.can_run_sheet_actions:
        raise HTTPException(status_code=403, detail="没有执行表格动作的权限")

    if payload.course_type == "nianzhu":
        from backend.core.attendance.nianzhu_course_sheets import (
            compact_nianzhu_course_sheet_step2,
            rebuild_nianzhu_attendance_from_course_sheets,
        )

        step2_summary = compact_nianzhu_course_sheet_step2(
            session,
            attendance_sheet_id=sheet_id,
            course_name=payload.course_name,
        )
        step3_summary = rebuild_nianzhu_attendance_from_course_sheets(
            session,
            attendance_sheet_id=sheet_id,
            active_only=True,
            course_name=payload.course_name,
        )
        session.commit()
        step3_response = {
            "sheet_id": int(sheet_id),
            "course_name": payload.course_name,
            **step3_summary,
            "message": (
                f"当前 CodeYun 实例已执行念住闯关 step3："
                f"从课程存储 sheet 重建 {step3_summary.get('rows', 0)} 行，"
                f"更新 {step3_summary.get('updated_rows', 0)} 行/"
                f"{step3_summary.get('updated_cells', 0)} 格，"
                f"渲染 {step3_summary.get('styled_cells', 0)} 格"
            ),
        }
        return {"step2": {"step2": step2_summary, "rebuild": None}, "step3": step3_response}

    from backend.core.attendance.fanbei_schedule import (
        _run_fanbei_attendance_step2_local,
        run_fanbei_attendance_step3_for_sheet,
    )

    step2_message = _run_fanbei_attendance_step2_local()
    step3_response = run_fanbei_attendance_step3_for_sheet()
    return {"step2": {"message": step2_message}, "step3": step3_response}


@router.patch("/sheets/{sheet_id}/table", response_model=NoteSheetTablePatchResponse)
def patch_note_sheet_table(
    sheet_id: int,
    payload: NoteSheetTablePatchRequest,
    workbook_id: int | None = Query(default=None, ge=1),
    session: Session = Depends(get_session),
    current_user: User | None = Depends(get_optional_current_user_from_token),
    trusted_device: Any | None = Depends(_get_optional_trusted_device),
):
    document, access, workbook = _get_note_sheet_for_table_or_404(
        session,
        current_user,
        trusted_device,
        sheet_id,
        required_role="editor",
        workbook_id=workbook_id,
    )
    if not access.capabilities.can_edit_data:
        raise HTTPException(status_code=403, detail="没有该资源权限")
    if current_user is None and trusted_device is None:
        raise HTTPException(status_code=403, detail="没有该资源权限")
    if payload.expected_version is not None and int(document.version or 1) != payload.expected_version:
        raise HTTPException(status_code=409, detail="表格版本已变化，请重新读取后再写入")
    if not payload.operations:
        raise HTTPException(status_code=400, detail="缺少表格操作")

    current_document = dict(document.document_json or {})
    next_document, updated_cell_count, updated_row_count = _apply_note_sheet_table_patch(
        current_document,
        payload.operations,
    )
    if _is_registration_sheet(document):
        next_document, _registration_header_changed = _normalize_registration_sheet_header_document(next_document)
    next_document, _formula_repaired_count = _normalize_attendance_dual_clockin_refund_formulas(next_document)
    if next_document != current_document:
        document.document_json = next_document
        document.version = max(int(document.version or 1), 1) + 1
        document.updated_by_user_id = current_user.id if current_user is not None else None
        document.updated_at = time.time()
        session.add(document)
        session.commit()
        session.refresh(document)
        _broadcast_sheet_resource_update(document)

    workbook_items = _list_workbook_refs_for_sheet_ids(session, [document.id], current_user).get(document.id, [])
    parent_workbook_id = _get_parent_workbook_id_for_sheet(session, document)
    sheet = NoteSheetDetailResponse.model_validate(
        _serialize_sheet_detail(
            document,
            workbook_items=workbook_items,
            parent_workbook_id=parent_workbook_id,
            access=access,
        ),
    )
    return NoteSheetTablePatchResponse(
        sheet=sheet,
        table=_build_note_sheet_table_response(
            document,
            workbook=workbook,
            include_grid=False,
            defined_names=_defined_names_for_formula(session, document, workbook),
        ),
        updated_cell_count=updated_cell_count,
        updated_row_count=updated_row_count,
    )


@router.patch("/sheets/{sheet_id}/cells", response_model=NoteSheetCellPatchResponse)
def patch_note_sheet_cells(
    sheet_id: int,
    payload: NoteSheetCellPatchRequest,
    workbook_id: int | None = Query(default=None, ge=1),
    session: Session = Depends(get_session),
    current_user: User | None = Depends(get_optional_current_user_from_token),
):
    document, access, _workbook = _get_note_sheet_or_404(
        session,
        current_user,
        sheet_id,
        required_role="viewer",
        workbook_id=workbook_id,
    )
    editable_columns = list(access.capabilities.editable_data_columns)
    if not access.capabilities.can_edit_data and not editable_columns:
        raise HTTPException(status_code=403, detail="没有该资源权限")
    if access.capabilities.can_edit_data and current_user is None:
        raise HTTPException(status_code=403, detail="没有该资源权限")
    if payload.base_version is not None and int(payload.base_version) != int(document.version or 1):
        raise HTTPException(status_code=409, detail="表格版本已变化，请重新读取后再写入")
    if not payload.operations:
        raise HTTPException(status_code=400, detail="缺少单元格操作")

    current_document = dict(document.document_json or {})
    normalized = _normalize_document_json(current_document)
    columns = _normalize_document_columns(normalized)
    rows = _extract_document_rows(normalized)
    allowed_columns = {
        int(index)
        for index in editable_columns
        if 0 <= int(index) < len(columns)
    }
    can_edit_all_data = bool(access.capabilities.can_edit_data)
    next_rows = list(rows)
    updated_cell_count = 0

    for operation in payload.operations:
        row_index = int(operation.row_index)
        column_index = int(operation.column_index)
        if column_index < 0 or column_index >= len(columns):
            raise HTTPException(status_code=400, detail="列号超出表格范围")
        if row_index < 0 or row_index >= len(next_rows):
            raise HTTPException(status_code=400, detail="行号超出表格范围")
        if not can_edit_all_data and column_index not in allowed_columns:
            raise HTTPException(status_code=403, detail="游客只能编辑开放列")

        current_row = next_rows[row_index]
        current_cells = _normalize_sheet_row(current_row, len(columns))
        next_value = operation.value
        if _normalize_restricted_cell_value(current_cells[column_index]) == _normalize_restricted_cell_value(next_value):
            continue
        next_rows[row_index] = _set_row_cell_value(current_row, columns, column_index, next_value)
        updated_cell_count += 1

    if updated_cell_count > 0:
        next_document = _replace_document_data_rows(normalized, next_rows)
        next_document = _strip_formula_cell_rich_text(next_document)
        next_document = _remove_orphan_document_entity_cells(next_document)
        next_document, _formula_repaired_count = _normalize_attendance_dual_clockin_refund_formulas(next_document)
        if _is_registration_sheet(document):
            next_document, _registration_header_changed = _normalize_registration_sheet_header_document(next_document)
        if _is_attendance_questionnaire_data_sheet(document):
            next_document, _links_changed = _sync_attendance_questionnaire_course_links(session, next_document)
        document.document_json = next_document
        document.version = max(int(document.version or 1), 1) + 1
        document.updated_by_user_id = current_user.id if current_user is not None else None
        document.updated_at = time.time()
        session.add(document)
        session.commit()
        session.refresh(document)
        _broadcast_sheet_resource_update(document)

    if _is_attendance_questionnaire_data_sheet(document):
        _sync_attendance_questionnaire_entry_statuses(session, dict(document.document_json or {}))

    return NoteSheetCellPatchResponse(
        sheet_id=_require_sheet_numeric_id(document),
        version=int(document.version or 1),
        updated_cell_count=updated_cell_count,
    )


@router.post("/sheets/{sheet_id}/patch", response_model=NoteSheetPatchResponse)
def patch_note_sheet(
    sheet_id: int,
    payload: NoteSheetPatchRequest,
    workbook_id: int | None = Query(default=None, ge=1),
    session: Session = Depends(get_session),
    current_user: User | None = Depends(get_optional_current_user_from_token),
):
    document, access, _workbook = _get_note_sheet_or_404(
        session,
        current_user,
        sheet_id,
        required_role="viewer",
        workbook_id=workbook_id,
    )
    if int(payload.base_version) != int(document.version or 1):
        raise HTTPException(status_code=409, detail="表格版本已变化，请重新读取后再写入")
    if not payload.ops:
        raise HTTPException(status_code=400, detail="缺少表格操作")

    current_document = deepcopy(dict(document.document_json or {}))
    normalized = _normalize_document_json(current_document)
    columns = _normalize_document_columns(normalized)
    rows = _extract_document_rows(normalized)
    _validate_note_sheet_patch_access(
        access=access,
        current_user=current_user,
        operations=payload.ops,
        columns=columns,
        rows=rows,
    )

    next_document, updated_cell_count = _apply_note_sheet_patch_ops(normalized, payload.ops)
    next_document = _strip_formula_cell_rich_text(next_document)
    next_document = _remove_orphan_document_entity_cells(next_document)
    next_document, _formula_repaired_count = _normalize_attendance_dual_clockin_refund_formulas(next_document)
    if _is_registration_sheet(document):
        next_document, _registration_header_changed = _normalize_registration_sheet_header_document(next_document)
    if _is_attendance_questionnaire_data_sheet(document):
        next_document, _links_changed = _sync_attendance_questionnaire_course_links(session, next_document)

    if next_document != normalized:
        document.document_json = next_document
        document.version = max(int(document.version or 1), 1) + 1
        document.updated_by_user_id = current_user.id if current_user is not None else None
        document.updated_at = time.time()
        session.add(document)
        session.commit()
        session.refresh(document)
        _broadcast_sheet_resource_update(document)

    if _is_attendance_questionnaire_data_sheet(document):
        _sync_attendance_questionnaire_entry_statuses(session, dict(document.document_json or {}))

    return NoteSheetPatchResponse(
        sheet_id=_require_sheet_numeric_id(document),
        version=int(document.version or 1),
        applied_op_count=len(payload.ops),
        updated_cell_count=updated_cell_count,
    )


@router.websocket("/ws/resources/sheet/{sheet_id}")
async def websocket_sheet_resource_updates(websocket: WebSocket, sheet_id: int):
    room = _sheet_resource_update_room(sheet_id)
    await ws_manager.connect(websocket, room)
    try:
        while True:
            await websocket.receive_text()
    except WebSocketDisconnect:
        ws_manager.disconnect(websocket, room)
    except Exception:
        ws_manager.disconnect(websocket, room)


@router.get("/sheets/{sheet_id}/access", response_model=NoteSheetResourceAccessResponse)
def get_sheet_access(
    sheet_id: int,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_active_user),
):
    document, access, _workbook = _get_note_sheet_or_404(session, current_user, sheet_id, required_role="manager")
    return _build_resource_access_response(
        session,
        resource_type=RESOURCE_TYPE_SHEET,
        numeric_id=_require_sheet_numeric_id(document),
        resource_id=_sheet_resource_id(document),
        access=access,
    )


@router.get("/sheets/{sheet_id}/defined-names", response_model=NoteSheetDefinedNamesResponse)
def get_sheet_defined_names_endpoint(
    sheet_id: int,
    workbook_id: int | None = Query(default=None, ge=1),
    session: Session = Depends(get_session),
    current_user: User | None = Depends(get_optional_current_user_from_token),
):
    document, _access, workbook = _get_note_sheet_or_404(
        session,
        current_user,
        sheet_id,
        required_role="viewer",
        workbook_id=workbook_id,
    )
    workbook_names = _get_workbook_defined_names(session, workbook)
    worksheet_names = _get_sheet_defined_names(dict(document.document_json or {}))
    return NoteSheetDefinedNamesResponse(
        workbook_id=_require_workbook_numeric_id(workbook) if workbook is not None else None,
        sheet_id=_require_sheet_numeric_id(document),
        sheet_version=int(document.version or 1),
        workbook=workbook_names,
        worksheet=worksheet_names,
        worksheets=[{
            "sheet_id": _require_sheet_numeric_id(document),
            "sheet_title": document.title or "未命名工作表",
            "sheet_version": int(document.version or 1),
            "names": worksheet_names,
        }],
        effective=_merge_effective_defined_names(workbook_names, worksheet_names),
    )


@router.put("/sheets/{sheet_id}/defined-names", response_model=NoteSheetDefinedNamesResponse)
def update_sheet_defined_names_endpoint(
    sheet_id: int,
    payload: NoteSheetDefinedNamesUpdateRequest,
    workbook_id: int | None = Query(default=None, ge=1),
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_active_user),
):
    document, access, workbook = _get_note_sheet_or_404(
        session,
        current_user,
        sheet_id,
        required_role="editor",
        workbook_id=workbook_id,
    )
    if not access.capabilities.can_edit_config:
        raise HTTPException(status_code=403, detail="没有该资源权限")
    if payload.base_version is not None and int(payload.base_version) != int(document.version or 1):
        raise HTTPException(status_code=409, detail="表格版本已变化，请重新读取后再写入")

    current_document = dict(document.document_json or {})
    next_document = _replace_sheet_defined_names(current_document, payload.names)
    if next_document != current_document:
        document.document_json = next_document
        document.version = max(int(document.version or 1), 1) + 1
        document.updated_by_user_id = current_user.id
        document.updated_at = time.time()
        session.add(document)
        session.commit()
        session.refresh(document)
        _broadcast_sheet_resource_update(document)

    workbook_names = _get_workbook_defined_names(session, workbook)
    worksheet_names = _get_sheet_defined_names(dict(document.document_json or {}))
    return NoteSheetDefinedNamesResponse(
        workbook_id=_require_workbook_numeric_id(workbook) if workbook is not None else None,
        sheet_id=_require_sheet_numeric_id(document),
        sheet_version=int(document.version or 1),
        workbook=workbook_names,
        worksheet=worksheet_names,
        worksheets=[{
            "sheet_id": _require_sheet_numeric_id(document),
            "sheet_title": document.title or "未命名工作表",
            "sheet_version": int(document.version or 1),
            "names": worksheet_names,
        }],
        effective=_merge_effective_defined_names(workbook_names, worksheet_names),
    )


@router.put("/sheets/{sheet_id}/access", response_model=NoteSheetResourceAccessResponse)
def update_sheet_access(
    sheet_id: int,
    payload: NoteSheetResourceAccessUpdateRequest,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_active_user),
):
    document, _access, _workbook = _get_note_sheet_or_404(session, current_user, sheet_id, required_role="manager")
    _save_resource_access_grants(
        session,
        resource_type=RESOURCE_TYPE_SHEET,
        resource_id=_sheet_resource_id(document),
        payload=payload,
        current_user=current_user,
    )
    session.commit()
    access = _resolve_sheet_resource_access(session, document, current_user)
    return _build_resource_access_response(
        session,
        resource_type=RESOURCE_TYPE_SHEET,
        numeric_id=_require_sheet_numeric_id(document),
        resource_id=_sheet_resource_id(document),
        access=access,
    )


@router.put("/sheets/{sheet_id}", response_model=NoteSheetDetailResponse)
def update_note_sheet(
    sheet_id: int,
    payload: NoteSheetUpdateRequest,
    workbook_id: int | None = Query(default=None, ge=1),
    session: Session = Depends(get_session),
    current_user: User | None = Depends(get_optional_current_user_from_token),
):
    document, access, _workbook = _get_note_sheet_or_404(
        session,
        current_user,
        sheet_id,
        required_role="viewer",
        workbook_id=workbook_id,
    )
    editable_columns = list(access.capabilities.editable_data_columns)
    if not access.capabilities.can_edit_data and not editable_columns:
        raise HTTPException(status_code=403, detail="没有该资源权限")
    if access.capabilities.can_edit_data and current_user is None:
        raise HTTPException(status_code=403, detail="没有该资源权限")
    next_title = (
        _normalize_title(payload.title, default_value=document.title or "未命名表格")
        if payload.title is not None and access.capabilities.can_edit_config
        else document.title
    )
    if payload.base_version is not None and int(payload.base_version) != int(document.version or 1):
        raise HTTPException(status_code=409, detail="工作表已更新，请刷新后重试")

    current_document = dict(document.document_json or {})
    if payload.document_json is None:
        next_document = current_document
    else:
        _reject_default_blank_overwrite(current_document, payload.document_json)
        if access.capabilities.can_edit_data:
            if payload.page_patch is None:
                next_document = _normalize_document_json(payload.document_json)
            else:
                next_document = _merge_paged_document(current_document, payload.document_json, payload.page_patch)
        else:
            incoming_document = (
                _normalize_document_json(payload.document_json)
                if payload.page_patch is None
                else _merge_paged_document(current_document, payload.document_json, payload.page_patch)
            )
            next_document = _apply_restricted_data_column_update(current_document, incoming_document, editable_columns)

    next_document = _strip_formula_cell_rich_text(next_document)
    next_document = _remove_orphan_document_entity_cells(next_document)
    next_document, _formula_repaired_count = _normalize_attendance_dual_clockin_refund_formulas(next_document)
    if _is_registration_sheet(document):
        next_document, _registration_header_changed = _normalize_registration_sheet_header_document(next_document)

    if _is_attendance_summary_document(session, document):
        next_document, _online_links_preserved_count = _preserve_attendance_summary_online_sheet_links(
            current_document,
            next_document,
        )
        next_document, _online_links_repaired_count = _repair_attendance_summary_online_sheet_links(next_document)

    if _is_attendance_questionnaire_data_sheet(document):
        next_document, _links_changed = _sync_attendance_questionnaire_course_links(session, next_document)

    if document.title != next_title or current_document != next_document:
        document.title = next_title
        document.document_json = next_document
        document.version = max(int(document.version or 1), 1) + 1
        document.updated_by_user_id = current_user.id if current_user is not None else None
        document.updated_at = time.time()
        session.add(document)
        session.commit()
        session.refresh(document)
        _broadcast_sheet_resource_update(document)

    if _is_attendance_questionnaire_data_sheet(document):
        _sync_attendance_questionnaire_entry_statuses(session, dict(document.document_json or {}))

    workbook_items = _list_workbook_refs_for_sheet_ids(session, [document.id], current_user).get(document.id, [])
    parent_workbook_id = _get_parent_workbook_id_for_sheet(session, document)
    response_document = dict(document.document_json or {})
    response_pagination: NoteSheetPaginationResponse | None = None
    response_paginate_enabled, _response_page_size = _get_document_pagination_settings(response_document)
    if payload.document_json is not None and payload.page_patch is not None:
        response_document = _normalize_document_json(payload.document_json)
        if response_paginate_enabled:
            response_pagination = _build_workspace_pagination(
                page_patch=payload.page_patch,
                total_rows=len(_extract_document_rows(dict(document.document_json or {}))),
                current_row_count=len(_extract_document_rows(response_document)),
            )

    return NoteSheetDetailResponse.model_validate(
        _serialize_sheet_detail(
            document,
            workbook_items=workbook_items,
            parent_workbook_id=parent_workbook_id,
            document_json=response_document,
            pagination=response_pagination,
            access=access,
        ),
    )


@router.post("/sheets/{sheet_id}/import-excel-reset", response_model=NoteSheetExcelImportResponse)
async def import_note_sheet_excel_reset(
    sheet_id: int,
    file: UploadFile = File(...),
    instruction: str = Form(default=""),
    mode: Literal["append", "reset"] = Form(default="reset"),
    action_document_row: int | None = Form(default=None),
    action_column: int | None = Form(default=None),
    base_version: int | None = Form(default=None, ge=1),
    workbook_id: int | None = Query(default=None, ge=1),
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_active_user),
):
    document, access, _workbook = _get_note_sheet_or_404(
        session,
        current_user,
        sheet_id,
        required_role="editor",
        workbook_id=workbook_id,
    )
    if not access.capabilities.can_edit_data:
        raise HTTPException(status_code=403, detail="导入 Excel 需要完整编辑权限")
    if base_version is not None and int(base_version) != int(document.version or 1):
        raise HTTPException(status_code=409, detail="表格已被其他人更新，请刷新后重试")

    filename = str(file.filename or "").strip()
    suffix = Path(filename).suffix.lower()
    if suffix not in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        raise HTTPException(status_code=400, detail="请上传 .xlsx 或 .xlsm Excel 文件")

    raw_bytes = await file.read()
    workbook_payload = _extract_excel_workbook_payload(raw_bytes, filename or "未命名.xlsx")
    current_document = _normalize_document_json(dict(document.document_json or {}))
    import_rows, extra_columns, warnings, mapping_notes = _run_note_sheet_excel_import_deepseek(
        document_json=current_document,
        workbook_payload=workbook_payload,
        instruction=instruction,
        session=session,
        current_user=current_user,
        action_document_row=action_document_row,
        action_column=action_column,
    )
    effective_document, _effective_extra_columns = _append_document_extra_columns_for_excel_import(
        current_document,
        extra_columns,
    )
    effective_columns = _normalize_document_columns(effective_document)
    import_rows, grouped_sequence_count, source_field_count = _prefer_registration_group_sequences_from_workbook(
        workbook_payload=workbook_payload,
        rows=import_rows,
        columns=effective_columns,
    )
    if grouped_sequence_count:
        mapping_notes = [
            *mapping_notes,
            f"检测到源 Excel 中存在分组学号，已优先使用分组序号覆盖 {grouped_sequence_count} 行",
        ]
    if source_field_count:
        mapping_notes = [
            *mapping_notes,
            f"已按源 Excel 补正提交时间、商户订单号、订单金额等字段 {source_field_count} 行",
        ]
    skipped_duplicate_count = 0
    if mode == "append":
        import_rows, skipped_duplicate_count = _filter_duplicate_excel_import_payment_order_rows(
            effective_document,
            import_rows,
            effective_columns,
        )
        if import_rows:
            next_document, preserved_row_count = _append_document_rows_for_excel_import(
                current_document,
                import_rows,
                extra_columns=extra_columns,
            )
            imported_count = max(0, len(_extract_document_rows(next_document)) - len(_extract_document_rows(current_document)))
        else:
            next_document = current_document
            preserved_row_count = len(_extract_document_rows(current_document))
            imported_count = 0
            extra_columns = []
    else:
        next_document, preserved_row_count = _replace_document_rows_for_excel_import(
            current_document,
            import_rows,
            extra_columns=extra_columns,
            action_document_row=action_document_row,
            action_column=action_column,
        )
        imported_count = len(import_rows)

    if _is_attendance_questionnaire_data_sheet(document):
        next_document, _links_changed = _sync_attendance_questionnaire_course_links(session, next_document)

    next_document = _remove_orphan_document_entity_cells(next_document)

    if current_document != next_document:
        document.document_json = next_document
        document.version = max(int(document.version or 1), 1) + 1
        document.updated_by_user_id = current_user.id
        document.updated_at = time.time()
        session.add(document)
        session.commit()
        session.refresh(document)
        _broadcast_sheet_resource_update(document)

    workbook_items = _list_workbook_refs_for_sheet_ids(session, [document.id], current_user).get(document.id, [])
    parent_workbook_id = _get_parent_workbook_id_for_sheet(session, document)
    detail = NoteSheetDetailResponse.model_validate(
        _serialize_sheet_detail(
            document,
            workbook_items=workbook_items,
            parent_workbook_id=parent_workbook_id,
            document_json=dict(document.document_json or {}),
            access=access,
        ),
    )
    return NoteSheetExcelImportResponse(
        sheet=detail,
        imported_count=imported_count,
        preserved_row_count=preserved_row_count,
        skipped_duplicate_count=skipped_duplicate_count,
        extra_columns=extra_columns,
        warnings=warnings,
        mapping_notes=mapping_notes,
    )


@router.post(
    "/sheets/{sheet_id}/clockin/link-detection-runs",
    response_model=NoteSheetClockinLinkDetectionRunResponse,
)
def start_note_sheet_clockin_link_detection_run(
    sheet_id: int,
    payload: NoteSheetClockinLinkDetectionRunRequest,
    workbook_id: int | None = Query(default=None, ge=1),
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_active_user),
):
    document, access, _workbook = _get_note_sheet_or_404(
        session,
        current_user,
        sheet_id,
        required_role="editor",
        workbook_id=workbook_id,
    )
    if not access.capabilities.can_edit_data or not access.capabilities.can_run_sheet_actions:
        raise HTTPException(status_code=403, detail="没有执行打卡链接检测的权限")
    if current_user.id is None:
        raise HTTPException(status_code=403, detail="当前用户无效")
    runtime = resolve_ai_app_runtime_config(
        session=session,
        current_user=current_user,
        app_id=AI_APP_NOTE_SHEET_CLOCKIN_LINK_DETECTION,
    )
    return _start_clockin_link_detection_run(
        sheet_id=_require_sheet_numeric_id(document),
        workbook_id=workbook_id,
        current_user=current_user,
        provider_id=payload.provider_id or str(runtime.get("provider") or NOTE_SHEET_CLOCKIN_LINK_DETECTION_PROVIDER_ID),
        model=payload.model or str(runtime.get("model") or NOTE_SHEET_CLOCKIN_LINK_DETECTION_MODEL),
        force_restart=payload.force_restart,
    )


@router.get(
    "/sheets/{sheet_id}/clockin/link-detection-runs/active",
    response_model=NoteSheetClockinLinkDetectionRunResponse,
)
def get_note_sheet_active_clockin_link_detection_run(
    sheet_id: int,
    workbook_id: int | None = Query(default=None, ge=1),
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_active_user),
):
    document, access, _workbook = _get_note_sheet_or_404(
        session,
        current_user,
        sheet_id,
        required_role="viewer",
        workbook_id=workbook_id,
    )
    run = _get_active_clockin_link_detection_run_snapshot(_require_sheet_numeric_id(document))
    sheet = _serialize_note_sheet_action_detail(session, document, access, current_user) if run is not None else None
    return _serialize_clockin_link_detection_run(
        run,
        sheet=sheet,
        sheet_id=_require_sheet_numeric_id(document),
        workbook_id=workbook_id,
    )


@router.get(
    "/sheets/{sheet_id}/clockin/link-detection-runs/{run_id}",
    response_model=NoteSheetClockinLinkDetectionRunResponse,
)
def get_note_sheet_clockin_link_detection_run(
    sheet_id: int,
    run_id: str,
    workbook_id: int | None = Query(default=None, ge=1),
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_active_user),
):
    document, access, _workbook = _get_note_sheet_or_404(
        session,
        current_user,
        sheet_id,
        required_role="viewer",
        workbook_id=workbook_id,
    )
    run = _get_clockin_link_detection_run_snapshot(run_id)
    if run is None or int(run.get("sheet_id") or 0) != _require_sheet_numeric_id(document):
        raise HTTPException(status_code=404, detail="打卡链接检测任务不存在")
    sheet = _serialize_note_sheet_action_detail(session, document, access, current_user)
    return _serialize_clockin_link_detection_run(run, sheet=sheet)


@router.post(
    "/sheets/{sheet_id}/registration/match-runs",
    response_model=NoteSheetRegistrationMatchRunResponse,
)
def start_note_sheet_registration_match_run(
    sheet_id: int,
    payload: NoteSheetRegistrationMatchRunRequest,
    workbook_id: int | None = Query(default=None, ge=1),
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_active_user),
):
    document, access, _workbook = _get_note_sheet_or_404(
        session,
        current_user,
        sheet_id,
        required_role="editor",
        workbook_id=workbook_id,
    )
    if not access.capabilities.can_edit_data or not access.capabilities.can_run_sheet_actions:
        raise HTTPException(status_code=403, detail="没有执行报名表动作的权限")
    if current_user.id is None:
        raise HTTPException(status_code=403, detail="当前用户无效")

    run_response = _start_registration_match_run(
        sheet_id=_require_sheet_numeric_id(document),
        workbook_id=workbook_id,
        action=payload.action,
        current_user=current_user,
        use_browser_fallback=(
            payload.use_browser_fallback
            if payload.use_browser_fallback is not None
            else _default_registration_match_browser_fallback(payload.action)
        ),
        force_restart=payload.force_restart,
    )
    return run_response


@router.get(
    "/sheets/{sheet_id}/registration/match-runs/active",
    response_model=NoteSheetRegistrationMatchRunResponse,
)
def get_note_sheet_active_registration_match_run(
    sheet_id: int,
    action: Literal["registration_order_match", "registration_user_match", "registration_composite_update"] | None = Query(default=None),
    workbook_id: int | None = Query(default=None, ge=1),
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_active_user),
):
    document, access, _workbook = _get_note_sheet_or_404(
        session,
        current_user,
        sheet_id,
        required_role="viewer",
        workbook_id=workbook_id,
    )
    run = _get_active_registration_match_run_snapshot(_require_sheet_numeric_id(document), action)
    sheet = _serialize_note_sheet_action_detail(session, document, access, current_user) if run is not None else None
    return _serialize_registration_match_run(
        run,
        sheet=sheet,
        sheet_id=_require_sheet_numeric_id(document),
        action=action or "",
        workbook_id=workbook_id,
    )


@router.get(
    "/sheets/{sheet_id}/registration/match-runs/{run_id}",
    response_model=NoteSheetRegistrationMatchRunResponse,
)
def get_note_sheet_registration_match_run(
    sheet_id: int,
    run_id: str,
    workbook_id: int | None = Query(default=None, ge=1),
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_active_user),
):
    document, access, _workbook = _get_note_sheet_or_404(
        session,
        current_user,
        sheet_id,
        required_role="viewer",
        workbook_id=workbook_id,
    )
    run = _get_registration_match_run_snapshot(run_id)
    if run is None or int(run.get("sheet_id") or 0) != _require_sheet_numeric_id(document):
        raise HTTPException(status_code=404, detail="匹配任务不存在")
    sheet = _serialize_note_sheet_action_detail(session, document, access, current_user)
    return _serialize_registration_match_run(run, sheet=sheet)


@router.post(
    "/sheets/{sheet_id}/registration/update-order-match",
    response_model=NoteSheetRegistrationMatchResponse,
)
def update_note_sheet_registration_order_match(
    sheet_id: int,
    use_browser_fallback: bool = Query(default=True),
    workbook_id: int | None = Query(default=None, ge=1),
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_active_user),
):
    return _run_registration_match_action(
        sheet_id=sheet_id,
        action=NOTE_SHEET_CELL_ACTION_REGISTRATION_ORDER_MATCH,
        workbook_id=workbook_id,
        session=session,
        current_user=current_user,
        use_browser_fallback=use_browser_fallback,
    )


@router.post(
    "/sheets/{sheet_id}/registration/update-user-match",
    response_model=NoteSheetRegistrationMatchResponse,
)
def update_note_sheet_registration_user_match(
    sheet_id: int,
    workbook_id: int | None = Query(default=None, ge=1),
    use_browser_fallback: bool = Query(default=NOTE_SHEET_REGISTRATION_USER_BROWSER_FALLBACK_DEFAULT),
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_active_user),
):
    return _run_registration_match_action(
        sheet_id=sheet_id,
        action=NOTE_SHEET_CELL_ACTION_REGISTRATION_USER_MATCH,
        workbook_id=workbook_id,
        session=session,
        current_user=current_user,
        use_browser_fallback=use_browser_fallback,
    )


@router.post(
    "/sheets/{sheet_id}/registration/detect-user-id",
    response_model=NoteSheetRegistrationUserIdDetectionResponse,
)
def detect_note_sheet_registration_user_id(
    sheet_id: int,
    payload: NoteSheetRegistrationUserIdDetectionRequest,
    workbook_id: int | None = Query(default=None, ge=1),
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_active_user),
):
    document, access, workbook = _get_note_sheet_or_404(
        session,
        current_user,
        sheet_id,
        required_role="editor",
        workbook_id=workbook_id,
    )
    if not access.capabilities.can_edit_data or not access.capabilities.can_run_sheet_actions:
        raise HTTPException(status_code=403, detail="没有执行报名表动作的权限")
    if payload.base_version is not None and int(payload.base_version) != int(document.version or 1):
        raise HTTPException(status_code=409, detail="表格已被其他人更新，请刷新后重试")
    if workbook is None:
        raise HTTPException(status_code=400, detail="检测用户ID需要在工作簿内执行")

    attendance = _get_workbook_sheet_by_key_or_title(
        session,
        workbook,
        sheet_key="attendance",
        title="考勤表",
    )
    video_data = _get_workbook_sheet_by_key_or_title(
        session,
        workbook,
        sheet_key="video_data",
        title="视频数据",
    )
    clockin_data = _get_workbook_sheet_by_key_or_title(
        session,
        workbook,
        sheet_key="clockin_data",
        title="打卡数据",
    )
    if attendance is None or (video_data is None and clockin_data is None):
        raise HTTPException(status_code=404, detail="当前工作簿缺少考勤表或课程数据表")

    current_document = _normalize_document_json(dict(document.document_json or {}))
    next_document, _linked_index = _ensure_registration_linked_user_id_column(current_document)
    columns = _normalize_document_columns(next_document)
    required_columns = ["用户ID", NOTE_SHEET_REGISTRATION_LINKED_USER_ID_COLUMN]
    missing_columns = [column for column in required_columns if column not in columns]
    if missing_columns:
        raise HTTPException(status_code=400, detail=f"报名表缺少字段：{', '.join(missing_columns)}")
    indexes = {column: columns.index(column) for column in columns}
    rows = [_normalize_sheet_row(row, len(columns)) for row in _extract_document_rows(next_document)]
    if payload.row_index >= len(rows):
        raise HTTPException(status_code=400, detail="报名表行不存在")

    target_row = list(rows[payload.row_index])
    row_map = dict(zip(columns, target_row))
    progress = _collect_registration_course_user_progress(
        dict(video_data.document_json or {}) if video_data is not None else None,
        dict(clockin_data.document_json or {}) if clockin_data is not None else None,
    )

    primary_user_id = _normalize_sheet_text(row_map.get("用户ID"))
    linked_user_ids: list[str] = []
    seen_linked_user_ids: set[str] = set()
    for user_id in re.split(
        r"[,，;；\s]+",
        _normalize_sheet_text(row_map.get(NOTE_SHEET_REGISTRATION_LINKED_USER_ID_COLUMN)),
    ):
        normalized_user_id = _normalize_sheet_text(user_id)
        if normalized_user_id and normalized_user_id not in seen_linked_user_ids:
            seen_linked_user_ids.add(normalized_user_id)
            linked_user_ids.append(normalized_user_id)

    primary_has_progress = bool(
        primary_user_id
        and int(progress.get(primary_user_id, {}).get("video_count") or 0)
        + int(progress.get(primary_user_id, {}).get("clockin_count") or 0)
        > 0
    )
    linked_progress_user_ids = [
        user_id
        for user_id in linked_user_ids
        if int(progress.get(user_id, {}).get("video_count") or 0) + int(progress.get(user_id, {}).get("clockin_count") or 0) > 0
    ]
    if primary_has_progress and linked_progress_user_ids:
        _append_registration_detection_note(
            target_row,
            indexes,
            f"当前用户ID和关联用户ID均有课程数据，未改ID，已重建考勤表："
            f"{primary_user_id};{';'.join(linked_progress_user_ids)}",
        )
        rows[payload.row_index] = target_row
        next_document = _replace_document_data_rows(next_document, rows)
        if next_document != current_document:
            document.document_json = next_document
            document.version = max(int(document.version or 1), 1) + 1
            document.updated_by_user_id = current_user.id
            document.updated_at = time.time()
            session.add(document)
        rebuild_summary = _rebuild_registration_attendance_after_user_id_detection(
            session,
            attendance=attendance,
            course_name=_get_registration_course_name(document, workbook),
        )
        attendance.updated_by_user_id = current_user.id
        attendance.updated_at = time.time()
        session.add(attendance)
        session.commit()
        session.refresh(document)
        session.refresh(attendance)
        _broadcast_sheet_resource_update(document)
        return NoteSheetRegistrationUserIdDetectionResponse(
            sheet=_serialize_note_sheet_action_detail(session, document, access, current_user),
            attendance_sheet=None,
            status="skipped",
            message="当前用户ID和关联用户ID均有课程数据，未改ID，已重建考勤表",
            rebuild_summary=rebuild_summary,
        )

    if len(linked_progress_user_ids) == 1:
        target_user_id = linked_progress_user_ids[0]
        user_id_index = indexes["用户ID"]
        linked_index = indexes[NOTE_SHEET_REGISTRATION_LINKED_USER_ID_COLUMN]
        previous_user_id = _normalize_sheet_text(target_row[user_id_index])
        target_row[user_id_index] = target_user_id
        target_row[linked_index] = ";".join(user_id for user_id in linked_user_ids if user_id != target_user_id)
        _append_registration_detection_note(
            target_row,
            indexes,
            f"关联用户ID已有唯一课程数据，已提升为用户ID：{target_user_id}"
            f"{f'；原用户ID：{previous_user_id}' if previous_user_id else ''}",
        )
        rows[payload.row_index] = target_row
        next_document = _replace_document_data_rows(next_document, rows)
        document.document_json = next_document
        document.version = max(int(document.version or 1), 1) + 1
        document.updated_by_user_id = current_user.id
        document.updated_at = time.time()
        session.add(document)
        rebuild_summary = _rebuild_registration_attendance_after_user_id_detection(
            session,
            attendance=attendance,
            course_name=_get_registration_course_name(document, workbook),
        )
        attendance.updated_by_user_id = current_user.id
        attendance.updated_at = time.time()
        session.add(attendance)
        session.commit()
        session.refresh(document)
        session.refresh(attendance)
        _broadcast_sheet_resource_update(document)
        return NoteSheetRegistrationUserIdDetectionResponse(
            sheet=_serialize_note_sheet_action_detail(session, document, access, current_user),
            attendance_sheet=None,
            status="applied",
            applied=True,
            message="已将关联用户ID提升为用户ID并重建考勤表",
            target_user_id=target_user_id,
            applied_to="用户ID",
            rebuild_summary=rebuild_summary,
        )

    candidates = _build_registration_user_id_detection_candidates(row_map, progress)
    high_candidates = [candidate for candidate in candidates if candidate.confidence == "high"]
    applied = False
    applied_to: Literal["用户ID", "关联用户ID", ""] = ""
    target_user_id = ""
    status: Literal["applied", "review", "skipped", "error"] = "review"
    message = ""
    rebuild_summary: dict[str, Any] | None = None

    if len(high_candidates) == 1:
        candidate = high_candidates[0]
        target_user_id = candidate.user_id
        user_id_index = indexes["用户ID"]
        linked_index = indexes[NOTE_SHEET_REGISTRATION_LINKED_USER_ID_COLUMN]
        previous_user_id = _normalize_sheet_text(target_row[user_id_index])
        if previous_user_id != target_user_id:
            if primary_has_progress:
                next_linked_user_ids = [
                    user_id
                    for user_id in linked_user_ids
                    if user_id and user_id != previous_user_id and user_id != target_user_id
                ]
                next_linked_user_ids.append(target_user_id)
                target_row[linked_index] = ";".join(dict.fromkeys(next_linked_user_ids))
                applied_to = "关联用户ID"
            else:
                target_row[user_id_index] = target_user_id
                target_row[linked_index] = ";".join(user_id for user_id in linked_user_ids if user_id != target_user_id)
                applied_to = "用户ID"
        else:
            applied_to = ""

        _append_registration_detection_note(
            target_row,
            indexes,
            f"已按唯一高置信候选写入{applied_to or '现有ID'}：{target_user_id}；"
            f"依据：{_format_registration_detection_candidate_summary([candidate])}"
            f"{f'；原用户ID：{previous_user_id}' if applied_to and previous_user_id else ''}",
        )
        applied = bool(applied_to)
        status = "applied"
        message = f"已写入{applied_to or '现有ID'}并重建考勤表"
    elif high_candidates:
        _append_registration_detection_note(
            target_row,
            indexes,
            f"发现多个高置信候选，未自动写入；请人工确认："
            f"{_format_registration_detection_candidate_summary(high_candidates)}",
        )
        message = "发现多个高置信候选，已写入参考信息"
    elif candidates:
        _append_registration_detection_note(
            target_row,
            indexes,
            f"仅找到姓名/昵称等中低置信候选，未自动写入；请人工确认："
            f"{_format_registration_detection_candidate_summary(candidates)}",
        )
        message = "候选证据不够明确，已写入参考信息"
    else:
        _append_registration_detection_note(
            target_row,
            indexes,
            "当前ID无课程数据，且未找到有课程数据的明确候选",
        )
        status = "skipped"
        message = "未找到可用候选，已写入参考信息"

    rows[payload.row_index] = target_row
    next_document = _replace_document_data_rows(next_document, rows)
    document_changed = next_document != current_document
    if document_changed:
        document.document_json = next_document
        document.version = max(int(document.version or 1), 1) + 1
        document.updated_by_user_id = current_user.id
        document.updated_at = time.time()
        session.add(document)

    if applied:
        rebuild_summary = _rebuild_registration_attendance_after_user_id_detection(
            session,
            attendance=attendance,
            course_name=_get_registration_course_name(document, workbook),
        )
        attendance.updated_by_user_id = current_user.id
        attendance.updated_at = time.time()
        session.add(attendance)

    session.commit()
    session.refresh(document)
    if document_changed:
        _broadcast_sheet_resource_update(document)
    if applied:
        session.refresh(attendance)

    attendance_sheet: NoteSheetDetailResponse | None = None
    if applied:
        try:
            attendance_document, attendance_access, _ = _get_note_sheet_or_404(
                session,
                current_user,
                int(attendance.numeric_id or attendance.id),
                required_role="viewer",
                workbook_id=workbook_id,
            )
            attendance_sheet = _serialize_note_sheet_action_detail(
                session,
                attendance_document,
                attendance_access,
                current_user,
            )
        except HTTPException:
            attendance_sheet = None

    return NoteSheetRegistrationUserIdDetectionResponse(
        sheet=_serialize_note_sheet_action_detail(session, document, access, current_user),
        attendance_sheet=attendance_sheet,
        status=status,
        applied=applied,
        message=message,
        target_user_id=target_user_id,
        applied_to=applied_to,
        candidates=candidates,
        rebuild_summary=rebuild_summary,
    )


@router.post("/sheets/{sheet_id}/sort", response_model=NoteSheetDetailResponse)
def sort_note_sheet(
    sheet_id: int,
    payload: NoteSheetSortRequest,
    workbook_id: int | None = Query(default=None, ge=1),
    session: Session = Depends(get_session),
    current_user: User | None = Depends(get_optional_current_user_from_token),
):
    document, access, _workbook = _get_note_sheet_or_404(
        session,
        current_user,
        sheet_id,
        required_role="editor",
        workbook_id=workbook_id,
    )
    if current_user is None:
        raise HTTPException(status_code=403, detail="没有该资源权限")
    if payload.base_version is not None and int(payload.base_version) != int(document.version or 1):
        raise HTTPException(status_code=409, detail="表格版本已变化，请重新读取后再写入")
    current_document = _normalize_document_json(dict(document.document_json or {}))
    columns = list(current_document.get("columns") or [])
    if payload.column_index >= len(columns):
        raise HTTPException(status_code=400, detail="排序字段不存在")

    next_document = _sort_sheet_document_rows(
        current_document,
        column_index=payload.column_index,
        direction=payload.direction,
    )

    if current_document != next_document:
        document.document_json = next_document
        document.version = max(int(document.version or 1), 1) + 1
        document.updated_by_user_id = current_user.id
        document.updated_at = time.time()
        session.add(document)
        session.commit()
        session.refresh(document)
        _broadcast_sheet_resource_update(document)

    workbook_items = _list_workbook_refs_for_sheet_ids(session, [document.id], current_user).get(document.id, [])
    parent_workbook_id = _get_parent_workbook_id_for_sheet(session, document)
    paginate_enabled, page_size = _get_document_pagination_settings(next_document)
    if paginate_enabled:
        response_document, pagination = _build_paged_document(next_document, page=1, page_size=page_size)
    else:
        response_document = _normalize_document_json(next_document)
        pagination = None

    return NoteSheetDetailResponse.model_validate(
        _serialize_sheet_detail(
            document,
            workbook_items=workbook_items,
            parent_workbook_id=parent_workbook_id,
            document_json=response_document,
            pagination=pagination,
            access=access,
        ),
    )


@router.post(
    "/sheets/{sheet_id}/attendance-summary/generate-next-month-templates",
    response_model=NoteSheetAttendanceTemplateGenerationResponse,
)
def generate_attendance_summary_next_month_templates(
    sheet_id: int,
    payload: NoteSheetAttendanceTemplateGenerationRequest = NoteSheetAttendanceTemplateGenerationRequest(),
    workbook_id: int | None = Query(default=None, ge=1),
    session: Session = Depends(get_session),
    current_user: User | None = Depends(get_optional_current_user_from_token),
):
    document, access, _workbook = _get_note_sheet_or_404(
        session,
        current_user,
        sheet_id,
        required_role="editor",
        workbook_id=workbook_id,
    )
    if current_user is None:
        raise HTTPException(status_code=403, detail="没有该资源权限")
    if not _is_attendance_summary_document(session, document):
        raise HTTPException(status_code=404, detail="当前工作表没有这个定制功能")
    _check_sheet_base_version(document, payload.base_version)

    target_date = _get_attendance_template_target_date(payload)
    current_document = _normalize_document_json(dict(document.document_json or {}))
    next_document, generated, skipped = _generate_attendance_next_month_templates(
        current_document,
        target_date=target_date,
        skip_course_types=_normalize_attendance_template_skip_course_types(payload.skip_course_types),
    )
    normalized_skip_course_types = _normalize_attendance_template_skip_course_types(payload.skip_course_types)
    targets = [
        (course_type, course_target_date)
        for course_type, course_target_date in _get_attendance_batch_course_targets(target_date)
        if course_type not in normalized_skip_course_types
    ]
    next_document, materialized_count = _materialize_attendance_template_workbooks_for_targets(
        session,
        source_document_json=current_document,
        generated_document_json=next_document,
        generated=generated,
        targets=targets,
        owner_user_id=current_user.id,
    )

    if (generated or materialized_count) and current_document != next_document:
        document.document_json = next_document
        document.version = max(int(document.version or 1), 1) + 1
        document.updated_by_user_id = current_user.id
        document.updated_at = time.time()
        session.add(document)
        session.commit()
        session.refresh(document)
        _broadcast_sheet_resource_update(document)
    else:
        next_document = current_document

    return NoteSheetAttendanceTemplateGenerationResponse(
        sheet=_build_attendance_template_detail_response(
            session,
            document,
            next_document,
            access=access,
            current_user=current_user,
        ),
        generated=generated,
        skipped=skipped,
    )


@router.post(
    "/sheets/{sheet_id}/attendance-summary/generate-course-template",
    response_model=NoteSheetAttendanceTemplateGenerationResponse,
)
def generate_attendance_summary_course_template(
    sheet_id: int,
    payload: NoteSheetAttendanceCourseTemplateGenerationRequest,
    workbook_id: int | None = Query(default=None, ge=1),
    session: Session = Depends(get_session),
    current_user: User | None = Depends(get_optional_current_user_from_token),
):
    document, access, _workbook = _get_note_sheet_or_404(
        session,
        current_user,
        sheet_id,
        required_role="editor",
        workbook_id=workbook_id,
    )
    if current_user is None:
        raise HTTPException(status_code=403, detail="没有该资源权限")
    if not _is_attendance_summary_document(session, document):
        raise HTTPException(status_code=404, detail="当前工作表没有这个定制功能")
    _check_sheet_base_version(document, payload.base_version)

    current_document = _normalize_document_json(dict(document.document_json or {}))
    course_type = _resolve_attendance_course_template_type(current_document, payload)
    target_date = _get_attendance_course_template_target_date(
        course_type,
        payload,
        document_json=current_document,
    )
    next_document, generated, skipped = _generate_attendance_course_templates(
        current_document,
        targets=[(course_type, target_date)],
    )
    next_document, materialized_count = _materialize_attendance_template_workbooks_for_targets(
        session,
        source_document_json=current_document,
        generated_document_json=next_document,
        generated=generated,
        targets=[(course_type, target_date)],
        owner_user_id=current_user.id,
    )

    if (generated or materialized_count) and current_document != next_document:
        document.document_json = next_document
        document.version = max(int(document.version or 1), 1) + 1
        document.updated_by_user_id = current_user.id
        document.updated_at = time.time()
        session.add(document)
        session.commit()
        session.refresh(document)
        _broadcast_sheet_resource_update(document)
    else:
        next_document = current_document

    return NoteSheetAttendanceTemplateGenerationResponse(
        sheet=_build_attendance_template_detail_response(
            session,
            document,
            next_document,
            access=access,
            current_user=current_user,
        ),
        generated=generated,
        skipped=skipped,
    )


@router.post(
    "/sheets/{sheet_id}/attendance-summary/repair-cell-meta",
    response_model=NoteSheetAttendanceTemplateGenerationResponse,
)
def repair_attendance_summary_cell_meta(
    sheet_id: int,
    payload: NoteSheetAttendanceTemplateGenerationRequest = NoteSheetAttendanceTemplateGenerationRequest(),
    workbook_id: int | None = Query(default=None, ge=1),
    session: Session = Depends(get_session),
    current_user: User | None = Depends(get_optional_current_user_from_token),
):
    """修复因插入课程模板导致的 cell_meta/entity_rows 行错位。"""
    document, access, _workbook = _get_note_sheet_or_404(
        session,
        current_user,
        sheet_id,
        required_role="editor",
        workbook_id=workbook_id,
    )
    if current_user is None:
        raise HTTPException(status_code=403, detail="没有该资源权限")
    if not _is_attendance_summary_document(session, document):
        raise HTTPException(status_code=404, detail="当前工作表没有这个定制功能")
    _check_sheet_base_version(document, payload.base_version)

    current_document = _normalize_document_json(dict(document.document_json or {}))
    next_document, repaired = _repair_attendance_summary_cell_meta(current_document)

    if repaired:
        document.document_json = next_document
        document.version = max(int(document.version or 1), 1) + 1
        document.updated_by_user_id = current_user.id
        document.updated_at = time.time()
        session.add(document)
        session.commit()
        session.refresh(document)
        _broadcast_sheet_resource_update(document)
    else:
        next_document = current_document

    return NoteSheetAttendanceTemplateGenerationResponse(
        sheet=_build_attendance_template_detail_response(
            session,
            document,
            next_document,
            access=access,
            current_user=current_user,
        ),
        generated=[],
        skipped=[],
    )


@router.get(
    "/sheets/{sheet_id}/attendance-summary/course-script-statuses",
    response_model=NoteSheetAttendanceCourseScriptStatusesResponse,
)
def list_attendance_summary_course_script_statuses(
    sheet_id: int,
    workbook_id: int | None = Query(default=None, ge=1),
    session: Session = Depends(get_session),
    current_user: User | None = Depends(get_optional_current_user_from_token),
):
    document, _access, _workbook = _get_note_sheet_or_404(
        session,
        current_user,
        sheet_id,
        required_role="editor",
        workbook_id=workbook_id,
    )
    if current_user is None:
        raise HTTPException(status_code=403, detail="没有该资源权限")
    if not _is_attendance_summary_document(session, document):
        raise HTTPException(status_code=404, detail="当前工作表没有这个定制功能")

    current_document = _normalize_document_json(dict(document.document_json or {}))
    return NoteSheetAttendanceCourseScriptStatusesResponse(
        statuses=_list_attendance_course_script_statuses(current_document),
    )


@router.post(
    "/sheets/{sheet_id}/attendance-summary/generate-course-script",
    response_model=NoteSheetAttendanceCourseScriptGenerationResponse,
)
def generate_attendance_summary_course_script(
    sheet_id: int,
    payload: NoteSheetAttendanceCourseScriptGenerationRequest,
    workbook_id: int | None = Query(default=None, ge=1),
    session: Session = Depends(get_session),
    current_user: User | None = Depends(get_optional_current_user_from_token),
):
    document, _access, _workbook = _get_note_sheet_or_404(
        session,
        current_user,
        sheet_id,
        required_role="editor",
        workbook_id=workbook_id,
    )
    if current_user is None:
        raise HTTPException(status_code=403, detail="没有该资源权限")
    if not _is_attendance_summary_document(session, document):
        raise HTTPException(status_code=404, detail="当前工作表没有这个定制功能")

    current_document = _normalize_document_json(dict(document.document_json or {}))
    return _generate_attendance_course_script(current_document, row_index=payload.row_index)


@router.post(
    "/sheets/{sheet_id}/attendance-summary/organize-course-scripts",
    response_model=NoteSheetAttendanceCourseScriptOrganizeResponse,
)
def organize_attendance_summary_course_scripts(
    sheet_id: int,
    workbook_id: int | None = Query(default=None, ge=1),
    session: Session = Depends(get_session),
    current_user: User | None = Depends(get_optional_current_user_from_token),
):
    document, _access, _workbook = _get_note_sheet_or_404(
        session,
        current_user,
        sheet_id,
        required_role="editor",
        workbook_id=workbook_id,
    )
    if current_user is None:
        raise HTTPException(status_code=403, detail="没有该资源权限")
    if not _is_attendance_summary_document(session, document):
        raise HTTPException(status_code=404, detail="当前工作表没有这个定制功能")

    current_document = _normalize_document_json(dict(document.document_json or {}))
    return _organize_attendance_course_scripts(current_document)


@router.post(
    "/sheets/{sheet_id}/attendance-summary/update-link-counts",
    response_model=NoteSheetAttendanceLinkCountUpdateResponse,
)
def update_attendance_summary_link_counts(
    sheet_id: int,
    payload: NoteSheetAttendanceLinkCountUpdateRequest,
    workbook_id: int | None = Query(default=None, ge=1),
    session: Session = Depends(get_session),
    current_user: User | None = Depends(get_optional_current_user_from_token),
):
    document, access, _workbook = _get_note_sheet_or_404(
        session,
        current_user,
        sheet_id,
        required_role="editor",
        workbook_id=workbook_id,
    )
    if current_user is None:
        raise HTTPException(status_code=403, detail="没有该资源权限")
    if not _is_attendance_summary_document(session, document):
        raise HTTPException(status_code=404, detail="当前工作表没有这个定制功能")
    _check_sheet_base_version(document, payload.base_version)

    current_document = _normalize_document_json(dict(document.document_json or {}))
    next_document, updated, skipped = _update_attendance_link_counts(
        session,
        current_document,
        field_key=payload.field_key,
        row_index=payload.row_index,
        repair_with_remote_browser=payload.repair_with_remote_browser,
    )

    if current_document != next_document:
        document.document_json = next_document
        document.version = max(int(document.version or 1), 1) + 1
        document.updated_by_user_id = current_user.id
        document.updated_at = time.time()
        session.add(document)
        session.commit()
        session.refresh(document)
        _broadcast_sheet_resource_update(document)
    else:
        next_document = current_document

    return NoteSheetAttendanceLinkCountUpdateResponse(
        sheet=_build_attendance_template_detail_response(
            session,
            document,
            next_document,
            access=access,
            current_user=current_user,
        ),
        updated=updated,
        skipped=skipped,
    )


@router.post(
    "/sheets/{sheet_id}/attendance-summary/set-completed",
    response_model=NoteSheetAttendanceCompletionResponse,
)
def set_attendance_summary_row_completed(
    sheet_id: int,
    payload: NoteSheetAttendanceCompletionRequest,
    workbook_id: int | None = Query(default=None, ge=1),
    session: Session = Depends(get_session),
    current_user: User | None = Depends(get_optional_current_user_from_token),
):
    document, access, _workbook = _get_note_sheet_or_404(
        session,
        current_user,
        sheet_id,
        required_role="editor",
        workbook_id=workbook_id,
    )
    if current_user is None:
        raise HTTPException(status_code=403, detail="没有该资源权限")
    if not _is_attendance_summary_document(session, document):
        raise HTTPException(status_code=404, detail="当前工作表没有这个定制功能")
    _check_sheet_base_version(document, payload.base_version)

    completion_date = _parse_attendance_date_text(payload.completion_date) if payload.completion_date else date.today()
    if completion_date is None:
        raise HTTPException(status_code=400, detail="completion_date 格式不正确")

    current_document = _normalize_document_json(dict(document.document_json or {}))
    next_document, next_row_index = _set_attendance_summary_row_completed(
        current_document,
        row_index=payload.row_index,
        completion_date=completion_date,
    )

    if current_document != next_document:
        document.document_json = next_document
        document.version = max(int(document.version or 1), 1) + 1
        document.updated_by_user_id = current_user.id
        document.updated_at = time.time()
        session.add(document)
        session.commit()
        session.refresh(document)
        _broadcast_sheet_resource_update(document)
    else:
        next_document = current_document

    return NoteSheetAttendanceCompletionResponse(
        sheet=_build_attendance_template_detail_response(
            session,
            document,
            next_document,
            access=access,
            current_user=current_user,
        ),
        row_index=next_row_index,
    )


@router.post(
    "/sheets/{sheet_id}/attendance/video-revision",
    response_model=NoteSheetAttendanceVideoRevisionResponse,
)
def revise_attendance_video_progress(
    sheet_id: int,
    payload: NoteSheetAttendanceVideoRevisionRequest,
    workbook_id: int | None = Query(default=None, ge=1),
    session: Session = Depends(get_session),
    current_user: User | None = Depends(get_optional_current_user_from_token),
):
    document, access, _workbook = _get_note_sheet_or_404(
        session,
        current_user,
        sheet_id,
        required_role="editor",
        workbook_id=workbook_id,
    )
    if current_user is None:
        raise HTTPException(status_code=403, detail="没有该资源权限")
    if not access.capabilities.can_edit_data or not access.capabilities.can_run_sheet_actions:
        raise HTTPException(status_code=403, detail="没有该资源权限")
    _check_sheet_base_version(document, payload.base_version)

    revision_label = payload.revision_label.strip()
    if not revision_label:
        raise HTTPException(status_code=400, detail="修订类型不能为空")

    try:
        from backend.core.attendance.nianzhu_course_sheets import apply_nianzhu_attendance_video_revision

        recalculation = apply_nianzhu_attendance_video_revision(
            session,
            attendance_sheet_id=int(document.numeric_id or sheet_id),
            cells=[cell.model_dump() for cell in payload.cells],
            revision_label=revision_label,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except RuntimeError as exc:
        raise HTTPException(status_code=404, detail="当前工作表没有这个定制功能") from exc

    document.updated_by_user_id = current_user.id
    document.updated_at = time.time()
    session.add(document)
    session.commit()
    session.refresh(document)

    next_document = _normalize_document_json(dict(document.document_json or {}))
    return NoteSheetAttendanceVideoRevisionResponse(
        sheet=_build_attendance_template_detail_response(
            session,
            document,
            next_document,
            access=access,
            current_user=current_user,
        ),
        revision_label=revision_label,
        updated_count=int(recalculation.get("revision_target_count") or 0),
        recalculation=recalculation,
    )


@router.delete("/sheets/{sheet_id}")
def delete_note_sheet(
    sheet_id: int,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_active_user),
):
    _require_note_sheets_feature(session, current_user)
    document, _access, _workbook = _get_note_sheet_or_404(session, current_user, sheet_id, required_role="manager")
    now = time.time()
    document.deleted_at = now
    document.deleted_by_user_id = current_user.id
    document.updated_by_user_id = current_user.id
    document.updated_at = now
    document.version = max(int(document.version or 1), 1) + 1
    session.add(document)
    session.commit()
    session.refresh(document)
    _broadcast_sheet_resource_update(document)
    return {"ok": True}


@router.post("/sheets/{sheet_id}/restore", response_model=NoteSheetDetailResponse)
def restore_note_sheet(
    sheet_id: int,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_active_user),
):
    _require_note_sheets_feature(session, current_user)
    document = _get_sheet_by_numeric_id_or_404(session, sheet_id, include_deleted=True)
    access = _resolve_sheet_resource_access(session, document, current_user)
    _require_resource_access(access, "manager")
    if document.deleted_at:
        now = time.time()
        document.deleted_at = None
        document.deleted_by_user_id = None
        document.updated_by_user_id = current_user.id
        document.updated_at = now
        document.version = max(int(document.version or 1), 1) + 1
        session.add(document)
        session.commit()
        session.refresh(document)
        _broadcast_sheet_resource_update(document)

    workbook_items = _list_workbook_refs_for_sheet_ids(session, [document.id], current_user).get(document.id, [])
    parent_workbook_id = _get_parent_workbook_id_for_sheet(session, document)
    return NoteSheetDetailResponse.model_validate(
        _serialize_sheet_detail(
            document,
            workbook_items=workbook_items,
            parent_workbook_id=parent_workbook_id,
            access=_resolve_sheet_resource_access(session, document, current_user),
        ),
    )


@router.get("/workbooks", response_model=list[WorkbookSummaryResponse])
def list_workbooks(
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_active_user),
):
    _require_note_sheets_feature(session, current_user)

    if current_user.is_superuser:
        candidate_workbooks = session.exec(select(WorkbookDocument).where(_active_workbook_condition())).all()
    else:
        owned_workbooks = session.exec(
            select(WorkbookDocument).where(WorkbookDocument.owner_user_id == current_user.id)
            .where(_active_workbook_condition())
        ).all()
        subject_keys = _current_user_subject_keys(current_user)
        grants = session.exec(
            select(ResourceAccessGrant)
            .where(ResourceAccessGrant.resource_type == RESOURCE_TYPE_WORKBOOK)
            .where(ResourceAccessGrant.subject_key.in_(subject_keys))
        ).all()
        granted_workbook_ids = sorted({int(grant.resource_id) for grant in grants if str(grant.resource_id).isdigit()})
        granted_workbooks = session.exec(
            select(WorkbookDocument)
            .where(WorkbookDocument.numeric_id.in_(granted_workbook_ids))
            .where(_active_workbook_condition())
        ).all() if granted_workbook_ids else []
        workbook_map = {workbook.id: workbook for workbook in [*owned_workbooks, *granted_workbooks]}
        candidate_workbooks = list(workbook_map.values())

    workbook_access_items = [
        (workbook, _resolve_workbook_resource_access(session, workbook, current_user))
        for workbook in candidate_workbooks
    ]
    workbook_access_items = [
        (workbook, access)
        for workbook, access in workbook_access_items
        if access.capabilities.can_read
    ]
    workbook_access_items.sort(
        key=lambda item: (float(item[0].updated_at or 0.0), float(item[0].created_at or 0.0)),
        reverse=True,
    )

    if not workbook_access_items:
        return []

    workbook_refs = [ref for workbook, _access in workbook_access_items for ref in workbook_ref_aliases(workbook)]
    links = session.exec(
        select(WorkbookSheetLink)
        .where(WorkbookSheetLink.workbook_id.in_(workbook_refs))
    ).all()
    counts: dict[str, int] = {}
    workbook_ref_map = {
        ref: workbook
        for workbook, _access in workbook_access_items
        for ref in workbook_ref_aliases(workbook)
    }
    linked_sheet_map = load_sheets_by_refs(session, [link.sheet_id for link in links])
    for link in links:
        workbook = workbook_ref_map.get(str(link.workbook_id))
        if workbook is None or linked_sheet_map.get(str(link.sheet_id)) is None:
            continue
        workbook_key = str(workbook.id)
        counts[workbook_key] = counts.get(workbook_key, 0) + 1

    return [
        WorkbookSummaryResponse.model_validate(
            _serialize_workbook_summary(
                workbook,
                sheet_count=counts.get(str(workbook.id), 0),
                access=access,
            ),
        )
        for workbook, access in workbook_access_items
    ]


@router.post("/workbooks", response_model=WorkbookDetailResponse)
def create_workbook(
    payload: WorkbookCreateRequest,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_active_user),
):
    _require_note_sheets_feature(session, current_user)
    now = time.time()
    workbook_identity = allocate_new_workbook_identity(session)
    workbook = WorkbookDocument(
        id=workbook_identity.primary_id,
        numeric_id=workbook_identity.numeric_id,
        legacy_id=workbook_identity.legacy_id,
        title=_normalize_title(payload.title, default_value="未命名工作簿"),
        owner_user_id=current_user.id,
        created_by_user_id=current_user.id,
        updated_by_user_id=current_user.id,
        created_at=now,
        updated_at=now,
    )
    session.add(workbook)
    session.flush()
    _ensure_workbook_identity(session, workbook)
    session.commit()
    session.refresh(workbook)
    access = _resolve_workbook_resource_access(session, workbook, current_user)
    return WorkbookDetailResponse.model_validate(
        _serialize_workbook_detail(session, workbook, current_user=current_user, access=access),
    )


@router.get("/workbooks/{workbook_id}", response_model=WorkbookDetailResponse)
def get_workbook(
    workbook_id: int,
    session: Session = Depends(get_session),
    current_user: User | None = Depends(get_optional_current_user_from_token),
):
    workbook, access = _get_workbook_or_404(session, current_user, workbook_id, required_role="viewer")
    return WorkbookDetailResponse.model_validate(
        _serialize_workbook_detail(session, workbook, current_user=current_user, access=access),
    )


@router.post("/workbooks/{workbook_id}/restore", response_model=WorkbookDetailResponse)
def restore_workbook(
    workbook_id: int,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_active_user),
):
    _require_note_sheets_feature(session, current_user)
    workbook = _get_workbook_by_numeric_id_or_404(session, workbook_id, include_deleted=True)
    access = _resolve_workbook_resource_access(session, workbook, current_user)
    _require_resource_access(access, "manager")
    now = time.time()
    changed = False
    if workbook.deleted_at:
        workbook.deleted_at = None
        workbook.deleted_by_user_id = None
        workbook.updated_by_user_id = current_user.id
        workbook.updated_at = now
        session.add(workbook)
        changed = True

    links = session.exec(
        select(WorkbookSheetLink)
        .where(WorkbookSheetLink.workbook_id.in_(workbook_ref_aliases(workbook)))
    ).all()
    sheet_map = load_sheets_by_refs(session, [link.sheet_id for link in links], include_deleted=True)
    restored_sheet_ids: set[str] = set()
    for sheet in {str(sheet.id): sheet for sheet in sheet_map.values()}.values():
        if not sheet.deleted_at or str(sheet.id) in restored_sheet_ids:
            continue
        restored_sheet_ids.add(str(sheet.id))
        sheet.deleted_at = None
        sheet.deleted_by_user_id = None
        sheet.updated_by_user_id = current_user.id
        sheet.updated_at = now
        session.add(sheet)
        changed = True

    if changed:
        session.commit()
        session.refresh(workbook)

    return WorkbookDetailResponse.model_validate(
        _serialize_workbook_detail(
            session,
            workbook,
            current_user=current_user,
            access=_resolve_workbook_resource_access(session, workbook, current_user),
        ),
    )


@router.put("/workbooks/{workbook_id}", response_model=WorkbookDetailResponse)
def update_workbook(
    workbook_id: int,
    payload: WorkbookUpdateRequest,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_active_user),
):
    _require_note_sheets_feature(session, current_user)
    workbook, access = _get_workbook_or_404(session, current_user, workbook_id, required_role="manager")
    next_title = _normalize_title(payload.title, default_value=workbook.title or "未命名工作簿")
    if workbook.title != next_title:
        workbook.title = next_title
        workbook.updated_by_user_id = current_user.id
        workbook.updated_at = time.time()
        session.add(workbook)
        session.commit()
        session.refresh(workbook)

    return WorkbookDetailResponse.model_validate(
        _serialize_workbook_detail(session, workbook, current_user=current_user, access=access),
    )


@router.get("/workbooks/{workbook_id}/defined-names", response_model=NoteSheetDefinedNamesResponse)
def get_workbook_defined_names_endpoint(
    workbook_id: int,
    session: Session = Depends(get_session),
    current_user: User | None = Depends(get_optional_current_user_from_token),
):
    workbook, _access = _get_workbook_or_404(session, current_user, workbook_id, required_role="viewer")
    names = _get_workbook_defined_names(session, workbook)
    worksheets = _list_workbook_defined_name_worksheets(session, workbook, current_user)
    return NoteSheetDefinedNamesResponse(
        workbook_id=_require_workbook_numeric_id(workbook),
        workbook=names,
        worksheets=worksheets,
        effective=names,
    )


@router.put("/workbooks/{workbook_id}/defined-names", response_model=NoteSheetDefinedNamesResponse)
def update_workbook_defined_names_endpoint(
    workbook_id: int,
    payload: NoteSheetDefinedNamesUpdateRequest,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_active_user),
):
    workbook, _access = _get_workbook_or_404(session, current_user, workbook_id, required_role="editor")
    names = _normalize_defined_names(payload.names, scope="workbook", strict=True)
    _set_workbook_defined_names(session, workbook, names)
    updated_documents: list[SheetDocument] = []

    if payload.worksheets:
        links = session.exec(
            select(WorkbookSheetLink)
            .where(WorkbookSheetLink.workbook_id.in_(workbook_ref_aliases(workbook)))
            .order_by(WorkbookSheetLink.order_index, WorkbookSheetLink.created_at)
        ).all()
        sheet_ids = [link.sheet_id for link in links]
        sheet_map = load_sheets_by_refs(session, sheet_ids)
        workbook_sheets_by_numeric_id: dict[int, SheetDocument] = {}
        for link in links:
            document = sheet_map.get(str(link.sheet_id))
            if document is None:
                continue
            workbook_sheets_by_numeric_id.setdefault(_require_sheet_numeric_id(document), document)

        seen_payload_sheet_ids: set[int] = set()
        for worksheet_payload in payload.worksheets:
            payload_sheet_id = int(worksheet_payload.sheet_id)
            if payload_sheet_id in seen_payload_sheet_ids:
                raise HTTPException(status_code=400, detail=f"工作表作用域重复: {payload_sheet_id}")
            seen_payload_sheet_ids.add(payload_sheet_id)
            document = workbook_sheets_by_numeric_id.get(payload_sheet_id)
            if document is None:
                raise HTTPException(status_code=404, detail=f"工作簿中不存在该工作表: {payload_sheet_id}")
            sheet_access = _resolve_sheet_resource_access(session, document, current_user, workbook=workbook)
            if not sheet_access.capabilities.can_edit_config:
                raise HTTPException(status_code=403, detail=f"没有权限修改工作表: {document.title or payload_sheet_id}")
            if (
                worksheet_payload.sheet_version is not None
                and int(worksheet_payload.sheet_version) != int(document.version or 1)
            ):
                raise HTTPException(status_code=409, detail=f"工作表版本已变化，请重新读取后再写入: {document.title or payload_sheet_id}")

            current_document = dict(document.document_json or {})
            next_document = _replace_sheet_defined_names(current_document, worksheet_payload.names)
            if next_document != current_document:
                document.document_json = next_document
                document.version = max(int(document.version or 1), 1) + 1
                document.updated_by_user_id = current_user.id
                document.updated_at = time.time()
                session.add(document)
                updated_documents.append(document)

    workbook.updated_by_user_id = current_user.id
    workbook.updated_at = time.time()
    session.add(workbook)
    session.commit()
    session.refresh(workbook)
    for document in updated_documents:
        session.refresh(document)
        _broadcast_sheet_resource_update(document)
    stored_names = _get_workbook_defined_names(session, workbook)
    worksheets = _list_workbook_defined_name_worksheets(session, workbook, current_user)
    return NoteSheetDefinedNamesResponse(
        workbook_id=_require_workbook_numeric_id(workbook),
        workbook=stored_names,
        worksheets=worksheets,
        effective=stored_names,
    )


@router.get("/workbooks/{workbook_id}/access", response_model=NoteSheetResourceAccessResponse)
def get_workbook_access(
    workbook_id: int,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_active_user),
):
    workbook, access = _get_workbook_or_404(session, current_user, workbook_id, required_role="manager")
    return _build_resource_access_response(
        session,
        resource_type=RESOURCE_TYPE_WORKBOOK,
        numeric_id=_require_workbook_numeric_id(workbook),
        resource_id=_workbook_resource_id(workbook),
        access=access,
    )


@router.put("/workbooks/{workbook_id}/access", response_model=NoteSheetResourceAccessResponse)
def update_workbook_access(
    workbook_id: int,
    payload: NoteSheetResourceAccessUpdateRequest,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_active_user),
):
    workbook, _access = _get_workbook_or_404(session, current_user, workbook_id, required_role="manager")
    _save_resource_access_grants(
        session,
        resource_type=RESOURCE_TYPE_WORKBOOK,
        resource_id=_workbook_resource_id(workbook),
        payload=payload,
        current_user=current_user,
    )
    session.commit()
    access = _resolve_workbook_resource_access(session, workbook, current_user)
    return _build_resource_access_response(
        session,
        resource_type=RESOURCE_TYPE_WORKBOOK,
        numeric_id=_require_workbook_numeric_id(workbook),
        resource_id=_workbook_resource_id(workbook),
        access=access,
    )


@router.post("/workbooks/{workbook_id}/save-as", response_model=WorkbookDetailResponse)
def save_as_workbook(
    workbook_id: int,
    payload: WorkbookSaveAsRequest,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_active_user),
):
    _require_note_sheets_feature(session, current_user)
    source_workbook, _source_access = _get_workbook_or_404(
        session,
        current_user,
        workbook_id,
        required_role="viewer",
    )
    links = session.exec(
        select(WorkbookSheetLink)
        .where(WorkbookSheetLink.workbook_id.in_(workbook_ref_aliases(source_workbook)))
        .order_by(WorkbookSheetLink.order_index, WorkbookSheetLink.created_at)
    ).all()
    source_sheet_ids = [link.sheet_id for link in links]
    source_sheet_map = load_sheets_by_refs(session, source_sheet_ids)

    now = time.time()
    workbook_identity = allocate_new_workbook_identity(session)
    workbook = WorkbookDocument(
        id=workbook_identity.primary_id,
        numeric_id=workbook_identity.numeric_id,
        legacy_id=workbook_identity.legacy_id,
        title=_normalize_title(payload.title, default_value="未命名工作簿"),
        owner_user_id=current_user.id,
        created_by_user_id=current_user.id,
        updated_by_user_id=current_user.id,
        created_at=now,
        updated_at=now,
    )
    session.add(workbook)
    session.flush()
    _ensure_workbook_identity(session, workbook)
    _set_workbook_defined_names(session, workbook, _get_workbook_defined_names(session, source_workbook))

    for link in links:
        source_sheet = source_sheet_map.get(str(link.sheet_id))
        if source_sheet is None:
            continue

        document_identity = allocate_new_sheet_identity(session)
        document = SheetDocument(
            id=document_identity.primary_id,
            numeric_id=document_identity.numeric_id,
            legacy_id=document_identity.legacy_id,
            scope=source_sheet.scope,
            owner_type=source_sheet.owner_type,
            owner_key=source_sheet.owner_key,
            sheet_key="pending",
            title=source_sheet.title,
            engine=source_sheet.engine,
            document_json=_clone_sheet_document_json(
                dict(source_sheet.document_json or {}),
                mode=payload.mode,
            ),
            version=1,
            owner_user_id=current_user.id,
            created_by_user_id=current_user.id,
            updated_by_user_id=current_user.id,
            created_at=now,
            updated_at=now,
        )
        session.add(document)
        session.flush()
        document.sheet_key = str(_require_sheet_numeric_id(document))
        session.add(document)
        ensure_attendance_sheet_anonymous_viewer(session, document)
        session.add(
            WorkbookSheetLink(
                workbook_id=_workbook_link_ref(workbook),
                sheet_id=_sheet_link_ref(document),
                order_index=link.order_index,
                created_at=now,
            ),
        )

    session.commit()
    session.refresh(workbook)
    access = _resolve_workbook_resource_access(session, workbook, current_user)
    return WorkbookDetailResponse.model_validate(
        _serialize_workbook_detail(session, workbook, current_user=current_user, access=access),
    )


@router.post("/workbooks/{workbook_id}/sheets", response_model=WorkbookDetailResponse)
def attach_sheet_to_workbook(
    workbook_id: int,
    payload: WorkbookAttachSheetRequest,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_active_user),
):
    _require_note_sheets_feature(session, current_user)
    workbook, access = _get_workbook_or_404(session, current_user, workbook_id, required_role="manager")
    document, _sheet_access, _ = _get_note_sheet_or_404(session, current_user, payload.sheet_id, required_role="manager")

    existing = session.exec(
        select(WorkbookSheetLink)
        .where(WorkbookSheetLink.workbook_id.in_(workbook_ref_aliases(workbook)))
        .where(WorkbookSheetLink.sheet_id.in_(sheet_ref_aliases(document)))
    ).first()

    if existing is None:
        session.add(
            WorkbookSheetLink(
                workbook_id=_workbook_link_ref(workbook),
                sheet_id=_sheet_link_ref(document),
                order_index=_get_next_workbook_link_order(session, _workbook_link_ref(workbook)),
            ),
        )
        workbook.updated_by_user_id = current_user.id
        workbook.updated_at = time.time()
        session.add(workbook)
        session.commit()
        session.refresh(workbook)

    return WorkbookDetailResponse.model_validate(
        _serialize_workbook_detail(session, workbook, current_user=current_user, access=access),
    )


@router.post("/workbooks/{workbook_id}/sheets/reorder", response_model=WorkbookDetailResponse)
def reorder_workbook_sheets(
    workbook_id: int,
    payload: WorkbookReorderSheetsRequest,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_active_user),
):
    _require_note_sheets_feature(session, current_user)
    workbook, access = _get_workbook_or_404(session, current_user, workbook_id, required_role="editor")
    requested_numeric_ids = [int(sheet_id) for sheet_id in payload.sheet_ids]
    if not requested_numeric_ids:
        raise HTTPException(status_code=400, detail="工作表顺序不能为空")
    if any(sheet_id <= 0 for sheet_id in requested_numeric_ids):
        raise HTTPException(status_code=400, detail="工作表编号无效")
    if len(set(requested_numeric_ids)) != len(requested_numeric_ids):
        raise HTTPException(status_code=400, detail="工作表顺序包含重复项")

    links = session.exec(
        select(WorkbookSheetLink)
        .where(WorkbookSheetLink.workbook_id.in_(workbook_ref_aliases(workbook)))
        .order_by(WorkbookSheetLink.order_index, WorkbookSheetLink.created_at)
    ).all()
    if not links:
        return WorkbookDetailResponse.model_validate(
            _serialize_workbook_detail(session, workbook, current_user=current_user, access=access),
        )

    linked_sheet_ids = [link.sheet_id for link in links]
    sheet_ref_map = load_sheets_by_refs(session, linked_sheet_ids)
    sheets = list({str(sheet.id): sheet for sheet in sheet_ref_map.values()}.values())
    sheet_by_numeric_id = {
        int(sheet.numeric_id): sheet
        for sheet in sheets
        if sheet.numeric_id is not None
    }
    link_by_sheet_id: dict[str, WorkbookSheetLink] = {}
    for link in links:
        sheet = sheet_ref_map.get(str(link.sheet_id))
        if sheet is not None:
            link_by_sheet_id[str(sheet.id)] = link

    requested_sheet_ids: list[str] = []
    for numeric_id in requested_numeric_ids:
        sheet = sheet_by_numeric_id.get(numeric_id)
        if sheet is None or sheet.id not in link_by_sheet_id:
            raise HTTPException(status_code=400, detail="工作表不在当前工作簿中")
        sheet_access = _resolve_sheet_resource_access(session, sheet, current_user, workbook=workbook)
        if not sheet_access.capabilities.can_read:
            raise HTTPException(status_code=403, detail="没有该工作表权限")
        requested_sheet_ids.append(sheet.id)

    requested_sheet_id_set = set(requested_sheet_ids)
    requested_links = [link_by_sheet_id[sheet_id] for sheet_id in requested_sheet_ids]
    requested_link_iter = iter(requested_links)
    next_links: list[WorkbookSheetLink] = []
    for link in links:
        linked_sheet = sheet_ref_map.get(str(link.sheet_id))
        linked_sheet_id = str(linked_sheet.id) if linked_sheet is not None else str(link.sheet_id)
        if linked_sheet_id in requested_sheet_id_set:
            next_links.append(next(requested_link_iter))
        else:
            next_links.append(link)

    now = time.time()
    for index, link in enumerate(next_links, start=1):
        next_order_index = index * 10
        if link.order_index != next_order_index:
            link.order_index = next_order_index
            session.add(link)
    workbook.updated_by_user_id = current_user.id
    workbook.updated_at = now
    session.add(workbook)
    session.commit()
    session.refresh(workbook)

    return WorkbookDetailResponse.model_validate(
        _serialize_workbook_detail(session, workbook, current_user=current_user, access=access),
    )


@router.delete("/workbooks/{workbook_id}/sheets/{sheet_id}", response_model=WorkbookDetailResponse)
def remove_sheet_from_workbook(
    workbook_id: int,
    sheet_id: int,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_active_user),
):
    _require_note_sheets_feature(session, current_user)
    workbook, access = _get_workbook_or_404(session, current_user, workbook_id, required_role="manager")
    document, _sheet_access, _ = _get_note_sheet_or_404(session, current_user, sheet_id, required_role="manager")

    link = session.exec(
        select(WorkbookSheetLink)
        .where(WorkbookSheetLink.workbook_id.in_(workbook_ref_aliases(workbook)))
        .where(WorkbookSheetLink.sheet_id.in_(sheet_ref_aliases(document)))
    ).first()
    if link is not None:
        session.delete(link)
        workbook.updated_by_user_id = current_user.id
        workbook.updated_at = time.time()
        session.add(workbook)
        session.commit()
        session.refresh(workbook)

    return WorkbookDetailResponse.model_validate(
        _serialize_workbook_detail(session, workbook, current_user=current_user, access=access),
    )


@router.post("/workbooks/{workbook_id}/unpack")
def unpack_workbook(
    workbook_id: int,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_active_user),
):
    _require_note_sheets_feature(session, current_user)
    workbook, _access = _get_workbook_or_404(session, current_user, workbook_id, required_role="manager")
    session.exec(delete(WorkbookSheetLink).where(WorkbookSheetLink.workbook_id.in_(workbook_ref_aliases(workbook))))
    _delete_resource_access_grants(
        session,
        resource_type=RESOURCE_TYPE_WORKBOOK,
        resource_ids=[_workbook_resource_id(workbook)],
    )
    session.delete(workbook)
    session.commit()
    return {"ok": True}


@router.delete("/workbooks/{workbook_id}")
def delete_workbook(
    workbook_id: int,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_active_user),
):
    _require_note_sheets_feature(session, current_user)
    workbook, _access = _get_workbook_or_404(session, current_user, workbook_id, required_role="manager")
    links = session.exec(
        select(WorkbookSheetLink).where(WorkbookSheetLink.workbook_id.in_(workbook_ref_aliases(workbook)))
    ).all()
    sheet_ids = sorted({link.sheet_id for link in links})
    owned_sheet_ids = _get_sheet_ids_owned_only_by_workbook(
        session,
        workbook_id=_workbook_link_ref(workbook),
        sheet_ids=sheet_ids,
    )
    sheets = list({str(sheet.id): sheet for sheet in load_sheets_by_refs(session, owned_sheet_ids).values()}.values()) if owned_sheet_ids else []

    now = time.time()
    for sheet in sheets:
        sheet.deleted_at = now
        sheet.deleted_by_user_id = current_user.id
        sheet.updated_by_user_id = current_user.id
        sheet.updated_at = now
        session.add(sheet)
    workbook.deleted_at = now
    workbook.deleted_by_user_id = current_user.id
    workbook.updated_by_user_id = current_user.id
    workbook.updated_at = now
    session.add(workbook)
    session.commit()
    return {"ok": True}
