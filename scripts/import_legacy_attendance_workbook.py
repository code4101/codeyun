from __future__ import annotations

import argparse
from datetime import date, datetime
import re
import sys
import time
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from sqlmodel import Session, select

REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from backend.db import engine
from backend.core.resource_identity import (
    RESOURCE_TYPE_SHEET,
    RESOURCE_TYPE_WORKBOOK,
    ensure_resource_identity,
)
from backend.core.sheet_identity import allocate_new_sheet_identity, allocate_new_workbook_identity
from backend.core.sheet_refs import sheet_public_id, sheet_ref_aliases, workbook_public_id, workbook_ref_aliases
from backend.models import SheetDocument, User, WorkbookDocument, WorkbookSheetLink


LEGACY_TEXT_COLUMNS = {"手机号", "错误手机号", "微信支付订单号", "商户订单号", "微信号"}
REGISTRATION_ACTION_LABELS = {
    "备注": ("excel_import_reset", "导入excel"),
    "微信支付订单号": ("registration_order_match", "更新订单匹配"),
    "用户ID": ("registration_user_match", "更新用户匹配"),
}
REGISTRATION_ACTION_ROW_NOTES = {
    "已返款": "仅看数据库已有情况，如果需要实时性需要用订单工具",
    "参考信息": "其他备注",
}
REGISTRATION_EXTRA_LEADING_COLUMNS = ["分组"]
REGISTRATION_GROUP_MARKERS = (
    "第1批",
    "第2批",
    "第3批",
    "以下",
    "添加",
)
ATTENDANCE_HEADER_RENAMES = {
    "微信昵称": "昵称",
}


def _normalize_text(value: Any) -> str:
    return str(value or "").strip()


def _normalize_header(value: Any, fallback: str) -> str:
    text = re.sub(r"\s+", " ", _normalize_text(value)).strip()
    return text or fallback


def _excel_color_to_hex(color: Any) -> str:
    if color is None:
        return ""
    if getattr(color, "type", None) != "rgb":
        return ""
    rgb = str(getattr(color, "rgb", "") or "")
    if len(rgb) == 8:
        alpha, rgb = rgb[:2], rgb[2:]
        if alpha == "00":
            return ""
    if len(rgb) != 6:
        return ""
    if rgb.upper() in {"FFFFFF", "000000"}:
        return ""
    return f"#{rgb.upper()}"


def _cell_meta_from_cell(cell: Cell) -> dict[str, Any]:
    meta: dict[str, Any] = {}
    style: dict[str, str] = {}

    fill = cell.fill
    fill_type = getattr(fill, "fill_type", None)
    if fill_type and fill_type != "none":
        background = _excel_color_to_hex(fill.fgColor)
        if background:
            style["background_color"] = background

    text_color = _excel_color_to_hex(cell.font.color)
    if text_color:
        style["text_color"] = text_color

    if style:
        meta["style"] = style

    hyperlink = cell.hyperlink.target if cell.hyperlink is not None else ""
    if hyperlink:
        meta["link"] = {"url": hyperlink}

    return meta


def _normalize_formula(value: str) -> str:
    return value.replace("_xlfn.", "").replace("_xlws.", "")


def _normalize_cell_value(value: Any, *, header: str = "") -> Any:
    if value is None:
        return ""
    if isinstance(value, str):
        text = _normalize_formula(value) if value.startswith("=") else value
        return text.lstrip("`'") if header in LEGACY_TEXT_COLUMNS else text
    if isinstance(value, datetime):
        if value.hour == value.minute == value.second == 0:
            return value.date().isoformat()
        return value.isoformat(sep=" ", timespec="minutes")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def _worksheet_last_used_column(ws: Worksheet, header_row: int = 2) -> int:
    last_column = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.value not in (None, ""):
                last_column = max(last_column, int(cell.column))
    header_last = 0
    for cell in ws[header_row]:
        if cell.value not in (None, ""):
            header_last = max(header_last, int(cell.column))
    return max(last_column, header_last, 1)


def _normalize_row(values: list[Any], size: int) -> list[Any]:
    row = list(values[:size])
    if len(row) < size:
        row.extend([""] * (size - len(row)))
    return row


def _column_widths(ws: Worksheet, column_count: int) -> list[int]:
    widths: list[int] = []
    for index in range(1, column_count + 1):
        letter = get_column_letter(index)
        width = ws.column_dimensions[letter].width
        if width is None:
            widths.append(88)
        else:
            widths.append(min(max(int(round(float(width) * 8)), 56), 260))
    return widths


def _append_meta(cell_meta: dict[str, Any], key: str, patch: dict[str, Any]) -> None:
    if not patch:
        return
    current = dict(cell_meta.get(key) or {})
    for field, value in patch.items():
        if field == "style" and isinstance(value, dict):
            current["style"] = {**dict(current.get("style") or {}), **value}
        else:
            current[field] = value
    if current:
        cell_meta[key] = current


def _is_registration_group_marker(value: str) -> bool:
    text = _normalize_text(value)
    if not text:
        return False
    if "退课" in text or "海外" in text:
        return False
    return any(marker in text for marker in REGISTRATION_GROUP_MARKERS)


def _normalize_registration_group(value: str) -> str:
    text = _normalize_text(value)
    match = re.search(r"第\s*(\d+)\s*批", text)
    if match:
        return f"第{match.group(1)}批"
    text = re.sub(r"^以下(?:是)?", "", text).strip()
    text = re.sub(r"^新增", "", text).strip()
    return text or value


def _registration_document(ws: Worksheet) -> dict[str, Any]:
    last_column = _worksheet_last_used_column(ws)
    source_headers = [
        _normalize_header(ws.cell(2, column).value, get_column_letter(column))
        for column in range(1, last_column + 1)
    ]
    while source_headers and source_headers[-1].startswith(get_column_letter(len(source_headers))):
        source_headers.pop()
    source_headers = source_headers or ["序号", "姓名"]
    columns = [*REGISTRATION_EXTRA_LEADING_COLUMNS, *source_headers]

    action_row = [""] * len(columns)
    for header, (_action_type, label) in REGISTRATION_ACTION_LABELS.items():
        if header in columns:
            action_row[columns.index(header)] = label
    for header, note in REGISTRATION_ACTION_ROW_NOTES.items():
        if header in columns:
            action_row[columns.index(header)] = note

    rows: list[list[Any]] = []
    current_group = ""
    note_index = source_headers.index("备注") if "备注" in source_headers else None
    for row_number in range(4, ws.max_row + 1):
        source_values = [
            _normalize_cell_value(ws.cell(row_number, column).value, header=source_headers[column - 1])
            for column in range(1, len(source_headers) + 1)
        ]
        if not any(_normalize_text(value) for value in source_values):
            continue

        note_value = _normalize_text(source_values[note_index]) if note_index is not None else ""
        if _is_registration_group_marker(note_value):
            current_group = _normalize_registration_group(note_value)
            source_values[note_index] = ""

        rows.append([current_group, *source_values])

    grid_rows = [columns, action_row, *rows]
    cell_meta: dict[str, Any] = {}
    for column_index, header in enumerate(columns):
        config_meta = _cell_meta_from_cell(ws.cell(2, max(column_index, 1)))
        if column_index == 0:
            config_meta = {"style": {"background_color": "#9DC3E6"}}
        _append_meta(cell_meta, f"0:{column_index}", config_meta)

    for header, (action_type, label) in REGISTRATION_ACTION_LABELS.items():
        if header in columns:
            column_index = columns.index(header)
            _append_meta(cell_meta, f"1:{column_index}", {"action": {"type": action_type, "label": label}})

    column_configs = _registration_column_configs(columns, action_row)
    return {
        "schema_version": 1,
        "columns": columns,
        "rows": rows,
        "grid_rows": grid_rows,
        "data_start_row": 2,
        "field_row_index": 0,
        "column_widths": [88, *_column_widths(ws, len(source_headers))],
        "column_configs": column_configs,
        "cell_meta": cell_meta,
        "merged_cells": [],
        "header_groups": [],
        "formula_reference_origin": "sheet_v2",
        "view_settings": _default_view_settings(row_marker_numbering="global"),
    }


def _registration_column_configs(columns: list[str], action_row: list[str]) -> dict[str, Any]:
    configs: dict[str, Any] = {}
    for index, header in enumerate(columns):
        config: dict[str, Any] = {}
        if header in {"序号", "订单日期", "订单金额", "已返款", "匹配得分"}:
            config["value_type"] = "number"
        if header in {"提交时间", "出生年月（必填）"}:
            config["value_type"] = "date"
        if header in {"姓名", "手机号", "微信支付订单号", "商户订单号", "用户ID"}:
            config["duplicate_value_highlight"] = True
        if header in {"微信支付订单号", "商户订单号", "用户ID"}:
            config["font_family"] = "monospace"
        if index <= 15:
            config["header_background_color"] = "#9DC3E6"
        if action_row[index]:
            config["note"] = action_row[index]
        if header in {"分组", "用户ID", "关联用户ID"}:
            config["width_mode"] = "fixed"
        if config:
            configs[header] = config
    return configs


def _attendance_headers(ws: Worksheet, column_count: int) -> list[str]:
    headers: list[str] = []
    used: dict[str, int] = {}
    for column in range(1, column_count + 1):
        header = _normalize_header(ws.cell(2, column).value, get_column_letter(column))
        header = ATTENDANCE_HEADER_RENAMES.get(header, header)
        if header in used:
            used[header] += 1
            header = f"{header}_{used[header]}"
        else:
            used[header] = 1
        headers.append(header)
    return headers


def _attendance_document(ws: Worksheet) -> dict[str, Any]:
    column_count = _worksheet_last_used_column(ws)
    columns = _attendance_headers(ws, column_count)
    rows: list[list[Any]] = []
    for row_number in range(4, ws.max_row + 1):
        row = [
            _normalize_cell_value(ws.cell(row_number, column).value, header=columns[column - 1])
            for column in range(1, column_count + 1)
        ]
        if any(_normalize_text(value) for value in row):
            rows.append(row)

    header_rows: list[list[Any]] = []
    for row_number in range(1, 4):
        row = [
            _normalize_cell_value(ws.cell(row_number, column).value, header=columns[column - 1])
            for column in range(1, column_count + 1)
        ]
        if row_number == 2:
            row = columns
        header_rows.append(_normalize_row(row, column_count))

    grid_rows = [*header_rows, *rows]
    cell_meta: dict[str, Any] = {}
    for row_number in range(1, ws.max_row + 1):
        document_row = row_number - 1
        if document_row >= len(grid_rows):
            break
        for column in range(1, column_count + 1):
            meta = _cell_meta_from_cell(ws.cell(row_number, column))
            _append_meta(cell_meta, f"{document_row}:{column - 1}", meta)

    merged_cells = _merged_cells(ws, row_count=len(grid_rows), column_count=column_count)
    column_configs = _attendance_column_configs(columns, header_rows)
    return {
        "schema_version": 1,
        "columns": columns,
        "rows": rows,
        "grid_rows": grid_rows,
        "data_start_row": 3,
        "field_row_index": 1,
        "column_widths": _column_widths(ws, column_count),
        "column_configs": column_configs,
        "cell_meta": cell_meta,
        "merged_cells": merged_cells,
        "header_groups": [[{"label": _normalize_text(value), "colspan": 1} for value in header_rows[0]]],
        "formula_reference_origin": "sheet_v2",
        "view_settings": _default_view_settings(row_marker_numbering="page"),
    }


def _attendance_column_configs(columns: list[str], header_rows: list[list[Any]]) -> dict[str, Any]:
    configs: dict[str, Any] = {}
    notes = header_rows[2] if len(header_rows) > 2 else []
    for index, header in enumerate(columns):
        config: dict[str, Any] = {}
        note = _normalize_text(notes[index] if index < len(notes) else "")
        if note:
            config["note"] = note
        if header in {"学号", "优秀学员评分", "完成视频数", "视频应返款", "打卡应返款", "总应返款", "已返款", "订单金额", "当前应返款", "打卡数"}:
            config["value_type"] = "number"
        if header in {"商户订单号", "用户ID"}:
            config["duplicate_value_highlight"] = True
            config["font_family"] = "monospace"
        if header in {"姓名", "昵称", "商户订单号", "用户ID"}:
            config["duplicate_value_highlight"] = True
        if index <= 7:
            config.setdefault("header_background_color", "#D9E1F2")
        elif index <= 14:
            config.setdefault("header_background_color", "#FFDCC4")
        elif index == 15:
            config.setdefault("header_background_color", "#FFF2CC")
        elif re.search(r"第\s*0*\d+\s*课", header):
            config.setdefault("header_background_color", "#E2F0D9")
        if header in {"用户ID"}:
            config["hidden"] = True
        if config:
            configs[header] = config
    return configs


def _merged_cells(ws: Worksheet, *, row_count: int, column_count: int) -> list[dict[str, int]]:
    cells: list[dict[str, int]] = []
    for merged in ws.merged_cells.ranges:
        row = int(merged.min_row) - 1
        col = int(merged.min_col) - 1
        if row < 0 or col < 0 or row >= row_count or col >= column_count:
            continue
        rowspan = min(int(merged.max_row) - int(merged.min_row) + 1, row_count - row)
        colspan = min(int(merged.max_col) - int(merged.min_col) + 1, column_count - col)
        if rowspan > 1 or colspan > 1:
            cells.append({"row": row, "col": col, "rowspan": rowspan, "colspan": colspan})
    return cells


def _default_view_settings(*, row_marker_numbering: str) -> dict[str, Any]:
    return {
        "show_row_numbers": True,
        "row_marker_numbering": row_marker_numbering,
        "row_marker_origin": "sheet",
        "show_column_markers": True,
        "column_marker_style": "letters",
        "column_note_display": "row",
        "height_mode": "fill",
        "frozen_column_count": 0,
        "pagination": {"enabled": False, "page_size": 100},
    }


def _sheet_key_for_title(title: str) -> str:
    if title == "报名表":
        return "registration"
    if title == "考勤表":
        return "attendance"
    if "5月22日以后" in title:
        return "attendance_after_20250522"
    if "5月22日以前" in title:
        return "attendance_before_20250522"
    return re.sub(r"[^a-zA-Z0-9_-]+", "-", title).strip("-").lower() or "sheet"


def _build_documents(xlsx_path: Path) -> list[tuple[str, str, dict[str, Any]]]:
    workbook = load_workbook(xlsx_path, read_only=False, data_only=False)
    try:
        documents: list[tuple[str, str, dict[str, Any]]] = []
        ordered_names = ["报名表", "考勤表"]
        ordered_names.extend(name for name in workbook.sheetnames if name not in ordered_names)
        for sheet_name in ordered_names:
            if sheet_name not in workbook.sheetnames:
                continue
            worksheet = workbook[sheet_name]
            if sheet_name == "报名表":
                document = _registration_document(worksheet)
            else:
                document = _attendance_document(worksheet)
            documents.append((_sheet_key_for_title(sheet_name), sheet_name, document))
        return documents
    finally:
        workbook.close()


def _ensure_workbook_identity(session: Session, workbook: WorkbookDocument) -> None:
    ensure_resource_identity(session, RESOURCE_TYPE_WORKBOOK, str(workbook.legacy_id or workbook.id), None)


def _ensure_sheet_identity(session: Session, sheet: SheetDocument) -> None:
    sheet.numeric_id = ensure_resource_identity(
        session,
        RESOURCE_TYPE_SHEET,
        str(sheet.legacy_id or sheet.id),
        int(sheet.numeric_id or 0) if sheet.numeric_id else None,
    )


def _find_or_create_workbook(
    session: Session,
    *,
    title: str,
    owner_user_id: int,
    replace: bool,
) -> WorkbookDocument:
    workbook = session.exec(
        select(WorkbookDocument)
        .where(WorkbookDocument.title == title)
        .where(WorkbookDocument.owner_user_id == owner_user_id)
    ).first()
    if workbook is not None:
        if not replace:
            raise SystemExit(f"工作簿已存在：{title}。如需覆盖请加 --replace")
        workbook.updated_by_user_id = owner_user_id
        workbook.updated_at = time.time()
        session.add(workbook)
        _ensure_workbook_identity(session, workbook)
        return workbook

    now = time.time()
    workbook_identity = allocate_new_workbook_identity(session)
    workbook = WorkbookDocument(
        id=workbook_identity.primary_id,
        numeric_id=workbook_identity.numeric_id,
        legacy_id=workbook_identity.legacy_id,
        title=title,
        owner_user_id=owner_user_id,
        created_by_user_id=owner_user_id,
        updated_by_user_id=owner_user_id,
        created_at=now,
        updated_at=now,
    )
    session.add(workbook)
    session.flush()
    _ensure_workbook_identity(session, workbook)
    return workbook


def _upsert_sheet(
    session: Session,
    *,
    owner_key: str,
    sheet_key: str,
    title: str,
    document_json: dict[str, Any],
    owner_user_id: int,
) -> SheetDocument:
    sheet = session.exec(
        select(SheetDocument)
        .where(SheetDocument.scope == "notes")
        .where(SheetDocument.owner_type == "course_workbook")
        .where(SheetDocument.owner_key == owner_key)
        .where(SheetDocument.sheet_key == sheet_key)
    ).first()
    now = time.time()
    if sheet is None:
        sheet_identity = allocate_new_sheet_identity(session)
        sheet = SheetDocument(
            id=sheet_identity.primary_id,
            numeric_id=sheet_identity.numeric_id,
            legacy_id=sheet_identity.legacy_id,
            scope="notes",
            owner_type="course_workbook",
            owner_key=owner_key,
            sheet_key=sheet_key,
            title=title,
            engine="handsontable",
            document_json=document_json,
            version=1,
            owner_user_id=owner_user_id,
            created_by_user_id=owner_user_id,
            updated_by_user_id=owner_user_id,
            created_at=now,
            updated_at=now,
        )
    else:
        sheet.title = title
        sheet.document_json = document_json
        sheet.version = max(int(sheet.version or 1), 1) + 1
        sheet.owner_user_id = owner_user_id
        sheet.updated_by_user_id = owner_user_id
        sheet.updated_at = now
    session.add(sheet)
    session.flush()
    _ensure_sheet_identity(session, sheet)
    return sheet


def _ensure_workbook_link(
    session: Session,
    *,
    workbook: WorkbookDocument,
    sheet: SheetDocument,
    order_index: int,
) -> None:
    workbook_ref = workbook_public_id(workbook)
    sheet_ref = sheet_public_id(sheet)
    link = session.exec(
        select(WorkbookSheetLink)
        .where(WorkbookSheetLink.workbook_id == workbook_ref)
        .where(WorkbookSheetLink.sheet_id == sheet_ref)
    ).first()
    if link is None:
        link = session.exec(
            select(WorkbookSheetLink)
            .where(WorkbookSheetLink.workbook_id.in_(workbook_ref_aliases(workbook)))
            .where(WorkbookSheetLink.sheet_id.in_(sheet_ref_aliases(sheet)))
        ).first()
    if link is None:
        session.add(
            WorkbookSheetLink(
                workbook_id=workbook_ref,
                sheet_id=sheet_ref,
                order_index=order_index,
                created_at=time.time(),
            )
        )
    else:
        changed = False
        if link.workbook_id != workbook_ref:
            link.workbook_id = workbook_ref
            changed = True
        if link.sheet_id != sheet_ref:
            link.sheet_id = sheet_ref
            changed = True
        if int(link.order_index or 0) != order_index:
            link.order_index = order_index
            changed = True
        if changed:
            session.add(link)


def _link_summary_row(
    session: Session,
    *,
    online_sheet_name: str,
    target_workbook_id: int,
    target_sheet_id: int,
    owner_user_id: int,
) -> bool:
    summary = session.exec(select(SheetDocument).where(SheetDocument.numeric_id == 4)).first()
    if summary is None:
        return False
    document = dict(summary.document_json or {})
    columns = list(document.get("columns") or [])
    try:
        online_index = columns.index("在线考勤表")
    except ValueError:
        return False
    rows = list(document.get("rows") or [])
    matched = False
    cell_meta = dict(document.get("cell_meta") or {})
    for row_index, source_row in enumerate(rows):
        row = list(source_row) if isinstance(source_row, list) else []
        value = _normalize_text(row[online_index] if online_index < len(row) else "")
        if value != online_sheet_name:
            continue
        key = f"{row_index}:{online_index}"
        meta = dict(cell_meta.get(key) or {})
        meta["link"] = {"url": f"/workbook/{target_workbook_id}?sheet={target_sheet_id}"}
        cell_meta[key] = meta
        matched = True
        break
    if not matched:
        return False
    document["cell_meta"] = cell_meta
    grid_rows = list(document.get("grid_rows") or [])
    data_start_row = int(document.get("data_start_row") or 0)
    if grid_rows:
        document["grid_rows"] = [*grid_rows[:data_start_row], *rows]
    summary.document_json = document
    summary.version = max(int(summary.version or 1), 1) + 1
    summary.updated_by_user_id = owner_user_id
    summary.updated_at = time.time()
    session.add(summary)
    return True


def import_workbook(
    *,
    xlsx_path: Path,
    title: str,
    owner_key: str,
    owner_user_id: int,
    replace: bool,
    link_summary: bool,
) -> dict[str, Any]:
    documents = _build_documents(xlsx_path)
    with Session(engine) as session:
        owner = session.get(User, owner_user_id)
        if owner is None:
            raise SystemExit(f"用户不存在：{owner_user_id}")
        workbook = _find_or_create_workbook(
            session,
            title=title,
            owner_user_id=owner_user_id,
            replace=replace,
        )

        sheets: list[SheetDocument] = []
        for index, (sheet_key, sheet_title, document_json) in enumerate(documents):
            sheet = _upsert_sheet(
                session,
                owner_key=owner_key,
                sheet_key=sheet_key,
                title=sheet_title,
                document_json=document_json,
                owner_user_id=owner_user_id,
            )
            _ensure_workbook_link(session, workbook=workbook, sheet=sheet, order_index=(index + 1) * 10)
            sheets.append(sheet)

        workbook.updated_by_user_id = owner_user_id
        workbook.updated_at = time.time()
        session.add(workbook)
        summary_linked = False
        if link_summary:
            attendance_sheet = next((sheet for sheet in sheets if sheet.sheet_key == "attendance"), sheets[0])
            session.flush()
            summary_linked = _link_summary_row(
                session,
                online_sheet_name=title,
                target_workbook_id=int(workbook.numeric_id or 0),
                target_sheet_id=int(attendance_sheet.numeric_id or 0),
                owner_user_id=owner_user_id,
            )
        session.commit()
        return {
            "workbook_id": workbook.numeric_id,
            "title": workbook.title,
            "sheets": [
                {
                    "id": sheet.numeric_id,
                    "title": sheet.title,
                    "sheet_key": sheet.sheet_key,
                    "rows": len((sheet.document_json or {}).get("rows") or []),
                }
                for sheet in sheets
            ],
            "summary_linked": summary_linked,
        }


def main() -> None:
    parser = argparse.ArgumentParser(description="Import legacy attendance workbook xlsx into CodeYun note sheets.")
    parser.add_argument("xlsx_path", type=Path)
    parser.add_argument("--title", default="20250106念住闯关")
    parser.add_argument("--owner-key", default="20250106-nianzhu-chuangguan")
    parser.add_argument("--owner-user-id", type=int, default=2)
    parser.add_argument("--replace", action="store_true", help="Replace an existing workbook with the same title.")
    parser.add_argument("--link-summary", action="store_true", help="Link workbook=2 sheet=4 online sheet cell to the imported workbook.")
    args = parser.parse_args()

    result = import_workbook(
        xlsx_path=args.xlsx_path,
        title=args.title,
        owner_key=args.owner_key,
        owner_user_id=args.owner_user_id,
        replace=args.replace,
        link_summary=args.link_summary,
    )
    print(result)


if __name__ == "__main__":
    main()
