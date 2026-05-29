"""把第41届念住重建为“40届业务结构 + 闯关技术框架”。

业务层以 `20260501第40届念住.xlsx` 为模板：
- 考勤表列结构、课程列、返款提示沿用第40届；
- 不引入 20250106 念住闯关里的“第N届答疑”课程列；
- 视频返款按第40届的“当堂/第1天/第2天/第3天/回放”文本口径。

技术层沿用念住闯关框架：
- 工作簿下保留/创建 考勤表、报名表、视频配置、视频数据、打卡配置、打卡数据；
- 源数据与配置 sheet 由 `backend.core.nianzhu_course_sheets` 物化生成。
"""
from __future__ import annotations

import argparse
from datetime import date, datetime, timedelta
import copy
from pathlib import Path
import re
import sys
import time
from typing import Any

from openpyxl import load_workbook
from sqlmodel import Session, select

REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from backend.core.nianzhu_course_sheets import (
    CLOCKIN_CONFIG_COLUMNS,
    CLOCKIN_CONFIG_SHEET_KEY,
    CLOCKIN_DATA_COLUMNS,
    CLOCKIN_DATA_SHEET_KEY,
    VIDEO_CONFIG_COLUMNS,
    VIDEO_CONFIG_SHEET_KEY,
    VIDEO_DATA_COLUMNS,
    VIDEO_DATA_SHEET_KEY,
)
from backend.core.sheet_identity import allocate_new_sheet_identity
from backend.core.sheet_refs import sheet_public_id, sheet_ref_aliases, workbook_public_id, workbook_ref_aliases
from backend.db import engine
from backend.models import SheetDocument, WorkbookDocument, WorkbookSheetLink
from scripts.import_legacy_attendance_workbook import _attendance_document, _registration_document


DEFAULT_TEMPLATE_PATH = Path(r"C:\Users\kzche\Downloads\20260501第40届念住.xlsx")
WORKBOOK_NUMERIC_ID = 10
ATTENDANCE_SHEET_NUMERIC_ID = 54605
REGISTRATION_SHEET_NUMERIC_ID = 54606
COURSE_NAME = "d260601第41届念住"
OWNER_KEY = "20260601-nianzhu-41"
TEMPLATE_COURSE_START_DATE = date(2026, 5, 1)
COURSE_START_DATE = date(2026, 6, 1)
COURSE_LESSON_CONFIG = [
    {"lesson_id2": "l_6a18441ce4b0694c5bcc1084", "start_date": "2026-06-01 05:20:00", "next_update": "2026-06-01 06:42:40", "video_duration": 4960},
    {"lesson_id2": "l_6a18441ee4b0694c35109bbb", "start_date": "2026-06-02 05:20:00", "next_update": "2026-06-02 06:51:11", "video_duration": 5471},
    {"lesson_id2": "l_6a184420e4b0694c5bcc108a", "start_date": "2026-06-03 05:20:00", "next_update": "2026-06-03 06:41:52", "video_duration": 4912},
    {"lesson_id2": "l_6a184422e4b0694c35109bbd", "start_date": "2026-06-04 05:20:00", "next_update": "2026-06-04 06:48:06", "video_duration": 5286},
    {"lesson_id2": "l_6a184424e4b0694c35109bbf", "start_date": "2026-06-05 05:20:00", "next_update": "2026-06-05 06:16:57", "video_duration": 3417},
    {"lesson_id2": "l_6a184426e4b0694c5bcc108c", "start_date": "2026-06-06 05:20:00", "next_update": "2026-06-06 06:29:55", "video_duration": 4195},
    {"lesson_id2": "l_6a184428e4b0694c5bcc108e", "start_date": "2026-06-07 05:20:00", "next_update": "2026-06-07 06:23:31", "video_duration": 3811},
    {"lesson_id2": "l_6a18442ae4b0694c5bcc1092", "start_date": "2026-06-08 05:20:00", "next_update": "2026-06-08 06:25:53", "video_duration": 3953},
    {"lesson_id2": "l_6a18442de4b0694c35109bc3", "start_date": "2026-06-09 05:20:00", "next_update": "2026-06-09 06:31:10", "video_duration": 4270},
    {"lesson_id2": "l_6a18442fe4b0694c5bcc1094", "start_date": "2026-06-10 05:20:00", "next_update": "2026-06-10 06:04:54", "video_duration": 2694},
    {"lesson_id2": "l_6a184431e4b0694c5bcc1096", "start_date": "2026-06-11 05:20:00", "next_update": "2026-06-11 06:22:06", "video_duration": 3726},
    {"lesson_id2": "l_6a184433e4b0694c35109bc7", "start_date": "2026-06-12 05:20:00", "next_update": "2026-06-12 06:31:30", "video_duration": 4290},
    {"lesson_id2": "l_6a184436e4b0694c5bcc109b", "start_date": "2026-06-13 05:20:00", "next_update": "2026-06-13 06:26:40", "video_duration": 4000},
    {"lesson_id2": "l_6a184439e4b0694c35109bc9", "start_date": "2026-06-14 05:20:00", "next_update": "2026-06-14 06:30:37", "video_duration": 4237},
    {"lesson_id2": "l_6a18443be4b0694c35109bcb", "start_date": "2026-06-15 05:20:00", "next_update": "2026-06-15 06:35:07", "video_duration": 4507},
    {"lesson_id2": "l_6a18443de4b0694c35109bcd", "start_date": "2026-06-16 05:20:00", "next_update": "2026-06-16 07:03:04", "video_duration": 6184},
    {"lesson_id2": "l_6a18443fe4b0694c5bcc109d", "start_date": "2026-06-17 05:20:00", "next_update": "2026-06-17 06:11:45", "video_duration": 3105},
    {"lesson_id2": "l_6a184441e4b0694c5bcc10a2", "start_date": "2026-06-18 05:20:00", "next_update": "2026-06-18 07:00:42", "video_duration": 6042},
    {"lesson_id2": "l_6a184443e4b0694c35109bcf", "start_date": "2026-06-19 05:20:00", "next_update": "2026-06-19 06:17:33", "video_duration": 3453},
    {"lesson_id2": "l_6a184446e4b0694c35109bd1", "start_date": "2026-06-20 05:20:00", "next_update": "2026-06-20 06:24:46", "video_duration": 3886},
    {"lesson_id2": "l_6a184448e4b0694c35109bd3", "start_date": "2026-06-21 05:20:00", "next_update": "2026-06-21 06:18:26", "video_duration": 3506},
]


def _normalize_text(value: Any) -> str:
    return str(value or "").strip()


def _shift_template_day(month: int, day: int) -> date | None:
    try:
        source_day = date(TEMPLATE_COURSE_START_DATE.year, month, day)
    except ValueError:
        return None
    return COURSE_START_DATE + (source_day - TEMPLATE_COURSE_START_DATE)


def _format_chinese_day(day: date) -> str:
    return f"{day.month}月{day.day}日"


def _shift_template_dates(text: str) -> str:
    def replace_chinese_day(match: re.Match[str]) -> str:
        shifted = _shift_template_day(5, int(match.group(1)))
        return _format_chinese_day(shifted) if shifted is not None else match.group(0)

    def replace_iso_day(match: re.Match[str]) -> str:
        separator = match.group(1)
        shifted = _shift_template_day(5, int(match.group(2)))
        if shifted is None:
            return match.group(0)
        return _format_chinese_day(shifted)

    def replace_formula_anchor(match: re.Match[str]) -> str:
        separator = match.group(1)
        anchor = COURSE_START_DATE - timedelta(days=1)
        return f"{anchor.year}{separator}{anchor.month}{separator}{anchor.day}"

    text = re.sub(r"5月(\d{1,2})日", replace_chinese_day, text)
    text = re.sub(r"2026([/-])0?5\1(\d{1,2})", replace_iso_day, text)
    text = re.sub(r"2026([/-])0?4[/-]30", replace_formula_anchor, text)
    return text


def _replace_text(value: Any) -> Any:
    if not isinstance(value, str):
        return value
    text = value.replace("第40届念住", "第41届念住")
    text = text.replace("第40届", "第41届")
    if "最近运行更新时间" in text:
        return "最近运行更新时间：\n待首次同步"
    return _shift_template_dates(text)


def _map_nested_text(value: Any) -> Any:
    if isinstance(value, list):
        return [_map_nested_text(item) for item in value]
    if isinstance(value, dict):
        return {key: _map_nested_text(item) for key, item in value.items()}
    return _replace_text(value)


def _empty_document(document: dict[str, Any]) -> dict[str, Any]:
    next_document = _map_nested_text(copy.deepcopy(document))
    next_document = _adapt_attendance_refund_area(next_document)
    columns = list(next_document.get("columns") or [])
    data_start_row = int(next_document.get("data_start_row") or 0)
    grid_rows = list(next_document.get("grid_rows") or [])
    prefix_rows = grid_rows[:data_start_row] if grid_rows else []
    cell_meta = dict(next_document.get("cell_meta") or {})
    next_document["cell_meta"] = {
        key: value
        for key, value in cell_meta.items()
        if int(str(key).split(":", 1)[0]) < data_start_row
    }
    next_document["rows"] = []
    next_document["grid_rows"] = prefix_rows
    next_document["source_meta"] = {
        **dict(next_document.get("source_meta") or {}),
        "course_name": COURSE_NAME,
        "business_template": "20260501第40届念住",
        "business_structure": "nianzhu_40",
        "columns": len(columns),
    }
    return next_document


def _remove_document_column(document: dict[str, Any], column_index: int) -> dict[str, Any]:
    next_document = copy.deepcopy(document)
    columns = list(next_document.get("columns") or [])
    if column_index < 0 or column_index >= len(columns):
        return next_document

    removed_column = columns[column_index]
    next_document["columns"] = [*columns[:column_index], *columns[column_index + 1:]]
    next_document["rows"] = [
        [*list(row)[:column_index], *list(row)[column_index + 1:]]
        for row in next_document.get("rows") or []
        if isinstance(row, list)
    ]
    next_document["grid_rows"] = [
        [*list(row)[:column_index], *list(row)[column_index + 1:]]
        for row in next_document.get("grid_rows") or []
        if isinstance(row, list)
    ]

    widths = list(next_document.get("column_widths") or [])
    if len(widths) > column_index:
        next_document["column_widths"] = [*widths[:column_index], *widths[column_index + 1:]]

    configs = dict(next_document.get("column_configs") or {})
    configs.pop(removed_column, None)
    next_document["column_configs"] = configs

    cell_meta = dict(next_document.get("cell_meta") or {})
    adjusted_meta: dict[str, Any] = {}
    for key, value in cell_meta.items():
        try:
            row_text, col_text = str(key).split(":", 1)
            col = int(col_text)
        except ValueError:
            adjusted_meta[key] = value
            continue
        if col == column_index:
            continue
        adjusted_col = col - 1 if col > column_index else col
        adjusted_meta[f"{row_text}:{adjusted_col}"] = value
    next_document["cell_meta"] = adjusted_meta

    merged_cells = []
    for item in next_document.get("merged_cells") or []:
        if not isinstance(item, dict):
            continue
        col = int(item.get("col") or 0)
        colspan = int(item.get("colspan") or 1)
        end_col = col + colspan
        if col <= column_index < end_col:
            colspan -= 1
            if colspan <= 0:
                continue
        elif col > column_index:
            col -= 1
        merged_cells.append({**item, "col": col, "colspan": colspan})
    next_document["merged_cells"] = merged_cells

    header_rows = next_document.get("grid_rows") or []
    if header_rows:
        next_document["header_groups"] = [[
            {"label": _normalize_text(value), "colspan": 1}
            for value in list(header_rows[0])
        ]]
    return next_document


def _adapt_attendance_refund_area(document: dict[str, Any]) -> dict[str, Any]:
    columns = list(document.get("columns") or [])
    removed_columns = [name for name in ("返款配置", "已返款") if name in columns]
    if not removed_columns:
        return document
    next_document = copy.deepcopy(document)
    grid_rows = [list(row) for row in next_document.get("grid_rows") or [] if isinstance(row, list)]
    if len(grid_rows) >= 3:
        row = grid_rows[2]
        if len(row) > 14:
            row[12] = row[12]
            row[13] = '="第"&返款周期&"天"'
            row[14] = ""
        next_document["grid_rows"] = grid_rows
    for column_name in removed_columns:
        current_columns = list(next_document.get("columns") or [])
        if column_name in current_columns:
            next_document = _remove_document_column(next_document, current_columns.index(column_name))
    _normalize_attendance_refund_layout(next_document)
    next_document["defined_names"] = _attendance_defined_names()
    _style_refund_faq_cell(next_document)
    _style_progress_note_cell(next_document)
    return next_document


def _normalize_attendance_refund_layout(document: dict[str, Any]) -> None:
    columns = list(document.get("columns") or [])
    required = ["完成视频数", "总应返款", "订单金额", "当前应返款", "打卡数"]
    if any(name not in columns for name in required):
        return

    refund_start = columns.index("完成视频数")
    refund_total_end = columns.index("总应返款")
    operation_start = columns.index("订单金额")
    operation_end = columns.index("当前应返款")
    clockin_index = columns.index("打卡数")

    grid_rows = [list(row) for row in document.get("grid_rows") or [] if isinstance(row, list)]
    if grid_rows:
        row = grid_rows[0]
        for index in range(refund_start, min(clockin_index + 1, len(row))):
            row[index] = ""
        row[refund_start] = "返款总计（数据仅在每天早上更新一次，不是实时更新！）"
        row[operation_start] = "返款操作"
        row[clockin_index] = "打卡"
        document["grid_rows"] = grid_rows

    merged_cells = []
    for item in document.get("merged_cells") or []:
        if not isinstance(item, dict):
            continue
        row_index = int(item.get("row") or 0)
        col = int(item.get("col") or 0)
        colspan = int(item.get("colspan") or 1)
        intersects_refund_header = (
            row_index == 0
            and col < clockin_index + 1
            and col + colspan > refund_start
        )
        if intersects_refund_header:
            continue
        merged_cells.append(item)
    merged_cells.append({
        "row": 0,
        "col": refund_start,
        "rowspan": 1,
        "colspan": refund_total_end - refund_start + 1,
    })
    merged_cells.append({
        "row": 0,
        "col": operation_start,
        "rowspan": 1,
        "colspan": operation_end - operation_start + 1,
    })
    document["merged_cells"] = merged_cells

    header_rows = document.get("grid_rows") or []
    if header_rows:
        document["header_groups"] = [[
            {"label": _normalize_text(value), "colspan": 1}
            for value in list(header_rows[0])
        ]]


def _style_refund_faq_cell(document: dict[str, Any]) -> None:
    grid_rows = document.get("grid_rows") or []
    if len(grid_rows) <= 2:
        return
    row = list(grid_rows[2])
    target_col = next(
        (index for index, value in enumerate(row) if _normalize_text(value) == "考勤返款常见问题解答"),
        -1,
    )
    if target_col < 0:
        return
    cell_meta = dict(document.get("cell_meta") or {})
    key = f"2:{target_col}"
    meta = dict(cell_meta.get(key) or {})
    style = dict(meta.get("style") or {})
    style.update({
        "background_color": "#D9D9D9",
        "text_color": "#FF0000",
        "font_size": 16,
        "text_align": "center",
        "vertical_align": "middle",
    })
    meta["style"] = style
    meta["cell_type"] = "rich_text"
    meta["rich_text"] = {
        "spans": [{
            "start": 0,
            "end": len("考勤返款常见问题解答"),
            "style": {"text_color": "#FF0000", "bold": True},
        }],
    }
    cell_meta[key] = meta
    document["cell_meta"] = cell_meta


def _style_progress_note_cell(document: dict[str, Any]) -> None:
    columns = list(document.get("columns") or [])
    first_lesson_index = next(
        (index for index, column in enumerate(columns) if re.search(r"第\s*0*1\s*课", _normalize_text(column))),
        -1,
    )
    grid_rows = document.get("grid_rows") or []
    if first_lesson_index < 0 or len(grid_rows) <= 2:
        return
    note = _normalize_text(grid_rows[2][first_lesson_index] if first_lesson_index < len(grid_rows[2]) else "")
    if not note:
        return

    highlight = "每节课的视频，观看时长至少达到视频时长的一半，即50%才视为完成。"
    start = note.find(highlight)
    spans = []
    if start >= 0:
        spans.append({
            "start": start,
            "end": start + len(highlight),
            "style": {"text_color": "#FF0000", "bold": True},
        })

    cell_meta = dict(document.get("cell_meta") or {})
    key = f"2:{first_lesson_index}"
    meta = dict(cell_meta.get(key) or {})
    style = dict(meta.get("style") or {})
    style["background_color"] = "#D9D9D9"
    style["text_align"] = "left"
    style["vertical_align"] = "middle"
    meta["style"] = style
    if spans:
        meta["cell_type"] = "rich_text"
        meta["rich_text"] = {"spans": spans}
    cell_meta[key] = meta
    document["cell_meta"] = cell_meta


def _attendance_defined_names() -> list[dict[str, str]]:
    anchor = COURSE_START_DATE - timedelta(days=1)
    return [
        {
            "name": "返款周期",
            "formula": f"=INT(TODAY()-DATE({anchor.year},{anchor.month},{anchor.day}))",
            "comment": "当前课程返款周期天数，按开课前一天到今天计算，支持开课前负数。",
            "scope": "worksheet",
        },
        {
            "name": "返款说明",
            "formula": '="第"&返款周期&"天返款"',
            "comment": "返款操作说明文本。",
            "scope": "worksheet",
        },
        {
            "name": "返款ID后缀",
            "formula": '="_day"&返款周期',
            "comment": "返款批次 ID 后缀。",
            "scope": "worksheet",
        },
    ]


def _simple_document(
    *,
    columns: list[str],
    rows: list[list[Any]],
    source_meta: dict[str, Any] | None = None,
) -> dict[str, Any]:
    return {
        "schema_version": 1,
        "columns": columns,
        "rows": rows,
        "grid_rows": [columns, *rows],
        "data_start_row": 1,
        "field_row_index": 0,
        "column_configs": {},
        "column_widths": [120] * len(columns),
        "cell_meta": {},
        "merged_cells": [],
        "header_groups": [],
        "formula_reference_origin": "sheet_v2",
        "view_settings": {
            "show_row_numbers": True,
            "row_marker_numbering": "page",
            "row_marker_origin": "sheet",
            "show_column_markers": True,
            "column_marker_style": "letters",
            "height_mode": "fill",
            "pagination": {"enabled": True, "page_size": 100},
        },
        "source_meta": dict(source_meta or {}),
    }


def _progress_columns(attendance_document: dict[str, Any]) -> list[str]:
    columns = list(attendance_document.get("columns") or [])
    try:
        start = columns.index("打卡数") + 1
    except ValueError:
        start = next((index for index, column in enumerate(columns) if re.search(r"第\s*0*\d+\s*课", column)), -1)
    if start < 0:
        return []
    return columns[start:]


def _video_config_lesson_name(progress_column: str) -> str:
    text = _normalize_text(progress_column)
    match = re.search(r"第\s*0*\d+\s*课", text)
    if not match:
        return ""
    return match.group(0).replace(" ", "")


def _lesson_end_date(config: dict[str, Any]) -> str:
    end_date = str(config.get("end_date") or "").strip()
    if end_date:
        return end_date
    next_update = str(config.get("next_update") or "").strip()
    if not next_update:
        return ""
    try:
        return (datetime.strptime(next_update, "%Y-%m-%d %H:%M:%S") + timedelta(days=7)).strftime("%Y-%m-%d %H:%M:%S")
    except ValueError:
        return ""


def _course_storage_documents(attendance_document: dict[str, Any]) -> dict[str, dict[str, Any]]:
    progress_columns = _progress_columns(attendance_document)
    lesson_names = [
        lesson_name
        for column in progress_columns
        if (lesson_name := _video_config_lesson_name(column))
    ]
    video_config_rows = [
        [
            index,
            str(config.get("start_date") or ""),
            _lesson_end_date(config),
            str(config.get("next_update") or ""),
            str(config.get("lesson_id2") or ""),
            int(config.get("shop_id") or 1),
            lesson_name,
            int(config.get("video_duration") or 0),
        ]
        for index, lesson_name in enumerate(lesson_names, start=1)
        for config in [COURSE_LESSON_CONFIG[index - 1] if index <= len(COURSE_LESSON_CONFIG) else {}]
    ]
    source_meta = {
        "course_name": COURSE_NAME,
        "business_template": "20260501第40届念住",
        "business_structure": "nianzhu_40",
        "video_refund_rule_mode": "timed_text",
    }
    return {
        VIDEO_CONFIG_SHEET_KEY: _simple_document(
            columns=VIDEO_CONFIG_COLUMNS,
            rows=video_config_rows,
            source_meta={**source_meta, "progress_columns": len(progress_columns), "lesson_count": len(lesson_names)},
        ),
        VIDEO_DATA_SHEET_KEY: _simple_document(
            columns=VIDEO_DATA_COLUMNS,
            rows=[],
            source_meta=source_meta,
        ),
        CLOCKIN_CONFIG_SHEET_KEY: _simple_document(
            columns=CLOCKIN_CONFIG_COLUMNS,
            rows=[[1, "打卡数", "", "", "", "", "", ""]],
            source_meta=source_meta,
        ),
        CLOCKIN_DATA_SHEET_KEY: _simple_document(
            columns=CLOCKIN_DATA_COLUMNS,
            rows=[],
            source_meta=source_meta,
        ),
    }


def _load_template_documents(template_path: Path) -> tuple[dict[str, Any], dict[str, Any]]:
    workbook = load_workbook(template_path, read_only=False, data_only=False)
    try:
        if "考勤表" not in workbook.sheetnames:
            raise RuntimeError(f"{template_path} 缺少 考勤表")
        if "报名表" not in workbook.sheetnames:
            raise RuntimeError(f"{template_path} 缺少 报名表")
        attendance = _empty_document(_attendance_document(workbook["考勤表"]))
        registration = _empty_document(_registration_document(workbook["报名表"]))
        return attendance, registration
    finally:
        workbook.close()


def _require_workbook(session: Session) -> WorkbookDocument:
    workbook = session.exec(
        select(WorkbookDocument).where(WorkbookDocument.numeric_id == WORKBOOK_NUMERIC_ID)
    ).first()
    if workbook is None:
        raise RuntimeError(f"找不到工作簿 numeric_id={WORKBOOK_NUMERIC_ID}")
    return workbook


def _require_sheet(session: Session, numeric_id: int) -> SheetDocument:
    sheet = session.exec(select(SheetDocument).where(SheetDocument.numeric_id == numeric_id)).first()
    if sheet is None:
        raise RuntimeError(f"找不到表格 numeric_id={numeric_id}")
    return sheet


def _update_sheet(
    sheet: SheetDocument,
    *,
    title: str,
    sheet_key: str,
    document: dict[str, Any],
    workbook: WorkbookDocument,
) -> None:
    now = time.time()
    sheet.title = title
    sheet.scope = "notes"
    sheet.owner_type = "course_workbook"
    sheet.owner_key = OWNER_KEY
    sheet.sheet_key = sheet_key
    sheet.engine = "handsontable"
    sheet.document_json = document
    sheet.version = max(int(sheet.version or 1), 1) + 1
    sheet.owner_user_id = sheet.owner_user_id or workbook.owner_user_id
    sheet.updated_by_user_id = workbook.updated_by_user_id or workbook.owner_user_id
    sheet.updated_at = now


def _ensure_link(session: Session, *, workbook: WorkbookDocument, sheet: SheetDocument, order_index: int) -> None:
    workbook_ref = workbook_public_id(workbook)
    sheet_ref = sheet_public_id(sheet)
    link = session.exec(
        select(WorkbookSheetLink)
        .where(WorkbookSheetLink.workbook_id.in_(workbook_ref_aliases(workbook)))
        .where(WorkbookSheetLink.sheet_id.in_(sheet_ref_aliases(sheet)))
    ).first()
    if link is None:
        link = WorkbookSheetLink(
            workbook_id=workbook_ref,
            sheet_id=sheet_ref,
            order_index=order_index,
            created_at=time.time(),
        )
    else:
        link.workbook_id = workbook_ref
        link.sheet_id = sheet_ref
        link.order_index = order_index
    session.add(link)


def _upsert_storage_sheet(
    session: Session,
    *,
    workbook: WorkbookDocument,
    sheet_key: str,
    title: str,
    order_index: int,
    document: dict[str, Any],
) -> SheetDocument:
    sheet = session.exec(
        select(SheetDocument)
        .where(SheetDocument.owner_type == "course_workbook")
        .where(SheetDocument.owner_key == OWNER_KEY)
        .where(SheetDocument.sheet_key == sheet_key)
    ).first()
    if sheet is None:
        identity = allocate_new_sheet_identity(session)
        sheet = SheetDocument(
            id=identity.primary_id,
            numeric_id=identity.numeric_id,
            legacy_id=identity.legacy_id,
            scope="notes",
            owner_type="course_workbook",
            owner_key=OWNER_KEY,
            sheet_key=sheet_key,
            title=title,
            engine="handsontable",
            document_json=document,
            version=1,
            owner_user_id=workbook.owner_user_id,
            created_by_user_id=workbook.created_by_user_id,
            updated_by_user_id=workbook.updated_by_user_id,
            created_at=time.time(),
            updated_at=time.time(),
        )
    else:
        sheet.title = title
        sheet.document_json = document
        sheet.version = max(int(sheet.version or 1), 1) + 1
        sheet.updated_at = time.time()
    sheet.document_json = document
    session.add(sheet)
    session.flush()
    _ensure_link(session, workbook=workbook, sheet=sheet, order_index=order_index)
    return sheet


def run(*, apply: bool, template_path: Path) -> dict[str, Any]:
    attendance_document, registration_document = _load_template_documents(template_path)
    with Session(engine) as session:
        workbook = _require_workbook(session)
        attendance = _require_sheet(session, ATTENDANCE_SHEET_NUMERIC_ID)
        registration = _require_sheet(session, REGISTRATION_SHEET_NUMERIC_ID)

        _update_sheet(
            attendance,
            title="考勤表",
            sheet_key="attendance",
            document=attendance_document,
            workbook=workbook,
        )
        _update_sheet(
            registration,
            title="报名表",
            sheet_key="registration",
            document=registration_document,
            workbook=workbook,
        )
        session.add(attendance)
        session.add(registration)
        _ensure_link(session, workbook=workbook, sheet=attendance, order_index=5)
        _ensure_link(session, workbook=workbook, sheet=registration, order_index=10)

        storage_documents = _course_storage_documents(attendance_document)
        storage_specs = [
            (VIDEO_CONFIG_SHEET_KEY, "视频配置", 30),
            (VIDEO_DATA_SHEET_KEY, "视频数据", 40),
            (CLOCKIN_CONFIG_SHEET_KEY, "打卡配置", 50),
            (CLOCKIN_DATA_SHEET_KEY, "打卡数据", 60),
        ]
        storage_sheets = []
        for sheet_key, title, order_index in storage_specs:
            sheet = _upsert_storage_sheet(
                session,
                workbook=workbook,
                sheet_key=sheet_key,
                title=title,
                order_index=order_index,
                document=storage_documents[sheet_key],
            )
            storage_sheets.append({
                "sheet_key": sheet.sheet_key,
                "title": sheet.title,
                "id": int(sheet.numeric_id or 0),
                "rows": len((sheet.document_json or {}).get("rows") or []),
            })

        workbook.title = COURSE_NAME
        workbook.updated_at = time.time()
        session.add(workbook)

        if apply:
            session.commit()
        else:
            session.rollback()

        return {
            "mode": "APPLY" if apply else "DRY-RUN",
            "workbook_id": WORKBOOK_NUMERIC_ID,
            "attendance_columns": len(attendance_document.get("columns") or []),
            "attendance_rows": len(attendance_document.get("rows") or []),
            "registration_columns": len(registration_document.get("columns") or []),
            "registration_rows": len(registration_document.get("rows") or []),
            "storage_sheets": storage_sheets,
        }


def main() -> None:
    parser = argparse.ArgumentParser(description="重建第41届念住为 40届业务结构 + 闯关技术框架。")
    parser.add_argument("--apply", action="store_true", help="实际写入数据库；默认 dry-run。")
    parser.add_argument("--template", type=Path, default=DEFAULT_TEMPLATE_PATH)
    args = parser.parse_args()
    summary = run(apply=args.apply, template_path=args.template)
    print(summary)


if __name__ == "__main__":
    main()
