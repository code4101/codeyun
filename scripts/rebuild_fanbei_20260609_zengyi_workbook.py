"""基于 20260409 梵呗增益业务表，生成 20260609 梵呗增益 CodeYun 工作簿。

业务层沿用 4 月梵呗增益：
- 考勤表列结构、22 课进度列、返款提示和公式口径来自 4 月模板；
- 不迁移 4 月学员数据、订单数据、视频数据和打卡数据；
- 不复用 4 月小鹅通课次链接，避免 6 月工作簿误指向旧课程。

技术层对齐 41 念住 / 47 觉观：
- 工作簿下创建/更新 考勤表、报名表、视频配置、视频数据、打卡配置、打卡数据；
- 计数源数据落到配置/数据 sheet，由 CodeYun 的梵呗 course_sheets 框架消费。
"""
from __future__ import annotations

import argparse
import copy
from datetime import date, datetime, timedelta
from pathlib import Path
import re
import sys
import time
from typing import Any

from openpyxl import load_workbook
from sqlmodel import Session, select

REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from backend.core.fanbei_course_sheets import (  # noqa: E402
    CLOCKIN_CONFIG_COLUMNS,
    CLOCKIN_CONFIG_SHEET_KEY,
    CLOCKIN_DATA_COLUMNS,
    CLOCKIN_DATA_SHEET_KEY,
    VIDEO_CONFIG_COLUMNS,
    VIDEO_CONFIG_SHEET_KEY,
    VIDEO_DATA_COLUMNS,
    VIDEO_DATA_SHEET_KEY,
)
from backend.core.note_sheet_access import ensure_attendance_sheet_anonymous_viewer  # noqa: E402
from backend.core.note_sheet_inline_links import extract_inline_cell_value  # noqa: E402
from backend.core.sheet_identity import allocate_new_sheet_identity, allocate_new_workbook_identity  # noqa: E402
from backend.core.sheet_refs import sheet_public_id, sheet_ref_aliases, workbook_public_id, workbook_ref_aliases  # noqa: E402
from backend.db import engine  # noqa: E402
from backend.models import SheetDocument, WorkbookDocument, WorkbookSheetLink  # noqa: E402
from scripts.import_legacy_attendance_workbook import _attendance_document, _registration_document  # noqa: E402


DEFAULT_TEMPLATE_PATH = Path(r"C:\Users\kzche\Downloads\20260409梵呗增益.xlsx")
TEMPLATE_COURSE_START_DATE = date(2026, 4, 9)
COURSE_START_DATE = date(2026, 6, 9)
COURSE_NAME = "d260609梵呗增益"
WORKBOOK_TITLE = "20260609梵呗增益"
SUMMARY_ONLINE_SHEET_NAME = "20260609梵呗增益"
OWNER_KEY = "20260609-fanbei-zengyi"
OFFICIAL_LESSON_COUNT = 22
SHOP_ID = 1


def _text(value: Any) -> str:
    return str(value or "").strip()


def _cell_text(value: Any) -> str:
    return _text(extract_inline_cell_value(value))


def _shift_date(value: date) -> date:
    return COURSE_START_DATE + (value - TEMPLATE_COURSE_START_DATE)


def _format_chinese_day(value: date) -> str:
    return f"{value.month}月{value.day}日"


def _shift_template_dates(text: str) -> str:
    def replace_chinese_day(match: re.Match[str]) -> str:
        try:
            source = date(TEMPLATE_COURSE_START_DATE.year, int(match.group("month")), int(match.group("day")))
        except ValueError:
            return match.group(0)
        return _format_chinese_day(_shift_date(source))

    def replace_iso_day(match: re.Match[str]) -> str:
        separator = match.group("sep")
        try:
            source = date(int(match.group("year")), int(match.group("month")), int(match.group("day")))
        except ValueError:
            return match.group(0)
        shifted = _shift_date(source)
        return f"{shifted.year}{separator}{shifted.month}{separator}{shifted.day}"

    text = re.sub(r"(?P<month>[45])月(?P<day>\d{1,2})日", replace_chinese_day, text)
    text = re.sub(
        r"(?P<year>2026)(?P<sep>[/-])0?(?P<month>[45])(?P=sep)(?P<day>\d{1,2})",
        replace_iso_day,
        text,
    )
    return (
        text.replace("20260409梵呗增益", SUMMARY_ONLINE_SHEET_NAME)
        .replace("4月梵呗增益", "6月梵呗增益")
        .replace("d260409梵呗增益", COURSE_NAME)
    )


def _replace_text(value: Any) -> Any:
    if not isinstance(value, str):
        return value
    if "最近运行更新时间" in value:
        return "待首次同步"
    return _shift_template_dates(value)


def _map_nested_text(value: Any) -> Any:
    if isinstance(value, list):
        return [_map_nested_text(item) for item in value]
    if isinstance(value, dict):
        return {key: _map_nested_text(item) for key, item in value.items()}
    return _replace_text(value)


def _clear_data_rows(document: dict[str, Any]) -> dict[str, Any]:
    next_document = _map_nested_text(copy.deepcopy(document))
    data_start_row = int(next_document.get("data_start_row") or 0)
    grid_rows = [list(row) for row in next_document.get("grid_rows") or [] if isinstance(row, list)]
    next_document["rows"] = []
    next_document["grid_rows"] = grid_rows[:data_start_row]
    next_document["cell_meta"] = {
        key: value
        for key, value in dict(next_document.get("cell_meta") or {}).items()
        if int(str(key).split(":", 1)[0]) < data_start_row
    }
    return next_document


def _set_cell_link(document: dict[str, Any], row: int, col: int, url: str | None) -> None:
    cell_meta = dict(document.get("cell_meta") or {})
    key = f"{row}:{col}"
    meta = dict(cell_meta.get(key) or {})
    if url:
        meta["link"] = {"url": url}
    else:
        meta.pop("link", None)
    cell_meta[key] = meta
    document["cell_meta"] = cell_meta


def _set_cell_style(document: dict[str, Any], row: int, col: int, style: dict[str, Any]) -> None:
    cell_meta = dict(document.get("cell_meta") or {})
    key = f"{row}:{col}"
    meta = dict(cell_meta.get(key) or {})
    current_style = dict(meta.get("style") or {})
    current_style.update(style)
    meta["style"] = current_style
    cell_meta[key] = meta
    document["cell_meta"] = cell_meta


def _remove_document_column(document: dict[str, Any], column_index: int) -> None:
    columns = list(document.get("columns") or [])
    if not 0 <= column_index < len(columns):
        return
    removed_column = columns[column_index]
    document["columns"] = columns[:column_index] + columns[column_index + 1:]
    for key in ("rows", "grid_rows"):
        document[key] = [
            list(row)[:column_index] + list(row)[column_index + 1:]
            for row in document.get(key) or []
            if isinstance(row, list)
        ]
    widths = list(document.get("column_widths") or [])
    if len(widths) > column_index:
        document["column_widths"] = widths[:column_index] + widths[column_index + 1:]
    configs = dict(document.get("column_configs") or {})
    configs.pop(removed_column, None)
    document["column_configs"] = configs

    adjusted_meta: dict[str, Any] = {}
    for key, value in dict(document.get("cell_meta") or {}).items():
        try:
            row_text, col_text = str(key).split(":", 1)
            col = int(col_text)
        except ValueError:
            adjusted_meta[key] = value
            continue
        if col == column_index:
            continue
        adjusted_meta[f"{row_text}:{col - 1 if col > column_index else col}"] = value
    document["cell_meta"] = adjusted_meta

    merged_cells = []
    for item in document.get("merged_cells") or []:
        if not isinstance(item, dict):
            continue
        col = int(item.get("col") or 0)
        colspan = int(item.get("colspan") or 1)
        if col <= column_index < col + colspan:
            colspan -= 1
            if colspan <= 0:
                continue
        elif col > column_index:
            col -= 1
        merged_cells.append({**item, "col": col, "colspan": colspan})
    document["merged_cells"] = merged_cells


def _normalize_refund_layout(document: dict[str, Any]) -> None:
    columns = list(document.get("columns") or [])
    required = ["完成视频数", "总应返款", "订单金额", "当前应返款", "打卡数"]
    if any(column not in columns for column in required):
        return

    refund_start = columns.index("完成视频数")
    refund_end = columns.index("总应返款")
    operation_start = columns.index("已返款") if "已返款" in columns else columns.index("订单金额")
    operation_end = columns.index("当前应返款")
    clockin_index = columns.index("打卡数")
    grid_rows = [list(row) for row in document.get("grid_rows") or [] if isinstance(row, list)]
    if grid_rows:
        for col in range(refund_start, min(clockin_index + 1, len(grid_rows[0]))):
            grid_rows[0][col] = ""
        grid_rows[0][refund_start] = "返款总计（数据仅在每天早上更新一次，不是实时更新！）"
        grid_rows[0][operation_start] = "返款操作"
        grid_rows[0][clockin_index] = "打卡"
        document["grid_rows"] = grid_rows

    merged = []
    for item in document.get("merged_cells") or []:
        if not isinstance(item, dict):
            continue
        is_refund_header = int(item.get("row") or 0) == 0 and int(item.get("col") or 0) < clockin_index + 1
        if is_refund_header and int(item.get("col") or 0) + int(item.get("colspan") or 1) > refund_start:
            continue
        merged.append(item)
    merged.append({"row": 0, "col": refund_start, "rowspan": 1, "colspan": refund_end - refund_start + 1})
    merged.append({"row": 0, "col": operation_start, "rowspan": 1, "colspan": operation_end - operation_start + 1})
    document["merged_cells"] = merged


def _attendance_defined_names() -> list[dict[str, str]]:
    anchor = COURSE_START_DATE - timedelta(days=1)
    return [
        {
            "name": "返款周期",
            "formula": f"=INT(TODAY()-DATE({anchor.year},{anchor.month},{anchor.day}))",
            "comment": "当前梵呗增益返款周期天数，按开课前一天到今天计算。",
            "scope": "worksheet",
        },
        {
            "name": "返款说明",
            "formula": '="6月梵呗增益第"&返款周期&"天返款"',
            "comment": "返款操作说明文本。",
            "scope": "worksheet",
        },
        {
            "name": "返款ID后缀",
            "formula": '="_daya"&返款周期',
            "comment": "梵呗增益返款批次 ID 后缀。",
            "scope": "worksheet",
        },
    ]


def _parse_lesson_number(value: Any) -> int | None:
    match = re.search(r"第\s*0*(\d+)\s*课", _text(value))
    return int(match.group(1)) if match else None


def _parse_lesson_times(value: Any) -> tuple[str, str] | None:
    match = re.search(r"(?P<start>\d{1,2}:\d{2})\s*~\s*(?P<end>\d{1,2}:\d{2})", _text(value))
    if not match:
        return None
    return match.group("start"), match.group("end")


def _lesson_duration_seconds(start_time: str, end_time: str) -> int:
    start = datetime.strptime(start_time, "%H:%M")
    end = datetime.strptime(end_time, "%H:%M")
    if end < start:
        end += timedelta(days=1)
    return int((end - start).total_seconds())


def _adapt_attendance_document(document: dict[str, Any]) -> dict[str, Any]:
    doc = _clear_data_rows(document)
    columns = list(doc.get("columns") or [])
    if "返款配置" in columns:
        _remove_document_column(doc, columns.index("返款配置"))
    columns = list(doc.get("columns") or [])
    grid_rows = [list(row) for row in doc.get("grid_rows") or [] if isinstance(row, list)]
    if len(grid_rows) < 3:
        raise RuntimeError("考勤表模板需要至少 3 行表头")

    lesson_columns = [
        (index, number)
        for index, column in enumerate(columns)
        if (number := _parse_lesson_number(column)) is not None
    ]
    for column_index, lesson_number in lesson_columns:
        lesson_date = COURSE_START_DATE + timedelta(days=lesson_number - 1)
        if column_index < len(grid_rows[0]):
            grid_rows[0][column_index] = f"{_format_chinese_day(lesson_date)}~{_format_chinese_day(lesson_date + timedelta(days=5))}"
        _set_cell_link(doc, 1, column_index, None)

    for column_name in ("打卡数",):
        if column_name in columns:
            _set_cell_link(doc, 1, columns.index(column_name), None)
    if "完成视频数" in columns:
        grid_rows[2][columns.index("完成视频数")] = ""
    if "已返款" in columns:
        grid_rows[2][columns.index("已返款")] = '="第"&返款周期&"天"'
    if "订单金额" in columns:
        grid_rows[2][columns.index("订单金额")] = 500
    if "当前应返款" in columns:
        grid_rows[2][columns.index("当前应返款")] = "待首次同步"

    doc["grid_rows"] = grid_rows
    _normalize_refund_layout(doc)
    doc["defined_names"] = _attendance_defined_names()
    doc["source_meta"] = {
        **dict(doc.get("source_meta") or {}),
        "course_name": COURSE_NAME,
        "workbook_title": WORKBOOK_TITLE,
        "business_template": "20260409梵呗增益",
        "business_structure": "fanbei_zengyi_20260409",
        "official_lesson_count": OFFICIAL_LESSON_COUNT,
        "video_refund_rule_mode": "timed_text",
        "columns": len(columns),
    }

    if "考勤返款常见问题解答" in grid_rows[2]:
        faq_col = grid_rows[2].index("考勤返款常见问题解答")
        _set_cell_style(doc, 2, faq_col, {"font_size": 16, "text_color": "#FF0000", "text_align": "center"})
    return doc


def _adapt_registration_document(document: dict[str, Any]) -> dict[str, Any]:
    doc = _clear_data_rows(document)
    doc["source_meta"] = {
        **dict(doc.get("source_meta") or {}),
        "course_name": COURSE_NAME,
        "business_template": "20260409梵呗增益",
        "business_structure": "fanbei_zengyi_registration",
        "columns": len(doc.get("columns") or []),
    }
    return doc


def _simple_document(columns: list[str], rows: list[list[Any]], source_meta: dict[str, Any] | None = None) -> dict[str, Any]:
    return {
        "schema_version": 1,
        "columns": columns,
        "rows": rows,
        "grid_rows": [columns, *rows],
        "data_start_row": 1,
        "field_row_index": 0,
        "column_configs": {},
        "column_widths": [120] * len(columns),
        "cell_meta": {},
        "merged_cells": [],
        "header_groups": [],
        "formula_reference_origin": "sheet_v2",
        "view_settings": {
            "show_row_numbers": True,
            "row_marker_numbering": "page",
            "row_marker_origin": "sheet",
            "show_column_markers": True,
            "column_marker_style": "letters",
            "height_mode": "fill",
            "pagination": {"enabled": True, "page_size": 100},
        },
        "source_meta": dict(source_meta or {}),
    }


def _course_storage_documents(attendance_document: dict[str, Any]) -> dict[str, dict[str, Any]]:
    columns = list(attendance_document.get("columns") or [])
    video_config_rows: list[list[Any]] = []
    for column in columns:
        lesson_number = _parse_lesson_number(column)
        times = _parse_lesson_times(column)
        if lesson_number is None or times is None or lesson_number > OFFICIAL_LESSON_COUNT:
            continue
        start_time, end_time = times
        lesson_date = COURSE_START_DATE + timedelta(days=lesson_number - 1)
        start_at = datetime.strptime(f"{lesson_date.isoformat()} {start_time}:00", "%Y-%m-%d %H:%M:%S")
        next_update = datetime.strptime(f"{lesson_date.isoformat()} {end_time}:00", "%Y-%m-%d %H:%M:%S")
        if next_update < start_at:
            next_update += timedelta(days=1)
        video_config_rows.append([
            lesson_number,
            start_at.strftime("%Y-%m-%d %H:%M:%S"),
            (next_update + timedelta(days=5)).strftime("%Y-%m-%d %H:%M:%S"),
            next_update.strftime("%Y-%m-%d %H:%M:%S"),
            "",
            SHOP_ID,
            f"第{lesson_number:02d}课",
            _lesson_duration_seconds(start_time, end_time),
        ])

    source_meta = {
        "course_name": COURSE_NAME,
        "workbook_title": WORKBOOK_TITLE,
        "business_template": "20260409梵呗增益",
        "business_structure": "fanbei_zengyi_20260409",
        "official_lesson_count": OFFICIAL_LESSON_COUNT,
        "video_refund_rule_mode": "timed_text",
        "video_config_note": "lesson_id2 需要在 6 月真实课次创建后补齐；不要复用 4 月模板链接。",
    }
    return {
        VIDEO_CONFIG_SHEET_KEY: _simple_document(
            VIDEO_CONFIG_COLUMNS,
            video_config_rows,
            {**source_meta, "lesson_count": len(video_config_rows)},
        ),
        VIDEO_DATA_SHEET_KEY: _simple_document(VIDEO_DATA_COLUMNS, [], source_meta),
        CLOCKIN_CONFIG_SHEET_KEY: _simple_document(
            CLOCKIN_CONFIG_COLUMNS,
            [[1, "打卡数", "", "", "", "", "", ""]],
            source_meta,
        ),
        CLOCKIN_DATA_SHEET_KEY: _simple_document(CLOCKIN_DATA_COLUMNS, [], source_meta),
    }


def _load_documents(template_path: Path) -> tuple[dict[str, Any], dict[str, Any]]:
    workbook = load_workbook(template_path, read_only=False, data_only=False)
    try:
        if "考勤表" not in workbook.sheetnames:
            raise RuntimeError(f"{template_path} 缺少 考勤表")
        if "报名表" not in workbook.sheetnames:
            raise RuntimeError(f"{template_path} 缺少 报名表")
        return (
            _adapt_attendance_document(_attendance_document(workbook["考勤表"])),
            _adapt_registration_document(_registration_document(workbook["报名表"])),
        )
    finally:
        workbook.close()


def _find_or_create_workbook(session: Session, owner_user_id: int) -> WorkbookDocument:
    workbook = session.exec(
        select(WorkbookDocument).where(WorkbookDocument.title == WORKBOOK_TITLE)
    ).first()
    if workbook is not None:
        return workbook

    identity = allocate_new_workbook_identity(session)
    now = time.time()
    workbook = WorkbookDocument(
        id=identity.primary_id,
        numeric_id=identity.numeric_id,
        legacy_id=identity.legacy_id,
        title=WORKBOOK_TITLE,
        owner_user_id=owner_user_id,
        created_by_user_id=owner_user_id,
        updated_by_user_id=owner_user_id,
        created_at=now,
        updated_at=now,
    )
    session.add(workbook)
    session.flush()
    return workbook


def _upsert_sheet(
    session: Session,
    workbook: WorkbookDocument,
    sheet_key: str,
    title: str,
    document: dict[str, Any],
) -> SheetDocument:
    sheet = session.exec(
        select(SheetDocument)
        .where(SheetDocument.owner_type == "course_workbook")
        .where(SheetDocument.owner_key == OWNER_KEY)
        .where(SheetDocument.sheet_key == sheet_key)
    ).first()
    now = time.time()
    if sheet is None:
        identity = allocate_new_sheet_identity(session)
        sheet = SheetDocument(
            id=identity.primary_id,
            numeric_id=identity.numeric_id,
            legacy_id=identity.legacy_id,
            scope="notes",
            owner_type="course_workbook",
            owner_key=OWNER_KEY,
            sheet_key=sheet_key,
            title=title,
            engine="handsontable",
            document_json=document,
            version=1,
            owner_user_id=workbook.owner_user_id,
            created_by_user_id=workbook.created_by_user_id,
            updated_by_user_id=workbook.updated_by_user_id,
            created_at=now,
            updated_at=now,
        )
    else:
        sheet.title = title
        sheet.document_json = document
        sheet.version = max(int(sheet.version or 1), 1) + 1
        sheet.updated_by_user_id = workbook.updated_by_user_id or workbook.owner_user_id
        sheet.updated_at = now
    session.add(sheet)
    session.flush()
    return sheet


def _ensure_link(session: Session, workbook: WorkbookDocument, sheet: SheetDocument, order_index: int) -> None:
    workbook_ref = workbook_public_id(workbook)
    sheet_ref = sheet_public_id(sheet)
    link = session.exec(
        select(WorkbookSheetLink)
        .where(WorkbookSheetLink.workbook_id.in_(workbook_ref_aliases(workbook)))
        .where(WorkbookSheetLink.sheet_id.in_(sheet_ref_aliases(sheet)))
    ).first()
    if link is None:
        link = WorkbookSheetLink(workbook_id=workbook_ref, sheet_id=sheet_ref, order_index=order_index, created_at=time.time())
    else:
        link.workbook_id = workbook_ref
        link.sheet_id = sheet_ref
        link.order_index = order_index
    session.add(link)


def _link_summary_row(session: Session, workbook: WorkbookDocument, attendance: SheetDocument) -> bool:
    summary = session.exec(select(SheetDocument).where(SheetDocument.numeric_id == 4)).first()
    if summary is None:
        return False
    document = copy.deepcopy(dict(summary.document_json or {}))
    columns = list(document.get("columns") or [])
    rows = [list(row) for row in document.get("rows") or [] if isinstance(row, list)]
    try:
        online_col = columns.index("在线考勤表")
    except ValueError:
        return False

    target_row = next(
        (
            index
            for index, row in enumerate(rows)
            if online_col < len(row) and _cell_text(row[online_col]) == SUMMARY_ONLINE_SHEET_NAME
        ),
        -1,
    )
    if target_row < 0:
        return False

    rows[target_row][online_col] = {
        "value": SUMMARY_ONLINE_SHEET_NAME,
        "link": {"url": f"/workbook/{int(workbook.numeric_id or 0)}?sheet={int(attendance.numeric_id or 0)}"},
    }
    cell_meta = dict(document.get("cell_meta") or {})
    cell_meta.pop(f"{target_row}:{online_col}", None)
    document["cell_meta"] = cell_meta
    document["rows"] = rows
    data_start = int(document.get("data_start_row") or 0)
    document["grid_rows"] = list(document.get("grid_rows") or [])[:data_start] + rows
    summary.document_json = document
    summary.version = max(int(summary.version or 1), 1) + 1
    summary.updated_by_user_id = workbook.updated_by_user_id or workbook.owner_user_id
    summary.updated_at = time.time()
    session.add(summary)
    return True


def run(*, apply: bool, template_path: Path, owner_user_id: int = 2) -> dict[str, Any]:
    attendance_document, registration_document = _load_documents(template_path)
    storage_documents = _course_storage_documents(attendance_document)
    with Session(engine) as session:
        workbook = _find_or_create_workbook(session, owner_user_id)
        workbook.title = WORKBOOK_TITLE
        workbook.updated_by_user_id = owner_user_id
        workbook.updated_at = time.time()
        session.add(workbook)

        specs = [
            ("attendance", "考勤表", 5, attendance_document),
            ("registration", "报名表", 10, registration_document),
            (VIDEO_CONFIG_SHEET_KEY, "视频配置", 30, storage_documents[VIDEO_CONFIG_SHEET_KEY]),
            (VIDEO_DATA_SHEET_KEY, "视频数据", 40, storage_documents[VIDEO_DATA_SHEET_KEY]),
            (CLOCKIN_CONFIG_SHEET_KEY, "打卡配置", 50, storage_documents[CLOCKIN_CONFIG_SHEET_KEY]),
            (CLOCKIN_DATA_SHEET_KEY, "打卡数据", 60, storage_documents[CLOCKIN_DATA_SHEET_KEY]),
        ]
        sheets = []
        for sheet_key, title, order_index, document in specs:
            sheet = _upsert_sheet(session, workbook, sheet_key, title, document)
            _ensure_link(session, workbook, sheet, order_index)
            if sheet_key == "attendance":
                ensure_attendance_sheet_anonymous_viewer(session, sheet)
            sheets.append(sheet)

        summary_linked = _link_summary_row(session, workbook, sheets[0])
        if apply:
            session.commit()
        else:
            session.rollback()

        return {
            "mode": "APPLY" if apply else "DRY-RUN",
            "workbook_id": int(workbook.numeric_id or 0),
            "attendance_sheet_id": int(sheets[0].numeric_id or 0),
            "summary_linked": summary_linked,
            "sheets": [
                {
                    "sheet_key": sheet.sheet_key,
                    "title": sheet.title,
                    "id": int(sheet.numeric_id or 0),
                    "rows": len((sheet.document_json or {}).get("rows") or []),
                }
                for sheet in sheets
            ],
        }


def main() -> None:
    parser = argparse.ArgumentParser(description="重建 20260609 梵呗增益为 CodeYun 课程工作簿。")
    parser.add_argument("--apply", action="store_true", help="实际写入数据库；默认 dry-run。")
    parser.add_argument("--template", type=Path, default=DEFAULT_TEMPLATE_PATH)
    parser.add_argument("--owner-user-id", type=int, default=2)
    args = parser.parse_args()
    print(run(apply=args.apply, template_path=args.template, owner_user_id=args.owner_user_id))


if __name__ == "__main__":
    main()
