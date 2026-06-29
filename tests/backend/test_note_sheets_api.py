from __future__ import annotations

import io
import json
import time
from datetime import date

from openpyxl import Workbook, load_workbook
from sqlmodel import Session, select

from backend.api import note_sheets as note_sheets_api
from backend.app import app
from backend.core.access.auth import get_current_user_from_token, get_optional_current_user_from_token
from backend.core.access.feature_access import FEATURE_ACCESS_SUBJECT_USER, save_feature_access_policy_overrides
from backend.migrations.manager import (
    v29_migrate_attendance_course_sheets_to_notes_workbook,
    v30_add_numeric_sheet_and_workbook_ids,
)
from backend.core.resources.sheet_identity import allocate_new_sheet_identity
from backend.models import AttendanceWjxDataEntry, ResourceAccessGrant, SheetDocument, User, WorkbookDocument, WorkbookSheetLink


def _override_user(user: User) -> None:
    app.dependency_overrides[get_current_user_from_token] = lambda: user
    app.dependency_overrides[get_optional_current_user_from_token] = lambda: user


def _clear_user_override() -> None:
    app.dependency_overrides.pop(get_current_user_from_token, None)
    app.dependency_overrides.pop(get_optional_current_user_from_token, None)


def _create_user(session: Session, *, username: str, nickname: str = "", is_superuser: bool = False) -> User:
    user = User(
        username=username,
        nickname=nickname,
        email=f"{username}@example.com",
        hashed_password="pw",
        is_active=True,
        is_superuser=is_superuser,
    )
    session.add(user)
    session.commit()
    session.refresh(user)
    return user


def _grant_feature_access(session: Session, *, user_id: int, feature_key: str) -> None:
    save_feature_access_policy_overrides(
        session,
        subject_type=FEATURE_ACCESS_SUBJECT_USER,
        subject_user_id=user_id,
        overrides={feature_key: "allow"},
    )


def _assert_pagination_contains(actual: dict, expected: dict) -> None:
    for key, value in expected.items():
        assert actual[key] == value


def test_normalize_attendance_refund_note_removes_duplicate_period_label() -> None:
    assert note_sheets_api._normalize_attendance_current_refund_note(
        "第3天返款\n最近运行更新时间：\n2026/06/03 07:18:34,6"
    ) == "2026/06/03 07:18:34,6"
    assert note_sheets_api._normalize_attendance_current_refund_note(
        "第6天返款最近运行更新时间： 2026/06/06 07:40:41,6"
    ) == "2026/06/06 07:40:41,6"
    assert note_sheets_api._normalize_attendance_current_refund_note(
        "网课第6天返款 最近运行更新时间：\n2026/06/06 07:40:41,6"
    ) == "2026/06/06 07:40:41,6"
    assert note_sheets_api._normalize_attendance_current_refund_note(
        "念住闯关每日返款\n最近运行更新时间：\n2026/06/03 07:11:38,6"
    ) == "2026/06/03 07:11:38,6"
    assert note_sheets_api._normalize_attendance_current_refund_note(
        "最近运行更新时间：\n待首次同步"
    ) == "待首次同步"


def test_attendance_formula_normalizer_does_not_touch_non_dual_clockin_config_row() -> None:
    columns = ["视频应返款", "打卡应返款", "总应返款", "已返款", "订单金额", "当前应返款", "打卡数"]
    document = {
        "schema_version": 1,
        "columns": columns,
        "rows": [[0, 0, 0, 620, 0, 0, ""]],
        "grid_rows": [
            [""] * len(columns),
            columns,
            ["", "", "", '="第"&返款周期&"天"', 620, "2026/06/04 11:00:46,6", ""],
            [0, 0, 0, 620, 0, 0, ""],
        ],
        "data_start_row": 3,
        "field_row_index": 1,
    }

    normalized, changed_count = note_sheets_api._normalize_attendance_dual_clockin_refund_formulas(document)

    assert changed_count == 0
    assert normalized["grid_rows"][2][columns.index("已返款")] == '="第"&返款周期&"天"'
    assert normalized["grid_rows"][2][columns.index("订单金额")] == 620


def test_attendance_formula_normalizer_removes_refund_label_from_config_row() -> None:
    columns = ["共学打卡", "共修打卡", "打卡应返款", "总应返款", "已返款", "订单金额", "当前应返款"]
    document = {
        "schema_version": 1,
        "columns": columns,
        "rows": [[1, 1, "=IF(AND(A4>0,B4>0),30,0)", "=C4", 110, 499, 0]],
        "grid_rows": [
            [""] * len(columns),
            columns,
            ["", "", "", "", '="第"&返款周期&"天"', 499, "第6天返款最近运行更新时间： 2026/06/06 07:40:41,6"],
            [1, 1, "=IF(AND(A4>0,B4>0),30,0)", "=C4", 110, 499, 0],
        ],
        "data_start_row": 3,
        "field_row_index": 1,
    }

    normalized, changed_count = note_sheets_api._normalize_attendance_dual_clockin_refund_formulas(document)

    assert changed_count >= 1
    assert normalized["grid_rows"][2][columns.index("已返款")] == '="第"&返款周期&"周"'
    assert normalized["grid_rows"][2][columns.index("当前应返款")] == "2026/06/06 07:40:41,6"


def test_attendance_formula_normalizer_removes_refund_label_without_dual_clockin_columns() -> None:
    columns = ["完成视频数", "视频应返款", "打卡应返款", "总应返款", "已返款", "订单金额", "当前应返款"]
    document = {
        "schema_version": 1,
        "columns": columns,
        "rows": [[5, 80, 30, 110, 110, 499, 0]],
        "grid_rows": [
            [""] * len(columns),
            columns,
            ["", "", "", "", "第6天", 499, "第6天返款最近运行更新时间： 2026/06/06 07:40:41,6"],
            [5, 80, 30, 110, 110, 499, 0],
        ],
        "data_start_row": 3,
        "field_row_index": 1,
    }

    normalized, changed_count = note_sheets_api._normalize_attendance_dual_clockin_refund_formulas(document)

    assert changed_count == 1
    assert normalized["grid_rows"][2][columns.index("当前应返款")] == "2026/06/06 07:40:41,6"


def test_note_sheet_access_user_options_searches_username_and_nickname(client, session):
    manager = _create_user(session, username="sheet-manager", nickname="表格负责人")
    alice = _create_user(session, username="alice", nickname="陈坤泽")
    bob = _create_user(session, username="bob", nickname="momo酱")
    inactive_user = _create_user(session, username="hidden-user", nickname="陈隐藏")
    inactive_user.is_active = False
    session.add(inactive_user)
    session.commit()
    _grant_feature_access(session, user_id=manager.id, feature_key="notes.sheets")
    _override_user(manager)

    try:
        response = client.get("/api/note-sheets/access-users", params={"q": "陈"})
        assert response.status_code == 200
        users = response.json()["users"]
        assert users == [{"id": alice.id, "username": "alice", "nickname": "陈坤泽"}]

        username_response = client.get("/api/note-sheets/access-users", params={"q": "bob"})
        assert username_response.status_code == 200
        assert username_response.json()["users"] == [{"id": bob.id, "username": "bob", "nickname": "momo酱"}]
    finally:
        _clear_user_override()


def test_note_sheet_workbook_mvp_flow(client, session):
    user = _create_user(session, username="note-sheet-user")
    _grant_feature_access(session, user_id=user.id, feature_key="notes.sheets")
    _override_user(user)

    try:
        create_workbook_response = client.post(
            "/api/note-sheets/workbooks",
            json={"title": "禅宗工作簿"},
        )
        assert create_workbook_response.status_code == 200
        workbook = create_workbook_response.json()
        assert isinstance(workbook["id"], int) and workbook["id"] > 0
        assert workbook["title"] == "禅宗工作簿"
        assert workbook["sheet_count"] == 0

        create_sheet_response = client.post(
            "/api/note-sheets/sheets",
            json={
                "title": "报名表",
                "workbook_id": workbook["id"],
                "document_json": {
                    "schema_version": 1,
                    "columns": ["姓名", "手机号"],
                    "rows": [["时秋菊", "15641363033"]],
                },
            },
        )
        assert create_sheet_response.status_code == 200
        sheet = create_sheet_response.json()
        assert isinstance(sheet["id"], int) and sheet["id"] > 0
        assert sheet["scope"] == "notes"
        assert sheet["title"] == "报名表"
        assert sheet["owner_user_id"] == user.id
        assert sheet["document_json"]["rows"] == [["时秋菊", "15641363033"]]
        assert sheet["workbook_items"] == [{"id": workbook["id"], "title": "禅宗工作簿"}]

        list_workbooks_response = client.get("/api/note-sheets/workbooks")
        assert list_workbooks_response.status_code == 200
        assert list_workbooks_response.json()[0]["sheet_count"] == 1

        list_sheets_response = client.get("/api/note-sheets/sheets")
        assert list_sheets_response.status_code == 200
        assert list_sheets_response.json()[0]["workbook_items"] == [{"id": workbook["id"], "title": "禅宗工作簿"}]
        assert list_sheets_response.json()[0]["version"] == sheet["version"]

        detail_workbook_response = client.get(f"/api/note-sheets/workbooks/{workbook['id']}")
        assert detail_workbook_response.status_code == 200
        assert detail_workbook_response.json()["sheets"][0]["id"] == sheet["id"]
        assert detail_workbook_response.json()["sheets"][0]["version"] == sheet["version"]

        rename_workbook_response = client.put(
            f"/api/note-sheets/workbooks/{workbook['id']}",
            json={"title": "禅宗工作簿-更新"},
        )
        assert rename_workbook_response.status_code == 200
        renamed_workbook = rename_workbook_response.json()
        assert renamed_workbook["title"] == "禅宗工作簿-更新"
        assert renamed_workbook["sheets"][0]["workbook_items"] == [
            {"id": workbook["id"], "title": "禅宗工作簿-更新"},
        ]

        list_sheets_after_rename_response = client.get("/api/note-sheets/sheets")
        assert list_sheets_after_rename_response.status_code == 200
        assert list_sheets_after_rename_response.json()[0]["workbook_items"] == [
            {"id": workbook["id"], "title": "禅宗工作簿-更新"},
        ]

        update_sheet_response = client.put(
            f"/api/note-sheets/sheets/{sheet['id']}",
            json={
                "title": "报名表-更新",
                "base_version": sheet["version"],
                "document_json": {
                    "schema_version": 1,
                    "columns": ["姓名", "手机号", "组号"],
                    "rows": [["时秋菊", "15641363033", "1"]],
                },
            },
        )
        assert update_sheet_response.status_code == 200
        updated_sheet = update_sheet_response.json()
        assert updated_sheet["title"] == "报名表-更新"
        assert updated_sheet["version"] == 2
        assert updated_sheet["document_json"]["columns"] == ["姓名", "手机号", "组号"]

        stale_rename_sheet_response = client.put(
            f"/api/note-sheets/sheets/{sheet['id']}",
            json={
                "title": "报名表-过期重命名",
                "base_version": sheet["version"],
            },
        )
        assert stale_rename_sheet_response.status_code == 409
        latest_sheet_response = client.get(f"/api/note-sheets/sheets/{sheet['id']}")
        assert latest_sheet_response.status_code == 200
        assert latest_sheet_response.json()["title"] == "报名表-更新"
        assert latest_sheet_response.json()["version"] == 2

        workbook_record = session.exec(
            select(WorkbookDocument).where(WorkbookDocument.numeric_id == workbook["id"])
        ).first()
        sheet_record = session.exec(
            select(SheetDocument).where(SheetDocument.numeric_id == sheet["id"])
        ).first()
        assert workbook_record is not None
        assert sheet_record is not None
        session.add(ResourceAccessGrant(
            resource_type="workbook",
            resource_id=str(workbook_record.numeric_id),
            subject_key="anonymous",
            subject_type="anonymous",
            role="viewer",
        ))
        session.add(ResourceAccessGrant(
            resource_type="sheet",
            resource_id=str(sheet_record.numeric_id),
            subject_key="anonymous",
            subject_type="anonymous",
            role="viewer",
        ))
        session.commit()

        delete_workbook_response = client.delete(f"/api/note-sheets/workbooks/{workbook['id']}")
        assert delete_workbook_response.status_code == 200

        session.refresh(workbook_record)
        session.refresh(sheet_record)
        workbook_count = len(session.exec(select(WorkbookDocument).where(note_sheets_api._active_workbook_condition())).all())
        sheet_count = len(session.exec(
            select(SheetDocument)
            .where(SheetDocument.scope == "notes")
            .where(note_sheets_api._active_sheet_condition())
        ).all())
        link_count = len(session.exec(select(WorkbookSheetLink)).all())
        grant_count = len(session.exec(select(ResourceAccessGrant)).all())
        assert workbook_record.deleted_at and workbook_record.deleted_at > 0
        assert sheet_record.deleted_at and sheet_record.deleted_at > 0
        assert workbook_count == 0
        assert sheet_count == 0
        assert link_count == 1
        assert grant_count == 2
    finally:
        _clear_user_override()


def test_note_sheet_workbook_reorders_sheets(client, session):
    user = _create_user(session, username="note-sheet-reorder-user")
    _grant_feature_access(session, user_id=user.id, feature_key="notes.sheets")
    _override_user(user)

    try:
        workbook_response = client.post("/api/note-sheets/workbooks", json={"title": "可排序工作簿"})
        assert workbook_response.status_code == 200
        workbook_id = workbook_response.json()["id"]

        created_sheets = []
        for title in ["报名表", "考勤表", "速查表"]:
            response = client.post(
                "/api/note-sheets/sheets",
                json={
                    "title": title,
                    "workbook_id": workbook_id,
                    "document_json": {"columns": ["列1"], "rows": []},
                },
            )
            assert response.status_code == 200
            created_sheets.append(response.json())

        target_order = [
            created_sheets[2]["id"],
            created_sheets[0]["id"],
            created_sheets[1]["id"],
        ]
        reorder_response = client.post(
            f"/api/note-sheets/workbooks/{workbook_id}/sheets/reorder",
            json={"sheet_ids": target_order},
        )

        assert reorder_response.status_code == 200
        assert [sheet["id"] for sheet in reorder_response.json()["sheets"]] == target_order

        workbook_record = session.exec(
            select(WorkbookDocument).where(WorkbookDocument.numeric_id == workbook_id)
        ).one()
        links = session.exec(
            select(WorkbookSheetLink)
            .where(WorkbookSheetLink.workbook_id == str(workbook_record.numeric_id))
            .order_by(WorkbookSheetLink.order_index)
        ).all()
        sheets = session.exec(
            select(SheetDocument).where(SheetDocument.numeric_id.in_([int(link.sheet_id) for link in links]))
        ).all()
        numeric_id_by_sheet_id = {str(sheet.numeric_id): sheet.numeric_id for sheet in sheets}
        assert [numeric_id_by_sheet_id[link.sheet_id] for link in links] == target_order
        assert [link.order_index for link in links] == [10, 20, 30]

        duplicate_response = client.post(
            f"/api/note-sheets/workbooks/{workbook_id}/sheets/reorder",
            json={"sheet_ids": [created_sheets[0]["id"], created_sheets[0]["id"]]},
        )
        assert duplicate_response.status_code == 400
    finally:
        _clear_user_override()


def test_note_sheet_defined_names_check_sheet_versions(client, session):
    user = _create_user(session, username="note-sheet-defined-name-version-user")
    _grant_feature_access(session, user_id=user.id, feature_key="notes.sheets")
    _override_user(user)

    try:
        workbook_response = client.post("/api/note-sheets/workbooks", json={"title": "名称管理器工作簿"})
        assert workbook_response.status_code == 200
        workbook_id = workbook_response.json()["id"]
        sheet_response = client.post(
            "/api/note-sheets/sheets",
            json={
                "title": "名称管理器表",
                "workbook_id": workbook_id,
                "document_json": {
                    "schema_version": 1,
                    "columns": ["A"],
                    "rows": [["1"]],
                },
            },
        )
        assert sheet_response.status_code == 200
        sheet_id = sheet_response.json()["id"]
        initial_version = sheet_response.json()["version"]

        update_response = client.put(
            f"/api/note-sheets/sheets/{sheet_id}/defined-names",
            json={
                "base_version": initial_version,
                "names": [{"name": "LocalValue", "formula": "=A1"}],
            },
            params={"workbook_id": workbook_id},
        )
        assert update_response.status_code == 200, update_response.text
        assert update_response.json()["sheet_version"] == initial_version + 1

        stale_sheet_response = client.put(
            f"/api/note-sheets/sheets/{sheet_id}/defined-names",
            json={
                "base_version": initial_version,
                "names": [{"name": "StaleLocalValue", "formula": "=A1"}],
            },
            params={"workbook_id": workbook_id},
        )
        assert stale_sheet_response.status_code == 409

        stale_workbook_response = client.put(
            f"/api/note-sheets/workbooks/{workbook_id}/defined-names",
            json={
                "names": [{"name": "GlobalValue", "formula": "=1"}],
                "worksheets": [{
                    "sheet_id": sheet_id,
                    "sheet_title": "名称管理器表",
                    "sheet_version": initial_version,
                    "names": [{"name": "StaleScopedValue", "formula": "=A1"}],
                }],
            },
        )
        assert stale_workbook_response.status_code == 409

        persisted = session.exec(select(SheetDocument).where(SheetDocument.numeric_id == sheet_id)).one()
        assert persisted.document_json["defined_names"] == [{"name": "LocalValue", "formula": "=A1", "comment": ""}]
    finally:
        _clear_user_override()


def test_note_sheet_unpack_workbook_preserves_sheets(client, session, monkeypatch):
    user = _create_user(session, username="note-sheet-unpack-user")
    _grant_feature_access(session, user_id=user.id, feature_key="notes.sheets")
    _override_user(user)

    try:
        workbook_response = client.post("/api/note-sheets/workbooks", json={"title": "待解包工作簿"})
        assert workbook_response.status_code == 200
        workbook_id = workbook_response.json()["id"]

        sheet_response = client.post(
            "/api/note-sheets/sheets",
            json={"title": "解包后保留的表", "workbook_id": workbook_id},
        )
        assert sheet_response.status_code == 200
        sheet_id = sheet_response.json()["id"]
        sheet_version = sheet_response.json()["version"]

        workbook_record = session.exec(
            select(WorkbookDocument).where(WorkbookDocument.numeric_id == workbook_id)
        ).first()
        assert workbook_record is not None
        session.add(ResourceAccessGrant(
            resource_type="workbook",
            resource_id=str(workbook_record.numeric_id),
            subject_key="anonymous",
            subject_type="anonymous",
            role="viewer",
        ))
        session.commit()

        unpack_response = client.post(f"/api/note-sheets/workbooks/{workbook_id}/unpack")
        assert unpack_response.status_code == 200

        workbooks = session.exec(select(WorkbookDocument)).all()
        sheets = session.exec(select(SheetDocument).where(SheetDocument.scope == "notes")).all()
        links = session.exec(select(WorkbookSheetLink)).all()
        workbook_grants = session.exec(
            select(ResourceAccessGrant).where(ResourceAccessGrant.resource_type == "workbook")
        ).all()
        assert workbooks == []
        assert [item.numeric_id for item in sheets] == [sheet_id]
        assert links == []
        assert workbook_grants == []

        broadcasts: list[tuple[str, dict]] = []

        async def fake_broadcast(room: str, message: dict) -> None:
            broadcasts.append((room, message))

        monkeypatch.setattr(note_sheets_api.ws_manager, "broadcast", fake_broadcast)

        delete_sheet_response = client.delete(f"/api/note-sheets/sheets/{sheet_id}")
        assert delete_sheet_response.status_code == 200
        active_sheets = session.exec(
            select(SheetDocument)
            .where(SheetDocument.scope == "notes")
            .where(note_sheets_api._active_sheet_condition())
        ).all()
        deleted_sheet = session.exec(select(SheetDocument).where(SheetDocument.numeric_id == sheet_id)).one()
        assert active_sheets == []
        assert deleted_sheet.deleted_at and deleted_sheet.deleted_at > 0
        assert deleted_sheet.version == sheet_version + 1
        assert broadcasts[-1][0] == f"resource:sheet:{sheet_id}"
        assert broadcasts[-1][1]["version"] == deleted_sheet.version

        restore_sheet_response = client.post(f"/api/note-sheets/sheets/{sheet_id}/restore")
        assert restore_sheet_response.status_code == 200
        restored_sheet = session.exec(select(SheetDocument).where(SheetDocument.numeric_id == sheet_id)).one()
        assert not restored_sheet.deleted_at
        assert restored_sheet.version == sheet_version + 2
        assert broadcasts[-1][0] == f"resource:sheet:{sheet_id}"
        assert broadcasts[-1][1]["version"] == restored_sheet.version
    finally:
        _clear_user_override()


def test_note_sheet_attach_existing_sheet_to_workbook(client, session):
    user = _create_user(session, username="note-sheet-link-user")
    _grant_feature_access(session, user_id=user.id, feature_key="notes.sheets")
    _override_user(user)

    try:
        workbook_response = client.post("/api/note-sheets/workbooks", json={"title": "工作簿A"})
        assert workbook_response.status_code == 200
        workbook_id = workbook_response.json()["id"]

        first_sheet_response = client.post("/api/note-sheets/sheets", json={"title": "表格A"})
        assert first_sheet_response.status_code == 200
        sheet_id = first_sheet_response.json()["id"]

        attach_response = client.post(
            f"/api/note-sheets/workbooks/{workbook_id}/sheets",
            json={"sheet_id": sheet_id},
        )
        assert attach_response.status_code == 200
        detail = attach_response.json()
        assert [item["id"] for item in detail["sheets"]] == [sheet_id]
    finally:
        _clear_user_override()


def test_delete_workbook_preserves_sheets_linked_from_other_workbooks(client, session):
    user = _create_user(session, username="note-sheet-shared-delete-user")
    _grant_feature_access(session, user_id=user.id, feature_key="notes.sheets")
    _override_user(user)

    try:
        first_workbook = client.post("/api/note-sheets/workbooks", json={"title": "工作簿A"}).json()
        second_workbook = client.post("/api/note-sheets/workbooks", json={"title": "工作簿B"}).json()
        sheet = client.post(
            "/api/note-sheets/sheets",
            json={"title": "共享表", "workbook_id": first_workbook["id"]},
        ).json()
        attach_response = client.post(
            f"/api/note-sheets/workbooks/{second_workbook['id']}/sheets",
            json={"sheet_id": sheet["id"]},
        )
        assert attach_response.status_code == 200

        sheet_record = session.exec(
            select(SheetDocument).where(SheetDocument.numeric_id == sheet["id"])
        ).first()
        assert sheet_record is not None
        session.add(ResourceAccessGrant(
            resource_type="sheet",
            resource_id=str(sheet_record.numeric_id),
            subject_key="anonymous",
            subject_type="anonymous",
            role="viewer",
        ))
        session.commit()

        delete_first_response = client.delete(f"/api/note-sheets/workbooks/{first_workbook['id']}")
        assert delete_first_response.status_code == 200
        first_record = session.exec(select(WorkbookDocument).where(WorkbookDocument.numeric_id == first_workbook["id"])).one()
        session.refresh(sheet_record)
        assert first_record.deleted_at and first_record.deleted_at > 0
        assert not sheet_record.deleted_at
        assert len(session.exec(select(WorkbookSheetLink).where(WorkbookSheetLink.workbook_id == str(second_workbook["id"]))).all()) == 1
        assert len(session.exec(select(ResourceAccessGrant)).all()) == 1

        delete_second_response = client.delete(f"/api/note-sheets/workbooks/{second_workbook['id']}")
        assert delete_second_response.status_code == 200
        session.refresh(sheet_record)
        assert sheet_record.deleted_at and sheet_record.deleted_at > 0
        assert len(session.exec(select(ResourceAccessGrant)).all()) == 1
    finally:
        _clear_user_override()


def test_note_sheet_workbook_library_lists_owned_shared_and_superuser_items(client, session):
    owner = _create_user(session, username="workbook-library-owner")
    viewer = _create_user(session, username="workbook-library-viewer")
    outsider = _create_user(session, username="workbook-library-outsider")
    superuser = _create_user(session, username="workbook-library-superuser", is_superuser=True)
    for user in [viewer, superuser]:
        _grant_feature_access(session, user_id=user.id, feature_key="notes.sheets")

    owned_by_viewer = WorkbookDocument(
        numeric_id=11,
        title="查看者自己的工作簿",
        owner_user_id=viewer.id,
        created_by_user_id=viewer.id,
        updated_by_user_id=viewer.id,
        created_at=10,
        updated_at=10,
    )
    shared_to_viewer = WorkbookDocument(
        numeric_id=12,
        title="共享给查看者的工作簿",
        owner_user_id=owner.id,
        created_by_user_id=owner.id,
        updated_by_user_id=owner.id,
        created_at=20,
        updated_at=20,
    )
    outsider_only = WorkbookDocument(
        numeric_id=13,
        title="其他人的工作簿",
        owner_user_id=outsider.id,
        created_by_user_id=outsider.id,
        updated_by_user_id=outsider.id,
        created_at=30,
        updated_at=30,
    )
    session.add(owned_by_viewer)
    session.add(shared_to_viewer)
    session.add(outsider_only)
    session.commit()
    session.refresh(shared_to_viewer)

    session.add(ResourceAccessGrant(
        resource_type="workbook",
        resource_id=str(shared_to_viewer.numeric_id),
        subject_key=f"user:{viewer.id}",
        subject_type="user",
        subject_user_id=viewer.id,
        role="viewer",
    ))
    session.commit()

    _override_user(viewer)
    try:
        viewer_response = client.get("/api/note-sheets/workbooks")
        assert viewer_response.status_code == 200
        assert [item["id"] for item in viewer_response.json()] == [12, 11]
    finally:
        _clear_user_override()

    _override_user(superuser)
    try:
        superuser_response = client.get("/api/note-sheets/workbooks")
        assert superuser_response.status_code == 200
        assert {item["id"] for item in superuser_response.json()} == {11, 12, 13}
    finally:
        _clear_user_override()


def test_note_sheet_blank_create_uses_sheet_address_defaults(client, session):
    user = _create_user(session, username="note-sheet-default-address-user")
    _grant_feature_access(session, user_id=user.id, feature_key="notes.sheets")
    _override_user(user)

    try:
        response = client.post("/api/note-sheets/sheets", json={"title": "默认地址表"})
        assert response.status_code == 200
        document = response.json()["document_json"]
        assert document["formula_reference_origin"] == "sheet_v2"
        assert document["data_start_row"] == 1
        assert document["field_row_index"] == 0
        assert document["grid_rows"] == [["列1", "列2", "列3"]]
        assert document["view_settings"]["row_marker_numbering"] == "global"
        assert document["view_settings"]["row_marker_origin"] == "sheet"
        assert document["view_settings"]["column_marker_style"] == "letters"
    finally:
        _clear_user_override()


def test_note_sheet_excel_import_reset_preserves_action_row(client, session, monkeypatch):
    monkeypatch.delenv("CODEYUN_NOTE_SHEET_EXCEL_IMPORT_MODEL", raising=False)
    user = _create_user(session, username="note-sheet-excel-import-user")
    _grant_feature_access(session, user_id=user.id, feature_key="notes.sheets")
    _override_user(user)

    target_columns = [
        "分组",
        "序号",
        "备注",
        "提交时间",
        "姓名",
        "微信昵称",
        "手机号",
        "错误手机号",
        "微信支付订单号",
        "订单日期",
        "商户订单号",
        "订单金额",
        "已退款",
        "用户ID",
        "匹配得分",
        "参考信息",
    ]
    action_row = ["", "", "导入报名", "", "", "", "数字加前缀", "", "更新订单匹配", "", "", "", "仅参考，有延迟", "", "", "其他备注"]

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "报名源表"
    worksheet.append(["序号", "促学金选择", "交易单号", "姓名", "微信昵称", "手机", "微信号"])
    worksheet.append(["", "1组", "", "", "", "", ""])
    worksheet.append([1, "自觉自律完成学修（无需支付促学金）", "", "阿丹", "阿丹", "15326693765", "adan-wx"])
    buffer = io.BytesIO()
    workbook.save(buffer)

    def fake_chat_with_provider(**kwargs):
        prompt = kwargs["messages"][0]["content"]
        assert "只导入报名学员" in prompt
        assert "excel_import_reset" in prompt
        assert "导入报名" in prompt
        assert "不要选择明显是系统生成的数字字母代码" in kwargs["system_prompt"]
        assert "真实昵称" in kwargs["system_prompt"]
        assert "不要把无法匹配的普通源字段塞进“备注”" in kwargs["system_prompt"]
        assert kwargs["provider_id"] == "deepseek"
        assert kwargs["model"] == "deepseek-v4-pro"
        assert kwargs["response_format"] == "json"
        return {
            "content": json.dumps(
                {
                    "extra_columns": ["促学金模式", "微信号"],
                    "rows": [
                        {
                            "分组": "1组",
                            "序号": "1",
                            "促学金模式": "自觉自律完成学修（无需支付促学金）",
                            "姓名": "阿丹",
                            "微信昵称": "阿丹",
                            "手机号": "15326693765",
                            "微信号": "adan-wx",
                        },
                    ],
                    "warnings": [],
                    "mapping_notes": ["按分组标题填充分组；未匹配字段追加为扩展列"],
                },
                ensure_ascii=False,
            ),
        }

    monkeypatch.setattr(note_sheets_api, "chat_with_provider", fake_chat_with_provider)

    try:
        workbook_response = client.post("/api/note-sheets/workbooks", json={"title": "报名工作簿"})
        workbook_id = workbook_response.json()["id"]
        sheet_response = client.post(
            "/api/note-sheets/sheets",
            json={
                "title": "报名表",
                "workbook_id": workbook_id,
                "document_json": {
                    "schema_version": 1,
                    "columns": target_columns,
                    "rows": [
                        ["旧组", "99", "旧数据", "", "旧姓名", "", "", "", "", "", "", "", "", "", "", ""],
                    ],
                    "grid_rows": [
                        target_columns,
                        action_row,
                        ["旧组", "99", "旧数据", "", "旧姓名", "", "", "", "", "", "", "", "", "", "", ""],
                    ],
                    "data_start_row": 2,
                    "field_row_index": 0,
                    "cell_meta": {
                        "1:2": {
                            "action": {"type": "excel_import_reset", "label": "导入excel"},
                            "style": {"background_color": "#eeeeee"},
                        },
                        "2:4": {"style": {"background_color": "#ff0000"}},
                    },
                    "entity_columns": [
                        {"id": f"col_{index}", "header": header}
                        for index, header in enumerate(target_columns)
                    ],
                    "entity_rows": [
                        {"id": "field-row", "kind": "field"},
                        {"id": "action-row", "kind": "field_note"},
                        {"id": "old-data-row", "kind": "data"},
                    ],
                    "entity_cells": {
                        "action-row": {
                            "col_2": {
                                "value": "导入报名",
                                "action": {"type": "excel_import_reset", "label": "导入excel"},
                            },
                        },
                        "old-data-row": {
                            "col_4": {
                                "value": "旧姓名",
                                "style": {"background_color": "#ff0000"},
                            },
                        },
                    },
                },
            },
        )
        sheet_id = sheet_response.json()["id"]
        initial_version = sheet_response.json()["version"]

        stale_response = client.post(
            f"/api/note-sheets/sheets/{sheet_id}/import-excel-reset",
            params={"workbook_id": workbook_id},
            data={"instruction": "旧版本导入", "base_version": str(initial_version + 1)},
            files={
                "file": (
                    "source.xlsx",
                    buffer.getvalue(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                ),
            },
        )
        assert stale_response.status_code == 409

        response = client.post(
            f"/api/note-sheets/sheets/{sheet_id}/import-excel-reset",
            params={"workbook_id": workbook_id},
            data={
                "instruction": "只导入报名学员",
                "action_document_row": "1",
                "action_column": "2",
                "base_version": str(initial_version),
            },
            files={
                "file": (
                    "source.xlsx",
                    buffer.getvalue(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                ),
            },
        )

        assert response.status_code == 200
        payload = response.json()
        assert payload["imported_count"] == 1
        assert payload["preserved_row_count"] == 0
        assert payload["extra_columns"] == ["促学金模式", "微信号"]
        document = payload["sheet"]["document_json"]
        assert document["columns"] == [*target_columns, "促学金模式", "微信号"]
        assert document["grid_rows"][0] == [*target_columns, "促学金模式", "微信号"]
        assert document["grid_rows"][1] == [*action_row, "", ""]
        assert document["rows"][0][target_columns.index("姓名")] == "阿丹"
        assert document["rows"][0][target_columns.index("手机号")] == "15326693765"
        assert document["rows"][0][target_columns.index("备注")] == ""
        assert document["rows"][0][target_columns.index("参考信息")] == ""
        assert document["rows"][0][len(target_columns)] == "自觉自律完成学修（无需支付促学金）"
        assert document["rows"][0][len(target_columns) + 1] == "adan-wx"
        assert document["column_configs"]["促学金模式"]["header_text_color"] == "#4B5563"
        assert document["column_configs"]["促学金模式"]["header_background_color"] == "#E5E7EB"
        assert document["column_widths"][-2:] == [132, 132]
        assert "1:2" in document["cell_meta"]
        assert document["cell_meta"]["1:2"]["action"]["type"] == "excel_import_reset"
        assert "2:4" not in document["cell_meta"]
        assert document["entity_rows"] == [
            {"id": "field-row", "kind": "field"},
            {"id": "action-row", "kind": "field_note"},
        ]
        assert set(document["entity_cells"]) == {"action-row"}
        assert "old-data-row" not in document["entity_cells"]
    finally:
        _clear_user_override()


def test_note_sheet_excel_import_append_keeps_existing_rows(client, session, monkeypatch):
    user = _create_user(session, username="note-sheet-excel-import-append-user")
    _grant_feature_access(session, user_id=user.id, feature_key="notes.sheets")
    _override_user(user)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "增量源表"
    worksheet.append(["姓名", "手机", "来源备注"])
    worksheet.append(["新姓名", "13900000000", "新报名"])
    buffer = io.BytesIO()
    workbook.save(buffer)

    def fake_chat_with_provider(**kwargs):
        prompt = kwargs["messages"][0]["content"]
        assert "增量导入" in prompt
        return {
            "content": json.dumps(
                {
                    "extra_columns": ["来源备注"],
                    "rows": [
                        {
                            "姓名": "新姓名",
                            "手机号": "13900000000",
                            "来源备注": "新报名",
                        },
                    ],
                    "warnings": [],
                    "mapping_notes": [],
                },
                ensure_ascii=False,
            ),
        }

    monkeypatch.setattr(note_sheets_api, "chat_with_provider", fake_chat_with_provider)

    try:
        workbook_response = client.post("/api/note-sheets/workbooks", json={"title": "增量导入工作簿"})
        workbook_id = workbook_response.json()["id"]
        sheet_response = client.post(
            "/api/note-sheets/sheets",
            json={
                "title": "增量报名表",
                "workbook_id": workbook_id,
                "document_json": {
                    "schema_version": 1,
                    "columns": ["姓名", "手机号"],
                    "rows": [["旧姓名", "13800000000"]],
                    "grid_rows": [["姓名", "手机号"], ["旧姓名", "13800000000"]],
                    "data_start_row": 1,
                    "field_row_index": 0,
                    "cell_meta": {
                        "1:0": {"style": {"background_color": "#ff0000"}},
                    },
                },
            },
        )
        sheet_id = sheet_response.json()["id"]

        response = client.post(
            f"/api/note-sheets/sheets/{sheet_id}/import-excel-reset",
            params={"workbook_id": workbook_id},
            data={"instruction": "增量导入", "mode": "append"},
            files={
                "file": (
                    "append.xlsx",
                    buffer.getvalue(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                ),
            },
        )

        assert response.status_code == 200
        payload = response.json()
        assert payload["imported_count"] == 1
        assert payload["preserved_row_count"] == 1
        assert payload["extra_columns"] == ["来源备注"]
        document = payload["sheet"]["document_json"]
        assert document["columns"] == ["姓名", "手机号", "来源备注"]
        assert document["rows"] == [
            ["旧姓名", "13800000000", ""],
            ["新姓名", "13900000000", "新报名"],
        ]
        assert document["grid_rows"] == [
            ["姓名", "手机号", "来源备注"],
            ["旧姓名", "13800000000", ""],
            ["新姓名", "13900000000", "新报名"],
        ]
        assert document["cell_meta"]["1:0"]["style"]["background_color"] == "#ff0000"
    finally:
        _clear_user_override()


def test_note_sheet_excel_import_retries_ai_before_success(client, session, monkeypatch):
    monkeypatch.setenv("CODEYUN_NOTE_SHEET_EXCEL_IMPORT_MAX_ATTEMPTS", "2")
    user = _create_user(session, username="note-sheet-excel-import-retry-user")
    _grant_feature_access(session, user_id=user.id, feature_key="notes.sheets")
    _override_user(user)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["姓名", "手机"])
    worksheet.append(["阿丹", "15326693765"])
    buffer = io.BytesIO()
    workbook.save(buffer)

    calls: list[str] = []

    def fake_chat_with_provider(**kwargs):
        calls.append(kwargs["messages"][0]["content"])
        if len(calls) == 1:
            return {"content": "not json"}
        assert "上一次 AI 导入失败" in kwargs["messages"][0]["content"]
        return {
            "content": json.dumps({
                "rows": [{"姓名": "阿丹", "手机号": "15326693765"}],
                "warnings": [],
                "mapping_notes": ["重试后返回合法 JSON"],
            }, ensure_ascii=False),
        }

    monkeypatch.setattr(note_sheets_api, "chat_with_provider", fake_chat_with_provider)

    try:
        workbook_response = client.post("/api/note-sheets/workbooks", json={"title": "AI 重试工作簿"})
        workbook_id = workbook_response.json()["id"]
        sheet_response = client.post(
            "/api/note-sheets/sheets",
            json={
                "title": "报名表",
                "workbook_id": workbook_id,
                "document_json": {
                    "schema_version": 1,
                    "columns": ["姓名", "手机号"],
                    "rows": [],
                },
            },
        )
        sheet_id = sheet_response.json()["id"]

        response = client.post(
            f"/api/note-sheets/sheets/{sheet_id}/import-excel-reset",
            params={"workbook_id": workbook_id},
            data={"instruction": "只导入报名学员"},
            files={
                "file": (
                    "retry.xlsx",
                    buffer.getvalue(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                ),
            },
        )

        assert response.status_code == 200
        assert len(calls) == 2
        payload = response.json()
        assert payload["imported_count"] == 1
        assert payload["mapping_notes"] == ["重试后返回合法 JSON"]
        assert payload["sheet"]["document_json"]["rows"] == [["阿丹", "15326693765"]]
    finally:
        _clear_user_override()


def test_note_sheet_excel_import_failure_reports_last_ai_error(client, session, monkeypatch):
    monkeypatch.setenv("CODEYUN_NOTE_SHEET_EXCEL_IMPORT_MAX_ATTEMPTS", "2")
    user = _create_user(session, username="note-sheet-excel-import-error-detail-user")
    _grant_feature_access(session, user_id=user.id, feature_key="notes.sheets")
    _override_user(user)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["姓名", "手机"])
    worksheet.append(["阿丹", "15326693765"])
    buffer = io.BytesIO()
    workbook.save(buffer)

    def fake_chat_with_provider(**_kwargs):
        return {"content": "not json"}

    monkeypatch.setattr(note_sheets_api, "chat_with_provider", fake_chat_with_provider)

    try:
        workbook_response = client.post("/api/note-sheets/workbooks", json={"title": "AI 错误提示工作簿"})
        workbook_id = workbook_response.json()["id"]
        sheet_response = client.post(
            "/api/note-sheets/sheets",
            json={
                "title": "报名表",
                "workbook_id": workbook_id,
                "document_json": {
                    "schema_version": 1,
                    "columns": ["姓名", "手机号"],
                    "rows": [],
                },
            },
        )
        sheet_id = sheet_response.json()["id"]

        response = client.post(
            f"/api/note-sheets/sheets/{sheet_id}/import-excel-reset",
            params={"workbook_id": workbook_id},
            data={"instruction": "只导入报名学员"},
            files={
                "file": (
                    "error-detail.xlsx",
                    buffer.getvalue(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                ),
            },
        )

        assert response.status_code == 502
        detail = response.json()["detail"]
        assert "AI 导入未完成：已尝试 2 次" in detail
        assert "最后错误发生在第 2 次" in detail
        assert "deepseek/deepseek-v4-pro" in detail
        assert "AI 返回解析错误" in detail
        assert "DeepSeek 未返回 JSON 对象" in detail
    finally:
        _clear_user_override()


def test_note_sheet_stale_save_does_not_overwrite_excel_import(client, session, monkeypatch):
    user = _create_user(session, username="note-sheet-stale-save-import-user")
    _grant_feature_access(session, user_id=user.id, feature_key="notes.sheets")
    _override_user(user)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["姓名", "手机号"])
    worksheet.append(["新姓名", "13900000000"])
    buffer = io.BytesIO()
    workbook.save(buffer)

    def fake_chat_with_provider(**kwargs):
        return {
            "content": json.dumps(
                {
                    "rows": [
                        {
                            "姓名": "新姓名",
                            "手机号": "13900000000",
                        },
                    ],
                    "warnings": [],
                    "mapping_notes": [],
                },
                ensure_ascii=False,
            ),
        }

    monkeypatch.setattr(note_sheets_api, "chat_with_provider", fake_chat_with_provider)

    try:
        create_response = client.post(
            "/api/note-sheets/sheets",
            json={
                "title": "导入竞态",
                "document_json": {
                    "schema_version": 1,
                    "columns": ["姓名", "手机号"],
                    "rows": [["旧姓名", "13800000000"]],
                    "grid_rows": [["姓名", "手机号"], ["旧姓名", "13800000000"]],
                    "data_start_row": 1,
                    "field_row_index": 0,
                },
            },
        )
        assert create_response.status_code == 200
        stale_payload = create_response.json()
        sheet_id = stale_payload["id"]
        stale_version = stale_payload["version"]
        stale_document = stale_payload["document_json"]

        import_response = client.post(
            f"/api/note-sheets/sheets/{sheet_id}/import-excel-reset",
            data={"instruction": "增量导入", "mode": "append"},
            files={
                "file": (
                    "append.xlsx",
                    buffer.getvalue(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                ),
            },
        )
        assert import_response.status_code == 200
        assert import_response.json()["sheet"]["document_json"]["rows"] == [
            ["旧姓名", "13800000000"],
            ["新姓名", "13900000000"],
        ]

        stale_save_response = client.put(
            f"/api/note-sheets/sheets/{sheet_id}",
            json={
                "base_version": stale_version,
                "document_json": stale_document,
            },
        )
        assert stale_save_response.status_code == 409

        final_response = client.get(f"/api/note-sheets/sheets/{sheet_id}", params={"paginate": False})
        assert final_response.status_code == 200
        assert final_response.json()["document_json"]["rows"] == [
            ["旧姓名", "13800000000"],
            ["新姓名", "13900000000"],
        ]
    finally:
        _clear_user_override()


def test_note_sheet_update_prunes_orphan_entity_cells(client, session):
    user = _create_user(session, username="note-sheet-orphan-entity-user")
    _grant_feature_access(session, user_id=user.id, feature_key="notes.sheets")
    _override_user(user)

    try:
        create_response = client.post(
            "/api/note-sheets/sheets",
            json={
                "title": "孤立实体",
                "document_json": {
                    "schema_version": 1,
                    "columns": ["姓名"],
                    "rows": [["旧姓名"]],
                    "grid_rows": [["姓名"], ["旧姓名"]],
                    "data_start_row": 1,
                    "field_row_index": 0,
                    "entity_columns": [{"id": "col_name", "header": "姓名"}],
                    "entity_rows": [
                        {"id": "field-row", "kind": "field"},
                        {"id": "row-old", "kind": "data"},
                    ],
                    "entity_cells": {
                        "field-row": {"col_name": {"value": "姓名"}},
                        "row-old": {"col_name": {"value": "旧姓名"}},
                        "row-orphan": {"col_name": {"value": "不应保留"}},
                    },
                },
            },
        )
        assert create_response.status_code == 200
        payload = create_response.json()
        sheet_id = payload["id"]

        save_response = client.put(
            f"/api/note-sheets/sheets/{sheet_id}",
            json={
                "base_version": payload["version"],
                "document_json": payload["document_json"],
            },
        )
        assert save_response.status_code == 200
        entity_cells = save_response.json()["document_json"]["entity_cells"]
        assert set(entity_cells) == {"field-row", "row-old"}
        assert "row-orphan" not in entity_cells
    finally:
        _clear_user_override()


def test_note_sheet_excel_import_append_inserts_registration_rows_before_archived_rows(client, session, monkeypatch):
    user = _create_user(session, username="note-sheet-excel-import-registration-append-user")
    _grant_feature_access(session, user_id=user.id, feature_key="notes.sheets")
    _override_user(user)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "新增报名"
    worksheet.append(["序号", "姓名", "微信昵称", "手机号", "提交时间"])
    worksheet.append([1, "新学员", "新昵称", "13900000000", "2026/5/14 09:15"])
    buffer = io.BytesIO()
    workbook.save(buffer)

    def fake_chat_with_provider(**kwargs):
        return {
            "content": json.dumps(
                {
                    "rows": [
                        {
                            "分组": "",
                            "序号": "1",
                            "提交时间": "2026/5/14 09:15",
                            "姓名": "新学员",
                            "微信昵称": "新昵称",
                            "手机号": "13900000000",
                        },
                    ],
                    "warnings": [],
                    "mapping_notes": [],
                },
                ensure_ascii=False,
            ),
        }

    monkeypatch.setattr(note_sheets_api, "chat_with_provider", fake_chat_with_provider)

    try:
        workbook_response = client.post("/api/note-sheets/workbooks", json={"title": "念住闯关"})
        workbook_id = workbook_response.json()["id"]
        columns = ["分组", "序号", "备注", "提交时间", "姓名", "微信昵称", "手机号"]
        sheet_response = client.post(
            "/api/note-sheets/sheets",
            json={
                "title": "报名表",
                "workbook_id": workbook_id,
                "document_json": {
                    "schema_version": 1,
                    "columns": columns,
                    "rows": [
                        ["5月4日", "123", "", "2026/5/14 09:15", "旧有效", "旧有效", "13800000000"],
                        ["5月4日", "115", "", "2026/3/11 11:41", "已冻结", "已冻结", "13700000000"],
                    ],
                    "grid_rows": [
                        columns,
                        ["", "", "导入excel", "", "", "", ""],
                        ["5月4日", "123", "", "2026/5/14 09:15", "旧有效", "旧有效", "13800000000"],
                        ["5月4日", "115", "", "2026/3/11 11:41", "已冻结", "已冻结", "13700000000"],
                    ],
                    "data_start_row": 2,
                    "field_row_index": 0,
                    "cell_meta": {
                        "1:2": {"action": {"type": "excel_import_reset", "label": "导入excel"}},
                        "3:0": {"style": {"background_color": "#f2f2f2", "text_color": "#6b7280"}},
                    },
                    "entity_columns": [
                        {"id": f"col_{index}", "header": header}
                        for index, header in enumerate(columns)
                    ],
                    "entity_rows": [
                        {"id": "field-row", "kind": "field"},
                        {"id": "action-row", "kind": "field_note"},
                        {"id": "active-row", "kind": "data"},
                        {"id": "archived-row", "kind": "data"},
                    ],
                    "entity_cells": {
                        "action-row": {
                            "col_2": {
                                "value": "导入excel",
                                "action": {"type": "excel_import_reset", "label": "导入excel"},
                            },
                        },
                        "archived-row": {
                            "col_0": {
                                "value": "5月4日",
                                "style": {"background_color": "#f2f2f2", "text_color": "#6b7280"},
                            },
                        },
                    },
                },
            },
        )
        sheet_id = sheet_response.json()["id"]

        response = client.post(
            f"/api/note-sheets/sheets/{sheet_id}/import-excel-reset",
            params={"workbook_id": workbook_id},
            data={"instruction": "新增报名", "mode": "append"},
            files={
                "file": (
                    "append-registration.xlsx",
                    buffer.getvalue(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                ),
            },
        )

        assert response.status_code == 200
        payload = response.json()
        assert payload["preserved_row_count"] == 1
        document = payload["sheet"]["document_json"]
        assert document["rows"] == [
            ["5月4日", "123", "", "2026/5/14 09:15", "旧有效", "旧有效", "13800000000"],
            ["5月4日", "124", "", "2026/5/14 09:15", "新学员", "新昵称", "13900000000"],
            ["5月4日", "115", "", "2026/3/11 11:41", "已冻结", "已冻结", "13700000000"],
        ]
        assert "3:0" not in document["cell_meta"]
        assert document["cell_meta"]["4:0"]["style"] == {"background_color": "#F2F2F2", "text_color": "#6B7280"}
        assert document["entity_rows"] == [
            {"id": "field-row", "kind": "field"},
            {"id": "action-row", "kind": "field_note"},
        ]
        assert set(document["entity_cells"]) == {"action-row"}
    finally:
        _clear_user_override()


def test_note_sheet_registration_dynamic_expiration_orders_rows_and_ignores_stale_status():
    columns = ["分组", "序号", "提交时间", "姓名", "追踪状态"]
    document = {
        "schema_version": 1,
        "columns": columns,
        "data_start_row": 1,
        "field_row_index": 0,
        "rows": [
            ["5月4日", "123", "2026/5/14 09:15", "有效较晚但状态过期", "已冻结"],
            ["5月4日", "120", "2026/4/6 15:14", "有效较早", ""],
            ["第3批", "20", "2026/2/1 17:20", "过期较早", ""],
            ["第3批", "21", "2026年3月11日 11:41", "过期较晚", ""],
        ],
        "grid_rows": [
            columns,
            ["5月4日", "123", "2026/5/14 09:15", "有效较晚但状态过期", "已冻结"],
            ["5月4日", "120", "2026/4/6 15:14", "有效较早", ""],
            ["第3批", "20", "2026/2/1 17:20", "过期较早", ""],
            ["第3批", "21", "2026/3/11 11:41", "过期较晚", ""],
        ],
        "cell_meta": {
            "1:3": {"style": {"background_color": "#ffffff"}},
            "2:0": {"style": {"background_color": "#F2F2F2", "text_color": "#6B7280"}},
            "4:3": {"style": {"background_color": "#f2f2f2"}},
        },
        "entity_rows": [
            {"id": "field-row", "kind": "field"},
            {"id": "active-late", "kind": "data"},
            {"id": "active-early", "kind": "data"},
            {"id": "expired-old", "kind": "data"},
            {"id": "expired-new", "kind": "data"},
        ],
        "entity_columns": [{"id": f"col_{index}", "header": header} for index, header in enumerate(columns)],
        "entity_cells": {
            "active-late": {"col_3": {"value": "有效较晚但状态过期"}},
            "active-early": {"col_3": {"value": "有效较早"}},
            "expired-old": {"col_3": {"value": "过期较早"}},
            "expired-new": {"col_3": {"value": "过期较晚"}},
        },
    }

    ordered = note_sheets_api._order_registration_rows_by_dynamic_expiration(
        document,
        now=date(2026, 5, 21),
    )

    assert [
        row[columns.index("姓名")]
        for row in ordered["rows"]
    ] == ["有效较早", "有效较晚但状态过期", "过期较晚", "过期较早"]
    assert [
        row[columns.index("追踪状态")]
        for row in ordered["rows"]
    ] == ["追踪中", "追踪中", "已冻结", "已冻结"]
    assert not note_sheets_api._is_archived_registration_row(
        document["rows"][0],
        columns,
        now=date(2026, 5, 21),
    )
    assert note_sheets_api._is_archived_registration_row(
        document["rows"][3],
        columns,
        now=date(2026, 5, 21),
    )
    assert "1:0" not in ordered["cell_meta"]
    assert ordered["cell_meta"]["2:3"]["style"]["background_color"] == "#ffffff"
    for document_row in (3, 4):
        for column_index in range(len(columns)):
            style = ordered["cell_meta"][f"{document_row}:{column_index}"]["style"]
            assert style["background_color"] == "#F2F2F2"
            assert style["text_color"] == "#6B7280"
    assert [row["id"] for row in ordered["entity_rows"][1:]] == [
        "active-early",
        "active-late",
        "expired-new",
        "expired-old",
    ]


def test_note_sheet_excel_import_prefers_grouped_sequence_from_source_workbook(client, session, monkeypatch):
    user = _create_user(session, username="note-sheet-excel-import-group-sequence-user")
    _grant_feature_access(session, user_id=user.id, feature_key="notes.sheets")
    _override_user(user)

    workbook = Workbook()
    grouped = workbook.active
    grouped.title = "考勤表"
    grouped.append(["组别", "备注", "学号", "姓名", "微信昵称", "提交日期", "手机号", "订单号", "", "订单金额"])
    grouped.append(["第1组", "", "1组组长：sand"])
    grouped.append(["", "", "1_01", "李媛", "李喵喵", "2026-05-07 12:35:09", "15923337197", "MA2026050712350910796457", "", "499"])
    grouped.append(["已退费", "已退费", "1_02", "陈秀银", "银子", "2026-05-08 13:43:37", "13705937742", "MA2026050813433772995767", "", "499"])
    grouped.append(["第2组", "", "2组组长：Yy"])
    grouped.append(["", "", "2_01", "周丽艳", "Julia", "2026-05-07 15:48:44", "17708888026", "MA2026050715484432032567", "", "499"])
    global_sheet = workbook.create_sheet("Sheet3")
    global_sheet.append(["序号", "备注", "提交时间", "真实姓名", "微信昵称", "手机号码", "商户订单号", "", "支付金额"])
    global_sheet.append([1, "", "2026-05-07 12:35:09", "李媛", "李喵喵", "15923337197", "MA2026050712350910796457", "", "499"])
    global_sheet.append([2, "已退费", "2026-05-08 13:43:37", "陈秀银", "银子", "13705937742", "MA2026050813433772995767", "", "499"])
    global_sheet.append([3, "", "2026-05-07 15:48:44", "周丽艳", "Julia", "17708888026", "MA2026050715484432032567", "", "499"])
    buffer = io.BytesIO()
    workbook.save(buffer)

    def fake_chat_with_provider(**_kwargs):
        return {
            "content": json.dumps(
                {
                    "rows": [
                        {"分组": "", "序号": "1", "姓名": "李媛", "微信昵称": "李喵喵", "手机号": "15923337197", "微信支付订单号": "MA2026050712350910796457"},
                        {"分组": "", "序号": "2", "备注": "已退费", "姓名": "陈秀银", "微信昵称": "银子", "手机号": "13705937742", "微信支付订单号": "MA2026050813433772995767"},
                        {"分组": "", "序号": "3", "姓名": "周丽艳", "微信昵称": "Julia", "手机号": "17708888026", "微信支付订单号": "MA2026050715484432032567"},
                    ],
                    "warnings": [],
                    "mapping_notes": ["AI 使用了全局流水源表"],
                },
                ensure_ascii=False,
            ),
        }

    monkeypatch.setattr(note_sheets_api, "chat_with_provider", fake_chat_with_provider)

    try:
        sheet_response = client.post(
            "/api/note-sheets/sheets",
            json={
                "title": "报名表",
                "document_json": {
                    "schema_version": 1,
                    "columns": ["分组", "序号", "备注", "提交时间", "姓名", "微信昵称", "手机号", "微信支付订单号", "商户订单号", "订单金额"],
                    "rows": [],
                    "grid_rows": [["分组", "序号", "备注", "提交时间", "姓名", "微信昵称", "手机号", "微信支付订单号", "商户订单号", "订单金额"]],
                    "data_start_row": 1,
                    "field_row_index": 0,
                    "column_configs": {"序号": {"value_type": "number", "header_background_color": "#9DC3E6"}},
                },
            },
        )
        sheet_id = sheet_response.json()["id"]

        response = client.post(
            f"/api/note-sheets/sheets/{sheet_id}/import-excel-reset",
            data={"instruction": "导入 47 届中心教室学员报名表"},
            files={
                "file": (
                    "47届中心教室学员信息等表v1.xlsx",
                    buffer.getvalue(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                ),
            },
        )

        assert response.status_code == 200
        payload = response.json()
        document = payload["sheet"]["document_json"]
        assert [row[1] for row in document["rows"]] == ["1_01", "1_02", "2_01"]
        assert [row[0] for row in document["rows"]] == ["第1组", "第1组", "第2组"]
        assert [row[3] for row in document["rows"]] == [
            "2026-05-07 12:35:09",
            "2026-05-08 13:43:37",
            "2026-05-07 15:48:44",
        ]
        assert [row[7] for row in document["rows"]] == ["", "", ""]
        assert [row[8] for row in document["rows"]] == [
            "MA2026050712350910796457",
            "MA2026050813433772995767",
            "MA2026050715484432032567",
        ]
        assert [row[9] for row in document["rows"]] == ["499", "499", "499"]
        assert document["column_configs"]["序号"] == {"header_background_color": "#9DC3E6"}
        assert any("分组序号" in note for note in payload["mapping_notes"])
        assert any("商户订单号" in note for note in payload["mapping_notes"])
    finally:
        _clear_user_override()


def test_registration_import_plain_sequences_become_grouped_when_groups_reset():
    columns = ["分组", "序号", "备注", "提交时间", "姓名"]
    rows = note_sheets_api._coerce_registration_import_rows(
        [],
        [
            ["一组", "1", "", "2026-05-07 12:35:09", "一组一号"],
            ["一组", "2", "", "2026-05-07 12:36:09", "一组二号"],
            ["二组", "1", "", "2026-05-07 12:37:09", "二组一号"],
        ],
        columns,
    )

    assert [row[1] for row in rows] == ["1_01", "1_02", "2_01"]


def test_registration_import_normalizes_group_sequence_separators():
    columns = ["分组", "序号", "备注", "提交时间", "姓名"]
    rows = note_sheets_api._coerce_registration_import_rows(
        [],
        [
            ["一组", "1-02", "", "2026-05-07 12:35:09", "横线"],
            ["一组", "1－03", "", "2026-05-07 12:36:09", "全角横线"],
            ["二组", "2组04号", "", "2026-05-07 12:37:09", "组号写法"],
        ],
        columns,
    )

    assert [row[1] for row in rows] == ["1_02", "1_03", "2_04"]


def test_note_sheet_excel_import_append_skips_duplicate_registration_order_id(client, session, monkeypatch):
    user = _create_user(session, username="note-sheet-excel-import-duplicate-order-user")
    _grant_feature_access(session, user_id=user.id, feature_key="notes.sheets")
    _override_user(user)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "重复报名"
    worksheet.append(["序号", "姓名", "微信支付订单号", "提交时间"])
    worksheet.append([1, "重复学员", "4200000000000000000000000000", "2026/5/21 09:46"])
    buffer = io.BytesIO()
    workbook.save(buffer)

    def fake_chat_with_provider(**kwargs):
        return {
            "content": json.dumps(
                {
                    "rows": [
                        {
                            "分组": "",
                            "序号": "1",
                            "提交时间": "2026/5/21 09:46",
                            "姓名": "重复学员",
                            "微信支付订单号": "4200000000000000000000000000",
                        },
                    ],
                    "warnings": [],
                    "mapping_notes": [],
                },
                ensure_ascii=False,
            ),
        }

    monkeypatch.setattr(note_sheets_api, "chat_with_provider", fake_chat_with_provider)

    try:
        sheet_response = client.post(
            "/api/note-sheets/sheets",
            json={
                "title": "报名表",
                "document_json": {
                    "schema_version": 1,
                    "columns": ["分组", "序号", "提交时间", "姓名", "微信支付订单号"],
                    "rows": [
                        ["5月4日", "123", "2026/5/14 09:15", "已有学员", "4200000000000000000000000000"],
                    ],
                    "grid_rows": [
                        ["分组", "序号", "提交时间", "姓名", "微信支付订单号"],
                        ["5月4日", "123", "2026/5/14 09:15", "已有学员", "4200000000000000000000000000"],
                    ],
                    "data_start_row": 1,
                    "field_row_index": 0,
                },
            },
        )
        assert sheet_response.status_code == 200
        sheet_id = sheet_response.json()["id"]

        response = client.post(
            f"/api/note-sheets/sheets/{sheet_id}/import-excel-reset",
            data={"instruction": "重复报名", "mode": "append"},
            files={
                "file": (
                    "duplicate.xlsx",
                    buffer.getvalue(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                ),
            },
        )

        assert response.status_code == 200
        payload = response.json()
        assert payload["imported_count"] == 0
        assert payload["skipped_duplicate_count"] == 1
        assert payload["sheet"]["document_json"]["rows"] == [
            ["5月4日", "123", "2026/5/14 09:15", "已有学员", "4200000000000000000000000000"],
        ]
    finally:
        _clear_user_override()


def test_note_sheet_excel_import_append_skips_duplicate_payment_order_id_on_plain_sheet(client, session, monkeypatch):
    user = _create_user(session, username="note-sheet-excel-import-plain-duplicate-order-user")
    _grant_feature_access(session, user_id=user.id, feature_key="notes.sheets")
    _override_user(user)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "订单源表"
    worksheet.append(["姓名", "微信支付订单号"])
    worksheet.append(["重复学员", "4200000000000000000000000000"])
    worksheet.append(["新学员", "4200000000000000000000000001"])
    buffer = io.BytesIO()
    workbook.save(buffer)

    def fake_chat_with_provider(**_kwargs):
        return {
            "content": json.dumps(
                {
                    "rows": [
                        {"姓名": "重复学员", "微信支付订单号": "4200000000000000000000000000"},
                        {"姓名": "新学员", "微信支付订单号": "4200000000000000000000000001"},
                    ],
                    "warnings": [],
                    "mapping_notes": [],
                },
                ensure_ascii=False,
            ),
        }

    monkeypatch.setattr(note_sheets_api, "chat_with_provider", fake_chat_with_provider)

    try:
        sheet_response = client.post(
            "/api/note-sheets/sheets",
            json={
                "title": "普通订单表",
                "document_json": {
                    "schema_version": 1,
                    "columns": ["姓名", "微信支付订单号"],
                    "rows": [["已有学员", "4200000000000000000000000000"]],
                },
            },
        )
        assert sheet_response.status_code == 200
        sheet_id = sheet_response.json()["id"]

        response = client.post(
            f"/api/note-sheets/sheets/{sheet_id}/import-excel-reset",
            data={"instruction": "普通增量导入", "mode": "append"},
            files={
                "file": (
                    "plain-duplicate.xlsx",
                    buffer.getvalue(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                ),
            },
        )

        assert response.status_code == 200
        payload = response.json()
        assert payload["imported_count"] == 1
        assert payload["skipped_duplicate_count"] == 1
        assert payload["sheet"]["document_json"]["rows"] == [
            ["已有学员", "4200000000000000000000000000"],
            ["新学员", "4200000000000000000000000001"],
        ]
    finally:
        _clear_user_override()


def test_note_sheet_excel_import_extracts_sparse_rows_without_cell_row_attribute():
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "稀疏表"
    worksheet["B2"] = "1组"
    worksheet["D2"] = "阿丹"
    buffer = io.BytesIO()
    workbook.save(buffer)

    payload = note_sheets_api._extract_excel_workbook_payload(buffer.getvalue(), "sparse.xlsx")

    assert payload["sheets"][0]["rows"] == [
        {"row_number": 2, "values": ["", "1组", "", "阿丹"]},
    ]


def test_note_sheet_attendance_progress_backgrounds_are_derived_from_cell_text_and_clockin_rules():
    columns = ["姓名", "打卡数", "05:20~06:43 第01课", "05:20~06:52 第02课", "追踪状态", "冻结时间"]
    document, changed_count = note_sheets_api._apply_attendance_progress_backgrounds({
        "schema_version": 1,
        "columns": columns,
        "rows": [
            ["学习者", 10, "学习中/21%", "1遍/100%", "追踪中", ""],
            ["空进度", "", "", "", "追踪中", ""],
            ["冻结者", 15, "1遍/100%", "学习中/33%", "已冻结", "2026-05-01 08:00:00"],
            ["觉观学员", 5, "当堂完成/98%", "暂不识别", "追踪中", ""],
            ["禅宗学员", 8, "准时完成", "延1周完成", "追踪中", ""],
            ["未返款", 4, "", "", "追踪中", ""],
        ],
        "grid_rows": [
            ["", "打卡", "6月1日~6月5日", "6月2日~6月6日", "", ""],
            columns,
            ["", '打卡达到"5/10/15"次，累计返回"30/60/100"元', "", "", "", ""],
        ],
        "data_start_row": 3,
        "field_row_index": 1,
        "source_meta": {
            "timed_video_rules": {"当堂": 19, "第1天": 14, "第2天": 9, "第3天": 4, "回放": 0},
        },
        "cell_meta": {
            "4:2": {"style": {"background_color": "#80FF80"}},
            "5:2": {"style": {"background_color": "#F2F2F2", "text_color": "#6B7280"}},
            "6:3": {"style": {"background_color": "#ABCDEF"}},
            "7:1": {"style": {"background_color": "#80FF80"}},
        },
    })

    assert changed_count == 8
    assert document["cell_meta"]["3:1"]["style"]["background_color"] == "#FFEBAD"
    assert "3:2" not in document["cell_meta"]
    assert document["cell_meta"]["3:3"]["style"]["background_color"] == "#80FF80"
    assert "4:2" not in document["cell_meta"]
    assert document["cell_meta"]["5:2"]["style"]["background_color"] == "#F2F2F2"
    assert document["cell_meta"]["6:1"]["style"]["background_color"] == "#FFF8DC"
    assert document["cell_meta"]["6:2"]["style"]["background_color"] == "#80FF80"
    assert document["cell_meta"]["6:3"]["style"]["background_color"] == "#ABCDEF"
    assert document["cell_meta"]["7:2"]["style"]["background_color"] == "#80FF80"
    assert document["cell_meta"]["7:3"]["style"]["background_color"] == "#D9D9D9"
    assert "8:1" not in document["cell_meta"]


def test_note_sheet_attendance_video_backgrounds_parse_refund_rules_from_existing_sheet_note():
    columns = ["姓名", "视频应返款", "05:20~06:18 第01课", "05:20~06:02 第02课", "05:20~06:10 第03课", "05:20~06:07 第04课"]
    document, _changed_count = note_sheets_api._apply_attendance_progress_backgrounds({
        "schema_version": 1,
        "columns": columns,
        "rows": [
            [
                "觉观学员",
                '=COUNTIF(C4:D4,"*当堂*")*19+COUNTIF(C4:D4,"*第1天*")*14+COUNTIF(C4:D4,"*第2天*")*9+COUNTIF(C4:D4,"*第3天*")*4',
                "当堂完成/98%",
                "第1天回放/99%",
                "学习中/39%",
                "第4天回放/99%",
            ],
        ],
        "grid_rows": [
            ["", "", "6月1日~6月5日", "6月2日~6月6日", "6月3日~6月7日", "6月4日~6月8日"],
            columns,
            [
                "",
                '21课*19元=399元。视频在"当堂(直播)/第1天(当天)/第2天/第3天/第4~5天"看完，对应返回"19/14/9/4/0"元',
                "",
                "",
                "",
                "",
            ],
            [
                "觉观学员",
                '=COUNTIF(C4:D4,"*当堂*")*19+COUNTIF(C4:D4,"*第1天*")*14+COUNTIF(C4:D4,"*第2天*")*9+COUNTIF(C4:D4,"*第3天*")*4',
                "当堂完成/98%",
                "第1天回放/99%",
                "学习中/39%",
                "第4天回放/99%",
            ],
        ],
        "data_start_row": 3,
        "field_row_index": 1,
    })

    assert document["cell_meta"]["3:2"]["style"]["background_color"] == "#80FF80"
    assert document["cell_meta"]["3:3"]["style"]["background_color"] == "#FFE18D"
    assert "3:4" not in document["cell_meta"]
    assert document["cell_meta"]["3:5"]["style"]["background_color"] == "#D9D9D9"


def test_note_sheet_attendance_challenge_progress_uses_completion_state_not_percent_depth():
    columns = ["姓名", "05:20~06:43 第01课", "05:20~06:52 第02课"]
    document, _changed_count = note_sheets_api._apply_attendance_progress_backgrounds({
        "schema_version": 1,
        "columns": columns,
        "rows": [
            ["念住学员", "1遍/89%", "学习中/88%"],
        ],
        "grid_rows": [
            columns,
            ["念住学员", "1遍/89%", "学习中/88%"],
        ],
        "data_start_row": 1,
    })

    assert document["cell_meta"]["1:1"]["style"]["background_color"] == "#80FF80"
    assert "1:2" not in document["cell_meta"]


def test_registration_standard_user_id_column_styles_include_linked_user_id():
    document, changed = note_sheets_api._apply_registration_standard_user_id_column_styles({
        "schema_version": 1,
        "columns": ["用户ID", "参考信息", "关联用户ID"],
        "rows": [],
        "grid_rows": [["用户ID", "参考信息", "关联用户ID"], ["", "", ""]],
        "data_start_row": 2,
        "field_row_index": 0,
        "column_configs": {
            "用户ID": {"width_mode": "fixed"},
            "关联用户ID": {"width_mode": "fixed"},
        },
        "cell_meta": {
            "0:2": {"style": {"background_color": "#F4B183"}},
        },
    })

    assert changed is True
    assert document["column_configs"]["用户ID"]["header_background_color"] == "#9DC3E6"
    assert document["column_configs"]["关联用户ID"]["header_background_color"] == "#9DC3E6"
    assert document["column_configs"]["关联用户ID"]["font_family"] == "monospace"
    assert (
        document["column_configs"]["关联用户ID"]["note"]
        == "有的用户账号数据源不统一，这里可以逗号隔开填写其他相关id，会合并到主id数据中汇总进度"
    )
    assert document["cell_meta"]["0:0"]["style"]["background_color"] == "#9DC3E6"
    assert document["cell_meta"]["0:2"]["style"]["background_color"] == "#9DC3E6"


def test_ensure_registration_linked_user_id_column_uses_standard_header_style():
    document, linked_index = note_sheets_api._ensure_registration_linked_user_id_column({
        "schema_version": 1,
        "columns": ["姓名", "用户ID", "参考信息"],
        "rows": [["学员", "u1", ""]],
        "grid_rows": [["姓名", "用户ID", "参考信息"], ["学员", "u1", ""]],
        "data_start_row": 1,
        "field_row_index": 0,
        "column_configs": {},
        "cell_meta": {},
    })

    assert document["columns"][linked_index] == "关联用户ID"
    assert document["column_configs"]["用户ID"]["header_background_color"] == "#9DC3E6"
    assert document["column_configs"]["关联用户ID"]["header_background_color"] == "#9DC3E6"
    assert document["cell_meta"][f"0:{linked_index}"]["style"]["background_color"] == "#9DC3E6"


def test_note_sheet_excel_import_unknown_row_keys_become_extra_columns():
    rows, extra_columns = note_sheets_api._coerce_note_sheet_excel_import_rows(
        {
            "rows": [
                {
                    "姓名": "阿丹",
                    "微信支付订单号": "`4200003070202605026584858144",
                    "微信号": "adan-wx",
                    "促学金模式": "自觉自律完成学修（无需支付促学金）",
                },
            ],
        },
        ["姓名", "微信支付订单号", "备注", "参考信息"],
    )

    assert extra_columns == ["微信号", "促学金模式"]
    assert rows == [["阿丹", "4200003070202605026584858144", "", "", "adan-wx", "自觉自律完成学修（无需支付促学金）"]]


def test_note_sheet_registration_order_match_updates_order_columns(client, session, monkeypatch):
    user = _create_user(session, username="note-sheet-order-match-user")
    _grant_feature_access(session, user_id=user.id, feature_key="notes.sheets")
    _override_user(user)

    target_columns = [
        "分组",
        "序号",
        "备注",
        "提交时间",
        "姓名",
        "微信昵称",
        "手机号",
        "错误手机号",
        "微信支付订单号",
        "订单日期",
        "商户订单号",
        "订单金额",
        "已返款",
        "用户ID",
        "关联用户ID",
        "匹配得分",
        "参考信息",
    ]
    order_index = target_columns.index("微信支付订单号")
    date_index = target_columns.index("订单日期")
    merchant_index = target_columns.index("商户订单号")
    amount_index = target_columns.index("订单金额")
    refunded_index = target_columns.index("已返款")

    calls: list[dict[str, object]] = []

    def fake_get_kqdb():
        return object()

    def fake_lookup_order(order_id, **kwargs):
        calls.append({"order_id": order_id, **kwargs})
        return {
            "微信支付订单号": f"`{order_id}",
            "订单日期": "202605",
            "商户订单号": "M20260509",
            "订单金额": 550.0,
            "已返款": 0.0,
        }

    monkeypatch.setattr(note_sheets_api, "_load_attendance_kqdb_provider", lambda: fake_get_kqdb)
    monkeypatch.setattr(note_sheets_api, "_load_attendance_order_lookup_provider", lambda: fake_lookup_order)

    try:
        workbook_response = client.post("/api/note-sheets/workbooks", json={"title": "20260509梵呗初阶"})
        workbook_id = workbook_response.json()["id"]
        row = [""] * len(target_columns)
        row[target_columns.index("姓名")] = "阿丹"
        row[order_index] = "4200003070202605026584858144"
        sheet_response = client.post(
            "/api/note-sheets/sheets",
            json={
                "title": "报名表",
                "workbook_id": workbook_id,
                "document_json": {
                    "schema_version": 1,
                    "columns": target_columns,
                    "rows": [row],
                    "grid_rows": [
                        target_columns,
                        [""] * order_index + ["更新订单匹配"] + [""] * (len(target_columns) - order_index - 1),
                        row,
                    ],
                    "data_start_row": 2,
                    "field_row_index": 0,
                },
            },
        )
        sheet_id = sheet_response.json()["id"]

        response = client.post(
            f"/api/note-sheets/sheets/{sheet_id}/registration/update-order-match",
            params={"workbook_id": workbook_id},
        )

        assert response.status_code == 200
        payload = response.json()
        assert payload["updated_count"] == 1
        assert payload["skipped_count"] == 0
        assert payload["error_count"] == 0
        updated_row = payload["sheet"]["document_json"]["rows"][0]
        assert updated_row[order_index] == "4200003070202605026584858144"
        assert updated_row[date_index] == "202605"
        assert updated_row[merchant_index] == "M20260509"
        assert updated_row[amount_index] == "550"
        assert updated_row[refunded_index] == "0"
        assert calls[0]["order_id"] == "4200003070202605026584858144"
        assert calls[0]["lookup_mode"] == "db_only"
        assert calls[0]["use_browser"] is False
    finally:
        _clear_user_override()


def test_note_sheet_registration_order_match_allows_missing_optional_refund_column(client, session, monkeypatch):
    user = _create_user(session, username="note-sheet-order-match-no-refund-user")
    _grant_feature_access(session, user_id=user.id, feature_key="notes.sheets")
    _override_user(user)

    target_columns = ["姓名", "微信支付订单号", "订单日期", "商户订单号", "订单金额"]
    order_index = target_columns.index("微信支付订单号")
    date_index = target_columns.index("订单日期")
    merchant_index = target_columns.index("商户订单号")
    amount_index = target_columns.index("订单金额")

    def fake_get_kqdb():
        return object()

    def fake_lookup_order(order_id, **kwargs):
        return {
            "微信支付订单号": order_id,
            "订单日期": "202605",
            "商户订单号": "M20260509",
            "订单金额": 550,
            "已返款": 0,
        }

    monkeypatch.setattr(note_sheets_api, "_load_attendance_kqdb_provider", lambda: fake_get_kqdb)
    monkeypatch.setattr(note_sheets_api, "_load_attendance_order_lookup_provider", lambda: fake_lookup_order)

    try:
        workbook_response = client.post("/api/note-sheets/workbooks", json={"title": "20260509梵呗初阶"})
        workbook_id = workbook_response.json()["id"]
        row = ["阿丹", "4200003070202605026584858144", "", "", ""]
        sheet_response = client.post(
            "/api/note-sheets/sheets",
            json={
                "title": "报名表",
                "workbook_id": workbook_id,
                "document_json": {
                    "schema_version": 1,
                    "columns": target_columns,
                    "rows": [row],
                    "grid_rows": [target_columns, row],
                    "data_start_row": 1,
                    "field_row_index": 0,
                },
            },
        )
        sheet_id = sheet_response.json()["id"]

        response = client.post(
            f"/api/note-sheets/sheets/{sheet_id}/registration/update-order-match",
            params={"workbook_id": workbook_id},
        )

        assert response.status_code == 200
        updated_row = response.json()["sheet"]["document_json"]["rows"][0]
        assert updated_row[order_index] == "4200003070202605026584858144"
        assert updated_row[date_index] == "202605"
        assert updated_row[merchant_index] == "M20260509"
        assert updated_row[amount_index] == "550"
    finally:
        _clear_user_override()


def test_note_sheet_registration_user_match_updates_user_columns(client, session, monkeypatch):
    user = _create_user(session, username="note-sheet-user-match-user")
    _grant_feature_access(session, user_id=user.id, feature_key="notes.sheets")
    _override_user(user)

    target_columns = ["姓名", "微信昵称", "手机号", "错误手机号", "用户ID", "匹配得分"]
    calls: list[dict[str, object]] = []

    class FakeKqdb:
        def 查找用户(self, names, phones, **kwargs):
            calls.append({"names": names, "phones": phones, **kwargs})
            return "u_adan", 92

    monkeypatch.setattr(note_sheets_api, "_load_attendance_kqdb_provider", lambda: (lambda: FakeKqdb()))

    try:
        workbook_response = client.post("/api/note-sheets/workbooks", json={"title": "20260509梵呗初阶"})
        workbook_id = workbook_response.json()["id"]
        sheet_response = client.post(
            "/api/note-sheets/sheets",
            json={
                "title": "报名表",
                "workbook_id": workbook_id,
                "document_json": {
                    "schema_version": 1,
                    "columns": target_columns,
                    "rows": [["阿丹", "阿丹", "15326693765", "", "", ""]],
                    "grid_rows": [
                        target_columns,
                        ["", "", "", "", "更新用户匹配", ""],
                        ["阿丹", "阿丹", "15326693765", "", "", ""],
                    ],
                    "data_start_row": 2,
                    "field_row_index": 0,
                },
            },
        )
        sheet_id = sheet_response.json()["id"]

        response = client.post(
            f"/api/note-sheets/sheets/{sheet_id}/registration/update-user-match",
            params={"workbook_id": workbook_id},
        )

        assert response.status_code == 200
        payload = response.json()
        assert payload["updated_count"] == 1
        assert payload["skipped_count"] == 0
        assert payload["error_count"] == 0
        updated_row = payload["sheet"]["document_json"]["rows"][0]
        assert updated_row[target_columns.index("用户ID")] == "u_adan"
        assert updated_row[target_columns.index("匹配得分")] == "92"
        assert calls[0]["names"] == ["阿丹", "阿丹"]
        assert calls[0]["phones"] == ["15326693765"]
        assert calls[0]["课程标准名"] == "d260509梵呗初阶"
        assert calls[0]["shop_id"] == 1
        assert calls[0]["return_mode"] == 1
    finally:
        _clear_user_override()


def test_note_sheet_registration_user_match_uses_browser_fallback_when_db_missing(client, session, monkeypatch):
    user = _create_user(session, username="note-sheet-user-match-browser-user")
    _grant_feature_access(session, user_id=user.id, feature_key="notes.sheets")
    _override_user(user)

    target_columns = ["姓名", "微信昵称", "手机号", "错误手机号", "用户ID", "匹配得分"]
    db_calls: list[dict[str, object]] = []
    remote_calls: list[dict[str, object]] = []

    class FakeKqdb:
        def 查找用户(self, names, phones, **kwargs):
            db_calls.append({"names": names, "phones": phones, **kwargs})
            return "", -1

    def fake_remote_lookup(session, current_user, *, course_name, shop_id, items):
        remote_calls.append(
            {
                "course_name": course_name,
                "shop_id": shop_id,
                "items": items,
            }
        )
        return {items[0]["key"]: {"key": items[0]["key"], "user_id": "u_live"}}

    monkeypatch.setattr(note_sheets_api, "_load_attendance_kqdb_provider", lambda: (lambda: FakeKqdb()))
    monkeypatch.setattr(note_sheets_api, "_lookup_registration_users_with_remote_browser", fake_remote_lookup)

    try:
        workbook_response = client.post("/api/note-sheets/workbooks", json={"title": "20260509梵呗初阶.xlsx"})
        workbook_id = workbook_response.json()["id"]
        sheet_response = client.post(
            "/api/note-sheets/sheets",
            json={
                "title": "报名表",
                "workbook_id": workbook_id,
                "document_json": {
                    "schema_version": 1,
                    "columns": target_columns,
                    "rows": [["阿丹", "阿丹", "`15326693765", "", "", ""]],
                    "grid_rows": [
                        target_columns,
                        ["", "", "", "", "更新用户匹配", ""],
                        ["阿丹", "阿丹", "`15326693765", "", "", ""],
                    ],
                    "data_start_row": 2,
                    "field_row_index": 0,
                },
            },
        )
        sheet_id = sheet_response.json()["id"]

        response = client.post(
            f"/api/note-sheets/sheets/{sheet_id}/registration/update-user-match",
            params={"workbook_id": workbook_id},
        )

        assert response.status_code == 200
        payload = response.json()
        assert payload["updated_count"] == 1
        assert payload["skipped_count"] == 0
        assert payload["error_count"] == 0
        updated_row = payload["sheet"]["document_json"]["rows"][0]
        assert updated_row[target_columns.index("用户ID")] == "u_live"
        assert updated_row[target_columns.index("匹配得分")] == "95"
        assert db_calls[0]["phones"] == ["15326693765"]
        assert db_calls[0]["课程标准名"] == "d260509梵呗初阶"
        assert remote_calls == [
            {
                "course_name": "d260509梵呗初阶",
                "shop_id": 1,
                "items": [
                        {
                            "key": "0",
                            "names": ["阿丹", "阿丹"],
                            "phones": ["15326693765"],
                        }
                ],
            }
        ]
    finally:
        _clear_user_override()


def test_note_sheet_registration_user_match_run_is_background_and_dedupes(client, session, engine, monkeypatch):
    user = _create_user(session, username="note-sheet-user-match-run-user")
    _grant_feature_access(session, user_id=user.id, feature_key="notes.sheets")
    _override_user(user)

    target_columns = ["姓名", "微信昵称", "手机号", "错误手机号", "用户ID", "匹配得分"]

    class FakeKqdb:
        def 查找用户(self, names, phones, **kwargs):
            time.sleep(0.15)
            return "u_async", 91

    monkeypatch.setattr(note_sheets_api, "engine", engine)
    monkeypatch.setattr(note_sheets_api, "_load_attendance_kqdb_provider", lambda: (lambda: FakeKqdb()))

    try:
        workbook_response = client.post("/api/note-sheets/workbooks", json={"title": "20260509梵呗初阶"})
        workbook_id = workbook_response.json()["id"]
        sheet_response = client.post(
            "/api/note-sheets/sheets",
            json={
                "title": "报名表",
                "workbook_id": workbook_id,
                "document_json": {
                    "schema_version": 1,
                    "columns": target_columns,
                    "rows": [["阿丹", "阿丹", "15326693765", "", "", ""]],
                    "grid_rows": [
                        target_columns,
                        ["", "", "", "", "更新用户匹配", ""],
                        ["阿丹", "阿丹", "15326693765", "", "", ""],
                    ],
                    "data_start_row": 2,
                    "field_row_index": 0,
                },
            },
        )
        sheet_id = sheet_response.json()["id"]

        start_response = client.post(
            f"/api/note-sheets/sheets/{sheet_id}/registration/match-runs",
            params={"workbook_id": workbook_id},
            json={"action": "registration_user_match"},
        )
        assert start_response.status_code == 200
        started = start_response.json()
        assert started["status"] in {"pending", "running"}
        assert started["already_running"] is False
        run_id = started["run_id"]

        duplicate_response = client.post(
            f"/api/note-sheets/sheets/{sheet_id}/registration/match-runs",
            params={"workbook_id": workbook_id},
            json={"action": "registration_user_match"},
        )
        assert duplicate_response.status_code == 200
        duplicate = duplicate_response.json()
        assert duplicate["already_running"] is True
        assert duplicate["run_id"] == run_id

        active_response = client.get(
            f"/api/note-sheets/sheets/{sheet_id}/registration/match-runs/active",
            params={"workbook_id": workbook_id, "action": "registration_user_match"},
        )
        assert active_response.status_code == 200
        assert active_response.json()["run_id"] == run_id

        deadline = time.time() + 3
        final = None
        while time.time() < deadline:
            status_response = client.get(
                f"/api/note-sheets/sheets/{sheet_id}/registration/match-runs/{run_id}",
                params={"workbook_id": workbook_id},
            )
            assert status_response.status_code == 200
            final = status_response.json()
            if final["status"] == "completed":
                break
            time.sleep(0.05)

        assert final is not None
        assert final["status"] == "completed"
        assert final["updated_count"] == 1
        updated_row = final["sheet"]["document_json"]["rows"][0]
        assert updated_row[target_columns.index("用户ID")] == "u_async"
        assert updated_row[target_columns.index("匹配得分")] == "91"
    finally:
        _clear_user_override()


def test_note_sheet_registration_background_run_blocks_related_actions(client, session, engine, monkeypatch):
    user = _create_user(session, username="note-sheet-order-run-user")
    _grant_feature_access(session, user_id=user.id, feature_key="notes.sheets")
    _override_user(user)

    target_columns = ["姓名", "微信支付订单号", "订单日期", "商户订单号", "订单金额", "已返款"]

    def fake_get_kqdb():
        return object()

    def fake_lookup_order(order_id, **kwargs):
        time.sleep(0.15)
        return {
            "微信支付订单号": order_id,
            "订单日期": "202605",
            "商户订单号": "M20260509",
            "订单金额": 620,
            "已返款": 0,
        }

    monkeypatch.setattr(note_sheets_api, "engine", engine)
    monkeypatch.setattr(note_sheets_api, "_load_attendance_kqdb_provider", lambda: fake_get_kqdb)
    monkeypatch.setattr(note_sheets_api, "_load_attendance_order_lookup_provider", lambda: fake_lookup_order)

    try:
        workbook_response = client.post("/api/note-sheets/workbooks", json={"title": "20260509梵呗初阶"})
        workbook_id = workbook_response.json()["id"]
        row = ["阿丹", "4200003070202605026584858144", "", "", "", ""]
        sheet_response = client.post(
            "/api/note-sheets/sheets",
            json={
                "title": "报名表",
                "workbook_id": workbook_id,
                "document_json": {
                    "schema_version": 1,
                    "columns": target_columns,
                    "rows": [row],
                    "grid_rows": [target_columns, row],
                    "data_start_row": 1,
                    "field_row_index": 0,
                },
            },
        )
        sheet_id = sheet_response.json()["id"]

        start_response = client.post(
            f"/api/note-sheets/sheets/{sheet_id}/registration/match-runs",
            params={"workbook_id": workbook_id},
            json={"action": "registration_order_match"},
        )
        assert start_response.status_code == 200
        started = start_response.json()
        assert started["status"] in {"pending", "running"}
        run_id = started["run_id"]

        related_response = client.post(
            f"/api/note-sheets/sheets/{sheet_id}/registration/match-runs",
            params={"workbook_id": workbook_id},
            json={"action": "registration_composite_update"},
        )
        assert related_response.status_code == 200
        related = related_response.json()
        assert related["already_running"] is True
        assert related["run_id"] == run_id
        assert related["action"] == "registration_order_match"

        active_response = client.get(
            f"/api/note-sheets/sheets/{sheet_id}/registration/match-runs/active",
            params={"workbook_id": workbook_id},
        )
        assert active_response.status_code == 200
        assert active_response.json()["run_id"] == run_id

        deadline = time.time() + 3
        final = None
        while time.time() < deadline:
            status_response = client.get(
                f"/api/note-sheets/sheets/{sheet_id}/registration/match-runs/{run_id}",
                params={"workbook_id": workbook_id},
            )
            assert status_response.status_code == 200
            final = status_response.json()
            if final["status"] == "completed":
                break
            time.sleep(0.05)

        assert final is not None
        assert final["status"] == "completed"
        assert final["updated_count"] == 1
        updated_row = final["sheet"]["document_json"]["rows"][0]
        assert updated_row[target_columns.index("商户订单号")] == "M20260509"
    finally:
        _clear_user_override()


def test_note_sheet_registration_order_run_uses_remote_fallback_when_db_missing(client, session, engine, monkeypatch):
    user = _create_user(session, username="note-sheet-order-run-remote-user")
    _grant_feature_access(session, user_id=user.id, feature_key="notes.sheets")
    _override_user(user)

    target_columns = ["姓名", "微信支付订单号", "订单日期", "商户订单号", "订单金额", "已返款"]
    remote_calls: list[list[str]] = []

    def fake_get_kqdb():
        return object()

    def fake_lookup_order(order_id, **kwargs):
        assert kwargs["lookup_mode"] == "db_only"
        assert kwargs["use_browser"] is False
        return {}

    def fake_remote_order_lookup(session, current_user, *, order_ids):
        remote_calls.append(list(order_ids))
        return [
            {
                "微信支付订单号": order_ids[0],
                "商户订单号": "M20260521",
                "订单金额": 620,
                "已返款": 0,
            }
        ]

    monkeypatch.setattr(note_sheets_api, "engine", engine)
    monkeypatch.setattr(note_sheets_api, "_load_attendance_kqdb_provider", lambda: fake_get_kqdb)
    monkeypatch.setattr(note_sheets_api, "_load_attendance_order_lookup_provider", lambda: fake_lookup_order)
    monkeypatch.setattr(note_sheets_api, "_lookup_registration_orders_with_remote_browser", fake_remote_order_lookup)

    try:
        workbook_response = client.post("/api/note-sheets/workbooks", json={"title": "20260521念住闯关"})
        workbook_id = workbook_response.json()["id"]
        row = ["赵誉博", "4200003079202605211268357324", "", "", "", ""]
        sheet_response = client.post(
            "/api/note-sheets/sheets",
            json={
                "title": "报名表",
                "workbook_id": workbook_id,
                "document_json": {
                    "schema_version": 1,
                    "columns": target_columns,
                    "rows": [row],
                    "grid_rows": [target_columns, row],
                    "data_start_row": 1,
                    "field_row_index": 0,
                },
            },
        )
        sheet_id = sheet_response.json()["id"]

        start_response = client.post(
            f"/api/note-sheets/sheets/{sheet_id}/registration/match-runs",
            params={"workbook_id": workbook_id},
            json={"action": "registration_order_match", "use_browser_fallback": True},
        )
        assert start_response.status_code == 200
        run_id = start_response.json()["run_id"]

        deadline = time.time() + 3
        final = None
        while time.time() < deadline:
            status_response = client.get(
                f"/api/note-sheets/sheets/{sheet_id}/registration/match-runs/{run_id}",
                params={"workbook_id": workbook_id},
            )
            assert status_response.status_code == 200
            final = status_response.json()
            if final["status"] == "completed":
                break
            time.sleep(0.05)

        assert final is not None
        assert final["status"] == "completed"
        assert final["updated_count"] == 1
        assert final["warning_count"] == 0
        assert remote_calls == [["4200003079202605211268357324"]]
        updated_row = final["sheet"]["document_json"]["rows"][0]
        assert updated_row[target_columns.index("订单日期")] == "202605"
        assert updated_row[target_columns.index("商户订单号")] == "M20260521"
        assert updated_row[target_columns.index("订单金额")] == "620"
        assert "订单匹配：补全 1/1" in final["message"]
    finally:
        _clear_user_override()


def test_note_sheet_registration_order_match_audits_completed_refund_remark_with_merchant_order(session, monkeypatch):
    user = _create_user(session, username="note-sheet-order-refund-audit-user")
    target_columns = ["姓名", "备注", "微信支付订单号", "订单日期", "商户订单号", "订单金额", "已返款"]
    remote_calls: list[list[str]] = []

    def fake_get_kqdb():
        return object()

    def fake_lookup_order(order_id, **kwargs):
        assert order_id == "MA2026051816560312333250"
        assert kwargs["lookup_mode"] == "db_only"
        assert kwargs["use_browser"] is False
        return {
            "微信支付订单号": "`4200003054202605183494484451",
            "订单日期": "202605",
            "商户订单号": "MA2026051816560312333250",
            "订单金额": 499,
            "已返款": 0,
        }

    def fake_remote_order_lookup(session_arg, current_user, *, order_ids):
        remote_calls.append(list(order_ids))
        return [
            {
                "微信支付订单号": "`4200003054202605183494484451",
                "商户订单号": "MA2026051816560312333250",
                "订单金额": 499,
                "已返款": 0,
            }
        ]

    monkeypatch.setattr(note_sheets_api, "_load_attendance_kqdb_provider", lambda: fake_get_kqdb)
    monkeypatch.setattr(note_sheets_api, "_load_attendance_order_lookup_provider", lambda: fake_lookup_order)
    monkeypatch.setattr(note_sheets_api, "_lookup_registration_orders_with_remote_browser", fake_remote_order_lookup)

    document = {
        "schema_version": 1,
        "columns": target_columns,
        "rows": [["刘海燕", "已退费", "", "202605", "MA2026051816560312333250", "499", "0"]],
    }

    next_doc, summary = note_sheets_api._update_registration_order_match_document(
        document,
        session=session,
        current_user=user,
        use_browser_fallback=True,
    )

    updated_row = next_doc["rows"][0]
    assert summary["target_count"] == 1
    assert summary["matched_count"] == 1
    assert summary["warning_count"] == 1
    assert remote_calls == [["MA2026051816560312333250"]]
    assert updated_row[target_columns.index("微信支付订单号")] == "4200003054202605183494484451"
    assert updated_row[target_columns.index("已返款")] == "0"
    assert "支付复核异常" in updated_row[target_columns.index("备注")]
    assert "0/499" in updated_row[target_columns.index("备注")]


def test_note_sheet_registration_order_run_defaults_to_remote_fallback_when_db_missing(client, session, engine, monkeypatch):
    user = _create_user(session, username="note-sheet-order-run-default-remote-user")
    _grant_feature_access(session, user_id=user.id, feature_key="notes.sheets")
    _override_user(user)

    target_columns = ["姓名", "微信支付订单号", "订单日期", "商户订单号", "订单金额"]
    remote_calls: list[list[str]] = []

    def fake_get_kqdb():
        return object()

    def fake_lookup_order(order_id, **kwargs):
        assert kwargs["lookup_mode"] == "db_only"
        assert kwargs["use_browser"] is False
        return {}

    def fake_remote_order_lookup(session, current_user, *, order_ids):
        remote_calls.append(list(order_ids))
        return [
            {
                "微信支付订单号": order_ids[0],
                "商户订单号": "M20260601",
                "订单金额": 620,
            }
        ]

    monkeypatch.setattr(note_sheets_api, "engine", engine)
    monkeypatch.setattr(note_sheets_api, "_load_attendance_kqdb_provider", lambda: fake_get_kqdb)
    monkeypatch.setattr(note_sheets_api, "_load_attendance_order_lookup_provider", lambda: fake_lookup_order)
    monkeypatch.setattr(note_sheets_api, "_lookup_registration_orders_with_remote_browser", fake_remote_order_lookup)

    try:
        workbook_response = client.post("/api/note-sheets/workbooks", json={"title": "20260601念住闯关"})
        workbook_id = workbook_response.json()["id"]
        row = ["饶彦", "4200003135202606011298497449", "", "", ""]
        sheet_response = client.post(
            "/api/note-sheets/sheets",
            json={
                "title": "报名表",
                "workbook_id": workbook_id,
                "document_json": {
                    "schema_version": 1,
                    "columns": target_columns,
                    "rows": [row],
                    "grid_rows": [target_columns, row],
                    "data_start_row": 1,
                    "field_row_index": 0,
                },
            },
        )
        sheet_id = sheet_response.json()["id"]

        start_response = client.post(
            f"/api/note-sheets/sheets/{sheet_id}/registration/match-runs",
            params={"workbook_id": workbook_id},
            json={"action": "registration_order_match"},
        )
        assert start_response.status_code == 200
        started = start_response.json()
        assert started["use_browser_fallback"] is True
        run_id = started["run_id"]

        deadline = time.time() + 3
        final = None
        while time.time() < deadline:
            status_response = client.get(
                f"/api/note-sheets/sheets/{sheet_id}/registration/match-runs/{run_id}",
                params={"workbook_id": workbook_id},
            )
            assert status_response.status_code == 200
            final = status_response.json()
            if final["status"] == "completed":
                break
            time.sleep(0.05)

        assert final is not None
        assert final["status"] == "completed"
        assert final["use_browser_fallback"] is True
        assert remote_calls == [["4200003135202606011298497449"]]
        updated_row = final["sheet"]["document_json"]["rows"][0]
        assert updated_row[target_columns.index("订单日期")] == "202606"
        assert updated_row[target_columns.index("商户订单号")] == "M20260601"
        assert updated_row[target_columns.index("订单金额")] == "620"
    finally:
        _clear_user_override()


def test_note_sheet_registration_order_match_derives_missing_order_month_without_remote(client, session, monkeypatch):
    user = _create_user(session, username="note-sheet-order-date-derive-user")
    _grant_feature_access(session, user_id=user.id, feature_key="notes.sheets")
    _override_user(user)

    target_columns = ["姓名", "微信支付订单号", "订单日期", "商户订单号", "订单金额", "已返款"]
    lookup_calls: list[str] = []

    def fake_get_kqdb():
        return object()

    def fake_lookup_order(order_id, **kwargs):
        lookup_calls.append(order_id)
        return {}

    def fake_remote_order_lookup(*args, **kwargs):
        raise AssertionError("只缺订单日期时不应启动远程查单")

    monkeypatch.setattr(note_sheets_api, "_load_attendance_kqdb_provider", lambda: fake_get_kqdb)
    monkeypatch.setattr(note_sheets_api, "_load_attendance_order_lookup_provider", lambda: fake_lookup_order)
    monkeypatch.setattr(note_sheets_api, "_lookup_registration_orders_with_remote_browser", fake_remote_order_lookup)

    try:
        workbook_response = client.post("/api/note-sheets/workbooks", json={"title": "20260521念住闯关"})
        workbook_id = workbook_response.json()["id"]
        sheet_response = client.post(
            "/api/note-sheets/sheets",
            json={
                "title": "报名表",
                "workbook_id": workbook_id,
                "document_json": {
                    "schema_version": 1,
                    "columns": target_columns,
                    "rows": [["赵誉博", "4200003079202605211268357324", "", "TFD61Z-0OZRE8O-GRZQ", "620", "0"]],
                    "grid_rows": [
                        target_columns,
                        ["赵誉博", "4200003079202605211268357324", "", "TFD61Z-0OZRE8O-GRZQ", "620", "0"],
                    ],
                    "data_start_row": 1,
                    "field_row_index": 0,
                },
            },
        )
        sheet_id = sheet_response.json()["id"]

        response = client.post(
            f"/api/note-sheets/sheets/{sheet_id}/registration/update-order-match",
            params={"workbook_id": workbook_id},
        )

        assert response.status_code == 200
        payload = response.json()
        assert payload["updated_count"] == 1
        assert lookup_calls == []
        updated_row = payload["sheet"]["document_json"]["rows"][0]
        assert updated_row[target_columns.index("订单日期")] == "202605"
    finally:
        _clear_user_override()


def test_note_sheet_registration_composite_run_updates_matches_and_attendance(client, session, engine, monkeypatch):
    user = _create_user(session, username="note-sheet-composite-run-user")
    _grant_feature_access(session, user_id=user.id, feature_key="notes.sheets")
    _override_user(user)

    registration_columns = [
        "分组",
        "序号",
        "备注",
        "提交时间",
        "姓名",
        "微信昵称",
        "手机号",
        "错误手机号",
        "微信支付订单号",
        "订单日期",
        "商户订单号",
        "订单金额",
        "已返款",
        "用户ID",
        "关联用户ID",
        "匹配得分",
    ]
    attendance_columns = [
        "报名日期",
        "学号",
        "姓名",
        "昵称",
        "商户订单号",
        "用户ID",
        "禅客",
        "优秀学员评分",
        "完成视频数",
        "视频应返款",
        "打卡应返款",
        "总应返款",
        "已返款",
        "订单金额",
        "当前应返款",
        "返款配置",
        "打卡数",
        "第01课",
        "第02课",
        "追踪状态",
        "冻结时间",
        "规则版本",
        "关联用户ID",
    ]

    def fake_get_kqdb():
        class FakeKqdb:
            def 查找用户(self, names, phones, **kwargs):
                return "u_new", 95

        return FakeKqdb()

    def fake_lookup_order(order_id, **kwargs):
        return {
            "微信支付订单号": order_id,
            "订单日期": "202605",
            "商户订单号": "M20260521",
            "订单金额": 620,
            "已返款": 0,
        }

    monkeypatch.setattr(note_sheets_api, "engine", engine)
    monkeypatch.setattr(note_sheets_api, "_load_attendance_kqdb_provider", lambda: fake_get_kqdb)
    monkeypatch.setattr(note_sheets_api, "_load_attendance_order_lookup_provider", lambda: fake_lookup_order)

    try:
        workbook_response = client.post("/api/note-sheets/workbooks", json={"title": "20260521念住闯关"})
        workbook_id = workbook_response.json()["id"]
        registration_row = ["5月4日", "124", "", "2026/5/21 09:46:30", "赵誉博", "赵玉博", "17702935475", "", "4200003079202605211268357324", "", "", "", "", "", ""]
        registration_response = client.post(
            "/api/note-sheets/sheets",
            json={
                "title": "报名表",
                "workbook_id": workbook_id,
                "document_json": {
                    "schema_version": 1,
                    "columns": registration_columns,
                    "rows": [registration_row],
                    "grid_rows": [registration_columns, registration_row],
                    "data_start_row": 1,
                    "field_row_index": 0,
                },
            },
        )
        registration_sheet_id = registration_response.json()["id"]
        attendance_rows = [
            [
                "05/14 09:15",
                "123",
                "旧学员",
                "旧昵称",
                "M20260514",
                "u_old",
                "=AND(J2>=11,Q2>=7)*1",
                "0",
                "8",
                "160",
                "100",
                '=MIN(IFERROR(J2+K2+N2-$N$1,0),N2)',
                "220",
                "620",
                '=OR(LEN(E2)=19,LEN(E2)=24)*(L2-M2)',
                '=IF(O2>0,TEXTJOIN(",",TRUE,E2,O2,"念住闯关每日返款",E2&"_day"&$O$1),"")',
                "7",
                "1遍/99%",
                "1遍/100%",
                "追踪中",
                "",
                "当前规则",
            ],
            [
                "03/11 11:41",
                "115",
                "历史学员",
                "历史昵称",
                "M20260311",
                "u_history",
                "0",
                "0",
                "0",
                "40",
                "0",
                "40",
                "40",
                "620",
                "0",
                "",
                "",
                "1遍/137%",
                "3遍/210%",
                "已冻结",
                "2026-05-01 08:00:00",
                "当前规则",
            ],
        ]
        attendance_response = client.post(
            "/api/note-sheets/sheets",
            json={
                "title": "考勤表",
                "workbook_id": workbook_id,
                "document_json": {
                    "schema_version": 1,
                    "columns": attendance_columns,
                    "rows": attendance_rows,
                    "grid_rows": [attendance_columns, *attendance_rows],
                    "data_start_row": 1,
                    "formula_reference_origin": "sheet_v2",
                    "field_row_index": 0,
                    "cell_meta": {
                        "1:16": {"style": {"background_color": "#FFFFBB"}},
                        "1:17": {"style": {"background_color": "#80FF80"}},
                        "1:18": {"style": {"background_color": "#80FF80"}},
                        "2:0": {"style": {"background_color": "#F2F2F2", "text_color": "#6B7280"}},
                        "2:1": {"style": {"background_color": "#F2F2F2", "text_color": "#6B7280"}},
                        "2:17": {"style": {"background_color": "#80FF80"}},
                        "2:18": {"style": {"background_color": "#80FF80"}},
                    },
                },
            },
        )
        attendance_sheet_id = attendance_response.json()["id"]

        start_response = client.post(
            f"/api/note-sheets/sheets/{registration_sheet_id}/registration/match-runs",
            params={"workbook_id": workbook_id},
            json={"action": "registration_composite_update"},
        )
        assert start_response.status_code == 200
        run_id = start_response.json()["run_id"]

        deadline = time.time() + 3
        final = None
        while time.time() < deadline:
            status_response = client.get(
                f"/api/note-sheets/sheets/{registration_sheet_id}/registration/match-runs/{run_id}",
                params={"workbook_id": workbook_id},
            )
            assert status_response.status_code == 200
            final = status_response.json()
            if final["status"] == "completed":
                break
            time.sleep(0.05)

        assert final is not None
        assert final["status"] == "completed"
        assert final["updated_count"] == 4
        registration_updated = final["sheet"]["document_json"]["rows"][0]
        assert registration_updated[registration_columns.index("商户订单号")] == "M20260521"
        assert registration_updated[registration_columns.index("用户ID")] == "u_new"
        assert registration_updated[registration_columns.index("匹配得分")] == "95"

        attendance_detail = client.get(
            f"/api/note-sheets/sheets/{attendance_sheet_id}",
            params={"workbook_id": workbook_id, "paginate": "false"},
        )
        assert attendance_detail.status_code == 200
        attendance_updated_rows = attendance_detail.json()["document_json"]["rows"]
        assert attendance_updated_rows[1][attendance_columns.index("学号")] == "124"
        assert attendance_updated_rows[1][attendance_columns.index("报名日期")] == "2026-05-21 09:46"
        assert attendance_updated_rows[1][attendance_columns.index("姓名")] == "赵誉博"
        assert attendance_updated_rows[1][attendance_columns.index("昵称")] == "赵玉博"
        assert attendance_updated_rows[1][attendance_columns.index("商户订单号")] == "M20260521"
        assert attendance_updated_rows[1][attendance_columns.index("用户ID")] == "u_new"
        assert attendance_updated_rows[1][attendance_columns.index("禅客")] == '=IF(AND(I3>=11,Q3>=7),"是","")'
        assert attendance_updated_rows[1][attendance_columns.index("打卡应返款")] == "0"
        assert attendance_updated_rows[1][attendance_columns.index("总应返款")] == '=MIN(IFERROR(J3+K3+N3-$N$1,0),N3)'
        assert attendance_updated_rows[1][attendance_columns.index("已返款")] == "0"
        assert attendance_updated_rows[1][attendance_columns.index("订单金额")] == "620"
        assert attendance_updated_rows[1][attendance_columns.index("当前应返款")] == '=OR(LEN(E3)=19,LEN(E3)=24)*(L3-M3)'
        assert attendance_updated_rows[1][attendance_columns.index("返款配置")] == '=IF(O3>0,TEXTJOIN(",",TRUE,E3,O3,"念住闯关每日返款",E3&"_day"&$O$1),"")'
        assert attendance_updated_rows[1][attendance_columns.index("打卡数")] == ""
        assert attendance_updated_rows[1][attendance_columns.index("第01课")] == ""
        assert attendance_updated_rows[1][attendance_columns.index("第02课")] == ""
        assert attendance_updated_rows[1][attendance_columns.index("追踪状态")] == "追踪中"
        assert attendance_updated_rows[1][attendance_columns.index("规则版本")] == "当前规则"
        assert attendance_updated_rows[2][attendance_columns.index("学号")] == "115"
        attendance_cell_meta = attendance_detail.json()["document_json"]["cell_meta"]
        assert "2:0" not in attendance_cell_meta
        assert "2:16" not in attendance_cell_meta
        assert "2:17" not in attendance_cell_meta
        assert "2:18" not in attendance_cell_meta
        assert attendance_cell_meta["3:0"]["style"]["background_color"] == "#F2F2F2"
    finally:
        _clear_user_override()


def test_note_sheet_registration_attendance_sync_repairs_incomplete_existing_row():
    registration_columns = [
        "分组",
        "序号",
        "提交时间",
        "姓名",
        "微信昵称",
        "手机号",
        "微信支付订单号",
        "订单日期",
        "商户订单号",
        "订单金额",
        "已返款",
        "用户ID",
        "关联用户ID",
        "匹配得分",
    ]
    attendance_columns = [
        "报名日期",
        "学号",
        "姓名",
        "昵称",
        "商户订单号",
        "用户ID",
        "禅客",
        "优秀学员评分",
        "完成视频数",
        "视频应返款",
        "打卡应返款",
        "总应返款",
        "已返款",
        "订单金额",
        "当前应返款",
        "返款配置",
        "打卡数",
        "第01课",
        "追踪状态",
        "冻结时间",
        "规则版本",
        "关联用户ID",
    ]
    registration_doc = {
        "schema_version": 1,
        "columns": registration_columns,
        "rows": [[
            "5月4日",
            "124",
            "2026/5/21 09:46:30",
            "赵誉博",
            "赵玉博",
            "17702935475",
            "4200003079202605211268357324",
            "202605",
            "M20260521",
            "620",
            "0",
            "u_new",
            "u_linked",
            "95",
        ]],
    }
    attendance_doc = {
        "schema_version": 1,
        "columns": attendance_columns,
        "data_start_row": 1,
        "formula_reference_origin": "sheet_v2",
        "rows": [
            [
                "05/14 09:15",
                "123",
                "旧学员",
                "旧昵称",
                "M20260514",
                "u_old",
                "=AND(J2>=11,Q2>=7)*1",
                "0",
                "8",
                "160",
                "100",
                '=MIN(IFERROR(J2+K2+N2-$N$1,0),N2)',
                "220",
                "620",
                '=OR(LEN(E2)=19,LEN(E2)=24)*(L2-M2)',
                '=IF(O2>0,TEXTJOIN(",",TRUE,E2,O2,"念住闯关每日返款",E2&"_day"&$O$1),"")',
                "7",
                "1遍/99%",
                "追踪中",
                "",
                "当前规则",
                "",
            ],
            ["05/21 09:46", "124", "赵誉博", "赵玉博", "M20260521", "u_new", "", "0", "0", "0", "0", "", "", "", "", "", "", "", "追踪中", "", "当前规则", ""],
            ["03/11 11:41", "115", "历史学员", "历史昵称", "M20260311", "u_history", "0", "0", "0", "40", "0", "40", "40", "620", "0", "", "", "1遍/137%", "已冻结", "2026-05-01 08:00:00", "当前规则", ""],
        ],
        "grid_rows": [attendance_columns],
        "entity_rows": [
            {"id": "row_header", "kind": "field"},
            {"id": "row_old", "kind": "data"},
            {"id": "row_new", "kind": "data"},
            {"id": "row_archived", "kind": "data"},
        ],
        "entity_columns": [
            {"id": f"col_{index}", "header": header}
            for index, header in enumerate(attendance_columns)
        ],
        "entity_cells": {
            "row_new": {
                "col_0": {"style": {"background_color": "#F2F2F2", "text_color": "#6B7280"}, "value": "05/21 09:46"},
                "col_1": {"style": {"background_color": "#F2F2F2", "text_color": "#6B7280"}, "value": "124"},
                "col_17": {"style": {"background_color": "#80FF80"}},
            },
            "row_archived": {
                "col_0": {"style": {"background_color": "#F2F2F2", "text_color": "#6B7280"}, "value": "03/11 11:41"},
            },
        },
        "cell_meta": {
            "2:0": {"style": {"background_color": "#F2F2F2", "text_color": "#6B7280"}},
            "2:1": {"style": {"background_color": "#F2F2F2", "text_color": "#6B7280"}},
            "3:0": {"style": {"background_color": "#F2F2F2", "text_color": "#6B7280"}},
        },
    }

    next_doc, summary = note_sheets_api._sync_registration_rows_to_attendance_document(registration_doc, attendance_doc)

    assert summary["inserted_count"] == 0
    assert summary["repaired_count"] == 2
    repaired_row = next_doc["rows"][1]
    assert repaired_row[attendance_columns.index("报名日期")] == "2026-05-21 09:46"
    assert repaired_row[attendance_columns.index("禅客")] == '=IF(AND(I3>=11,Q3>=7),"是","")'
    assert repaired_row[attendance_columns.index("总应返款")] == '=MIN(IFERROR(J3+K3+N3-$N$1,0),N3)'
    assert repaired_row[attendance_columns.index("已返款")] == "0"
    assert repaired_row[attendance_columns.index("订单金额")] == "620"
    assert repaired_row[attendance_columns.index("关联用户ID")] == "u_linked"
    assert repaired_row[attendance_columns.index("当前应返款")] == '=OR(LEN(E3)=19,LEN(E3)=24)*(L3-M3)'
    assert repaired_row[attendance_columns.index("打卡数")] == ""
    assert "2:0" not in next_doc["cell_meta"]
    assert next_doc["cell_meta"]["3:0"]["style"]["background_color"] == "#F2F2F2"
    assert "style" not in next_doc["entity_cells"]["row_new"]["col_0"]
    assert "style" not in next_doc["entity_cells"]["row_new"]["col_1"]
    assert "col_17" not in next_doc["entity_cells"]["row_new"]
    assert next_doc["entity_cells"]["row_archived"]["col_0"]["style"]["background_color"] == "#F2F2F2"


def test_note_sheet_registration_user_id_detection_replaces_stale_primary_id(client, session, monkeypatch):
    user = _create_user(session, username="note-sheet-detect-replace-stale-user-id")
    _grant_feature_access(session, user_id=user.id, feature_key="notes.sheets")
    _override_user(user)
    rebuild_calls: list[str] = []

    monkeypatch.setattr(note_sheets_api, "_query_registration_detection_user_ids_by_phone", lambda phone: ["u_real"])

    def fake_rebuild(_session, *, attendance, course_name):
        rebuild_calls.append(course_name)
        return {"ok": True, "attendance_sheet_id": int(attendance.numeric_id or attendance.id)}

    monkeypatch.setattr(note_sheets_api, "_rebuild_registration_attendance_after_user_id_detection", fake_rebuild)

    try:
        workbook_response = client.post("/api/note-sheets/workbooks", json={"title": "第47届觉观"})
        workbook_id = workbook_response.json()["id"]
        registration_columns = ["姓名", "微信昵称", "手机号", "用户ID", "参考信息", "关联用户ID"]
        registration_row = ["学员", "学员昵称", "18800000000", "u_stale", "", ""]
        registration_response = client.post(
            "/api/note-sheets/sheets",
            json={
                "title": "报名表",
                "workbook_id": workbook_id,
                "document_json": {
                    "schema_version": 1,
                    "columns": registration_columns,
                    "rows": [registration_row],
                    "grid_rows": [registration_columns, registration_row],
                    "data_start_row": 1,
                    "field_row_index": 0,
                },
            },
        )
        assert registration_response.status_code == 200
        registration_sheet_id = registration_response.json()["id"]
        initial_version = registration_response.json()["version"]
        for title, document_json in [
            ("考勤表", {"columns": ["姓名", "用户ID"], "rows": [], "grid_rows": [["姓名", "用户ID"]], "data_start_row": 1}),
            ("视频数据", {"columns": ["user_id2", "nickname"], "rows": [["u_real", "学员昵称"]], "grid_rows": [["user_id2", "nickname"], ["u_real", "学员昵称"]], "data_start_row": 1}),
        ]:
            response = client.post(
                "/api/note-sheets/sheets",
                json={
                    "title": title,
                    "workbook_id": workbook_id,
                    "document_json": {"schema_version": 1, **document_json},
                },
            )
            assert response.status_code == 200

        stale_response = client.post(
            f"/api/note-sheets/sheets/{registration_sheet_id}/registration/detect-user-id",
            params={"workbook_id": workbook_id},
            json={"row_index": 0, "base_version": initial_version + 1},
        )
        assert stale_response.status_code == 409

        response = client.post(
            f"/api/note-sheets/sheets/{registration_sheet_id}/registration/detect-user-id",
            params={"workbook_id": workbook_id},
            json={"row_index": 0, "base_version": initial_version},
        )

        assert response.status_code == 200
        payload = response.json()
        assert payload["status"] == "applied"
        assert payload["applied_to"] == "用户ID"
        updated_row = payload["sheet"]["document_json"]["rows"][0]
        assert updated_row[registration_columns.index("用户ID")] == "u_real"
        assert updated_row[registration_columns.index("关联用户ID")] == ""
        assert "原用户ID：u_stale" in updated_row[registration_columns.index("参考信息")]
        assert rebuild_calls == ["第47届觉观"]
    finally:
        _clear_user_override()


def test_note_sheet_registration_user_id_detection_adds_alias_when_primary_has_progress(client, session, monkeypatch):
    user = _create_user(session, username="note-sheet-detect-add-linked-user-id")
    _grant_feature_access(session, user_id=user.id, feature_key="notes.sheets")
    _override_user(user)
    rebuild_calls: list[str] = []

    monkeypatch.setattr(note_sheets_api, "_query_registration_detection_user_ids_by_phone", lambda phone: ["u_alias"])

    def fake_rebuild(_session, *, attendance, course_name):
        rebuild_calls.append(course_name)
        return {"ok": True}

    monkeypatch.setattr(note_sheets_api, "_rebuild_registration_attendance_after_user_id_detection", fake_rebuild)

    try:
        workbook_response = client.post("/api/note-sheets/workbooks", json={"title": "第47届觉观"})
        workbook_id = workbook_response.json()["id"]
        registration_columns = ["姓名", "微信昵称", "手机号", "用户ID", "参考信息", "关联用户ID"]
        registration_row = ["闫子翼", "阿紫", "18800000000", "u_primary", "", ""]
        registration_response = client.post(
            "/api/note-sheets/sheets",
            json={
                "title": "报名表",
                "workbook_id": workbook_id,
                "document_json": {
                    "schema_version": 1,
                    "columns": registration_columns,
                    "rows": [registration_row],
                    "grid_rows": [registration_columns, registration_row],
                    "data_start_row": 1,
                    "field_row_index": 0,
                },
            },
        )
        assert registration_response.status_code == 200
        registration_sheet_id = registration_response.json()["id"]
        for title, document_json in [
            ("考勤表", {"columns": ["姓名", "用户ID"], "rows": [], "grid_rows": [["姓名", "用户ID"]], "data_start_row": 1}),
            (
                "视频数据",
                {
                    "columns": ["user_id2", "nickname", "lesson_id"],
                    "rows": [["u_primary", "阿紫", 1], ["u_alias", "阿紫", 3]],
                    "grid_rows": [["user_id2", "nickname", "lesson_id"], ["u_primary", "阿紫", 1], ["u_alias", "阿紫", 3]],
                    "data_start_row": 1,
                },
            ),
        ]:
            response = client.post(
                "/api/note-sheets/sheets",
                json={
                    "title": title,
                    "workbook_id": workbook_id,
                    "document_json": {"schema_version": 1, **document_json},
                },
            )
            assert response.status_code == 200

        response = client.post(
            f"/api/note-sheets/sheets/{registration_sheet_id}/registration/detect-user-id",
            params={"workbook_id": workbook_id},
            json={"row_index": 0},
        )

        assert response.status_code == 200
        payload = response.json()
        assert payload["status"] == "applied"
        assert payload["applied_to"] == "关联用户ID"
        updated_row = payload["sheet"]["document_json"]["rows"][0]
        assert updated_row[registration_columns.index("用户ID")] == "u_primary"
        assert updated_row[registration_columns.index("关联用户ID")] == "u_alias"
        assert "原用户ID：u_primary" in updated_row[registration_columns.index("参考信息")]
        assert rebuild_calls == ["第47届觉观"]
    finally:
        _clear_user_override()


def test_note_sheet_registration_user_id_detection_promotes_existing_linked_progress_id(client, session, monkeypatch):
    user = _create_user(session, username="note-sheet-detect-promote-linked-user-id")
    _grant_feature_access(session, user_id=user.id, feature_key="notes.sheets")
    _override_user(user)
    monkeypatch.setattr(
        note_sheets_api,
        "_rebuild_registration_attendance_after_user_id_detection",
        lambda _session, *, attendance, course_name: {"ok": True},
    )

    try:
        workbook_response = client.post("/api/note-sheets/workbooks", json={"title": "第47届觉观"})
        workbook_id = workbook_response.json()["id"]
        registration_columns = ["姓名", "手机号", "用户ID", "参考信息", "关联用户ID"]
        registration_row = ["学员", "18800000000", "u_stale", "", "u_real"]
        registration_response = client.post(
            "/api/note-sheets/sheets",
            json={
                "title": "报名表",
                "workbook_id": workbook_id,
                "document_json": {
                    "schema_version": 1,
                    "columns": registration_columns,
                    "rows": [registration_row],
                    "grid_rows": [registration_columns, registration_row],
                    "data_start_row": 1,
                    "field_row_index": 0,
                },
            },
        )
        assert registration_response.status_code == 200
        registration_sheet_id = registration_response.json()["id"]
        for title, document_json in [
            ("考勤表", {"columns": ["姓名", "用户ID"], "rows": [], "grid_rows": [["姓名", "用户ID"]], "data_start_row": 1}),
            ("视频数据", {"columns": ["user_id2"], "rows": [["u_real"]], "grid_rows": [["user_id2"], ["u_real"]], "data_start_row": 1}),
        ]:
            response = client.post(
                "/api/note-sheets/sheets",
                json={
                    "title": title,
                    "workbook_id": workbook_id,
                    "document_json": {"schema_version": 1, **document_json},
                },
            )
            assert response.status_code == 200

        response = client.post(
            f"/api/note-sheets/sheets/{registration_sheet_id}/registration/detect-user-id",
            params={"workbook_id": workbook_id},
            json={"row_index": 0},
        )

        assert response.status_code == 200
        payload = response.json()
        assert payload["status"] == "applied"
        updated_row = payload["sheet"]["document_json"]["rows"][0]
        assert updated_row[registration_columns.index("用户ID")] == "u_real"
        assert updated_row[registration_columns.index("关联用户ID")] == ""
        assert "关联用户ID已有唯一课程数据" in updated_row[registration_columns.index("参考信息")]
    finally:
        _clear_user_override()


def test_note_sheet_registration_attendance_sync_refreshes_expired_tracking_rows():
    columns = [
        "报名日期",
        "学号",
        "姓名",
        "昵称",
        "当前应返款",
        "追踪分组",
        "追踪状态",
        "追踪截止日",
        "冻结时间",
        "规则版本",
    ]
    attendance_doc = {
        "schema_version": 1,
        "columns": columns,
        "data_start_row": 1,
        "formula_reference_origin": "sheet_v2",
        "rows": [
            ["2026-03-22 23:17", "116", "李龙", "李龙", "=(E2>0)*1", "B组", "追踪中", "2026-05-22", "", "当前规则"],
            ["2026-05-21 09:46", "124", "赵誉博", "赵玉博", "=(E3>0)*1", "5月4日", "追踪中", "", "", "当前规则"],
            ["2026-03-11 11:41", "115", "历史学员", "历史昵称", "=(E4>0)*1", "A组", "已冻结", "2026-05-11", "2026-05-16 14:07:07", "当前规则"],
        ],
        "grid_rows": [columns],
    }

    next_doc, repaired_count = note_sheets_api._order_attendance_rows_by_dynamic_expiration(
        attendance_doc,
        now=date(2026, 6, 1),
    )

    assert repaired_count > 0
    rows = next_doc["rows"]
    assert [row[columns.index("学号")] for row in rows] == ["124", "116", "115"]
    assert rows[0][columns.index("追踪分组")] == "5月4日"
    assert rows[0][columns.index("追踪状态")] == "追踪中"
    assert rows[0][columns.index("追踪截止日")] == "2026-07-21"
    assert rows[1][columns.index("追踪分组")] == "A组"
    assert rows[1][columns.index("追踪状态")] == "已冻结"
    assert rows[1][columns.index("追踪截止日")] == "2026-05-22"
    assert rows[1][columns.index("冻结时间")]
    assert next_doc["cell_meta"]["2:0"]["style"]["background_color"] == "#F2F2F2"
    assert next_doc["cell_meta"]["2:0"]["style"]["text_color"] == "#6B7280"
    assert next_doc["cell_meta"]["2:1"]["style"]["background_color"] == "#F2F2F2"
    assert rows[0][columns.index("当前应返款")] == "=(E2>0)*1"


def test_note_sheet_registration_attendance_sync_skips_tracking_refresh_without_new_rows():
    registration_columns = ["序号", "提交时间", "姓名", "微信昵称", "用户ID"]
    attendance_columns = ["报名日期", "学号", "姓名", "昵称", "追踪分组", "追踪状态", "追踪截止日", "冻结时间"]
    registration_doc = {
        "schema_version": 1,
        "columns": registration_columns,
        "rows": [["116", "2026-03-22 23:17", "李龙", "李龙", "u_old"]],
    }
    attendance_doc = {
        "schema_version": 1,
        "columns": attendance_columns,
        "data_start_row": 1,
        "rows": [["2026-03-22 23:17", "116", "李龙", "李龙", "B组", "追踪中", "2026-05-22", ""]],
        "grid_rows": [attendance_columns],
    }

    next_doc, summary = note_sheets_api._sync_registration_rows_to_attendance_document(registration_doc, attendance_doc)

    assert summary["updated_count"] == 0
    assert next_doc["rows"] == attendance_doc["rows"]


def test_note_sheet_registration_attendance_sync_allows_attendance_without_user_id():
    registration_columns = ["序号", "提交时间", "姓名", "微信昵称", "商户订单号", "用户ID", "匹配得分"]
    attendance_columns = ["报名日期", "学号", "姓名", "昵称", "商户订单号", "打卡数"]
    registration_doc = {
        "schema_version": 1,
        "columns": registration_columns,
        "rows": [["124", "2026/5/21 09:46:30", "赵誉博", "赵玉博", "M20260521", "u_new", "95"]],
    }
    attendance_doc = {
        "schema_version": 1,
        "columns": attendance_columns,
        "data_start_row": 1,
        "formula_reference_origin": "sheet_v2",
        "rows": [],
        "grid_rows": [attendance_columns],
    }

    next_doc, summary = note_sheets_api._sync_registration_rows_to_attendance_document(registration_doc, attendance_doc)

    assert summary["inserted_count"] == 1
    assert next_doc["columns"] == attendance_columns
    assert next_doc["rows"] == [["2026-05-21 09:46", "124", "赵誉博", "赵玉博", "M20260521", ""]]


def test_note_sheet_registration_attendance_sync_skips_refunded_rows():
    registration_columns = ["序号", "备注", "提交时间", "姓名", "微信昵称", "商户订单号", "用户ID"]
    attendance_columns = ["报名日期", "学号", "姓名", "昵称", "商户订单号", "用户ID"]
    registration_doc = {
        "schema_version": 1,
        "columns": registration_columns,
        "rows": [
            ["1_01", "", "2026/5/21 09:46:30", "正常学员", "正常", "M20260521", "u_active"],
            ["1_02", "已退费", "2026/5/22 09:46:30", "退费学员", "退费", "M20260522", "u_refunded"],
        ],
    }
    attendance_doc = {
        "schema_version": 1,
        "columns": attendance_columns,
        "data_start_row": 1,
        "formula_reference_origin": "sheet_v2",
        "rows": [],
        "grid_rows": [attendance_columns],
        "column_configs": {"学号": {"value_type": "number", "header_background_color": "#D9E1F2"}},
    }

    next_doc, summary = note_sheets_api._sync_registration_rows_to_attendance_document(registration_doc, attendance_doc)

    assert summary["inserted_count"] == 1
    assert summary["skipped_count"] == 1
    assert next_doc["rows"] == [["2026-05-21 09:46", "1_01", "正常学员", "正常", "M20260521", "u_active"]]
    assert next_doc["column_configs"]["学号"] == {"header_background_color": "#D9E1F2"}


def test_note_sheet_registration_attendance_sync_inserts_identified_row_without_user_id():
    registration_columns = ["序号", "备注", "提交时间", "姓名", "微信昵称", "商户订单号", "用户ID", "匹配得分"]
    attendance_columns = ["报名日期", "学号", "姓名", "昵称", "商户订单号", "用户ID", "匹配得分"]
    registration_doc = {
        "schema_version": 1,
        "columns": registration_columns,
        "rows": [["2_15", "", "2026/5/29 10:08:01", "伍苗", "wm", "M20260529", "", "-1"]],
    }
    attendance_doc = {
        "schema_version": 1,
        "columns": attendance_columns,
        "data_start_row": 1,
        "formula_reference_origin": "sheet_v2",
        "rows": [],
        "grid_rows": [attendance_columns],
    }

    next_doc, summary = note_sheets_api._sync_registration_rows_to_attendance_document(registration_doc, attendance_doc)

    assert summary["inserted_count"] == 1
    assert next_doc["rows"] == [["2026-05-29 10:08", "2_15", "伍苗", "wm", "M20260529", "", "-1"]]


def test_note_sheet_registration_attendance_sync_styles_identity_columns_by_group():
    registration_columns = ["分组", "序号", "备注", "提交时间", "姓名", "微信昵称", "商户订单号", "用户ID"]
    attendance_columns = ["分组", "学号", "姓名", "昵称", "商户订单号", "用户ID", "禅客", "完成视频数"]
    registration_doc = {
        "schema_version": 1,
        "columns": registration_columns,
        "rows": [
            ["1组", "1_01", "", "2026/5/21 09:46:30", "一组学员", "一", "M20260521", "u_1"],
            ["二组", "2_01", "", "2026/5/22 09:46:30", "二组学员", "二", "M20260522", "u_2"],
        ],
    }
    attendance_doc = {
        "schema_version": 1,
        "columns": attendance_columns,
        "data_start_row": 3,
        "formula_reference_origin": "sheet_v2",
        "rows": [],
        "grid_rows": [
            ["用户信息", "", "", "", "", "", "", "统计"],
            attendance_columns,
            ["说明", "", "", "", "", "", "", ""],
        ],
        "cell_meta": {"0:0": {"style": {"background_color": "#B4C6E7"}}},
    }

    next_doc, summary = note_sheets_api._sync_registration_rows_to_attendance_document(registration_doc, attendance_doc)

    assert summary["inserted_count"] == 2
    assert next_doc["cell_meta"]["3:0"]["style"]["background_color"] == "#DDEBF7"
    assert next_doc["cell_meta"]["3:6"]["style"]["background_color"] == "#DDEBF7"
    assert next_doc["cell_meta"]["4:0"]["style"]["background_color"] == "#FCE4D6"
    assert next_doc["cell_meta"]["4:6"]["style"]["background_color"] == "#FCE4D6"
    assert "3:7" not in next_doc["cell_meta"]


def test_note_sheet_registration_attendance_sync_derives_order_amount_without_attendance_order_column():
    registration_columns = ["序号", "提交时间", "姓名", "微信昵称", "商户订单号", "订单金额", "用户ID"]
    attendance_columns = ["报名日期", "学号", "姓名", "昵称", "订单金额", "当前应返款"]
    registration_doc = {
        "schema_version": 1,
        "columns": registration_columns,
        "rows": [
            ["124", "2026/5/21 09:46:30", "赵誉博", "赵玉博", "", "620", "u_no_order"],
            ["125", "2026/5/22 09:46:30", "李同学", "同学", "M20260522", "620", "u_with_order"],
        ],
    }
    attendance_doc = {
        "schema_version": 1,
        "columns": attendance_columns,
        "data_start_row": 1,
        "formula_reference_origin": "sheet_v2",
        "rows": [
            ["2026-05-21 09:46", "124", "赵誉博", "赵玉博", "620", "0"],
            ["2026-05-22 09:46", "125", "李同学", "同学", "0", "0"],
        ],
        "grid_rows": [attendance_columns],
    }

    next_doc, summary = note_sheets_api._sync_registration_rows_to_attendance_document(registration_doc, attendance_doc)

    assert summary["repaired_count"] == 2
    assert next_doc["rows"][0][attendance_columns.index("订单金额")] == "0"
    assert next_doc["rows"][1][attendance_columns.index("订单金额")] == "620"


def test_note_sheet_attendance_export_inserts_registration_phone_after_nickname(client, session):
    user = _create_user(session, username="attendance-export-user")
    _grant_feature_access(session, user_id=user.id, feature_key="notes.sheets")
    _override_user(user)

    try:
        workbook_response = client.post("/api/note-sheets/workbooks", json={"title": "修道班7期5阶"})
        workbook_id = workbook_response.json()["id"]
        registration_response = client.post(
            "/api/note-sheets/sheets",
            json={
                "title": "报名表",
                "workbook_id": workbook_id,
                "document_json": {
                    "schema_version": 1,
                    "columns": ["序号", "姓名", "微信昵称", "手机号"],
                    "rows": [
                        ["1", "张三", "三三", "13800138000"],
                        ["2", "李四", "四四", "13900139000"],
                    ],
                },
            },
        )
        assert registration_response.status_code == 200
        attendance_response = client.post(
            "/api/note-sheets/sheets",
            json={
                "title": "考勤表",
                "workbook_id": workbook_id,
                "document_json": {
                    "schema_version": 1,
                    "columns": ["学号", "姓名", "昵称", "当前应返款"],
                    "rows": [
                        ["1", "张三", "三三", "0"],
                        ["2", "李四", "四四", "50"],
                    ],
                },
            },
        )
        assert attendance_response.status_code == 200
        attendance_id = attendance_response.json()["id"]

        response = client.get(
            f"/api/note-sheets/sheets/{attendance_id}/attendance-export",
            params={"workbook_id": workbook_id},
        )

        assert response.status_code == 200
        workbook = load_workbook(io.BytesIO(response.content), data_only=True)
        worksheet = workbook.active
        assert [worksheet.cell(1, column).value for column in range(1, 6)] == [
            "学号",
            "姓名",
            "昵称",
            "手机号",
            "当前应返款",
        ]
        assert [worksheet.cell(2, column).value for column in range(1, 6)] == ["1", "张三", "三三", "13800138000", "0"]
        assert [worksheet.cell(3, column).value for column in range(1, 6)] == ["2", "李四", "四四", "13900139000", "50"]
    finally:
        _clear_user_override()


def test_note_sheet_registration_user_match_can_disable_browser_fallback(client, session, monkeypatch):
    user = _create_user(session, username="note-sheet-user-match-db-only-user")
    _grant_feature_access(session, user_id=user.id, feature_key="notes.sheets")
    _override_user(user)

    target_columns = ["姓名", "微信昵称", "手机号", "错误手机号", "用户ID", "匹配得分"]

    class FakeKqdb:
        def 查找用户(self, names, phones, **kwargs):
            return "", -1

    def fake_remote_lookup(*args, **kwargs):
        raise AssertionError("显式关闭远程回查后不应调用小鹅通")

    monkeypatch.setattr(note_sheets_api, "_load_attendance_kqdb_provider", lambda: (lambda: FakeKqdb()))
    monkeypatch.setattr(note_sheets_api, "_lookup_registration_users_with_remote_browser", fake_remote_lookup)

    try:
        workbook_response = client.post("/api/note-sheets/workbooks", json={"title": "20260509梵呗初阶"})
        workbook_id = workbook_response.json()["id"]
        sheet_response = client.post(
            "/api/note-sheets/sheets",
            json={
                "title": "报名表",
                "workbook_id": workbook_id,
                "document_json": {
                    "schema_version": 1,
                    "columns": target_columns,
                    "rows": [["阿丹", "阿丹", "15326693765", "", "", ""]],
                    "grid_rows": [
                        target_columns,
                        ["", "", "", "", "更新用户匹配", ""],
                        ["阿丹", "阿丹", "15326693765", "", "", ""],
                    ],
                    "data_start_row": 2,
                    "field_row_index": 0,
                },
            },
        )
        sheet_id = sheet_response.json()["id"]

        response = client.post(
            f"/api/note-sheets/sheets/{sheet_id}/registration/update-user-match",
            params={"workbook_id": workbook_id, "use_browser_fallback": "false"},
        )

        assert response.status_code == 200
        payload = response.json()
        updated_row = payload["sheet"]["document_json"]["rows"][0]
        assert updated_row[target_columns.index("用户ID")] == ""
        assert updated_row[target_columns.index("匹配得分")] == "-1"
    finally:
        _clear_user_override()


def test_note_sheet_resource_access_acl_flow(client, session):
    owner = _create_user(session, username="note-sheet-acl-owner")
    editor = _create_user(session, username="note-sheet-acl-editor")
    superuser = _create_user(session, username="note-sheet-acl-superuser", is_superuser=True)

    workbook = WorkbookDocument(
        numeric_id=2,
        title="共享工作簿",
        owner_user_id=owner.id,
        created_by_user_id=owner.id,
        updated_by_user_id=owner.id,
    )
    sheet = SheetDocument(
        numeric_id=4,
        scope="notes",
        owner_type="user",
        owner_key=str(owner.id),
        sheet_key="4",
        title="共享工作表",
        owner_user_id=owner.id,
        created_by_user_id=owner.id,
        updated_by_user_id=owner.id,
        document_json={
            "schema_version": 1,
            "columns": ["姓名", "金额"],
            "rows": [["时秋菊", "620"]],
        },
    )
    session.add(workbook)
    session.add(sheet)
    session.commit()
    session.refresh(workbook)
    session.refresh(sheet)
    session.add(WorkbookSheetLink(workbook_id=workbook.id, sheet_id=sheet.id, order_index=0))
    session.commit()

    _override_user(owner)
    try:
        owner_response = client.get("/api/note-sheets/workbooks/2")
        assert owner_response.status_code == 200
        assert owner_response.json()["access"]["role"] == "manager"

        share_response = client.put(
            "/api/note-sheets/workbooks/2/access",
            json={"grants": [{"subject_type": "anonymous", "role": "viewer"}]},
        )
        assert share_response.status_code == 200
    finally:
        _clear_user_override()

    anonymous_workbook_response = client.get("/api/note-sheets/workbooks/2")
    assert anonymous_workbook_response.status_code == 200
    anonymous_workbook = anonymous_workbook_response.json()
    assert anonymous_workbook["access"]["role"] == "viewer"
    assert [item["id"] for item in anonymous_workbook["sheets"]] == [4]

    anonymous_sheet_response = client.get("/api/note-sheets/sheets/4", params={"workbook_id": 2})
    assert anonymous_sheet_response.status_code == 200
    assert anonymous_sheet_response.json()["access"]["role"] == "viewer"

    anonymous_update_response = client.put(
        "/api/note-sheets/sheets/4",
        json={
            "document_json": {
                "schema_version": 1,
                "columns": ["姓名", "金额"],
                "rows": [["匿名改写", "1"]],
            },
        },
    )
    assert anonymous_update_response.status_code == 403

    session.add(ResourceAccessGrant(
        resource_type="workbook",
        resource_id=str(workbook.numeric_id),
        subject_key=f"user:{editor.id}",
        subject_type="user",
        subject_user_id=editor.id,
        role="editor",
    ))
    session.commit()

    _override_user(editor)
    try:
        inherited_editor_sheet_response = client.get("/api/note-sheets/sheets/4", params={"workbook_id": 2})
        assert inherited_editor_sheet_response.status_code == 200
        inherited_editor_access = inherited_editor_sheet_response.json()["access"]
        assert inherited_editor_access["role"] == "editor"
        assert inherited_editor_access["capabilities"]["can_edit_data"] is True
        assert inherited_editor_access["capabilities"]["can_run_sheet_actions"] is True
    finally:
        _clear_user_override()

    assert client.get("/api/note-sheets/workbooks").status_code == 401

    _override_user(owner)
    try:
        sheet_deny_response = client.put(
            "/api/note-sheets/sheets/4/access",
            json={"grants": [{"subject_type": "anonymous", "role": "deny"}]},
        )
        assert sheet_deny_response.status_code == 200
    finally:
        _clear_user_override()

    denied_sheet_response = client.get("/api/note-sheets/sheets/4", params={"workbook_id": 2})
    assert denied_sheet_response.status_code == 403
    denied_workbook_response = client.get("/api/note-sheets/workbooks/2")
    assert denied_workbook_response.status_code == 200
    assert denied_workbook_response.json()["sheets"] == []

    _override_user(owner)
    try:
        editor_share_response = client.put(
            "/api/note-sheets/sheets/4/access",
            json={
                "grants": [
                    {"subject_type": "anonymous", "role": "deny"},
                    {"subject_type": "user", "username": editor.username, "role": "editor"},
                ],
            },
        )
        assert editor_share_response.status_code == 200
    finally:
        _clear_user_override()

    _override_user(editor)
    try:
        editor_update_response = client.put(
            "/api/note-sheets/sheets/4",
            json={
                "document_json": {
                    "schema_version": 1,
                    "columns": ["姓名", "金额"],
                    "rows": [["编辑者改写", "999"]],
                },
            },
        )
        assert editor_update_response.status_code == 200
        assert editor_update_response.json()["access"]["role"] == "editor"
    finally:
        _clear_user_override()

    _override_user(superuser)
    try:
        superuser_workbook_response = client.get("/api/note-sheets/workbooks/2")
        assert superuser_workbook_response.status_code == 200
        assert superuser_workbook_response.json()["access"]["role"] == "manager"

        superuser_response = client.get("/api/note-sheets/sheets/4")
        assert superuser_response.status_code == 200
        assert superuser_response.json()["access"]["role"] == "manager"
    finally:
        _clear_user_override()


def test_note_sheet_resource_creator_is_implicit_manager_when_owner_missing(client, session):
    creator = _create_user(session, username="note-sheet-creator-principal")

    workbook = WorkbookDocument(
        numeric_id=2,
        title="迁移工作簿",
        owner_user_id=None,
        created_by_user_id=creator.id,
        updated_by_user_id=creator.id,
    )
    sheet = SheetDocument(
        numeric_id=4,
        scope="notes",
        owner_type="user",
        owner_key=str(creator.id),
        sheet_key="4",
        title="迁移工作表",
        owner_user_id=None,
        created_by_user_id=creator.id,
        updated_by_user_id=creator.id,
        document_json={
            "schema_version": 1,
            "columns": ["姓名"],
            "rows": [["时秋菊"]],
        },
    )
    session.add(workbook)
    session.add(sheet)
    session.commit()
    session.refresh(workbook)
    session.refresh(sheet)
    session.add(WorkbookSheetLink(workbook_id=workbook.id, sheet_id=sheet.id, order_index=0))
    session.commit()

    _override_user(creator)
    try:
        workbook_response = client.get("/api/note-sheets/workbooks/2")
        assert workbook_response.status_code == 200
        assert workbook_response.json()["access"]["role"] == "manager"

        sheet_response = client.get("/api/note-sheets/sheets/4")
        assert sheet_response.status_code == 200
        assert sheet_response.json()["access"]["role"] == "manager"
    finally:
        _clear_user_override()


def test_sheet_detail_only_lists_accessible_parent_workbooks(client, session):
    owner = _create_user(session, username="note-sheet-parent-owner")
    workbook = WorkbookDocument(
        numeric_id=2,
        title="私有工作簿",
        owner_user_id=owner.id,
        created_by_user_id=owner.id,
        updated_by_user_id=owner.id,
    )
    sheet = SheetDocument(
        numeric_id=4,
        scope="notes",
        owner_type="user",
        owner_key=str(owner.id),
        sheet_key="4",
        title="独立共享工作表",
        owner_user_id=owner.id,
        created_by_user_id=owner.id,
        updated_by_user_id=owner.id,
        document_json={
            "schema_version": 1,
            "columns": ["姓名"],
            "rows": [["时秋菊"]],
        },
    )
    session.add(workbook)
    session.add(sheet)
    session.commit()
    session.refresh(workbook)
    session.refresh(sheet)
    session.add(WorkbookSheetLink(workbook_id=workbook.id, sheet_id=sheet.id, order_index=0))
    session.add(ResourceAccessGrant(
        resource_type="sheet",
        resource_id=str(sheet.numeric_id),
        subject_key="anonymous",
        subject_type="anonymous",
        role="viewer",
    ))
    session.commit()

    sheet_response = client.get("/api/note-sheets/sheets/4")
    assert sheet_response.status_code == 200
    sheet_payload = sheet_response.json()
    assert sheet_payload["parent_workbook_id"] == 2
    assert sheet_payload["workbook_items"] == []
    assert client.get("/api/note-sheets/workbooks/2").status_code == 403

    session.add(ResourceAccessGrant(
        resource_type="workbook",
        resource_id=str(workbook.numeric_id),
        subject_key="anonymous",
        subject_type="anonymous",
        role="viewer",
    ))
    session.commit()

    shared_parent_response = client.get("/api/note-sheets/sheets/4")
    assert shared_parent_response.status_code == 200
    shared_parent_payload = shared_parent_response.json()
    assert shared_parent_payload["parent_workbook_id"] == 2
    assert shared_parent_payload["workbook_items"] == [{"id": 2, "title": "私有工作簿"}]


def test_attendance_questionnaire_sheet_allows_anonymous_status_column_edit(client, session, monkeypatch):
    owner = _create_user(session, username="note-sheet-public-status-owner")
    sheet = SheetDocument(
        numeric_id=5,
        scope="notes",
        owner_type="attendance_questionnaire",
        owner_key="wjx-data",
        sheet_key="data",
        title="问卷数据",
        owner_user_id=owner.id,
        created_by_user_id=owner.id,
        updated_by_user_id=owner.id,
        document_json={
            "schema_version": 1,
            "columns": [
                "序号",
                "提交时间",
                "来源",
                "课程",
                "考勤负责人",
                "学号",
                "姓名",
                "修正需求",
                "补充说明",
                "处理状态",
            ],
            "rows": [
                ["651", "2026/4/20", "微信", "课程A", "", "5组6号", "陈香米", "第一堂视频课", "-", "确认中"],
                ["650", "2026/4/19", "微信", "课程B", "", "4--49", "曾玉清", "请问作业", "其实也没什么", "已处理"],
            ],
            "view_settings": {
                "pagination": {
                    "enabled": True,
                    "page_size": 100,
                },
            },
            "column_configs": {
                "姓名": {"display_mode": "single_line"},
            },
            "entity_rows": [
                {"id": "field_1", "kind": "field"},
            ],
        },
    )
    session.add(sheet)
    entry = AttendanceWjxDataEntry(
        activity_id="264266843",
        seq=651,
        submitted_at_text="2026/4/20",
        source="微信",
        course_name="课程A",
        student_id_text="5组6号",
        student_name="陈香米",
        correction_request="第一堂视频课",
        extra_note="-",
        process_status="",
        process_note="",
        synced_at=1713426373.0,
        created_at=1713426373.0,
        updated_at=1713426373.0,
    )
    session.add(entry)
    session.commit()
    session.refresh(sheet)
    session.add(ResourceAccessGrant(
        resource_type="sheet",
        resource_id=str(sheet.numeric_id),
        subject_key="anonymous",
        subject_type="anonymous",
        role="viewer",
    ))
    session.commit()
    broadcasts: list[tuple[str, dict]] = []

    async def fake_broadcast(room: str, message: dict) -> None:
        broadcasts.append((room, message))

    monkeypatch.setattr(note_sheets_api.ws_manager, "broadcast", fake_broadcast)

    detail_response = client.get("/api/note-sheets/sheets/5")
    assert detail_response.status_code == 200
    detail = detail_response.json()
    assert detail["access"]["role"] == "viewer"
    assert detail["access"]["capabilities"]["can_edit_data"] is False
    assert detail["access"]["capabilities"]["editable_data_columns"] == [9]
    session.refresh(entry)
    assert entry.process_status == ""
    assert entry.process_note == ""

    rows = detail["document_json"]["rows"]
    rows[0][9] = "用户d问题，已修正"
    page_patch = {key: value for key, value in detail["pagination"].items() if value is not None}
    update_response = client.put(
        "/api/note-sheets/sheets/5",
        json={
            "title": "不应被部分编辑保存改名",
            "document_json": {
                **detail["document_json"],
                "column_configs": {
                    "序号": {"display_mode": "single_line"},
                },
                "rows": rows,
            },
            "page_patch": page_patch,
        },
    )
    assert update_response.status_code == 200, update_response.json()
    session.refresh(sheet)
    session.refresh(entry)
    assert sheet.title == "问卷数据"
    assert sheet.document_json["rows"][0][9] == "用户d问题，已修正"
    assert sheet.document_json["rows"][0][7] == "第一堂视频课"
    assert entry.process_status == "用户d问题，已修正"
    assert entry.process_note == "用户d问题，已修正"
    assert broadcasts[-1][0] == "resource:sheet:5"
    assert broadcasts[-1][1]["type"] == "resource-updated"
    assert broadcasts[-1][1]["resource_type"] == "sheet"
    assert broadcasts[-1][1]["resource_id"] == "5"
    assert broadcasts[-1][1]["version"] == sheet.version

    broadcasts_before_cell_patch = len(broadcasts)
    cell_patch_base_version = sheet.version
    cell_patch_response = client.patch(
        "/api/note-sheets/sheets/5/cells",
        json={
            "base_version": cell_patch_base_version,
            "operations": [
                {"row_index": 0, "column_index": 9, "value": "单元格patch已处理"},
            ],
        },
    )
    assert cell_patch_response.status_code == 200, cell_patch_response.json()
    cell_patch_payload = cell_patch_response.json()
    assert cell_patch_payload["updated_cell_count"] == 1
    session.refresh(sheet)
    session.refresh(entry)
    assert sheet.document_json["rows"][0][9] == "单元格patch已处理"
    assert entry.process_status == "单元格patch已处理"
    assert entry.process_note == "单元格patch已处理"
    assert len(broadcasts) == broadcasts_before_cell_patch + 1
    assert broadcasts[-1][0] == "resource:sheet:5"
    assert broadcasts[-1][1]["version"] == sheet.version

    stale_cell_patch_response = client.patch(
        "/api/note-sheets/sheets/5/cells",
        json={
            "base_version": cell_patch_base_version,
            "operations": [
                {"row_index": 0, "column_index": 9, "value": "过期单元格patch"},
            ],
        },
    )
    assert stale_cell_patch_response.status_code == 409
    session.refresh(sheet)
    session.refresh(entry)
    assert sheet.document_json["rows"][0][9] == "单元格patch已处理"
    assert entry.process_status == "单元格patch已处理"
    assert len(broadcasts) == broadcasts_before_cell_patch + 1

    op_patch_response = client.post(
        "/api/note-sheets/sheets/5/patch",
        json={
            "base_version": sheet.version,
            "ops": [
                {"op": "set-cell-value", "row_index": 0, "column_index": 9, "value": ""},
            ],
        },
    )
    assert op_patch_response.status_code == 200, op_patch_response.json()
    op_patch_payload = op_patch_response.json()
    assert op_patch_payload["updated_cell_count"] == 1
    session.refresh(sheet)
    session.refresh(entry)
    assert sheet.document_json["rows"][0][9] == ""
    assert entry.process_status == ""
    assert entry.process_note == ""

    forbidden_cell_patch_response = client.patch(
        "/api/note-sheets/sheets/5/cells",
        json={
            "operations": [
                {"row_index": 0, "column_index": 7, "value": "越权patch"},
            ],
        },
    )
    assert forbidden_cell_patch_response.status_code == 403

    forbidden_op_patch_response = client.post(
        "/api/note-sheets/sheets/5/patch",
        json={
            "base_version": sheet.version,
            "ops": [
                {"op": "set-column-width", "column_index": 9, "width": 180},
            ],
        },
    )
    assert forbidden_op_patch_response.status_code == 403

    forbidden_config_response = client.put(
        "/api/note-sheets/sheets/5",
        json={
            "document_json": {
                **detail["document_json"],
                "column_widths": [120] * len(detail["document_json"]["columns"]),
            },
            "page_patch": page_patch,
        },
    )
    assert forbidden_config_response.status_code == 403

    forbidden_identity_response = client.put(
        "/api/note-sheets/sheets/5",
        json={
            "document_json": {
                **detail["document_json"],
                "row_ids": ["tampered-row-1", "tampered-row-2"],
                "column_ids": [f"tampered-col-{index}" for index, _ in enumerate(detail["document_json"]["columns"])],
            },
            "page_patch": page_patch,
        },
    )
    assert forbidden_identity_response.status_code == 403

    rows[0][7] = "越权修改"
    forbidden_response = client.put(
        "/api/note-sheets/sheets/5",
        json={
            "document_json": {
                **detail["document_json"],
                "rows": rows,
            },
            "page_patch": page_patch,
        },
    )
    assert forbidden_response.status_code == 403


def test_note_sheet_operation_patch_applies_atomically_and_checks_version(client, session, monkeypatch):
    owner = _create_user(session, username="note-sheet-op-patch-owner")
    sheet = SheetDocument(
        numeric_id=61,
        scope="notes",
        owner_type="note_sheet",
        owner_key="op-patch",
        sheet_key="main",
        title="操作 Patch",
        owner_user_id=owner.id,
        created_by_user_id=owner.id,
        updated_by_user_id=owner.id,
        document_json={
            "schema_version": 1,
            "columns": ["姓名", "状态", "备注"],
            "rows": [["张三", "待处理", ""]],
            "grid_rows": [["姓名", "状态", "备注"], ["张三", "待处理", ""]],
            "data_start_row": 1,
            "field_row_index": 0,
            "column_widths": [100, 100, 100],
            "column_configs": {},
            "merged_cells": [],
        },
    )
    session.add(sheet)
    session.commit()
    _override_user(owner)
    broadcasts: list[tuple[str, dict]] = []

    async def fake_broadcast(room: str, message: dict) -> None:
        broadcasts.append((room, message))

    monkeypatch.setattr(note_sheets_api.ws_manager, "broadcast", fake_broadcast)

    try:
        response = client.post(
            "/api/note-sheets/sheets/61/patch",
            json={
                "base_version": sheet.version,
                "ops": [
                    {"op": "set-cell-value", "row_index": 0, "column_index": 1, "value": ""},
                    {"op": "set-cell-meta", "row_index": 0, "column_index": 1, "meta": {"cell_type": "rich_text"}},
                    {"op": "set-column-width", "column_index": 1, "width": 168},
                    {"op": "set-column-hidden", "column_index": 2, "hidden": True},
                    {"op": "set-column-config", "column_index": 1, "config": {"display_mode": "single_line"}},
                    {"op": "merge-cells", "row": 0, "col": 0, "rowspan": 1, "colspan": 2},
                ],
            },
        )
        assert response.status_code == 200, response.json()
        payload = response.json()
        assert payload["applied_op_count"] == 6
        assert payload["updated_cell_count"] == 1
        session.refresh(sheet)
        assert sheet.version == 2
        document = sheet.document_json
        assert document["rows"][0][1] == ""
        assert document["cell_meta"]["1:1"] == {"cell_type": "rich_text"}
        assert document["column_widths"][1] == 168
        assert document["column_configs"]["备注"]["hidden"] is True
        assert document["column_configs"]["状态"]["display_mode"] == "single_line"
        assert document["merged_cells"] == [{"row": 0, "col": 0, "rowspan": 1, "colspan": 2}]
        assert broadcasts[-1][0] == "resource:sheet:61"
        assert broadcasts[-1][1]["type"] == "resource-updated"
        assert broadcasts[-1][1]["resource_type"] == "sheet"
        assert broadcasts[-1][1]["resource_id"] == "61"
        assert broadcasts[-1][1]["version"] == 2

        stale_response = client.post(
            "/api/note-sheets/sheets/61/patch",
            json={
                "base_version": 1,
                "ops": [
                    {"op": "set-cell-value", "row_index": 0, "column_index": 1, "value": "旧版本覆盖"},
                ],
            },
        )
        assert stale_response.status_code == 409
        session.refresh(sheet)
        assert sheet.document_json["rows"][0][1] == ""

        atomic_response = client.post(
            "/api/note-sheets/sheets/61/patch",
            json={
                "base_version": sheet.version,
                "ops": [
                    {"op": "set-cell-value", "row_index": 0, "column_index": 1, "value": "不应落库"},
                    {"op": "set-column-width", "column_index": 99, "width": 120},
                ],
            },
        )
        assert atomic_response.status_code == 400
        session.refresh(sheet)
        assert sheet.document_json["rows"][0][1] == ""

        unmerge_response = client.post(
            "/api/note-sheets/sheets/61/patch",
            json={
                "base_version": sheet.version,
                "ops": [
                    {"op": "unmerge-cells", "row": 0, "col": 1},
                ],
            },
        )
        assert unmerge_response.status_code == 200, unmerge_response.json()
        session.refresh(sheet)
        assert sheet.document_json["merged_cells"] == []
    finally:
        _clear_user_override()


def test_note_sheet_get_backfills_stable_row_and_column_ids(client, session):
    owner = _create_user(session, username="note-sheet-identity-owner")
    sheet = SheetDocument(
        numeric_id=62,
        scope="notes",
        owner_type="note_sheet",
        owner_key="identity",
        sheet_key="main",
        title="旧表身份补齐",
        owner_user_id=owner.id,
        created_by_user_id=owner.id,
        updated_by_user_id=owner.id,
        document_json={
            "schema_version": 1,
            "columns": ["姓名", "状态"],
            "rows": [["张三", "待处理"], ["李四", "已处理"]],
            "grid_rows": [["姓名", "状态"], ["张三", "待处理"], ["李四", "已处理"]],
            "data_start_row": 1,
            "field_row_index": 0,
        },
    )
    session.add(sheet)
    session.commit()
    _override_user(owner)

    try:
        response = client.get("/api/note-sheets/sheets/62")
        assert response.status_code == 200, response.json()
        document = response.json()["document_json"]
        assert len(document["row_ids"]) == 2
        assert len(document["column_ids"]) == 2
        assert all(str(row_id).startswith("row_") for row_id in document["row_ids"])
        assert all(str(column_id).startswith("col_") for column_id in document["column_ids"])

        session.refresh(sheet)
        assert sheet.document_json["row_ids"] == document["row_ids"]
        assert sheet.document_json["column_ids"] == document["column_ids"]
    finally:
        _clear_user_override()


def test_note_sheet_structure_patch_uses_stable_ids_and_keeps_metadata_aligned(client, session):
    owner = _create_user(session, username="note-sheet-structure-owner")
    sheet = SheetDocument(
        numeric_id=63,
        scope="notes",
        owner_type="note_sheet",
        owner_key="structure-patch",
        sheet_key="main",
        title="结构 Patch",
        owner_user_id=owner.id,
        created_by_user_id=owner.id,
        updated_by_user_id=owner.id,
        document_json={
            "schema_version": 1,
            "columns": ["姓名", "状态", "备注"],
            "column_ids": ["col_name", "col_status", "col_note"],
            "rows": [["张三", "待处理", "a"], ["李四", "已处理", "b"]],
            "row_ids": ["row_zhang", "row_li"],
            "grid_rows": [["姓名", "状态", "备注"], ["张三", "待处理", "a"], ["李四", "已处理", "b"]],
            "data_start_row": 1,
            "field_row_index": 0,
            "cell_meta": {"1:1": {"cell_type": "rich_text"}, "2:2": {"cell_type": "note"}},
            "column_widths": [80, 90, 100],
            "column_configs": {"状态": {"display_mode": "single_line"}, "备注": {"hidden": True}},
            "merged_cells": [],
        },
    )
    session.add(sheet)
    session.commit()
    _override_user(owner)

    try:
        response = client.post(
            "/api/note-sheets/sheets/63/patch",
            json={
                "base_version": sheet.version,
                "ops": [
                    {"op": "insert-row", "after_row_id": "row_zhang", "row_id": "row_wang", "row": ["王五", "新建", "c"]},
                    {"op": "delete-row", "row_id": "row_li"},
                    {
                        "op": "insert-column",
                        "after_column_id": "col_name",
                        "column_id": "col_priority",
                        "column": {"header": "优先级", "width": 144, "config": {"value_type": "number"}},
                    },
                    {"op": "delete-column", "column_id": "col_note"},
                ],
            },
        )
        assert response.status_code == 200, response.json()
        session.refresh(sheet)
        document = sheet.document_json
        assert sheet.version == 2
        assert document["row_ids"] == ["row_zhang", "row_wang"]
        assert document["column_ids"] == ["col_name", "col_priority", "col_status"]
        assert document["columns"] == ["姓名", "优先级", "状态"]
        assert document["rows"] == [["张三", "", "待处理"], ["王五", "", "新建"]]
        assert document["grid_rows"] == [["姓名", "优先级", "状态"], ["张三", "", "待处理"], ["王五", "", "新建"]]
        assert document["cell_meta"] == {"1:2": {"cell_type": "rich_text"}}
        assert document["column_widths"] == [80, 144, 90]
        assert document["column_configs"]["优先级"] == {"value_type": "number"}
        assert document["column_configs"]["状态"] == {"display_mode": "single_line"}
        assert "备注" not in document["column_configs"]

        stale_response = client.post(
            "/api/note-sheets/sheets/63/patch",
            json={
                "base_version": 1,
                "ops": [{"op": "delete-row", "row_id": "row_zhang"}],
            },
        )
        assert stale_response.status_code == 409
        session.refresh(sheet)
        assert sheet.document_json["row_ids"] == ["row_zhang", "row_wang"]
    finally:
        _clear_user_override()


def test_attendance_questionnaire_sheet_get_backfills_course_links(client, session):
    owner = _create_user(session, username="note-sheet-questionnaire-link-owner")
    source_sheet = SheetDocument(
        numeric_id=4,
        scope="notes",
        owner_type="note_sheet",
        owner_key="attendance-summary",
        sheet_key="courses",
        title="课程",
        document_json={
            "schema_version": 1,
            "columns": [
                "课程类型",
                "课程名称",
                "在线考勤表",
                "考勤负责人",
                "课次链接",
                "打卡链接",
                "备注",
                "课程开始日期",
                "课程结束日期",
                "考勤实际完成结点",
            ],
                "rows": [
                    [
                        "禅宗5阶",
                        "20260412禅宗12期一阶",
                        {
                            "value": "20260412禅宗12期一阶",
                            "link": {
                                "url": "https://www.kdocs.cn/l/nianzhu12",
                            },
                        },
                        "陈坤泽",
                        "",
                        "",
                        "",
                    "04/12",
                    "06/14",
                    "",
                ],
            ],
            },
        )
    data_sheet = SheetDocument(
        numeric_id=5,
        scope="notes",
        owner_type="attendance_questionnaire",
        owner_key="wjx-data",
        sheet_key="data",
        title="问卷数据",
        owner_user_id=owner.id,
        created_by_user_id=owner.id,
        updated_by_user_id=owner.id,
        version=1,
        document_json={
            "schema_version": 1,
            "columns": [
                "序号",
                "提交时间",
                "来源",
                "课程",
                "学号",
                "姓名",
                "修正需求",
                "补充说明",
                "处理状态",
            ],
            "rows": [
                [
                    "651",
                    "2026/4/20 07:51:46",
                    "微信",
                    "20260412禅宗12期一阶",
                    "5组6号",
                    "陈香米",
                    "第一堂视频课",
                    "-",
                    "确认中",
                ],
            ],
            "grid_rows": [
                [
                    "序号",
                    "提交时间",
                    "来源",
                    "课程",
                    "学号",
                    "姓名",
                    "修正需求",
                    "补充说明",
                    "处理状态",
                ],
                [
                    "651",
                    "2026/4/20 07:51:46",
                    "微信",
                    "20260412禅宗12期一阶",
                    "5组6号",
                    "陈香米",
                    "第一堂视频课",
                    "-",
                    "确认中",
                ],
            ],
            "data_start_row": 1,
            "field_row_index": 0,
            "column_configs": {
                "AI初判": {"display_mode": "single_line"},
            },
            "cell_meta": {
                "0:3": {"link": {"url": "https://example.com/legacy-data-row-coordinate"}},
                "1:3": {"link": {"url": "https://example.com/stale-next-row-link"}},
            },
        },
    )
    session.add(source_sheet)
    session.add(data_sheet)
    session.commit()
    session.refresh(data_sheet)

    _override_user(owner)
    try:
        response = client.get("/api/note-sheets/sheets/5")
        assert response.status_code == 200
        detail = response.json()
        assert detail["document_json"]["columns"] == [
            "序号",
            "提交时间",
            "来源",
            "课程",
            "考勤负责人",
            "学号",
            "姓名",
            "修正需求",
            "补充说明",
            "处理状态",
            "AI初判",
        ]
        assert detail["document_json"]["rows"][0][4] == ""
        assert detail["document_json"]["rows"][0][5] == "5组6号"
        assert detail["document_json"]["grid_rows"][0] == detail["document_json"]["columns"]
        assert detail["document_json"]["grid_rows"][1][4] == ""
        assert detail["document_json"]["grid_rows"][1][5] == "5组6号"
        assert detail["document_json"]["column_configs"]["AI初判"]["display_mode"] == "single_line"
        assert "0:3" not in detail["document_json"]["cell_meta"]
        assert detail["document_json"]["rows"][0][3] == {
            "value": "20260412禅宗12期一阶",
            "link": {"url": "https://www.kdocs.cn/l/nianzhu12"},
        }
        assert detail["document_json"]["grid_rows"][1][3] == detail["document_json"]["rows"][0][3]

        persisted = session.exec(select(SheetDocument).where(SheetDocument.numeric_id == 5)).first()
        assert persisted is not None
        assert persisted.version == 2
        assert persisted.document_json["rows"][0][4] == ""
        assert persisted.document_json["column_configs"]["AI初判"]["display_mode"] == "single_line"
        assert persisted.document_json["rows"][0][3]["link"]["url"] == "https://www.kdocs.cn/l/nianzhu12"

        second_response = client.get("/api/note-sheets/sheets/5")
        assert second_response.status_code == 200
        second_detail = second_response.json()
        assert second_detail["version"] == 2
        persisted = session.exec(select(SheetDocument).where(SheetDocument.numeric_id == 5)).first()
        assert persisted is not None
        assert persisted.version == 2
    finally:
        _clear_user_override()


def test_attendance_questionnaire_sheet_get_reconciles_rows_from_entries(client, session):
    from backend.api import attendance as attendance_api

    owner = _create_user(session, username="note-sheet-questionnaire-reconcile-owner")
    source_sheet = SheetDocument(
        numeric_id=4,
        scope="notes",
        owner_type="note_sheet",
        owner_key="attendance-summary",
        sheet_key="courses",
        title="课程",
        document_json={
            "schema_version": 1,
            "columns": ["课程类型", "课程名称", "在线考勤表", "考勤负责人"],
            "rows": [
                [
                    "禅宗二阶",
                    "禅宗11期二阶",
                    {
                        "value": "20260412禅宗11期二阶",
                        "link": {"url": "https://www.kdocs.cn/l/zen11-stage2"},
                    },
                    "陈坤泽, 白玄度",
                ],
                [
                    "禅宗一阶",
                    "禅宗12期一阶",
                    {
                        "value": "20260412禅宗12期一阶",
                        "link": {"url": "https://www.kdocs.cn/l/zen12-stage1"},
                    },
                    "陈坤泽, Judy chen",
                ],
            ],
        },
    )
    data_sheet = SheetDocument(
        numeric_id=5,
        scope="notes",
        owner_type="attendance_questionnaire",
        owner_key="wjx-data",
        sheet_key="data",
        title="问卷数据",
        owner_user_id=owner.id,
        created_by_user_id=owner.id,
        updated_by_user_id=owner.id,
        version=1,
        document_json={
            "schema_version": 1,
            "columns": [
                "序号",
                "提交时间",
                "来源",
                "课程",
                "考勤负责人",
                "学号",
                "姓名",
                "修正需求",
                "补充说明",
                "处理状态",
                "AI初判",
            ],
            "rows": [[
                "683",
                "2026/06/14 12:31:47",
                "采集系统",
                {"value": "20260412禅宗12期一阶", "link": {"url": "https://www.kdocs.cn/l/zen12-stage1"}},
                "Judy chen",
                "306",
                "韩羽婷",
                "旧问题",
                "",
                "人工备注保留",
                "",
            ]],
            "grid_rows": [
                [
                    "序号",
                    "提交时间",
                    "来源",
                    "课程",
                    "考勤负责人",
                    "学号",
                    "姓名",
                    "修正需求",
                    "补充说明",
                    "处理状态",
                    "AI初判",
                ],
                [
                    "683",
                    "2026/06/14 12:31:47",
                    "采集系统",
                    {"value": "20260412禅宗12期一阶", "link": {"url": "https://www.kdocs.cn/l/zen12-stage1"}},
                    "Judy chen",
                    "306",
                    "韩羽婷",
                    "旧问题",
                    "",
                    "人工备注保留",
                    "",
                ],
            ],
            "data_start_row": 1,
            "field_row_index": 0,
        },
    )
    entry = AttendanceWjxDataEntry(
        activity_id=attendance_api.LOCAL_FEEDBACK_ACTIVITY_ID,
        seq=683,
        submitted_at_text="2026/06/14 12:31:47",
        source="采集系统",
        source_detail="CodeYun反馈表",
        course_name="20260412禅宗11期二阶",
        student_id_text="306",
        student_name="韩羽婷",
        correction_request="第九周5阿含经 完成当堂学习",
        extra_note="",
        process_status="",
    )
    session.add(source_sheet)
    session.add(data_sheet)
    session.add(entry)
    session.commit()
    session.refresh(data_sheet)

    _override_user(owner)
    try:
        response = client.get("/api/note-sheets/sheets/5")
        assert response.status_code == 200
        detail = response.json()
        row = detail["document_json"]["rows"][0]
        assert row[3] == {
            "value": "20260412禅宗11期二阶",
            "link": {"url": "https://www.kdocs.cn/l/zen11-stage2"},
        }
        assert row[4] == "白玄度"
        assert row[7] == "第九周5阿含经 完成当堂学习"
        assert row[9] == "人工备注保留"

        persisted = session.exec(select(SheetDocument).where(SheetDocument.numeric_id == 5)).first()
        assert persisted is not None
        assert persisted.version == 2
        assert persisted.document_json["rows"][0][3]["value"] == "20260412禅宗11期二阶"
        assert persisted.document_json["rows"][0][4] == "白玄度"
    finally:
        _clear_user_override()


def test_attendance_questionnaire_sheet_normalize_preserves_frontend_default_ai_display():
    from backend.api.attendance import _normalize_attendance_wjx_sheet_document

    normalized = _normalize_attendance_wjx_sheet_document({
        "schema_version": 1,
        "columns": [
            "序号",
            "提交时间",
            "来源",
            "课程",
            "考勤负责人",
            "学号",
            "姓名",
            "修正需求",
            "补充说明",
            "处理状态",
            "AI初判",
        ],
        "rows": [],
        "column_configs": {},
    })
    ai_config = normalized["column_configs"].get("AI初判") or {}
    assert ai_config.get("display_mode") is None

    legacy_normalized = _normalize_attendance_wjx_sheet_document({
        "schema_version": 1,
        "columns": [
            "序号",
            "提交时间",
            "来源",
            "课程",
            "学号",
            "姓名",
            "修正需求",
            "补充说明",
            "处理状态",
        ],
        "rows": [],
        "column_configs": {},
    })
    assert legacy_normalized["column_configs"]["AI初判"]["display_mode"] == "wrap"


def test_note_sheet_rejects_default_blank_overwrite(client, session):
    user = _create_user(session, username="note-sheet-blank-guard-user")
    _grant_feature_access(session, user_id=user.id, feature_key="notes.sheets")
    _override_user(user)

    try:
        create_sheet_response = client.post(
            "/api/note-sheets/sheets",
            json={
                "title": "课程",
                "document_json": {
                    "schema_version": 1,
                    "columns": ["课程类型", "课程名称", "在线考勤表"],
                    "rows": [["", "", "20250106念住闯关"]],
                },
            },
        )
        assert create_sheet_response.status_code == 200
        sheet = create_sheet_response.json()

        overwrite_response = client.put(
            f"/api/note-sheets/sheets/{sheet['id']}",
            json={
                "document_json": {
                    "schema_version": 1,
                    "columns": ["列1", "列2", "列3"],
                    "rows": [],
                },
            },
        )
        assert overwrite_response.status_code == 409
        assert overwrite_response.json()["detail"] == "拒绝使用默认空表覆盖已有表格数据"

        persisted = session.exec(select(SheetDocument).where(SheetDocument.numeric_id == sheet["id"])).one()
        assert persisted.document_json["columns"] == ["课程类型", "课程名称", "在线考勤表"]
        assert persisted.document_json["rows"] == [["", "", "20250106念住闯关"]]
    finally:
        _clear_user_override()


def test_note_sheet_sort_shifts_moved_formula_references(client, session):
    user = _create_user(session, username="note-sheet-formula-sort-user")
    _grant_feature_access(session, user_id=user.id, feature_key="notes.sheets")
    _override_user(user)

    try:
        create_sheet_response = client.post(
            "/api/note-sheets/sheets",
            json={
                "title": "公式排序表",
                "document_json": {
                    "schema_version": 1,
                    "columns": ["名称", "课程", "公式"],
                    "rows": [
                        ["b", "课程乙", "=B1 + B2"],
                        ["a", "课程甲", "=$B2 + B$1 + $B$1"],
                    ],
                },
            },
        )
        assert create_sheet_response.status_code == 200
        sheet = create_sheet_response.json()

        sort_response = client.post(
            f"/api/note-sheets/sheets/{sheet['id']}/sort",
            json={"column_index": 0, "direction": "asc"},
        )
        assert sort_response.status_code == 200

        assert sort_response.json()["document_json"]["rows"] == [
            ["a", "课程甲", "=$B1 + B$2 + $B$2"],
            ["b", "课程乙", "=B2 + B1"],
        ]
    finally:
        _clear_user_override()


def test_note_sheet_sort_date_formula_uses_typed_date_value(client, session):
    user = _create_user(session, username="note-sheet-date-formula-sort-user")
    _grant_feature_access(session, user_id=user.id, feature_key="notes.sheets")
    _override_user(user)

    try:
        create_sheet_response = client.post(
            "/api/note-sheets/sheets",
            json={
                "title": "日期公式排序表",
                "document_json": {
                    "schema_version": 1,
                    "columns": ["在线考勤表", "课程开始日期"],
                    "rows": [
                        ["20260409梵呗增益", '=DATE_PARSE(A1, "yyyymmdd")'],
                        ["20250106念住闯关", '=DATE_PARSE(A2, "yyyymmdd")'],
                        ["20260301禅宗46期五阶", '=DATE_PARSE(A3, "yyyymmdd")'],
                    ],
                    "column_configs": {
                        "课程开始日期": {
                            "value_type": "date",
                            "display_format": "m/d",
                        },
                    },
                },
            },
        )
        assert create_sheet_response.status_code == 200
        sheet = create_sheet_response.json()

        sort_response = client.post(
            f"/api/note-sheets/sheets/{sheet['id']}/sort",
            json={"column_index": 1, "direction": "asc"},
        )
        assert sort_response.status_code == 200

        assert sort_response.json()["document_json"]["rows"] == [
            ["20250106念住闯关", '=DATE_PARSE(A1, "yyyymmdd")'],
            ["20260301禅宗46期五阶", '=DATE_PARSE(A2, "yyyymmdd")'],
            ["20260409梵呗增益", '=DATE_PARSE(A3, "yyyymmdd")'],
        ]
    finally:
        _clear_user_override()


def test_note_sheet_sort_date_formula_uses_sheet_v2_addresses(client, session):
    user = _create_user(session, username="note-sheet-sheet-v2-formula-sort-user")
    _grant_feature_access(session, user_id=user.id, feature_key="notes.sheets")
    _override_user(user)

    try:
        columns = ["在线考勤表", "课程开始日期"]
        rows = [
            ["20260409梵呗增益", '=DATE_PARSE(A2, "yyyymmdd")'],
            ["20250106念住闯关", '=DATE_PARSE(A3, "yyyymmdd")'],
            ["20260301禅宗46期五阶", '=DATE_PARSE(A4, "yyyymmdd")'],
        ]
        create_sheet_response = client.post(
            "/api/note-sheets/sheets",
            json={
                "title": "工作表地址公式排序表",
                "document_json": {
                    "schema_version": 1,
                    "columns": columns,
                    "rows": rows,
                    "grid_rows": [columns, *rows],
                    "data_start_row": 1,
                    "field_row_index": 0,
                    "formula_reference_origin": "sheet_v2",
                    "column_configs": {
                        "课程开始日期": {
                            "value_type": "date",
                            "display_format": "m/d",
                        },
                    },
                },
            },
        )
        assert create_sheet_response.status_code == 200
        sheet = create_sheet_response.json()

        sort_response = client.post(
            f"/api/note-sheets/sheets/{sheet['id']}/sort",
            json={"column_index": 1, "direction": "asc"},
        )
        assert sort_response.status_code == 200

        assert sort_response.json()["document_json"]["rows"] == [
            ["20250106念住闯关", '=DATE_PARSE(A2, "yyyymmdd")'],
            ["20260301禅宗46期五阶", '=DATE_PARSE(A3, "yyyymmdd")'],
            ["20260409梵呗增益", '=DATE_PARSE(A4, "yyyymmdd")'],
        ]
    finally:
        _clear_user_override()


def test_note_sheet_sort_date_offset_formula_uses_referenced_value(client, session):
    user = _create_user(session, username="note-sheet-date-offset-sort-user")
    _grant_feature_access(session, user_id=user.id, feature_key="notes.sheets")
    _override_user(user)

    try:
        create_sheet_response = client.post(
            "/api/note-sheets/sheets",
            json={
                "title": "日期偏移排序表",
                "document_json": {
                    "schema_version": 1,
                    "columns": ["在线考勤表", "课程开始日期", "课程结束日期"],
                    "rows": [
                        ["20260409梵呗增益", '=DATE_PARSE(A1, "yyyymmdd")', "=B1+5.5"],
                        ["20250106念住闯关", '=DATE_PARSE(A2, "yyyymmdd")', "=B2+5.5"],
                        ["20260301禅宗46期五阶", '=DATE_PARSE(A3, "yyyymmdd")', "=B3+5.5"],
                    ],
                    "column_configs": {
                        "课程开始日期": {"value_type": "date", "display_format": "m/d"},
                        "课程结束日期": {"value_type": "date", "display_format": "m/d"},
                    },
                },
            },
        )
        assert create_sheet_response.status_code == 200
        sheet = create_sheet_response.json()

        sort_response = client.post(
            f"/api/note-sheets/sheets/{sheet['id']}/sort",
            json={"column_index": 2, "direction": "asc"},
        )
        assert sort_response.status_code == 200

        assert sort_response.json()["document_json"]["rows"] == [
            ["20250106念住闯关", '=DATE_PARSE(A1, "yyyymmdd")', "=B1+5.5"],
            ["20260301禅宗46期五阶", '=DATE_PARSE(A2, "yyyymmdd")', "=B2+5.5"],
            ["20260409梵呗增益", '=DATE_PARSE(A3, "yyyymmdd")', "=B3+5.5"],
        ]
    finally:
        _clear_user_override()


def test_note_sheet_sort_percent_uses_numeric_ratio(client, session):
    user = _create_user(session, username="note-sheet-percent-sort-user")
    _grant_feature_access(session, user_id=user.id, feature_key="notes.sheets")
    _override_user(user)

    try:
        create_sheet_response = client.post(
            "/api/note-sheets/sheets",
            json={
                "title": "百分比排序表",
                "document_json": {
                    "schema_version": 1,
                    "columns": ["课程", "返款率"],
                    "rows": [
                        ["甲", "75%"],
                        ["乙", "0.2"],
                        ["丙", "100%"],
                    ],
                    "column_configs": {
                        "返款率": {"value_type": "percent"},
                    },
                },
            },
        )
        assert create_sheet_response.status_code == 200
        sheet = create_sheet_response.json()

        sort_response = client.post(
            f"/api/note-sheets/sheets/{sheet['id']}/sort",
            json={"column_index": 1, "direction": "asc"},
        )
        assert sort_response.status_code == 200

        assert sort_response.json()["document_json"]["rows"] == [
            ["乙", "0.2"],
            ["甲", "75%"],
            ["丙", "100%"],
        ]
    finally:
        _clear_user_override()


def test_attendance_summary_generates_next_month_templates_idempotently(client, session):
    user = _create_user(session, username="note-sheet-attendance-template-user")
    _grant_feature_access(session, user_id=user.id, feature_key="notes.sheets")
    _override_user(user)

    def serial(year: int, month: int, day: int) -> str:
        return str((date(year, month, day) - date(1970, 1, 1)).days + 25569)

    try:
        workbook = WorkbookDocument(
            numeric_id=2,
            title="武陵禅寺网课考勤汇总",
            owner_user_id=user.id,
            created_by_user_id=user.id,
            updated_by_user_id=user.id,
        )
        sheet = SheetDocument(
            numeric_id=4,
            scope="notes",
            owner_type="note_sheet",
            owner_key="4",
            sheet_key="4",
            title="课程",
            owner_user_id=user.id,
            created_by_user_id=user.id,
            updated_by_user_id=user.id,
            document_json={
                "schema_version": 1,
                "columns": [
                    "课程类型",
                    "课程名称",
                    "在线考勤表",
                    "考勤负责人",
                    "备注",
                    "返款频次",
                    "课程开始日期",
                    "课程结束日期",
                    "考勤实际完成结点",
                    "报名费",
                    "报名人数",
                    "总报名费",
                    "退课人数",
                    "实际总报名费",
                    "促学金矫正",
                    "已返款",
                    "剩余促学金",
                    "返款率",
                ],
                "rows": [
                    ["念住闯关", "2025念住闯关第2部分", "20250106念住闯关", "如如, 陈坤泽"],
                    [
                        "念住",
                        "第39届念住",
                        {"value": "20260401第39届念住", "link": {"url": "https://www.kdocs.cn/l/source-nianzhu"}},
                        "如如, 陈坤泽",
                        "",
                        "每天上午6:14~7:07之后",
                        serial(2026, 4, 1),
                        "=G2+26",
                        "",
                        "620",
                        "7",
                        "=J2*K2",
                    ],
                    [
                        "觉观",
                        "第45届觉观",
                        {"value": "20260401第45届觉观", "link": {"url": "https://www.kdocs.cn/l/source-jueguan"}},
                        "月上, 陈成, 陈坤泽",
                        "",
                        "每天上午6:00~6:49之后",
                        serial(2026, 4, 1),
                        "=G3+24",
                        "",
                        "499",
                        "54",
                        "=J3*K3",
                    ],
                    ["梵呗初阶", "梵呗初阶", "20260309梵呗初阶", "王秀芹, 乐道行音, 陈坤泽", "", "数据每晚21点更新，次日返款", serial(2026, 3, 9), serial(2026, 3, 24), "", "550", "3", "=J4*K4"],
                    ["梵呗增益", "梵呗增益", "20260409梵呗增益", "王秀芹, 卓尔不凡, 陈坤泽", "", "数据每晚21点更新，次日返款", serial(2026, 4, 9), serial(2026, 5, 4), "", "500", "0", "=J5*K5"],
                ],
                "cell_meta": {},
                "entity_columns": [
                    {"id": f"col_{index}", "header": header}
                    for index, header in enumerate([
                        "课程类型",
                        "课程名称",
                        "在线考勤表",
                        "考勤负责人",
                        "备注",
                        "返款频次",
                        "课程开始日期",
                        "课程结束日期",
                        "考勤实际完成结点",
                        "报名费",
                        "报名人数",
                        "总报名费",
                        "退课人数",
                        "实际总报名费",
                        "促学金矫正",
                        "已返款",
                        "剩余促学金",
                        "返款率",
                    ])
                ],
                "entity_rows": [
                    {},
                    {"id": "row-nianzhu", "kind": "data"},
                    {"id": "row-jueguan", "kind": "data"},
                    {"id": "row-fanbei-chujie", "kind": "data"},
                    {"id": "row-fanbei-zengyi", "kind": "data"},
                ],
                "entity_cells": {},
            },
        )
        session.add(workbook)
        session.add(sheet)
        session.commit()
        session.refresh(workbook)
        session.refresh(sheet)
        session.add(WorkbookSheetLink(workbook_id=workbook.id, sheet_id=sheet.id, order_index=0))
        session.commit()
        initial_version = sheet.version

        stale_response = client.post(
            "/api/note-sheets/sheets/4/attendance-summary/generate-next-month-templates",
            json={"base_version": initial_version + 1, "target_year": 2026, "target_month": 5},
        )
        assert stale_response.status_code == 409

        response = client.post(
            "/api/note-sheets/sheets/4/attendance-summary/generate-next-month-templates",
            json={"base_version": initial_version, "target_year": 2026, "target_month": 5},
        )
        assert response.status_code == 200
        payload = response.json()
        assert [item["course_name"] for item in payload["generated"]] == ["第40届念住", "第46届觉观", "梵呗初阶"]
        rows = payload["sheet"]["document_json"]["rows"]
        assert rows[0][0:4] == ["念住闯关", "2025念住闯关第2部分", "20250106念住闯关", "如如, 陈坤泽"]
        assert rows[1] == [
            "念住",
            "第40届念住",
            "第40届念住",
            "如如, 陈坤泽",
            "",
            "每天上午6:14~7:07之后",
            serial(2026, 5, 1),
            "=G2+26",
            "",
            "620",
            "",
            "=J2*K2",
            "",
            "",
            "",
            "",
            "",
            "",
        ]
        assert rows[2][0:10] == [
            "觉观",
            "第46届觉观",
            "第46届觉观",
            "月上, 陈成, 陈坤泽",
            "",
            "每天上午6:00~6:49之后",
            serial(2026, 5, 1),
            "=G3+24",
            "",
            "499",
        ]
        assert rows[2][10:12] == ["", "=J3*K3"]
        assert rows[3][0:12] == [
            "梵呗初阶",
            "梵呗初阶",
            "20260509梵呗初阶",
            "王秀芹, 乐道行音, 陈坤泽",
            "",
            "数据每晚21点更新，次日返款",
            serial(2026, 5, 9),
            serial(2026, 5, 24),
            "",
            "550",
            "",
            "=J4*K4",
        ]
        assert rows[4][7] == "=G5+26"
        assert rows[5][7] == "=G6+24"

        persisted = session.exec(select(SheetDocument).where(SheetDocument.numeric_id == 4)).one()
        assert not any(
            "link" in entry
            for entry in persisted.document_json.get("cell_meta", {}).values()
            if isinstance(entry, dict)
        )
        assert persisted.document_json["rows"][4][2] == {
            "value": "20260401第39届念住",
            "link": {"url": "https://www.kdocs.cn/l/source-nianzhu"},
        }
        assert persisted.document_json["rows"][5][2] == {
            "value": "20260401第45届觉观",
            "link": {"url": "https://www.kdocs.cn/l/source-jueguan"},
        }
        assert persisted.document_json["entity_rows"][:6] == [
            {},
            {},
            {},
            {},
            {"id": "row-nianzhu", "kind": "data"},
            {"id": "row-jueguan", "kind": "data"},
        ]
        assert (
            note_sheets_api._get_document_cell_link_url(persisted.document_json, 4, 2)
            == "https://www.kdocs.cn/l/source-nianzhu"
        )
        assert (
            note_sheets_api._get_document_cell_link_url(persisted.document_json, 5, 2)
            == "https://www.kdocs.cn/l/source-jueguan"
        )

        second_response = client.post(
            "/api/note-sheets/sheets/4/attendance-summary/generate-next-month-templates",
            json={"base_version": payload["sheet"]["version"], "target_year": 2026, "target_month": 5},
        )
        assert second_response.status_code == 200
        second_payload = second_response.json()
        assert second_payload["generated"] == []
        assert [item["reason"] for item in second_payload["skipped"]] == ["目标课程已存在", "目标课程已存在", "目标课程已存在"]
        assert len(second_payload["sheet"]["document_json"]["rows"]) == 8

        fanbei_zengyi_row_index = next(
            index
            for index, row in enumerate(second_payload["sheet"]["document_json"]["rows"])
            if row[0] == "梵呗增益"
        )
        generic_response = client.post(
            "/api/note-sheets/sheets/4/attendance-summary/generate-course-template",
            json={"base_version": second_payload["sheet"]["version"], "row_index": fanbei_zengyi_row_index},
        )
        assert generic_response.status_code == 200
        generic_payload = generic_response.json()
        assert [item["course_name"] for item in generic_payload["generated"]] == ["梵呗增益"]
        generic_rows = generic_payload["sheet"]["document_json"]["rows"]
        assert generic_rows[7][0:12] == [
            "梵呗增益",
            "梵呗增益",
            "20260609梵呗增益",
            "王秀芹, 卓尔不凡, 陈坤泽",
            "",
            "数据每晚21点更新，次日返款",
            serial(2026, 6, 9),
            serial(2026, 7, 4),
            "",
            "500",
            "",
            "=J8*K8",
        ]
    finally:
        _clear_user_override()


def test_attendance_summary_next_month_templates_can_skip_monthly_course_type(client, session):
    user = _create_user(session, username="note-sheet-attendance-template-skip-user")
    _grant_feature_access(session, user_id=user.id, feature_key="notes.sheets")
    _override_user(user)

    def serial(year: int, month: int, day: int) -> str:
        return str((date(year, month, day) - date(1970, 1, 1)).days + 25569)

    try:
        workbook = WorkbookDocument(
            numeric_id=2,
            title="武陵禅寺网课考勤汇总",
            owner_user_id=user.id,
            created_by_user_id=user.id,
            updated_by_user_id=user.id,
        )
        sheet = SheetDocument(
            numeric_id=4,
            scope="notes",
            owner_type="note_sheet",
            owner_key="4",
            sheet_key="4",
            title="课程",
            owner_user_id=user.id,
            created_by_user_id=user.id,
            updated_by_user_id=user.id,
            document_json={
                "schema_version": 1,
                "columns": [
                    "课程类型",
                    "课程名称",
                    "在线考勤表",
                    "考勤负责人",
                    "返款频次",
                    "课程开始日期",
                    "课程结束日期",
                    "考勤实际完成结点",
                    "报名费",
                    "报名人数",
                    "总报名费",
                ],
                "rows": [
                    ["念住", "第41届念住", "第41届念住", "如如, 陈坤泽", "每天上午", serial(2026, 6, 1), "=F1+26", "", "620", "1", "=I1*J1"],
                    ["觉观", "第47届觉观", "第47届觉观", "陈成, 陈坤泽", "每天上午", serial(2026, 6, 1), "=F2+24", "", "499", "1", "=I2*J2"],
                    ["梵呗初阶", "梵呗初阶", "20260609梵呗初阶", "王秀芹, 陈坤泽", "数据每晚21点更新", serial(2026, 6, 9), serial(2026, 6, 24), "", "550", "1", "=I3*J3"],
                ],
                "cell_meta": {},
            },
        )
        session.add(workbook)
        session.add(sheet)
        session.commit()
        session.refresh(workbook)
        session.refresh(sheet)
        session.add(WorkbookSheetLink(workbook_id=workbook.id, sheet_id=sheet.id, order_index=0))
        session.commit()

        response = client.post(
            "/api/note-sheets/sheets/4/attendance-summary/generate-next-month-templates",
            json={
                "base_version": sheet.version,
                "target_year": 2026,
                "target_month": 7,
                "skip_course_types": ["梵呗初阶"],
            },
        )
        assert response.status_code == 200
        payload = response.json()
        assert [item["course_name"] for item in payload["generated"]] == ["第42届念住", "第48届觉观"]
        assert [(item["course_type"], item["reason"]) for item in payload["skipped"]] == [("梵呗初阶", "本月未排课")]
        rows = payload["sheet"]["document_json"]["rows"]
        assert not any(row[0] == "梵呗初阶" and row[2] == "20260709梵呗初阶" for row in rows)
    finally:
        _clear_user_override()


def test_attendance_summary_next_month_templates_materialize_local_course_workbook(client, session, monkeypatch):
    user = _create_user(session, username="note-sheet-attendance-template-workbook-user")
    template_owner = _create_user(session, username="note-sheet-attendance-template-ru-ru")
    _grant_feature_access(session, user_id=user.id, feature_key="notes.sheets")
    _override_user(user)

    def serial(year: int, month: int, day: int) -> str:
        return str((date(year, month, day) - date(1970, 1, 1)).days + 25569)

    try:
        from backend.core.attendance import nianzhu_course_sheets

        monkeypatch.setattr(
            nianzhu_course_sheets,
            "_query_legacy_lesson_rows",
            lambda course_name: (
                [
                    {
                        "lesson_id": 9001,
                        "start_date": "2026-07-01 05:20:00",
                        "end_date": "2026-07-08 06:20:00",
                        "next_update": "9999-12-31 23:59:59",
                        "lesson_id2": "l_july_nianzhu_01",
                        "shop_id": 1,
                        "lesson_name": f"{course_name}-第01课",
                        "video_duration": 3600,
                    }
                ],
                None,
            ),
        )
        monkeypatch.setattr(nianzhu_course_sheets, "_query_legacy_lesson_data_rows", lambda lesson_ids: ([], None))
        monkeypatch.setattr(
            nianzhu_course_sheets,
            "_query_legacy_clockin_rows",
            lambda course_name: (
                [
                    {
                        "clockin_id": 7001,
                        "name": f"{course_name}-打卡数",
                        "url": "https://example.com/july-clockin",
                        "start_date": "2026-07-01",
                        "end_date": "2026-07-27",
                        "days": 27,
                        "clockin_user_num": 0,
                        "total_user_num": 0,
                    }
                ],
                None,
            ),
        )
        monkeypatch.setattr(nianzhu_course_sheets, "_query_legacy_clockin_data_rows", lambda clockin_ids: ([], None))

        summary_workbook = WorkbookDocument(
            numeric_id=2,
            title="武陵禅寺网课考勤汇总",
            owner_user_id=user.id,
            created_by_user_id=user.id,
            updated_by_user_id=user.id,
        )
        source_workbook = WorkbookDocument(
            numeric_id=10,
            title="第41届念住",
            owner_user_id=user.id,
            created_by_user_id=user.id,
            updated_by_user_id=user.id,
        )
        summary_sheet = SheetDocument(
            numeric_id=4,
            scope="notes",
            owner_type="note_sheet",
            owner_key="4",
            sheet_key="4",
            title="课程",
            owner_user_id=user.id,
            created_by_user_id=user.id,
            updated_by_user_id=user.id,
            document_json={
                "schema_version": 1,
                "columns": [
                    "课程类型",
                    "课程名称",
                    "在线考勤表",
                    "考勤负责人",
                    "返款频次",
                    "课程开始日期",
                    "课程结束日期",
                    "考勤实际完成结点",
                    "报名费",
                    "报名人数",
                    "总报名费",
                ],
                "rows": [
                    [
                        "念住",
                        "第41届念住",
                        {"value": "第41届念住", "link": {"url": "/workbook/10?sheet=54605"}},
                        "如如, 陈坤泽",
                        "每天上午",
                        serial(2026, 6, 1),
                        "=F1+26",
                        "",
                        "620",
                        "1",
                        "=I1*J1",
                    ],
                ],
                "grid_rows": [
                    [
                        "课程类型",
                        "课程名称",
                        "在线考勤表",
                        "考勤负责人",
                        "返款频次",
                        "课程开始日期",
                        "课程结束日期",
                        "考勤实际完成结点",
                        "报名费",
                        "报名人数",
                        "总报名费",
                    ],
                    [
                        "念住",
                        "第41届念住",
                        {"value": "第41届念住", "link": {"url": "/workbook/10?sheet=54605"}},
                        "如如, 陈坤泽",
                        "每天上午",
                        serial(2026, 6, 1),
                        "=F1+26",
                        "",
                        "620",
                        "1",
                        "=I1*J1",
                    ],
                ],
                "data_start_row": 1,
                "field_row_index": 0,
                "cell_meta": {},
            },
        )
        source_attendance = SheetDocument(
            numeric_id=54605,
            scope="notes",
            owner_type="course_workbook",
            owner_key="20260601-nianzhu-41",
            sheet_key="attendance",
            title="考勤表",
            owner_user_id=user.id,
            created_by_user_id=user.id,
            updated_by_user_id=user.id,
            document_json={
                "schema_version": 1,
                "columns": ["学号", "姓名", "打卡数", "05:20~06:20 第01课", "20:00~22:00 觉观同学会", "第01课", "已返款"],
                "rows": [["1", "甲", "", "", "", "", "0"]],
                "header_groups": [[
                    {"label": "", "colspan": 1},
                    {"label": "", "colspan": 1},
                    {"label": "", "colspan": 1},
                    {"label": "6月1日~6月5日", "colspan": 1},
                    {"label": "6月22日", "colspan": 1},
                    {"label": "", "colspan": 1},
                    {"label": "", "colspan": 1},
                ]],
                "grid_rows": [
                    ["", "", "", "6月1日~6月5日", "6月22日", "", ""],
                    ["学号", "姓名", "打卡数", "05:20~06:20 第01课", "20:00~22:00 觉观同学会", "第01课", "已返款"],
                    ["", "", "只统计正课的打卡次数", "", "结营分享+答疑解惑，不做考勤", "", ""],
                    ["1", "甲", "", "", "", "", "0"],
                ],
                "data_start_row": 3,
                "field_row_index": 1,
                "cell_meta": {
                    "0:2": {"style": {"background_color": "#C5E0B4"}},
                    "1:2": {"style": {"background_color": "#E2F0D9", "text_color": "#0000FF"}},
                    "2:2": {"style": {"background_color": "#D8D8D8", "text_color": "#FF0000"}},
                    "0:3": {"style": {"background_color": "#FFE699"}},
                    "1:3": {"style": {"background_color": "#FFF2CC"}},
                    "2:3": {"style": {"background_color": "#D8D8D8"}},
                    "0:4": {"style": {"background_color": "#FFE699"}},
                    "1:4": {"style": {"background_color": "#FFF2CC"}},
                    "2:4": {"style": {"background_color": "#D8D8D8"}},
                },
                "entity_columns": [
                    {"id": "col_student_no", "header": "学号"},
                    {"id": "col_name", "header": "姓名"},
                    {"id": "col_clockin", "header": "打卡数"},
                    {"id": "col_lesson_01", "header": "05:20~06:20 第01课"},
                    {"id": "col_meeting", "header": "20:00~22:00 觉观同学会"},
                    {"id": "col_plain_01", "header": "第01课"},
                    {"id": "col_refunded", "header": "已返款"},
                ],
                "entity_rows": [
                    {"id": "header_group", "kind": "header_group"},
                    {"id": "field", "kind": "field"},
                    {"id": "field_note", "kind": "field_note"},
                    {"id": "row_student_1", "kind": "data"},
                ],
                "entity_cells": {
                    "header_group": {
                        "col_clockin": {"style": {"background_color": "#C5E0B4"}},
                        "col_lesson_01": {"value": "6月1日~6月5日", "style": {"background_color": "#FFE699"}},
                        "col_meeting": {"value": "6月22日", "style": {"background_color": "#FFE699"}},
                    },
                    "field": {
                        "col_clockin": {"style": {"background_color": "#E2F0D9", "text_color": "#0000FF"}},
                        "col_lesson_01": {"value": "05:20~06:20 第01课", "style": {"background_color": "#FFF2CC"}},
                        "col_meeting": {"value": "20:00~22:00 觉观同学会", "style": {"background_color": "#FFF2CC"}},
                    },
                    "field_note": {
                        "col_clockin": {"style": {"background_color": "#D8D8D8", "text_color": "#FF0000"}},
                        "col_lesson_01": {"style": {"background_color": "#D8D8D8"}},
                        "col_meeting": {"value": "结营分享+答疑解惑，不做考勤", "style": {"background_color": "#D8D8D8"}},
                    },
                    "row_student_1": {
                        "col_student_no": {"value": "1"},
                        "col_name": {"value": "甲"},
                        "col_refunded": {"value": "0"},
                    },
                },
                "column_configs": {
                    "05:20~06:20 第01课": {
                        "header_background_color": "#E2F0D9",
                        "note": "视频观看说明",
                    },
                    "打卡数": {
                        "header_background_color": "#FFF2CC",
                        "value_type": "number",
                    },
                },
            },
        )
        source_registration = SheetDocument(
            numeric_id=54606,
            scope="notes",
            owner_type="course_workbook",
            owner_key="20260601-nianzhu-41",
            sheet_key="registration",
            title="报名表",
            owner_user_id=user.id,
            created_by_user_id=user.id,
            updated_by_user_id=user.id,
            document_json={
                "schema_version": 1,
                "columns": ["序号", "姓名"],
                "rows": [["1", "甲"]],
                "grid_rows": [["序号", "姓名"], ["1", "甲"]],
                "data_start_row": 1,
                "field_row_index": 0,
            },
        )
        session.add(summary_workbook)
        session.add(source_workbook)
        session.add(summary_sheet)
        session.add(source_attendance)
        session.add(source_registration)
        session.commit()
        session.refresh(summary_workbook)
        session.refresh(source_workbook)
        session.refresh(summary_sheet)
        session.refresh(source_attendance)
        session.refresh(source_registration)
        session.add(WorkbookSheetLink(workbook_id=summary_workbook.id, sheet_id=summary_sheet.id, order_index=0))
        session.add(WorkbookSheetLink(workbook_id=source_workbook.id, sheet_id=source_attendance.id, order_index=5))
        session.add(WorkbookSheetLink(workbook_id=source_workbook.id, sheet_id=source_registration.id, order_index=10))
        session.add(ResourceAccessGrant(
            resource_type="workbook",
            resource_id=str(source_workbook.numeric_id),
            subject_key=f"user:{template_owner.id}",
            subject_type="user",
            subject_user_id=template_owner.id,
            role="editor",
        ))
        session.add(ResourceAccessGrant(
            resource_type="sheet",
            resource_id=str(source_attendance.numeric_id),
            subject_key=f"user:{template_owner.id}",
            subject_type="user",
            subject_user_id=template_owner.id,
            role="editor",
        ))
        session.add(ResourceAccessGrant(
            resource_type="sheet",
            resource_id=str(source_registration.numeric_id),
            subject_key=f"user:{template_owner.id}",
            subject_type="user",
            subject_user_id=template_owner.id,
            role="viewer",
        ))
        session.commit()

        response = client.post(
            "/api/note-sheets/sheets/4/attendance-summary/generate-next-month-templates",
            json={
                "base_version": summary_sheet.version,
                "target_year": 2026,
                "target_month": 7,
                "skip_course_types": ["觉观", "梵呗初阶"],
            },
        )

        assert response.status_code == 200
        payload = response.json()
        assert [item["course_name"] for item in payload["generated"]] == ["第42届念住"]
        linked_cell = payload["sheet"]["document_json"]["rows"][0][2]
        assert linked_cell["value"] == "第42届念住"
        assert linked_cell["link"]["url"].startswith("/workbook/")

        workbook_id_text = linked_cell["link"]["url"].split("/workbook/", 1)[1].split("?", 1)[0]
        sheet_id_text = linked_cell["link"]["url"].split("sheet=", 1)[1]
        created_workbook = session.exec(
            select(WorkbookDocument).where(WorkbookDocument.numeric_id == int(workbook_id_text))
        ).one()
        created_attendance = session.exec(
            select(SheetDocument).where(SheetDocument.numeric_id == int(sheet_id_text))
        ).one()
        assert created_workbook.title == "第42届念住"
        assert created_attendance.owner_key == "20260701-nianzhu-42"
        assert created_attendance.sheet_key == "attendance"
        assert created_attendance.document_json["columns"] == ["学号", "姓名", "打卡数", "05:20~06:20 第01课", "20:00~22:00 觉观同学会", "已返款"]
        assert created_attendance.document_json["rows"] == []
        assert created_attendance.document_json["grid_rows"][0][3] == "7月1日~7月5日"
        assert created_attendance.document_json["grid_rows"][0][4] == "7月22日"
        assert created_attendance.document_json["grid_rows"][2][4] == "结营分享+答疑解惑，不做考勤"
        assert created_attendance.document_json["header_groups"][0][3]["label"] == "7月1日~7月5日"
        assert created_attendance.document_json["header_groups"][0][4]["label"] == "7月22日"
        created_cell_meta = created_attendance.document_json.get("cell_meta") or {}
        assert created_cell_meta["0:3"]["style"]["background_color"] == "#C5E0B4"
        assert created_cell_meta["1:3"]["style"]["background_color"] == "#E2F0D9"
        assert created_cell_meta["2:3"]["style"]["background_color"] == "#D8D8D8"
        assert created_cell_meta["0:4"]["style"]["background_color"] == "#FFE699"
        entity_rows = created_attendance.document_json["entity_rows"]
        entity_columns = created_attendance.document_json["entity_columns"]
        entity_cells = created_attendance.document_json["entity_cells"]
        assert [row["kind"] for row in entity_rows] == ["header_group", "field", "field_note"]
        lesson_entity_column = entity_columns[3]["id"]
        meeting_entity_column = entity_columns[4]["id"]
        assert entity_cells["header_group"][lesson_entity_column]["value"] == "7月1日~7月5日"
        assert entity_cells["header_group"][lesson_entity_column]["style"]["background_color"] == "#C5E0B4"
        assert entity_cells["field"][lesson_entity_column]["style"]["background_color"] == "#E2F0D9"
        assert entity_cells["field_note"][lesson_entity_column]["style"]["background_color"] == "#D8D8D8"
        assert entity_cells["header_group"][meeting_entity_column]["value"] == "7月22日"
        assert entity_cells["header_group"][meeting_entity_column]["style"]["background_color"] == "#FFE699"
        assert created_attendance.document_json["column_configs"]["05:20~06:20 第01课"] == {"note": "视频观看说明"}
        assert created_attendance.document_json["column_configs"]["打卡数"] == {"value_type": "number"}
        created_video_config = session.exec(
            select(SheetDocument)
            .where(SheetDocument.owner_key == "20260701-nianzhu-42")
            .where(SheetDocument.sheet_key == "video_config")
        ).one()
        created_clockin_config = session.exec(
            select(SheetDocument)
            .where(SheetDocument.owner_key == "20260701-nianzhu-42")
            .where(SheetDocument.sheet_key == "clockin_config")
        ).one()
        assert created_video_config.document_json["rows"][0][4] == "l_july_nianzhu_01"
        assert all("同学会" not in str(row[6]) for row in created_video_config.document_json["rows"])
        assert created_clockin_config.document_json["rows"][0][2] == "https://example.com/july-clockin"
        created_registration = session.exec(
            select(SheetDocument)
            .where(SheetDocument.owner_key == "20260701-nianzhu-42")
            .where(SheetDocument.sheet_key == "registration")
        ).one()
        copied_workbook_grant = session.exec(
            select(ResourceAccessGrant)
            .where(ResourceAccessGrant.resource_type == "workbook")
            .where(ResourceAccessGrant.resource_id == str(created_workbook.numeric_id))
            .where(ResourceAccessGrant.subject_key == f"user:{template_owner.id}")
        ).one()
        copied_attendance_grant = session.exec(
            select(ResourceAccessGrant)
            .where(ResourceAccessGrant.resource_type == "sheet")
            .where(ResourceAccessGrant.resource_id == str(created_attendance.numeric_id))
            .where(ResourceAccessGrant.subject_key == f"user:{template_owner.id}")
        ).one()
        copied_registration_grant = session.exec(
            select(ResourceAccessGrant)
            .where(ResourceAccessGrant.resource_type == "sheet")
            .where(ResourceAccessGrant.resource_id == str(created_registration.numeric_id))
            .where(ResourceAccessGrant.subject_key == f"user:{template_owner.id}")
        ).one()
        assert copied_workbook_grant.role == "editor"
        assert copied_attendance_grant.role == "editor"
        assert copied_registration_grant.role == "viewer"
    finally:
        _clear_user_override()


def test_attendance_summary_existing_next_month_template_can_materialize_workbook(client, session):
    user = _create_user(session, username="note-sheet-attendance-existing-template-workbook-user")
    _grant_feature_access(session, user_id=user.id, feature_key="notes.sheets")
    _override_user(user)

    def serial(year: int, month: int, day: int) -> str:
        return str((date(year, month, day) - date(1970, 1, 1)).days + 25569)

    try:
        summary_workbook = WorkbookDocument(numeric_id=2, title="武陵禅寺网课考勤汇总", owner_user_id=user.id)
        source_workbook = WorkbookDocument(numeric_id=10, title="第41届念住", owner_user_id=user.id)
        summary_sheet = SheetDocument(
            numeric_id=4,
            scope="notes",
            owner_type="note_sheet",
            owner_key="4",
            sheet_key="4",
            title="课程",
            owner_user_id=user.id,
            document_json={
                "schema_version": 1,
                "columns": ["课程类型", "课程名称", "在线考勤表", "考勤负责人", "课程开始日期", "课程结束日期"],
                "rows": [
                    ["念住", "第42届念住", "第42届念住", "如如, 陈坤泽", serial(2026, 7, 1), "=E1+26"],
                    [
                        "念住",
                        "第41届念住",
                        {"value": "第41届念住", "link": {"url": "/workbook/10?sheet=54605"}},
                        "如如, 陈坤泽",
                        serial(2026, 6, 1),
                        "=E2+26",
                    ],
                ],
                "grid_rows": [
                    ["课程类型", "课程名称", "在线考勤表", "考勤负责人", "课程开始日期", "课程结束日期"],
                    ["念住", "第42届念住", "第42届念住", "如如, 陈坤泽", serial(2026, 7, 1), "=E1+26"],
                    [
                        "念住",
                        "第41届念住",
                        {"value": "第41届念住", "link": {"url": "/workbook/10?sheet=54605"}},
                        "如如, 陈坤泽",
                        serial(2026, 6, 1),
                        "=E2+26",
                    ],
                ],
                "data_start_row": 1,
                "field_row_index": 0,
            },
        )
        source_attendance = SheetDocument(
            numeric_id=54605,
            scope="notes",
            owner_type="course_workbook",
            owner_key="20260601-nianzhu-41",
            sheet_key="attendance",
            title="考勤表",
            owner_user_id=user.id,
            document_json={
                "schema_version": 1,
                "columns": ["学号", "姓名"],
                "rows": [["1", "甲"]],
                "grid_rows": [["学号", "姓名"], ["1", "甲"]],
                "data_start_row": 1,
                "field_row_index": 0,
            },
        )
        session.add(summary_workbook)
        session.add(source_workbook)
        session.add(summary_sheet)
        session.add(source_attendance)
        session.commit()
        session.refresh(summary_workbook)
        session.refresh(source_workbook)
        session.refresh(summary_sheet)
        session.refresh(source_attendance)
        session.add(WorkbookSheetLink(workbook_id=summary_workbook.id, sheet_id=summary_sheet.id, order_index=0))
        session.add(WorkbookSheetLink(workbook_id=source_workbook.id, sheet_id=source_attendance.id, order_index=5))
        session.commit()

        response = client.post(
            "/api/note-sheets/sheets/4/attendance-summary/generate-next-month-templates",
            json={
                "base_version": summary_sheet.version,
                "target_year": 2026,
                "target_month": 7,
                "skip_course_types": ["觉观", "梵呗初阶"],
            },
        )

        assert response.status_code == 200
        payload = response.json()
        assert payload["generated"] == []
        assert payload["sheet"]["document_json"]["rows"][0][2]["value"] == "第42届念住"
        link_url = payload["sheet"]["document_json"]["rows"][0][2]["link"]["url"]
        assert link_url.startswith("/workbook/")
        sheet_id = int(link_url.split("sheet=", 1)[1])
        created_attendance = session.exec(select(SheetDocument).where(SheetDocument.numeric_id == sheet_id)).one()
        assert created_attendance.owner_key == "20260701-nianzhu-42"
        assert created_attendance.document_json["rows"] == []
    finally:
        _clear_user_override()


def test_attendance_course_template_uses_column_binding_fallback(client, session):
    user = _create_user(session, username="note-sheet-attendance-binding-user")
    _grant_feature_access(session, user_id=user.id, feature_key="notes.sheets")
    _override_user(user)

    def serial(year: int, month: int, day: int) -> str:
        return str((date(year, month, day) - date(1970, 1, 1)).days + 25569)

    try:
        workbook = WorkbookDocument(
            numeric_id=2,
            title="武陵禅寺网课考勤汇总",
            owner_user_id=user.id,
            created_by_user_id=user.id,
            updated_by_user_id=user.id,
        )
        sheet = SheetDocument(
            numeric_id=4,
            scope="notes",
            owner_type="note_sheet",
            owner_key="4",
            sheet_key="4",
            title="课程",
            owner_user_id=user.id,
            created_by_user_id=user.id,
            updated_by_user_id=user.id,
            document_json={
                "schema_version": 1,
                "columns": [
                    "类型别名",
                    "名称别名",
                    "考勤链接别名",
                    "负责人",
                    "备注",
                    "返款频次",
                    "开始日期别名",
                    "结束日期别名",
                    "完成日期别名",
                    "报名费",
                    "人数别名",
                    "总报名费",
                ],
                "rows": [
                    [
                        "禅宗二阶",
                        "禅宗9期二阶",
                        "20251026禅宗9期二阶",
                        "陈坤泽, 王颖",
                        "",
                        "周日开课",
                        serial(2025, 10, 26),
                        serial(2025, 12, 28),
                        "",
                        "910",
                        "29",
                        "=J1*K1",
                    ],
                ],
            },
        )
        session.add(workbook)
        session.add(sheet)
        session.commit()
        session.refresh(workbook)
        session.refresh(sheet)
        session.add(WorkbookSheetLink(workbook_id=workbook.id, sheet_id=sheet.id, order_index=0))
        session.commit()

        response = client.post(
            "/api/note-sheets/sheets/4/attendance-summary/generate-course-template",
            json={"row_index": 0, "target_date": "2026-05-03"},
        )
        assert response.status_code == 200
        rows = response.json()["sheet"]["document_json"]["rows"]
        assert rows[0][0:12] == [
            "禅宗二阶",
            "禅宗9期二阶",
            "20260503禅宗9期二阶",
            "陈坤泽, 王颖",
            "",
            "周日开课",
            serial(2026, 5, 3),
            serial(2026, 7, 5),
            "",
            "910",
            "",
            "=J1*K1",
        ]
    finally:
        _clear_user_override()


def test_attendance_summary_set_completed_moves_row_to_completion_boundary(client, session):
    user = _create_user(session, username="note-sheet-attendance-complete-user")
    _grant_feature_access(session, user_id=user.id, feature_key="notes.sheets")
    _override_user(user)

    def serial(year: int, month: int, day: int) -> str:
        return str((date(year, month, day) - date(1970, 1, 1)).days + 25569)

    try:
        workbook = WorkbookDocument(
            numeric_id=2,
            title="武陵禅寺网课考勤汇总",
            owner_user_id=user.id,
            created_by_user_id=user.id,
            updated_by_user_id=user.id,
        )
        sheet = SheetDocument(
            numeric_id=4,
            scope="notes",
            owner_type="note_sheet",
            owner_key="4",
            sheet_key="4",
            title="课程",
            owner_user_id=user.id,
            created_by_user_id=user.id,
            updated_by_user_id=user.id,
            document_json={
                "schema_version": 1,
                "columns": [
                    "课程类型",
                    "课程名称",
                    "在线考勤表",
                    "考勤负责人",
                    "备注",
                    "返款频次",
                    "课程开始日期",
                    "课程结束日期",
                    "考勤实际完成结点",
                    "报名费",
                    "报名人数",
                    "总报名费",
                ],
                "rows": [
                    ["念住", "第40届念住", "20260501第40届念住", "", "", "", serial(2026, 5, 1), "=G1+26", "", "620", "7", "=J1*K1"],
                    [
                        "觉观",
                        "第46届觉观",
                        {"value": "20260501第46届觉观", "link": {"url": "https://example.com/jueguan"}},
                        "",
                        "",
                        "",
                        serial(2026, 5, 1),
                        "=G2+24",
                        "",
                        "499",
                        "54",
                        "=J2*K2",
                    ],
                    [
                        "梵呗初阶",
                        "梵呗初阶",
                        {"value": "20260501梵呗初阶", "link": {"url": "https://example.com/fanbei"}},
                        "",
                        "",
                        "",
                        serial(2026, 5, 1),
                        serial(2026, 5, 16),
                        "",
                        "550",
                        "3",
                        "=J3*K3",
                    ],
                    ["旧完结", "旧完结", "20260401旧完结", "", "", "", serial(2026, 4, 1), serial(2026, 4, 20), serial(2026, 4, 20), "1", "1", "=J4*K4"],
                ],
                "cell_meta": {},
            },
        )
        session.add(workbook)
        session.add(sheet)
        session.commit()
        session.refresh(workbook)
        session.refresh(sheet)
        session.add(WorkbookSheetLink(workbook_id=workbook.id, sheet_id=sheet.id, order_index=0))
        session.commit()
        initial_version = sheet.version

        stale_response = client.post(
            "/api/note-sheets/sheets/4/attendance-summary/set-completed",
            json={"base_version": initial_version + 1, "row_index": 1, "completion_date": "2026-04-30"},
        )
        assert stale_response.status_code == 409

        response = client.post(
            "/api/note-sheets/sheets/4/attendance-summary/set-completed",
            json={"base_version": initial_version, "row_index": 1, "completion_date": "2026-04-30"},
        )
        assert response.status_code == 200
        payload = response.json()
        assert payload["row_index"] == 2
        rows = payload["sheet"]["document_json"]["rows"]
        assert [row[0] for row in rows] == ["念住", "梵呗初阶", "觉观", "旧完结"]
        assert rows[1][11] == "=J2*K2"
        assert rows[2][7] == "=G3+24"
        assert rows[2][8] == serial(2026, 4, 30)
        assert rows[2][11] == "=J3*K3"
        assert payload["sheet"]["document_json"]["rows"][2][2]["link"]["url"] == "https://example.com/jueguan"
        assert payload["sheet"]["document_json"]["rows"][1][2]["link"]["url"] == "https://example.com/fanbei"
        assert not any(
            "link" in entry
            for entry in payload["sheet"]["document_json"].get("cell_meta", {}).values()
            if isinstance(entry, dict)
        )
    finally:
        _clear_user_override()


def test_attendance_summary_generates_course_script_from_nearest_local_template(
    client,
    session,
    tmp_path,
    monkeypatch,
):
    user = _create_user(session, username="note-sheet-attendance-script-user")
    _override_user(user)

    def serial(year: int, month: int, day: int) -> str:
        return str((date(year, month, day) - date(1970, 1, 1)).days + 25569)

    courses_dir = tmp_path / "courses"
    completed_dir = courses_dir / "已完结"
    completed_dir.mkdir(parents=True)
    monkeypatch.setattr(note_sheets_api, "ATTENDANCE_COURSE_SCRIPT_DIR", courses_dir)
    completed_source = completed_dir / "d260401第39届念住.py"
    completed_zen_source = completed_dir / "d251130禅宗7期4点5阶.py"
    completed_source.write_text(
        "\n".join([
            "from xlsln.kq5034.courses.kqcourse import *",
            "",
            "",
            "class 考勤课程(KqCourse):",
            "",
            "    def __init__(self):",
            "        super().__init__(1,",
            "                         XlPath(__file__).stem,",
            "                         'oldKdocsToken',",
            "                         'V2-source-script',",
            "                         7,",
            "                         课程商品名='第39届念住禅法（初阶）【中心教室】')",
            "",
        ]),
        encoding="utf-8",
    )
    completed_zen_source.write_text("# zen source\n", encoding="utf-8")

    try:
        workbook = WorkbookDocument(
            numeric_id=2,
            title="武陵禅寺网课考勤汇总",
            owner_user_id=user.id,
            created_by_user_id=user.id,
            updated_by_user_id=user.id,
        )
        sheet = SheetDocument(
            numeric_id=4,
            scope="notes",
            owner_type="note_sheet",
            owner_key="4",
            sheet_key="4",
            title="课程",
            owner_user_id=user.id,
            created_by_user_id=user.id,
            updated_by_user_id=user.id,
            document_json={
                "schema_version": 1,
                "columns": [
                    "课程类型",
                    "课程名称",
                    "在线考勤表",
                    "考勤负责人",
                    "备注",
                    "返款频次",
                    "课程开始日期",
                    "课程结束日期",
                    "考勤实际完成结点",
                    "报名费",
                    "报名人数",
                    "总报名费",
                ],
                "rows": [
                    [
                        "念住",
                        "第39届念住",
                        {"value": "20260401第39届念住", "link": {"url": "https://www.kdocs.cn/l/existingToken"}},
                        "",
                        "",
                        "",
                        serial(2026, 4, 1),
                        "=G1+26",
                        "",
                        "620",
                        "7",
                        "=J1*K1",
                    ],
                    [
                        "念住",
                        "第40届念住",
                        {"value": "20260501第40届念住", "link": {"url": "https://www.kdocs.cn/l/newKdocsToken"}},
                        "",
                        "",
                        "",
                        serial(2026, 5, 1),
                        "=G2+26",
                        "",
                        "620",
                        "",
                        "=J2*K2",
                    ],
                    [
                        "禅宗4.5阶",
                        "禅宗8期4.5阶",
                        {"value": "20260503禅宗8期4.5阶.xlsx", "link": {"url": "https://www.kdocs.cn/l/newZenToken"}},
                        "",
                        "",
                        "",
                        serial(2026, 5, 3),
                        serial(2026, 7, 5),
                        "",
                        "1030",
                        "",
                        "=J3*K3",
                    ],
                ],
                "cell_meta": {},
            },
        )
        session.add(workbook)
        session.add(sheet)
        session.commit()
        session.refresh(workbook)
        session.refresh(sheet)
        session.add(WorkbookSheetLink(workbook_id=workbook.id, sheet_id=sheet.id, order_index=0))
        session.commit()

        status_response = client.get("/api/note-sheets/sheets/4/attendance-summary/course-script-statuses")
        assert status_response.status_code == 200
        statuses = {item["row_index"]: item for item in status_response.json()["statuses"]}
        assert statuses[0]["exists"] is True
        assert statuses[0]["target_filename"] == "d260401第39届念住.py"
        assert statuses[1]["can_generate"] is True
        assert statuses[1]["target_filename"] == "d260501第40届念住.py"
        assert statuses[2]["can_generate"] is True
        assert statuses[2]["target_filename"] == "d260503禅宗8期4点5阶.py"

        generate_response = client.post(
            "/api/note-sheets/sheets/4/attendance-summary/generate-course-script",
            json={"row_index": 1},
        )
        assert generate_response.status_code == 200
        payload = generate_response.json()
        assert payload["source_filename"] == "d260401第39届念住.py"
        assert payload["status"]["target_filename"] == "d260501第40届念住.py"

        created_file = courses_dir / "d260501第40届念住.py"
        assert created_file.exists()
        created_text = created_file.read_text(encoding="utf-8")
        assert "'newKdocsToken'" in created_text
        assert "oldKdocsToken" not in created_text
        assert "课程商品名='第40届念住禅法（初阶）【中心教室】'" in created_text
        assert "课程商品名='第39届念住禅法（初阶）【中心教室】'" not in created_text
        assert "XlPath(__file__).stem" in created_text

        duplicate_response = client.post(
            "/api/note-sheets/sheets/4/attendance-summary/generate-course-script",
            json={"row_index": 1},
        )
        assert duplicate_response.status_code == 409
    finally:
        _clear_user_override()


def test_attendance_summary_organizes_course_scripts_by_completion(
    client,
    session,
    tmp_path,
    monkeypatch,
):
    user = _create_user(session, username="note-sheet-attendance-script-organize-user")
    _override_user(user)

    def serial(year: int, month: int, day: int) -> str:
        return str((date(year, month, day) - date(1970, 1, 1)).days + 25569)

    courses_dir = tmp_path / "courses"
    completed_dir = courses_dir / "已完结"
    completed_dir.mkdir(parents=True)
    monkeypatch.setattr(note_sheets_api, "ATTENDANCE_COURSE_SCRIPT_DIR", courses_dir)

    active_completed = courses_dir / "d260401第39届念住.py"
    completed_active = completed_dir / "d260401第45届觉观.py"
    active_completed.write_text("# should move to completed\n", encoding="utf-8")
    completed_active.write_text("# should move to active\n", encoding="utf-8")

    try:
        workbook = WorkbookDocument(
            numeric_id=2,
            title="武陵禅寺网课考勤汇总",
            owner_user_id=user.id,
            created_by_user_id=user.id,
            updated_by_user_id=user.id,
        )
        sheet = SheetDocument(
            numeric_id=4,
            scope="notes",
            owner_type="note_sheet",
            owner_key="4",
            sheet_key="4",
            title="课程",
            owner_user_id=user.id,
            created_by_user_id=user.id,
            updated_by_user_id=user.id,
            document_json={
                "schema_version": 1,
                "columns": [
                    "课程类型",
                    "课程名称",
                    "在线考勤表",
                    "考勤负责人",
                    "备注",
                    "返款频次",
                    "课程开始日期",
                    "课程结束日期",
                    "考勤实际完成结点",
                    "报名费",
                    "报名人数",
                    "总报名费",
                ],
                "rows": [
                    [
                        "念住",
                        "第39届念住",
                        {"value": "20260401第39届念住", "link": {"url": "https://www.kdocs.cn/l/tokenDone"}},
                        "",
                        "",
                        "",
                        serial(2026, 4, 1),
                        "=G1+26",
                        serial(2026, 4, 30),
                        "620",
                        "7",
                        "=J1*K1",
                    ],
                    [
                        "觉观",
                        "第45届觉观",
                        {"value": "20260401第45届觉观", "link": {"url": "https://www.kdocs.cn/l/tokenActive"}},
                        "",
                        "",
                        "",
                        serial(2026, 4, 1),
                        "=G2+24",
                        "",
                        "499",
                        "54",
                        "=J2*K2",
                    ],
                    [
                        "念住",
                        "第40届念住",
                        {"value": "20260501第40届念住", "link": {"url": "https://www.kdocs.cn/l/tokenMissing"}},
                        "",
                        "",
                        "",
                        serial(2026, 5, 1),
                        "=G3+26",
                        "",
                        "620",
                        "",
                        "=J3*K3",
                    ],
                ],
                "cell_meta": {},
            },
        )
        session.add(workbook)
        session.add(sheet)
        session.commit()
        session.refresh(workbook)
        session.refresh(sheet)
        session.add(WorkbookSheetLink(workbook_id=workbook.id, sheet_id=sheet.id, order_index=0))
        session.commit()

        response = client.post("/api/note-sheets/sheets/4/attendance-summary/organize-course-scripts")
        assert response.status_code == 200
        payload = response.json()
        moved = {item["target_filename"]: item for item in payload["moved"]}
        skipped = {item["target_filename"]: item for item in payload["skipped"]}

        assert set(moved) == {"d260401第39届念住.py", "d260401第45届觉观.py"}
        assert moved["d260401第39届念住.py"]["completed"] is True
        assert moved["d260401第45届觉观.py"]["completed"] is False
        assert (completed_dir / "d260401第39届念住.py").exists()
        assert not active_completed.exists()
        assert (courses_dir / "d260401第45届觉观.py").exists()
        assert not completed_active.exists()
        assert skipped["d260501第40届念住.py"]["reason"] == "脚本不存在"
    finally:
        _clear_user_override()


def test_attendance_summary_updates_link_count_fields(client, session, monkeypatch):
    user = _create_user(session, username="note-sheet-attendance-link-count-user")
    _override_user(user)

    def serial(year: int, month: int, day: int) -> str:
        return str((date(year, month, day) - date(1970, 1, 1)).days + 25569)

    def fake_query_link_count(field_key: str, course_name: str):
        assert course_name == "d260401第39届念住"
        if field_key == "lesson_links":
            return 21, 21
        return 4, 3

    monkeypatch.setattr(note_sheets_api, "_query_attendance_link_count", fake_query_link_count)

    try:
        workbook = WorkbookDocument(
            numeric_id=2,
            title="武陵禅寺网课考勤汇总",
            owner_user_id=user.id,
            created_by_user_id=user.id,
            updated_by_user_id=user.id,
        )
        sheet = SheetDocument(
            numeric_id=4,
            scope="notes",
            owner_type="note_sheet",
            owner_key="4",
            sheet_key="4",
            title="课程",
            owner_user_id=user.id,
            created_by_user_id=user.id,
            updated_by_user_id=user.id,
            document_json={
                "schema_version": 1,
                "columns": [
                    "课程类型",
                    "课程名称",
                    "在线考勤表",
                    "考勤负责人",
                    "备注",
                    "返款频次",
                    "课程开始日期",
                    "课程结束日期",
                    "考勤实际完成结点",
                    "报名费",
                    "报名人数",
                    "总报名费",
                ],
                "rows": [
                    [
                        "念住",
                        "第39届念住",
                        {"value": "20260401第39届念住", "link": {"url": "https://www.kdocs.cn/l/token"}},
                        "",
                        "",
                        "",
                        serial(2026, 4, 1),
                        "=G1+26",
                        "",
                        "620",
                        "7",
                        "=J1*K1",
                    ],
                ],
                "header_groups": [[{"label": "课程基本信息", "colspan": 5}, {"label": "时间节点", "colspan": 4}, {"label": "促学金情况", "colspan": 3}]],
                "cell_meta": {
                    "0:11": {"style": {"text_color": "#ff0000"}},
                },
                "column_widths": [100] * 12,
            },
        )
        session.add(workbook)
        session.add(sheet)
        session.commit()
        session.refresh(workbook)
        session.refresh(sheet)
        session.add(WorkbookSheetLink(workbook_id=workbook.id, sheet_id=sheet.id, order_index=0))
        session.commit()
        initial_version = sheet.version

        stale_response = client.post(
            "/api/note-sheets/sheets/4/attendance-summary/update-link-counts",
            json={
                "base_version": initial_version + 1,
                "field_key": "lesson_links",
                "repair_with_remote_browser": False,
            },
        )
        assert stale_response.status_code == 409

        lesson_response = client.post(
            "/api/note-sheets/sheets/4/attendance-summary/update-link-counts",
            json={
                "base_version": initial_version,
                "field_key": "lesson_links",
                "repair_with_remote_browser": False,
            },
        )
        assert lesson_response.status_code == 200
        lesson_payload = lesson_response.json()
        assert lesson_payload["updated"][0]["lookup_name"] == "d260401第39届念住"
        assert lesson_payload["updated"][0]["value"] == "21"
        lesson_document = lesson_payload["sheet"]["document_json"]
        assert lesson_document["columns"][:7] == [
            "课程类型",
            "课程名称",
            "在线考勤表",
            "考勤负责人",
            "课次链接",
            "打卡链接",
            "备注",
        ]
        assert lesson_document["rows"][0][4] == "21"
        assert lesson_document["rows"][0][13] == "=L1*M1"
        assert lesson_document["cell_meta"]["0:13"]["style"]["text_color"] == "#ff0000"
        assert lesson_document["header_groups"][0][0]["colspan"] == 7

        clockin_response = client.post(
            "/api/note-sheets/sheets/4/attendance-summary/update-link-counts",
            json={
                "base_version": lesson_payload["sheet"]["version"],
                "field_key": "clockin_links",
                "repair_with_remote_browser": False,
            },
        )
        assert clockin_response.status_code == 200
        clockin_payload = clockin_response.json()
        assert clockin_payload["updated"][0]["value"] == "3/4"
        assert clockin_payload["sheet"]["document_json"]["rows"][0][5] == "3/4"
    finally:
        _clear_user_override()


def test_attendance_summary_link_count_repairs_nianzhu_jueguan_with_remote_step1(client, session, monkeypatch):
    user = _create_user(session, username="note-sheet-attendance-link-count-remote-user")
    _override_user(user)

    def serial(year: int, month: int, day: int) -> str:
        return str((date(year, month, day) - date(1970, 1, 1)).days + 25569)

    query_results = [(0, 0), (21, 21)]
    remote_calls: list[dict[str, object]] = []

    def fake_query_link_count(field_key: str, course_name: str):
        assert field_key == "lesson_links"
        assert course_name == "d260701第48届觉观"
        return query_results.pop(0)

    class FakeRemoteEntry:
        entry_id = "mi15-entry"
        name = "codepc_mi15"
        device_id = "mi15-device"
        mode = "remote"
        server_url = "http://192.168.31.15:8000"
        token = "token"

    def fake_step_runner_device(*args, **kwargs):
        return FakeRemoteEntry()

    def fake_remote_post(entry, *, path: str, payload: dict, timeout: int):
        remote_calls.append({"entry": entry, "path": path, "payload": payload, "timeout": timeout})
        return {"lesson_update_count": 21, "clockin_update_count": 0}

    monkeypatch.setattr(note_sheets_api, "_query_attendance_link_count", fake_query_link_count)
    monkeypatch.setattr(note_sheets_api, "get_attendance_course_data_step_runner_device", fake_step_runner_device)
    monkeypatch.setattr(note_sheets_api, "_post_remote_attendance_device_json", fake_remote_post)

    try:
        workbook = WorkbookDocument(numeric_id=2, title="武陵禅寺网课考勤汇总", owner_user_id=user.id)
        sheet = SheetDocument(
            numeric_id=4,
            scope="notes",
            owner_type="note_sheet",
            owner_key="4",
            sheet_key="4",
            title="课程",
            owner_user_id=user.id,
            document_json={
                "schema_version": 1,
                "columns": [
                    "课程类型",
                    "课程名称",
                    "在线考勤表",
                    "考勤负责人",
                    "备注",
                    "返款频次",
                    "课程开始日期",
                    "课程结束日期",
                    "考勤实际完成结点",
                    "报名费",
                    "报名人数",
                    "总报名费",
                ],
                "rows": [
                    [
                        "觉观",
                        "第48届觉观",
                        "第48届觉观",
                        "",
                        "",
                        "",
                        serial(2026, 7, 1),
                        "=G1+24",
                        "",
                        "499",
                        "",
                        "=J1*K1",
                    ],
                ],
                "cell_meta": {},
            },
        )
        session.add(workbook)
        session.add(sheet)
        session.commit()
        session.refresh(workbook)
        session.refresh(sheet)
        session.add(WorkbookSheetLink(workbook_id=workbook.id, sheet_id=sheet.id, order_index=0))
        session.commit()

        response = client.post(
            "/api/note-sheets/sheets/4/attendance-summary/update-link-counts",
            json={"base_version": sheet.version, "field_key": "lesson_links"},
        )

        assert response.status_code == 200
        payload = response.json()
        assert payload["updated"][0]["remote_repair_attempted"] is True
        assert payload["updated"][0]["value"] == "21"
        assert payload["sheet"]["document_json"]["rows"][0][4] == "21"
        assert remote_calls == [
            {
                "entry": remote_calls[0]["entry"],
                "path": "/api/device-control/attendance/nianzhu/step1",
                "payload": {
                    "course_name": "d260701第48届觉观",
                    "shop_id": 1,
                    "update_lessons": True,
                    "update_clockins": False,
                    "clockin_pattern": "",
                    "dynamic_clockin_plugin": "",
                    "close_browser": True,
                },
                "timeout": 1200,
            }
        ]
    finally:
        _clear_user_override()


def test_attendance_summary_link_count_keeps_existing_value_when_remote_repair_fails(client, session, monkeypatch):
    user = _create_user(session, username="note-sheet-attendance-link-count-remote-fail-user")
    _override_user(user)

    def serial(year: int, month: int, day: int) -> str:
        return str((date(year, month, day) - date(1970, 1, 1)).days + 25569)

    monkeypatch.setattr(note_sheets_api, "_query_attendance_link_count", lambda _field_key, _course_name: (0, 0))

    class FakeRemoteEntry:
        entry_id = "mi15-entry"
        name = "codepc_mi15"
        device_id = "mi15-device"
        mode = "remote"
        server_url = "http://192.168.31.15:8000"
        token = "token"

    monkeypatch.setattr(note_sheets_api, "get_attendance_course_data_step_runner_device", lambda *args, **kwargs: FakeRemoteEntry())

    def fake_remote_post(*args, **kwargs):
        raise RuntimeError("mi15 offline")

    monkeypatch.setattr(note_sheets_api, "_post_remote_attendance_device_json", fake_remote_post)

    try:
        workbook = WorkbookDocument(numeric_id=2, title="武陵禅寺网课考勤汇总", owner_user_id=user.id)
        sheet = SheetDocument(
            numeric_id=4,
            scope="notes",
            owner_type="note_sheet",
            owner_key="4",
            sheet_key="4",
            title="课程",
            owner_user_id=user.id,
            document_json={
                "schema_version": 1,
                "columns": [
                    "课程类型",
                    "课程名称",
                    "在线考勤表",
                    "考勤负责人",
                    "课次链接",
                    "打卡链接",
                    "备注",
                    "返款频次",
                    "课程开始日期",
                    "课程结束日期",
                    "考勤实际完成结点",
                    "报名费",
                    "报名人数",
                    "总报名费",
                ],
                "rows": [
                    [
                        "念住",
                        "第42届念住",
                        "第42届念住",
                        "",
                        "21",
                        "1",
                        "",
                        "",
                        serial(2026, 7, 1),
                        "=I1+26",
                        "",
                        "620",
                        "",
                        "=L1*M1",
                    ],
                ],
                "cell_meta": {},
            },
        )
        session.add(workbook)
        session.add(sheet)
        session.commit()
        session.refresh(workbook)
        session.refresh(sheet)
        session.add(WorkbookSheetLink(workbook_id=workbook.id, sheet_id=sheet.id, order_index=0))
        session.commit()

        response = client.post(
            "/api/note-sheets/sheets/4/attendance-summary/update-link-counts",
            json={"base_version": sheet.version, "field_key": "lesson_links"},
        )

        assert response.status_code == 200
        payload = response.json()
        assert payload["updated"] == []
        assert payload["skipped"][0]["remote_repair_attempted"] is True
        assert payload["skipped"][0]["reason"] == "远程补抓失败，已保留原链接数"
        assert payload["sheet"]["document_json"]["rows"][0][4] == "21"
    finally:
        _clear_user_override()


def test_workbook_save_as_template_and_duplicate(client, session):
    user = _create_user(session, username="note-sheet-save-as-user")
    _grant_feature_access(session, user_id=user.id, feature_key="notes.sheets")
    _override_user(user)

    try:
        workbook_response = client.post("/api/note-sheets/workbooks", json={"title": "原工作簿"})
        assert workbook_response.status_code == 200
        workbook_id = workbook_response.json()["id"]

        create_sheet_response = client.post(
            "/api/note-sheets/sheets",
            json={
                "title": "报名表",
                "workbook_id": workbook_id,
                "document_json": {
                    "schema_version": 1,
                    "columns": ["姓名", "手机号"],
                    "rows": [["时秋菊", "15641363033"]],
                    "column_configs": {"手机号": {"note": "保留原始手机号", "hidden": True}},
                },
            },
        )
        assert create_sheet_response.status_code == 200
        source_sheet = create_sheet_response.json()

        template_response = client.post(
            f"/api/note-sheets/workbooks/{workbook_id}/save-as",
            json={"mode": "template", "title": "原工作簿 模版"},
        )
        assert template_response.status_code == 200
        template_workbook = template_response.json()
        assert template_workbook["id"] != workbook_id
        assert template_workbook["title"] == "原工作簿 模版"
        assert len(template_workbook["sheets"]) == 1
        assert template_workbook["sheets"][0]["id"] != source_sheet["id"]

        template_sheet_detail = client.get(
            f"/api/note-sheets/sheets/{template_workbook['sheets'][0]['id']}",
            params={"paginate": False},
        )
        assert template_sheet_detail.status_code == 200
        assert template_sheet_detail.json()["document_json"]["rows"] == []
        assert template_sheet_detail.json()["document_json"]["column_configs"]["手机号"] == {
            "note": "保留原始手机号",
            "hidden": True,
        }

        duplicate_response = client.post(
            f"/api/note-sheets/workbooks/{workbook_id}/save-as",
            json={"mode": "duplicate", "title": "原工作簿 副本"},
        )
        assert duplicate_response.status_code == 200
        duplicate_workbook = duplicate_response.json()
        assert duplicate_workbook["id"] != workbook_id
        assert duplicate_workbook["title"] == "原工作簿 副本"
        assert len(duplicate_workbook["sheets"]) == 1
        assert duplicate_workbook["sheets"][0]["id"] not in {workbook_id, source_sheet["id"]}

        duplicate_sheet_detail = client.get(
            f"/api/note-sheets/sheets/{duplicate_workbook['sheets'][0]['id']}",
            params={"paginate": False},
        )
        assert duplicate_sheet_detail.status_code == 200
        assert duplicate_sheet_detail.json()["document_json"]["rows"] == [["时秋菊", "15641363033"]]
        assert duplicate_sheet_detail.json()["document_json"]["column_configs"]["手机号"] == {
            "note": "保留原始手机号",
            "hidden": True,
        }
    finally:
        _clear_user_override()


def test_workbook_save_as_template_preserves_linked_user_id_as_standard_field(client, session):
    user = _create_user(session, username="note-sheet-save-as-registration-template-user")
    _grant_feature_access(session, user_id=user.id, feature_key="notes.sheets")
    _override_user(user)

    try:
        workbook_response = client.post("/api/note-sheets/workbooks", json={"title": "报名模板源"})
        assert workbook_response.status_code == 200
        workbook_id = workbook_response.json()["id"]

        columns = [
            "分组",
            "序号",
            "提交时间",
            "姓名",
            "用户ID",
            "匹配得分",
            "参考信息",
            "关联用户ID",
            "追踪状态",
            "冻结时间",
            "规则版本",
        ]
        create_sheet_response = client.post(
            "/api/note-sheets/sheets",
            json={
                "title": "报名表",
                "workbook_id": workbook_id,
                "document_json": {
                    "schema_version": 1,
                    "columns": columns,
                    "rows": [["5月4日", "1", "06/01 08:00", "阿紫", "u_old", "90", "", "u_new", "追踪中", "", "当前规则"]],
                    "data_start_row": 2,
                    "grid_rows": [
                        columns,
                        ["", "", "", "", "", "综合更新", "其他备注", "", "", "", ""],
                        ["5月4日", "1", "06/01 08:00", "阿紫", "u_old", "90", "", "u_new", "追踪中", "", "当前规则"],
                    ],
                    "column_configs": {
                        "用户ID": {"header_background_color": "#9DC3E6"},
                        "关联用户ID": {"header_background_color": "#F4B183", "width_mode": "fixed"},
                        "追踪状态": {"hidden": True},
                        "规则版本": {"hidden": True},
                    },
                    "cell_meta": {
                        "0:7": {"style": {"background_color": "#F4B183"}},
                        "2:0": {"style": {"background_color": "#FFFF00"}},
                    },
                },
            },
        )
        assert create_sheet_response.status_code == 200

        template_response = client.post(
            f"/api/note-sheets/workbooks/{workbook_id}/save-as",
            json={"mode": "template", "title": "报名模板"},
        )
        assert template_response.status_code == 200
        template_sheet_id = template_response.json()["sheets"][0]["id"]

        template_sheet_detail = client.get(
            f"/api/note-sheets/sheets/{template_sheet_id}",
            params={"paginate": False},
        )
        assert template_sheet_detail.status_code == 200
        document = template_sheet_detail.json()["document_json"]
        assert document["rows"] == []
        assert len(document["grid_rows"]) == 2
        assert "关联用户ID" in document["columns"]
        assert "追踪状态" not in document["columns"]
        assert "冻结时间" not in document["columns"]
        assert "规则版本" not in document["columns"]
        assert document["column_configs"]["关联用户ID"]["header_background_color"] == "#9DC3E6"
        assert document["column_configs"]["关联用户ID"]["font_family"] == "monospace"
        assert (
            document["column_configs"]["关联用户ID"]["note"]
            == "有的用户账号数据源不统一，这里可以逗号隔开填写其他相关id，会合并到主id数据中汇总进度"
        )
        assert "追踪状态" not in document["column_configs"]
        assert "规则版本" not in document["column_configs"]
        assert "2:0" not in document["cell_meta"]
        linked_index = document["columns"].index("关联用户ID")
        assert document["cell_meta"][f"0:{linked_index}"]["style"]["background_color"] == "#9DC3E6"
    finally:
        _clear_user_override()


def test_workbook_save_as_generated_attendance_sheet_defaults_to_anonymous_viewer(client, session):
    user = _create_user(session, username="note-sheet-save-as-attendance-user")
    _grant_feature_access(session, user_id=user.id, feature_key="notes.sheets")
    _override_user(user)

    try:
        workbook_response = client.post("/api/note-sheets/workbooks", json={"title": "课程工作簿"})
        assert workbook_response.status_code == 200
        workbook_id = workbook_response.json()["id"]

        create_sheet_response = client.post(
            "/api/note-sheets/sheets",
            json={
                "title": "考勤表",
                "workbook_id": workbook_id,
                "document_json": {"schema_version": 1, "columns": ["姓名"], "rows": []},
            },
        )
        assert create_sheet_response.status_code == 200
        source_sheet_id = create_sheet_response.json()["id"]
        source_sheet = session.exec(select(SheetDocument).where(SheetDocument.numeric_id == source_sheet_id)).one()
        source_sheet.owner_type = "course_workbook"
        source_sheet.owner_key = "course-a"
        source_sheet.sheet_key = "attendance"
        session.add(source_sheet)
        session.commit()

        response = client.post(
            f"/api/note-sheets/workbooks/{workbook_id}/save-as",
            json={"mode": "duplicate", "title": "课程工作簿 副本"},
        )

        assert response.status_code == 200
        copied_sheet_id = response.json()["sheets"][0]["id"]
        grant = session.exec(
            select(ResourceAccessGrant)
            .where(ResourceAccessGrant.resource_type == "sheet")
            .where(ResourceAccessGrant.resource_id == str(copied_sheet_id))
            .where(ResourceAccessGrant.subject_key == "anonymous")
        ).first()
        assert grant is not None
        assert grant.subject_type == "anonymous"
        assert grant.role == "viewer"
    finally:
        _clear_user_override()


def test_note_sheet_pagination_load_and_page_patch_save(client, session):
    user = _create_user(session, username="note-sheet-pagination-user")
    _grant_feature_access(session, user_id=user.id, feature_key="notes.sheets")
    _override_user(user)

    try:
        create_sheet_response = client.post(
            "/api/note-sheets/sheets",
            json={
                "title": "大表格",
                "document_json": {
                    "schema_version": 1,
                    "columns": ["序号", "内容"],
                    "column_ids": ["cid-seq", "cid-content"],
                    "rows": [[str(index), f"row-{index}"] for index in range(1, 251)],
                    "row_ids": [f"rid-{index}" for index in range(1, 251)],
                    "view_settings": {
                        "pagination": {
                            "enabled": True,
                            "page_size": 100,
                        },
                    },
                },
            },
        )
        assert create_sheet_response.status_code == 200
        sheet_id = create_sheet_response.json()["id"]

        page2_response = client.get(
            f"/api/note-sheets/sheets/{sheet_id}",
            params={"page": 2, "page_size": 100},
        )
        assert page2_response.status_code == 200
        page2_detail = page2_response.json()
        _assert_pagination_contains(page2_detail["pagination"], {
            "page": 2,
            "page_size": 100,
            "total_rows": 250,
            "page_count": 3,
            "row_offset": 100,
            "loaded_row_count": 100,
        })
        assert len(page2_detail["document_json"]["row_ids"]) == 250
        assert page2_detail["document_json"]["rows"][0] == ["101", "row-101"]
        assert page2_detail["document_json"]["rows"][-1] == ["200", "row-200"]

        edited_rows = page2_detail["document_json"]["rows"][2:]
        edited_row_ids = page2_detail["document_json"]["row_ids"][102:200]
        save_response = client.put(
            f"/api/note-sheets/sheets/{sheet_id}",
            json={
                "document_json": {
                    "schema_version": 1,
                    "columns": ["序号", "内容"],
                    "column_ids": ["cid-seq", "cid-content"],
                    "rows": edited_rows,
                    "row_ids": edited_row_ids,
                },
                "page_patch": {
                    "page": 2,
                    "page_size": 100,
                    "row_offset": 100,
                    "loaded_row_count": 100,
                },
            },
        )
        assert save_response.status_code == 200
        save_detail = save_response.json()
        _assert_pagination_contains(save_detail["pagination"], {
            "page": 2,
            "page_size": 100,
            "total_rows": 248,
            "page_count": 3,
            "row_offset": 100,
            "loaded_row_count": 98,
        })
        assert len(save_detail["document_json"]["rows"]) == 98
        assert save_detail["document_json"]["rows"][0] == ["103", "row-103"]
        stored_sheet = session.exec(select(SheetDocument).where(SheetDocument.numeric_id == sheet_id)).one()
        assert stored_sheet.document_json["row_ids"][:100] == [f"rid-{index}" for index in range(1, 101)]
        assert stored_sheet.document_json["row_ids"][100:198] == [f"rid-{index}" for index in range(103, 201)]
        assert stored_sheet.document_json["row_ids"][198:] == [f"rid-{index}" for index in range(201, 251)]
        assert stored_sheet.document_json["column_ids"] == ["cid-seq", "cid-content"]

        normalized_page2_response = client.get(
            f"/api/note-sheets/sheets/{sheet_id}",
            params={"page": 2, "page_size": 100},
        )
        assert normalized_page2_response.status_code == 200
        normalized_page2 = normalized_page2_response.json()
        assert normalized_page2["pagination"]["loaded_row_count"] == 100
        assert normalized_page2["document_json"]["rows"][0] == ["103", "row-103"]
        assert normalized_page2["document_json"]["rows"][-1] == ["202", "row-202"]

        page3_response = client.get(
            f"/api/note-sheets/sheets/{sheet_id}",
            params={"page": 3, "page_size": 100},
        )
        assert page3_response.status_code == 200
        page3_detail = page3_response.json()
        assert page3_detail["pagination"]["loaded_row_count"] == 48
        assert page3_detail["document_json"]["rows"][0] == ["203", "row-203"]
    finally:
        _clear_user_override()


def test_note_sheet_query_filters_before_pagination_and_saves_by_row_indexes(client, session):
    user = _create_user(session, username="note-sheet-filtered-pagination-user")
    _grant_feature_access(session, user_id=user.id, feature_key="notes.sheets")
    _override_user(user)

    try:
        create_sheet_response = client.post(
            "/api/note-sheets/sheets",
            json={
                "title": "筛选分页表格",
                "document_json": {
                    "schema_version": 1,
                    "columns": ["分类", "内容", "金额"],
                    "rows": [
                        ["支出", f"expense-{index}", str(index)]
                        if index % 2 == 0
                        else ["收入", f"income-{index}", str(index)]
                        for index in range(1, 121)
                    ],
                    "column_configs": {
                        "分类": {
                            "filter_enabled": True,
                            "value_mode": "fixed_options",
                        },
                    },
                    "view_settings": {
                        "pagination": {
                            "enabled": True,
                            "page_size": 50,
                        },
                    },
                },
            },
        )
        assert create_sheet_response.status_code == 200
        sheet_id = create_sheet_response.json()["id"]

        query_response = client.post(
            f"/api/note-sheets/sheets/{sheet_id}/query",
            json={
                "page": 1,
                "page_size": 50,
                "paginate": True,
                "column_filters": {
                    "分类": {
                        "excludedValues": ["收入"],
                    },
                },
            },
        )
        assert query_response.status_code == 200
        detail = query_response.json()
        assert detail["pagination"] == {
            "page": 1,
            "page_size": 50,
            "total_rows": 60,
            "unfiltered_total_rows": 120,
            "page_count": 2,
            "row_offset": 0,
            "loaded_row_count": 50,
            "row_indexes": list(range(1, 100, 2)),
        }
        assert len(detail["document_json"]["rows"]) == 50
        assert detail["document_json"]["rows"][0] == ["支出", "expense-2", "2"]
        assert detail["document_json"]["rows"][-1] == ["支出", "expense-100", "100"]

        edited_rows = detail["document_json"]["rows"]
        edited_rows[0] = ["支出", "expense-2-edited", "2"]
        save_response = client.put(
            f"/api/note-sheets/sheets/{sheet_id}",
            json={
                "document_json": {
                    **detail["document_json"],
                    "rows": edited_rows,
                },
                "page_patch": {
                    "page": 1,
                    "page_size": 50,
                    "row_offset": 0,
                    "loaded_row_count": 50,
                    "row_indexes": detail["pagination"]["row_indexes"],
                },
            },
        )
        assert save_response.status_code == 200

        full_response = client.get(
            f"/api/note-sheets/sheets/{sheet_id}",
            params={"paginate": False},
        )
        assert full_response.status_code == 200
        full_rows = full_response.json()["document_json"]["rows"]
        assert full_rows[0] == ["收入", "income-1", "1"]
        assert full_rows[1] == ["支出", "expense-2-edited", "2"]
        assert full_rows[2] == ["收入", "income-3", "3"]
    finally:
        _clear_user_override()


def test_note_sheet_filtered_page_patch_deletes_by_row_indexes(client, session):
    user = _create_user(session, username="note-sheet-filtered-delete-user")
    _grant_feature_access(session, user_id=user.id, feature_key="notes.sheets")
    _override_user(user)

    try:
        create_sheet_response = client.post(
            "/api/note-sheets/sheets",
            json={
                "title": "筛选分页删除",
                "document_json": {
                    "schema_version": 1,
                    "columns": ["分类", "内容"],
                    "rows": [
                        ["收入", "income-1"],
                        ["支出", "expense-2"],
                        ["收入", "income-3"],
                        ["支出", "expense-4"],
                        ["收入", "income-5"],
                        ["支出", "expense-6"],
                    ],
                    "grid_rows": [
                        ["分类", "内容"],
                        ["收入", "income-1"],
                        ["支出", "expense-2"],
                        ["收入", "income-3"],
                        ["支出", "expense-4"],
                        ["收入", "income-5"],
                        ["支出", "expense-6"],
                    ],
                    "data_start_row": 1,
                    "field_row_index": 0,
                    "column_configs": {
                        "分类": {
                            "filter_enabled": True,
                            "value_mode": "fixed_options",
                        },
                    },
                    "view_settings": {
                        "pagination": {
                            "enabled": True,
                            "page_size": 10,
                        },
                    },
                },
            },
        )
        assert create_sheet_response.status_code == 200
        sheet_id = create_sheet_response.json()["id"]

        query_response = client.post(
            f"/api/note-sheets/sheets/{sheet_id}/query",
            json={
                "page": 1,
                "page_size": 10,
                "paginate": True,
                "column_filters": {
                    "分类": {
                        "excludedValues": ["收入"],
                    },
                },
            },
        )
        assert query_response.status_code == 200
        detail = query_response.json()
        assert detail["pagination"]["row_indexes"] == [1, 3, 5]
        assert detail["document_json"]["rows"] == [
            ["支出", "expense-2"],
            ["支出", "expense-4"],
            ["支出", "expense-6"],
        ]

        save_response = client.put(
            f"/api/note-sheets/sheets/{sheet_id}",
            json={
                "document_json": {
                    **detail["document_json"],
                    "rows": [
                        ["支出", "expense-2-edited"],
                        ["支出", "expense-6"],
                    ],
                    "grid_rows": [
                        ["分类", "内容"],
                        ["支出", "expense-2-edited"],
                        ["支出", "expense-6"],
                    ],
                },
                "page_patch": {
                    "page": 1,
                    "page_size": 10,
                    "row_offset": 0,
                    "loaded_row_count": 2,
                    "row_indexes": [1, 5],
                    "deleted_row_indexes": [3],
                },
            },
        )
        assert save_response.status_code == 200

        full_response = client.get(
            f"/api/note-sheets/sheets/{sheet_id}",
            params={"paginate": False},
        )
        assert full_response.status_code == 200
        full_document = full_response.json()["document_json"]
        assert full_document["rows"] == [
            ["收入", "income-1"],
            ["支出", "expense-2-edited"],
            ["收入", "income-3"],
            ["收入", "income-5"],
            ["支出", "expense-6"],
        ]
        assert full_document["grid_rows"] == [
            ["分类", "内容"],
            ["收入", "income-1"],
            ["支出", "expense-2-edited"],
            ["收入", "income-3"],
            ["收入", "income-5"],
            ["支出", "expense-6"],
        ]
    finally:
        _clear_user_override()


def test_note_sheet_filtered_page_patch_preserves_unloaded_cell_meta(client, session):
    user = _create_user(session, username="note-sheet-filtered-cell-meta-user")
    _grant_feature_access(session, user_id=user.id, feature_key="notes.sheets")
    _override_user(user)

    try:
        create_sheet_response = client.post(
            "/api/note-sheets/sheets",
            json={
                "title": "筛选分页样式",
                "document_json": {
                    "schema_version": 1,
                    "columns": ["分类", "内容"],
                    "rows": [
                        ["收入", "income-1"],
                        ["支出", "expense-2"],
                        ["收入", "income-3"],
                        ["支出", "expense-4"],
                        ["收入", "income-5"],
                        ["支出", "expense-6"],
                    ],
                    "grid_rows": [
                        ["分类", "内容"],
                        ["收入", "income-1"],
                        ["支出", "expense-2"],
                        ["收入", "income-3"],
                        ["支出", "expense-4"],
                        ["收入", "income-5"],
                        ["支出", "expense-6"],
                    ],
                    "data_start_row": 1,
                    "field_row_index": 0,
                    "cell_meta": {
                        "1:1": {"style": {"background_color": "#eeeeee"}},
                        "2:1": {"style": {"background_color": "#ff0000"}},
                        "4:1": {"style": {"background_color": "#00ff00"}},
                        "5:0": {"style": {"background_color": "#ffff00"}},
                    },
                    "column_configs": {
                        "分类": {
                            "filter_enabled": True,
                            "value_mode": "fixed_options",
                        },
                    },
                    "view_settings": {
                        "pagination": {
                            "enabled": True,
                            "page_size": 10,
                        },
                    },
                },
            },
        )
        assert create_sheet_response.status_code == 200
        sheet_id = create_sheet_response.json()["id"]

        query_response = client.post(
            f"/api/note-sheets/sheets/{sheet_id}/query",
            json={
                "page": 1,
                "page_size": 10,
                "paginate": True,
                "column_filters": {
                    "分类": {
                        "excludedValues": ["收入"],
                    },
                },
            },
        )
        assert query_response.status_code == 200
        detail = query_response.json()
        assert detail["pagination"]["row_indexes"] == [1, 3, 5]

        save_response = client.put(
            f"/api/note-sheets/sheets/{sheet_id}",
            json={
                "document_json": {
                    **detail["document_json"],
                    "rows": [
                        ["支出", "expense-2-edited"],
                        ["支出", "expense-6"],
                    ],
                    "grid_rows": [
                        ["分类", "内容"],
                        ["支出", "expense-2-edited"],
                        ["支出", "expense-6"],
                    ],
                    "cell_meta": {
                        "2:1": {"style": {"background_color": "#0000ff"}},
                    },
                },
                "page_patch": {
                    "page": 1,
                    "page_size": 10,
                    "row_offset": 0,
                    "loaded_row_count": 2,
                    "row_indexes": [1, 5],
                    "deleted_row_indexes": [3],
                },
            },
        )
        assert save_response.status_code == 200

        full_response = client.get(
            f"/api/note-sheets/sheets/{sheet_id}",
            params={"paginate": False},
        )
        assert full_response.status_code == 200
        cell_meta = full_response.json()["document_json"]["cell_meta"]
        assert cell_meta == {
            "1:1": {"style": {"background_color": "#eeeeee"}},
            "2:1": {"style": {"background_color": "#0000ff"}},
            "4:0": {"style": {"background_color": "#ffff00"}},
        }
    finally:
        _clear_user_override()


def test_note_sheet_respects_document_pagination_settings(client, session):
    user = _create_user(session, username="note-sheet-pagination-settings-user")
    _grant_feature_access(session, user_id=user.id, feature_key="notes.sheets")
    _override_user(user)

    try:
        create_sheet_response = client.post(
            "/api/note-sheets/sheets",
            json={
                "title": "可配置分页表格",
                "document_json": {
                    "schema_version": 1,
                    "columns": ["序号", "内容"],
                    "rows": [[str(index), f"row-{index}"] for index in range(1, 251)],
                    "view_settings": {
                        "pagination": {
                            "enabled": True,
                            "page_size": 50,
                        },
                    },
                },
            },
        )
        assert create_sheet_response.status_code == 200
        sheet_id = create_sheet_response.json()["id"]

        auto_page_response = client.get(f"/api/note-sheets/sheets/{sheet_id}")
        assert auto_page_response.status_code == 200
        auto_page_detail = auto_page_response.json()
        _assert_pagination_contains(auto_page_detail["pagination"], {
            "page": 1,
            "page_size": 50,
            "total_rows": 250,
            "page_count": 5,
            "row_offset": 0,
            "loaded_row_count": 50,
        })
        assert len(auto_page_detail["document_json"]["rows"]) == 50

        disable_pagination_response = client.put(
            f"/api/note-sheets/sheets/{sheet_id}",
            json={
                "document_json": {
                    **auto_page_detail["document_json"],
                    "view_settings": {
                        "pagination": {
                            "enabled": False,
                            "page_size": 50,
                        },
                    },
                },
                "page_patch": {
                    "page": 1,
                    "page_size": 50,
                    "row_offset": 0,
                    "loaded_row_count": 50,
                },
            },
        )
        assert disable_pagination_response.status_code == 200
        disable_detail = disable_pagination_response.json()
        assert disable_detail["pagination"] is None
        assert disable_detail["document_json"]["view_settings"]["pagination"]["enabled"] is False

        full_sheet_response = client.get(f"/api/note-sheets/sheets/{sheet_id}")
        assert full_sheet_response.status_code == 200
        full_sheet_detail = full_sheet_response.json()
        assert full_sheet_detail["pagination"] is None
        assert len(full_sheet_detail["document_json"]["rows"]) == 250
        assert full_sheet_detail["document_json"]["rows"][0] == ["1", "row-1"]
        assert full_sheet_detail["document_json"]["rows"][-1] == ["250", "row-250"]
    finally:
        _clear_user_override()


def test_note_sheet_column_options_use_full_sheet_when_paginated(client, session):
    user = _create_user(session, username="note-sheet-column-options-user")
    _grant_feature_access(session, user_id=user.id, feature_key="notes.sheets")
    _override_user(user)

    try:
        create_sheet_response = client.post(
            "/api/note-sheets/sheets",
            json={
                "title": "选项统计表格",
                "document_json": {
                    "schema_version": 1,
                    "columns": ["来源", "金额"],
                    "rows": [
                        ["支付宝", "10"],
                        ["微信", "20"],
                        ["支付宝", "30"],
                        ["微信", "40"],
                        ["支付宝", "50"],
                        ["", "60"],
                    ],
                    "view_settings": {
                        "pagination": {
                            "enabled": True,
                            "page_size": 2,
                        },
                    },
                },
            },
        )
        assert create_sheet_response.status_code == 200
        sheet_id = create_sheet_response.json()["id"]

        page_response = client.get(f"/api/note-sheets/sheets/{sheet_id}")
        assert page_response.status_code == 200
        assert page_response.json()["document_json"]["rows"] == [
            ["支付宝", "10"],
            ["微信", "20"],
        ]

        options_response = client.get(
            f"/api/note-sheets/sheets/{sheet_id}/column-options",
            params={"column_index": 0},
        )
        assert options_response.status_code == 200
        assert options_response.json() == {
            "column_index": 0,
            "header": "来源",
            "total_rows": 6,
            "options": [
                {"value": "支付宝", "label": "支付宝", "count": 3},
                {"value": "微信", "label": "微信", "count": 2},
                {"value": "", "label": "(空白)", "count": 1},
            ],
        }
    finally:
        _clear_user_override()


def test_note_sheet_inline_cell_helpers_parse_stringified_link_cells():
    cell = "{'value': '20260412禅宗12期一阶', 'link': {'url': 'https://www.kdocs.cn/l/copnS6juyN2T'}}"

    assert note_sheets_api._extract_cell_value(cell) == "20260412禅宗12期一阶"
    assert note_sheets_api._normalize_sheet_text(cell) == "20260412禅宗12期一阶"
    assert note_sheets_api._inline_cell_link_url(cell) == "https://www.kdocs.cn/l/copnS6juyN2T"


def test_attendance_summary_update_repairs_missing_online_sheet_links(client, session, tmp_path, monkeypatch):
    user = _create_user(session, username="note-sheet-attendance-online-link-repair-user")
    _grant_feature_access(session, user_id=user.id, feature_key="notes.sheets")
    _override_user(user)

    def serial(year: int, month: int, day: int) -> str:
        return str((date(year, month, day) - date(1970, 1, 1)).days + 25569)

    courses_dir = tmp_path / "courses"
    courses_dir.mkdir()
    monkeypatch.setattr(note_sheets_api, "ATTENDANCE_COURSE_SCRIPT_DIR", courses_dir)
    (courses_dir / "d260601第41届念住.py").write_text(
        "\n".join([
            "from xlsln.kq5034.courses.codeyun_course import CodeYunCourseSheets, CodeYunSheetRef",
            "",
            "class 考勤课程:",
            "    def __init__(self):",
            "        self.sheets = CodeYunCourseSheets(",
            "            registration=CodeYunSheetRef(10, 54606, \"报名表\"),",
            "            attendance=CodeYunSheetRef(10, 54605, \"考勤表\"),",
            "        )",
        ]),
        encoding="utf-8",
    )

    current_document_json = {
        "schema_version": 1,
        "columns": ["课程类型", "课程名称", "在线考勤表", "课程开始日期"],
        "rows": [
            ["念住", "第41届念住", "第41届念住", serial(2026, 6, 1)],
            ["念住", "第40届念住", {"value": "第40届念住", "link": {"url": "https://www.kdocs.cn/l/existingOld"}}, serial(2026, 5, 1)],
        ],
        "grid_rows": [
            ["课程类型", "课程名称", "在线考勤表", "课程开始日期"],
            ["念住", "第41届念住", "第41届念住", serial(2026, 6, 1)],
            ["念住", "第40届念住", {"value": "第40届念住", "link": {"url": "https://www.kdocs.cn/l/existingOld"}}, serial(2026, 5, 1)],
        ],
        "data_start_row": 1,
        "field_row_index": 0,
    }
    incoming_document_json = {
        **current_document_json,
        "rows": [
            ["念住", "第41届念住", "第41届念住", serial(2026, 6, 1)],
            ["念住", "第40届念住", "第40届念住", serial(2026, 5, 1)],
        ],
        "grid_rows": [
            ["课程类型", "课程名称", "在线考勤表", "课程开始日期"],
            ["念住", "第41届念住", "第41届念住", serial(2026, 6, 1)],
            ["念住", "第40届念住", "第40届念住", serial(2026, 5, 1)],
        ],
    }

    try:
        workbook = WorkbookDocument(
            numeric_id=2,
            title="武陵禅寺网课考勤汇总",
            owner_user_id=user.id,
            created_by_user_id=user.id,
            updated_by_user_id=user.id,
        )
        sheet = SheetDocument(
            numeric_id=4,
            scope="notes",
            owner_type="note_sheet",
            owner_key="4",
            sheet_key="4",
            title="课程",
            owner_user_id=user.id,
            created_by_user_id=user.id,
            updated_by_user_id=user.id,
            document_json=current_document_json,
        )
        session.add(workbook)
        session.add(sheet)
        session.commit()
        session.refresh(workbook)
        session.refresh(sheet)
        session.add(WorkbookSheetLink(workbook_id=workbook.id, sheet_id=sheet.id, order_index=0))
        session.commit()

        response = client.put(
            "/api/note-sheets/sheets/4",
            params={"workbook_id": 2},
            json={
                "title": "课程",
                "document_json": incoming_document_json,
                "base_version": sheet.version,
            },
        )

        assert response.status_code == 200
        saved_document = response.json()["document_json"]
        assert saved_document["rows"][0][2] == {
            "value": "第41届念住",
            "link": {"url": "/workbook/10?sheet=54605"},
        }
        assert saved_document["grid_rows"][1][2] == saved_document["rows"][0][2]
        assert saved_document["rows"][1][2] == {
            "value": "第40届念住",
            "link": {"url": "https://www.kdocs.cn/l/existingOld"},
        }
        assert saved_document["grid_rows"][2][2] == saved_document["rows"][1][2]
    finally:
        _clear_user_override()


def test_note_sheet_legacy_link_migration_is_one_way():
    document = {
        "schema_version": 1,
        "columns": ["序号", "链接"],
        "rows": [["1", "row-1"], ["2", "row-2"]],
        "grid_rows": [["序号", "链接"], ["1", "row-1"], ["2", "row-2"]],
        "data_start_row": 1,
        "cell_meta": {
            "1:1": {"link": {"url": "https://example.com/meta-1"}},
            "2:1": {"style": {"background_color": "#f8fafc"}, "link": {"url": "https://example.com/meta-2"}},
        },
        "entity_columns": [{"id": "c0", "header": "序号"}, {"id": "c1", "header": "链接"}],
        "entity_rows": [{"id": "h0", "kind": "field"}, {"id": "r1", "kind": "data"}, {"id": "r2", "kind": "data"}],
        "entity_cells": {
            "r1": {"c1": {"value": "row-1", "link": {"url": "https://example.com/entity-1"}}},
        },
    }

    migrated, stats = note_sheets_api.note_sheet_inline_links.canonicalize_sheet_document_inline_links(
        document,
        migrate_legacy_links=True,
        strip_legacy_links=True,
    )

    assert stats["legacy"] == 2
    assert migrated["rows"] == [
        ["1", {"value": "row-1", "link": {"url": "https://example.com/entity-1"}}],
        ["2", {"value": "row-2", "link": {"url": "https://example.com/meta-2"}}],
    ]
    assert migrated["grid_rows"][1:] == migrated["rows"]
    assert migrated["cell_meta"] == {"2:1": {"style": {"background_color": "#f8fafc"}}}
    assert migrated["entity_cells"] == {"r1": {"c1": {"value": "row-1"}}}

    runtime_normalized = note_sheets_api._normalize_document_json(document)
    assert runtime_normalized["rows"] == [
        ["1", {"value": "row-1", "link": {"url": "https://example.com/entity-1"}}],
        ["2", {"value": "row-2", "link": {"url": "https://example.com/meta-2"}}],
    ]
    assert note_sheets_api._get_document_cell_link_url(runtime_normalized, 0, 1) == "https://example.com/entity-1"
    assert runtime_normalized["cell_meta"] == {"2:1": {"style": {"background_color": "#f8fafc"}}}
    assert runtime_normalized["entity_cells"] == {"r1": {"c1": {"value": "row-1"}}}


def test_note_sheet_create_migrates_legacy_links_before_stripping(client, session):
    user = _create_user(session, username="note-sheet-legacy-link-create-user")
    _grant_feature_access(session, user_id=user.id, feature_key="notes.sheets")
    _override_user(user)

    try:
        response = client.post(
            "/api/note-sheets/sheets",
            json={
                "title": "旧链接兼容表",
                "document_json": {
                    "schema_version": 1,
                    "columns": ["序号", "在线考勤表"],
                    "rows": [["1", "20260601第41届念住"]],
                    "grid_rows": [["序号", "在线考勤表"], ["1", "20260601第41届念住"]],
                    "data_start_row": 1,
                    "field_row_index": 0,
                    "cell_meta": {
                        "1:1": {"link": {"url": "https://www.kdocs.cn/l/courseSheetToken"}},
                    },
                    "entity_columns": [{"id": "c0", "header": "序号"}, {"id": "c1", "header": "在线考勤表"}],
                    "entity_rows": [{"id": "h0", "kind": "field"}, {"id": "r1", "kind": "data"}],
                    "entity_cells": {
                        "r1": {"c1": {"value": "20260601第41届念住", "link": {"url": "https://www.kdocs.cn/l/courseSheetToken"}}},
                    },
                },
            },
        )

        assert response.status_code == 200
        document = response.json()["document_json"]
        linked_cell = {
            "value": "20260601第41届念住",
            "link": {"url": "https://www.kdocs.cn/l/courseSheetToken"},
        }
        assert document["rows"] == [["1", linked_cell]]
        assert document["grid_rows"][1] == ["1", linked_cell]
        assert document["cell_meta"] == {}
        assert document["entity_cells"] == {"r1": {"c1": {"value": "20260601第41届念住"}}}

        persisted = session.exec(select(SheetDocument).where(SheetDocument.id == str(response.json()["id"]))).one()
        assert persisted.document_json["rows"] == [["1", linked_cell]]
        assert persisted.document_json["cell_meta"] == {}
        assert persisted.document_json["entity_cells"] == {"r1": {"c1": {"value": "20260601第41届念住"}}}
    finally:
        _clear_user_override()


def test_note_sheet_legacy_links_without_rows_are_preserved():
    document = {
        "cell_meta": {"0:0": {"link": {"url": "https://example.com/orphan-link"}}},
        "entity_cells": {"r1": {"c1": {"link": {"url": "https://example.com/entity-link"}}}},
    }

    normalized = note_sheets_api._normalize_document_json(document)

    assert normalized == document


def test_note_sheet_sort_action_reorders_full_sheet_and_returns_first_page(client, session):
    user = _create_user(session, username="note-sheet-sort-user")
    _grant_feature_access(session, user_id=user.id, feature_key="notes.sheets")
    _override_user(user)

    try:
        create_sheet_response = client.post(
            "/api/note-sheets/sheets",
            json={
                "title": "排序表格",
                "document_json": {
                    "schema_version": 1,
                    "columns": ["序号", "内容"],
                    "rows": [
                        ["10", {"value": "row-10", "link": {"url": "https://example.com/row-10"}}],
                        ["2", "row-2"],
                        ["1", {"value": "row-1", "link": {"url": "https://example.com/row-1"}}],
                        ["", "row-empty"],
                        ["11", "row-11"],
                    ],
                    "cell_meta": {},
                    "view_settings": {
                        "pagination": {
                            "enabled": True,
                            "page_size": 2,
                        },
                    },
                },
            },
        )
        assert create_sheet_response.status_code == 200
        sheet_id = create_sheet_response.json()["id"]
        initial_version = create_sheet_response.json()["version"]

        asc_response = client.post(
            f"/api/note-sheets/sheets/{sheet_id}/sort",
            json={
                "base_version": initial_version,
                "column_index": 0,
                "direction": "asc",
            },
        )
        assert asc_response.status_code == 200
        asc_detail = asc_response.json()
        _assert_pagination_contains(asc_detail["pagination"], {
            "page": 1,
            "page_size": 2,
            "total_rows": 5,
            "page_count": 3,
            "row_offset": 0,
            "loaded_row_count": 2,
        })
        assert asc_detail["document_json"]["rows"] == [
            ["1", {"value": "row-1", "link": {"url": "https://example.com/row-1"}}],
            ["2", "row-2"],
        ]
        assert asc_detail["document_json"]["cell_meta"] == {}
        assert asc_detail["version"] == initial_version + 1

        stale_sort_response = client.post(
            f"/api/note-sheets/sheets/{sheet_id}/sort",
            json={
                "base_version": initial_version,
                "column_index": 0,
                "direction": "desc",
            },
        )
        assert stale_sort_response.status_code == 409

        page2_response = client.get(
            f"/api/note-sheets/sheets/{sheet_id}",
            params={"page": 2, "page_size": 2},
        )
        assert page2_response.status_code == 200
        assert page2_response.json()["document_json"]["rows"] == [
            ["10", {"value": "row-10", "link": {"url": "https://example.com/row-10"}}],
            ["11", "row-11"],
        ]
        assert page2_response.json()["document_json"]["cell_meta"] == {}

        desc_response = client.post(
            f"/api/note-sheets/sheets/{sheet_id}/sort",
            json={
                "column_index": 0,
                "direction": "desc",
            },
        )
        assert desc_response.status_code == 200
        desc_detail = desc_response.json()
        assert desc_detail["document_json"]["rows"] == [
            ["11", "row-11"],
            ["10", {"value": "row-10", "link": {"url": "https://example.com/row-10"}}],
        ]
        assert desc_detail["document_json"]["cell_meta"] == {}

        desc_page2_response = client.get(
            f"/api/note-sheets/sheets/{sheet_id}",
            params={"page": 2, "page_size": 2},
        )
        assert desc_page2_response.status_code == 200
        assert desc_page2_response.json()["document_json"]["rows"] == [
            ["2", "row-2"],
            ["1", {"value": "row-1", "link": {"url": "https://example.com/row-1"}}],
        ]
        assert desc_page2_response.json()["document_json"]["cell_meta"] == {}

        last_page_response = client.get(
            f"/api/note-sheets/sheets/{sheet_id}",
            params={"page": 3, "page_size": 2},
        )
        assert last_page_response.status_code == 200
        assert last_page_response.json()["document_json"]["rows"] == [
            ["", "row-empty"],
        ]
    finally:
        _clear_user_override()


def test_note_sheet_paged_document_slices_row_metadata():
    document = {
        "schema_version": 1,
        "columns": ["序号", "内容"],
        "rows": [
            ["1", {"value": "row-1", "link": {"url": "https://example.com/1"}}],
            ["2", {"value": "row-2", "link": {"url": "https://example.com/2"}}],
            ["3", {"value": "row-3", "link": {"url": "https://example.com/3"}}],
            ["4", {"value": "row-4", "link": {"url": "https://example.com/4"}}],
        ],
        "grid_rows": [
            ["序号", "内容"],
            ["1", {"value": "row-1", "link": {"url": "https://example.com/1"}}],
            ["2", {"value": "row-2", "link": {"url": "https://example.com/2"}}],
            ["3", {"value": "row-3", "link": {"url": "https://example.com/3"}}],
            ["4", {"value": "row-4", "link": {"url": "https://example.com/4"}}],
        ],
        "data_start_row": 1,
        "field_row_index": 0,
        "cell_meta": {
            "0:0": {"style": {"background_color": "#e0f2fe"}},
            "legacy": {"note": "keep"},
        },
        "entity_rows": [
            {"id": "h0", "kind": "field"},
            {"id": "r1", "kind": "data"},
            {"id": "r2", "kind": "data"},
            {"id": "r3", "kind": "data"},
            {"id": "r4", "kind": "data"},
        ],
        "entity_cells": {
            "h0": {"c1": {"value": "内容"}},
            "r1": {"c1": {"value": "row-1"}},
            "r2": {"c1": {"value": "row-2"}},
            "r3": {"c1": {"value": "row-3"}},
            "r4": {"c1": {"value": "row-4"}},
        },
    }

    page_document, pagination = note_sheets_api._build_paged_document(document, page=2, page_size=2)

    assert page_document["rows"] == [
        ["3", {"value": "row-3", "link": {"url": "https://example.com/3"}}],
        ["4", {"value": "row-4", "link": {"url": "https://example.com/4"}}],
    ]
    assert page_document["grid_rows"] == [
        ["序号", "内容"],
        ["3", {"value": "row-3", "link": {"url": "https://example.com/3"}}],
        ["4", {"value": "row-4", "link": {"url": "https://example.com/4"}}],
    ]
    assert page_document["cell_meta"] == {
        "0:0": {"style": {"background_color": "#e0f2fe"}},
        "legacy": {"note": "keep"},
    }
    assert page_document["entity_rows"] == [
        {"id": "h0", "kind": "field"},
        {"id": "r3", "kind": "data"},
        {"id": "r4", "kind": "data"},
    ]
    assert set(page_document["entity_cells"]) == {"h0", "r3", "r4"}
    assert pagination.row_offset == 2
    assert pagination.loaded_row_count == 2


def test_note_sheet_filtered_paged_document_slices_row_metadata():
    document = {
        "schema_version": 1,
        "columns": ["分类", "内容"],
        "rows": [
            ["收入", "income-1"],
            ["支出", {"value": "expense-2", "link": {"url": "https://example.com/2"}}],
            ["收入", "income-3"],
            ["支出", {"value": "expense-4", "link": {"url": "https://example.com/4"}}],
            ["收入", "income-5"],
            ["支出", {"value": "expense-6", "link": {"url": "https://example.com/6"}}],
        ],
        "grid_rows": [
            ["分类", "内容"],
            ["收入", "income-1"],
            ["支出", {"value": "expense-2", "link": {"url": "https://example.com/2"}}],
            ["收入", "income-3"],
            ["支出", {"value": "expense-4", "link": {"url": "https://example.com/4"}}],
            ["收入", "income-5"],
            ["支出", {"value": "expense-6", "link": {"url": "https://example.com/6"}}],
        ],
        "data_start_row": 1,
        "field_row_index": 0,
        "cell_meta": {
            "0:0": {"style": {"background_color": "#e0f2fe"}},
        },
        "entity_rows": [
            {"id": "h0", "kind": "field"},
            {"id": "r1", "kind": "data"},
            {"id": "r2", "kind": "data"},
            {"id": "r3", "kind": "data"},
            {"id": "r4", "kind": "data"},
            {"id": "r5", "kind": "data"},
            {"id": "r6", "kind": "data"},
        ],
        "entity_cells": {
            "h0": {"c1": {"value": "内容"}},
            "r2": {"c1": {"value": "expense-2"}},
            "r4": {"c1": {"value": "expense-4"}},
            "r6": {"c1": {"value": "expense-6"}},
        },
        "column_configs": {
            "分类": {
                "filter_enabled": True,
                "value_mode": "fixed_options",
            },
        },
    }

    page_document, pagination = note_sheets_api._build_filtered_paged_document(
        document,
        page=2,
        page_size=2,
        column_filters={"分类": {"excludedValues": ["收入"]}},
        row_filter_programs=[],
    )

    assert page_document["rows"] == [["支出", {"value": "expense-6", "link": {"url": "https://example.com/6"}}]]
    assert page_document["grid_rows"] == [
        ["分类", "内容"],
        ["支出", {"value": "expense-6", "link": {"url": "https://example.com/6"}}],
    ]
    assert page_document["cell_meta"] == {
        "0:0": {"style": {"background_color": "#e0f2fe"}},
    }
    assert page_document["entity_rows"] == [
        {"id": "h0", "kind": "field"},
        {"id": "r6", "kind": "data"},
    ]
    assert set(page_document["entity_cells"]) == {"h0", "r6"}
    assert pagination["row_indexes"] == [5]
    assert pagination["row_offset"] == 2
    assert pagination["loaded_row_count"] == 1


def test_note_sheet_unified_grid_page_patch_preserves_headers_and_merges(client, session):
    user = _create_user(session, username="note-sheet-grid-page-user")
    _grant_feature_access(session, user_id=user.id, feature_key="notes.sheets")
    _override_user(user)

    try:
        document = {
            "schema_version": 1,
            "columns": ["序号", "内容"],
            "rows": [["1", "row-1"], ["2", "row-2"], ["3", "row-3"]],
            "grid_rows": [
                ["分组", ""],
                ["序号", "内容"],
                ["1", "row-1"],
                ["2", "row-2"],
                ["3", "row-3"],
            ],
            "data_start_row": 2,
            "field_row_index": 1,
            "merged_cells": [{"row": 0, "col": 0, "rowspan": 1, "colspan": 2}],
            "cell_meta": {"0:0": {"style": {"background_color": "#e0f2fe"}}},
            "view_settings": {"pagination": {"enabled": True, "page_size": 2}},
        }
        create_response = client.post(
            "/api/note-sheets/sheets",
            json={"title": "统一网格分页", "document_json": document},
        )
        assert create_response.status_code == 200
        sheet_id = create_response.json()["id"]

        page2_response = client.get(f"/api/note-sheets/sheets/{sheet_id}", params={"page": 2, "page_size": 2})
        assert page2_response.status_code == 200
        page2_document = page2_response.json()["document_json"]
        assert page2_document["rows"] == [["3", "row-3"]]
        assert page2_document["grid_rows"] == [["分组", ""], ["序号", "内容"], ["3", "row-3"]]

        save_response = client.put(
            f"/api/note-sheets/sheets/{sheet_id}",
            json={
                "document_json": {
                    **page2_document,
                    "rows": [["30", "row-30"]],
                    "grid_rows": [["分组", ""], ["序号", "内容"], ["30", "row-30"]],
                },
                "page_patch": {
                    "page": 2,
                    "page_size": 2,
                    "row_offset": 2,
                    "loaded_row_count": 1,
                },
            },
        )
        assert save_response.status_code == 200
        full_response = client.get(f"/api/note-sheets/sheets/{sheet_id}", params={"paginate": False})
        assert full_response.status_code == 200
        saved_document = full_response.json()["document_json"]
        assert saved_document["rows"] == [["1", "row-1"], ["2", "row-2"], ["30", "row-30"]]
        assert saved_document["grid_rows"] == [
            ["分组", ""],
            ["序号", "内容"],
            ["1", "row-1"],
            ["2", "row-2"],
            ["30", "row-30"],
        ]
        assert saved_document["merged_cells"] == [{"row": 0, "col": 0, "rowspan": 1, "colspan": 2}]
        assert saved_document["cell_meta"] == {"0:0": {"style": {"background_color": "#e0f2fe"}}}
    finally:
        _clear_user_override()


def test_note_sheet_page_patch_merges_entity_cells_without_overwriting_unloaded_rows(client, session):
    user = _create_user(session, username="note-sheet-entity-page-user")
    _grant_feature_access(session, user_id=user.id, feature_key="notes.sheets")
    _override_user(user)

    try:
        document = {
            "schema_version": 1,
            "columns": ["序号", "链接"],
            "rows": [
                ["1", {"value": "row-1", "link": {"url": "https://example.com/1"}}],
                ["2", {"value": "row-2", "link": {"url": "https://example.com/2"}}],
                ["3", {"value": "row-3", "link": {"url": "https://example.com/3"}}],
            ],
            "grid_rows": [
                ["序号", "链接"],
                ["1", {"value": "row-1", "link": {"url": "https://example.com/1"}}],
                ["2", {"value": "row-2", "link": {"url": "https://example.com/2"}}],
                ["3", {"value": "row-3", "link": {"url": "https://example.com/3"}}],
            ],
            "data_start_row": 1,
            "field_row_index": 0,
            "entity_columns": [{"id": "c0", "header": "序号"}, {"id": "c1", "header": "链接"}],
            "entity_rows": [
                {"id": "h0", "kind": "field"},
                {"id": "r1", "kind": "data"},
                {"id": "r2", "kind": "data"},
                {"id": "r3", "kind": "data"},
            ],
            "entity_cells": {
                "r1": {"c1": {"value": "row-1"}},
                "r2": {"c1": {"value": "row-2"}},
                "r3": {"c1": {"value": "row-3"}},
            },
            "view_settings": {"pagination": {"enabled": True, "page_size": 2}},
        }
        create_response = client.post(
            "/api/note-sheets/sheets",
            json={"title": "实体分页", "document_json": document},
        )
        assert create_response.status_code == 200
        sheet_id = create_response.json()["id"]

        page2_response = client.get(f"/api/note-sheets/sheets/{sheet_id}", params={"page": 2, "page_size": 2})
        assert page2_response.status_code == 200
        page2_document = page2_response.json()["document_json"]
        save_response = client.put(
            f"/api/note-sheets/sheets/{sheet_id}",
            json={
                "document_json": {
                    **page2_document,
                    "rows": [["3", {"value": "row-30", "link": {"url": "https://example.com/30"}}]],
                    "grid_rows": [
                        ["序号", "链接"],
                        ["3", {"value": "row-30", "link": {"url": "https://example.com/30"}}],
                    ],
                    "entity_rows": [{"id": "h0", "kind": "field"}, {"id": "r3", "kind": "data"}],
                    "entity_cells": {
                        "r3": {
                            "c1": {
                                "value": "row-30",
                            },
                        },
                    },
                },
                "page_patch": {
                    "page": 2,
                    "page_size": 2,
                    "row_offset": 2,
                    "loaded_row_count": 1,
                },
            },
        )
        assert save_response.status_code == 200

        full_response = client.get(f"/api/note-sheets/sheets/{sheet_id}", params={"paginate": False})
        assert full_response.status_code == 200
        saved_document = full_response.json()["document_json"]
        assert saved_document["rows"] == [
            ["1", {"value": "row-1", "link": {"url": "https://example.com/1"}}],
            ["2", {"value": "row-2", "link": {"url": "https://example.com/2"}}],
            ["3", {"value": "row-30", "link": {"url": "https://example.com/30"}}],
        ]
        assert saved_document["entity_rows"] == [
            {"id": "h0", "kind": "field"},
            {"id": "r1", "kind": "data"},
            {"id": "r2", "kind": "data"},
            {"id": "r3", "kind": "data"},
        ]
        assert "link" not in saved_document["entity_cells"]["r1"]["c1"]
        assert "link" not in saved_document["entity_cells"]["r2"]["c1"]
        assert saved_document["entity_cells"]["r3"]["c1"]["value"] == "row-30"
        assert "link" not in saved_document["entity_cells"]["r3"]["c1"]
    finally:
        _clear_user_override()


def test_attendance_summary_link_lookup_prefers_entity_and_document_row_meta():
    document = {
        "schema_version": 1,
        "columns": ["课程类型", "课程名称", "在线考勤表"],
        "rows": [[
            "禅宗4.5阶",
            "禅宗8期4.5阶",
            {
                "value": "20260308禅宗8期4.5阶",
                "link": {"url": "https://www.kdocs.cn/l/inline-token"},
            },
        ]],
        "grid_rows": [
            ["课程类型", "课程名称", "在线考勤表"],
            ["", "", "备注"],
            ["禅宗4.5阶", "禅宗8期4.5阶", "20260308禅宗8期4.5阶"],
        ],
        "data_start_row": 2,
        "entity_columns": [
            {"id": "c0", "header": "课程类型"},
            {"id": "c1", "header": "课程名称"},
            {"id": "c2", "header": "在线考勤表"},
        ],
        "entity_rows": [
            {"id": "h0", "kind": "field"},
            {"id": "h1", "kind": "field_note"},
            {"id": "r0", "kind": "data"},
        ],
        "entity_cells": {
            "r0": {
                "c2": {
                    "value": "20260308禅宗8期4.5阶",
                    "link": {"url": "https://www.kdocs.cn/l/entity-token"},
                },
            },
        },
        }

    assert (
        note_sheets_api._get_document_cell_link_url(document, 0, 2)
        == "https://www.kdocs.cn/l/inline-token"
    )

    document_without_inline = {
        **document,
        "rows": [["禅宗4.5阶", "禅宗8期4.5阶", "20260308禅宗8期4.5阶"]],
    }
    assert note_sheets_api._get_document_cell_link_url(document_without_inline, 0, 2) == ""

    document_without_inline_or_entity = {
        **document_without_inline,
        "entity_cells": {},
    }
    assert note_sheets_api._get_document_cell_link_url(document_without_inline_or_entity, 0, 2) == ""


def test_course_attendance_header_links_are_derived_from_config_sheets(client, session):
    user = _create_user(session, username="course-header-link-user")
    _grant_feature_access(session, user_id=user.id, feature_key="notes.sheets")
    _override_user(user)
    owner_key = "test-course-header-links"

    def add_sheet(sheet_key: str, title: str, document_json: dict) -> SheetDocument:
        identity = allocate_new_sheet_identity(session)
        sheet = SheetDocument(
            id=identity.primary_id,
            numeric_id=identity.numeric_id,
            legacy_id=identity.legacy_id,
            scope="notes",
            owner_type="course_workbook",
            owner_key=owner_key,
            sheet_key=sheet_key,
            title=title,
            engine="handsontable",
            document_json=document_json,
            owner_user_id=user.id,
            created_by_user_id=user.id,
            updated_by_user_id=user.id,
            created_at=time.time(),
            updated_at=time.time(),
        )
        session.add(sheet)
        session.flush()
        return sheet

    try:
        attendance = add_sheet(
            "attendance",
            "考勤表",
            {
                "schema_version": 1,
                "columns": ["姓名", "打卡数", "05:20~06:18 第01课", "05:20~06:02 第02课"],
                "rows": [],
                "grid_rows": [
                    ["", "打卡", "6月1日~6月5日", "6月2日~6月6日"],
                    ["姓名", "打卡数", "05:20~06:18 第01课", "05:20~06:02 第02课"],
                    ["", "只统计正课的打卡次数", "", ""],
                ],
                "data_start_row": 3,
                "field_row_index": 1,
                "cell_meta": {},
            },
        )
        add_sheet(
            "video_config",
            "视频配置",
            {
                "columns": ["lesson_id", "lesson_id2", "lesson_name", "video_duration"],
                "rows": [
                    [1, "l_testLesson01", "第01课", 3510],
                    [2, "https://example.com/video-02", "第02课", 2538],
                ],
                "grid_rows": [
                    ["lesson_id", "lesson_id2", "lesson_name", "video_duration"],
                    [1, "l_testLesson01", "第01课", 3510],
                    [2, "https://example.com/video-02", "第02课", 2538],
                ],
                "data_start_row": 1,
            },
        )
        add_sheet(
            "video_data",
            "视频数据",
            {"columns": ["lesson_id"], "rows": [], "grid_rows": [["lesson_id"]], "data_start_row": 1},
        )
        add_sheet(
            "clockin_config",
            "打卡配置",
            {
                "columns": ["clockin_id", "name", "url"],
                "rows": [[1, "打卡数", "https://example.com/clockin"]],
                "grid_rows": [["clockin_id", "name", "url"], [1, "打卡数", "https://example.com/clockin"]],
                "data_start_row": 1,
            },
        )
        add_sheet(
            "clockin_data",
            "打卡数据",
            {"columns": ["clockin_id"], "rows": [], "grid_rows": [["clockin_id"]], "data_start_row": 1},
        )
        session.commit()

        response = client.get(f"/api/note-sheets/sheets/{attendance.numeric_id}", params={"paginate": False})
        assert response.status_code == 200
        field_row = response.json()["document_json"]["grid_rows"][1]
        assert field_row[1]["link"]["url"] == "https://example.com/clockin"
        assert (
            field_row[2]["link"]["url"]
            == "https://admin.xiaoe-tech.com/t/live_management#/userOperation?id=l_testLesson01&tabName=UserManage"
        )
        assert field_row[3]["link"]["url"] == "https://example.com/video-02"
    finally:
        _clear_user_override()


def test_note_sheet_unified_grid_sort_remaps_sheet_meta_and_rejects_rowspan_merge(client, session):
    user = _create_user(session, username="note-sheet-grid-sort-user")
    _grant_feature_access(session, user_id=user.id, feature_key="notes.sheets")
    _override_user(user)

    try:
        create_response = client.post(
            "/api/note-sheets/sheets",
            json={
                "title": "统一网格排序",
                "document_json": {
                    "schema_version": 1,
                    "columns": ["序号", "内容"],
                    "rows": [
                        ["2", {"value": "row-2", "link": {"url": "https://example.com/row-2"}}],
                        ["1", "row-1"],
                    ],
                    "grid_rows": [
                        ["分组", ""],
                        ["序号", "内容"],
                        ["2", {"value": "row-2", "link": {"url": "https://example.com/row-2"}}],
                        ["1", "row-1"],
                    ],
                    "data_start_row": 2,
                    "field_row_index": 1,
                    "merged_cells": [{"row": 2, "col": 0, "rowspan": 1, "colspan": 2}],
                    "cell_meta": {},
                },
            },
        )
        assert create_response.status_code == 200
        sheet_id = create_response.json()["id"]

        sort_response = client.post(
            f"/api/note-sheets/sheets/{sheet_id}/sort",
            json={"column_index": 0, "direction": "asc"},
        )
        assert sort_response.status_code == 200
        sorted_document = sort_response.json()["document_json"]
        assert sorted_document["rows"] == [
            ["1", "row-1"],
            ["2", {"value": "row-2", "link": {"url": "https://example.com/row-2"}}],
        ]
        assert sorted_document["grid_rows"] == [
            ["分组", ""],
            ["序号", "内容"],
            ["1", "row-1"],
            ["2", {"value": "row-2", "link": {"url": "https://example.com/row-2"}}],
        ]
        assert sorted_document["cell_meta"] == {}

        reject_response = client.put(
            f"/api/note-sheets/sheets/{sheet_id}",
            json={
                "document_json": {
                    **sorted_document,
                    "merged_cells": [{"row": 2, "col": 0, "rowspan": 2, "colspan": 1}],
                },
            },
        )
        assert reject_response.status_code == 200
        reject_sort_response = client.post(
            f"/api/note-sheets/sheets/{sheet_id}/sort",
            json={"column_index": 0, "direction": "asc"},
        )
        assert reject_sort_response.status_code == 400
        assert reject_sort_response.json()["detail"] == "数据区存在跨行合并，不能排序"
    finally:
        _clear_user_override()


def test_migrate_attendance_course_sheets_to_notes_workbook_is_idempotent(session):
    owner_user = _create_user(session, username="attendance-workbook-owner")
    source_document = SheetDocument(
        scope="attendance",
        owner_type="course_session",
        owner_key="20260412-chanzong-12qi-1jie",
        sheet_key="registration",
        title="报名表",
        engine="handsontable",
        document_json={
            "schema_version": 1,
            "columns": ["姓名", "手机号"],
            "rows": [["时秋菊", "15641363033"]],
        },
        version=3,
        owner_user_id=owner_user.id,
        created_by_user_id=owner_user.id,
        updated_by_user_id=owner_user.id,
    )
    session.add(source_document)
    session.commit()

    v29_migrate_attendance_course_sheets_to_notes_workbook(session)
    v29_migrate_attendance_course_sheets_to_notes_workbook(session)

    workbooks = session.exec(select(WorkbookDocument)).all()
    note_sheets = session.exec(
        select(SheetDocument)
        .where(SheetDocument.scope == "notes")
        .where(SheetDocument.owner_type == "course_workbook")
        .where(SheetDocument.owner_key == "20260412-chanzong-12qi-1jie")
    ).all()
    links = session.exec(select(WorkbookSheetLink)).all()

    assert len(workbooks) == 1
    assert workbooks[0].title == "20260412禅宗12期一阶"
    assert workbooks[0].owner_user_id == owner_user.id

    note_sheet_map = {item.sheet_key: item for item in note_sheets}
    assert set(note_sheet_map) == {"registration", "attendance"}
    assert note_sheet_map["registration"].document_json["rows"] == [["时秋菊", "15641363033"]]
    assert note_sheet_map["attendance"].title == "考勤表"
    assert len(links) == 2


def test_v30_backfills_numeric_sheet_and_workbook_ids(session):
    owner_user = _create_user(session, username="numeric-id-owner")
    workbook = WorkbookDocument(
        title="工作簿A",
        owner_user_id=owner_user.id,
        created_by_user_id=owner_user.id,
        updated_by_user_id=owner_user.id,
    )
    sheet = SheetDocument(
        scope="notes",
        owner_type="user",
        owner_key=str(owner_user.id),
        sheet_key="sheet-a",
        title="表格A",
        engine="handsontable",
        document_json={"schema_version": 1, "columns": [], "rows": []},
        owner_user_id=owner_user.id,
        created_by_user_id=owner_user.id,
        updated_by_user_id=owner_user.id,
    )
    session.add(workbook)
    session.add(sheet)
    session.commit()

    assert workbook.numeric_id is None
    assert sheet.numeric_id is None

    v30_add_numeric_sheet_and_workbook_ids(session)
    session.refresh(workbook)
    session.refresh(sheet)

    assert workbook.numeric_id == 1
    assert sheet.numeric_id == 1
